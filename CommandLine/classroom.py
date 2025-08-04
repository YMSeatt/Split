import json
import os
import sys
from datetime import datetime
from pathlib import Path
import re
import cryptography.fernet
from openpyxl import Workbook
from openpyxl.styles import Font as OpenpyxlFont, Alignment as OpenpyxlAlignment
from openpyxl.utils import get_column_letter

# Add the parent directory to the path to allow importing from the root
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from data_encryption import _read_and_decrypt_file, _encrypt_and_write_file

# --- Constants copied from seatingchartmain.py ---
APP_NAME = "BehaviorLogger"
CURRENT_DATA_VERSION_TAG = "v10"
DATA_FILE_PATTERN = f"classroom_data_{CURRENT_DATA_VERSION_TAG}.json"

def get_app_data_path(filename):
    """Simplified version of get_app_data_path from the main app."""
    # For the TUI, we'll just use the root directory for now.
    # This can be enhanced later if needed.
    base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    return os.path.join(base_path, filename)

DATA_FILE = get_app_data_path(DATA_FILE_PATTERN)
LAYOUT_TEMPLATES_DIR_NAME = "layout_templates"
LAYOUT_TEMPLATES_DIR = get_app_data_path(LAYOUT_TEMPLATES_DIR_NAME)

DEFAULT_BEHAVIORS_LIST = [
    "Talking", "Off Task", "Out of Seat", "Uneasy", "Placecheck",
    "Great Participation", "Called On", "Complimented", "Fighting", "Other"
]

DEFAULT_HOMEWORK_TYPES_LIST = [
    "Reading Assignment", "Worksheet", "Math Problems", "Project Work", "Study for Test"
]

DEFAULT_HOMEWORK_STATUSES = [
    "Done", "Not Done", "Partially Done", "Signed", "Returned", "Late", "Excellent Work"
]


class Classroom:
    """
    Manages the classroom data, including students, furniture, logs, and settings.
    This class is designed to be compatible with the data structure of the main Tkinter application.
    """
    def __init__(self, data_file=DATA_FILE):
        self.data_file = Path(data_file)

        # Ensure layout templates directory exists
        if not os.path.exists(LAYOUT_TEMPLATES_DIR):
            os.makedirs(LAYOUT_TEMPLATES_DIR, exist_ok=True)

        self.students = {}
        self.furniture = {}
        self.behavior_log = []
        self.homework_log = []
        self.student_groups = {}
        self.quiz_templates = {}
        self.homework_templates = {}
        self.guides = {}
        self.settings = self._get_default_settings()
        self.undo_stack = []
        self.redo_stack = []
        self._per_student_last_cleared = {}
        self.last_excel_export_path = None
        self.next_guide_id_num = 1

        self._load_data()

    def _get_default_settings(self):
        """
        Returns a dictionary of default settings, mirroring the main application.
        This ensures that if no settings file is found, the app starts with a consistent state.
        """
        # This is a subset of the settings from the main app.
        # It can be expanded as more features are added to the TUI.
        return {
            "encrypt_data_files": True,
            "show_recent_incidents_on_boxes": True,
            "num_recent_incidents_to_show": 2,
            "recent_incident_time_window_hours": 24,
            "autosave_interval_ms": 30000,
            "default_student_box_width": 130,
            "default_student_box_height": 80,
            "next_student_id_num": 1,
            "next_furniture_id_num": 1,
            "next_group_id_num": 1,
            "next_quiz_template_id_num": 1,
            "next_homework_template_id_num": 1,
            "next_custom_homework_type_id_num": 1,
            "theme": "System",
        }

    def _load_data(self):
        """
        Loads the main application data from the JSON file.
        It handles decryption and populates the classroom object's attributes.
        """
        data = _read_and_decrypt_file(self.data_file)

        if data:
            # Basic migration check can be added here in the future if needed

            default_settings_copy = self._get_default_settings()
            # Merge loaded settings with defaults to ensure all keys are present
            final_settings = default_settings_copy.copy()
            final_settings.update(data.get("settings", {}))

            self.students = data.get("students", {})
            self.furniture = data.get("furniture", {})
            self.behavior_log = data.get("behavior_log", [])
            self.homework_log = data.get("homework_log", [])
            self.settings = final_settings
            self.guides = data.get("guides", {})
            self.next_guide_id_num = data.get("next_guide_id_num", 1)

            # More attributes can be loaded here as they are implemented
            # e.g., self.student_groups, self.quiz_templates, etc.

            # For now, we don't handle the undo/redo stack in the TUI
            self.undo_stack = []
            self.redo_stack = []

            self._ensure_next_ids()
            return True
        return False

    def _save_data(self):
        """
        Saves the current state of the classroom to the JSON data file.
        Handles encryption before writing.
        """
        self._ensure_next_ids()

        data_to_save = {
            "students": self.students,
            "furniture": self.furniture,
            "behavior_log": self.behavior_log,
            "homework_log": self.homework_log,
            "settings": self.settings,
            "guides": self.guides,
            "next_guide_id_num": self.next_guide_id_num,
            # Empty placeholders for compatibility
            "student_groups": self.student_groups,
            "quiz_templates": self.quiz_templates,
            "homework_templates": self.homework_templates,
            "undo_stack": [], # Not implemented yet
            "redo_stack": [], # Not implemented yet
            "_per_student_last_cleared": self._per_student_last_cleared,
            "last_excel_export_path": self.last_excel_export_path,
        }

        # The encryption rule is read from settings
        should_encrypt = self.settings.get("encrypt_data_files", True)
        _encrypt_and_write_file(self.data_file, data_to_save, should_encrypt)

    def _ensure_next_ids(self):
        """Ensures the next_id counters in settings are up to date."""
        max_s_id = 0
        for sid in self.students:
            if sid.startswith("student_"):
                try: max_s_id = max(max_s_id, int(sid.split("_")[1]))
                except (ValueError, IndexError): pass
        self.settings["next_student_id_num"] = max(self.settings.get("next_student_id_num", 1), max_s_id + 1)

        max_f_id = 0
        for fid in self.furniture:
            if fid.startswith("furniture_"):
                try: max_f_id = max(max_f_id, int(fid.split("_")[1]))
                except (ValueError, IndexError): pass
        self.settings["next_furniture_id_num"] = max(self.settings.get("next_furniture_id_num", 1), max_f_id + 1)

    def add_student(self, first_name, last_name, nickname="", gender="Boy"):
        """
        Adds a new student to the classroom with full details.
        This is a more comprehensive version of the original method.
        """
        next_id_num = self.settings.get("next_student_id_num", 1)
        student_id_str = f"student_{next_id_num}"

        full_name = f"{first_name} \"{nickname}\" {last_name}" if nickname else f"{first_name} {last_name}"

        student_data = {
            "id": student_id_str,
            "first_name": first_name,
            "last_name": last_name,
            "nickname": nickname,
            "full_name": full_name,
            "gender": gender,
            "x": 50, "y": 50, # Default position
            "width": self.settings.get("default_student_box_width"),
            "height": self.settings.get("default_student_box_height"),
            "group_id": None,
            "style_overrides": {}
        }

        self.students[student_id_str] = student_data
        self.settings["next_student_id_num"] = next_id_num + 1
        self._save_data()
        return student_id_str

    def remove_student(self, student_id):
        """Removes a student by their ID."""
        if student_id in self.students:
            del self.students[student_id]
            # Also remove associated logs
            self.behavior_log = [log for log in self.behavior_log if log.get("student_id") != student_id]
            self.homework_log = [log for log in self.homework_log if log.get("student_id") != student_id]
            self._save_data()
            return True
        return False

    def _make_safe_sheet_name(self, name_str, id_fallback="Sheet"):
        invalid_chars = r'[\\/?*\[\]:]' # Excel invalid sheet name characters
        safe_name = re.sub(invalid_chars, '_', str(name_str))
        if not safe_name: safe_name = str(id_fallback)
        return safe_name[:31] # Max 31 chars for sheet names

    def export_data_to_excel(self, file_path, filter_settings=None):
        # This is a simplified version of the export logic from the main app.
        # It does not support all the advanced filtering and formatting yet.
        wb = Workbook()
        wb.remove(wb.active) # Remove default sheet

        student_data_for_export = {sid: {"first_name": s["first_name"], "last_name": s["last_name"], "full_name": s["full_name"]} for sid, s in self.students.items()}

        logs_to_process = self.behavior_log + self.homework_log

        # Create a combined log sheet
        ws = wb.create_sheet(title="Combined Log")
        headers = ["Timestamp", "Date", "Time", "Day", "Student ID", "First Name", "Last Name", "Log Type", "Item Name", "Comment"]
        ws.append(headers)

        for entry in sorted(logs_to_process, key=lambda x: x["timestamp"]):
            student_info = student_data_for_export.get(entry["student_id"], {"first_name": "N/A", "last_name": "N/A"})
            try:
                dt_obj = datetime.fromisoformat(entry["timestamp"])
            except ValueError:
                dt_obj = datetime.now()

            log_type = entry.get("type", "behavior").capitalize()
            item_name = entry.get("behavior")
            if log_type == "Homework":
                item_name = f"{entry.get('homework_type')}: {entry.get('homework_status')}"

            row_data = [
                entry["timestamp"],
                dt_obj.strftime('%Y-%m-%d'),
                dt_obj.strftime('%H:%M:%S'),
                entry.get("day", dt_obj.strftime('%A')),
                entry["student_id"],
                student_info["first_name"],
                student_info["last_name"],
                log_type,
                item_name,
                entry.get("comment", "")
            ]
            ws.append(row_data)

        try:
            wb.save(filename=file_path)
            return True
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            return False

    # --- Logging Methods ---

    def get_all_behaviors(self):
        """
        Returns a list of all available behaviors.
        For now, this is just the default list.
        """
        # In the future, this can be expanded to include custom behaviors from settings.
        return DEFAULT_BEHAVIORS_LIST

    def log_behavior(self, student_id, behavior, comment=""):
        """Logs a behavior for a student."""
        student = self.get_student(student_id)
        if not student:
            return False

        log_entry = {
            "timestamp": datetime.now().isoformat(),
            "student_id": student_id,
            "student_first_name": student["first_name"],
            "student_last_name": student["last_name"],
            "behavior": behavior,
            "comment": comment,
            "type": "behavior",
            "day": datetime.now().strftime('%A')
        }
        self.behavior_log.append(log_entry)
        self._save_data()
        return True

    def log_quiz_score(self, student_id, quiz_name, marks_data, num_questions, comment=""):
        """Logs a quiz score for a student."""
        student = self.get_student(student_id)
        if not student:
            return False

        log_entry = {
            "timestamp": datetime.now().isoformat(),
            "student_id": student_id,
            "student_first_name": student["first_name"],
            "student_last_name": student["last_name"],
            "behavior": quiz_name,
            "comment": comment,
            "marks_data": marks_data,
            "num_questions": num_questions,
            "type": "quiz",
            "day": datetime.now().strftime('%A')
        }
        self.behavior_log.append(log_entry) # Quiz logs are also in behavior_log
        self._save_data()
        return True

    def get_all_homework_types(self):
        return DEFAULT_HOMEWORK_TYPES_LIST

    def get_all_homework_statuses(self):
        return DEFAULT_HOMEWORK_STATUSES

    def log_homework(self, student_id, homework_type, status, comment=""):
        """Logs a homework entry for a student."""
        student = self.get_student(student_id)
        if not student:
            return False

        log_entry = {
            "timestamp": datetime.now().isoformat(),
            "student_id": student_id,
            "student_first_name": student["first_name"],
            "student_last_name": student["last_name"],
            "behavior": f"{homework_type}: {status}",
            "homework_type": homework_type,
            "homework_status": status,
            "comment": comment,
            "type": "homework",
            "day": datetime.now().strftime('%A')
        }
        self.homework_log.append(log_entry)
        self._save_data()
        return True

    # --- Layout Template Methods ---

    def list_layout_templates(self):
        """Returns a list of available layout template filenames."""
        if not os.path.exists(LAYOUT_TEMPLATES_DIR):
            return []
        return [f for f in os.listdir(LAYOUT_TEMPLATES_DIR) if f.endswith(".json")]

    def save_layout_template(self, template_name: str):
        """Saves the current layout of students and furniture to a template file."""
        if not template_name.endswith(".json"):
            template_name += ".json"

        file_path = os.path.join(LAYOUT_TEMPLATES_DIR, template_name)

        layout_data = {
            "students": {
                sid: {
                    "x": s["x"], "y": s["y"],
                    "width": s.get("width"), "height": s.get("height"),
                    "style_overrides": s.get("style_overrides", {}).copy(),
                    "first_name": s.get("first_name", ""), "last_name": s.get("last_name", ""),
                    "nickname": s.get("nickname", "")
                } for sid, s in self.students.items()
            },
            "furniture": {
                fid: {
                    "x": f["x"], "y": f["y"],
                    "width": f.get("width"), "height": f.get("height")
                } for fid, f in self.furniture.items()
            }
        }

        should_encrypt = self.settings.get("encrypt_data_files", True)
        _encrypt_and_write_file(file_path, layout_data, should_encrypt)
        return True

    def load_layout_template(self, template_name: str):
        """Loads a layout from a template file and applies it."""
        if not template_name.endswith(".json"):
            template_name += ".json"

        file_path = os.path.join(LAYOUT_TEMPLATES_DIR, template_name)
        if not os.path.exists(file_path):
            return False

        template_data = _read_and_decrypt_file(file_path)
        if not template_data:
            return False

        template_students = template_data.get("students", {})

        # Match by name first, then by ID as a fallback
        # This is a simplified version of the main app's logic
        for t_sid, t_sdata in template_students.items():
            found_student = None
            # Try to find a matching student in the current classroom
            for c_sid, c_sdata in self.students.items():
                if (c_sdata.get("first_name") == t_sdata.get("first_name") and
                        c_sdata.get("last_name") == t_sdata.get("last_name")):
                    found_student = c_sdata
                    break

            if not found_student and t_sid in self.students:
                found_student = self.students[t_sid]

            if found_student:
                found_student['x'] = t_sdata.get('x', found_student['x'])
                found_student['y'] = t_sdata.get('y', found_student['y'])
                found_student['width'] = t_sdata.get('width', found_student['width'])
                found_student['height'] = t_sdata.get('height', found_student['height'])

        # Apply furniture layout by ID
        template_furniture = template_data.get("furniture", {})
        for f_id, f_data in template_furniture.items():
            if f_id in self.furniture:
                self.furniture[f_id]['x'] = f_data.get('x', self.furniture[f_id]['x'])
                self.furniture[f_id]['y'] = f_data.get('y', self.furniture[f_id]['y'])

        self._save_data()
        return True

    # --- Furniture Methods ---

    def add_furniture(self, name, item_type, width=200, height=100):
        """Adds a new piece of furniture."""
        next_id_num = self.settings.get("next_furniture_id_num", 1)
        furniture_id_str = f"furniture_{next_id_num}"

        furniture_data = {
            "id": furniture_id_str,
            "name": name,
            "type": item_type,
            "x": 50, "y": 200, # Default position
            "width": width,
            "height": height,
            "fill_color": "lightgray",
            "outline_color": "dimgray",
        }

        self.furniture[furniture_id_str] = furniture_data
        self.settings["next_furniture_id_num"] = next_id_num + 1
        self._save_data()
        return furniture_id_str

    def remove_furniture(self, furniture_id):
        """Removes a piece of furniture by its ID."""
        if furniture_id in self.furniture:
            del self.furniture[furniture_id]
            self._save_data()
            return True
        return False

    def list_furniture(self):
        """Returns a list of furniture data dictionaries."""
        return sorted(self.furniture.values(), key=lambda f: f.get('name', '').lower())

    def get_furniture(self, furniture_id):
        """Retrieves a single piece of furniture's data."""
        return self.furniture.get(furniture_id)

    def update_furniture(self, furniture_id, data):
        """Updates a piece of furniture's data."""
        if furniture_id in self.furniture:
            item = self.furniture[furniture_id]
            item.update(data)
            self._save_data()
            return True
        return False

    def list_students(self):
        """Returns a list of student data dictionaries."""
        # Sort students by last name, then first name
        return sorted(self.students.values(), key=lambda s: (s.get('last_name', '').lower(), s.get('first_name', '').lower()))

    def get_student(self, student_id):
        """Retrieves a single student's data."""
        return self.students.get(student_id)

    def update_student(self, student_id, data):
        """Updates a student's data."""
        if student_id in self.students:
            student = self.students[student_id]
            student.update(data)
            self._save_data()
            return True
        return False
