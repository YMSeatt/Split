import tkinter as tk
from tkinter import ttk, simpledialog, messagebox, filedialog, colorchooser, font as tkfont
import json
import os
import sys
import subprocess
from datetime import datetime, timedelta, date as datetime_date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font as OpenpyxlFont, Alignment as OpenpyxlAlignment
from openpyxl.utils import get_column_letter
import re
import shutil
import zipfile
import csv
import hashlib # For password hashing
import PIL
from PIL import Image 
try: 
    import sv_ttk # For themed widgets
    import darkdetect # For dark mode detection
    from PIL import ImageTk 
except:
    pass


try:
    import portalocker # For file locking
except ModuleNotFoundError: print("Install portalocker")
# Conditional import for platform-specific screenshot capability
import threading


# def listener(callback: typing.Callable[[str], None]) -> None: ...

# TODO: make conditional formatting work by quizzes. add thing for homework also.



try:
    if sys.platform == "win32":
        import win32gui
        import win32ui
        import win32con
    else:
        # For non-Windows, these modules won't be available or needed for the current screenshot method.
        # The postscript-based method is platform-independent.
        win32gui = None
except ImportError:
    win32gui = None # Explicitly set to None if import fails
    print("Warning: win32gui/win32ui/win32con not found. Full window screenshot (deprecated) might not work if called.")

import io
import tempfile

try:
    from tkcalendar import DateEntry
except ImportError:
    DateEntry = None
    print("Warning: tkcalendar library not found. Date pickers in export filter will be simple text entries.")
    print("Consider installing it: pip install tkcalendar")

# --- Application Constants ---
APP_NAME = "BehaviorLogger"
APP_VERSION = "v52.0" # Version incremented
CURRENT_DATA_VERSION_TAG = "v9" # Incremented for new homework/marks features

# --- Default Configuration ---
DEFAULT_STUDENT_BOX_WIDTH = 130
DEFAULT_STUDENT_BOX_HEIGHT = 80
MIN_STUDENT_BOX_WIDTH = 60
MIN_STUDENT_BOX_HEIGHT = 40
REBBI_DESK_WIDTH = 200
REBBI_DESK_HEIGHT = 100

DEFAULT_FONT_FAMILY = "TkDefaultFont"
DEFAULT_FONT_SIZE = 10
DEFAULT_FONT_COLOR = "black"
DEFAULT_BOX_FILL_COLOR = "skyblue"
DEFAULT_BOX_OUTLINE_COLOR = "blue"
DEFAULT_QUIZ_SCORE_FONT_COLOR = "darkgreen"
DEFAULT_QUIZ_SCORE_FONT_STYLE_BOLD = True
DEFAULT_HOMEWORK_SCORE_FONT_COLOR = "purple" # New for homework scores
DEFAULT_HOMEWORK_SCORE_FONT_STYLE_BOLD = True # New
GROUP_COLOR_INDICATOR_SIZE = 12
DEFAULT_THEME = "System"
THEME_LIST = ["Light", "Dark", "System"]

DRAG_THRESHOLD = 5
DEFAULT_GRID_SIZE = 20
MAX_UNDO_HISTORY_DAYS = 90
LAYOUT_COLLISION_OFFSET = 5
RESIZE_HANDLE_SIZE = 10 # World units for resize handle

# --- Path Handling ---
def get_app_data_path(filename):
    try:
        # Determine base path based on whether the app is frozen (packaged) or running from script
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            # Running as a PyInstaller bundle
            if os.name == 'win32': # Windows
                base_path = os.path.join(os.getenv('APPDATA'), APP_NAME)
            elif sys.platform == 'darwin': # macOS
                base_path = os.path.join(os.path.expanduser('~'), 'Library', 'Application Support', APP_NAME)
            else: # Linux and other Unix-like
                xdg_config_home = os.getenv('XDG_CONFIG_HOME')
                if xdg_config_home:
                    base_path = os.path.join(xdg_config_home, APP_NAME)
                else:
                    base_path = os.path.join(os.path.expanduser('~'), '.config', APP_NAME)
        else:
            # Running as a script, use the script's directory
            base_path = os.path.dirname(os.path.abspath(__file__))

        # Create the base directory if it doesn't exist
        if not os.path.exists(base_path):
            os.makedirs(base_path, exist_ok=True)
        return os.path.join(base_path, filename)
    except Exception as e:
        # Fallback to current working directory if standard paths fail
        print(f"Warning: Could not determine standard app data path due to {e}. Using current working directory as fallback.")
        fallback_path = os.path.join(os.getcwd(), APP_NAME) # Create a subfolder in CWD
        if not os.path.exists(fallback_path):
             os.makedirs(fallback_path, exist_ok=True)
        return os.path.join(fallback_path, filename)

DATA_FILE_PATTERN = f"classroom_data_{CURRENT_DATA_VERSION_TAG}.json"
CUSTOM_BEHAVIORS_FILE_PATTERN = f"custom_behaviors_{CURRENT_DATA_VERSION_TAG}.json"
CUSTOM_HOMEWORKS_FILE_PATTERN = f"custom_homeworks_{CURRENT_DATA_VERSION_TAG}.json" # New
AUTOSAVE_EXCEL_FILE_PATTERN = f"autosave_log_{CURRENT_DATA_VERSION_TAG}.xlsx" # Renamed for clarity
LAYOUT_TEMPLATES_DIR_NAME = "layout_templates"
STUDENT_GROUPS_FILE_PATTERN = f"student_groups_{CURRENT_DATA_VERSION_TAG}.json"
QUIZ_TEMPLATES_FILE_PATTERN = f"quiz_templates_{CURRENT_DATA_VERSION_TAG}.json"
HOMEWORK_TEMPLATES_FILE_PATTERN = f"homework_templates_{CURRENT_DATA_VERSION_TAG}.json" # New

DATA_FILE = get_app_data_path(DATA_FILE_PATTERN)
CUSTOM_BEHAVIORS_FILE = get_app_data_path(CUSTOM_BEHAVIORS_FILE_PATTERN)
CUSTOM_HOMEWORKS_FILE = get_app_data_path(CUSTOM_HOMEWORKS_FILE_PATTERN) # New
AUTOSAVE_EXCEL_FILE = get_app_data_path(AUTOSAVE_EXCEL_FILE_PATTERN)
LAYOUT_TEMPLATES_DIR = get_app_data_path(LAYOUT_TEMPLATES_DIR_NAME)
STUDENT_GROUPS_FILE = get_app_data_path(STUDENT_GROUPS_FILE_PATTERN)
QUIZ_TEMPLATES_FILE = get_app_data_path(QUIZ_TEMPLATES_FILE_PATTERN)
HOMEWORK_TEMPLATES_FILE = get_app_data_path(HOMEWORK_TEMPLATES_FILE_PATTERN) # New
LOCK_FILE_PATH = get_app_data_path(f"{APP_NAME}.lock") # Lock file

if not os.path.exists(LAYOUT_TEMPLATES_DIR):
    os.makedirs(LAYOUT_TEMPLATES_DIR, exist_ok=True)

DEFAULT_BEHAVIORS_LIST = [
    "Talking", "Off Task", "Out of Seat", "Uneasy", "Placecheck",
    "Great Participation", "Called On", "Complimented", "Fighting", "Other"
]


# for manual detailed logging, i would want these | currently used for sessions (yes/no)
DEFAULT_HOMEWORK_TYPES_LIST = [ # For live session "Yes/No" mode
    "Reading Assignment", "Worksheet", "Math Problems", "Project Work", "Study for Test"
]


# these would only be for behavior-like logging (not detailed)
DEFAULT_HOMEWORK_LOG_BEHAVIORS = [ # For manual homework logging (like behavior logging)
    "Done", "Not Done", "Partially Done", "Signed", "Returned", "Late", "Excellent Work"
]


# Make these be used for the select AND Yes/No Live Homework session modes.
DEFAULT_HOMEWORK_SESSION_BUTTONS = [ # For live session "Select" mode
    {"name": "Done"}, {"name": "Not Done"}, {"name": "Signed"}, {"name": "Returned"}
]

DEFAULT_HOMEWORK_SESSION_BUTTONS2 = [ # For live session "Select" mode
    "Done", "Not Done", "Signed", "Returned"
]

DEFAULT_GROUP_COLORS = ["#FFADAD", "#FFD6A5", "#FDFFB6", "#CAFFBF", "#9BF6FF", "#A0C4FF", "#BDB2FF", "#FFC6FF", "#E0E0E0"]

DEFAULT_QUIZ_MARK_TYPES = [
    {"id": "mark_correct", "name": "Correct", "contributes_to_total": True, "is_extra_credit": False, "default_points": 1},
    {"id": "mark_incorrect", "name": "Incorrect", "contributes_to_total": True, "is_extra_credit": False, "default_points": 0},
    {"id": "mark_partial", "name": "Partial Credit", "contributes_to_total": True, "is_extra_credit": False, "default_points": 0.5},
    {"id": "extra_credit", "name": "Bonus", "contributes_to_total": False, "is_extra_credit": True, "default_points": 1}
]
DEFAULT_HOMEWORK_MARK_TYPES = [ # New for homework marks
    {"id": "hmark_complete", "name": "Complete", "default_points": 10},
    {"id": "hmark_incomplete", "name": "Incomplete", "default_points": 5},
    {"id": "hmark_notdone", "name": "Not Done", "default_points": 0},
    {"id": "hmark_effort", "name": "Effort Score (1-5)", "default_points": 3} # Example
]

MAX_CUSTOM_TYPES = 90 # Max for custom behaviors, homeworks, mark types

MASTER_RECOVERY_PASSWORD_HASH = "d3c01af653d8940fc36ea1e1f33a8dc03f47dd864d2cd0d8814e2643fa37e70de0a2228e58d7d591eb2f124e2f4f9ff7c98686f4f5da3de6bbfc0267db3c1a0e" # SHA256 of "RecoverMyData123!"
#Recovery1Master2Password!Jaffe1
# --- File Lock Manager ---
class FileLockManager:
    def __init__(self, lock_file_path):
        self.lock_file_path = lock_file_path
        self.lock = None

    def acquire_lock(self):
        try:
            # Open the lock file in exclusive mode, creating it if it doesn't exist.
            # portalocker will raise an exception if the lock cannot be acquired.
            self.lock = open(self.lock_file_path, 'w')
            portalocker.lock(self.lock, portalocker.LOCK_EX | portalocker.LOCK_NB)
            # Write PID to lock file for informational purposes (optional)
            self.lock.write(str(os.getpid()))
            self.lock.flush()
            return True
        except ( IOError, BlockingIOError) as e:
            # Check if lock file exists and try to read PID
            pid_in_lock = None
            if os.path.exists(self.lock_file_path):
                try:
                    with open(self.lock_file_path, 'r') as f_read:
                        pid_in_lock = f_read.read().strip()
                except IOError:
                    pass # Could not read PID
            
            error_message = f"Another instance of {APP_NAME} may already be running."
            if pid_in_lock:
                error_message += f" (PID in lock file: {pid_in_lock})"
            else:
                error_message += " (Could not read PID from lock file)."
            error_message += f"\n\nIf you are sure no other instance is running, you can manually delete the lock file:\n{self.lock_file_path}\n\nError details: {e}"
            
            messagebox.showerror("Application Already Running?", error_message)
            if self.lock: # Close the file handle if it was opened but locking failed
                portalocker.unlock(self.lock)
                self.lock.close()
                self.lock = None
            return False
        
    def release_lock(self):
        if self.lock:
            try:
                portalocker.unlock(self.lock)
                self.lock.close()
            except Exception as e:
                print(f"Warning: Error releasing file lock: {e}")
            finally:
                self.lock = None
                # Attempt to delete the lock file
                try:
                    if os.path.exists(self.lock_file_path):
                        os.remove(self.lock_file_path)
                except OSError as e:
                    print(f"Warning: Could not delete lock file {self.lock_file_path}: {e}")
# --- Password Management ---
class PasswordManager:
    def __init__(self, app_settings):
        self.app_settings = app_settings
        self.is_locked = False
        self.last_activity_time = datetime.now()

    def _hash_password(self, password):
        return hashlib.sha3_512(password.encode('utf-8')).hexdigest()

    def set_password(self, password):
        if not password:
            self.app_settings["app_password_hash"] = None
            return True
        self.app_settings["app_password_hash"] = self._hash_password(password)
        return True

    def check_password(self, password):
        stored_hash = self.app_settings.get("app_password_hash")
        if not stored_hash: return True
        return self._hash_password(password) == stored_hash

    def check_recovery_password(self, recovery_password):
        return self._hash_password(recovery_password) == MASTER_RECOVERY_PASSWORD_HASH

    def is_password_set(self):
        return bool(self.app_settings.get("app_password_hash"))

    def lock_application(self):
        if self.is_password_set():
            self.is_locked = True
            return True
        return False

    def unlock_application(self, password_attempt):
        if self.check_password(password_attempt) or self.check_recovery_password(password_attempt):
            self.is_locked = False
            self.last_activity_time = datetime.now()
            return True
        return False

    def check_auto_lock(self):
        if self.is_locked or not self.is_password_set() or not self.app_settings.get("password_auto_lock_enabled", False):
            return
        timeout_minutes = self.app_settings.get("password_auto_lock_timeout_minutes", 15)
        if timeout_minutes > 0:
            if (datetime.now() - self.last_activity_time).total_seconds() / 60 > timeout_minutes:
                self.lock_application()

    def record_activity(self):
        self.last_activity_time = datetime.now()

# --- Command Pattern for Undo/Redo ---
class Command:
    def __init__(self, app, timestamp=None):
        self.app = app
        self.timestamp = timestamp or datetime.now().isoformat()

    def execute(self): raise NotImplementedError
    def undo(self): raise NotImplementedError
    def to_dict(self): return {'type': self.__class__.__name__, 'timestamp': self.timestamp, 'data': self._get_data_for_serialization()}
    def _get_data_for_serialization(self): raise NotImplementedError
    @classmethod
    def from_dict(cls, app, data_dict):
        command_type_name = data_dict['type']
        command_class = getattr(sys.modules[__name__], command_type_name, None)
        if command_class and issubclass(command_class, Command):
            try:
                return command_class._from_serializable_data(app, data_dict['data'], data_dict['timestamp'])
            except KeyError as e:
                print(f"Warning: Missing key '{e}' in data for command type '{command_type_name}'. Skipping command.")
                return None
        print(f"Warning: Unknown command type '{command_type_name}' in undo/redo history.")
        return None
    @classmethod
    def _from_serializable_data(cls, app, data, timestamp): raise NotImplementedError

class MoveItemsCommand(Command):
    def __init__(self, app, items_moves, timestamp=None):
        super().__init__(app, timestamp)
        self.items_moves = items_moves # List of dicts: {'id', 'type', 'old_x', 'old_y', 'new_x', 'new_y'}

    def execute(self):
        for item_move in self.items_moves:
            item_id, item_type, new_x, new_y = item_move['id'], item_move['type'], item_move['new_x'], item_move['new_y']
            data_source = self.app.students if item_type == 'student' else self.app.furniture
            if item_id in data_source:
                data_source[item_id]['x'] = new_x
                data_source[item_id]['y'] = new_y
        self.app.update_status(f"Moved {len(self.items_moves)} item(s).")
        self.app.draw_all_items(check_collisions_on_redraw=True)

    def undo(self):
        for item_move in self.items_moves:
            item_id, item_type, old_x, old_y = item_move['id'], item_move['type'], item_move['old_x'], item_move['old_y']
            data_source = self.app.students if item_type == 'student' else self.app.furniture
            if item_id in data_source:
                data_source[item_id]['x'] = old_x
                data_source[item_id]['y'] = old_y
        self.app.update_status(f"Undid move of {len(self.items_moves)} item(s).")
        self.app.draw_all_items(check_collisions_on_redraw=True)

    def _get_data_for_serialization(self): return {'items_moves': self.items_moves}
    @classmethod
    def _from_serializable_data(cls, app, data, timestamp): return cls(app, data['items_moves'], timestamp)

class AddItemCommand(Command):
    def __init__(self, app, item_id, item_type, item_data, old_next_id_num, timestamp=None):
        super().__init__(app, timestamp)
        self.item_id = item_id
        self.item_type = item_type
        self.item_data = item_data
        self.old_next_id_num = old_next_id_num

    def execute(self):
        data_source = self.app.students if self.item_type == 'student' else self.app.furniture
        data_source[self.item_id] = self.item_data.copy()
        if self.item_type == 'student':
            self.app.next_student_id_num = self.item_data.get('original_next_id_num_after_add', self.app.next_student_id_num)
            self.app.update_student_display_text(self.item_id)
            self.app.update_status(f"Student '{self.item_data['full_name']}' added.")
        else:
            self.app.next_furniture_id_num = self.item_data.get('original_next_id_num_after_add', self.app.next_furniture_id_num)
            self.app.update_status(f"Furniture '{self.item_data['name']}' added.")
        self.app.draw_all_items(check_collisions_on_redraw=True)

    def undo(self):
        data_source = self.app.students if self.item_type == 'student' else self.app.furniture
        if self.item_id in data_source:
            item_name = data_source[self.item_id].get('full_name', data_source[self.item_id].get('name'))
            del data_source[self.item_id]
            self.app.canvas.delete(self.item_id)
            if self.item_type == 'student':
                self.app.next_student_id_num = self.old_next_id_num
                self.app.update_status(f"Undid add of student '{item_name}'.")
            else:
                self.app.next_furniture_id_num = self.old_next_id_num
                self.app.update_status(f"Undid add of furniture '{item_name}'.")
        self.app.draw_all_items(check_collisions_on_redraw=True)

    def _get_data_for_serialization(self): return {'item_id': self.item_id, 'item_type': self.item_type, 'item_data': self.item_data, 'old_next_id_num': self.old_next_id_num}
    @classmethod
    def _from_serializable_data(cls, app, data, timestamp): return cls(app, data['item_id'], data['item_type'], data['item_data'], data['old_next_id_num'], timestamp)

class DeleteItemCommand(Command):
    def __init__(self, app, item_id, item_type, item_data, associated_logs=None, timestamp=None):
        super().__init__(app, timestamp)
        self.item_id = item_id
        self.item_type = item_type
        self.item_data = item_data
        self.associated_logs = associated_logs or [] # For behavior and quiz logs
        self.associated_homework_logs = [] # New for homework logs

        if item_type == 'student': # Separate homework logs for students
            self.associated_homework_logs = [log.copy() for log in app.homework_log if log["student_id"] == item_id]


    def execute(self):
        data_source = self.app.students if self.item_type == 'student' else self.app.furniture
        item_name = "Item"
        if self.item_id in data_source:
            item_name = data_source[self.item_id].get('full_name', data_source[self.item_id].get('name'))
            del data_source[self.item_id]
        self.app.canvas.delete(self.item_id)
        if self.item_id in self.app.selected_items: self.app.selected_items.remove(self.item_id)

        if self.item_type == 'student':
            original_log_count = len(self.app.behavior_log)
            self.app.behavior_log = [log for log in self.app.behavior_log if log["student_id"] != self.item_id]
            logs_removed_count = original_log_count - len(self.app.behavior_log)

            original_homework_log_count = len(self.app.homework_log)
            self.app.homework_log = [log for log in self.app.homework_log if log["student_id"] != self.item_id]
            homework_logs_removed_count = original_homework_log_count - len(self.app.homework_log)

            self.app.update_status(f"Student '{item_name}', {logs_removed_count} behavior/quiz log(s), and {homework_logs_removed_count} homework log(s) deleted.")
        else:
            self.app.update_status(f"Furniture '{item_name}' deleted.")
        self.app.draw_all_items(check_collisions_on_redraw=True)

    def undo(self):
        data_source = self.app.students if self.item_type == 'student' else self.app.furniture
        data_source[self.item_id] = self.item_data.copy()
        if self.item_type == 'student':
            self.app.update_student_display_text(self.item_id)
            for log_entry in self.associated_logs:
                if log_entry not in self.app.behavior_log: self.app.behavior_log.append(log_entry.copy())
            self.app.behavior_log.sort(key=lambda x: x.get("timestamp", ""))

            for hw_log_entry in self.associated_homework_logs: # Restore homework logs
                if hw_log_entry not in self.app.homework_log: self.app.homework_log.append(hw_log_entry.copy())
            self.app.homework_log.sort(key=lambda x: x.get("timestamp", ""))

            self.app.update_status(f"Undid delete of student '{self.item_data['full_name']}'. Logs restored.")
        else:
            self.app.update_status(f"Undid delete of furniture '{self.item_data['name']}'.")
        self.app.draw_all_items(check_collisions_on_redraw=True)

    def _get_data_for_serialization(self):
        return {
            'item_id': self.item_id, 'item_type': self.item_type,
            'item_data': self.item_data, 'associated_logs': self.associated_logs,
            'associated_homework_logs': self.associated_homework_logs # Serialize homework logs
        }
    @classmethod
    def _from_serializable_data(cls, app, data, timestamp):
        cmd = cls(app, data['item_id'], data['item_type'], data['item_data'], data.get('associated_logs'), timestamp)
        cmd.associated_homework_logs = data.get('associated_homework_logs', []) # Deserialize homework logs
        return cmd

class LogEntryCommand(Command): # For Behavior and Quiz logs
    def __init__(self, app, log_entry, student_id, timestamp=None):
        super().__init__(app, timestamp)
        self.log_entry = log_entry
        self.student_id = student_id

    def execute(self):
        # Behavior/Quiz logs go into self.app.behavior_log
        if not any(le == self.log_entry for le in self.app.behavior_log):
            self.app.behavior_log.append(self.log_entry.copy())
            self.app.behavior_log.sort(key=lambda x: x.get("timestamp", ""))
        self.app.update_student_display_text(self.student_id)
        log_type = self.log_entry.get("type", "behavior")
        behavior_name = self.log_entry.get("behavior", "Unknown")
        student_name = self.app.students.get(self.student_id, {}).get('full_name', 'Unknown Student')
        self.app.update_status(f"{log_type.capitalize()} '{behavior_name}' logged for {student_name}.")

    def undo(self):
        try:
            self.app.behavior_log.remove(self.log_entry)
        except ValueError:
            for i, entry in enumerate(self.app.behavior_log):
                if entry["timestamp"] == self.log_entry["timestamp"] and \
                   entry["student_id"] == self.log_entry["student_id"] and \
                   entry["behavior"] == self.log_entry["behavior"]:
                    del self.app.behavior_log[i]; break
        self.app.update_student_display_text(self.student_id)
        log_type = self.log_entry.get("type", "behavior")
        behavior_name = self.log_entry.get("behavior", "Unknown")
        student_name = self.app.students.get(self.student_id, {}).get('full_name', 'Unknown Student')
        self.app.update_status(f"Undid log of {log_type.capitalize()} '{behavior_name}' for {student_name}.")

    def _get_data_for_serialization(self): return {'log_entry': self.log_entry, 'student_id': self.student_id}
    @classmethod
    def _from_serializable_data(cls, app, data, timestamp): return cls(app, data['log_entry'], data['student_id'], timestamp)

class LogHomeworkEntryCommand(Command): # New for Homework logs
    def __init__(self, app, log_entry, student_id, timestamp=None):
        super().__init__(app, timestamp)
        self.log_entry = log_entry
        self.student_id = student_id

    def execute(self):
        # Homework logs go into self.app.homework_log
        if not any(le == self.log_entry for le in self.app.homework_log):
            self.app.homework_log.append(self.log_entry.copy())
            self.app.homework_log.sort(key=lambda x: x.get("timestamp", ""))
        self.app.update_student_display_text(self.student_id) # Redraw student box
        homework_name = self.log_entry.get("homework_type", self.log_entry.get("behavior", "Unknown Homework")) # Use "homework_type" or "behavior"
        student_name = self.app.students.get(self.student_id, {}).get('full_name', 'Unknown Student')
        self.app.update_status(f"Homework '{homework_name}' logged for {student_name}.")

    def undo(self):
        try:
            self.app.homework_log.remove(self.log_entry)
        except ValueError:
            for i, entry in enumerate(self.app.homework_log):
                 # Match based on key fields for homework
                if entry["timestamp"] == self.log_entry["timestamp"] and \
                   entry["student_id"] == self.log_entry["student_id"] and \
                   entry.get("homework_type", entry.get("behavior")) == self.log_entry.get("homework_type", self.log_entry.get("behavior")):
                    del self.app.homework_log[i]; break
        self.app.update_student_display_text(self.student_id)
        homework_name = self.log_entry.get("homework_type", self.log_entry.get("behavior", "Unknown Homework"))
        student_name = self.app.students.get(self.student_id, {}).get('full_name', 'Unknown Student')
        self.app.update_status(f"Undid log of homework '{homework_name}' for {student_name}.")

    def _get_data_for_serialization(self): return {'log_entry': self.log_entry, 'student_id': self.student_id}
    @classmethod
    def _from_serializable_data(cls, app, data, timestamp): return cls(app, data['log_entry'], data['student_id'], timestamp)


class EditItemCommand(Command):
    def __init__(self, app, item_id, item_type, old_item_data, new_item_data_changes, timestamp=None):
        super().__init__(app, timestamp)
        self.item_id = item_id
        self.item_type = item_type
        self.old_item_data_snapshot = old_item_data
        self.new_item_data_changes = new_item_data_changes

    def execute(self):
        data_source = self.app.students if self.item_type == 'student' else self.app.furniture
        if self.item_id in data_source:
            data_source[self.item_id].update(self.new_item_data_changes)
            if self.item_type == 'student':
                self.app.update_student_display_text(self.item_id)
                self.app.update_status(f"Student '{data_source[self.item_id]['full_name']}' edited.")
            else:
                self.app.update_status(f"Furniture '{data_source[self.item_id]['name']}' edited.")
        self.app.draw_all_items(check_collisions_on_redraw=True)

    def undo(self):
        data_source = self.app.students if self.item_type == 'student' else self.app.furniture
        if self.item_id in data_source:
            data_source[self.item_id] = self.old_item_data_snapshot.copy()
            if self.item_type == 'student':
                self.app.update_student_display_text(self.item_id)
                self.app.update_status(f"Undid edit for student '{data_source[self.item_id]['full_name']}'.")
            else:
                self.app.update_status(f"Undid edit for furniture '{data_source[self.item_id]['name']}'.")
        self.app.draw_all_items(check_collisions_on_redraw=True)

    def _get_data_for_serialization(self):
        return {
            'item_id': self.item_id, 'item_type': self.item_type,
            'old_item_data_snapshot': self.old_item_data_snapshot,
            'new_item_data_changes': self.new_item_data_changes
        }
    @classmethod
    def _from_serializable_data(cls, app, data, timestamp):
        return cls(app, data['item_id'], data['item_type'],
                   data['old_item_data_snapshot'], data['new_item_data_changes'], timestamp)

class ChangeItemsSizeCommand(Command):
    def __init__(self, app, items_sizes_changes, timestamp=None):
        super().__init__(app, timestamp)
        self.items_sizes_changes = items_sizes_changes

    def _apply_sizes(self, use_new_sizes):
        changed_item_names = []
        for item_size_info in self.items_sizes_changes:
            item_id, item_type = item_size_info['id'], item_size_info['type']
            w = item_size_info['new_w'] if use_new_sizes else item_size_info['old_w']
            h = item_size_info['new_h'] if use_new_sizes else item_size_info['old_h']
            data_source = self.app.students if item_type == 'student' else self.app.furniture
            if item_id in data_source:
                if item_type == 'student':
                    if 'style_overrides' not in data_source[item_id]: data_source[item_id]['style_overrides'] = {}
                    data_source[item_id]['style_overrides']['width'] = w
                    data_source[item_id]['style_overrides']['height'] = h
                    data_source[item_id]['width'] = w # Keep base in sync
                    data_source[item_id]['height'] = h
                else:
                    data_source[item_id]['width'] = w
                    data_source[item_id]['height'] = h
                changed_item_names.append(data_source[item_id].get('full_name', data_source[item_id].get('name', item_id)))
        return changed_item_names

    def execute(self):
        names = self._apply_sizes(use_new_sizes=True)
        self.app.update_status(f"Size changed for {len(names)} item(s): {', '.join(names[:3])}{'...' if len(names)>3 else ''}.")
        self.app.draw_all_items(check_collisions_on_redraw=True)

    def undo(self):
        names = self._apply_sizes(use_new_sizes=False)
        self.app.update_status(f"Undid size change for {len(names)} item(s): {', '.join(names[:3])}{'...' if len(names)>3 else ''}.")
        self.app.draw_all_items(check_collisions_on_redraw=True)

    def _get_data_for_serialization(self): return {'items_sizes_changes': self.items_sizes_changes}
    @classmethod
    def _from_serializable_data(cls, app, data, timestamp): return cls(app, data['items_sizes_changes'], timestamp)

class MarkLiveQuizQuestionCommand(Command):
    def __init__(self, app, student_id, action_taken, timestamp=None):
        super().__init__(app, timestamp)
        self.student_id = student_id
        self.action_taken = action_taken
        self.previous_student_score_state = None

    def execute(self):
        if self.previous_student_score_state is None:
            self.previous_student_score_state = self.app.live_quiz_scores.get(self.student_id, {"correct": 0, "total_asked": 0}).copy()
        current_score = self.app.live_quiz_scores.get(self.student_id, {"correct": 0, "total_asked": 0}).copy()
        current_score["total_asked"] += 1
        if self.action_taken == "correct": current_score["correct"] += 1
        self.app.live_quiz_scores[self.student_id] = current_score
        self.app.draw_single_student(self.student_id)
        student_name = self.app.students[self.student_id]['full_name']
        self.app.update_status(f"Live Quiz: '{self.action_taken.capitalize()}' for {student_name}. Score: {current_score['correct']}/{current_score['total_asked']}")

    def undo(self):
        if self.previous_student_score_state is not None:
            self.app.live_quiz_scores[self.student_id] = self.previous_student_score_state.copy()
            if self.app.live_quiz_scores[self.student_id]["total_asked"] == 0 and self.app.live_quiz_scores[self.student_id]["correct"] == 0:
                del self.app.live_quiz_scores[self.student_id]
        elif self.student_id in self.app.live_quiz_scores:
            current_score = self.app.live_quiz_scores[self.student_id]
            current_score["total_asked"] -= 1
            if self.action_taken == "correct": current_score["correct"] -= 1
            if current_score["total_asked"] <= 0: del self.app.live_quiz_scores[self.student_id]
        self.app.draw_single_student(self.student_id)
        student_name = self.app.students[self.student_id]['full_name']
        score_info = self.app.live_quiz_scores.get(self.student_id)
        status = f"Undo Live Quiz Mark for {student_name}. Score: {score_info['correct']}/{score_info['total_asked']}" if score_info else f"Undo Live Quiz Mark for {student_name}. No questions marked."
        self.app.update_status(status)

    def _get_data_for_serialization(self):
        return {'student_id': self.student_id, 'action_taken': self.action_taken, 'previous_student_score_state': self.previous_student_score_state}
    @classmethod
    def _from_serializable_data(cls, app, data, timestamp):
        cmd = cls(app, data['student_id'], data['action_taken'], timestamp)
        cmd.previous_student_score_state = data.get('previous_student_score_state')
        return cmd

class MarkLiveHomeworkCommand(Command): # New for Live Homework
    def __init__(self, app, student_id, homework_actions, session_mode, timestamp=None):
        super().__init__(app, timestamp)
        self.student_id = student_id
        self.homework_actions = homework_actions # Dict: {"homework_type_name": "yes/no/selected_option"} or list of selected options
        self.session_mode = session_mode # "Yes/No" or "Select"
        self.previous_homework_state = None # To store previous state for undo

    def execute(self):
        if self.previous_homework_state is None:
            self.previous_homework_state = self.app.live_homework_scores.get(self.student_id, {}).copy()

        # Update live_homework_scores based on homework_actions and session_mode
        # For "Yes/No", homework_actions might be {"Reading": "yes", "Math": "no"}
        # For "Select", homework_actions might be ["Done", "Signed"]
        current_hw_data = self.app.live_homework_scores.get(self.student_id, {}).copy()
        if self.session_mode == "Yes/No":
            current_hw_data.update(self.homework_actions) # Overwrite/add specific yes/no statuses
        elif self.session_mode == "Select":
            # Store as a list of selected actions for this student for this session
            current_hw_data["selected_options"] = list(self.homework_actions) # Ensure it's a list

        self.app.live_homework_scores[self.student_id] = current_hw_data
        self.app.draw_single_student(self.student_id) # Redraw to update display
        student_name = self.app.students[self.student_id]['full_name']
        self.app.update_status(f"Live Homework updated for {student_name}.")

    def undo(self):
        if self.previous_homework_state is not None:
            self.app.live_homework_scores[self.student_id] = self.previous_homework_state.copy()
            if not self.app.live_homework_scores[self.student_id]: # If restored to empty
                del self.app.live_homework_scores[self.student_id]
        elif self.student_id in self.app.live_homework_scores: # Should not happen if previous_homework_state was set
            del self.app.live_homework_scores[self.student_id]

        self.app.draw_single_student(self.student_id)
        student_name = self.app.students[self.student_id]['full_name']
        self.app.update_status(f"Undo Live Homework update for {student_name}.")

    def _get_data_for_serialization(self):
        return {
            'student_id': self.student_id,
            'homework_actions': self.homework_actions,
            'session_mode': self.session_mode,
            'previous_homework_state': self.previous_homework_state
        }
    @classmethod
    def _from_serializable_data(cls, app, data, timestamp):
        cmd = cls(app, data['student_id'], data['homework_actions'], data['session_mode'], timestamp)
        cmd.previous_homework_state = data.get('previous_homework_state')
        return cmd


class ChangeStudentStyleCommand(Command):
    def __init__(self, app, student_id, style_property, old_value, new_value, timestamp=None):
        super().__init__(app, timestamp)
        self.student_id = student_id
        self.style_property = style_property
        self.old_value = old_value
        self.new_value = new_value

    def execute(self):
        student = self.app.students.get(self.student_id)
        if student:
            if "style_overrides" not in student: student["style_overrides"] = {}
            if self.new_value is None:
                if self.style_property in student["style_overrides"]: del student["style_overrides"][self.style_property]
            else: student["style_overrides"][self.style_property] = self.new_value
            self.app.update_student_display_text(self.student_id)
            self.app.draw_single_student(self.student_id, check_collisions=True)
            self.app.update_status(f"Style '{self.style_property}' updated for {student['full_name']}.")

    def undo(self):
        student = self.app.students.get(self.student_id)
        if student:
            if "style_overrides" not in student: student["style_overrides"] = {}
            if self.old_value is None:
                if self.style_property in student["style_overrides"]: del student["style_overrides"][self.style_property]
            else: student["style_overrides"][self.style_property] = self.old_value
            if not student["style_overrides"]: del student["style_overrides"]
            self.app.update_student_display_text(self.student_id)
            self.app.draw_single_student(self.student_id, check_collisions=True)
            self.app.update_status(f"Undid style '{self.style_property}' change for {student['full_name']}.")

    def _get_data_for_serialization(self):
        return {'student_id': self.student_id, 'style_property': self.style_property, 'old_value': self.old_value, 'new_value': self.new_value}
    @classmethod
    def _from_serializable_data(cls, app, data, timestamp):
        return cls(app, data['student_id'], data['style_property'], data['old_value'], data['new_value'], timestamp)

class ManageStudentGroupCommand(Command):
    def __init__(self, app, old_groups_snapshot, new_groups_snapshot,
                 old_student_group_assignments, new_student_group_assignments,
                 old_next_group_id_num, new_next_group_id_num, timestamp=None):
        super().__init__(app, timestamp)
        self.old_groups_snapshot = old_groups_snapshot
        self.new_groups_snapshot = new_groups_snapshot
        self.old_student_group_assignments = old_student_group_assignments
        self.new_student_group_assignments = new_student_group_assignments
        self.old_next_group_id_num = old_next_group_id_num
        self.new_next_group_id_num = new_next_group_id_num

    def execute(self):
        self.app.student_groups = self.new_groups_snapshot.copy()
        for student_id, group_id in self.new_student_group_assignments.items():
            if student_id in self.app.students: self.app.students[student_id]['group_id'] = group_id
        for student_id in self.app.students:
            if student_id not in self.new_student_group_assignments and 'group_id' in self.app.students[student_id]:
                del self.app.students[student_id]['group_id']
        self.app.next_group_id_num = self.new_next_group_id_num
        self.app.settings["next_group_id_num"] = self.new_next_group_id_num
        self.app.save_student_groups()
        self.app.draw_all_items(check_collisions_on_redraw=True)
        self.app.update_status("Student groups updated.")

    def undo(self):
        self.app.student_groups = self.old_groups_snapshot.copy()
        for student_id, group_id in self.old_student_group_assignments.items():
            if student_id in self.app.students: self.app.students[student_id]['group_id'] = group_id
        for student_id in self.app.students:
            if student_id not in self.old_student_group_assignments and 'group_id' in self.app.students[student_id]:
                del self.app.students[student_id]['group_id']
        self.app.next_group_id_num = self.old_next_group_id_num
        self.app.settings["next_group_id_num"] = self.old_next_group_id_num
        self.app.save_student_groups()
        self.app.draw_all_items(check_collisions_on_redraw=True)
        self.app.update_status("Student group update undone.")

    def _get_data_for_serialization(self):
        return {
            'old_groups_snapshot': self.old_groups_snapshot, 'new_groups_snapshot': self.new_groups_snapshot,
            'old_student_group_assignments': self.old_student_group_assignments, 'new_student_group_assignments': self.new_student_group_assignments,
            'old_next_group_id_num': self.old_next_group_id_num, 'new_next_group_id_num': self.new_next_group_id_num,
        }
    @classmethod
    def _from_serializable_data(cls, app, data, timestamp):
        return cls(app, data['old_groups_snapshot'], data['new_groups_snapshot'],
                   data['old_student_group_assignments'], data['new_student_group_assignments'],
                   data['old_next_group_id_num'], data['new_next_group_id_num'], timestamp)




# --- Main Application Class ---
class SeatingChartApp:
    def __init__(self, root_window):
        self.root = root_window
        self.root.title(f"Classroom Behavior Tracker - {APP_NAME} - {APP_VERSION}")
        self.root.geometry("1380x980") # Adjusted for new mode button

        self.file_lock_manager = FileLockManager(LOCK_FILE_PATH)
        if not self.file_lock_manager.acquire_lock():
            self.root.destroy() # Exit if lock cannot be acquired
            sys.exit(1)


        self.students = {}
        self.furniture = {}
        self.behavior_log = [] # For behavior and quiz logs
        self.homework_log = [] # New for homework logs
        self.student_groups = {}
        self.quiz_templates = {}
        self.homework_templates = {} # New

        self.next_student_id_num = 1
        self.next_furniture_id_num = 1
        self.next_group_id_num = 1
        self.next_quiz_template_id_num = 1
        self.next_homework_template_id_num = 1 # New

        self.all_behaviors = []
        self.custom_behaviors = []
        self.all_homework_log_behaviors = [] # For manual homework logging options
        self.custom_homework_log_behaviors = []
        self.all_homework_session_types = [] # For live homework "Yes/No" mode options
        self.custom_homework_session_types = []


        self.last_excel_export_path = None
        self.selected_items = set()
        self.undo_stack = []
        self.redo_stack = []
        try:
            self.theme_style_using = sv_ttk.get_theme() #"Light"
        except:
            self.theme_style_using = "System"
        #print(self.theme_style_using, "Ini")

        self.settings = self._get_default_settings()
        self.password_manager = PasswordManager(self.settings)

        self.canvas_frame = None; self.canvas = None; self.h_scrollbar = None; self.v_scrollbar = None
        self.status_bar_label = None; self.zoom_display_label = None
        self.mode_var = tk.StringVar(value=self.settings["current_mode"])
        self.edit_mode_var = tk.BooleanVar(value=False)

        self.drag_data = {"x": 0, "y": 0, "item_id": None, "item_type": None,
                          "start_x_world": 0, "start_y_world": 0, # Start of drag in world coords
                          "original_positions": {}, "is_resizing": False, "resize_handle_corner": None,
                          "original_size_world": {}} # Original W/H in world coords
        self._potential_click_target = None
        self._drag_started_on_item = False
        self._recent_incidents_hidden_globally = False
        self._recent_homeworks_hidden_globally = False # New
        self._per_student_last_cleared = {}

        self.last_used_quiz_name = ""
        self.initial_num_questions = "" # For quiz
        self.last_used_quiz_name_timestamp = None
        self.last_used_homework_name = "" # New for homework
        self.initial_num_homework_items = "" # New for homework
        self.last_used_homework_name_timestamp = None # New

        self.is_live_quiz_active = False
        self.current_live_quiz_name = ""
        self.live_quiz_scores = {}

        self.is_live_homework_active = False # New
        self.current_live_homework_name = "" # New
        self.live_homework_scores = {} # New {student_id: {"homework_type_name": "yes/no/selected_option", ...} or {"selected_options": [...]}}

        self.current_zoom_level = 1.0
        self.canvas_orig_width = 2000
        self.canvas_orig_height = 1500
        self.custom_canvas_color = None
        try:
            if self.custom_canvas_color != None and self.custom_canvas_color != "Default" and self.custom_canvas_color != "": self.canvas_color = self.custom_canvas_color
            elif self.theme_style_using == "dark": self.canvas_color = "#1F1F1F"
            elif self.theme_style_using == "System": self.canvas_color = "lightgrey" if darkdetect.theme() == "Light" else "#1F1F1F"
            else: self.canvas_color = "lightgrey"
        #self.canvas_color = 
        except:
            if self.custom_canvas_color != None and self.custom_canvas_color != "Default" and self.custom_canvas_color != "": self.canvas_color = self.custom_canvas_color
            elif self.theme_style_using == "dark": self.canvas_color = "#1F1F1F"
            #elif self.theme_style_using == "System": self.canvas_color = "lightgrey" if darkdetect.theme() == "Light" else "#1F1F1F"
            else: self.canvas_color = "lightgrey"
            
        self.load_custom_behaviors()
        self.load_custom_homework_log_behaviors() # New
        self.load_custom_homework_session_types() # New
        self.load_student_groups()
        self.load_quiz_templates()
        self.load_homework_templates() # New
        self.update_all_behaviors()
        self.update_all_homework_log_behaviors() # New
        self.update_all_homework_session_types() # New

        self.load_data() # Loads main data, including settings
        self._ensure_next_ids()

        self.setup_ui()
        self.draw_all_items()
        self.update_status(f"Application started. Data loaded from: {os.path.dirname(DATA_FILE)}")
        self.update_undo_redo_buttons_state()
        self.toggle_mode() # Apply initial mode

        self.root.after(30000, self.periodic_checks)
        self.root.after(self.settings.get("autosave_interval_ms", 30000), self.autosave_data_wrapper)
        self.root.protocol("WM_DELETE_WINDOW", self.on_exit_protocol)

        if self.password_manager.is_password_set() and self.settings.get("password_on_open", False):
            self.root.withdraw()
            if not self.prompt_for_password("Application Locked", "Enter password to open:"):
                self.on_exit_protocol(force_quit=True) # Ensure lock is released if exit fails here
            self.root.deiconify()



    def capture_tkinter_window(self, filename="tkinter_screenshot.png"):
        """
        Captures the content of a specific Tkinter window by its handle.
        This method is now deprecated in favor of export_layout_as_image for full canvas capture.
        """
        root_window = self.root
        hwnd = root_window.winfo_id() # Get the window handle (HWND) of the Tkinter root

        # Get the window dimensions
        left, top, right, bot = win32gui.GetWindowRect(hwnd)
        width = right - left
        height = bot - top

        # Create a device context (DC) for the window
        hdc = win32gui.GetWindowDC(hwnd)
        dc_obj = win32ui.CreateDCFromHandle(hdc)
        mem_dc = dc_obj.CreateCompatibleDC()

        # Create a bitmap object
        bitmap = win32ui.CreateBitmap()
        bitmap.CreateCompatibleBitmap(dc_obj, width, height)
        mem_dc.SelectObject(bitmap)

        # Copy the window content to the bitmap
        # Note: PrintWindow is generally better as it can capture minimized/obscured windows,
        # but GetWindowDC + BitBlt works well for visible windows.
        # For PrintWindow, you might need to find the exact flags.
        mem_dc.BitBlt((0, 0), (width, height), dc_obj, (0, 0), win32con.SRCCOPY)

        # Get the bitmap bits and create a PIL Image
        bmpinfo = bitmap.GetInfo()
        bmpstr = bitmap.GetBitmapBits(True)
        img = Image.frombuffer(
            'RGB',
            (bmpinfo['bmWidth'], bmpinfo['bmHeight']),
            bmpstr, 'raw', 'BGRX', 0, 1)

        # Clean up the DCs
        dc_obj.DeleteDC()
        mem_dc.DeleteDC()
        win32gui.ReleaseDC(hwnd, hdc)

        #print(os.listdir(os.path.dirname(get_app_data_path(filename))))
        default_filename = f"app_screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        file_path = filedialog.asksaveasfilename(defaultextension=".png", initialfile=default_filename + ".png",
                                                       filetypes=[("Image files", "*.png"), ("All Files", "*.*")], parent=self.root)
        
        #if filename in os.listdir(os.path.dirname(get_app_data_path(filename))): print(filename, 1); filename = "tkinter_screenshot3.png"; print(filename, 2) 
            
        
        #print(path)
        
        img.save(file_path)
        print(f"Screenshot saved to {filename}")
        self.update_status(f"Screenshot saved to {file_path}")

    def _get_default_settings(self):
        return {
            "show_recent_incidents_on_boxes": True,
            "num_recent_incidents_to_show": 2,
            "recent_incident_time_window_hours": 24,
            "show_full_recent_incidents": False,
            "reverse_incident_order": True,
            "selected_recent_behaviors_filter": None, # List of behavior names, or None for all

            "show_recent_homeworks_on_boxes": True, # New
            "num_recent_homeworks_to_show": 2, # New
            "recent_homework_time_window_hours": 24, # New
            "show_full_recent_homeworks": False, # New
            "reverse_homework_order": True, # New
            "selected_recent_homeworks_filter": None, # New

            "autosave_interval_ms": 30000,
            "default_student_box_width": DEFAULT_STUDENT_BOX_WIDTH,
            "default_student_box_height": DEFAULT_STUDENT_BOX_HEIGHT,
            "student_box_fill_color": DEFAULT_BOX_FILL_COLOR,
            "student_box_outline_color": DEFAULT_BOX_OUTLINE_COLOR,
            "student_font_family": DEFAULT_FONT_FAMILY,
            "student_font_size": DEFAULT_FONT_SIZE,
            "student_font_color": DEFAULT_FONT_COLOR,
            "grid_snap_enabled": False,
            "grid_size": DEFAULT_GRID_SIZE,
            "behavior_initial_overrides": {},
            "homework_initial_overrides": {}, # New for homework display initials
            "current_mode": "behavior", # "behavior", "quiz", or "homework"
            "max_undo_history_days": MAX_UNDO_HISTORY_DAYS,
            "conditional_formatting_rules": [],
            "student_groups_enabled": True,
            "show_zoom_level_display": True,
            "available_fonts": sorted(list(tkfont.families())),

            # Quiz specific
            "default_quiz_name": "Pop Quiz",
            "last_used_quiz_name_timeout_minutes": 60,
            "show_recent_incidents_during_quiz": True,
            "live_quiz_score_font_color": DEFAULT_QUIZ_SCORE_FONT_COLOR,
            "live_quiz_score_font_style_bold": DEFAULT_QUIZ_SCORE_FONT_STYLE_BOLD,
            "quiz_mark_types": DEFAULT_QUIZ_MARK_TYPES.copy(),
            "default_quiz_questions": 10,
            "quiz_score_calculation": "percentage",
            "combine_marks_for_display": True,

            # Homework specific (New)
            "default_homework_name": "Homework Check", # For live session name
            "live_homework_session_mode": "Yes/No", # "Yes/No" or "Select"
            "log_homework_marks_enabled": True, # Enable/disable detailed marks for manual log
            "homework_mark_types": DEFAULT_HOMEWORK_MARK_TYPES.copy(),
            "default_homework_items_for_yes_no_mode": 5, # For live session "Yes/No"
            "live_homework_score_font_color": DEFAULT_HOMEWORK_SCORE_FONT_COLOR,
            "live_homework_score_font_style_bold": DEFAULT_HOMEWORK_SCORE_FONT_STYLE_BOLD,


            # Password settings
            "app_password_hash": None,
            "password_on_open": False,
            "password_on_edit_action": False,
            "password_auto_lock_enabled": False,
            "password_auto_lock_timeout_minutes": 15,

            # Next ID counters (managed by _ensure_next_ids but good to have defaults)
            "next_student_id_num": 1,
            "next_furniture_id_num": 1,
            "next_group_id_num": 1,
            "next_quiz_template_id_num": 1,
            "next_homework_template_id_num": 1, # New
            "next_custom_homework_type_id_num": 1, # For custom homework types in Yes/No mode

            # Internal state storage (prefixed with underscore)
            "_last_used_quiz_name_for_session": "",
            "_last_used_quiz_name_timestamp_for_session": None,
            "_last_used_q_num_for_session": 10,
            "_last_used_homework_name_for_session": "", # New
            "_last_used_homework_name_timestamp_for_session": None, # New
            "_last_used_hw_items_for_session": 5, # 
            "theme": "System", # Newer
        }

    def _ensure_next_ids(self):
        # Student IDs
        max_s_id = 0
        for sid in self.students:
            if sid.startswith("student_"):
                try: max_s_id = max(max_s_id, int(sid.split("_")[1]))
                except (ValueError, IndexError): pass
        self.next_student_id_num = max(self.settings.get("next_student_id_num", 1), max_s_id + 1)
        self.settings["next_student_id_num"] = self.next_student_id_num

        # Furniture IDs
        max_f_id = 0
        for fid in self.furniture:
            if fid.startswith("furniture_"):
                try: max_f_id = max(max_f_id, int(fid.split("_")[1]))
                except (ValueError, IndexError): pass
        self.next_furniture_id_num = max(self.settings.get("next_furniture_id_num", 1), max_f_id + 1)
        self.settings["next_furniture_id_num"] = self.next_furniture_id_num

        # Group IDs
        max_g_id = 0
        for gid in self.student_groups:
            if gid.startswith("group_"):
                try: max_g_id = max(max_g_id, int(gid.split("_")[1]))
                except (ValueError, IndexError): pass
        self.next_group_id_num = max(self.settings.get("next_group_id_num", 1), max_g_id + 1)
        self.settings["next_group_id_num"] = self.next_group_id_num

        # Quiz Template IDs
        max_qt_id = 0
        for qtid in self.quiz_templates:
            if qtid.startswith("quiztemplate_"):
                try: max_qt_id = max(max_qt_id, int(qtid.split("_")[1]))
                except (ValueError, IndexError): pass
        self.next_quiz_template_id_num = max(self.settings.get("next_quiz_template_id_num", 1), max_qt_id + 1)
        self.settings["next_quiz_template_id_num"] = self.next_quiz_template_id_num

        # Homework Template IDs (New)
        max_ht_id = 0
        for htid in self.homework_templates:
            if htid.startswith("hwtemplate_"): # Consistent prefix
                try: max_ht_id = max(max_ht_id, int(htid.split("_")[1]))
                except (ValueError, IndexError): pass
        self.next_homework_template_id_num = max(self.settings.get("next_homework_template_id_num", 1), max_ht_id + 1)
        self.settings["next_homework_template_id_num"] = self.next_homework_template_id_num

        # Custom Homework Type IDs (for Yes/No mode live session) - New
        max_chwt_id = 0
        for chwt in self.custom_homework_session_types: # Assuming these are dicts with an 'id' field
            if isinstance(chwt, dict) and chwt.get('id', '').startswith("hwtype_"):
                try: max_chwt_id = max(max_chwt_id, int(chwt['id'].split("_")[1]))
                except (ValueError, IndexError): pass
        self.settings["next_custom_homework_type_id_num"] = max(self.settings.get("next_custom_homework_type_id_num", 1), max_chwt_id + 1)


    def periodic_checks(self):
        self.password_manager.check_auto_lock()
        if self.password_manager.is_locked and not hasattr(self, '_lock_screen_active'):
            self.show_lock_screen()
        self.root.after(30000, self.periodic_checks)

    def show_lock_screen(self):
        if hasattr(self, '_lock_screen_active') and self._lock_screen_active.winfo_exists(): return
        self._lock_screen_active = tk.Toplevel(self.root)
        self._lock_screen_active.title("Application Locked")
        self._lock_screen_active.transient(self.root)
        self._lock_screen_active.grab_set()
        self._lock_screen_active.protocol("WM_DELETE_WINDOW", lambda: None)
        self.root.update_idletasks()
        px, py, pw, ph = self.root.winfo_x(), self.root.winfo_y(), self.root.winfo_width(), self.root.winfo_height()
        dw, dh = 300, 150
        x, y = px + (pw // 2) - (dw // 2), py + (ph // 2) - (dh // 2)
        self._lock_screen_active.geometry(f"{dw}x{dh}+{x}+{y}")
        self._lock_screen_active.resizable(False, False)
        ttk.Label(self._lock_screen_active, text="Application is locked. Enter password:", font=("", 12)).pack(pady=10)
        password_var = tk.StringVar()
        password_entry = ttk.Entry(self._lock_screen_active, textvariable=password_var, show="*", width=30)
        password_entry.pack(pady=5); password_entry.focus_set()
        status_label_lock = ttk.Label(self._lock_screen_active, text="", foreground="red"); status_label_lock.pack(pady=2)
        def attempt_unlock():
            if self.password_manager.unlock_application(password_var.get()):
                self._lock_screen_active.destroy(); del self._lock_screen_active
                self.update_status("Application unlocked.")
            else:
                status_label_lock.config(text="Incorrect password."); password_entry.select_range(0, tk.END)
        ttk.Button(self._lock_screen_active, text="Unlock", command=attempt_unlock).pack(pady=10)
        self._lock_screen_active.bind('<Return>', lambda e: attempt_unlock())

    def prompt_for_password(self, title, prompt_message, for_editing=False):
        if self.password_manager.is_locked:
             if not hasattr(self, '_lock_screen_active') or not self._lock_screen_active.winfo_exists(): self.show_lock_screen()
             return not self.password_manager.is_locked
        if for_editing and not self.settings.get("password_on_edit_action", False) and not self.password_manager.is_password_set(): return True
        if not self.password_manager.is_password_set(): return True
        dialog = PasswordPromptDialog(self.root, title, prompt_message, self.password_manager)
        return dialog.result

    def execute_command(self, command: Command):
        is_sensitive_edit = isinstance(command, (AddItemCommand, DeleteItemCommand, EditItemCommand, ChangeItemsSizeCommand, ManageStudentGroupCommand))
        if is_sensitive_edit and self.settings.get("password_on_edit_action", False) and self.password_manager.is_password_set():
            if not self.prompt_for_password("Confirm Action", "Enter password to make this change:", for_editing=True):
                self.update_status("Action cancelled: Password not provided or incorrect."); return
        try:
            command.execute()
            self.undo_stack.append(command)
            self.redo_stack.clear()
            self.update_undo_redo_buttons_state()
            if not isinstance(command, (MarkLiveQuizQuestionCommand, MarkLiveHomeworkCommand)):
                self.save_data_wrapper(source="command_execution")
            self.password_manager.record_activity()
        except Exception as e:
            messagebox.showerror("Command Error", f"Error executing command: {e}\nCommand Type: {type(command).__name__}", parent=self.root)
            print(f"Command execution error: {e}\n{type(command)}")

    def undo_last_action(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock Required", "Enter password to undo action:"): return
        if self.undo_stack:
            command = self.undo_stack.pop()
            try:
                command.undo()
                self.redo_stack.append(command)
                self.update_undo_redo_buttons_state()
                if not isinstance(command, (MarkLiveQuizQuestionCommand, MarkLiveHomeworkCommand)):
                    self.save_data_wrapper(source="undo_command")
                self.password_manager.record_activity()
            except Exception as e:
                messagebox.showerror("Undo Error", f"Error undoing action: {e}", parent=self.root)
                self.undo_stack.append(command); print(f"Undo error: {e}\n{type(command)}")

    def redo_last_action(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock Required", "Enter password to redo action:"): return
        if self.redo_stack:
            command = self.redo_stack.pop()
            try:
                command.execute()
                self.undo_stack.append(command)
                self.update_undo_redo_buttons_state()
                if not isinstance(command, (MarkLiveQuizQuestionCommand, MarkLiveHomeworkCommand)):
                    self.save_data_wrapper(source="redo_command")
                self.password_manager.record_activity()
            except Exception as e:
                messagebox.showerror("Redo Error", f"Error redoing action: {e}", parent=self.root)
                self.redo_stack.append(command); print(f"Redo error: {e}\n{type(command)}")

    def update_undo_redo_buttons_state(self):
        if hasattr(self, 'undo_btn'): self.undo_btn.config(state=tk.NORMAL if self.undo_stack else tk.DISABLED)
        if hasattr(self, 'redo_btn'): self.redo_btn.config(state=tk.NORMAL if self.redo_stack else tk.DISABLED)

    def get_new_student_id(self):
        current_id_to_assign = self.next_student_id_num
        return f"student_{current_id_to_assign}", self.next_student_id_num + 1
    def get_new_furniture_id(self):
        current_id_to_assign = self.next_furniture_id_num
        return f"furniture_{current_id_to_assign}", self.next_furniture_id_num + 1
    def get_new_group_id(self):
        current_id_to_assign = self.next_group_id_num
        return f"group_{current_id_to_assign}", self.next_group_id_num + 1
    def get_new_quiz_template_id(self):
        current_id_to_assign = self.next_quiz_template_id_num
        return f"quiztemplate_{current_id_to_assign}", self.next_quiz_template_id_num + 1
    def get_new_homework_template_id(self): # New
        current_id_to_assign = self.next_homework_template_id_num
        return f"hwtemplate_{current_id_to_assign}", self.next_homework_template_id_num + 1
    def get_new_custom_homework_type_id(self): # New
        current_id_to_assign = self.settings.get("next_custom_homework_type_id_num", 1)
        return f"hwtype_{current_id_to_assign}", current_id_to_assign + 1


    def update_status(self, message):
        if self.status_bar_label: self.status_bar_label.configure(text=message)

    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="5"); main_frame.pack(fill=tk.BOTH, expand=True)
        top_controls_frame_row1 = ttk.Frame(main_frame); top_controls_frame_row1.pack(side=tk.TOP, fill=tk.X, pady=(0, 2))
        self.undo_btn = ttk.Button(top_controls_frame_row1, text="Undo", command=self.undo_last_action, state=tk.DISABLED); self.undo_btn.pack(side=tk.LEFT, padx=2)
        self.redo_btn = ttk.Button(top_controls_frame_row1, text="Redo", command=self.redo_last_action, state=tk.DISABLED); self.redo_btn.pack(side=tk.LEFT, padx=2)
        ttk.Button(top_controls_frame_row1, text="Add Student", command=self.add_student_dialog).pack(side=tk.LEFT, padx=2)
        ttk.Button(top_controls_frame_row1, text="Add Furniture", command=self.add_furniture_dialog).pack(side=tk.LEFT, padx=2)

        self.mode_frame = ttk.LabelFrame(top_controls_frame_row1, text="Mode", padding=2); self.mode_frame.pack(side=tk.LEFT, padx=3); self.mode_frame.pack_propagate(True)
        ttk.Radiobutton(self.mode_frame, text="Behavior", variable=self.mode_var, value="behavior", command=self.toggle_mode).pack(side=tk.LEFT)
        ttk.Radiobutton(self.mode_frame, text="Quiz", variable=self.mode_var, value="quiz", command=self.toggle_mode).pack(side=tk.LEFT)
        ttk.Radiobutton(self.mode_frame, text="Homework", variable=self.mode_var, value="homework", command=self.toggle_mode).pack(side=tk.LEFT) # New Homework mode

        self.live_quiz_button_frame = ttk.LabelFrame(top_controls_frame_row1, text="Class Quiz")
        self.start_live_quiz_btn = ttk.Button(self.live_quiz_button_frame, text="Start Session", command=self.start_live_quiz_session_dialog); self.start_live_quiz_btn.pack(side=tk.LEFT, padx=3, pady=3)
        self.end_live_quiz_btn = ttk.Button(self.live_quiz_button_frame, text="End Session", command=self.end_live_quiz_session, state=tk.DISABLED); self.end_live_quiz_btn.pack(side=tk.LEFT, padx=3, pady=3)

        self.live_homework_button_frame = ttk.LabelFrame(top_controls_frame_row1, text="Homework Session") # New
        self.start_live_homework_btn = ttk.Button(self.live_homework_button_frame, text="Start Session", command=self.start_live_homework_session_dialog); self.start_live_homework_btn.pack(side=tk.LEFT, padx=3, pady=3)
        self.end_live_homework_btn = ttk.Button(self.live_homework_button_frame, text="End Session", command=self.end_live_homework_session, state=tk.DISABLED); self.end_live_homework_btn.pack(side=tk.LEFT, padx=3, pady=3)

        templates_groups_frame = ttk.LabelFrame(top_controls_frame_row1, text="Layout & Groups", padding=2); templates_groups_frame.pack(side=tk.LEFT, padx=0)
        ttk.Button(templates_groups_frame, text="Save Layout...", command=self.save_layout_template_dialog).pack(side=tk.LEFT,pady=1, padx=1)
        ttk.Button(templates_groups_frame, text="Load Layout...", command=self.load_layout_template_dialog).pack(side=tk.LEFT,pady=1, padx=1)
        self.manage_groups_btn = ttk.Button(templates_groups_frame, text="Manage Groups...", command=self.manage_student_groups_dialog); self.manage_groups_btn.pack(side=tk.LEFT,pady=1, padx=1)

        top_controls_frame_row2 = ttk.Frame(main_frame); top_controls_frame_row2.pack(side=tk.TOP, fill=tk.X, pady=(2, 5))
        self.file_menu_btn = ttk.Menubutton(top_controls_frame_row2, text="File"); self.file_menu = tk.Menu(self.file_menu_btn, tearoff=0)
        self.file_menu.add_command(label="Save Now", command=self.save_data_wrapper, accelerator="Ctrl+S")
        self.file_menu.add_command(label="Import Students from Excel...", command=self.import_students_from_excel_dialog)
        self.file_menu.add_separator(); self.file_menu.add_command(label="Open Data Folder", command=self.open_data_folder)
        self.open_export_folder_menu_entry_index = self.file_menu.index(tk.END)
        self.file_menu.add_command(label="Open Last Export Folder (None)", command=self.open_last_export_folder, state=tk.DISABLED)
        self.open_export_folder_menu_entry_index = self.file_menu.index(tk.END)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Backup All Application Data (.zip)...", command=self.backup_all_data_dialog)
        self.file_menu.add_command(label="Restore All Application Data...", command=self.restore_all_data_dialog)
        self.file_menu.add_separator(); self.file_menu.add_command(label="Reset Application (Caution!)...", command=self.reset_application_dialog)
        self.file_menu.add_separator(); self.file_menu.add_command(label="Exit", command=self.on_exit_protocol, accelerator="Ctrl+Q")
        self.file_menu_btn["menu"] = self.file_menu; self.file_menu_btn.pack(side=tk.LEFT, padx=2)
        self.update_open_last_export_folder_menu_item()
        self.root.bind_all("<Control-s>", lambda event: self.save_data_wrapper())
        self.root.bind_all("<Control-q>", lambda event: self.save_and_quit_app())
        self.root.bind_all("<Control-Q>", lambda event: self.save_and_quit_app())
        self.root.bind_all("<Control-z>", lambda event: self.undo_last_action())
        self.root.bind_all("<Control-y>", lambda event: self.redo_last_action())

        self.export_menu_btn = ttk.Menubutton(top_controls_frame_row2, text="Export Log"); self.export_menu = tk.Menu(self.export_menu_btn, tearoff=0)
        self.export_menu.add_command(label="To Excel (.xlsx)", command=lambda: self.export_log_dialog_with_filter(export_type="xlsx"))
        self.export_menu.add_command(label="To Excel Macro-Enabled (.xlsm)", command=lambda: self.export_log_dialog_with_filter(export_type="xlsm"))
        self.export_menu.add_command(label="To CSV Files (.zip)", command=lambda: self.export_log_dialog_with_filter(export_type="csv"))
        self.export_menu.add_separator()
        self.export_menu.add_command(label="Export Layout as Image (see Help)...", command=self.export_layout_as_image)
        self.export_menu.add_command(label="Generate Attendance Report...", command=self.generate_attendance_report_dialog)
        self.export_menu_btn["menu"] = self.export_menu; self.export_menu_btn.pack(side=tk.LEFT, padx=2)
        ttk.Button(top_controls_frame_row2, text="Settings", command=self.open_settings_dialog).pack(side=tk.LEFT, padx=2)


        self.zoom_var = tk.StringVar(value=str(float(self.current_zoom_level)*100))
        #print(self.zoom_var.get())

        view_controls_frame = ttk.LabelFrame(top_controls_frame_row2, text="View & Edit", padding=2); view_controls_frame.pack(side=tk.LEFT, padx=5)
        ttk.Button(view_controls_frame, text="Zoom In", command=lambda: self.zoom_canvas(1.1)).pack(side=tk.LEFT, padx=2)
        self.zoom_display_label = ttk.Entry(view_controls_frame, textvariable=self.zoom_var, width=5)
        if self.settings.get("show_zoom_level_display", True): self.zoom_display_label.pack(side=tk.LEFT, padx=1)
        ttk.Button(view_controls_frame, text="Zoom Out", command=lambda: self.zoom_canvas(0.9)).pack(side=tk.LEFT, padx=2)
        ttk.Button(view_controls_frame, text="Reset Zoom", command=lambda: self.zoom_canvas(0)).pack(side=tk.LEFT, padx=2)
        self.edit_mode_checkbutton = ttk.Checkbutton(view_controls_frame, text="Edit Mode (Resize)", variable=self.edit_mode_var, command=self.toggle_edit_mode); self.edit_mode_checkbutton.pack(side=tk.LEFT, padx=5)
        self.toggle_incidents_btn = ttk.Button(view_controls_frame, text="Hide Recent Logs", command=self.toggle_global_recent_logs_visibility); self.toggle_incidents_btn.pack(side=tk.LEFT, padx=2) # Renamed
        self.update_toggle_incidents_button_text()

        layout_tools_frame = ttk.LabelFrame(top_controls_frame_row2, text="Layout Tools", padding=2); layout_tools_frame.pack(side=tk.LEFT, padx=0)
        ttk.Button(layout_tools_frame, text="Align Top", command=lambda: self.align_selected_items("top")).pack(side=tk.LEFT,pady=1, padx=1)
        ttk.Button(layout_tools_frame, text="Align Bottom", command=lambda: self.align_selected_items("bottom")).pack(side=tk.LEFT,pady=1, padx=1)
        ttk.Button(layout_tools_frame, text="Align Left", command=lambda: self.align_selected_items("left")).pack(side=tk.LEFT,pady=1, padx=1)
        ttk.Button(layout_tools_frame, text="Align Right", command=lambda: self.align_selected_items("right")).pack(side=tk.LEFT,pady=1, padx=1)

        self.lock_app_btn = ttk.Button(top_controls_frame_row1, text="Lock", command=self.lock_application_ui_triggered); self.lock_app_btn.pack(side=tk.RIGHT, padx=5)
        self.update_lock_button_state()
        self.root.bind_all("<Control-l>", lambda e: self.lock_application_ui_triggered())
        
        self.zoom_display_label.bind("<FocusOut>", lambda e: self.update_zoom_display2())
        self.zoom_display_label.bind("<Return>", lambda e: self.update_zoom_display2())
        
        
        ttk.Button(top_controls_frame_row1, text="Help", command=self.show_help_dialog).pack(side=tk.RIGHT, padx=2)
        #self.theme_style_using = "dark"
        try:
            if self.theme_style_using == "dark": var4 = "#1F1F1F"
            elif self.theme_style_using == "System": var4 = "lightgrey" if darkdetect.theme() == "Light" else "#1F1F1F"
            else: var4 = "lightgrey"
        except:
            if self.theme_style_using == "dark": var4 = "#1F1F1F"
            #elif self.theme_style_using == "System": var4 = "lightgrey" if darkdetect.theme() == "Light" else "#1F1F1F"
            else: var4 = "lightgrey"
        try:
            if self.custom_canvas_color != "Default": self.canvas_color = self.custom_canvas_color
            elif self.theme_style_using == "Dark": self.canvas_color = "#1F1F1F"
            elif self.theme_style_using == "System": self.canvas_color = "lightgrey" if darkdetect.theme() == "Light" else "#1F1F1F"
            else: self.canvas_color = "lightgrey"
        except:
            if self.custom_canvas_color != "Default": self.canvas_color = self.custom_canvas_color
            elif self.theme_style_using == "Dark": self.canvas_color = "#1F1F1F"
            #elif self.theme_style_using == "System": self.canvas_color = "lightgrey" if darkdetect.theme() == "Light" else "#1F1F1F"
            else: self.canvas_color = "lightgrey"
            
        self.canvas_frame = ttk.Frame(main_frame); self.canvas_frame.pack(fill=tk.BOTH, expand=True)
        self.h_scrollbar = ttk.Scrollbar(self.canvas_frame, orient=tk.HORIZONTAL, command=self.canvas_xview_custom)
        self.v_scrollbar = ttk.Scrollbar(self.canvas_frame, orient=tk.VERTICAL, command=self.canvas_yview_custom) #else "#1F1F1F"
        self.canvas = tk.Canvas(self.canvas_frame, bg=self.canvas_color, relief=tk.SUNKEN, borderwidth=1, xscrollcommand=self.h_scrollbar.set, yscrollcommand=self.v_scrollbar.set)
        self.h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X); self.v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y); self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.canvas.config(scrollregion=(0, 0, self.canvas_orig_width * self.current_zoom_level, self.canvas_orig_height * self.current_zoom_level))
        self.canvas.bind("<ButtonPress-1>", self.on_canvas_left_press); self.canvas.bind("<ButtonPress-3>", self.on_canvas_right_press)
        self.canvas.bind("<B1-Motion>", self.on_canvas_drag); self.canvas.bind("<ButtonRelease-1>", self.on_canvas_release)
        self.canvas.bind("<Control-Button-1>", self.on_canvas_ctrl_click); self.canvas.bind("<KeyPress-Delete>", self.on_delete_key_press); self.canvas.bind("<BackSpace>", self.on_delete_key_press)
        self.canvas.bind("<Control-MouseWheel>", self.on_mousewheel_zoom); self.canvas.bind("<Control-Button-4>", self.on_mousewheel_zoom); self.canvas.bind("<Control-Button-5>", self.on_mousewheel_zoom)
        self.canvas.bind("<ButtonPress-2>", self.on_pan_start); self.canvas.bind("<B2-Motion>", self.on_pan_move); self.canvas.bind("<ButtonRelease-2>", self.on_pan_end)
        self.canvas.bind("<MouseWheel>", self.on_mousewheel_scroll); self.canvas.bind("<Button-4>", self.on_mousewheel_scroll); self.canvas.bind("<Button-5>", self.on_mousewheel_scroll)
        if sys.platform == "darwin": self.canvas.bind("<Shift-MouseWheel>", self.on_mousewheel_scroll_horizontal_mac)
        else: self.canvas.bind("<Shift-MouseWheel>", self.on_mouse_wheel_horizontal) # For Windows/Linux with Shift

        self.status_bar_label = ttk.Label(self.root, text="Welcome!", relief=tk.SUNKEN, anchor=tk.W, padding=5); self.status_bar_label.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.focus_set()
        self.toggle_student_groups_ui_visibility()

    def canvas_xview_custom(self, *args): self.canvas.xview(*args); self.password_manager.record_activity()
    def canvas_yview_custom(self, *args): self.canvas.yview(*args); self.password_manager.record_activity()
    def on_mousewheel_scroll(self, event):
        if self.password_manager.is_locked: return
        self.password_manager.record_activity()
        if event.num == 5 or event.delta < 0: self.canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta > 0: self.canvas.yview_scroll(-1, "units")
    def on_mouse_wheel_horizontal(self, event): # For Shift+Wheel on Windows/Linux
        if self.password_manager.is_locked: return
        self.password_manager.record_activity()
        if event.delta < 0: self.canvas.xview_scroll(1, "units") # Scroll right
        elif event.delta > 0: self.canvas.xview_scroll(-1, "units") # Scroll left
    def on_mousewheel_scroll_horizontal_mac(self, event):
        if self.password_manager.is_locked: return
        self.password_manager.record_activity()
        if event.delta < 0: self.canvas.xview_scroll(1, "units")
        elif event.delta > 0: self.canvas.xview_scroll(-1, "units")

    def lock_application_ui_triggered(self):
        if self.password_manager.is_password_set():
            if self.password_manager.lock_application():
                self.update_status("Application locked."); self.show_lock_screen(); self.update_lock_button_state()
            else: self.update_status("Failed to lock: No password set or already locked.")
        else:
            messagebox.showinfo("Password Not Set", "Please set an application password in Settings first.", parent=self.root)
            self.open_settings_dialog()
    def update_lock_button_state(self):
        if hasattr(self, 'lock_app_btn'): self.lock_app_btn.config(state=tk.NORMAL if self.password_manager.is_password_set() else tk.DISABLED)
    def save_and_quit_app(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Save & Quit", "Enter password to save and quit:"): return
        if self.is_live_quiz_active and not self.prompt_end_live_session_on_mode_switch("quiz"): return
        if self.is_live_homework_active and not self.prompt_end_live_session_on_mode_switch("homework"): return # New
        self.save_data_wrapper(source="save_and_quit")
        self.on_exit_protocol(force_quit=True) # Call main exit to release lock

    def on_delete_key_press(self, event=None):
        if self.password_manager.is_locked: return
        if self.selected_items: self.delete_selected_items_confirm()
        self.password_manager.record_activity()
    def update_open_last_export_folder_menu_item(self):
        if hasattr(self, 'file_menu') and self.file_menu:
            label_text, state = "Open Last Export Folder (None)", tk.DISABLED
            if self.last_excel_export_path and os.path.exists(os.path.dirname(self.last_excel_export_path)):
                label_text, state = f"Open Last Export Folder ({os.path.basename(os.path.dirname(self.last_excel_export_path))})", tk.NORMAL
            try:
                if self.open_export_folder_menu_entry_index is not None and self.open_export_folder_menu_entry_index <= self.file_menu.index(tk.END):
                    self.file_menu.entryconfigure(self.open_export_folder_menu_entry_index, label=label_text, state=state)
            except tk.TclError as e: print(f"Error updating 'Open Last Export Folder' menu item: {e}.")

    def prompt_end_live_session_on_mode_switch(self, session_type_to_check): # "quiz" or "homework"
        is_active_flag = self.is_live_quiz_active if session_type_to_check == "quiz" else self.is_live_homework_active
        current_name = self.current_live_quiz_name if session_type_to_check == "quiz" else self.current_live_homework_name
        end_function = self.end_live_quiz_session if session_type_to_check == "quiz" else self.end_live_homework_session
        start_btn = self.start_live_quiz_btn if session_type_to_check == "quiz" else self.start_live_homework_btn
        end_btn = self.end_live_quiz_btn if session_type_to_check == "quiz" else self.end_live_homework_btn
        scores_dict = self.live_quiz_scores if session_type_to_check == "quiz" else self.live_homework_scores

        if is_active_flag:
            msg = f"A Class {session_type_to_check.capitalize()} session ('{current_name}') is active.\n\nDo you want to end and log this session?"
            response = messagebox.askyesnocancel(f"Class {session_type_to_check.capitalize()} Active", msg, parent=self.root)
            if response is True: end_function(confirm=False); return True
            elif response is False: # Discard
                if session_type_to_check == "quiz": self.is_live_quiz_active = False; self.current_live_quiz_name = ""
                else: self.is_live_homework_active = False; self.current_live_homework_name = ""
                scores_dict.clear()
                start_btn.config(state=tk.NORMAL); end_btn.config(state=tk.DISABLED)
                self.update_status(f"Class {session_type_to_check.capitalize()} session discarded.")
                self.draw_all_items(check_collisions_on_redraw=True); return True
            else: return False # Cancel
        return True # No active session of this type


    def toggle_mode(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Change Mode", "Enter password to change mode:"):
                self.mode_var.set(self.settings["current_mode"]); return

        new_mode = self.mode_var.get()
        old_mode = self.settings["current_mode"]

        if old_mode == "quiz" and new_mode != "quiz" and self.is_live_quiz_active:
            if not self.prompt_end_live_session_on_mode_switch("quiz"): self.mode_var.set("quiz"); return
        if old_mode == "homework" and new_mode != "homework" and self.is_live_homework_active: # New
            if not self.prompt_end_live_session_on_mode_switch("homework"): self.mode_var.set("homework"); return


        self.settings["current_mode"] = new_mode
        self.update_status(f"Mode switched to {new_mode.capitalize()}.")

        if hasattr(self, 'live_quiz_button_frame'):
            if new_mode == "quiz":
                self.live_quiz_button_frame.pack(side=tk.LEFT, padx=(0,3), after=self.mode_frame)
                self.start_live_quiz_btn.config(state=tk.DISABLED if self.is_live_quiz_active else tk.NORMAL)
                self.end_live_quiz_btn.config(state=tk.NORMAL if self.is_live_quiz_active else tk.DISABLED)
            else: self.live_quiz_button_frame.pack_forget()

        if hasattr(self, 'live_homework_button_frame'): # New
            if new_mode == "homework":
                self.live_homework_button_frame.pack(side=tk.LEFT, padx=(0,3), after=self.mode_frame)
                self.start_live_homework_btn.config(state=tk.DISABLED if self.is_live_homework_active else tk.NORMAL)
                self.end_live_homework_btn.config(state=tk.NORMAL if self.is_live_homework_active else tk.DISABLED)
            else: self.live_homework_button_frame.pack_forget()

        self.draw_all_items(check_collisions_on_redraw=True)
        self.save_data_wrapper(source="toggle_mode")
        self.password_manager.record_activity()

    def toggle_edit_mode(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Toggle Edit Mode", "Enter password to change edit mode:"):
                self.edit_mode_var.set(not self.edit_mode_var.get()); return
        is_edit_mode = self.edit_mode_var.get()
        self.update_status(f"Edit Mode (Resize) {'Enabled' if is_edit_mode else 'Disabled'}. Click item corners to resize.")
        self.draw_all_items(check_collisions_on_redraw=True)
        self.password_manager.record_activity()

    def start_live_quiz_session_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Start Quiz", "Enter password to start quiz session:"): return
        if self.is_live_quiz_active:
            messagebox.showinfo("Class Quiz Active", "A Class Quiz session is already active.", parent=self.root); return
        quiz_name = simpledialog.askstring("Start Class Quiz", "Enter a name for this quiz session:", initialvalue=self.settings.get("default_quiz_name", "Class Quiz"), parent=self.root)
        if quiz_name and quiz_name.strip():
            self.current_live_quiz_name = quiz_name.strip()
            self.is_live_quiz_active = True; self.live_quiz_scores.clear()
            self.start_live_quiz_btn.config(state=tk.DISABLED); self.end_live_quiz_btn.config(state=tk.NORMAL)
            self.update_status(f"Class Quiz '{self.current_live_quiz_name}' started. Click a student to mark.")
            self.draw_all_items(check_collisions_on_redraw=True); self.password_manager.record_activity()
        else: self.update_status("Class Quiz start cancelled.")

    def end_live_quiz_session(self, confirm=True):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to End Quiz", "Enter password to end quiz session:"): return
        if not self.is_live_quiz_active: self.update_status("No active Class Quiz session to end."); return
        if confirm and not messagebox.askyesno("End Class Quiz", f"End Class Quiz session: '{self.current_live_quiz_name}'?\nScores will be logged.", parent=self.root): return

        log_commands = []
        for student_id, score_data in self.live_quiz_scores.items():
            if student_id in self.students and score_data["total_asked"] > 0:
                student = self.students[student_id]
                log_entry = {"timestamp": datetime.now().isoformat(), "student_id": student_id,
                             "student_first_name": student["first_name"], "student_last_name": student["last_name"],
                             "behavior": self.current_live_quiz_name, "score_details": score_data.copy(),
                             "comment": "From Class Quiz session.", "type": "quiz", "day": datetime.now().strftime('%A')}
                log_commands.append(LogEntryCommand(self, log_entry, student_id))
        for cmd in log_commands: self.execute_command(cmd)
        if log_commands: self.save_data_wrapper(source="end_live_quiz")
        self.update_status(f"Class Quiz '{self.current_live_quiz_name}' ended. {len(log_commands)} student scores logged.")
        self.is_live_quiz_active = False; self.current_live_quiz_name = ""; self.live_quiz_scores.clear()
        self.start_live_quiz_btn.config(state=tk.NORMAL); self.end_live_quiz_btn.config(state=tk.DISABLED)
        self.draw_all_items(check_collisions_on_redraw=True); self.password_manager.record_activity()

    def handle_live_quiz_tap(self, student_id):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Mark Quiz", "Enter password to mark quiz:"): return
        if not self.is_live_quiz_active or student_id not in self.students: return
        dialog = LiveQuizMarkDialog(self.root, student_id, self, session_type="Quiz")
        if dialog.result:
            self.execute_command(MarkLiveQuizQuestionCommand(self, student_id, dialog.result))
            self.password_manager.record_activity()

    # --- Live Homework Session Methods (New) ---
    def start_live_homework_session_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Start Homework Session", "Enter password to start homework session:"): return
        if self.is_live_homework_active:
            messagebox.showinfo("Homework Session Active", "A Homework session is already active.", parent=self.root); return

        homework_session_name = simpledialog.askstring("Start Homework Session", "Enter a name for this homework session:",
                                                       initialvalue=self.settings.get("default_homework_name", "Homework Check"), parent=self.root)
        if homework_session_name and homework_session_name.strip():
            self.current_live_homework_name = homework_session_name.strip()
            self.is_live_homework_active = True
            self.live_homework_scores.clear() # {student_id: {hw_type_id: "yes"/"no"} or {student_id: {"selected_options": [...]}}}
            self.start_live_homework_btn.config(state=tk.DISABLED)
            self.end_live_homework_btn.config(state=tk.NORMAL)
            self.update_status(f"Homework Session '{self.current_live_homework_name}' started. Click a student to mark.")
            self.draw_all_items(check_collisions_on_redraw=True)
            self.password_manager.record_activity()
        else:
            self.update_status("Homework Session start cancelled.")

    def end_live_homework_session(self, confirm=True):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to End Homework Session", "Enter password to end homework session:"): return
        if not self.is_live_homework_active:
            self.update_status("No active Homework Session to end."); return
        if confirm and not messagebox.askyesno("End Homework Session", f"End Homework Session: '{self.current_live_homework_name}'?\nData will be logged.", parent=self.root):
            return

        log_commands = []
        for student_id, hw_data in self.live_homework_scores.items():
            if student_id in self.students and hw_data: # Ensure there's some data to log
                student = self.students[student_id]
                # Log entry structure depends on session mode
                if self.settings.get('live_homework_session_mode') == "Yes/No":
                    type2 = "y"
                else: type2 = "s"
                log_entry = {
                    "timestamp": datetime.now().isoformat(), "student_id": student_id,
                    "student_first_name": student["first_name"], "student_last_name": student["last_name"],
                    "behavior": self.current_live_homework_name, # Session name
                    "homework_details": hw_data.copy(), # Store the collected data
                    "comment": f"From Live Homework Session ({self.settings.get('live_homework_session_mode', 'Yes/No')} mode).",
                    "type": f"homework_session_{type2}", # Distinguish from manual homework log
                    "day": datetime.now().strftime('%A')
                }
                log_commands.append(LogHomeworkEntryCommand(self, log_entry, student_id)) # Use homework log command

        for cmd in log_commands: self.execute_command(cmd)
        if log_commands: self.save_data_wrapper(source="end_live_homework_session")

        self.update_status(f"Homework Session '{self.current_live_homework_name}' ended. {len(log_commands)} student entries logged.")
        self.is_live_homework_active = False; self.current_live_homework_name = ""; self.live_homework_scores.clear()
        self.start_live_homework_btn.config(state=tk.NORMAL); self.end_live_homework_btn.config(state=tk.DISABLED)
        self.draw_all_items(check_collisions_on_redraw=True); self.password_manager.record_activity()

    def handle_live_homework_tap(self, student_id):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Mark Homework", "Enter password to mark homework:"): return
        if not self.is_live_homework_active or student_id not in self.students: return

        session_mode = self.settings.get("live_homework_session_mode", "Yes/No")
        current_student_hw_data = self.live_homework_scores.get(student_id, {}).copy()

        dialog = LiveHomeworkMarkDialog(self.root, student_id, self,
                                        session_mode=session_mode,
                                        current_hw_data=current_student_hw_data)
        if dialog.result_actions is not None: # Dialog returns actions or None if cancelled
            cmd = MarkLiveHomeworkCommand(self, student_id, dialog.result_actions, session_mode)
            self.execute_command(cmd)
            self.password_manager.record_activity()


    def add_student_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Add Student", "Enter password to add student:"): return
        dialog = AddEditStudentDialog(self.root, "Add Student", app=self)
        if dialog.result:
            first_name, last_name, nickname, gender, group_id_selection = dialog.result
            if first_name and last_name:
                old_next_student_id_num_for_command = self.next_student_id_num
                student_id_str, next_id_val_for_app_state_after_this = self.get_new_student_id()
                full_name = f"{first_name} \"{nickname}\" {last_name}" if nickname else f"{first_name} {last_name}"
                x, y = self.canvas_to_world_coords(50 + (len(self.students) % 10) * 20, 50 + ((len(self.students) // 10) * 20))
                student_data = {"first_name": first_name, "last_name": last_name, "nickname": nickname, "full_name": full_name, "gender": gender,
                                "x": x, "y": y, "id": student_id_str, "width": self.settings.get("default_student_box_width"),
                                "height": self.settings.get("default_student_box_height"), "original_next_id_num_after_add": next_id_val_for_app_state_after_this,
                                "group_id": group_id_selection if group_id_selection else None, "style_overrides": {}}
                self.execute_command(AddItemCommand(self, student_id_str, 'student', student_data, old_next_student_id_num_for_command))
                self.password_manager.record_activity()
            else: messagebox.showwarning("Invalid Name", "First and Last names cannot be empty.", parent=self.root)

    def add_furniture_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Add Furniture", "Enter password to add furniture:"): return
        dialog = AddFurnitureDialog(self.root, "Add Furniture Item")
        if dialog.result:
            name, item_type, width, height = dialog.result
            if name and item_type:
                old_next_furniture_id_num_for_command = self.next_furniture_id_num
                furniture_id_str, next_id_val_for_app_state_after_this = self.get_new_furniture_id()
                x, y = self.canvas_to_world_coords(70 + (len(self.furniture) % 10) * 20, 70 + ((len(self.furniture) // 10) * 20))
                furniture_data = {"name": name, "type": item_type, "x": x, "y": y, "id": furniture_id_str, "width": width, "height": height,
                                  "fill_color": "lightgray", "outline_color": "dimgray", "original_next_id_num_after_add": next_id_val_for_app_state_after_this}
                self.execute_command(AddItemCommand(self, furniture_id_str, 'furniture', furniture_data, old_next_furniture_id_num_for_command))
                self.password_manager.record_activity()

    def toggle_global_recent_logs_visibility(self): # Renamed for clarity
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Change Setting", "Enter password to change log visibility:"): return
        # This now toggles both behavior and homework visibility together for simplicity,
        # or could be split into two separate global toggles if needed.
        self._recent_incidents_hidden_globally = not self._recent_incidents_hidden_globally
        self._recent_homeworks_hidden_globally = self._recent_incidents_hidden_globally # Link them for now
        self.draw_all_items(check_collisions_on_redraw=True)
        self.update_toggle_incidents_button_text() # Button text reflects combined state
        status_msg = "Recent behavior/homework logs hidden globally." if self._recent_incidents_hidden_globally else "Recent behavior/homework logs shown globally."
        self.update_status(status_msg)
        self.password_manager.record_activity()

    def update_toggle_incidents_button_text(self): # Renamed
        if hasattr(self, "toggle_incidents_btn"):
            # Text reflects combined state of behavior and homework logs
            text = "Show Recent Logs" if self._recent_incidents_hidden_globally or self._recent_homeworks_hidden_globally else "Hide Recent Logs"
            self.toggle_incidents_btn.config(text=text)

    def clear_recent_logs_for_student(self, student_id): # Renamed
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Clear Logs", "Enter password to clear logs for student:"): return
        self._per_student_last_cleared[student_id] = datetime.now().isoformat()
        self.draw_single_student(student_id, check_collisions=True)
        student = self.students.get(student_id)
        if student: self.update_status(f"Recent behavior/homework logs cleared for {student['full_name']}.")
        self.save_data_wrapper(); self.password_manager.record_activity()

    def show_recent_logs_for_student(self, student_id): # Renamed
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Show Logs", "Enter password to show logs for student:"): return
        if student_id in self._per_student_last_cleared:
            del self._per_student_last_cleared[student_id]
            self.draw_single_student(student_id, check_collisions=True)
            student = self.students.get(student_id)
            if student: self.update_status(f"Recent behavior/homework logs will now show for {student['full_name']}.")
            self.save_data_wrapper(); self.password_manager.record_activity()

    def _get_recent_logs_for_student(self, student_id, log_type_key): # "behavior" or "homework"
        student = self.students.get(student_id)
        if not student: return []

        summary_lines_list = []
        log_source = self.behavior_log if log_type_key == "behavior" else self.homework_log
        setting_prefix = "recent_incidents" if log_type_key == "behavior" else "recent_homeworks" # For settings keys
        global_hidden_flag = self._recent_incidents_hidden_globally if log_type_key == "behavior" else self._recent_homeworks_hidden_globally
        behavior_key_in_log = "behavior" if log_type_key == "behavior" else "homework_type" # Or "behavior" for manual homework log

        if not global_hidden_flag and self.settings.get(f"show_{setting_prefix}_on_boxes", True):
            num_to_show = self.settings.get(f"num_{setting_prefix}_to_show", 0)
            if num_to_show > 0:
                time_window_hours = self.settings.get(f"{setting_prefix}_time_window_hours", 24)
                cutoff_time = datetime.now() - timedelta(hours=time_window_hours)
                last_cleared_iso = self._per_student_last_cleared.get(student_id)
                cleared_dt = datetime.fromisoformat(last_cleared_iso) if last_cleared_iso else None

                # Adjust filter for log type. Behavior logs have "type":"behavior", quiz logs have "type":"quiz".
                # Homework logs have "type":"homework" or "type":"homework_session".
                type_filter_values = []
                if log_type_key == "behavior": type_filter_values = ["behavior", "quiz"] # Behavior tab shows both
                elif log_type_key == "homework": type_filter_values = ["homework", "homework_session_y", "homework_session_s"]


                all_recent_logs = sorted(
                    [log for log in log_source
                     if log["student_id"] == student_id and log.get("type") in type_filter_values and
                        datetime.fromisoformat(log["timestamp"]) >= cutoff_time and
                        (not cleared_dt or datetime.fromisoformat(log["timestamp"]) > cleared_dt)],
                    key=lambda x: x["timestamp"], reverse=True
                )

                specific_filter_list = self.settings.get(f"selected_{setting_prefix}_filter", None)
                filtered_logs = []
                if specific_filter_list is None: filtered_logs = all_recent_logs
                elif isinstance(specific_filter_list, list) and not specific_filter_list: filtered_logs = [] # Empty list means filter all
                elif isinstance(specific_filter_list, list):
                    filtered_logs = [log for log in all_recent_logs if log.get(behavior_key_in_log, log.get("behavior")) in specific_filter_list]


                recent_to_display = filtered_logs[:num_to_show]
                if self.settings.get(f"reverse_{setting_prefix.replace('s', '')}_order", True): recent_to_display.reverse()

                initial_overrides_key = "behavior_initial_overrides" if log_type_key == "behavior" else "homework_initial_overrides"
                initial_overrides = self.settings.get(initial_overrides_key, {})
                
                #string_for_get = 
                
                if self.settings.get(f"show_full_{setting_prefix}", False):
                    summary_lines_list = [log.get(behavior_key_in_log, log.get("behavior")) for log in recent_to_display if log.get(behavior_key_in_log, log.get("behavior"))]
                else:
                    temp_initials = []
                    for entry in recent_to_display:
                        name = entry.get(behavior_key_in_log, entry.get("behavior"))
                        if name in initial_overrides and initial_overrides[name]: temp_initials.append(initial_overrides[name])
                        elif name: temp_initials.append(''.join(part[0].upper() for part in name.split() if part))
                        else: temp_initials.append("?")
                    if temp_initials: summary_lines_list.append('  '.join(temp_initials))
        return summary_lines_list


    
    def _get__logs_for_student(self, student_id, log_type_key, num_max, window, name_of_spec): # "behavior" or "homework"
        student = self.students.get(student_id)
        if not student: return []
        print(num_max)
        summary_lines_list = []
        log_source = self.behavior_log if log_type_key == "behavior" else self.homework_log
        setting_prefix = "recent_incidents" if log_type_key == "behavior" else "recent_homeworks" # For settings keys
        global_hidden_flag = self._recent_incidents_hidden_globally if log_type_key == "behavior" else self._recent_homeworks_hidden_globally
        behavior_key_in_log = "behavior" if log_type_key == "behavior" else "homework_type" # Or "behavior" for manual homework log

        if not global_hidden_flag and self.settings.get(f"show_{setting_prefix}_on_boxes", True):
            num_to_show = num_max
            if num_to_show > 0:
                time_window_hours = window
                cutoff_time = datetime.now() - timedelta(hours=time_window_hours)

                # Adjust filter for log type. Behavior logs have "type":"behavior", quiz logs have "type":"quiz".
                # Homework logs have "type":"homework" or "type":"homework_session".
                type_filter_values = []
                if log_type_key == "behavior": type_filter_values = ["behavior", "quiz"] # Behavior tab shows both
                elif log_type_key == "homework": type_filter_values = ["homework", "homework_session_y", "homework_session_s"]


                all_recent_logs = sorted(
                    [log for log in log_source
                     if log["student_id"] == student_id and log.get("type") in type_filter_values and
                        datetime.fromisoformat(log["timestamp"]) >= cutoff_time],
                    key=lambda x: x["timestamp"], reverse=True
                )

                specific_filter_list = [name_of_spec]
                filtered_logs = []
                if specific_filter_list is None: filtered_logs = all_recent_logs
                elif isinstance(specific_filter_list, list) and not specific_filter_list: filtered_logs = [] # Empty list means filter all
                elif isinstance(specific_filter_list, list):
                    filtered_logs = [log for log in all_recent_logs if log.get(behavior_key_in_log, log.get("behavior")) in specific_filter_list]


                recent_to_display = filtered_logs[:num_to_show]
                
                
                summary_lines_list = [log.get(behavior_key_in_log, log.get("behavior")) for log in recent_to_display if log.get(behavior_key_in_log, log.get("behavior"))]
                
        return len(summary_lines_list)


    

    def update_student_display_text(self, student_id):
        student = self.students.get(student_id)
        if not student: return

        name_to_display = student.get('nickname') or student['first_name']
        main_content_lines = [name_to_display, student['last_name']]
        incident_display_lines = [] # List of dicts: {"text": "...", "type": "incident/quiz_score/homework_score"}

        current_mode = self.mode_var.get()
        if current_mode == "quiz" and self.is_live_quiz_active:
            score_text = f"{self.current_live_quiz_name}: (Pending)"
            if student_id in self.live_quiz_scores:
                score_info = self.live_quiz_scores[student_id]
                score_text = f"{self.current_live_quiz_name}: {score_info['correct']} of {score_info['total_asked']}"
            incident_display_lines.append({"text": score_text, "type": "quiz_score"})
            if self.settings.get("show_recent_incidents_during_quiz", True):
                for line_text in self._get_recent_logs_for_student(student_id, "behavior"):
                    if line_text: incident_display_lines.append({"text": line_text, "type": "incident"})

        elif current_mode == "homework" and self.is_live_homework_active: # New for live homework
            hw_session_mode = self.settings.get("live_homework_session_mode", "Yes/No")
            hw_score_data = self.live_homework_scores.get(student_id, {})
            hw_display_texts = []
            if hw_session_mode == "Yes/No":
                for hw_type_id, status in hw_score_data.items():
                    # Find homework type name from custom_homework_session_types or defaults
                    hw_type_obj = next((ht for ht in self.all_homework_session_types if isinstance(ht, dict) and ht.get('id') == hw_type_id), None)
                    hw_name_display = hw_type_obj['name'] if hw_type_obj else hw_type_id
                    hw_display_texts.append(f"{hw_name_display}: {status.capitalize()}")
            elif hw_session_mode == "Select" and "selected_options" in hw_score_data:
                hw_display_texts = list(hw_score_data["selected_options"])

            if hw_display_texts:
                incident_display_lines.append({"text": f"{self.current_live_homework_name}:", "type": "homework_score_header"})
                for text in hw_display_texts:
                    incident_display_lines.append({"text": f"  {text}", "type": "homework_score_item"})
            else:
                incident_display_lines.append({"text": f"{self.current_live_homework_name}: (Pending)", "type": "homework_score_header"})


        else: # Behavior mode or non-live quiz/homework mode
            # Show recent behavior/quiz logs
            for line_text in self._get_recent_logs_for_student(student_id, "behavior"):
                if line_text: incident_display_lines.append({"text": line_text, "type": "incident"})
            # Show recent homework logs
            homework_lines = self._get_recent_logs_for_student(student_id, "homework")
            if homework_lines and incident_display_lines: # Add separator if both types of logs exist
                incident_display_lines.append({"text": "--- Homework ---", "type": "separator"})
            for line_text in homework_lines:
                if line_text: incident_display_lines.append({"text": line_text, "type": "homework_log"})


        student["display_lines"] = main_content_lines
        student["incident_display_lines"] = incident_display_lines


    def applies_to_conditional(self, student_id, rule):
        
        student_data = self.students.get(student_id)
        if not student_data: return None
        
        type_ = rule.get("type", "")
        
        if type_ == "behavior_count":
            time_window = rule.get("time_window_hours", "")
            count_threshhold = rule.get("count_threshhold", 1)
            behavior_name = rule.get("behavior_name", "")
            result = self._get__logs_for_student(student_id, "behavior", count_threshhold,time_window, behavior_name)
            if result >=count_threshhold: # type: ignore
                return True
        
        elif type_ == "quiz_score_threshhold":
            
            operator = rule.get("operator", "")
            contains = rule.get("quiz_name_contains", "")
            score_threshhold = rule.get("score_threshold_percent", 0)
            
            return None
        
        
        elif type_ == "group": return None
        
        return None
        #if student_data.get()
        
        




    def draw_single_student(self, student_id, check_collisions=False):
        # ... (largely same as v51, but needs to handle new "homework_score_header/item" and "separator" types for drawing)
        # This method is long, so I'll highlight the key change area for incident_display_lines
        try:
            student_data = self.students.get(student_id)
            if not student_data: return
            self.update_student_display_text(student_id)

            style_overrides = student_data.get("style_overrides", {})
            world_x, world_y = student_data["x"], student_data["y"]
            world_width = style_overrides.get("width", student_data.get("width", self.settings.get("default_student_box_width")))
            world_base_height = style_overrides.get("height", student_data.get("height", self.settings.get("default_student_box_height")))
            canvas_x, canvas_y = self.world_to_canvas_coords(world_x, world_y)
            canvas_width = world_width * self.current_zoom_level
            canvas_base_height = world_base_height * self.current_zoom_level

            fill_color = style_overrides.get("fill_color", self.settings.get("student_box_fill_color"))
            outline_color_orig = style_overrides.get("outline_color", self.settings.get("student_box_outline_color"))
            font_family = style_overrides.get("font_family", self.settings.get("student_font_family"))
            font_size_world = style_overrides.get("font_size", self.settings.get("student_font_size"))
            font_size_canvas = int(max(6, font_size_world * self.current_zoom_level))
            font_color = style_overrides.get("font_color", self.settings.get("student_font_color"))

            group_id = student_data.get("group_id"); group_indicator_color = None
            if self.settings.get("student_groups_enabled", True) and group_id and group_id in self.student_groups:
                group_data = self.student_groups[group_id]
                group_indicator_color = group_data.get("color")
                for rule in self.settings.get("conditional_formatting_rules", []):
                    if rule.get("type") == "group" and rule.get("group_id") == group_id:
                        if "color" in rule and rule["color"]: fill_color = rule["color"]
                        if "outline" in rule and rule["outline"]: outline_color_orig = rule["outline"]
                        break
                    
            
            if self.settings.get("conditonal_formatting_rules") != []:
                #pass
                for rule in self.settings.get("conditional_formatting_rules", []):
                    if self.applies_to_conditional(student_id, rule):
                        
                        if "color" in rule and rule["color"]: fill_color = rule["color"]
                        if "outline" in rule and rule["outline"]: outline_color_orig = rule["outline"]
                        break
            
                    
                    

            name_font_obj = tkfont.Font(family=font_family, size=font_size_canvas, weight="bold")
            incident_font_obj = tkfont.Font(family=font_family, size=max(5, font_size_canvas -1))
            quiz_score_font_color_setting = self.settings.get("live_quiz_score_font_color")
            quiz_score_font_bold_setting = self.settings.get("live_quiz_score_font_style_bold")
            quiz_score_font_weight = "bold" if quiz_score_font_bold_setting else "normal"
            quiz_score_font_obj = tkfont.Font(family=font_family, size=font_size_canvas, weight=quiz_score_font_weight)

            # New: Homework Score Font
            hw_score_font_color_setting = self.settings.get("live_homework_score_font_color", DEFAULT_HOMEWORK_SCORE_FONT_COLOR)
            hw_score_font_bold_setting = self.settings.get("live_homework_score_font_style_bold", DEFAULT_HOMEWORK_SCORE_FONT_STYLE_BOLD)
            hw_score_font_weight = "bold" if hw_score_font_bold_setting else "normal"
            hw_score_font_obj = tkfont.Font(family=font_family, size=font_size_canvas, weight=hw_score_font_weight)
            hw_score_item_font_obj = tkfont.Font(family=font_family, size=max(5, font_size_canvas -1), weight=hw_score_font_weight) # Slightly smaller for items


            self.canvas.delete(student_id)
            rect_tag = ("student_item", student_id, "rect")

            world_padding = 5; canvas_padding = world_padding * self.current_zoom_level
            current_y_offset_for_calc_world = world_padding
            for name_line_text in student_data.get("display_lines", []):
                font_for_calc = tkfont.Font(family=font_family, size=font_size_world, weight="bold")
                current_y_offset_for_calc_world += font_for_calc.metrics('linespace')

            if student_data.get("incident_display_lines"):
                current_y_offset_for_calc_world += world_padding / 2
                for line_info in student_data.get("incident_display_lines", []):
                    line_text, line_type = line_info["text"], line_info["type"]
                    current_font_world_calc = tkfont.Font(family=font_family, size=font_size_world -1)
                    if line_type == "quiz_score": current_font_world_calc = tkfont.Font(family=font_family, size=font_size_world, weight=quiz_score_font_weight)
                    elif line_type == "homework_score_header": current_font_world_calc = tkfont.Font(family=font_family, size=font_size_world, weight=hw_score_font_weight)
                    elif line_type == "homework_score_item": current_font_world_calc = tkfont.Font(family=font_family, size=max(5, font_size_world -1), weight=hw_score_font_weight)
                    elif line_type == "separator": current_font_world_calc = tkfont.Font(family=font_family, size=max(4, font_size_world -2)) # Smaller for separator

                    text_width_pixels_world = current_font_world_calc.measure(line_text)
                    visual_lines_world = 1
                    available_width_for_text_world = world_width - 2 * world_padding
                    if available_width_for_text_world > 0 and text_width_pixels_world > available_width_for_text_world:
                        visual_lines_world = -(-text_width_pixels_world // available_width_for_text_world)
                    current_y_offset_for_calc_world += visual_lines_world * current_font_world_calc.metrics('linespace')

            world_text_content_height_with_padding = current_y_offset_for_calc_world + world_padding
            world_dynamic_height = max(world_base_height, world_text_content_height_with_padding)
            canvas_dynamic_height = world_dynamic_height * self.current_zoom_level
            student_data['_current_world_height'] = world_dynamic_height
            student_data['_current_world_width'] = world_width

            self.canvas.create_rectangle(canvas_x, canvas_y, canvas_x + canvas_width, canvas_y + canvas_dynamic_height,
                                         fill=fill_color, outline=outline_color_orig, width=max(1, int(2 * self.current_zoom_level)), tags=rect_tag)

            current_y_text_draw_canvas = canvas_y + canvas_padding
            available_text_width_canvas = canvas_width - 2 * canvas_padding
            for name_line_text in student_data.get("display_lines", []):
                self.canvas.create_text(canvas_x + canvas_width / 2, current_y_text_draw_canvas, text=name_line_text,
                                        fill=font_color, font=name_font_obj, tags=("student_item", student_id, "text", "student_name"),
                                        anchor=tk.N, width=max(1, available_text_width_canvas), justify=tk.CENTER)
                current_y_text_draw_canvas += name_font_obj.metrics('linespace')

            if student_data.get("incident_display_lines"):
                current_y_text_draw_canvas += canvas_padding / 2
                for line_info in student_data.get("incident_display_lines", []):
                    line_text, line_type = line_info["text"], line_info["type"]
                    current_font_canvas_draw, current_color_canvas_draw = incident_font_obj, font_color
                    text_anchor_canvas, text_justify_canvas = tk.N, tk.CENTER
                    text_x_pos_canvas = canvas_x + canvas_width / 2

                    if line_type == "quiz_score": current_font_canvas_draw, current_color_canvas_draw = quiz_score_font_obj, quiz_score_font_color_setting
                    elif line_type == "homework_score_header": current_font_canvas_draw, current_color_canvas_draw = hw_score_font_obj, hw_score_font_color_setting
                    elif line_type == "homework_score_item":
                        current_font_canvas_draw, current_color_canvas_draw = hw_score_item_font_obj, hw_score_font_color_setting
                        text_anchor_canvas, text_justify_canvas = tk.NW, tk.LEFT # Align items to left
                        text_x_pos_canvas = canvas_x + canvas_padding # Start from left padding
                    elif line_type == "separator":
                        current_font_canvas_draw = tkfont.Font(family=font_family, size=max(4, int((font_size_world-2)*self.current_zoom_level)))
                        current_color_canvas_draw = "gray"


                    self.canvas.create_text(text_x_pos_canvas, current_y_text_draw_canvas, text=line_text,
                                            fill=current_color_canvas_draw, font=current_font_canvas_draw,
                                            tags=("student_item", student_id, "text", f"student_{line_type}"),
                                            anchor=text_anchor_canvas, width=max(1, available_text_width_canvas if text_anchor_canvas == tk.N else available_text_width_canvas - canvas_padding), # Adjust width for NW anchor
                                            justify=text_justify_canvas)
                    text_width_pixels_canvas = current_font_canvas_draw.measure(line_text)
                    visual_lines_canvas = 1
                    if available_text_width_canvas > 0 and text_width_pixels_canvas > available_text_width_canvas:
                        visual_lines_canvas = -(-text_width_pixels_canvas // available_text_width_canvas)
                    current_y_text_draw_canvas += visual_lines_canvas * current_font_canvas_draw.metrics('linespace')


            if self.settings.get("student_groups_enabled", True) and group_indicator_color:
                indicator_size_canvas = GROUP_COLOR_INDICATOR_SIZE * self.current_zoom_level
                indicator_padding_canvas = 2 * self.current_zoom_level
                indicator_x = canvas_x + canvas_width - indicator_size_canvas - indicator_padding_canvas
                indicator_y = canvas_y + indicator_padding_canvas
                self.canvas.create_rectangle(indicator_x, indicator_y, indicator_x + indicator_size_canvas, indicator_y + indicator_size_canvas,
                                            fill=group_indicator_color, outline=outline_color_orig, tags=("student_item", student_id, "group_indicator"))
            if student_id in self.selected_items:
                sel_outline_width = max(1, int(2 * self.current_zoom_level))
                self.canvas.create_rectangle(canvas_x - sel_outline_width, canvas_y - sel_outline_width,
                                             canvas_x + canvas_width + sel_outline_width, canvas_y + canvas_dynamic_height + sel_outline_width,
                                             outline="red", width=sel_outline_width, tags=("student_item", student_id, "selection_highlight"))
            if self.edit_mode_var.get() and student_id in self.selected_items:
                handle_size_canvas = RESIZE_HANDLE_SIZE * self.current_zoom_level
                br_x = canvas_x + canvas_width - handle_size_canvas / 2 # Center handle on corner
                br_y = canvas_y + canvas_dynamic_height - handle_size_canvas / 2
                self.canvas.create_rectangle(br_x - handle_size_canvas/2, br_y - handle_size_canvas/2,
                                             br_x + handle_size_canvas/2, br_y + handle_size_canvas/2,
                                             fill="gray", outline="black", tags=("student_item", student_id, "resize_handle", "br_handle"))
            if check_collisions: self.handle_layout_collision(student_id)
        except AttributeError: pass # Canvas might not be fully initialized during early calls

    def draw_single_furniture(self, furniture_id):
        # ... (same as v51)
        item_data = self.furniture.get(furniture_id)
        if not item_data: return
        world_x, world_y = item_data["x"], item_data["y"]
        world_width = item_data.get("width", DEFAULT_STUDENT_BOX_WIDTH)
        world_height = item_data.get("height", DEFAULT_STUDENT_BOX_HEIGHT)
        canvas_x, canvas_y = self.world_to_canvas_coords(world_x, world_y)
        canvas_width = world_width * self.current_zoom_level
        canvas_height = world_height * self.current_zoom_level
        item_data['_current_world_height'] = world_height
        item_data['_current_world_width'] = world_width
        fill_color = item_data.get("fill_color", "lightgrey")
        outline_color = item_data.get("outline_color", "dimgray")
        name = item_data.get("name", "Furniture")
        try:
            self.canvas.delete(furniture_id)
            rect_tag = ("furniture_item", furniture_id, "rect")
            self.canvas.create_rectangle(canvas_x, canvas_y, canvas_x + canvas_width, canvas_y + canvas_height,
                                         fill=fill_color, outline=outline_color, width=max(1, int(2*self.current_zoom_level)), tags=rect_tag)
            font_size_canvas = int(max(6, (self.settings.get("student_font_size", DEFAULT_FONT_SIZE) -1) * self.current_zoom_level))
            font_spec = (self.settings.get("student_font_family", DEFAULT_FONT_FAMILY), font_size_canvas)
            self.canvas.create_text(canvas_x + canvas_width / 2, canvas_y + canvas_height / 2, text=name,
                                    fill=self.settings.get("student_font_color", DEFAULT_FONT_COLOR), font=font_spec,
                                    tags=("furniture_item", furniture_id, "text"), anchor=tk.CENTER,
                                    width=max(1, canvas_width - int(10*self.current_zoom_level)), justify=tk.CENTER)
            if furniture_id in self.selected_items:
                sel_outline_width = max(1, int(2 * self.current_zoom_level))
                self.canvas.create_rectangle(canvas_x - sel_outline_width, canvas_y - sel_outline_width,
                                             canvas_x + canvas_width + sel_outline_width, canvas_y + canvas_height + sel_outline_width,
                                             outline="red", width=sel_outline_width, tags=("furniture_item", furniture_id, "selection_highlight"))
            if self.edit_mode_var.get() and furniture_id in self.selected_items:
                handle_size_canvas = RESIZE_HANDLE_SIZE * self.current_zoom_level
                br_x = canvas_x + canvas_width - handle_size_canvas / 2
                br_y = canvas_y + canvas_height - handle_size_canvas / 2
                self.canvas.create_rectangle(br_x - handle_size_canvas/2, br_y - handle_size_canvas/2,
                                             br_x + handle_size_canvas/2, br_y + handle_size_canvas/2,
                                             fill="gray", outline="black", tags=("furniture_item", furniture_id, "resize_handle", "br_handle"))
        except AttributeError: pass

    def draw_all_items(self, check_collisions_on_redraw=False):
        # ... (same as v51)
        try: self.canvas.delete("student_item"); self.canvas.delete("furniture_item")
        except AttributeError: pass
        all_items_data = list(self.students.values()) + list(self.furniture.values())
        if not all_items_data:
            try:
                default_sr_w = self.canvas_orig_width * self.current_zoom_level; default_sr_h = self.canvas_orig_height * self.current_zoom_level
                self.canvas.configure(scrollregion=(0, 0, default_sr_w, default_sr_h))
            except AttributeError: pass
        else:
            min_x_world, min_y_world = float('inf'), float('inf'); max_x_world_br, max_y_world_br = float('-inf'), float('-inf')
            for item_data in all_items_data:
                item_world_x, item_world_y = item_data['x'], item_data['y']
                item_world_w = item_data.get('_current_world_width', item_data.get('width', DEFAULT_STUDENT_BOX_WIDTH))
                item_world_h = item_data.get('_current_world_height', item_data.get('height', DEFAULT_STUDENT_BOX_HEIGHT))
                min_x_world = min(min_x_world, item_world_x); min_y_world = min(min_y_world, item_world_y)
                max_x_world_br = max(max_x_world_br, item_world_x + item_world_w); max_y_world_br = max(max_y_world_br, item_world_y + item_world_h)
            padding_world = 100
            scroll_min_x_canvas, scroll_min_y_canvas = self.world_to_canvas_coords(min_x_world - padding_world, min_y_world - padding_world)
            scroll_max_x_canvas, scroll_max_y_canvas = self.world_to_canvas_coords(max_x_world_br + padding_world, max_y_world_br + padding_world)
            try:
                visible_canvas_width = self.canvas.winfo_width(); visible_canvas_height = self.canvas.winfo_height()
                final_scroll_max_x = max(scroll_max_x_canvas, scroll_min_x_canvas + visible_canvas_width)
                final_scroll_max_y = max(scroll_max_y_canvas, scroll_min_y_canvas + visible_canvas_height)
            except AttributeError: final_scroll_max_x, final_scroll_max_y = scroll_max_x_canvas, scroll_max_y_canvas # Fallback if winfo fails
            final_scroll_min_x = min(scroll_min_x_canvas, 0); final_scroll_min_y = min(scroll_min_y_canvas, 0)
            try: self.canvas.config(scrollregion=(final_scroll_min_x, final_scroll_min_y, final_scroll_max_x, final_scroll_max_y))
            except AttributeError: pass
        for student_id in self.students: self.draw_single_student(student_id, check_collisions=check_collisions_on_redraw)
        for furniture_id in self.furniture: self.draw_single_furniture(furniture_id)
        self.update_toggle_incidents_button_text(); self.update_zoom_display()

    def handle_layout_collision(self, moved_item_id):
        # ... (same as v51)
        if moved_item_id not in self.students: return
        moved_item_data = self.students[moved_item_id]
        moved_x1, moved_y1 = moved_item_data['x'], moved_item_data['y']
        moved_w = moved_item_data.get('_current_world_width', moved_item_data.get('width', DEFAULT_STUDENT_BOX_WIDTH))
        moved_h = moved_item_data.get('_current_world_height', moved_item_data.get('height', DEFAULT_STUDENT_BOX_HEIGHT))
        moved_x2, moved_y2 = moved_x1 + moved_w, moved_y1 + moved_h
        items_to_shift_data = []
        all_other_items = []
        for sid, sdata in self.students.items():
            if sid != moved_item_id: all_other_items.append({'id': sid, 'type': 'student', 'data': sdata})
        for fid, fdata in self.furniture.items(): all_other_items.append({'id': fid, 'type': 'furniture', 'data': fdata})
        for other_item_info in all_other_items:
            other_id, other_data = other_item_info['id'], other_item_info['data']
            other_x1, other_y1 = other_data['x'], other_data['y']
            other_w = other_data.get('_current_world_width', other_data.get('width', DEFAULT_STUDENT_BOX_WIDTH))
            other_h = other_data.get('_current_world_height', other_data.get('height', DEFAULT_STUDENT_BOX_HEIGHT))
            other_x2, other_y2 = other_x1 + other_w, other_y1 + other_h
            is_colliding = not (moved_x2 <= other_x1 or moved_x1 >= other_x2 or moved_y2 <= other_y1 or moved_y1 >= other_y2)
            if is_colliding:
                vertical_overlap = moved_y2 - other_y1
                if vertical_overlap > 0 :
                    shift_by = vertical_overlap + LAYOUT_COLLISION_OFFSET
                    items_to_shift_data.append({'id': other_id, 'type': other_item_info['type'], 'old_x': other_x1, 'old_y': other_y1, 'new_x': other_x1, 'new_y': other_y1 + shift_by})
        if items_to_shift_data:
            self.execute_command(MoveItemsCommand(self, items_to_shift_data))
            self.update_status(f"Adjusted layout for {len(items_to_shift_data)} items due to overlap with {moved_item_data['full_name']}.")

    def world_to_canvas_coords(self, world_x, world_y):
        return world_x * self.current_zoom_level, world_y * self.current_zoom_level
     
    def canvas_to_world_coords(self, canvas_x_on_screen, canvas_y_on_screen):
        if self.current_zoom_level == 0: return canvas_x_on_screen, canvas_y_on_screen
        true_canvas_x = self.canvas.canvasx(canvas_x_on_screen)
        true_canvas_y = self.canvas.canvasy(canvas_y_on_screen)
        return true_canvas_x / self.current_zoom_level, true_canvas_y / self.current_zoom_level

    def update_zoom_display(self):
        if hasattr(self, 'zoom_display_label') and self.zoom_display_label:
            if self.settings.get("show_zoom_level_display", True):
                #self.zoom_display_label.config(text=f"{self.current_zoom_level*100:.0f}%")
                self.zoom_var.set(value=str(self.current_zoom_level*100.0))
                if not self.zoom_display_label.winfo_ismapped():
                    zoom_in_btn = next((child for child in self.zoom_display_label.master.winfo_children() if isinstance(child, ttk.Button) and "Zoom In" in child.cget("text")), None)
                    if zoom_in_btn: self.zoom_display_label.pack(side=tk.LEFT, padx=1, after=zoom_in_btn)
                    else: self.zoom_display_label.pack(side=tk.LEFT, padx=1)
            elif self.zoom_display_label.winfo_ismapped(): self.zoom_display_label.pack_forget()

    def update_zoom_display2(self):
        #print("Return")
        self.current_zoom_level = float(self.zoom_var.get())/100.0
        self.zoom_var.set(value=str(float(self.current_zoom_level)*100.0))
        self.zoom_canvas(1)


    def zoom_canvas(self, factor):
        # ... (same as v51)
        if self.password_manager.is_locked: return
        world_center_x_before, world_center_y_before = self.canvas_to_world_coords(self.canvas.winfo_width() // 2, self.canvas.winfo_height() // 2)
        if factor == 0: self.current_zoom_level = 1.0
        else: self.current_zoom_level = max(0.1, min(self.current_zoom_level * factor, 10.0))
        self.draw_all_items(check_collisions_on_redraw=False)
        # Centering logic after zoom could be added here if desired, similar to v50/v51
        self.update_status(f"Zoom level: {self.current_zoom_level:.2f}x"); self.update_zoom_display(); self.password_manager.record_activity()
        self.zoom_var.set(value=str(self.current_zoom_level*100.0))
        self.zoom_display_label.configure(textvariable=self.zoom_var)
        #print(self.zoom_var.get())

    def on_mousewheel_zoom(self, event):
        if self.password_manager.is_locked: return
        factor = 0.9 if (event.num == 5 or event.delta < 0) else 1.1
        self.zoom_canvas(factor)
    def on_pan_start(self, event):
        # ... (same as v51)
        if self.password_manager.is_locked: return
        world_event_x, world_event_y = self.canvas_to_world_coords(event.x, event.y)
        item_ids_under_cursor = self.canvas.find_overlapping(event.x-1, event.y-1, event.x+1, event.y+1) # Use screen coords for find_overlapping
        clicked_on_item = False
        for item_canvas_id in item_ids_under_cursor:
            tags = self.canvas.gettags(item_canvas_id)
            if any(t.startswith("student_") or t.startswith("furniture_") for t in tags if "rect" in tags or "resize_handle" in tags):
                clicked_on_item = True; break
        if not clicked_on_item:
            self.canvas.scan_mark(event.x, event.y); self.update_status("Panning canvas..."); self._drag_started_on_item = False
        else: self._drag_started_on_item = True
        self.password_manager.record_activity()

    def on_pan_move(self, event):
        if self.password_manager.is_locked: return
        if not self._drag_started_on_item: self.canvas.scan_dragto(event.x, event.y, gain=1)
        self.password_manager.record_activity()
    def on_pan_end(self, event):
        if self.password_manager.is_locked: return
        if not self._drag_started_on_item: self.update_status("Canvas panned.")
        self._drag_started_on_item = False; self.password_manager.record_activity()
    def on_canvas_ctrl_click(self, event):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Select", "Enter password to select items:"): return
        world_click_x, world_click_y = self.canvas_to_world_coords(event.x, event.y)
        item_canvas_ids = self.canvas.find_overlapping(event.x -1, event.y -1, event.x +1, event.y+1) # Use screen coords
        topmost_item_id, topmost_item_type = None, None
        for item_c_id in reversed(item_canvas_ids):
            tags = self.canvas.gettags(item_c_id); current_item_id, current_item_type = None, None
            for tag in tags:
                if tag.startswith("student_") and tag in self.students: current_item_id, current_item_type = tag, "student"; break
                elif tag.startswith("furniture_") and tag in self.furniture: current_item_id, current_item_type = tag, "furniture"; break
            if current_item_id and any("rect" in t for t in tags): topmost_item_id, topmost_item_type = current_item_id, current_item_type; break
        if topmost_item_id:
            if topmost_item_id in self.selected_items: self.selected_items.remove(topmost_item_id)
            else: self.selected_items.add(topmost_item_id)
            if topmost_item_type == "student": self.draw_single_student(topmost_item_id)
            elif topmost_item_type == "furniture": self.draw_single_furniture(topmost_item_id)
            self.update_status(f"{len(self.selected_items)} items selected.")
        self.password_manager.record_activity()

    def on_canvas_left_press(self, event):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Interact", "Enter password to interact with canvas:"): return
        self.canvas.focus_set()
        world_event_x, world_event_y = self.canvas_to_world_coords(event.x, event.y)
        self.drag_data = {"x": world_event_x, "y": world_event_y, "item_id": None, "item_type": None,
                          "start_x_world": world_event_x, "start_y_world": world_event_y, # Store start in world
                          "original_positions": {}, "is_resizing": False, "resize_handle_corner": None,
                          "original_size_world": {}}
        self._potential_click_target = None; self._drag_started_on_item = False

        if self.edit_mode_var.get():
            # Use screen coordinates for find_overlapping with canvas items
            item_ids_under_cursor_screen = self.canvas.find_overlapping(world_event_x - RESIZE_HANDLE_SIZE/2, world_event_y - RESIZE_HANDLE_SIZE/2,
                                                                        world_event_x + RESIZE_HANDLE_SIZE/2, world_event_y + RESIZE_HANDLE_SIZE/2)
            for item_c_id in reversed(item_ids_under_cursor_screen):
                tags = self.canvas.gettags(item_c_id)
                if any("resize_handle" in tag for tag in tags):
                    item_id, item_type, handle_corner = None, None, None
                    for tag in tags:
                        if tag.startswith("student_") and tag in self.students: item_id, item_type = tag, "student"
                        elif tag.startswith("furniture_") and tag in self.furniture: item_id, item_type = tag, "furniture"
                        if "br_handle" in tag: handle_corner = "br"; break
                    if item_id and item_id in self.selected_items:
                        self.drag_data.update({"item_id": item_id, "item_type": item_type, "is_resizing": True, "resize_handle_corner": handle_corner})
                        data_src = self.students if item_type == "student" else self.furniture
                        # Store original world width/height
                        orig_w = data_src[item_id].get("width", DEFAULT_STUDENT_BOX_WIDTH)
                        orig_h = data_src[item_id].get("height", DEFAULT_STUDENT_BOX_HEIGHT)
                        if item_type == "student":
                           orig_w = data_src[item_id].get("style_overrides",{}).get("width", orig_w)
                           orig_h = data_src[item_id].get("style_overrides",{}).get("height", orig_h)
                        self.drag_data["original_size_world"] = {"width": orig_w, "height": orig_h}
                        self._drag_started_on_item = True; self.update_status(f"Resizing {item_type} '{item_id}'..."); self.password_manager.record_activity(); return

        item_canvas_ids_click = self.canvas.find_overlapping(world_event_x -1, world_event_y -1, world_event_x +1, world_event_y +1) # Screen coords
        clicked_item_id, clicked_item_type = None, None
        for item_c_id in reversed(item_canvas_ids_click):
            tags = self.canvas.gettags(item_c_id); temp_id, temp_type, is_main_rect = None, None, False
            for tag in tags:
                if tag.startswith("student_") and tag in self.students: temp_id, temp_type = tag, "student"
                elif tag.startswith("furniture_") and tag in self.furniture: temp_id, temp_type = tag, "furniture"
                if "rect" in tag: is_main_rect = True
            if temp_id and is_main_rect: clicked_item_id, clicked_item_type = temp_id, temp_type; break
        clicked_on_selected_item = clicked_item_id and clicked_item_id in self.selected_items
        if not (event.state & 0x0004): # Ctrl NOT pressed
            if clicked_item_id and not clicked_on_selected_item:
                self.deselect_all_items(); self.selected_items.add(clicked_item_id)
                if clicked_item_type == "student": self.draw_single_student(clicked_item_id)
                else: self.draw_single_furniture(clicked_item_id)
            elif not clicked_item_id: self.deselect_all_items()
        if clicked_item_id:
            self._potential_click_target = clicked_item_id; self.drag_data["item_id"] = clicked_item_id; self.drag_data["item_type"] = clicked_item_type; self._drag_started_on_item = True
            self.drag_data["original_positions"] = {}
            for sel_id in self.selected_items:
                sel_item_type_drag = "student" if sel_id in self.students else "furniture"
                data_src_drag = self.students if sel_id in self.students else self.furniture
                if sel_id in data_src_drag: self.drag_data["original_positions"][sel_id] = {"x": data_src_drag[sel_id]["x"], "y": data_src_drag[sel_id]["y"], "type": sel_item_type_drag}
        self.password_manager.record_activity()

    def on_canvas_drag(self, event):
        if self.password_manager.is_locked: return
        if not self.drag_data.get("item_id") or not self._drag_started_on_item: return

        world_event_x, world_event_y = self.canvas_to_world_coords(event.x, event.y)

        if self.drag_data.get("is_resizing"):
            item_id, item_type = self.drag_data["item_id"], self.drag_data["item_type"]
            data_src = self.students if item_type == "student" else self.furniture
            item_data = data_src[item_id]

            # Calculate total delta from the drag start in world coordinates
            dx_total_world = world_event_x - self.drag_data["start_x_world"]
            dy_total_world = world_event_y - self.drag_data["start_y_world"]

            orig_world_w = self.drag_data["original_size_world"]["width"]
            orig_world_h = self.drag_data["original_size_world"]["height"]

            new_world_w = orig_world_w + dx_total_world
            new_world_h = orig_world_h + dy_total_world

            min_w = MIN_STUDENT_BOX_WIDTH if item_type == "student" else 20
            min_h = MIN_STUDENT_BOX_HEIGHT if item_type == "student" else 20
            new_world_w = max(min_w, new_world_w); new_world_h = max(min_h, new_world_h)

            if item_type == "student":
                if "style_overrides" not in item_data: item_data["style_overrides"] = {}
                item_data["style_overrides"]["width"] = new_world_w; item_data["style_overrides"]["height"] = new_world_h
                item_data["width"] = new_world_w; item_data["height"] = new_world_h # Sync base
            else: item_data["width"] = new_world_w; item_data["height"] = new_world_h
            if item_type == "student": self.draw_single_student(item_id)
            else: self.draw_single_furniture(item_id)
        else: # Moving
            if not self.drag_data.get("_actual_drag_initiated"):
                # Use world coordinates for drag threshold calculation
                total_dx_world_from_start = abs(world_event_x - self.drag_data["start_x_world"])
                total_dy_world_from_start = abs(world_event_y - self.drag_data["start_y_world"])
                # Convert threshold to world units (approximately, as zoom might affect perception)
                threshold_world = DRAG_THRESHOLD / self.current_zoom_level
                if total_dx_world_from_start > threshold_world or total_dy_world_from_start > threshold_world:
                    self.drag_data["_actual_drag_initiated"] = True; self._potential_click_target = None
                else: return

            # For moving, dx_world and dy_world are deltas from the *last* event's world position
            dx_world_move = world_event_x - self.drag_data["x"]
            dy_world_move = world_event_y - self.drag_data["y"]
            dx_canvas_move = dx_world_move * self.current_zoom_level
            dy_canvas_move = dy_world_move * self.current_zoom_level
            for selected_id in self.selected_items: self.canvas.move(selected_id, dx_canvas_move, dy_canvas_move)

        self.drag_data["x"] = world_event_x # Update last world position for next delta
        self.drag_data["y"] = world_event_y
        self.password_manager.record_activity()


    def on_canvas_release(self, event):
        # ... (largely same as v51, but uses start_x_world/start_y_world for move calculations)
        if self.password_manager.is_locked: return
        clicked_item_id_at_press = self._potential_click_target
        dragged_item_id = self.drag_data.get("item_id")
        was_resizing = self.drag_data.get("is_resizing", False)
        actual_drag_initiated = self.drag_data.get("_actual_drag_initiated", False)
        world_event_x, world_event_y = self.canvas_to_world_coords(event.x, event.y)

        if was_resizing and dragged_item_id:
            item_type = self.drag_data["item_type"]
            data_src = self.students if item_type == "student" else self.furniture
            item_data = data_src[dragged_item_id]
            final_w = item_data.get("width"); final_h = item_data.get("height") # These were updated during drag
            if item_type == "student":
                final_w = item_data.get("style_overrides", {}).get("width", final_w)
                final_h = item_data.get("style_overrides", {}).get("height", final_h)
            old_w = self.drag_data["original_size_world"]["width"]; old_h = self.drag_data["original_size_world"]["height"]
            if final_w != old_w or final_h != old_h:
                size_change_info = [{'id': dragged_item_id, 'type': item_type, 'old_w': old_w, 'old_h': old_h, 'new_w': final_w, 'new_h': final_h}]
                self.execute_command(ChangeItemsSizeCommand(self, size_change_info))
            else: self.draw_all_items(check_collisions_on_redraw=True)
            self.update_status(f"Resized {item_type} '{dragged_item_id}'.")
        elif clicked_item_id_at_press and not actual_drag_initiated:
            item_type_of_clicked = "student" if clicked_item_id_at_press in self.students else "furniture"
            if self.edit_mode_var.get(): pass # No action on click in edit mode unless on handle
            elif item_type_of_clicked == "student":
                current_mode = self.mode_var.get()
                if current_mode == "quiz" and self.is_live_quiz_active: self.handle_live_quiz_tap(clicked_item_id_at_press)
                elif current_mode == "homework" and self.is_live_homework_active: self.handle_live_homework_tap(clicked_item_id_at_press) # New
                elif current_mode == "quiz": self.log_quiz_score_dialog(clicked_item_id_at_press)
                elif current_mode == "homework": self.log_homework_dialog(clicked_item_id_at_press) # New
                elif current_mode == "behavior": self.log_behavior_dialog(clicked_item_id_at_press)
        elif actual_drag_initiated and dragged_item_id and not was_resizing:
            grid_size_world = self.settings.get("grid_size", DEFAULT_GRID_SIZE); snap_to_grid = self.settings.get("grid_snap_enabled", False)
            items_moves_for_command = []
            for item_id_moved in self.selected_items:
                original_pos_info = self.drag_data.get("original_positions", {}).get(item_id_moved)
                if not original_pos_info: continue
                current_item_type = original_pos_info["type"]
                # Use total delta from drag_start_world for accurate final position
                total_dx_world = world_event_x - self.drag_data["start_x_world"]
                total_dy_world = world_event_y - self.drag_data["start_y_world"]
                new_world_x = original_pos_info["x"] + total_dx_world
                new_world_y = original_pos_info["y"] + total_dy_world
                final_x_world, final_y_world = new_world_x, new_world_y
                if snap_to_grid and grid_size_world > 0:
                    final_x_world = round(new_world_x / grid_size_world) * grid_size_world
                    final_y_world = round(new_world_y / grid_size_world) * grid_size_world
                if abs(final_x_world - original_pos_info["x"]) > 0.01 or abs(final_y_world - original_pos_info["y"]) > 0.01:
                    items_moves_for_command.append({'id': item_id_moved, 'type': current_item_type, 'old_x': original_pos_info["x"], 'old_y': original_pos_info["y"], 'new_x': final_x_world, 'new_y': final_y_world})
                else: # No change after snap, ensure it's redrawn to its original spot
                    data_s = self.students if current_item_type == "student" else self.furniture
                    data_s[item_id_moved]['x'] = original_pos_info["x"]; data_s[item_id_moved]['y'] = original_pos_info["y"]
            if items_moves_for_command: self.execute_command(MoveItemsCommand(self, items_moves_for_command))
            else: self.draw_all_items(check_collisions_on_redraw=True)
        self.drag_data.clear(); self._potential_click_target = None; self._drag_started_on_item = False; self.password_manager.record_activity()

    def on_canvas_right_press(self, event):
        # ... (same as v51, but ensure context menu for homework log is added)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Open Menu", "Enter password to open context menu:"): return
        self.canvas.focus_set()
        world_event_x, world_event_y = self.canvas_to_world_coords(event.x, event.y)
        item_canvas_ids_context = self.canvas.find_overlapping(event.x -1, event.y -1, event.x +1, event.y +1) # Screen coords
        context_item_id, context_item_type = None, None
        for item_c_id in reversed(item_canvas_ids_context):
            tags = self.canvas.gettags(item_c_id); temp_id, temp_type, is_main_rect = None, None, False
            for tag in tags:
                if tag.startswith("student_") and tag in self.students: temp_id, temp_type = tag, "student"
                elif tag.startswith("furniture_") and tag in self.furniture: temp_id, temp_type = tag, "furniture"
                if "rect" in tag: is_main_rect = True
            if temp_id and is_main_rect: context_item_id, context_item_type = temp_id, temp_type; break
        if context_item_id:
            if context_item_id not in self.selected_items:
                self.deselect_all_items(); self.selected_items.add(context_item_id)
                if context_item_type == "student": self.draw_single_student(context_item_id)
                else: self.draw_single_furniture(context_item_id)
            if context_item_type == "student": self.show_student_context_menu(event, context_item_id)
            elif context_item_type == "furniture": self.show_furniture_context_menu(event, context_item_id)
        else: self.show_general_context_menu(event)
        self.password_manager.record_activity()

    def show_general_context_menu(self, event):
        # ... (same as v51)
        context_menu = tk.Menu(self.canvas, tearoff=0)
        context_menu.add_command(label="Add Student...", command=self.add_student_dialog)
        context_menu.add_command(label="Add Furniture...", command=self.add_furniture_dialog)
        context_menu.add_separator()
        if self.selected_items:
            context_menu.add_command(label=f"Change Size of Selected ({len(self.selected_items)})...", command=self.change_size_selected_dialog)
            context_menu.add_command(label=f"Delete Selected ({len(self.selected_items)})", command=self.delete_selected_items_confirm)
            #context_menu.add_command(label=f"Assign selected ({len(self.selected_items)}) to group")
            if self.settings.get("student_groups_enabled", True) and self.student_groups and not ('furniture' in str(self.selected_items)):
                previous_group = ""
                #print(list(self.selected_items))
                for student in self.selected_items:
                    try:
                        student_data = self.students[student]; student_name = student_data["full_name"]
                        if student_data.get("group_id") == previous_group and previous_group != "":
                            pass
                        elif previous_group == "" and student_data.get("group_id"):
                            previous_group = student_data.get("group_id")
                        else:
                            previous_group = None
                            break
                    except KeyError: previous_group = None
                current_group_id = previous_group
                group_menu = tk.Menu(context_menu, tearoff=0) #; current_group_id = student_data.get("group_id")
                group_var_menu = tk.StringVar(value=current_group_id if current_group_id else "NONE_GROUP_SENTINEL")
                group_menu.add_radiobutton(label="No Group", variable=group_var_menu, value="NONE_GROUP_SENTINEL", command=lambda sid=self.selected_items: self.assign_students_to_group_via_menu(sid, None))
                for gid, gdata in sorted(self.student_groups.items(), key=lambda item: item[1]['name']):
                    group_menu.add_radiobutton(label=gdata['name'], variable=group_var_menu, value=gid, command=lambda sid=self.selected_items, new_gid=gid: self.assign_students_to_group_via_menu(sid, new_gid))
                context_menu.add_cascade(label="Assign selected to Group", menu=group_menu)
            context_menu.add_separator()
            
            if len(self.selected_items) > 1:
                align_menu = tk.Menu(context_menu, tearoff=0)
                align_menu.add_command(label="Align Top", command=lambda: self.align_selected_items("top"))
                align_menu.add_command(label="Align Bottom", command=lambda: self.align_selected_items("bottom"))
                align_menu.add_command(label="Align Left", command=lambda: self.align_selected_items("left"))
                align_menu.add_command(label="Align Right", command=lambda: self.align_selected_items("right"))
                align_menu.add_command(label="Align Horizontal Center", command=lambda: self.align_selected_items("center_h"))
                align_menu.add_command(label="Align Vertical Center", command=lambda: self.align_selected_items("center_v"))
                context_menu.add_cascade(label="Align Selected", menu=align_menu); context_menu.add_separator()
        context_menu.add_command(label="Select All Students", command=self.select_all_students)
        context_menu.add_command(label="Select All Furniture", command=self.select_all_furniture)
        context_menu.add_command(label="Select All Items", command=self.select_all_items)
        context_menu.add_command(label="Deselect All", command=self.deselect_all_items)
        try: context_menu.tk_popup(event.x_root, event.y_root)
        finally: context_menu.grab_release()

    def show_student_context_menu(self, event, student_id):
        # ... (updated for homework mode)
        if student_id not in self.students: return
        student_data = self.students[student_id]; student_name = student_data["full_name"]
        context_menu = tk.Menu(self.canvas, tearoff=0); current_mode = self.mode_var.get()
        if current_mode == "quiz":
            if self.is_live_quiz_active: context_menu.add_command(label=f"Mark Quiz for {student_name}", command=lambda: self.handle_live_quiz_tap(student_id))
            else: context_menu.add_command(label=f"Log Quiz Score for {student_name}", command=lambda: self.log_quiz_score_dialog(student_id))
            context_menu.add_command(label=f"Log Behavior for {student_name}", command=lambda: self.log_behavior_dialog(student_id))
            context_menu.add_command(label=f"Log Homework for {student_name}", command=lambda: self.log_homework_dialog(student_id)) # New
        elif current_mode == "homework": # New
            if self.is_live_homework_active: context_menu.add_command(label=f"Mark Homework for {student_name}", command=lambda: self.handle_live_homework_tap(student_id))
            else: context_menu.add_command(label=f"Log Homework for {student_name}", command=lambda: self.log_homework_dialog(student_id))
            context_menu.add_command(label=f"Log Behavior for {student_name}", command=lambda: self.log_behavior_dialog(student_id))
            context_menu.add_command(label=f"Log Quiz Score for {student_name}", command=lambda: self.log_quiz_score_dialog(student_id))
        else: # Behavior mode
            context_menu.add_command(label=f"Log Behavior for {student_name}", command=lambda: self.log_behavior_dialog(student_id))
            context_menu.add_command(label=f"Log Homework for {student_name}", command=lambda: self.log_homework_dialog(student_id)) # New
            context_menu.add_command(label=f"Log Quiz Score for {student_name}", command=lambda: self.log_quiz_score_dialog(student_id))

        context_menu.add_command(label=f"Edit {student_name}...", command=lambda: self.edit_student_dialog(student_id))
        context_menu.add_command(label=f"Customize Style for {student_name}...", command=lambda: self.customize_student_style_dialog(student_id))
        context_menu.add_command(label=f"Change Size for {student_name}...", command=lambda: self.change_item_size_dialog(student_id, "student"))
        if self.settings.get("student_groups_enabled", True) and self.student_groups:
            group_menu = tk.Menu(context_menu, tearoff=0); current_group_id = student_data.get("group_id")
            group_var_menu = tk.StringVar(value=current_group_id if current_group_id else "NONE_GROUP_SENTINEL")
            group_menu.add_radiobutton(label="No Group", variable=group_var_menu, value="NONE_GROUP_SENTINEL", command=lambda sid=student_id: self.assign_student_to_group_via_menu(sid, None))
            for gid, gdata in sorted(self.student_groups.items(), key=lambda item: item[1]['name']):
                group_menu.add_radiobutton(label=gdata['name'], variable=group_var_menu, value=gid, command=lambda sid=student_id, new_gid=gid: self.assign_student_to_group_via_menu(sid, new_gid))
            context_menu.add_cascade(label="Assign to Group", menu=group_menu)
        context_menu.add_command(label=f"Delete {student_name}", command=lambda: self.delete_student_confirm(student_id))
        context_menu.add_separator()
        if student_id in self._per_student_last_cleared: context_menu.add_command(label="Show Recent Logs (This Student)", command=lambda: self.show_recent_logs_for_student(student_id))
        else: context_menu.add_command(label="Hide Recent Logs (This Student)", command=lambda: self.clear_recent_logs_for_student(student_id))
        context_menu.add_separator()
        global_toggle_text = "Show Recent Logs (All)" if self._recent_incidents_hidden_globally or self._recent_homeworks_hidden_globally else "Hide Recent Logs (All)"
        context_menu.add_command(label=global_toggle_text, command=self.toggle_global_recent_logs_visibility)
        try: context_menu.tk_popup(event.x_root, event.y_root)
        finally: context_menu.grab_release()

    def show_furniture_context_menu(self, event, furniture_id):
        # ... (same as v51)
        if furniture_id not in self.furniture: return
        item_data = self.furniture[furniture_id]; item_name = item_data["name"]
        context_menu = tk.Menu(self.canvas, tearoff=0)
        context_menu.add_command(label=f"Edit {item_name}...", command=lambda: self.edit_furniture_dialog(furniture_id))
        context_menu.add_command(label=f"Change Size for {item_name}...", command=lambda: self.change_item_size_dialog(furniture_id, "furniture"))
        context_menu.add_command(label=f"Delete {item_name}", command=lambda: self.delete_furniture_confirm(furniture_id))
        try: context_menu.tk_popup(event.x_root, event.y_root)
        finally: context_menu.grab_release()

    def _select_items_by_type(self, item_type_key):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Select", "Enter password to select items:"): return
        self.deselect_all_items(); items_to_select = []
        if item_type_key == "students" or item_type_key == "all": items_to_select.extend(self.students.keys())
        if item_type_key == "furniture" or item_type_key == "all": items_to_select.extend(self.furniture.keys())
        for item_id in items_to_select:
            self.selected_items.add(item_id)
            if item_id in self.students: self.draw_single_student(item_id)
            elif item_id in self.furniture: self.draw_single_furniture(item_id)
        self.update_status(f"{len(self.selected_items)} items selected."); self.password_manager.record_activity()

    def select_all_students(self): self._select_items_by_type("students")
    def select_all_furniture(self): self._select_items_by_type("furniture")
    def select_all_items(self): self._select_items_by_type("all")
    def deselect_all_items(self):
        # ... (same as v51)
        if self.password_manager.is_locked: pass
        items_to_redraw = list(self.selected_items); self.selected_items.clear()
        for item_id in items_to_redraw:
            if item_id in self.students: self.draw_single_student(item_id)
            elif item_id in self.furniture: self.draw_single_furniture(item_id)
        if items_to_redraw: self.update_status("Selection cleared.")

    def change_item_size_dialog(self, item_id, item_type):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Change Size", "Enter password to change item size:"): return
        item_data_source = self.students if item_type == "student" else self.furniture
        if item_id not in item_data_source: return
        item_data = item_data_source[item_id]
        current_w = item_data.get("width", DEFAULT_STUDENT_BOX_WIDTH); current_h = item_data.get("height", DEFAULT_STUDENT_BOX_HEIGHT)
        status = False
        if item_type == "student":
            status = True
            style_overrides = item_data.get("style_overrides", {})
            current_w = style_overrides.get("width", current_w); current_h = style_overrides.get("height", current_h)
        dialog = SizeInputDialog(self.root, f"Set Size for {item_data.get('full_name', item_data.get('name'))}", current_w, current_h, status)
        if dialog.result:
            new_w, new_h = dialog.result
            min_w = MIN_STUDENT_BOX_WIDTH if item_type == "student" else 20; min_h = MIN_STUDENT_BOX_HEIGHT if item_type == "student" else 20
            final_w, final_h = max(min_w, new_w), max(min_h, new_h)
            if final_w != current_w or final_h != current_h:
                item_size_info = [{'id': item_id, 'type': item_type, 'old_w': current_w, 'old_h': current_h, 'new_w': final_w, 'new_h': final_h}]
                self.execute_command(ChangeItemsSizeCommand(self, item_size_info))
            self.password_manager.record_activity()

    def change_size_selected_dialog(self):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Change Size", "Enter password to change size of selected items:"): return
        if not self.selected_items: messagebox.showinfo("No Selection", "No items are selected to resize.", parent=self.root); return
        first_selected_id = next(iter(self.selected_items))
        item_type_first = "student" if first_selected_id in self.students else "furniture"
        item_data_source_first = self.students if item_type_first == "student" else self.furniture
        first_item_data = item_data_source_first[first_selected_id]
        current_w_default = first_item_data.get("width", DEFAULT_STUDENT_BOX_WIDTH); current_h_default = first_item_data.get("height", DEFAULT_STUDENT_BOX_HEIGHT)
        status = False
        if item_type_first == "student":
            status = True
            first_item_overrides = first_item_data.get("style_overrides", {})
            current_w_default = first_item_overrides.get("width", current_w_default); current_h_default = first_item_overrides.get("height", current_h_default)
        dialog = SizeInputDialog(self.root, f"Set Size for Selected ({len(self.selected_items)}) Items", current_w_default, current_h_default, status)
        if dialog.result:
            new_w_dialog, new_h_dialog = dialog.result; items_size_changes_for_command = []
            for item_id in self.selected_items:
                item_type_current = "student" if item_id in self.students else "furniture"
                data_src_current = self.students if item_type_current == "student" else self.furniture
                item_data_current = data_src_current[item_id]
                old_w_eff = item_data_current.get("width", DEFAULT_STUDENT_BOX_WIDTH); old_h_eff = item_data_current.get("height", DEFAULT_STUDENT_BOX_HEIGHT)
                if item_type_current == "student":
                    style_ov = item_data_current.get("style_overrides", {})
                    old_w_eff = style_ov.get("width", old_w_eff); old_h_eff = style_ov.get("height", old_h_eff)
                min_w, min_h = (MIN_STUDENT_BOX_WIDTH, MIN_STUDENT_BOX_HEIGHT) if item_type_current == "student" else (20, 20)
                final_w, final_h = max(min_w, new_w_dialog), max(min_h, new_h_dialog)
                if final_w != old_w_eff or final_h != old_h_eff:
                    items_size_changes_for_command.append({'id': item_id, 'type': item_type_current, 'old_w': old_w_eff, 'old_h': old_h_eff, 'new_w': final_w, 'new_h': final_h})
            if items_size_changes_for_command: self.execute_command(ChangeItemsSizeCommand(self, items_size_changes_for_command))
            self.password_manager.record_activity()

    def edit_student_dialog(self, student_id):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Edit", "Enter password to edit student:"): return
        student = self.students.get(student_id);
        if not student: return
        old_student_data_snapshot = student.copy()
        if "style_overrides" in old_student_data_snapshot: old_student_data_snapshot["style_overrides"] = old_student_data_snapshot["style_overrides"].copy()
        dialog = AddEditStudentDialog(self.root, f"Edit Student: {student['full_name']}", student_data=student, app=self)
        if dialog.result:
            fn, ln, nick, gend, grp_id = dialog.result; new_full_name = f"{fn} \"{nick}\" {ln}" if nick else f"{fn} {ln}"
            changes_for_command = {}
            if fn != student.get("first_name"): changes_for_command["first_name"] = fn
            if ln != student.get("last_name"): changes_for_command["last_name"] = ln
            if nick != student.get("nickname", ""): changes_for_command["nickname"] = nick
            if gend != student.get("gender", "Boy"): changes_for_command["gender"] = gend
            if grp_id != student.get("group_id"): changes_for_command["group_id"] = grp_id
            if new_full_name != student.get("full_name"): changes_for_command["full_name"] = new_full_name
            if changes_for_command: self.execute_command(EditItemCommand(self, student_id, "student", old_student_data_snapshot, changes_for_command))
            self.password_manager.record_activity()

    def edit_furniture_dialog(self, furniture_id):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Edit", "Enter password to edit furniture:"): return
        item = self.furniture.get(furniture_id);
        if not item: return
        old_item_data_snapshot = item.copy()
        dialog = AddFurnitureDialog(self.root, f"Edit Furniture: {item['name']}", furniture_data=item)
        if dialog.result:
            name, item_type, width, height = dialog.result; changes_for_command = {}
            if name != item.get("name"): changes_for_command["name"] = name
            if item_type != item.get("type"): changes_for_command["type"] = item_type
            if width != item.get("width"): changes_for_command["width"] = width
            if height != item.get("height"): changes_for_command["height"] = height
            if changes_for_command: self.execute_command(EditItemCommand(self, furniture_id, "furniture", old_item_data_snapshot, changes_for_command))
            self.password_manager.record_activity()

    def customize_student_style_dialog(self, student_id):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Customize Style", "Enter password to customize style:"): return
        student = self.students.get(student_id);
        if not student: return
        dialog = StudentStyleDialog(self.root, f"Customize Style: {student['full_name']}", student, self)
        if dialog.result:
            for prop, old_val, new_val in dialog.result:
                self.execute_command(ChangeStudentStyleCommand(self, student_id, prop, old_val, new_val))
            self.password_manager.record_activity()

    def delete_student_confirm(self, student_id):
        # ... (updated to include homework_log)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Delete", "Enter password to delete student:"): return
        if student_id not in self.students: return
        student_name = self.students[student_id]["full_name"]
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete {student_name}?\nThis will also remove their behavior, quiz, and homework log entries.", parent=self.root):
            student_data_copy = self.students[student_id].copy()
            if "style_overrides" in student_data_copy: student_data_copy["style_overrides"] = student_data_copy["style_overrides"].copy()
            associated_logs = [log.copy() for log in self.behavior_log if log["student_id"] == student_id]
            # DeleteItemCommand now handles associated_homework_logs internally
            cmd = DeleteItemCommand(self, student_id, "student", student_data_copy, associated_logs)
            self.execute_command(cmd); self.password_manager.record_activity()

    def delete_furniture_confirm(self, furniture_id):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Delete", "Enter password to delete furniture:"): return
        if furniture_id not in self.furniture: return
        item_name = self.furniture[furniture_id]["name"]
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete furniture item '{item_name}'?", parent=self.root):
            item_data_copy = self.furniture[furniture_id].copy()
            self.execute_command(DeleteItemCommand(self, furniture_id, "furniture", item_data_copy))
            self.password_manager.record_activity()

    def delete_selected_items_confirm(self):
        # ... (updated to include homework_log in message)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Delete", "Enter password to delete selected items:"): return
        if not self.selected_items: messagebox.showinfo("No Selection", "No items are selected to delete.", parent=self.root); return
        num_students = sum(1 for sid in self.selected_items if sid in self.students)
        num_furniture = sum(1 for fid in self.selected_items if fid in self.furniture)
        message = f"Are you sure you want to delete {len(self.selected_items)} selected items"
        details = []
        if num_students > 0: details.append(f"{num_students} student(s) (all logs will also be removed)") # Updated message
        if num_furniture > 0: details.append(f"{num_furniture} furniture item(s)")
        if details: message += f" ({', '.join(details)})"; message += "?"
        if messagebox.askyesno("Confirm Delete Selected", message, parent=self.root):
            commands_to_execute = []
            for item_id in list(self.selected_items):
                if item_id in self.students:
                    student_data_copy = self.students[item_id].copy()
                    if "style_overrides" in student_data_copy: student_data_copy["style_overrides"] = student_data_copy["style_overrides"].copy()
                    associated_logs = [log.copy() for log in self.behavior_log if log["student_id"] == item_id]
                    # DeleteItemCommand now handles associated_homework_logs internally
                    commands_to_execute.append(DeleteItemCommand(self, item_id, "student", student_data_copy, associated_logs))
                elif item_id in self.furniture:
                    item_data_copy = self.furniture[item_id].copy()
                    commands_to_execute.append(DeleteItemCommand(self, item_id, "furniture", item_data_copy))
            for cmd in commands_to_execute: self.execute_command(cmd)
            self.password_manager.record_activity()

    def log_behavior_dialog(self, student_id):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Log Behavior", "Enter password to log behavior:"): return
        student = self.students.get(student_id);
        if not student: return
        dialog = BehaviorDialog(self.root, f"Log Behavior for {student['full_name']}", self.all_behaviors, self.custom_behaviors)
        if dialog.result:
            behavior, comment = dialog.result
            log_entry = {"timestamp": datetime.now().isoformat(), "student_id": student_id, "student_first_name": student["first_name"],
                         "student_last_name": student["last_name"], "behavior": behavior, "comment": comment, "type": "behavior", "day": datetime.now().strftime('%A')}
            self.execute_command(LogEntryCommand(self, log_entry, student_id))
            self.draw_all_items(check_collisions_on_redraw=True); self.password_manager.record_activity()

    def log_homework_dialog(self, student_id): # New for manual homework logging
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Log Homework", "Enter password to log homework:"): return
        student = self.students.get(student_id)
        if not student: return

        dialog = ManualHomeworkLogDialog(self.root, f"Log Homework for {student['full_name']}",
                                         self.all_homework_log_behaviors, # Use specific list for manual log options
                                         self.custom_homework_log_behaviors,
                                         log_marks_enabled=self.settings.get("log_homework_marks_enabled", True),
                                         homework_mark_types=self.settings.get("homework_mark_types", DEFAULT_HOMEWORK_MARK_TYPES.copy()),
                                         homework_templates=self.homework_templates) # Pass templates
        if dialog.result:
            homework_type, comment, marks_data, num_items = dialog.result # Marks data is new
            log_entry = {
                "timestamp": datetime.now().isoformat(), "student_id": student_id,
                "student_first_name": student["first_name"], "student_last_name": student["last_name"],
                "behavior": homework_type, # Using "behavior" key for consistency with export, but it's homework_type
                "homework_type": homework_type, # Explicitly store as homework_type
                "comment": comment, "type": "homework", # Distinguish from "homework_session"
                "day": datetime.now().strftime('%A')
            }
            if self.settings.get("log_homework_marks_enabled", True) and marks_data:
                log_entry["marks_data"] = marks_data
                log_entry["num_items"] = num_items # Number of items/questions for this homework

            self.execute_command(LogHomeworkEntryCommand(self, log_entry, student_id))
            self.draw_all_items(check_collisions_on_redraw=True)
            self.password_manager.record_activity()


    def log_quiz_score_dialog(self, student_id):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Log Quiz Score", "Enter password to log quiz score:"): return
        student = self.students.get(student_id);
        if not student: return
        initial_quiz_name = self.settings.get("default_quiz_name", "Pop Quiz")
        initial_num_questions = self.settings.get("_last_used_q_num_for_session", self.settings.get("default_quiz_questions", 10))
        timeout_minutes = self.settings.get("last_used_quiz_name_timeout_minutes", 60)
        if self.last_used_quiz_name and self.last_used_quiz_name_timestamp:
            try:
                if (datetime.now() - datetime.fromisoformat(self.last_used_quiz_name_timestamp)).total_seconds() / 60 < timeout_minutes:
                    initial_quiz_name = self.last_used_quiz_name; initial_num_questions = self.initial_num_questions
            except ValueError: print("Warning: Could not parse last_used_quiz_name_timestamp.")
        dialog = QuizScoreDialog(self.root, f"Log Quiz Score for {student['full_name']}", initial_quiz_name,
                                 self.settings.get("quiz_mark_types", DEFAULT_QUIZ_MARK_TYPES.copy()), self.quiz_templates,
                                 self.settings.get("default_quiz_questions", 10), initial_num_questions)
        if dialog.result:
            quiz_name, marks_data, comment, num_questions_actual = dialog.result
            log_entry = {"timestamp": datetime.now().isoformat(), "student_id": student_id, "student_first_name": student["first_name"],
                         "student_last_name": student["last_name"], "behavior": quiz_name, "comment": comment, "marks_data": marks_data,
                         "num_questions": num_questions_actual, "type": "quiz", "day": datetime.now().strftime('%A')}
            self.execute_command(LogEntryCommand(self, log_entry, student_id))
            self.last_used_quiz_name = quiz_name; self.last_used_quiz_name_timestamp = datetime.now().isoformat(); self.initial_num_questions = num_questions_actual
            self.settings["_last_used_quiz_name_for_session"] = self.last_used_quiz_name
            self.settings["_last_used_quiz_name_timestamp_for_session"] = self.last_used_quiz_name_timestamp
            self.settings["_last_used_q_num_for_session"] = self.initial_num_questions
            self.password_manager.record_activity()

    def save_data_wrapper(self, event=None, source="manual"):
        # ... (updated to include homework_log)
        self._ensure_next_ids()
        serializable_undo_stack = [cmd.to_dict() for cmd in self.undo_stack]
        serializable_redo_stack = [cmd.to_dict() for cmd in self.redo_stack]
        data_to_save = {"students": self.students, "furniture": self.furniture, "behavior_log": self.behavior_log,
                        "homework_log": self.homework_log, # Save homework log
                        "settings": self.settings, "last_excel_export_path": self.last_excel_export_path,
                        "_per_student_last_cleared": self._per_student_last_cleared,
                        "undo_stack": serializable_undo_stack, "redo_stack": serializable_redo_stack}
        try:
            with open(DATA_FILE, 'w', encoding='utf-8') as f: json.dump(data_to_save, f, indent=4)
            verbose_save = source not in ["autosave", "command_execution", "undo_command", "redo_command", "toggle_mode",
                                          "end_live_quiz", "end_live_homework_session", "reset", "assign_group_menu", "load_template", "save_and_quit"]
            if verbose_save: self.update_status(f"Data saved to {os.path.basename(DATA_FILE)}")
            elif source == "autosave": self.update_status(f"Autosaved data at {datetime.now().strftime('%H:%M:%S')}")
        except IOError as e:
            self.update_status(f"Error saving data: {e}")
            messagebox.showerror("Save Error", f"Could not save data to {DATA_FILE}: {e}", parent=self.root)
        self.save_student_groups(); self.save_custom_behaviors(); self.save_custom_homework_log_behaviors(); self.save_custom_homework_session_types()
        self.save_quiz_templates(); self.save_homework_templates() # New


    def _migrate_v8_data(self, data): # New migration for v8 -> v9
        """Migration for data version 8 (APP_VERSION v51) to v9 (APP_VERSION v52)."""
        # Add new homework-related settings if missing
        if "settings" in data:
            data["settings"].setdefault("homework_log", []) # Initialize if not present
            data["settings"].setdefault("show_recent_homeworks_on_boxes", True)
            data["settings"].setdefault("num_recent_homeworks_to_show", 2)
            data["settings"].setdefault("recent_homework_time_window_hours", 24)
            data["settings"].setdefault("show_full_recent_homeworks", False)
            data["settings"].setdefault("reverse_homework_order", True)
            data["settings"].setdefault("selected_recent_homeworks_filter", None)
            data["settings"].setdefault("homework_initial_overrides", {})
            data["settings"].setdefault("default_homework_name", "Homework Check")
            data["settings"].setdefault("live_homework_session_mode", "Yes/No")
            data["settings"].setdefault("log_homework_marks_enabled", True)
            data["settings"].setdefault("homework_mark_types", DEFAULT_HOMEWORK_MARK_TYPES.copy())
            data["settings"].setdefault("default_homework_items_for_yes_no_mode", 5)
            data["settings"].setdefault("live_homework_score_font_color", DEFAULT_HOMEWORK_SCORE_FONT_COLOR)
            data["settings"].setdefault("live_homework_score_font_style_bold", DEFAULT_HOMEWORK_SCORE_FONT_STYLE_BOLD)
            data["settings"].setdefault("next_homework_template_id_num", 1)
            data["settings"].setdefault("next_custom_homework_type_id_num", 1)
            data["settings"].setdefault("_last_used_homework_name_for_session", "")
            data["settings"].setdefault("_last_used_homework_name_timestamp_for_session", None)
            data["settings"].setdefault("_last_used_hw_items_for_session", 5)

        # Ensure homework_log list exists at the top level of data
        if "homework_log" not in data:
            data["homework_log"] = []

        # Migrate existing behavior_log entries that might have been intended as homework
        # This is heuristic; might need adjustment based on how users previously logged homework.
        # For now, assume if behavior name contains "Homework" or "HW", it might be a homework log.
        # This is a very basic migration attempt.
        temp_behavior_log = []
        temp_homework_log = list(data.get("homework_log", [])) # Start with existing homework log

        for log_entry in data.get("behavior_log", []):
            behavior_name = log_entry.get("behavior", "").lower()
            log_type = log_entry.get("type", "behavior").lower()
            if ("homework" in behavior_name or "hw" in behavior_name) and log_type == "behavior":
                # Convert this to a homework log entry
                new_hw_entry = log_entry.copy()
                new_hw_entry["type"] = "homework" # Change type
                new_hw_entry["homework_type"] = log_entry.get("behavior") # Store original behavior as homework_type
                if "marks_data" not in new_hw_entry and self.settings.get("log_homework_marks_enabled", True):
                    # If marks are enabled but not present, add a placeholder or try to infer
                    new_hw_entry["marks_data"] = {} # Add empty marks_data
                temp_homework_log.append(new_hw_entry)
            else:
                temp_behavior_log.append(log_entry)

        data["behavior_log"] = temp_behavior_log
        data["homework_log"] = temp_homework_log
        return data

    def load_data(self, file_path=None, is_restore=False):
        # ... (updated migration chain)
        target_file = file_path or DATA_FILE
        default_settings_copy = self._get_default_settings()
        data_loaded_successfully = False

        if os.path.exists(target_file):
            try:
                with open(target_file, 'r', encoding='utf-8') as f: data = json.load(f)
                file_basename = os.path.basename(target_file)
                data_version_from_filename = None
                if "_v3" in file_basename or "_v4" in file_basename or file_basename == f"classroom_data.json": data_version_from_filename = 3
                elif "_v5" in file_basename: data_version_from_filename = 5
                elif "_v6" in file_basename: data_version_from_filename = 6
                elif "_v7" in file_basename: data_version_from_filename = 7
                elif "_v8" in file_basename: data_version_from_filename = 8 # Previous version

                if data_version_from_filename is None or data_version_from_filename <= 3:
                    print(f"Migrating data from v3/v4 format (or older) from {target_file}")
                    data = self._migrate_v3_edited_data(data); data = self._migrate_v4_data(data); data = self._migrate_v5_data(data)
                    data = self._migrate_v6_data(data); data = self._migrate_v7_data(data); data = self._migrate_v8_data(data) # Add v8 migration
                elif data_version_from_filename == 5:
                    print(f"Migrating data from v5 format from {target_file}")
                    data = self._migrate_v5_data(data); data = self._migrate_v6_data(data); data = self._migrate_v7_data(data); data = self._migrate_v8_data(data)
                elif data_version_from_filename == 6:
                    print(f"Migrating data from v6 format from {target_file}")
                    data = self._migrate_v6_data(data); data = self._migrate_v7_data(data); data = self._migrate_v8_data(data)
                elif data_version_from_filename == 7:
                    print(f"Migrating data from v7 format from {target_file}")
                    data = self._migrate_v7_data(data); data = self._migrate_v8_data(data)
                elif data_version_from_filename == 8: # New: If loading v8 data
                    print(f"Migrating data from v8 format from {target_file}")
                    data = self._migrate_v8_data(data)

                final_settings = default_settings_copy.copy(); final_settings.update(data.get("settings", {}))
                data["settings"] = final_settings
                self.students = data.get("students", {}); self.furniture = data.get("furniture", {})
                self.behavior_log = data.get("behavior_log", []); self.homework_log = data.get("homework_log", []) # Load homework log
                self.settings = data.get("settings", default_settings_copy)
                self.last_excel_export_path = data.get("last_excel_export_path", None)
                self._per_student_last_cleared = data.get("_per_student_last_cleared", {})
                self.last_used_quiz_name = self.settings.get("_last_used_quiz_name_for_session", "")
                self.last_used_quiz_name_timestamp = self.settings.get("_last_used_quiz_name_timestamp_for_session", None)
                self.initial_num_questions = self.settings.get("_last_used_q_num_for_session", 10)
                self.last_used_homework_name = self.settings.get("_last_used_homework_name_for_session", "") # New
                self.last_used_homework_name_timestamp = self.settings.get("_last_used_homework_name_timestamp_for_session", None) # New
                self.initial_num_homework_items = self.settings.get("_last_used_hw_items_for_session", 5) # New
                self.theme_style_using = self.settings.get("theme", "System") # Newer
                self.custom_canvas_color = self.settings.get("canvas_color", "Default")

                self.undo_stack.clear(); self.redo_stack.clear()
                loaded_undo_stack = data.get("undo_stack", [])
                loaded_redo_stack = data.get("redo_stack", [])
                cutoff_date_iso = (datetime.now() - timedelta(days=self.settings.get("max_undo_history_days", MAX_UNDO_HISTORY_DAYS))).isoformat()
                for cmd_data in loaded_undo_stack:
                    if cmd_data.get('timestamp', '0') >= cutoff_date_iso:
                        cmd_obj = Command.from_dict(self, cmd_data)
                        if cmd_obj: self.undo_stack.append(cmd_obj)
                for cmd_data in loaded_redo_stack:
                    if cmd_data.get('timestamp', '0') >= cutoff_date_iso:
                        cmd_obj = Command.from_dict(self, cmd_data)
                        if cmd_obj: self.redo_stack.append(cmd_obj)
                self.update_undo_redo_buttons_state()
                self.password_manager = PasswordManager(self.settings) # Re-initialize with loaded settings
                self.update_lock_button_state()
                data_loaded_successfully = True
            except (json.JSONDecodeError, KeyError, IOError, TypeError) as e:
                print(f"Error loading data from {target_file}: {e}. Using defaults or attempting recovery.")
                if is_restore:
                    messagebox.showerror("Restore Error", f"Failed to load restored data from {target_file}: {e}\n\nApplication will use default data or attempt to load the standard data file.", parent=self.root)
                    # Fallback to loading standard DATA_FILE if restore failed and target_file was not DATA_FILE
                    if target_file != DATA_FILE: self.load_data(DATA_FILE, is_restore=False); return
                else:
                    messagebox.showwarning("Load Error", f"Error loading data file: {e}.\nDefault settings and empty classroom will be used.", parent=self.root)
                self.students, self.furniture, self.behavior_log, self.homework_log = {}, {}, [], []
                self.settings = default_settings_copy.copy()
                self.last_excel_export_path, self._per_student_last_cleared = None, {}
                self.undo_stack.clear(); self.redo_stack.clear()
        else:
            if not is_restore: print(f"Data file {target_file} not found. Using default settings and empty classroom.")
            self.students, self.furniture, self.behavior_log, self.homework_log = {}, {}, [], []
            self.settings = default_settings_copy.copy()
            self.last_excel_export_path, self._per_student_last_cleared = None, {}
            self.undo_stack.clear(); self.redo_stack.clear()

        # Ensure essential settings are present if a very old or corrupted file was loaded
        for key, value in default_settings_copy.items():
            if key not in self.settings: self.settings[key] = value
        
        # Ensure next ID counters are robustly initialized/updated after data load
        self._ensure_next_ids()
        if self.theme_style_using != "System":
            try:
                sv_ttk.set_theme(self.theme_style_using)
            except: pass
        else:
            try:
                sv_ttk.set_theme(darkdetect.theme())
            except:
                pass
        if data_loaded_successfully and not is_restore and file_path is None and \
           (os.path.basename(DATA_FILE) != f"classroom_data_{CURRENT_DATA_VERSION_TAG}.json" or \
            (data_version_from_filename is not None and data_version_from_filename < int(CURRENT_DATA_VERSION_TAG[1:]))):
            # If the main data file was from an older version, save it immediately in the new version format
            print(f"Data file loaded from an older version ({file_basename}). Saving in new format: classroom_data_{CURRENT_DATA_VERSION_TAG}.json")
            self.save_data_wrapper(source="migration_save")
            # Optionally, attempt to delete the old version file if migration was successful
            # Be cautious with this, maybe offer as a user option later
            if os.path.exists(target_file) and target_file != DATA_FILE:
                try:
                    # os.remove(target_file)
                    # print(f"Old data file {target_file} removed after successful migration.")
                    print(f"Old data file {target_file} can be manually removed if no longer needed.")
                except OSError as e_del:
                    print(f"Could not remove old data file {target_file}: {e_del}")


    def _migrate_v7_data(self, data):
        """Migration for data version 7 (APP_VERSION v50) to v8 (APP_VERSION v51)."""
        # For v50 -> v51 (data v7 -> v8)
        if "settings" in data:
            data["settings"].setdefault("student_groups_enabled", True)
            data["settings"].setdefault("show_zoom_level_display", True)
            data["settings"].setdefault("next_group_id_num", 1) # Initialize if missing
            data["settings"].setdefault("next_quiz_template_id_num", 1) # Initialize if missing
            data["settings"].setdefault("available_fonts", sorted(list(tkfont.families())))
            data["settings"].setdefault("default_quiz_questions", 10)
            data["settings"].setdefault("quiz_score_calculation", "percentage")
            data["settings"].setdefault("combine_marks_for_display", True)

            # Password settings
            data["settings"].setdefault("app_password_hash", None)
            data["settings"].setdefault("password_on_open", False)
            data["settings"].setdefault("password_on_edit_action", False)
            data["settings"].setdefault("password_auto_lock_enabled", False)
            data["settings"].setdefault("password_auto_lock_timeout_minutes", 15)
             # Quiz session internal state storage
            data["settings"].setdefault("_last_used_quiz_name_for_session", "")
            data["settings"].setdefault("_last_used_quiz_name_timestamp_for_session", None)
            data["settings"].setdefault("_last_used_q_num_for_session", 10)

            # Quiz Mark Types (ensure all fields are present)
            migrated_mark_types = []
            default_mark_type_map = {d["id"]: d for d in DEFAULT_QUIZ_MARK_TYPES}
            for mt in data["settings"].get("quiz_mark_types", []):
                # If it's an old string-based mark type, skip or attempt conversion if possible (not done here)
                if isinstance(mt, dict) and "id" in mt and "name" in mt:
                    default_entry = default_mark_type_map.get(mt["id"])
                    if default_entry: # Use defaults for missing fields if id matches a default
                        migrated_mt = default_entry.copy()
                        migrated_mt.update(mt) # Override with user's existing values
                    else: # Custom entry, ensure all fields are present
                        migrated_mt = mt.copy()
                        migrated_mt.setdefault("contributes_to_total", True)
                        migrated_mt.setdefault("is_extra_credit", False)
                        migrated_mt.setdefault("default_points", 1 if mt.get("name","").lower() == "correct" else (0 if mt.get("name","").lower() == "incorrect" else 0.5))
                    migrated_mark_types.append(migrated_mt)
            if not migrated_mark_types and not data["settings"].get("quiz_mark_types"): # If empty or missing
                 data["settings"]["quiz_mark_types"] = DEFAULT_QUIZ_MARK_TYPES.copy()
            elif migrated_mark_types:
                 data["settings"]["quiz_mark_types"] = migrated_mark_types


        if "students" in data:
            for student_id, student_data in data["students"].items():
                student_data.setdefault("group_id", None)
                student_data.setdefault("style_overrides", {})

        # Remove obsolete _undo_history_file from settings if present
        if "settings" in data and "_undo_history_file" in data["settings"]:
            del data["settings"]["_undo_history_file"]

        # Migrate undo/redo stack if it exists (it might be from a very old direct load)
        if "undo_stack" not in data: data["undo_stack"] = []
        if "redo_stack" not in data: data["redo_stack"] = []

        print("Applied v7 (to v8) data migration.")
        return data

    def _migrate_v6_data(self, data):
        """Migration for data version 6 (APP_VERSION v0.6.x) to v7 (APP_VERSION v50)."""
        # For v0.6.x -> v50 (data v6 -> v7)
        if "settings" in data:
            data["settings"].setdefault("conditional_formatting_rules", [])
            data["settings"].setdefault("live_quiz_score_font_color", DEFAULT_QUIZ_SCORE_FONT_COLOR)
            data["settings"].setdefault("live_quiz_score_font_style_bold", DEFAULT_QUIZ_SCORE_FONT_STYLE_BOLD)
            data["settings"].setdefault("quiz_mark_types", DEFAULT_QUIZ_MARK_TYPES) # Previously just strings
            data["settings"].setdefault("last_used_quiz_name_timeout_minutes", 60)
            data["settings"].setdefault("show_recent_incidents_during_quiz", True)

        if "behavior_log" in data:
            for log_entry in data["behavior_log"]:
                log_entry.setdefault("type", "behavior") # Old logs are behavior
                if log_entry["type"] == "quiz" and "marks_data" not in log_entry:
                    # Simple migration for old quiz logs to new marks_data structure
                    score_str = log_entry.get("score", "") # e.g., "7/10"
                    marks = {}
                    if score_str and "/" in score_str:
                        try:
                            correct, total = map(int, score_str.split('/'))
                            marks[DEFAULT_QUIZ_MARK_TYPES[0]["id"]] = correct # Assume first type is "Correct"
                            marks[DEFAULT_QUIZ_MARK_TYPES[1]["id"]] = total - correct # Assume second is "Incorrect"
                            log_entry["num_questions"] = total
                        except ValueError:
                            log_entry["num_questions"] = log_entry.get("num_questions", 0) # Keep if already exists
                    log_entry["marks_data"] = marks

        print("Applied v6 (to v7) data migration.")
        return data

    def _migrate_v5_data(self, data):
        """Migration for data version 5 (APP_VERSION v0.5.x) to v6 (APP_VERSION v0.6.x)."""
        if "settings" in data:
            data["settings"].setdefault("max_undo_history_days", MAX_UNDO_HISTORY_DAYS)
            data["settings"].setdefault("selected_recent_behaviors_filter", None) # None means all
            data["settings"].setdefault("current_mode", "behavior")
            data["settings"].setdefault("default_quiz_name", "Pop Quiz")
        if "students" in data:
            for student_id, student_data in data["students"].items():
                student_data.setdefault("nickname", "")
                student_data.setdefault("gender", "Boy") # Default to Boy if missing
        if "_per_student_last_cleared" not in data:
            data["_per_student_last_cleared"] = {}
        if "undo_stack" not in data: data["undo_stack"] = []
        if "redo_stack" not in data: data["redo_stack"] = []
        print("Applied v5 (to v6) data migration.")
        return data

    def _migrate_v4_data(self, data):
        """Migration for data version 4 (APP_VERSION v0.4.x) to v5."""
        if "settings" in data:
            data["settings"].setdefault("reverse_incident_order", True) # Default new setting
            data["settings"].setdefault("show_full_recent_incidents", False)
        print("Applied v4 (to v5) data migration.")
        return data

    def _migrate_v3_edited_data(self, data):
        """Migration for data version 3 and older (APP_VERSION < v0.3.x) to v4."""
        if "settings" in data:
            data["settings"].setdefault("show_recent_incidents_on_boxes", True)
            data["settings"].setdefault("num_recent_incidents_to_show", 2)
            data["settings"].setdefault("recent_incident_time_window_hours", 24)
            data["settings"].setdefault("autosave_interval_ms", 30000)
            data["settings"].setdefault("default_student_box_width", DEFAULT_STUDENT_BOX_WIDTH)
            data["settings"].setdefault("default_student_box_height", DEFAULT_STUDENT_BOX_HEIGHT)
            data["settings"].setdefault("student_box_fill_color", DEFAULT_BOX_FILL_COLOR)
            data["settings"].setdefault("student_box_outline_color", DEFAULT_BOX_OUTLINE_COLOR)
            data["settings"].setdefault("student_font_family", DEFAULT_FONT_FAMILY)
            data["settings"].setdefault("student_font_size", DEFAULT_FONT_SIZE)
            data["settings"].setdefault("student_font_color", DEFAULT_FONT_COLOR)
            data["settings"].setdefault("grid_snap_enabled", False)
            data["settings"].setdefault("grid_size", DEFAULT_GRID_SIZE)
            data["settings"].setdefault("behavior_initial_overrides", {})

            # Student ID migration: from simple numbers to student_NUM
            # Furniture ID migration: from simple numbers to furniture_NUM
            if "next_student_id" in data["settings"]:
                data["settings"]["next_student_id_num"] = data["settings"].pop("next_student_id")
            if "next_furniture_id" in data["settings"]:
                data["settings"]["next_furniture_id_num"] = data["settings"].pop("next_furniture_id")

        migrated_students = {}
        if "students" in data:
            for k, v in data["students"].items():
                new_id = f"student_{k}" if not str(k).startswith("student_") else str(k)
                v["id"] = new_id # Ensure 'id' field matches key
                if 'name' in v and 'full_name' not in v: v['full_name'] = v['name'] # Old format had 'name'
                if 'first_name' not in v or 'last_name' not in v: # Try to split 'full_name'
                    parts = v.get('full_name', '').split(' ', 1)
                    v['first_name'] = parts[0]
                    v['last_name'] = parts[1] if len(parts) > 1 else "Lastname"
                v.setdefault("nickname", "")
                v.setdefault("gender", "Boy")
                v.setdefault("width", data.get("settings", {}).get("default_student_box_width", DEFAULT_STUDENT_BOX_WIDTH))
                v.setdefault("height", data.get("settings", {}).get("default_student_box_height", DEFAULT_STUDENT_BOX_HEIGHT))
                migrated_students[new_id] = v
        data["students"] = migrated_students

        migrated_furniture = {}
        if "furniture" in data:
            for k, v in data["furniture"].items():
                new_id = f"furniture_{k}" if not str(k).startswith("furniture_") else str(k)
                v["id"] = new_id
                v.setdefault("width", REBBI_DESK_WIDTH) # Example default, adjust as needed
                v.setdefault("height", REBBI_DESK_HEIGHT)
                migrated_furniture[new_id] = v
        data["furniture"] = migrated_furniture

        if "behavior_log" in data:
            for entry in data["behavior_log"]:
                if "student_id" in entry and not str(entry["student_id"]).startswith("student_"):
                    entry["student_id"] = f"student_{entry['student_id']}"
                if "student_name" in entry and ("student_first_name" not in entry or "student_last_name" not in entry):
                    parts = entry["student_name"].split(" ", 1)
                    entry["student_first_name"] = parts[0]
                    entry["student_last_name"] = parts[1] if len(parts) > 1 else ""
                entry.setdefault("day", datetime.fromisoformat(entry["timestamp"]).strftime('%A') if "timestamp" in entry else "Unknown")
        print("Applied v3 (to v4) data migration (includes older formats).")
        return data

    def autosave_data_wrapper(self):
        self.save_data_wrapper(source="autosave")
        if hasattr(self, 'autosave_excel_log') and callable(self.autosave_excel_log):
             self.autosave_excel_log() # Call autosave for Excel if it exists
        self.root.after(self.settings.get("autosave_interval_ms", 30000), self.autosave_data_wrapper)

    def autosave_excel_log(self):
        if self.settings.get("enable_excel_autosave", False):
            if not self.behavior_log and not self.homework_log: # Don't save if empty
                # print("Autosave Excel: Log is empty, skipping.")
                return

            filename = AUTOSAVE_EXCEL_FILE
            #try:
            # Use current settings for export filters if they exist
            filter_settings = {
                "start_date": None, "end_date": None,
                "selected_students": "all", "student_ids": [],
                "selected_behaviors": "all", "behaviors_list": [],
                "selected_homework_types": "all", "homework_types_list": [], # New
                "include_behavior_logs": True,
                "include_quiz_logs": True,
                "include_homework_logs": True, # New
                "include_summaries": self.settings.get("excel_export_include_summaries_by_default", True),
                "separate_sheets_by_log_type": self.settings.get("excel_export_separate_sheets_by_default", True)
            }
            self.export_data_to_excel(filename, "xlsx", filter_settings, is_autosave=True)
                # self.update_status(f"Log autosaved to {os.path.basename(filename)} at {datetime.now().strftime('%H:%M:%S')}")
            #except Exception as e:
            #    print(f"Error during Excel autosave: {e}")
            #   # self.update_status(f"Error during Excel autosave: {e}")

    def load_custom_behaviors(self):
        if os.path.exists(CUSTOM_BEHAVIORS_FILE):
            try:
                with open(CUSTOM_BEHAVIORS_FILE, 'r', encoding='utf-8') as f:
                    self.custom_behaviors = json.load(f)
            except (json.JSONDecodeError, IOError) as e:
                print(f"Error loading custom behaviors: {e}")
                self.custom_behaviors = []
        else: self.custom_behaviors = []
    def save_custom_behaviors(self):
        try:
            with open(CUSTOM_BEHAVIORS_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.custom_behaviors, f, indent=4)
        except IOError as e: print(f"Error saving custom behaviors: {e}")

    def load_custom_homework_log_behaviors(self): # New
        if os.path.exists(CUSTOM_HOMEWORKS_FILE): # Deprecated, now CUSTOM_HOMEWORK_LOG_BEHAVIORS_FILE
            try:
                with open(CUSTOM_HOMEWORKS_FILE, 'r', encoding='utf-8') as f:
                    loaded_data = json.load(f)
                    # Check if it's the old list of strings format or new list of dicts
                    if loaded_data and isinstance(loaded_data[0], str): # Old format
                        self.custom_homework_log_behaviors = [{"name": name} for name in loaded_data]
                        self.save_custom_homework_log_behaviors() # Save in new format
                    else:
                        self.custom_homework_log_behaviors = loaded_data
            except (json.JSONDecodeError, IOError, IndexError) as e:
                print(f"Error loading custom homework log behaviors from {CUSTOM_HOMEWORKS_FILE}: {e}")
                self.custom_homework_log_behaviors = []
        else:
            self.custom_homework_log_behaviors = [] # e.g., [{"name": "Project A Submitted"}, {"name": "Reading Log Signed"}]

    def save_custom_homework_log_behaviors(self): # New
        try:
            with open(CUSTOM_HOMEWORKS_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.custom_homework_log_behaviors, f, indent=4)
        except IOError as e:
            print(f"Error saving custom homework log behaviors: {e}")

    def load_custom_homework_session_types(self): # New for "Yes/No" mode in Live Homework
        # This will store list of dicts like: [{"id": "hwtype_1", "name": "Reading Assignment"}, {"id": "hwtype_2", "name": "Math Worksheet"}]
        # Path: CUSTOM_HOMEWORK_SESSION_TYPES_FILE_PATTERN
        file_path = get_app_data_path(f"custom_homework_session_types_{CURRENT_DATA_VERSION_TAG}.json")
        if os.path.exists(file_path):
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    self.custom_homework_session_types = json.load(f)
                    # Ensure next ID num is updated
                    max_id = 0
                    for ht in self.custom_homework_session_types:
                        if isinstance(ht, dict) and ht.get('id','').startswith("hwtype_"):
                            try: max_id = max(max_id, int(ht['id'].split("_")[1]))
                            except: pass
                    self.settings["next_custom_homework_type_id_num"] = max(self.settings.get("next_custom_homework_type_id_num",1), max_id + 1)

            except (json.JSONDecodeError, IOError) as e:
                print(f"Error loading custom homework session types: {e}")
                self.custom_homework_session_types = []
        else:
            self.custom_homework_session_types = []

    def save_custom_homework_session_types(self): # New
        file_path = get_app_data_path(f"custom_homework_session_types_{CURRENT_DATA_VERSION_TAG}.json")
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(self.custom_homework_session_types, f, indent=4)
        except IOError as e:
            print(f"Error saving custom homework session types: {e}")

    def load_student_groups(self):
        if os.path.exists(STUDENT_GROUPS_FILE):
            try:
                with open(STUDENT_GROUPS_FILE, 'r', encoding='utf-8') as f: self.student_groups = json.load(f)
                max_g_id = 0
                for gid in self.student_groups:
                    if gid.startswith("group_"):
                        try: max_g_id = max(max_g_id, int(gid.split("_")[1]))
                        except (ValueError, IndexError): pass
                self.next_group_id_num = max(self.settings.get("next_group_id_num",1), max_g_id + 1)
                self.settings["next_group_id_num"] = self.next_group_id_num

            except (json.JSONDecodeError, IOError) as e:
                print(f"Error loading student groups: {e}"); self.student_groups = {}
        else: self.student_groups = {}
    def save_student_groups(self):
        try:
            with open(STUDENT_GROUPS_FILE, 'w', encoding='utf-8') as f: json.dump(self.student_groups, f, indent=4)
        except IOError as e: print(f"Error saving student groups: {e}")

    def load_quiz_templates(self):
        if os.path.exists(QUIZ_TEMPLATES_FILE):
            try:
                with open(QUIZ_TEMPLATES_FILE, 'r', encoding='utf-8') as f: self.quiz_templates = json.load(f)
                max_qt_id = 0
                for qtid in self.quiz_templates:
                    if qtid.startswith("quiztemplate_"):
                        try: max_qt_id = max(max_qt_id, int(qtid.split("_")[1]))
                        except (ValueError, IndexError): pass
                self.next_quiz_template_id_num = max(self.settings.get("next_quiz_template_id_num",1), max_qt_id + 1)
                self.settings["next_quiz_template_id_num"] = self.next_quiz_template_id_num
            except (json.JSONDecodeError, IOError) as e:
                print(f"Error loading quiz templates: {e}"); self.quiz_templates = {}
        else: self.quiz_templates = {}
    def save_quiz_templates(self):
        try:
            with open(QUIZ_TEMPLATES_FILE, 'w', encoding='utf-8') as f: json.dump(self.quiz_templates, f, indent=4)
        except IOError as e: print(f"Error saving quiz templates: {e}")

    def load_homework_templates(self): # New
        if os.path.exists(HOMEWORK_TEMPLATES_FILE):
            try:
                with open(HOMEWORK_TEMPLATES_FILE, 'r', encoding='utf-8') as f:
                    self.homework_templates = json.load(f)
                max_ht_id = 0
                for htid in self.homework_templates:
                    if htid.startswith("hwtemplate_"):
                        try: max_ht_id = max(max_ht_id, int(htid.split("_")[1]))
                        except (ValueError, IndexError): pass
                self.next_homework_template_id_num = max(self.settings.get("next_homework_template_id_num",1), max_ht_id + 1)
                self.settings["next_homework_template_id_num"] = self.next_homework_template_id_num
            except (json.JSONDecodeError, IOError) as e:
                print(f"Error loading homework templates: {e}")
                self.homework_templates = {}
        else:
            self.homework_templates = {}

    def save_homework_templates(self): # New
        try:
            with open(HOMEWORK_TEMPLATES_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.homework_templates, f, indent=4)
        except IOError as e:
            print(f"Error saving homework templates: {e}")


    def update_all_behaviors(self):
        #self.all_behaviors = DEFAULT_BEHAVIORS_LIST + [b["name"] for b in self.custom_behaviors if isinstance(b,dict) and "name" in b else b] # Handle old string list
        self.all_behaviors = DEFAULT_BEHAVIORS_LIST + [b["name"] if isinstance(b, dict) and "name" in b else str(b) for b in self.custom_behaviors]
    def update_all_homework_log_behaviors(self): # New
        self.all_homework_log_behaviors = DEFAULT_HOMEWORK_LOG_BEHAVIORS + [b["name"] for b in self.custom_homework_log_behaviors if "name" in b]
    def update_all_homework_session_types(self): # New
        # Combines default types (strings) with custom types (dicts with 'name' and 'id')
        # For display in dialogs, we'll need the names. For internal use, IDs are important for custom ones.
        default_as_dicts = [{"id": f"default_{name.lower().replace(' ','_')}", "name": name} for name in DEFAULT_HOMEWORK_SESSION_BUTTONS2]
        default_as_dicts2 = [name for name in (DEFAULT_HOMEWORK_TYPES_LIST)]
        plus_names = [ct.get("name") for ct in self.custom_homework_session_types if isinstance(ct, dict) and "name" in ct and "id" in ct]
        
        
        self.all_homework_session_types = default_as_dicts + [ct for ct in self.custom_homework_session_types if isinstance(ct, dict) and "name" in ct and "id" in ct]
        
        self.all_homework_session_types2 = default_as_dicts2 + plus_names
        
        #print(self.custom_homework_session_types, [ct.get("name") for ct in self.custom_homework_session_types if isinstance(ct, dict) and "name" in ct and "id" in ct] )

    def export_log_dialog_with_filter(self, export_type="xlsx"):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Export", "Enter password to export log data:"): return

        dialog = ExportFilterDialog(self.root, self.students, self.all_behaviors,
                                    self.all_homework_session_types + self.all_homework_log_behaviors, # Combine all possible homework type names for filter
                                    default_settings=self.settings)
        if dialog.result:
            filter_settings = dialog.result
            default_filename = f"behavior_log_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            if export_type == "xlsx":
                file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_filename + ".xlsx",
                                                       filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")], parent=self.root)
            elif export_type == "xlsm":
                file_path = filedialog.asksaveasfilename(defaultextension=".xlsm", initialfile=default_filename + ".xlsm",
                                                       filetypes=[("Excel Macro-Enabled files", "*.xlsm"), ("All files", "*.*")], parent=self.root)
            elif export_type == "csv":
                 file_path = filedialog.asksaveasfilename(defaultextension=".zip", initialfile=default_filename + "_csv.zip",
                                                       filetypes=[("ZIP archives", "*.zip"), ("All files", "*.*")], parent=self.root)
            else: return

            if file_path:
                try:
                    if export_type in ["xlsx", "xlsm"]:
                        self.export_data_to_excel(file_path, export_type, filter_settings)
                    elif export_type == "csv":
                        self.export_data_to_csv_zip(file_path, filter_settings)

                    self.last_excel_export_path = file_path # Store path even for CSV for "Open Last Export Folder"
                    self.update_open_last_export_folder_menu_item()
                    self.save_data_wrapper(source="export_log")
                    self.update_status(f"Log exported to {os.path.basename(file_path)}")
                    if messagebox.askyesno("Export Successful", f"Log exported successfully to:\n{file_path}\n\nDo you want to open the file location?", parent=self.root):
                        self.open_last_export_folder()
                except Exception as e:
                    messagebox.showerror("Export Error", f"Failed to export log: {e}", parent=self.root)
                    self.update_status(f"Error exporting log: {e}")
                    print(f"Error: {e}")
            else: self.update_status("Export cancelled.")
            self.password_manager.record_activity()


    def _make_safe_sheet_name(self, name_str, id_fallback="Sheet"):
        invalid_chars = r'[\\/?*\[\]:]' # Excel invalid sheet name characters
        safe_name = re.sub(invalid_chars, '_', str(name_str))
        if not safe_name: safe_name = str(id_fallback)
        return safe_name[:31] # Max 31 chars for sheet names



    def export_data_to_excel(self, file_path, export_format="xlsx", filter_settings=None, is_autosave=False, export_all_students_info = True):
        # ... (substantially updated for new log types, summaries, and filtering)
        wb = Workbook()
        wb.remove(wb.active) # Remove default sheet
        mark_type_configs = self.settings.get("quiz_mark_types", [])
        mark_type_configs_h = self.settings.get("homework_mark_types", [])
        quiz_mark_type_headers = [mt["name"] for mt in mark_type_configs]
        homework_mark_type_headers = [mt["name"] for mt in mark_type_configs_h]
        homework_session_types_headers = [mt["name"] for mt in self.all_homework_session_types]
        
        #print(filter_settings)

        student_data_for_export = {sid: {"first_name": s["first_name"], "last_name": s["last_name"], "full_name": s["full_name"]} for sid, s in self.students.items()}

        logs_to_process = []
        if filter_settings.get("include_behavior_logs", True):
            logs_to_process.extend([log for log in self.behavior_log if log.get("type") == "behavior"])
        if filter_settings.get("include_quiz_logs", True):
            logs_to_process.extend([log for log in self.behavior_log if log.get("type") == "quiz"])
        if filter_settings.get("include_homework_logs", True): # New
            logs_to_process.extend([log for log in self.homework_log if log.get("type") == "homework" or log.get("type") == "homework_session_y" or log.get("type") == "homework_session_s"])
        #print(self.homework_log)
        # Apply filters
        filtered_stud_ids = set()
        filtered_log = []
        start_date = filter_settings.get("start_date")
        end_date = filter_settings.get("end_date")
        selected_students_option = filter_settings.get("selected_students", "all")
        student_ids_filter = filter_settings.get("student_ids", [])
        selected_behaviors_option = filter_settings.get("selected_behaviors", "all")
        behaviors_list_filter = filter_settings.get("behaviors_list", [])
        selected_homework_types_option = filter_settings.get("selected_homework_types", "all") # New
        homework_types_list_filter = filter_settings.get("homework_types_list", []) # New
        #print("hi",selected_homework_types_option)
        for entry in logs_to_process:
            try:
                entry_date = datetime.fromisoformat(entry["timestamp"]).date()
                if start_date and entry_date < start_date: continue
                if end_date and entry_date > end_date: continue
            except ValueError: continue # Skip if timestamp is invalid

            if selected_students_option == "specific" and entry["student_id"] not in student_ids_filter: continue
            filtered_stud_ids.add(entry["student_id"])
            log_type = entry.get("type", "behavior")
            
            entry_name_field = entry.get("behavior") # Default for behavior and quiz
            if log_type == "homework" or log_type == "homework_session_y" or log_type == "homework_session_s":
                entry_name_field = entry.get("homework_type", entry.get("behavior")) # For homework logs
                #entry_name_field2 = entry.get("home")
            #print("list", homework_types_list_filter)
            #print("entry", entry_name_field)
            if log_type == "behavior" or log_type == "quiz":
                if selected_behaviors_option == "specific" and entry_name_field not in behaviors_list_filter: continue
            elif log_type == "homework" or log_type == "homework_session_s":
                if selected_homework_types_option == "specific" and entry_name_field not in homework_types_list_filter: continue
                elif selected_homework_types_option == "specific" and entry_name_field in homework_types_list_filter: continue
            elif log_type == "homework_session_y":
                pass
            filtered_log.append(entry)

        filtered_log.sort(key=lambda x: x["timestamp"])

        # Determine sheet strategy
        separate_sheets = filter_settings.get("separate_sheets_by_log_type", True)
        sheets_data = {} # {sheet_name: [entries]}

        if separate_sheets:
            if filter_settings.get("include_behavior_logs", True): sheets_data["Behavior Log"] = []
            if filter_settings.get("include_quiz_logs", True): sheets_data["Quiz Log"] = []
            if filter_settings.get("include_homework_logs", True): sheets_data["Homework Log"] = [] # New
            for entry in filtered_log:
                log_type = entry.get("type")
                if log_type == "behavior" and "Behavior Log" in sheets_data: sheets_data["Behavior Log"].append(entry)
                elif log_type == "quiz" and "Quiz Log" in sheets_data: sheets_data["Quiz Log"].append(entry)
                elif (log_type == "homework" or log_type == "homework_session_y" or log_type == "homework_session_s") and "Homework Log" in sheets_data: sheets_data["Homework Log"].append(entry)
        else:
            sheets_data["Combined Log"] = filtered_log


        bold_font = OpenpyxlFont(bold=True)
        center_alignment = OpenpyxlAlignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = OpenpyxlAlignment(horizontal='left', vertical='center', wrap_text=True)
        right_alignment = OpenpyxlAlignment(horizontal='right', vertical='center', wrap_text=False)

        for sheet_name, entries_for_sheet in sheets_data.items():
            if not entries_for_sheet and (sheet_name != "Combined Log" or not filtered_log) : continue # Skip empty specific sheets

            ws = wb.create_sheet(title=sheet_name)
            headers = ["Timestamp", "Date", "Time", "Day", "Student ID", "First Name", "Last Name"]
            if sheet_name == "Behavior Log" or not separate_sheets: headers.append("Behavior")
            if sheet_name == "Quiz Log" or not separate_sheets:
                headers.extend(["Quiz Name", "Num Questions"])
                # Add headers for each mark type (e.g., Correct, Incorrect, Bonus)
                for mt in self.settings.get("quiz_mark_types", []): headers.append(mt["name"])
                headers.append("Quiz Score (%)")
            if sheet_name == "Homework Log" or not separate_sheets: # New headers for Homework
                headers.extend(["Homework Type/Session Name", "Num Items"])
                # Add headers for each homework mark type
                for hmt in self.settings.get("homework_mark_types", []): headers.append(hmt["name"])
                headers.extend(["Homework Score (Total Pts)", "Homework Effort"]) # Example summary fields
                headers.extend(homework_session_types_headers)
            headers.append("Comment")
            if not separate_sheets: headers.append("Log Type")


            for col_num, header_title in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header_title)
                cell.font = bold_font; cell.alignment = center_alignment
            ws.freeze_panes = 'A2'

            row_num = 2
            for entry in entries_for_sheet:
                student_info = student_data_for_export.get(entry["student_id"], {"first_name": "N/A", "last_name": "N/A"})
                try: dt_obj = datetime.fromisoformat(entry["timestamp"])
                except ValueError: dt_obj = datetime.now() # Fallback
                col_num = 1
                ws.cell(row=row_num, column=col_num, value=entry["timestamp"]); col_num+=1
                ws.cell(row=row_num, column=col_num, value=dt_obj.strftime('%Y-%m-%d')).alignment = right_alignment; col_num+=1
                ws.cell(row=row_num, column=col_num, value=dt_obj.strftime('%H:%M:%S')).alignment = right_alignment; col_num+=1
                ws.cell(row=row_num, column=col_num, value=entry.get("day", dt_obj.strftime('%A'))); col_num+=1
                ws.cell(row=row_num, column=col_num, value=entry["student_id"]); col_num+=1
                ws.cell(row=row_num, column=col_num, value=student_info["first_name"]); col_num+=1
                ws.cell(row=row_num, column=col_num, value=student_info["last_name"]); col_num+=1

                entry_type = entry.get("type", "behavior")

                if sheet_name == "Behavior Log" or (not separate_sheets and entry_type == "behavior"):
                    ws.cell(row=row_num, column=col_num, value=entry.get("behavior")); col_num+=1
                elif sheet_name == "Quiz Log" or (not separate_sheets and entry_type == "quiz"):
                    ws.cell(row=row_num, column=col_num, value=entry.get("behavior")); col_num+=1 # Quiz Name
                    num_q = entry.get("num_questions", 0); ws.cell(row=row_num, column=col_num, value=num_q).alignment = right_alignment; col_num+=1
                    marks_data = entry.get("marks_data", {})
                    total_possible_points_for_calc = 0; total_earned_points_for_calc = 0; extra_credit_earned = 0
                    for mt in self.settings.get("quiz_mark_types", []):
                        points = marks_data.get(mt["id"], 0)
                        ws.cell(row=row_num, column=col_num, value=points).alignment = right_alignment; col_num+=1
                        if mt.get("contributes_to_total", True): total_possible_points_for_calc += mt.get("default_points",1) * num_q # Simplified: assumes each question can get this mark type
                        if points > 0 : # Only add earned if student got this mark
                            if mt.get("is_extra_credit", False): extra_credit_earned += points * mt.get("default_points",1)
                            else: total_earned_points_for_calc += points * mt.get("default_points",1)
                    # More robust score calculation needed based on how num_questions and marks_data relate
                    score_percent = 0
                    if num_q > 0: # Use num_questions as the basis for total possible points from main Qs
                        # Calculate total possible for main questions based on default points of contributing mark types
                        # This is a simplification; assumes each question has a potential max based on one 'correct' type
                        main_q_total_possible = 0
                        correct_type = next((m for m in self.settings.get("quiz_mark_types",[]) if m.get("id") == "mark_correct"), None)
                        if correct_type: main_q_total_possible = correct_type.get("default_points", 1) * num_q

                        if main_q_total_possible > 0:
                            score_percent = ((total_earned_points_for_calc + extra_credit_earned) / main_q_total_possible) * 100
                        elif total_earned_points_for_calc + extra_credit_earned > 0 : # Scored only on EC or non-standard
                            score_percent = 100 # Or some other representation
                    ws.cell(row=row_num, column=col_num, value=round(score_percent,2) if score_percent else "").alignment = right_alignment; col_num+=1
                elif sheet_name == "Homework Log" or (not separate_sheets and (entry_type == "homework" or entry_type == "homework_session_y" or entry_type == "homework_session_s")): # New Homework
                    ws.cell(row=row_num, column=col_num, value=entry.get("homework_type", entry.get("behavior"))); col_num+=1 # Homework Type/Session Name
                    num_items = entry.get("num_items") # For manually logged with marks
                    if entry.get("type") == "homework_session_s": # For live sessions
                        # Try to count items from details if Yes/No mode
                        homework_details = entry.get("homework_details", {})
                        if not is_autosave:
                            num_items = len(homework_details.get("selected_options",[])) if isinstance(homework_details, dict) else 0
                    elif entry.get("type") == "homework_session_y":
                        num_items = None
                    ws.cell(row=row_num, column=col_num, value=num_items if num_items is not None else "").alignment = right_alignment; col_num+=1
                    total_hw_points = 0; effort_score_val = "" # For summary columns
                    if entry_type == "homework" and "marks_data" in entry: # Graded manual log
                        hw_marks_data = entry.get("marks_data", {})
                        for hmt in self.settings.get("homework_mark_types", []):
                            val = hw_marks_data.get(hmt["id"], "")
                            ws.cell(row=row_num, column=col_num, value=val).alignment = right_alignment; col_num+=1
                            if isinstance(val, (int,float)): total_hw_points += val # Sum points if numeric
                            if hmt["id"] == "hmark_effort": effort_score_val = val # Capture effort score
                    elif entry_type == "homework_session_s" or entry_type == "homework_session_y": # Live session log
                        # For live sessions, fill placeholders for mark type columns or try to map
                        session_details = entry.get("homework_details", {})
                        live_session_mode = entry.get("type")
                        if live_session_mode == "homework_session_y":
                            """
                            for hmt in self.settings.get("homework_mark_types", []): # Fill placeholders
                                # Could try to map "Yes" to complete, "No" to not done, etc.
                                # For now, just leave blank or show raw status if one of the types matches the key
                                found_status_for_mark_type = ""
                                for hw_type_id_key, status_val in session_details.items():
                                    # This mapping is very approximate.
                                    #print(status_val)
                                    if hmt["name"].lower() in hw_type_id_key.lower() or hmt["name"].lower() == status_val.lower():
                                        found_status_for_mark_type = status_val
                                        break
                                    elif "complete" in hmt["name"].lower() and status_val.lower() == "yes":
                                        found_status_for_mark_type = "Yes" # Or map to points
                                        if "default_points" in hmt: total_hw_points += hmt["default_points"]
                                        break
                                    elif "not done" in hmt["name"].lower() and status_val.lower() == "no":
                                        found_status_for_mark_type = "No"
                                        if "default_points" in hmt: total_hw_points += hmt["default_points"]
                                        break
                                ws.cell(row=row_num, column=col_num, value=found_status_for_mark_type).alignment = right_alignment; col_num+=1
                            """
                            i = 0
                            #print(self.all_homework_session_types)
                            found_status_for_mark_type2 = ""
                            col_num += (((len(headers)-col_num)-len(homework_session_types_headers))-1) if not is_autosave else (((len(headers)-col_num)-len(homework_session_types_headers)))
                            for typeh in entry.get("homework_details"):
                                #print(typeh)
                                for hwtype in self.all_homework_session_types:
                                    h_id = hwtype.get("id")
                                    name = hwtype.get("name")
                                    if typeh == h_id:
                                        found_status_for_mark_type2 = entry.get("homework_details").get(typeh).capitalize()
                                i += 1
                                ws.cell(row=row_num, column=col_num, value=found_status_for_mark_type2).alignment = right_alignment; col_num+=1
                        elif live_session_mode == "homework_session_s":
                            selected_options = session_details.get("selected_options", [])
                            for hmt in self.settings.get("homework_mark_types", []): # Fill placeholders based on selected options
                                val_to_put = ""
                                if hmt["name"] in selected_options: # If a mark type name matches a selected option
                                    val_to_put = "Selected" # or hmt["default_points"]
                                    if "default_points" in hmt: total_hw_points += hmt["default_points"]
                                ws.cell(row=row_num, column=col_num, value=val_to_put).alignment = right_alignment; col_num+=1
                        else: # Unknown live mode or no details
                            for _ in self.settings.get("homework_mark_types", []): ws.cell(row=row_num, column=col_num, value="").alignment = right_alignment; col_num+=1

                    ws.cell(row=row_num, column=col_num, value=total_hw_points if total_hw_points else "").alignment = right_alignment; col_num+=1 # Total Points
                    ws.cell(row=row_num, column=col_num, value=effort_score_val).alignment = right_alignment; col_num+=1 # Effort

                comment_col = headers.index("Comment") + 1
                ws.cell(row=row_num, column=comment_col, value=entry.get("comment", "")).alignment = left_alignment
                if not separate_sheets:
                    log_type_col = headers.index("Log Type") + 1
                    ws.cell(row=row_num, column=log_type_col, value=entry.get("type", "behavior").capitalize())
                row_num += 1

            # Auto-size columns
            for col_letter in [get_column_letter(i) for i in range(1, ws.max_column + 1)]:
                max_length = 0
                column_values = [cell.value for cell in ws[col_letter]]
                for cell_val in column_values:
                    if cell_val is not None:
                        try: max_length = max(max_length, len(str(cell_val)))
                        except: pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[col_letter].width = min(max(adjusted_width, 10), 50) # Min/Max width

        log_data_to_export = filtered_log

        # --- Individual Student Log Sheets ---
        if export_all_students_info: # Only create these if full export
            student_worksheets = {} # {student_id: worksheet_object}
            student_headers = ["Timestamp", "Type", "Behavior/Homework/Quiz Name",
                               "Correct/Did", "Total Qs/Total Selected", "Percentage", "Comment", "Day"]
            student_headers.extend(quiz_mark_type_headers) # Add mark types here too
            student_headers.extend(homework_mark_type_headers)
            student_headers.extend(homework_session_types_headers)
            # Put this line below back if it puts something in those columns
            #student_headers.extend(["Homework Type/Session Name", "Num Items"])

            for entry in log_data_to_export:
                student_id = entry["student_id"]
                student_data = self.students.get(student_id)
                student_name_for_sheet = self._make_safe_sheet_name(
                    f"{student_data['first_name']}_{student_data['last_name']}" if student_data else f"Unknown_{student_id}",
                    student_id
                )
                s_homework_marks_data = [""] * len(homework_mark_type_headers)
                
                if entry.get("type") == "homework" and "marks_data" in entry: # Graded manual log
                    total_hw_points = 0; effort_score_val = "" # For summary columns
                    hw_marks_data = entry.get("marks_data", {})
                    i=0
                    for hmt in self.settings.get("homework_mark_types", []):
                        val = hw_marks_data.get(hmt["id"], "")
                        #ws.cell(row=row_num, column=col_num, value=val).alignment = right_alignment; col_num+=1
                        if isinstance(val, (int,float)): total_hw_points += val # Sum points if numeric
                        if hmt["id"] == "hmark_effort": effort_score_val = val # Capture effort score
                        #for i, mt_config_h in enumerate(mark_type_configs_h):
                        s_homework_marks_data[i] = hw_marks_data.get(hmt["id"], "")
                        i+=1
                    #print(s_homework_marks_data)
                    
                    # Add headers for each homework mark type
                    for hmt in self.settings.get("homework_mark_types", []): student_headers.append(hmt["name"])
                    
                    
                if student_id not in student_worksheets:
                    ws_student = wb.create_sheet(title=student_name_for_sheet)
                    student_worksheets[student_id] = ws_student
                    ws_student.append(student_headers)
                    for col_num, header_text in enumerate(student_headers, 1):
                        cell = ws_student.cell(row=1, column=col_num)
                        cell.font = OpenpyxlFont(bold=True)
                        cell.alignment = OpenpyxlAlignment(horizontal="center")
                        width = len(header_text) + 5 # Basic width
                        if header_text == "Timestamp": width = 20
                        elif header_text == "Behavior/Homework/Quiz Name": width = 30
                        elif header_text == "Type": width = 20
                        elif header_text == "Comment": width = 40
                        elif header_text == "Day": width = 12
                        ws_student.column_dimensions[get_column_letter(col_num)].width = width

                ws_student = student_worksheets[student_id]
                ts_obj_s = datetime.fromisoformat(entry["timestamp"])
                s_correct, s_total, s_perc = "", "", ""
                s_quiz_marks_data = [""] * len(quiz_mark_type_headers)
                all_h_types = self.all_homework_session_types
                s_homework_marks_data_2 = [""] * len(all_h_types)
                #print(all_h_types)
                
                
                if entry.get("type") == "quiz":
                    s_marks_data = entry.get("marks_data")
                    s_num_q = entry.get("num_questions", self.settings.get("default_quiz_questions",10))
                    if "score_details" in entry: # Live quiz
                        s_correct = entry["score_details"].get("correct", "")
                        s_total = entry["score_details"].get("total_asked", "")
                        if isinstance(s_correct, (int, float)) and isinstance(s_total, (int, float)) and s_total > 0:
                            s_perc = f"{round((s_correct / s_total) * 100)}%"
                    elif isinstance(s_marks_data, dict):
                        primary_correct_id_s = next((mt["id"] for mt in mark_type_configs if mt["name"].lower() == "correct"), "mark1")
                        s_correct = s_marks_data.get(primary_correct_id_s, "")
                        s_total = s_num_q
                        if isinstance(s_correct, (int,float)) and isinstance(s_total, (int,float)) and s_total > 0:
                            s_perc = f"{round((s_correct / s_total) * 100)}%"
                        for i, mt_config_s in enumerate(mark_type_configs):
                            s_quiz_marks_data[i] = s_marks_data.get(mt_config_s["id"], "")
                elif entry.get("type") == "homework_session_s":
                    s_correct = str(entry["homework_details"].get("selected_options")).removeprefix("[").removesuffix("]")
                    s_total = len(entry["homework_details"]["selected_options"])
                    #pass #print("Homework_session")
                elif entry.get("type") == "homework_session_y":
                    i = 0
                    #print(self.all_homework_session_types)
                    for typeh in entry.get("homework_details"):
                        #print(typeh)
                        for hwtype in all_h_types:
                            h_id = hwtype.get("id")
                            name = hwtype.get("name")
                            if typeh == h_id:
                                s_homework_marks_data_2[i] = entry.get("homework_details").get(typeh).capitalize()
                        i += 1
                            
                s_row_base = [
                    ts_obj_s.strftime('%Y-%m-%d %H:%M:%S'),
                    entry.get("type", "behavior").capitalize(),
                    entry.get("behavior", ""),
                    #entry.get("homework", ""),
                    s_correct, s_total, s_perc,
                    entry.get("comment", "").replace("\n", " "),
                    entry.get("day", "")
                ]
                s_row_base.extend(s_quiz_marks_data)
                s_row_base.extend(s_homework_marks_data)
                s_row_base.extend(s_homework_marks_data_2)
                
                ws_student.append(s_row_base)

        # --- Student Information Sheet ---
        #print((filtered_stud_ids))
        if export_all_students_info and len(filtered_stud_ids) > 1:
            students_info_ws = wb.create_sheet(title="Students Info")
            student_info_headers = ["Student ID", "First Name", "Last Name", "Nickname", "Full Name", "Gender", "Group Name"]
            students_info_ws.append(student_info_headers)
            for col_num, header in enumerate(student_info_headers, 1):
                cell = students_info_ws.cell(row=1, column=col_num)
                cell.font = OpenpyxlFont(bold=True); cell.alignment = OpenpyxlAlignment(horizontal="center")
                info_widths = {"Student ID": 15, "First Name": 15, "Last Name": 15, "Nickname": 15,
                               "Full Name": 25, "Gender": 10, "Group Name": 20}
                students_info_ws.column_dimensions[get_column_letter(col_num)].width = info_widths.get(header, 12)
            
            sorted_students_info = sorted(self.students.values(), key=lambda s: (s.get("last_name", "").lower(), s.get("first_name", "").lower()))

            for student_data in sorted_students_info:
                if student_data.get("id", "") in filtered_stud_ids:
                    group_id = student_data.get("group_id")
                    group_name = ""
                    if self.settings.get("student_groups_enabled", True) and group_id and group_id in self.student_groups:
                        group_name = self.student_groups[group_id].get("name", "")

                    info_row = [
                        student_data.get("id", ""), student_data.get("first_name", ""),
                        student_data.get("last_name", ""), student_data.get("nickname", ""),
                        student_data.get("full_name", ""), student_data.get("gender", ""), group_name
                    ]
                    students_info_ws.append(info_row)


        # Add Summary Sheet if requested
        if filter_settings.get("include_summaries", True) and filtered_log:
            ws_summary = wb.create_sheet(title="Summary")
            current_row = 1
            ws_summary.cell(row=current_row, column=1, value="Log Summary").font = OpenpyxlFont(bold=True, size=14); current_row += 2

            # Behavior Summary
            if filter_settings.get("include_behavior_logs", True):
                ws_summary.cell(row=current_row, column=1, value="Behavior Summary by Student").font = bold_font; current_row += 1
                b_headers = ["Student", "Behavior", "Count"]
                for c_num, h_title in enumerate(b_headers, 1): ws_summary.cell(row=current_row, column=c_num, value=h_title).font = OpenpyxlFont(italic=True)
                current_row += 1
                behavior_counts = {} # {student_id: {behavior_name: count}}
                for entry in filtered_log:
                    if entry.get("type") == "behavior":
                        sid = entry["student_id"]; b_name = entry.get("behavior")
                        behavior_counts.setdefault(sid, {}).setdefault(b_name, 0)
                        behavior_counts[sid][b_name] += 1
                for sid in sorted(behavior_counts.keys(), key=lambda x: student_data_for_export.get(x, {}).get("last_name","")):
                    s_info = student_data_for_export.get(sid, {"full_name": "Unknown"})
                    for b_name, count in sorted(behavior_counts[sid].items()):
                        ws_summary.cell(row=current_row, column=1, value=s_info["full_name"])
                        ws_summary.cell(row=current_row, column=2, value=b_name)
                        ws_summary.cell(row=current_row, column=3, value=count).alignment = right_alignment
                        current_row +=1
                current_row +=1 # Spacer

            # Quiz Summary
            if filter_settings.get("include_quiz_logs", True):
                ws_summary.cell(row=current_row, column=1, value="Quiz Averages by Student").font = bold_font; current_row += 1
                q_headers = ["Student", "Quiz Name", "Avg Score (%)", "Times Taken"]
                for c_num, h_title in enumerate(q_headers, 1): ws_summary.cell(row=current_row, column=c_num, value=h_title).font = OpenpyxlFont(italic=True)
                current_row += 1
                quiz_scores_summary = {} # {student_id: {quiz_name: [scores]}}
                for entry in filtered_log:
                    if entry.get("type") == "quiz":
                        sid = entry["student_id"]; q_name = entry.get("behavior"); num_q_s = entry.get("num_questions",0)
                        marks_d = entry.get("marks_data", {})
                        total_earned_s = 0; extra_credit_s = 0
                        for mt_s in self.settings.get("quiz_mark_types", []):
                            pts_s = marks_d.get(mt_s["id"], 0)
                            if pts_s > 0:
                                if mt_s.get("is_extra_credit", False): extra_credit_s += pts_s * mt_s.get("default_points",1)
                                else: total_earned_s += pts_s * mt_s.get("default_points",1)
                        main_q_total_possible_s = 0
                        correct_type_s = next((m for m in self.settings.get("quiz_mark_types",[]) if m.get("id") == "mark_correct"), None)
                        if correct_type_s and num_q_s > 0: main_q_total_possible_s = correct_type_s.get("default_points", 1) * num_q_s
                        score_val = ((total_earned_s + extra_credit_s) / main_q_total_possible_s) * 100 if main_q_total_possible_s > 0 else (100 if total_earned_s + extra_credit_s > 0 else 0)
                        quiz_scores_summary.setdefault(sid, {}).setdefault(q_name, []).append(score_val)
                for sid in sorted(quiz_scores_summary.keys(), key=lambda x: student_data_for_export.get(x, {}).get("last_name","")):
                    s_info = student_data_for_export.get(sid, {"full_name": "Unknown"})
                    for q_name, scores_list in sorted(quiz_scores_summary[sid].items()):
                        avg_score = sum(scores_list) / len(scores_list) if scores_list else 0
                        ws_summary.cell(row=current_row, column=1, value=s_info["full_name"])
                        ws_summary.cell(row=current_row, column=2, value=q_name)
                        ws_summary.cell(row=current_row, column=3, value=f"{avg_score:.2f}%").alignment = right_alignment
                        ws_summary.cell(row=current_row, column=4, value=len(scores_list)).alignment = right_alignment
                        current_row+=1
                current_row +=1

            # Homework Summary (New)
            if filter_settings.get("include_homework_logs", True):
                ws_summary.cell(row=current_row, column=1, value="Homework Completion by Student").font = bold_font; current_row += 1
                hw_headers = ["Student", "Homework Type/Session", "Count", "Total Points (if applicable)"]
                for c_num, h_title in enumerate(hw_headers, 1): ws_summary.cell(row=current_row, column=c_num, value=h_title).font = OpenpyxlFont(italic=True)
                current_row += 1
                homework_summary = {} # {student_id: {hw_type: {"count": 0, "total_points": 0}}}
                for entry in filtered_log:
                    if entry.get("type") == "homework" or entry.get("type") == "homework_session_s" or entry.get("type") == "homework_session_y":
                        sid = entry["student_id"]
                        hw_name = entry.get("homework_type", entry.get("behavior"))
                        summary_entry = homework_summary.setdefault(sid, {}).setdefault(hw_name, {"count": 0, "total_points": 0.0})
                        summary_entry["count"] += 1
                        # Sum points from marks_data for "homework" type
                        if entry.get("type") == "homework" and "marks_data" in entry:
                            for mark_id, mark_val in entry["marks_data"].items():
                                if isinstance(mark_val, (int, float)): summary_entry["total_points"] += mark_val
                        # Sum points from live session details (approximate)
                        elif entry.get("type") == "homework_session_y" or entry.get("type") == "homework_session_s":
                            hw_details = entry.get("homework_details", {})
                            live_mode = entry.get("type")
                            if live_mode == "homework_session_y":
                                for ht_id_key, status_val in hw_details.items():
                                     if status_val.lower() == "yes": # Simplified: 'yes' adds default points of 'complete' mark type
                                        complete_mark_type = next((m for m in self.settings.get("homework_mark_types",[]) if m["id"] == "hmark_complete"), None)
                                        if complete_mark_type: summary_entry["total_points"] += complete_mark_type.get("default_points",0)
                            elif live_mode == "homework_session_s":
                                selected_opts = hw_details.get("selected_options", [])
                                for opt_name in selected_opts:
                                    opt_mark_type = next((m for m in self.settings.get("homework_mark_types",[]) if m["name"] == opt_name), None)
                                    if opt_mark_type: summary_entry["total_points"] += opt_mark_type.get("default_points",0)

                for sid in sorted(homework_summary.keys(), key=lambda x: student_data_for_export.get(x, {}).get("last_name","")):
                    s_info = student_data_for_export.get(sid, {"full_name": "Unknown"})
                    for hw_name, data in sorted(homework_summary[sid].items()):
                        ws_summary.cell(row=current_row, column=1, value=s_info["full_name"])
                        ws_summary.cell(row=current_row, column=2, value=hw_name)
                        ws_summary.cell(row=current_row, column=3, value=data["count"]).alignment = right_alignment
                        ws_summary.cell(row=current_row, column=4, value=f"{data['total_points']:.2f}" if data['total_points'] else "").alignment = right_alignment
                        current_row += 1
                current_row += 1

            for col_letter_s in [get_column_letter(i) for i in range(1, ws_summary.max_column + 1)]:
                 ws_summary.column_dimensions[col_letter_s].width = 25


        # Save workbook
        try:
            wb.save(filename=file_path)
        except PermissionError as e:
            if is_autosave:
                print(f"Autosave PermissionError: {e}. File might be open.")
                # Don't show messagebox for autosave, just print
            else:
                messagebox.showerror("Save Error", f"Permission denied. Could not save to '{file_path}'.\nPlease ensure the file is not open in another program and you have write permissions.", parent=self.root)
            raise # Re-raise to be caught by the calling function for status update
        except Exception as e_save:
            if is_autosave: print(f"Autosave error: {e_save}")
            else: messagebox.showerror("Save Error", f"An unexpected error occurred while saving Excel file: {e_save}", parent=self.root)
            raise

    def export_data_to_csv_zip(self, zip_file_path, filter_settings=None):
        # ... (updated for new log types and filtering)
        temp_dir = tempfile.mkdtemp()
        try:
            student_data_for_export = {sid: {"first_name": s["first_name"], "last_name": s["last_name"], "full_name": s["full_name"]} for sid, s in self.students.items()}
            logs_to_process_csv = []
            if filter_settings.get("include_behavior_logs", True): logs_to_process_csv.extend([log for log in self.behavior_log if log.get("type") == "behavior"])
            if filter_settings.get("include_quiz_logs", True): logs_to_process_csv.extend([log for log in self.behavior_log if log.get("type") == "quiz"])
            if filter_settings.get("include_homework_logs", True): logs_to_process_csv.extend([log for log in self.homework_log if log.get("type") == "homework" or log.get("type") == "homework_session_y" or log.get("type") == "homework_session_s"])

            filtered_log_csv = []
            start_date_csv, end_date_csv = filter_settings.get("start_date"), filter_settings.get("end_date")
            sel_students_opt_csv, student_ids_flt_csv = filter_settings.get("selected_students", "all"), filter_settings.get("student_ids", [])
            sel_behaviors_opt_csv, behaviors_flt_csv = filter_settings.get("selected_behaviors", "all"), filter_settings.get("behaviors_list", [])
            sel_hw_opt_csv, hw_flt_csv = filter_settings.get("selected_homework_types", "all"), filter_settings.get("homework_types_list", [])

            for entry in logs_to_process_csv:
                try:
                    entry_date = datetime.fromisoformat(entry["timestamp"]).date()
                    if start_date_csv and entry_date < start_date_csv: continue
                    if end_date_csv and entry_date > end_date_csv: continue
                except ValueError: continue
                if sel_students_opt_csv == "specific" and entry["student_id"] not in student_ids_flt_csv: continue
                log_type_csv = entry.get("type", "behavior")
                entry_name_csv = entry.get("behavior")
                if log_type_csv == "homework" or log_type_csv == "homework_session": entry_name_csv = entry.get("homework_type", entry.get("behavior"))

                if log_type_csv == "behavior" or log_type_csv == "quiz":
                    if sel_behaviors_opt_csv == "specific" and entry_name_csv not in behaviors_flt_csv: continue
                elif log_type_csv == "homework" or log_type_csv == "homework_session":
                    if sel_hw_opt_csv == "specific" and entry_name_csv not in hw_flt_csv: continue
                filtered_log_csv.append(entry)
            filtered_log_csv.sort(key=lambda x: x["timestamp"])

            # CSV file for all logs (or separate if preferred, but Excel handles separation better)
            all_logs_csv_path = os.path.join(temp_dir, "all_logs.csv")
            with open(all_logs_csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ["Timestamp", "Date", "Time", "Day", "Student_ID", "First_Name", "Last_Name",
                              "Log_Type", "Item_Name", "Comment", "Num_Questions_Items",
                              "Marks_Data_JSON", "Score_Details_JSON", "Homework_Details_JSON"]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, extrasaction='ignore')
                writer.writeheader()
                for entry in filtered_log_csv:
                    student_info = student_data_for_export.get(entry["student_id"], {"first_name": "N/A", "last_name": "N/A"})
                    try: dt_obj = datetime.fromisoformat(entry["timestamp"])
                    except ValueError: dt_obj = datetime.now()
                    row_data = {
                        "Timestamp": entry["timestamp"], "Date": dt_obj.strftime('%Y-%m-%d'), "Time": dt_obj.strftime('%H:%M:%S'),
                        "Day": entry.get("day", dt_obj.strftime('%A')), "Student_ID": entry["student_id"],
                        "First_Name": student_info["first_name"], "Last_Name": student_info["last_name"],
                        "Log_Type": entry.get("type", "").capitalize(),
                        "Item_Name": entry.get("behavior", entry.get("homework_type", "")),
                        "Comment": entry.get("comment", ""),
                        "Num_Questions_Items": entry.get("num_questions", entry.get("num_items")),
                        "Marks_Data_JSON": json.dumps(entry.get("marks_data")) if "marks_data" in entry else "",
                        "Score_Details_JSON": json.dumps(entry.get("score_details")) if "score_details" in entry else "",
                        "Homework_Details_JSON": json.dumps(entry.get("homework_details")) if "homework_details" in entry else ""
                    }
                    writer.writerow(row_data)

            # CSV file for student list
            students_csv_path = os.path.join(temp_dir, "students.csv")
            with open(students_csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames_s = ["Student_ID", "First_Name", "Last_Name", "Nickname", "Gender", "Group_ID"]
                writer_s = csv.DictWriter(csvfile, fieldnames=fieldnames_s, extrasaction='ignore')
                writer_s.writeheader()
                for sid, sdata in self.students.items():
                     writer_s.writerow({"Student_ID": sid, "First_Name": sdata["first_name"], "Last_Name": sdata["last_name"],
                                        "Nickname": sdata.get("nickname",""), "Gender": sdata.get("gender",""), "Group_ID": sdata.get("group_id","")})

            # Create ZIP file
            with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.write(all_logs_csv_path, arcname="all_logs.csv")
                zf.write(students_csv_path, arcname="students.csv")
                if filter_settings.get("include_summaries", False): # Basic summary text file
                    summary_txt_path = os.path.join(temp_dir, "summary.txt")
                    with open(summary_txt_path, 'w', encoding='utf-8')as f_sum:
                        f_sum.write(f"Log Export Summary - {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
                        f_sum.write(f"Date Range: {start_date_csv or 'Any'} to {end_date_csv or 'Any'}\n")
                        f_sum.write(f"Total Log Entries Exported: {len(filtered_log_csv)}\n")
                        # Further summary details could be added here
                    zf.write(summary_txt_path, arcname="summary.txt")

        finally: shutil.rmtree(temp_dir) # Clean up temp directory


    def export_layout_as_image(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Export Image", "Enter password to export layout as image:"): return
        file_path = filedialog.asksaveasfilename(defaultextension=".png", initialfile=f"layout_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
                                               filetypes=[("PNG Image", "*.png"), ("All files", "*.*")], parent=self.root)
        if not file_path: self.update_status("Image export cancelled."); return
        try:
            # Determine current bounds of drawn items on canvas (in canvas coordinates)
            # This uses the scrollregion which should be set by draw_all_items
            s_region = self.canvas.cget("scrollregion")
            if not s_region: # Fallback if scrollregion is not set (e.g. empty canvas)
                 x1, y1, x2, y2 = 0,0, self.canvas.winfo_width(), self.canvas.winfo_height()
            else:
                try: x1,y1,x2,y2 = map(int, s_region.split())
                except: x1, y1, x2, y2 = 0,0, self.canvas.winfo_width(), self.canvas.winfo_height()

            # Ensure x1, y1 are not negative for postscript (though typically they are 0 or positive)
            # If they are negative, it means content is scrolled left/up off screen.
            # We want to capture from the top-leftmost content.
            postscript_x_offset = -x1 if x1 < 0 else 0
            postscript_y_offset = -y1 if y1 < 0 else 0

            # Create PostScript of the entire scrollable region
            ps_io = io.BytesIO()
            self.canvas.postscript(
                x=x1 + postscript_x_offset,
                y=y1 + postscript_y_offset,
                width=x2 - x1, # Width of the scrollable area
                height=y2 - y1, # Height of the scrollable area
                colormode='color',
                file=ps_io # Write to BytesIO object
            )
            ps_io.seek(0)

            # Use PIL/Pillow to open the PostScript data and save as PNG
            # This requires Ghostscript to be installed and in PATH for PIL to use it
            try:
                img = Image.open(ps_io)
                img.save(file_path, "png")
                self.update_status(f"Layout exported as image: {os.path.basename(file_path)}")
                if messagebox.askyesno("Export Successful", f"Layout image saved to:\n{file_path}\n\nDo you want to open the file location?", parent=self.root):
                    self.open_specific_export_folder(file_path)
            except (OSError, PIL.UnidentifiedImageError) as e_pil:
                print(f"PIL error processing PostScript: {e_pil}")
                if "gs" in str(e_pil).lower() or "ghostscript" in str(e_pil).lower():
                     messagebox.showerror("Image Export Error", "Failed to convert PostScript to image. Ghostscript might not be installed or found in your system's PATH. Please install Ghostscript to enable image export.", parent=self.root)
                else:
                     messagebox.showerror("Image Export Error", f"Failed to save image: {e_pil}.\nEnsure you have image processing libraries like Pillow and its dependencies (e.g., Ghostscript for EPS/PS) installed.", parent=self.root)
            finally:
                ps_io.close()

        except tk.TclError as e_tk:
            messagebox.showerror("Image Export Error", f"Tkinter error during PostScript generation: {e_tk}", parent=self.root)
        except Exception as e:
            messagebox.showerror("Image Export Error", f"An unexpected error occurred: {e}", parent=self.root)
        finally:
            self.password_manager.record_activity()

    def _import_data_from_excel_logic(self, file_path, import_incidents_flag, student_sheet_name_to_import):
        # This function needs significant updates if we want to import detailed quiz scores.
        # For now, it will import students and basic incident info as before.
        # Importing complex quiz scores from Excel would require a well-defined column mapping.
        workbook = load_workbook(filename=file_path, data_only=True)
        imported_student_count = 0

        # --- Import Students ---
        if student_sheet_name_to_import:
            if student_sheet_name_to_import not in workbook.sheetnames:
                messagebox.showerror("Import Error", f"Selected student sheet '{student_sheet_name_to_import}' not found.", parent=self.root)
                return 0, 0

            sheet = workbook[student_sheet_name_to_import]
            header_row_values = [str(cell.value).lower().strip() if cell.value else "" for cell in sheet[1]]

            # Try to find columns for student data
            col_indices = {}
            common_headers = {
                "first_name": ["first", "first name", "firstname"],
                "last_name": ["last", "last name", "lastname", "surname"],
                "full_name": ["full name", "name", "student name"],
                "nickname": ["nickname", "preferred name", "nick"],
                "gender": ["gender", "sex"],
                "group_name": ["group", "group name", "student group"] # For importing group assignment by name
            }
            for key, common_list in common_headers.items():
                for idx, header_val in enumerate(header_row_values):
                    if header_val in common_list:
                        col_indices[key] = idx; break

            # Fallback for full name if "name" exists
            if "name" in header_row_values and "full_name" not in col_indices:
                try: col_indices["full_name"] = header_row_values.index("name")
                except ValueError: pass

            # Basic name column check
            if "first_name" not in col_indices or "last_name" not in col_indices:
                if "full_name" not in col_indices:
                    if messagebox.askyesno("Column Ambiguity", f"Could not auto-detect specific name columns in '{student_sheet_name_to_import}'.\n"
                                           "Assume Col A = First, Col B = Last? \n(No assumes Col A = 'Last, First' or 'First Last')", parent=self.root):
                        col_indices["first_name"], col_indices["last_name"] = 0, 1
                    else: col_indices["full_name"] = 0


            existing_full_names_in_app = {s['full_name'].lower().strip(): s['id'] for s in self.students.values()}
            commands_to_add_students = []
            current_id_num_for_batch = self.next_student_id_num # Use app's current next ID

            for row_idx, row_values_tuple in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                row_values = list(row_values_tuple)
                first_name, last_name, nickname, gender, group_id_to_assign = None, None, "", "Boy", None

                def get_val(key_idx, default=""):
                    idx = col_indices.get(key_idx)
                    if idx is not None and idx < len(row_values) and row_values[idx] is not None:
                        return str(row_values[idx]).strip()
                    return default

                first_name = get_val("first_name", None)
                last_name = get_val("last_name", None)
                nickname = get_val("nickname", "")
                gender_str = get_val("gender", "Boy").lower()
                if gender_str in ["girl", "female", "f"]: gender = "Girl"

                group_name_from_excel = get_val("group_name", "")
                if group_name_from_excel and self.settings.get("student_groups_enabled", True):
                    # Find group_id by name
                    for gid, gdata in self.student_groups.items():
                        if gdata.get("name", "").lower() == group_name_from_excel.lower():
                            group_id_to_assign = gid; break


                if not first_name or not last_name: # Try parsing from full_name if specific cols missing
                    if "full_name" in col_indices:
                        full_name_str = get_val("full_name", None)
                        if full_name_str:
                            if "," in full_name_str:
                                parts = full_name_str.split(",", 1)
                                last_name = parts[0].strip()
                                first_name = parts[1].strip() if len(parts) > 1 else ""
                            elif " " in full_name_str: # Assume "First Last"
                                parts = full_name_str.split(" ", 1)
                                first_name = parts[0].strip()
                                last_name = parts[1].strip() if len(parts) > 1 else ""
                            else: first_name = full_name_str # Single word as first name

                if first_name: # Must have at least a first name
                    if not last_name: last_name = "" # Ensure last_name is a string

                    full_name_display = f"{first_name} \"{nickname}\" {last_name}" if nickname else f"{first_name} {last_name}"
                    full_name_key = full_name_display.lower().strip()

                    if full_name_key not in existing_full_names_in_app:
                        student_id_str = f"student_{current_id_num_for_batch}"
                        old_next_id_for_command = current_id_num_for_batch # For AddItemCommand's undo logic
                        next_id_for_app_after_this = current_id_num_for_batch + 1

                        x_pos = 50 + (imported_student_count % 10) * (self.settings.get("default_student_box_width", DEFAULT_STUDENT_BOX_WIDTH) + 10)
                        y_pos = 50 + (imported_student_count // 10) * (self.settings.get("default_student_box_height", DEFAULT_STUDENT_BOX_HEIGHT) + 10)

                        s_data = {
                            "first_name": first_name, "last_name": last_name, "nickname": nickname,
                            "gender": gender, "full_name": full_name_display,
                            "x": x_pos, "y": y_pos, "id": student_id_str,
                            "width": self.settings.get("default_student_box_width"),
                            "height": self.settings.get("default_student_box_height"),
                            "original_next_id_num_after_add": next_id_for_app_after_this,
                            "group_id": group_id_to_assign,
                            "style_overrides": {}
                        }
                        cmd = AddItemCommand(self, student_id_str, 'student', s_data, old_next_id_for_command)
                        commands_to_add_students.append(cmd)
                        existing_full_names_in_app[full_name_key] = student_id_str # Add to check for this batch
                        imported_student_count += 1
                        current_id_num_for_batch += 1
                    else: # Student exists, maybe update their group?
                        existing_student_id = existing_full_names_in_app[full_name_key]
                        if group_id_to_assign and self.students[existing_student_id].get("group_id") != group_id_to_assign:
                            # Create an EditItemCommand to update group_id
                            old_data_snapshot = self.students[existing_student_id].copy()
                            changes = {"group_id": group_id_to_assign}
                            # No need for AddItemCommand here, it's an edit.
                            # This part could be more complex if we want to batch these edits.
                            # For now, let's assume import primarily adds new students.
                            print(f"Student {full_name_display} already exists. Group update from Excel not yet fully implemented here.")


            for cmd in commands_to_add_students:
                self.execute_command(cmd) # This will update self.next_student_id_num via AddItemCommand

            if commands_to_add_students: # If any students were added
                 self.next_student_id_num = current_id_num_for_batch # Ensure app's counter is past the last used ID

        # --- Import Incidents (Simplified - does not import detailed quiz marks yet) ---
        imported_incident_count = 0
        if import_incidents_flag:
            incident_commands_to_add = []
            for sheet_name_excel in workbook.sheetnames:
                # Try to match Excel sheet name (e.g., "FirstName_LastName") to an existing student
                matched_student_id, matched_student_first_name, matched_student_last_name = None, "", ""
                normalized_excel_sheet_name_for_match = sheet_name_excel.replace("_", " ").lower()

                for s_id_app, s_data_app in self.students.items():
                    # Check against "FirstName LastName" and "FirstName_LastName" formats
                    app_student_full_name_match = s_data_app['full_name'].lower()
                    app_student_export_format_match = f"{s_data_app['first_name']}_{s_data_app['last_name']}".lower()
                    if normalized_excel_sheet_name_for_match == app_student_full_name_match or \
                       sheet_name_excel.lower() == app_student_export_format_match:
                        matched_student_id = s_id_app
                        matched_student_first_name = s_data_app['first_name']
                        matched_student_last_name = s_data_app['last_name']
                        break

                if matched_student_id:
                    student_sheet_incidents = workbook[sheet_name_excel]
                    s_header_values = [str(cell.value).lower().strip() if cell.value else "" for cell in student_sheet_incidents[1]]
                    col_map_incidents = {}
                    try: # Basic incident columns
                        col_map_incidents["ts"] = s_header_values.index("timestamp")
                        col_map_incidents["type"] = s_header_values.index("type")
                        col_map_incidents["beh_quiz_name"] = s_header_values.index("behavior/quiz name")
                        # For quiz scores, we'd look for "Correct", "Total Qs" or specific mark type columns
                        # This part is simplified for now.
                        col_map_incidents["score_correct"] = s_header_values.index("correct") if "correct" in s_header_values else -1
                        col_map_incidents["score_total_qs"] = s_header_values.index("total qs") if "total qs" in s_header_values else -1
                        col_map_incidents["comment"] = s_header_values.index("comment")
                        col_map_incidents["day"] = s_header_values.index("day")
                    except ValueError:
                        print(f"Skipping sheet '{sheet_name_excel}' for incident import: missing expected basic headers.")
                        continue

                    for row_idx_inc, s_row_values_tuple_inc in enumerate(student_sheet_incidents.iter_rows(min_row=2, values_only=True)):
                        s_row_values_inc = list(s_row_values_tuple_inc)
                        try:
                            timestamp_str = str(s_row_values_inc[col_map_incidents["ts"]]) if col_map_incidents["ts"] < len(s_row_values_inc) and s_row_values_inc[col_map_incidents["ts"]] else None
                            log_type_str = str(s_row_values_inc[col_map_incidents["type"]]).lower() if col_map_incidents["type"] < len(s_row_values_inc) and s_row_values_inc[col_map_incidents["type"]] else "behavior"
                            behavior_quiz_name_str = str(s_row_values_inc[col_map_incidents["beh_quiz_name"]]) if col_map_incidents["beh_quiz_name"] < len(s_row_values_inc) and s_row_values_inc[col_map_incidents["beh_quiz_name"]] else ""
                            comment_str = str(s_row_values_inc[col_map_incidents["comment"]]) if col_map_incidents["comment"] < len(s_row_values_inc) and s_row_values_inc[col_map_incidents["comment"]] else ""
                            day_str = str(s_row_values_inc[col_map_incidents["day"]]) if col_map_incidents["day"] < len(s_row_values_inc) and s_row_values_inc[col_map_incidents["day"]] else ""

                            if not timestamp_str or not behavior_quiz_name_str: continue

                            parsed_dt = None
                            if isinstance(s_row_values_inc[col_map_incidents["ts"]], datetime): parsed_dt = s_row_values_inc[col_map_incidents["ts"]]
                            else:
                                try: parsed_dt = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                                except ValueError:
                                    try: parsed_dt = datetime.fromisoformat(timestamp_str)
                                    except ValueError:
                                        print(f"Skipping incident in '{sheet_name_excel}', row {row_idx_inc+2}: bad timestamp format '{timestamp_str}'")
                                        continue
                            iso_timestamp = parsed_dt.isoformat()

                            # Check for duplicates before adding
                            is_duplicate = False
                            for existing_log in self.behavior_log:
                                if existing_log["student_id"] == matched_student_id and \
                                   existing_log["timestamp"] == iso_timestamp and \
                                   existing_log["behavior"] == behavior_quiz_name_str and \
                                   existing_log.get("type","behavior") == log_type_str:
                                    is_duplicate = True; break

                            if not is_duplicate:
                                log_entry_data = {
                                    "timestamp": iso_timestamp, "student_id": matched_student_id,
                                    "student_first_name": matched_student_first_name,
                                    "student_last_name": matched_student_last_name,
                                    "behavior": behavior_quiz_name_str, "comment": comment_str,
                                    "type": log_type_str, "day": day_str
                                }
                                # Simplified quiz score import (just basic correct/total if available)
                                if log_type_str == "quiz":
                                    correct_val_imp = None
                                    total_qs_val_imp = None
                                    if col_map_incidents["score_correct"] != -1 and col_map_incidents["score_correct"] < len(s_row_values_inc) and s_row_values_inc[col_map_incidents["score_correct"]]:
                                        try: correct_val_imp = int(s_row_values_inc[col_map_incidents["score_correct"]])
                                        except ValueError: pass
                                    if col_map_incidents["score_total_qs"] != -1 and col_map_incidents["score_total_qs"] < len(s_row_values_inc) and s_row_values_inc[col_map_incidents["score_total_qs"]]:
                                        try: total_qs_val_imp = int(s_row_values_inc[col_map_incidents["score_total_qs"]])
                                        except ValueError: pass

                                    if correct_val_imp is not None and total_qs_val_imp is not None:
                                        # Store as simple score_details for now, similar to live quiz
                                        log_entry_data["score_details"] = {"correct": correct_val_imp, "total_asked": total_qs_val_imp}
                                        log_entry_data["num_questions"] = total_qs_val_imp # Assume total_asked is num_questions
                                    elif correct_val_imp is not None: # If only correct is found, log it as a raw score string
                                        log_entry_data["score"] = str(correct_val_imp)


                                cmd = LogEntryCommand(self, log_entry_data, matched_student_id, timestamp=iso_timestamp)
                                incident_commands_to_add.append(cmd)
                                imported_incident_count += 1
                        except IndexError:
                            print(f"Skipping row {row_idx_inc + 2} in '{sheet_name_excel}' for incident import: missing data columns.")
                            continue
            for cmd in incident_commands_to_add:
                self.execute_command(cmd)

        return imported_student_count, imported_incident_count


    def import_students_from_excel_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Import", "Enter password to import from Excel:"): return

        dialog = ImportExcelOptionsDialog(self.root, app_instance=self)
        if dialog.result:
            file_path, import_incidents_flag, student_sheet_name = dialog.result
            if not file_path: return
            try:
                imported_student_count, imported_incident_count = self._import_data_from_excel_logic(file_path, import_incidents_flag, student_sheet_name)
                status_msg = f"Imported {imported_student_count} new students"
                if import_incidents_flag: status_msg += f" and {imported_incident_count} new incidents"
                status_msg += ". Duplicates were skipped."
                self.update_status(status_msg)
                self.draw_all_items(check_collisions_on_redraw=True)
                self.save_data_wrapper(source="import_excel") # Save after successful import
                self.password_manager.record_activity()
            except Exception as e:
                messagebox.showerror("Import Error", f"Failed to import from Excel: {e}", parent=self.root)
                self.update_status(f"Error during Excel import: {e}")
                import traceback
                traceback.print_exc()



    def save_layout_template_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Save Layout", "Enter password to save layout template:"): return
        template_name = simpledialog.askstring("Save Layout Template", "Enter a name for this layout template:", parent=self.root)
        if template_name and template_name.strip():
            filename = "".join(c if c.isalnum() or c in (' ', '_', '-') else '_' for c in template_name.strip()) + ".json"
            file_path = os.path.join(LAYOUT_TEMPLATES_DIR, filename)
            layout_data = {
                "students": {sid: {"x": s["x"], "y": s["y"], "width": s.get("width"), "height": s.get("height"),
                                   "style_overrides": s.get("style_overrides",{}).copy()}
                             for sid, s in self.students.items()},
                "furniture": {fid: {"x": f["x"], "y": f["y"], "width": f.get("width"), "height": f.get("height")}
                              for fid, f in self.furniture.items()}
            }
            try:
                with open(file_path, 'w', encoding='utf-8') as f: json.dump(layout_data, f, indent=4)
                self.update_status(f"Layout template '{template_name}' saved.")
            except IOError as e: messagebox.showerror("Save Error", f"Could not save layout template: {e}", parent=self.root)
        else: self.update_status("Layout template save cancelled.")
        self.password_manager.record_activity()

    def load_layout_template_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Load Layout", "Enter password to load layout template:"): return
        if not os.path.exists(LAYOUT_TEMPLATES_DIR) or not os.listdir(LAYOUT_TEMPLATES_DIR):
            messagebox.showinfo("No Templates", "No layout templates found.", parent=self.root); return
        file_path = filedialog.askopenfilename(initialdir=LAYOUT_TEMPLATES_DIR, title="Select Layout Template",
                                               filetypes=[("JSON files", "*.json"), ("All files", "*.*")], parent=self.root)
        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f: template_data = json.load(f)
                if messagebox.askyesno("Confirm Load", "Loading this template will overwrite current item positions and sizes. Student data (names, logs) will be preserved. Continue?", parent=self.root):
                    # Use EditItemCommand for existing items, AddItemCommand if template has items not in current classroom (less likely for pure layout)
                    # For simplicity, direct modification and one large MoveItemsCommand and ChangeItemsSizeCommand
                    move_commands_data = []
                    size_commands_data = []
                    template_students = template_data.get("students", {})
                    template_furniture = template_data.get("furniture", {})

                    for item_id, t_data in template_students.items():
                        if item_id in self.students:
                            s_current = self.students[item_id]
                            old_x, old_y = s_current["x"], s_current["y"]
                            new_x, new_y = t_data["x"], t_data["y"]
                            if old_x != new_x or old_y != new_y : move_commands_data.append({'id':item_id, 'type':'student', 'old_x':old_x, 'old_y':old_y, 'new_x':new_x, 'new_y':new_y})
                            
                            old_w = s_current.get("style_overrides",{}).get("width", s_current.get("width"))
                            old_h = s_current.get("style_overrides",{}).get("height", s_current.get("height"))
                            new_w = t_data.get("width", old_w)
                            new_h = t_data.get("height", old_h)
                            if old_w != new_w or old_h != new_h: size_commands_data.append({'id':item_id, 'type':'student', 'old_w':old_w, 'old_h':old_h, 'new_w':new_w, 'new_h':new_h})
                            
                            # Apply style overrides (color, font size etc.) from template
                            # This is more complex for undo, might need a dedicated StyleApplyCommand or enhance EditItemCommand
                            t_style_overrides = t_data.get("style_overrides", {})
                            if t_style_overrides:
                                old_style_snapshot = s_current.get("style_overrides", {}).copy()
                                # Create a set of specific changes for EditItemCommand-like application for styles
                                style_changes_for_cmd = {}
                                for k,v_new in t_style_overrides.items():
                                    v_old = old_style_snapshot.get(k)
                                    if v_old != v_new: style_changes_for_cmd[k] = v_new
                                # Need to handle removal of keys present in old but not new if template dictates reset
                                for k_old in old_style_snapshot:
                                    if k_old not in t_style_overrides: style_changes_for_cmd[k_old] = None # Sentinel for removal

                                if style_changes_for_cmd:
                                    # Create snapshot of full student data for EditItemCommand
                                    full_old_student_data = s_current.copy()
                                    full_old_student_data["style_overrides"] = old_style_snapshot
                                    self.execute_command(EditItemCommand(self,item_id,"student",full_old_student_data, {"style_overrides": t_style_overrides}))


                    for item_id, t_data in template_furniture.items():
                         if item_id in self.furniture:
                            f_current = self.furniture[item_id]
                            old_x, old_y = f_current["x"], f_current["y"]
                            new_x, new_y = t_data["x"], t_data["y"]
                            if old_x != new_x or old_y != new_y : move_commands_data.append({'id':item_id, 'type':'furniture', 'old_x':old_x, 'old_y':old_y, 'new_x':new_x, 'new_y':new_y})

                            old_w = f_current.get("width") ; old_h = f_current.get("height")
                            new_w = t_data.get("width", old_w); new_h = t_data.get("height", old_h)
                            if old_w != new_w or old_h != new_h: size_commands_data.append({'id':item_id, 'type':'furniture', 'old_w':old_w, 'old_h':old_h, 'new_w':new_w, 'new_h':new_h})

                    if move_commands_data: self.execute_command(MoveItemsCommand(self, move_commands_data))
                    if size_commands_data: self.execute_command(ChangeItemsSizeCommand(self, size_commands_data))

                    self.update_status(f"Layout template '{os.path.basename(file_path)}' loaded.")
                    self.draw_all_items(check_collisions_on_redraw=True)
                    self.save_data_wrapper(source="load_template")
            except (json.JSONDecodeError, IOError) as e: messagebox.showerror("Load Error", f"Could not load layout template: {e}", parent=self.root)
        else: self.update_status("Layout template load cancelled.")
        self.password_manager.record_activity()

    def generate_attendance_report_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Generate Report", "Enter password to generate attendance report:"): return

        dialog = AttendanceReportDialog(self.root, self.students)
        if dialog.result:
            start_date, end_date, selected_student_ids = dialog.result
            if not selected_student_ids:
                messagebox.showinfo("No Students", "No students selected for the report.", parent=self.root)
                return

            report_data = self.generate_attendance_data(start_date, end_date, selected_student_ids)
            if not report_data:
                messagebox.showinfo("No Data", "No attendance-relevant log data found for the selected criteria.", parent=self.root)
                return

            default_filename = f"attendance_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_filename,
                                                   filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")], parent=self.root)
            if file_path:
                try:
                    self.export_attendance_to_excel(file_path, report_data, start_date, end_date)
                    self.update_status(f"Attendance report saved to {os.path.basename(file_path)}.")
                    if messagebox.askyesno("Export Successful", f"Attendance report saved to:\n{file_path}\n\nDo you want to open the file location?", parent=self.root):
                        self.open_specific_export_folder(file_path)
                except Exception as e:
                    messagebox.showerror("Export Error", f"Failed to save attendance report: {e}", parent=self.root)
            else:
                self.update_status("Attendance report export cancelled.")
        self.password_manager.record_activity()

    def generate_attendance_data(self, start_date, end_date, student_ids):
        attendance = {} # {date_obj: {student_id: "Present"}}
        all_logs = self.behavior_log + self.homework_log # Combine logs for presence check

        current_date = start_date
        while current_date <= end_date:
            attendance[current_date] = {}
            for student_id in student_ids:
                # Check if any log entry exists for this student on this date
                present = any(
                    log["student_id"] == student_id and
                    datetime.fromisoformat(log["timestamp"]).date() == current_date
                    for log in all_logs
                )
                attendance[current_date][student_id] = "P" if present else "A" # Present / Absent
            current_date += timedelta(days=1)
        return attendance

    def export_attendance_to_excel(self, file_path, attendance_data, report_start_date, report_end_date):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        headers = ["Student Name"]
        date_columns_map = {} # date_obj -> column_index
        current_col = 2
        d_iter = report_start_date
        while d_iter <= report_end_date:
            headers.append(d_iter.strftime("%Y-%m-%d (%a)"))
            date_columns_map[d_iter] = current_col
            current_col += 1
        headers.append("Total Present")
        headers.append("Total Absent")

        for col_num, header_title in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header_title).font = OpenpyxlFont(bold=True)
            ws.column_dimensions[get_column_letter(col_num)].width = 15 if col_num > 1 else 25
        ws.freeze_panes = 'B2'

        current_row = 2
        sorted_student_ids = sorted(
            list(set(sid for day_data in attendance_data.values() for sid in day_data.keys())),
            key=lambda sid: (self.students.get(sid, {}).get("last_name", ""), self.students.get(sid, {}).get("first_name", ""))
        )

        for student_id in sorted_student_ids:
            student_name = self.students.get(student_id, {}).get("full_name", student_id)
            ws.cell(row=current_row, column=1, value=student_name)
            total_present = 0
            total_absent = 0
            for date_obj, col_idx in date_columns_map.items():
                status = attendance_data.get(date_obj, {}).get(student_id, "A")
                ws.cell(row=current_row, column=col_idx, value=status).alignment = OpenpyxlAlignment(horizontal='center')
                if status == "P": total_present += 1
                else: total_absent += 1
            ws.cell(row=current_row, column=len(headers)-1, value=total_present).alignment = OpenpyxlAlignment(horizontal='center')
            ws.cell(row=current_row, column=len(headers), value=total_absent).alignment = OpenpyxlAlignment(horizontal='center')
            current_row += 1
        wb.save(file_path)

    def align_selected_items(self, edge):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Align", "Enter password to align items:"): return
        if len(self.selected_items) < 2:
            self.update_status("Select at least two items to align."); return

        items_data_for_align = []
        for item_id in self.selected_items:
            item_type = "student" if item_id in self.students else "furniture"
            data_source = self.students if item_type == "student" else self.furniture
            if item_id in data_source:
                item = data_source[item_id]
                items_data_for_align.append({
                    "id": item_id, "type": item_type, "x": item["x"], "y": item["y"],
                    "width": item.get("_current_world_width", item.get("width", DEFAULT_STUDENT_BOX_WIDTH)), # Use dynamic if available
                    "height": item.get("_current_world_height", item.get("height", DEFAULT_STUDENT_BOX_HEIGHT))
                })
        if not items_data_for_align: return

        target_coord = 0
        if edge == "left": target_coord = min(item["x"] for item in items_data_for_align)
        elif edge == "right": target_coord = max(item["x"] + item["width"] for item in items_data_for_align)
        elif edge == "top": target_coord = min(item["y"] for item in items_data_for_align)
        elif edge == "bottom": target_coord = max(item["y"] + item["height"] for item in items_data_for_align)
        elif edge == "center_h": # Align to average horizontal center of the selection box
            min_x = min(it["x"] for it in items_data_for_align)
            max_x_br = max(it["x"] + it["width"] for it in items_data_for_align)
            target_coord = (min_x + max_x_br) / 2 # This is the center of the bounding box of selected items
        elif edge == "center_v": # Align to average vertical center
            min_y = min(it["y"] for it in items_data_for_align)
            max_y_br = max(it["y"] + it["height"] for it in items_data_for_align)
            target_coord = (min_y + max_y_br) / 2

        move_commands_for_align = []
        for item_to_align in items_data_for_align:
            old_x_align, old_y_align = item_to_align["x"], item_to_align["y"]
            new_x_align, new_y_align = old_x_align, old_y_align
            if edge == "left": new_x_align = target_coord
            elif edge == "right": new_x_align = target_coord - item_to_align["width"]
            elif edge == "top": new_y_align = target_coord
            elif edge == "bottom": new_y_align = target_coord - item_to_align["height"]
            elif edge == "center_h": new_x_align = target_coord - item_to_align["width"] / 2
            elif edge == "center_v": new_y_align = target_coord - item_to_align["height"] / 2

            if abs(new_x_align - old_x_align) > 0.01 or abs(new_y_align - old_y_align) > 0.01:
                move_commands_for_align.append({'id': item_to_align["id"], 'type': item_to_align["type"], 'old_x': old_x_align, 'old_y': old_y_align, 'new_x': new_x_align, 'new_y': new_y_align})

        if move_commands_for_align:
            self.execute_command(MoveItemsCommand(self, move_commands_for_align))
            self.update_status(f"Aligned {len(move_commands_for_align)} items to {edge}.")
        else: self.update_status("Items already aligned."); self.draw_all_items(check_collisions_on_redraw=True)
        self.password_manager.record_activity()

    def assign_student_to_group_via_menu(self, student_id, group_id):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Assign Group", "Enter password to assign group:"): return
        student = self.students.get(student_id)
        if not student: return

        old_group_id = student.get("group_id")
        new_group_id_val = None if group_id == "NONE_GROUP_SENTINEL" else group_id

        if old_group_id != new_group_id_val:
            # For undo/redo, we need snapshots of all group assignments if ManageStudentGroupCommand is used broadly.
            # For a single assignment, EditItemCommand is simpler.
            old_student_data_snapshot = student.copy()
            if "style_overrides" in old_student_data_snapshot: old_student_data_snapshot["style_overrides"] = old_student_data_snapshot["style_overrides"].copy()
            
            changes_for_command = {"group_id": new_group_id_val}
            self.execute_command(EditItemCommand(self, student_id, "student", old_student_data_snapshot, changes_for_command))
            # Command will call draw_all_items, which will redraw the student
            group_name = self.student_groups[new_group_id_val]['name'] if new_group_id_val and new_group_id_val in self.student_groups else "No Group"
            self.update_status(f"Assigned {student['full_name']} to group: {group_name}.")
            self.save_data_wrapper(source="assign_group_menu") # Save student data which now includes group_id
        self.password_manager.record_activity()



    def assign_students_to_group_via_menu(self, student_ids, group_id):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Assign Group", "Enter password to assign group:"): return
        
        for student in student_ids:
            student_id = student
            student = self.students.get(student_id)
            if not student: return

            old_group_id = student.get("group_id")
            new_group_id_val = None if group_id == "NONE_GROUP_SENTINEL" else group_id

            if old_group_id != new_group_id_val:
                # For undo/redo, we need snapshots of all group assignments if ManageStudentGroupCommand is used broadly.
                # For a single assignment, EditItemCommand is simpler.
                old_student_data_snapshot = student.copy()
                if "style_overrides" in old_student_data_snapshot: old_student_data_snapshot["style_overrides"] = old_student_data_snapshot["style_overrides"].copy()
                
                changes_for_command = {"group_id": new_group_id_val}
                self.execute_command(EditItemCommand(self, student_id, "student", old_student_data_snapshot, changes_for_command))
                # Command will call draw_all_items, which will redraw the student
                group_name = self.student_groups[new_group_id_val]['name'] if new_group_id_val and new_group_id_val in self.student_groups else "No Group"
                self.update_status(f"Assigned {student['full_name']} to group: {group_name}.")
                self.save_data_wrapper(source="assign_group_menu") # Save student data which now includes group_id
        self.password_manager.record_activity()



    
    def manage_student_groups_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Manage Groups", "Enter password to manage student groups:"): return
        # Take snapshots for ManageStudentGroupCommand
        old_groups_snap = {gid: gdata.copy() for gid, gdata in self.student_groups.items()}
        old_student_assignments_snap = {sid: sdata.get("group_id") for sid, sdata in self.students.items() if "group_id" in sdata and sdata["group_id"] is not None}
        old_next_group_id_num_snap = self.next_group_id_num

        dialog = ManageStudentGroupsDialog(self.root, self.student_groups, self.students, self, default_colors=DEFAULT_GROUP_COLORS)
        if dialog.groups_changed_flag: # Check if dialog indicated changes were made
            # The dialog directly modifies self.student_groups, self.students (group_id), and self.next_group_id_num
            new_groups_snap = {gid: gdata.copy() for gid, gdata in self.student_groups.items()} # Current state after dialog
            new_student_assignments_snap = {sid: sdata.get("group_id") for sid, sdata in self.students.items() if "group_id" in sdata and sdata["group_id"] is not None}
            new_next_group_id_num_snap = self.next_group_id_num

            cmd = ManageStudentGroupCommand(self, old_groups_snap, new_groups_snap,
                                            old_student_assignments_snap, new_student_assignments_snap,
                                            old_next_group_id_num_snap, new_next_group_id_num_snap)
            # Execute command will apply the new state (which is already set by dialog) and handle saving/drawing
            # The command's execute() will effectively re-apply what the dialog did, which is fine.
            # The crucial part is that undo() will restore the old snapshots.
            self.execute_command(cmd) # This will also call save_student_groups and draw_all_items
            self.update_status("Student groups updated via dialog.") # Status from command is more generic
        else:
            self.update_status("Manage student groups cancelled or no changes made.")
        self.password_manager.record_activity()

    def toggle_student_groups_ui_visibility(self):
        enabled = self.settings.get("student_groups_enabled", True)
        if hasattr(self, 'manage_groups_btn'):
            self.manage_groups_btn.config(state=tk.NORMAL if enabled else tk.DISABLED)
        self.draw_all_items(check_collisions_on_redraw=False) # Redraw to show/hide indicators


    def manage_quiz_templates_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Manage Templates", "Enter password to manage quiz templates:"): return
        dialog = ManageQuizTemplatesDialog(self.root, self) # Pass app instance
        if dialog.templates_changed_flag:
            self.save_quiz_templates() # Dialog modifies self.quiz_templates directly for now
            self.update_status("Quiz templates updated.")
        else:
            self.update_status("Quiz template management cancelled or no changes made.")
        self.password_manager.record_activity()

    def manage_homework_templates_dialog(self): # New
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Manage Templates", "Enter password to manage homework templates:"): return
        dialog = ManageHomeworkTemplatesDialog(self.root, self) # Pass app instance
        if dialog.templates_changed_flag:
            self.save_homework_templates() # Dialog modifies self.homework_templates directly
            self.update_status("Homework templates updated.")
        else:
            self.update_status("Homework template management cancelled or no changes made.")
        self.password_manager.record_activity()


    def set_theme(self, theme, canvas_color):
        if theme != "System":
            sv_ttk.set_theme(theme)
        else:
            sv_ttk.set_theme(darkdetect.theme())

        self.theme_style_using = theme
        
        if canvas_color == "Default" or canvas_color == "" or canvas_color == None:
            canvas_color = None; self.custom_canvas_color = None
        else:
            self.custom_canvas_color = canvas_color
            #print(self.custom_canvas_color)
            self.canvas_color = canvas_color
        
        if self.custom_canvas_color: self.canvas_color = self.custom_canvas_color
        elif self.theme_style_using == "Dark": self.canvas_color = "#1F1F1F"
        elif self.theme_style_using == "System": self.canvas_color = "lightgrey" if darkdetect.theme() == "Light" else "#1F1F1F"
        else: self.canvas_color = "lightgrey"
        self.canvas.configure(bg=self.canvas_color)

    def theme_auto(self, theme):
        #print(theme)
        if self.theme_style_using != "System":
            sv_ttk.set_theme(self.theme_style_using)
        else:
            sv_ttk.set_theme(darkdetect.theme())
            #self.theme_style_using = darkdetect.theme()
        
        
        if self.custom_canvas_color: self.canvas_color = self.custom_canvas_color
        elif self.theme_style_using == "Dark": self.canvas_color = "#1F1F1F"
        elif self.theme_style_using == "System": self.canvas_color = "lightgrey" if darkdetect.theme() == "Light" else "#1F1F1F"
        else: self.canvas_color = "lightgrey"
        self.canvas.configure(bg=self.canvas_color)




    def open_settings_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Open Settings", "Enter password to open settings:"): return
        # Store a copy of settings for potential revert or detailed change tracking for undo (complex)
        # For now, settings dialog applies changes directly and saves.
        # old_settings_snapshot = self.settings.copy() # For a potential future SettingsChangeCommand
        dialog = SettingsDialog(self.root, self.settings, self.custom_behaviors, self.all_behaviors, self,
                                self.custom_homework_log_behaviors, self.all_homework_log_behaviors, # Homework log behaviors
                                self.custom_homework_session_types, self.all_homework_session_types, # Homework session types (Yes/No mode)
                                self.password_manager, self.theme_style_using, self.custom_canvas_color)
        if dialog.settings_changed_flag: # Check if dialog indicated changes
            # Settings are applied directly by the dialog for most parts
            self.save_data_wrapper(source="settings_dialog") # Save all data as settings are part of it
            self.update_all_behaviors(); self.update_all_homework_log_behaviors(); self.update_all_homework_session_types()
            self.draw_all_items(check_collisions_on_redraw=True)
            self.update_status("Settings updated.")
            self.update_zoom_display()
            self.update_lock_button_state()
            self.toggle_student_groups_ui_visibility()
            self.set_theme(self.theme_style_using, self.custom_canvas_color)
            # Re-schedule autosave if interval changed
            self.root.after_cancel(self.autosave_data_wrapper) # Cancel existing if any (might need to store the after_id)
            self.root.after(self.settings.get("autosave_interval_ms", 30000), self.autosave_data_wrapper)
        else:
            try:
                self.update_status("Settings dialog closed, no changes applied through dialog confirm.")
            except: pass
        

        self.password_manager.record_activity()

    def reset_application_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Reset Application", "Enter password to reset application:"): return
        msg = "This will reset ALL application data including students, furniture, logs, settings, custom behaviors, templates, and groups. This action CANNOT be undone.\n\nAre you absolutely sure you want to reset the application to its default state?"
        if messagebox.askyesno("Confirm Application Reset", msg, icon='error', parent=self.root, default=messagebox.NO):
            # Second, more direct confirmation
            if messagebox.askyesno("Final Confirmation", "Really reset everything? This is your last chance to cancel.", icon='error', parent=self.root, default=messagebox.NO):
                self._perform_reset()
            else: self.update_status("Application reset cancelled (final prompt).")
        else: self.update_status("Application reset cancelled.")
        self.password_manager.record_activity()

    def _perform_reset(self):
        try:
            self.backup_all_data_dialog(force=True)
        except Exception as e:
            print(e)
        try:
            # Clear current data in memory
            self.students.clear(); self.furniture.clear(); self.behavior_log.clear(); self.homework_log.clear()
            self.student_groups.clear(); self.quiz_templates.clear(); self.homework_templates.clear()
            self.custom_behaviors.clear(); self.custom_homework_log_behaviors.clear(); self.custom_homework_session_types.clear()
            self.undo_stack.clear(); self.redo_stack.clear()
            self._per_student_last_cleared.clear()
            self.last_excel_export_path = None
            self.settings = self._get_default_settings() # Reset to defaults
            self._ensure_next_ids() # Reset ID counters based on default settings
            self.password_manager = PasswordManager(self.settings) # Reset password manager with fresh settings

            # Delete data files
            files_to_delete = [
                DATA_FILE, CUSTOM_BEHAVIORS_FILE, CUSTOM_HOMEWORKS_FILE,
                STUDENT_GROUPS_FILE, QUIZ_TEMPLATES_FILE, HOMEWORK_TEMPLATES_FILE,
                get_app_data_path(f"custom_homework_session_types_{CURRENT_DATA_VERSION_TAG}.json"),
                AUTOSAVE_EXCEL_FILE # Delete autosaved excel too
            ]
            # Attempt to delete old version files if they exist from previous versions
            for i in range(1, int(CURRENT_DATA_VERSION_TAG[1:])):
                files_to_delete.append(get_app_data_path(f"classroom_data_v{i}.json"))
                files_to_delete.append(get_app_data_path(f"custom_behaviors_v{i}.json"))
                files_to_delete.append(get_app_data_path(f"custom_homeworks_v{i}.json"))
                files_to_delete.append(get_app_data_path(f"student_groups_v{i}.json"))
                files_to_delete.append(get_app_data_path(f"quiz_templates_v{i}.json"))
                files_to_delete.append(get_app_data_path(f"homework_templates_v{i}.json"))
                files_to_delete.append(get_app_data_path(f"custom_homework_session_types_v{i}.json"))
                files_to_delete.append(get_app_data_path(f"autosave_log_v{i}.xlsx"))


            for f_path in files_to_delete:
                if os.path.exists(f_path):
                    try: os.remove(f_path)
                    except OSError as e: print(f"Warning: Could not delete file {f_path} during reset: {e}")
            
            # Delete layout templates directory contents
            if os.path.exists(LAYOUT_TEMPLATES_DIR):
                for item_name in os.listdir(LAYOUT_TEMPLATES_DIR):
                    item_path = os.path.join(LAYOUT_TEMPLATES_DIR, item_name)
                    try:
                        if os.path.isfile(item_path) or os.path.islink(item_path): os.unlink(item_path)
                        elif os.path.isdir(item_path): shutil.rmtree(item_path)
                    except Exception as e_del_layout: print(f"Failed to delete {item_path}: {e_del_layout}")


            # Save fresh default data (which will create new empty files)
            self.save_data_wrapper(source="reset")
            self.update_all_behaviors(); self.update_all_homework_log_behaviors(); self.update_all_homework_session_types()
            self.draw_all_items(check_collisions_on_redraw=True)
            self.update_undo_redo_buttons_state()
            self.update_lock_button_state()
            self.update_status("Application has been reset to default state.")
            messagebox.showinfo("Reset Complete", "Application reset successfully. All data and settings are now at their defaults.", parent=self.root)
        except Exception as e:
            messagebox.showerror("Reset Error", f"An error occurred during application reset: {e}", parent=self.root)
            self.update_status(f"Error during reset: {e}")

    def backup_all_data_dialog(self, force=False):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Backup", "Enter password to create a backup:"): return
        default_filename = f"{APP_NAME}_Backup_{CURRENT_DATA_VERSION_TAG}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        if force == False:
            backup_zip_path = filedialog.asksaveasfilename(
                title="Save Backup As",
                defaultextension=".zip",
                initialfile=default_filename,
                filetypes=[("ZIP archive", "*.zip")],
                parent=self.root
            )
            if not backup_zip_path:
                self.update_status("Backup cancelled."); return
        elif force == True:
            backup_zip_path = os.path.abspath(os.path.join(os.path.dirname(DATA_FILE), default_filename))
        # Ensure latest data is saved before backup
        self.save_data_wrapper(source="backup_preparation")

        files_to_backup = [
            DATA_FILE, CUSTOM_BEHAVIORS_FILE, CUSTOM_HOMEWORKS_FILE, STUDENT_GROUPS_FILE,
            QUIZ_TEMPLATES_FILE, HOMEWORK_TEMPLATES_FILE,
            get_app_data_path(f"custom_homework_session_types_{CURRENT_DATA_VERSION_TAG}.json"),
            # AUTOSAVE_EXCEL_FILE # Optional: include autosaved excel log
        ]
        # Also include all files in LAYOUT_TEMPLATES_DIR
        layout_template_files = []
        if os.path.exists(LAYOUT_TEMPLATES_DIR):
            for fname in os.listdir(LAYOUT_TEMPLATES_DIR):
                fpath = os.path.join(LAYOUT_TEMPLATES_DIR, fname)
                if os.path.isfile(fpath): layout_template_files.append(fpath)
        
        try:
            with zipfile.ZipFile(backup_zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for file_path in files_to_backup:
                    if os.path.exists(file_path) and os.path.isfile(file_path):
                        zf.write(file_path, arcname=os.path.basename(file_path))
                for file_path in layout_template_files:
                     zf.write(file_path, arcname=os.path.join(LAYOUT_TEMPLATES_DIR_NAME, os.path.basename(file_path)))
            self.update_status(f"Backup created: {os.path.basename(backup_zip_path)}")
            messagebox.showinfo("Backup Successful", f"All application data backed up to:\n{backup_zip_path}", parent=self.root)
        except Exception as e:
            messagebox.showerror("Backup Error", f"Failed to create backup: {e}", parent=self.root)
            self.update_status(f"Error creating backup: {e}")
        finally:
            self.password_manager.record_activity()

    def restore_all_data_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Restore", "Enter password to restore data:"): return

        if not messagebox.askyesno("Confirm Restore", "Restoring data will OVERWRITE all current application data (students, logs, settings, etc.) with the contents of the backup.\nThis action CANNOT BE UNDONE directly from the backup itself.\n\nAre you sure you want to proceed?", parent=self.root, icon='warning', default=messagebox.NO):
            self.update_status("Restore cancelled."); return

        backup_zip_path = filedialog.askopenfilename(
            title="Select Backup File to Restore",
            filetypes=[("ZIP archives", "*.zip")],
            parent=self.root
        )
        if not backup_zip_path:
            self.update_status("Restore cancelled."); return

        app_data_dir = os.path.dirname(DATA_FILE) # Get the directory where app data is stored
        
        try:
            with zipfile.ZipFile(backup_zip_path, 'r') as zf:
                # Preliminary check for main data file to guess version compatibility (optional)
                main_data_filename_in_zip = None
                expected_main_data_filename = os.path.basename(DATA_FILE) # e.g. classroom_data_v9.json
                
                # List of all possible data file names across versions to check against
                possible_main_data_files = [f"classroom_data_v{i}.json" for i in range(1,10)] + ["classroom_data.json"]
                
                found_compatible_main_file = False
                for name_in_zip in zf.namelist():
                    if os.path.basename(name_in_zip) in possible_main_data_files:
                        main_data_filename_in_zip = os.path.basename(name_in_zip)
                        found_compatible_main_file = True
                        break
                
                if not found_compatible_main_file:
                    messagebox.showerror("Restore Error", "The selected ZIP file does not appear to be a valid application backup (missing main data file).", parent=self.root)
                    return

                # Clear existing layout templates before extraction
                if os.path.exists(LAYOUT_TEMPLATES_DIR):
                    for item_name in os.listdir(LAYOUT_TEMPLATES_DIR):
                        item_path = os.path.join(LAYOUT_TEMPLATES_DIR, item_name)
                        try:
                            if os.path.isfile(item_path) or os.path.islink(item_path): os.unlink(item_path)
                            elif os.path.isdir(item_path): shutil.rmtree(item_path)
                        except Exception as e_del_layout: print(f"Failed to delete old layout item {item_path}: {e_del_layout}")
                else:
                    os.makedirs(LAYOUT_TEMPLATES_DIR, exist_ok=True)


                # Extract files directly into the application data directory
                # This will overwrite existing files with the same names.
                for member in zf.infolist():
                    # Handle paths correctly: extract to app_data_dir, but if member.filename includes a dir (like layout_templates), recreate that structure
                    target_path = os.path.join(app_data_dir, member.filename)
                    # Ensure parent directory exists for the target_path
                    target_dir_for_file = os.path.dirname(target_path)
                    if not os.path.exists(target_dir_for_file):
                        os.makedirs(target_dir_for_file, exist_ok=True)

                    if not member.is_dir(): # Check if it's a file
                        with open(target_path, "wb") as outfile:
                            outfile.write(zf.read(member.filename))
            
            # After extraction, reload data from the potentially new main data file.
            # The main data file name in the backup might be an older version.
            # The load_data method handles migration.
            path_to_load_after_restore = os.path.join(app_data_dir, main_data_filename_in_zip) if main_data_filename_in_zip else DATA_FILE

            self.load_data(file_path=path_to_load_after_restore, is_restore=True) # Reload all data from extracted files
            self.load_custom_behaviors(); self.load_custom_homework_log_behaviors(); self.load_custom_homework_session_types()
            self.load_student_groups(); self.load_quiz_templates(); self.load_homework_templates()
            self._ensure_next_ids() # Crucial after loading potentially old data
            self.update_all_behaviors(); self.update_all_homework_log_behaviors(); self.update_all_homework_session_types()
            self.draw_all_items(check_collisions_on_redraw=True)
            self.update_undo_redo_buttons_state()
            self.update_lock_button_state()
            self.toggle_student_groups_ui_visibility()
            self.mode_var.set(self.settings.get("current_mode", "behavior")); self.toggle_mode()


            self.update_status("Data restored successfully. Application reloaded.")
            messagebox.showinfo("Restore Successful", "Data restored from backup. The application has reloaded the restored data.", parent=self.root)

        except FileNotFoundError:
            messagebox.showerror("Restore Error", "Backup file not found.", parent=self.root)
        except zipfile.BadZipFile:
            messagebox.showerror("Restore Error", "Invalid or corrupted backup ZIP file.", parent=self.root)
        except Exception as e:
            messagebox.showerror("Restore Error", f"Failed to restore data: {e}", parent=self.root)
            self.update_status(f"Error restoring data: {e}")
            # Attempt to reload current (pre-restore attempt) data to stabilize
            self.load_data(DATA_FILE, is_restore=False)
            self.draw_all_items()
        finally:
            self.password_manager.record_activity()

    def open_data_folder(self):
        folder_path = os.path.dirname(DATA_FILE)
        try:
            if sys.platform == "win32": os.startfile(folder_path)
            elif sys.platform == "darwin": subprocess.Popen(["open", folder_path])
            else: subprocess.Popen(["xdg-open", folder_path])
            self.update_status(f"Opened data folder: {folder_path}")
        except Exception as e:
            self.update_status(f"Error opening data folder: {e}")
            messagebox.showerror("Error", f"Could not open data folder: {e}\nPath: {folder_path}", parent=self.root)

    def open_last_export_folder(self):
        if self.last_excel_export_path and os.path.exists(os.path.dirname(self.last_excel_export_path)):
            self.open_specific_export_folder(self.last_excel_export_path)
        else: self.update_status("Last export path not set or not found.")

    def open_specific_export_folder(self, file_path_in_folder):
        folder_path = os.path.dirname(file_path_in_folder)
        try:
            if sys.platform == "win32": os.startfile(folder_path)
            elif sys.platform == "darwin": subprocess.Popen(["open", folder_path])
            else: subprocess.Popen(["xdg-open", folder_path])
            self.update_status(f"Opened folder: {folder_path}")
        except Exception as e:
            self.update_status(f"Error opening folder {folder_path}: {e}")
            messagebox.showerror("Error", f"Could not open folder: {e}\nPath: {folder_path}", parent=self.root)


    def show_help_dialog(self):
        HelpDialog(self.root, APP_VERSION)


    def on_exit_protocol(self, force_quit=False):

        #dialog = ExitConfirmationDialog(self.root, "Exit Confirmation")
        #if dialog.result == "save_quit":
        #    self.save_data_wrapper(source="exit_save")
        #    self.root.destroy()
        #elif dialog.result == "no_save_quit":
        #    self.root.destroy()
        # If dialog.result is None (Cancel), do nothing.
        
        try:
            if not force_quit:
                if self.password_manager.is_locked:
                    if not self.prompt_for_password("Unlock to Save & Quit", "Enter password to save and quit:"): return

                if self.is_live_quiz_active and not self.prompt_end_live_session_on_mode_switch("quiz"): return
                if self.is_live_homework_active and not self.prompt_end_live_session_on_mode_switch("homework"): return

                #if messagebox.askyesno("Exit", "Save changes and exit application?", parent=self.root, ):
                #    self.save_data_wrapper(source="exit_protocol")
                #else: # User chose not to save, but still wants to exit
                #    self.update_status("Exited without saving.")
                dialog = ExitConfirmationDialog(self.root, "Exit Confirmation")
                if dialog.result == "save_quit":
                    self.save_data_wrapper(source="exit_protocol")
                    self.root.destroy()
                    sys.exit(0) # Ensure clean exit
                elif dialog.result == "no_save_quit":
                    #if self.file_lock_manager: self.file_lock_manager.release_lock()
                    self.update_status("Exited without saving.")
                    self.root.destroy()
                    sys.exit(0) # Ensure clean exit
            else: # Force quit (e.g. after save_and_quit or if lock fails)
                self.root.destroy()
                sys.exit(0) # Ensure clean exit # Data should have been saved by save_and_quit if called from there
        except Exception as e:
            print(f"Error during exit procedure: {e}") # Log error but proceed with exit
            self.root.destroy()
            sys.exit(0) # Ensure clean exit
        #finally:
        #    
        #    self.root.destroy()
        #    sys.exit(0) # Ensure clean exit
    




# --- Dialog Classes ---
class PasswordPromptDialog(simpledialog.Dialog):
    def __init__(self, parent, title, prompt, password_manager_instance):
        self.prompt = prompt
        self.password_manager = password_manager_instance
        self.result = False # True if password correct or not set, False otherwise
        super().__init__(parent, title)

    def body(self, master):
        ttk.Label(master, text=self.prompt, wraplength=280).pack(pady=5)
        self.password_entry = ttk.Entry(master, show="*", width=30)
        self.password_entry.pack(pady=5)
        self.status_label = ttk.Label(master, text="", foreground="red")
        self.status_label.pack(pady=2)
        return self.password_entry # initial focus

    def buttons(self):
        box = ttk.Frame(self)
        ttk.Button(box, text="OK", width=10, command=self.ok, default=tk.ACTIVE).pack(side=tk.LEFT, padx=5, pady=5)
        ttk.Button(box, text="Cancel", width=10, command=self.cancel).pack(side=tk.LEFT, padx=5, pady=5)
        self.bind("<Return>", lambda e: self.ok())
        self.bind("<Escape>", lambda e: self.cancel())
        box.pack()

    def apply(self):
        password_attempt = self.password_entry.get()
        if self.password_manager.unlock_application(password_attempt): # unlock_application handles recovery pw too
            self.result = True
        else:
            self.status_label.config(text="Incorrect password.")
            self.password_entry.focus_set()
            self.password_entry.select_range(0, tk.END)
            self.result = False # Explicitly set to false, stay open
            # To prevent dialog from closing on incorrect password, we don't call super().ok() or cancel
            # Instead, the dialog stays open. The caller of this dialog will check self.result.
            # This means the typical simpledialog auto-close on OK needs to be managed carefully.
            # For this use case, we want it to close only on success or cancel.
            # The `ok` method will be called. If result is False, the dialog won't close.
            # We need to ensure `ok` can prevent closing.
            # A simple way is to override `ok` more directly or ensure `validate` fails.
            # simpledialog.Dialog closes if validate() returns true.

    def ok(self, event=None): # Override ok to control closing
        if not self.validate(): # validate calls apply
            self.password_entry.focus_set()
            return # Don't close if validation (password check) fails
        self.withdraw()
        self.update_idletasks()
        # apply has already been called by validate
        self.parent.focus_set() # Give focus back to parent
        self.destroy()


    def validate(self): # This is called by simpledialog's ok method
        self.apply() # Calls our apply which sets self.result
        return self.result # If True, dialog closes. If False, stays open due to logic in apply.

class AddEditStudentDialog(simpledialog.Dialog):
    # ... (similar to v51, but needs group dropdown to be populated from app.student_groups)
    def __init__(self, parent, title, student_data=None, app=None):
        self.student_data = student_data
        self.app = app # To access student_groups
        self.result = None
        super().__init__(parent, title)

    def body(self, master):
        ttk.Label(master, text="First Name:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        self.fn_var = tk.StringVar(value=self.student_data["first_name"] if self.student_data else "")
        self.fn_entry = ttk.Entry(master, textvariable=self.fn_var, width=30); self.fn_entry.grid(row=0, column=1, padx=5, pady=2)

        ttk.Label(master, text="Last Name:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        self.ln_var = tk.StringVar(value=self.student_data["last_name"] if self.student_data else "")
        self.ln_entry = ttk.Entry(master, textvariable=self.ln_var, width=30); self.ln_entry.grid(row=1, column=1, padx=5, pady=2)

        ttk.Label(master, text="Nickname (Optional):").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        self.nick_var = tk.StringVar(value=self.student_data.get("nickname", "") if self.student_data else "")
        self.nick_entry = ttk.Entry(master, textvariable=self.nick_var, width=30); self.nick_entry.grid(row=2, column=1, padx=5, pady=2)

        ttk.Label(master, text="Gender:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=2)
        self.gender_var = tk.StringVar(value=self.student_data.get("gender", "Boy") if self.student_data else "Boy")
        gender_frame = ttk.Frame(master)
        ttk.Radiobutton(gender_frame, text="Boy", variable=self.gender_var, value="Boy").pack(side=tk.LEFT)
        ttk.Radiobutton(gender_frame, text="Girl", variable=self.gender_var, value="Girl").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(gender_frame, text="Other", variable=self.gender_var, value="Other").pack(side=tk.LEFT)
        gender_frame.grid(row=3, column=1, sticky=tk.W, padx=5, pady=2)

        if self.app and self.app.settings.get("student_groups_enabled", True):
            ttk.Label(master, text="Group:").grid(row=4, column=0, sticky=tk.W, padx=5, pady=2)
            self.group_var = tk.StringVar()
            group_options = {"NONE_GROUP_SENTINEL": "No Group"} # Sentinel for no group
            for gid, gdata in sorted(self.app.student_groups.items(), key=lambda item: item[1]['name']):
                group_options[gid] = gdata['name']
            
            self.group_combobox = ttk.Combobox(master, textvariable=self.group_var, values=list(group_options.values()), state="readonly", width=28)
            self.group_combobox_map = {name: gid for gid, name in group_options.items()} # Map display name back to ID
            
            current_group_id = self.student_data.get("group_id") if self.student_data else None
            if current_group_id and current_group_id in self.app.student_groups:
                self.group_var.set(self.app.student_groups[current_group_id]['name'])
            else:
                self.group_var.set("No Group")
            self.group_combobox.grid(row=4, column=1, padx=5, pady=2, sticky=tk.W)
            self.group_combobox.bind("<MouseWheel>", lambda event: "break") # Prevent main canvas scroll

        return self.fn_entry

    def apply(self):
        first_name = self.fn_var.get().strip()
        last_name = self.ln_var.get().strip()
        nickname = self.nick_var.get().strip()
        gender = self.gender_var.get()
        group_id_selection = None
        if self.app and self.app.settings.get("student_groups_enabled", True) and hasattr(self, 'group_var'):
            selected_group_name = self.group_var.get()
            group_id_selection = self.group_combobox_map.get(selected_group_name)
            if group_id_selection == "NONE_GROUP_SENTINEL": group_id_selection = None

        if first_name and last_name:
            self.result = (first_name, last_name, nickname, gender, group_id_selection)
        else:
            messagebox.showwarning("Missing Information", "First Name and Last Name are required.", parent=self)
            self.result = None # Stay open

class AddFurnitureDialog(simpledialog.Dialog):
    # ... (same as v51)
    def __init__(self, parent, title, furniture_data=None):
        self.furniture_data = furniture_data
        self.result = None
        super().__init__(parent, title)
    def body(self, master):
        ttk.Label(master, text="Name:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        self.name_var = tk.StringVar(value=self.furniture_data["name"] if self.furniture_data else "Desk")
        self.name_entry = ttk.Entry(master, textvariable=self.name_var); self.name_entry.grid(row=0, column=1, padx=5, pady=2)
        ttk.Label(master, text="Type:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        self.type_var = tk.StringVar(value=self.furniture_data["type"] if self.furniture_data else "Rebbi's Desk")
        self.type_entry = ttk.Entry(master, textvariable=self.type_var); self.type_entry.grid(row=1, column=1, padx=5, pady=2)
        ttk.Label(master, text="Width:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        self.width_var = tk.IntVar(value=self.furniture_data["width"] if self.furniture_data else REBBI_DESK_WIDTH)
        self.width_spinbox = ttk.Spinbox(master, from_=20, to=1000, textvariable=self.width_var, width=5); self.width_spinbox.grid(row=2, column=1, sticky=tk.W, padx=5, pady=2)
        ttk.Label(master, text="Height:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=2)
        self.height_var = tk.IntVar(value=self.furniture_data["height"] if self.furniture_data else REBBI_DESK_HEIGHT)
        self.height_spinbox = ttk.Spinbox(master, from_=20, to=1000, textvariable=self.height_var, width=5); self.height_spinbox.grid(row=3, column=1, sticky=tk.W, padx=5, pady=2)
        return self.name_entry
    def apply(self):
        name = self.name_var.get().strip(); item_type = self.type_var.get().strip()
        width = self.width_var.get(); height = self.height_var.get()
        if name and item_type and width > 0 and height > 0: self.result = (name, item_type, width, height)
        else: messagebox.showwarning("Invalid Input", "Name, Type, Width, and Height are required and must be positive.", parent=self); self.result = None

class BehaviorDialog(simpledialog.Dialog):
    # ... (same as v51)
    def __init__(self, parent, title, all_behaviors, custom_behaviors):
        self.all_behaviors = all_behaviors
        self.custom_behaviors = custom_behaviors # For potential editing in future, not used now
        self.result = None
        self.selected_behavior_var = tk.StringVar()
        super().__init__(parent, title)
    def body(self, master):
        
        master.grid_columnconfigure(0, weight=1) # Allow master frame to expand
        top_label = ttk.Label(master, text="Select Behavior:")
        top_label.grid(row=0, column=0, columnspan=6, sticky="nw", pady=2, padx=5)

        # Frame for buttons with scrollbar
        button_frame_outer = ttk.Frame(master)
        button_frame_outer.grid(row=1, column=0, columnspan=6, sticky="nsew", pady=(5,10), )
        master.grid_rowconfigure(1, weight=1) # Allow button area to expand

        theheight=0
        for i in range(0,len(self.all_behaviors),6): theheight += 115
        
        btn_canvas = tk.Canvas(button_frame_outer, borderwidth=0,width=1000,height= theheight)
        btn_scrollbar = ttk.Scrollbar(button_frame_outer, orient="vertical", command=btn_canvas.yview)
        scrollable_frame_for_buttons = ttk.Frame(btn_canvas)

        scrollable_frame_for_buttons.bind("<Configure>", lambda e: btn_canvas.configure(scrollregion=btn_canvas.bbox("all")))
        btn_canvas.create_window((0, 0), window=scrollable_frame_for_buttons, anchor="nw")
        btn_canvas.configure(yscrollcommand=btn_scrollbar.set)

        btn_canvas.pack(side="left", fill="both", expand=True)
        btn_scrollbar.pack(side="right", fill="y")

        # Populate buttons
        row_idx, col_idx = 0, 0
        max_cols = 6 # Adjust number of columns as desired
        button_width_chars = 14 # Approximate characters
        button_height_lines = 5 # Approximate lines
        button_width = 17
        button_height = 5
        wraplength=button_width_chars*7
        button_font = ('TkDefaultFont', 15)
        b = ttk.Style(); b.configure("Keypad.TButton", font=button_font, wraplength=button_width_chars*10,justify='center',padding=(0))
        for i, behavior_text in enumerate(self.all_behaviors):
            btn = ttk.Button(scrollable_frame_for_buttons, text=behavior_text,
                            padding=(0,19),# Rough estimate for pixel width
                            width=button_width_chars, style="Keypad.TButton",
                            command=lambda n=behavior_text: self.behavior_selected(n))
            btn.grid(row=row_idx, column=col_idx, sticky="nsew", padx=2, pady=3)
            scrollable_frame_for_buttons.grid_columnconfigure(col_idx, weight=1) # Allow columns to expand
            col_idx += 1
            if col_idx >= max_cols:
                col_idx = 0
                row_idx += 1
        if col_idx != 0: # Ensure last row also configures row weight if partially filled
             scrollable_frame_for_buttons.grid_rowconfigure(row_idx, weight=1)
        

        comment_label = ttk.Label(master, text="Additional Comment (Optional):")
        comment_label.grid(row=2,column=0,columnspan=6, sticky="sw", pady=(8,2), padx=5)
        self.comment_text_widget = tk.Text(master, width=50, height=4, wrap=tk.WORD)
        self.comment_text_widget.grid(row=3, column=0,columnspan=6, pady=(0,5), padx=5, sticky="sew")
        master.grid_rowconfigure(3, weight=0) # Comment text doesn't expand as much

        # Set initial focus if desired, e.g., to the first button or comment box
        # For now, default Tk focus handling.
        return self.comment_text_widget # Or another widget for initial focus


        """        
        
        ttk.Label(master, text="Select Behavior:").pack(pady=5)
        self.behavior_var = tk.StringVar()
        self.behavior_combobox = ttk.Combobox(master, textvariable=self.behavior_var, values=self.all_behaviors, width=30)
        if self.all_behaviors: self.behavior_combobox.set(self.all_behaviors[0])
        self.behavior_combobox.pack(pady=5)
        self.behavior_combobox.bind("<MouseWheel>", lambda event: "break")
        ttk.Label(master, text="Comment (Optional):").pack(pady=5)
        self.comment_entry = ttk.Entry(master, width=33); self.comment_entry.pack(pady=5)
        return self.behavior_combobox
        """
    
    
    
    def behavior_selected(self, behavior_name):
        self.selected_behavior_var.set(behavior_name)
        self.ok() # Trigger apply and close

    def buttonbox(self): # Standard OK/Cancel are usually handled by simpledialog.Dialog
        # We only need a cancel button as OK is triggered by behavior selection
        cancel_button_frame = ttk.Frame(self)
        cancel_button_frame.pack(fill=tk.X, padx=5, pady=(0,5)) # Pad bottom
        ttk.Button(cancel_button_frame, text="Cancel", width=10, command=self.cancel).pack(side=tk.RIGHT, padx=5)
        self.bind("<Escape>", self.cancel)

    def apply(self):
        behavior = self.selected_behavior_var.get()
        comment = self.comment_text_widget.get("1.0", tk.END).strip()
        if not behavior: # Should not happen if ok() is only called on selection
            messagebox.showwarning("Input Error", "Please select a behavior.", parent=self)
            self.result = None
            return
        self.result = (behavior, comment)
    
    
    #def apply(self):
    #    behavior = self.behavior_var.get().strip(); comment = self.comment_entry.get().strip()
    #    if behavior: self.result = (behavior, comment)
    #    else: messagebox.showwarning("Input Required", "Please select or enter a behavior.", parent=self); self.result = None




class ManualHomeworkLogDialog(simpledialog.Dialog): # New
    def __init__(self, parent, title, all_homework_types, custom_homework_types, log_marks_enabled, homework_mark_types, homework_templates):
        self.all_homework_types = all_homework_types # List of strings (behavior names for homework)
        self.custom_homework_types = custom_homework_types # List of dicts (unused for now, for future editing)
        self.log_marks_enabled = log_marks_enabled
        self.homework_mark_types = homework_mark_types # List of dicts, e.g. {"id": "hmark_complete", "name":"Complete", "default_points":10}
        self.homework_templates = homework_templates # Dict of templates
        self.result = None
        self.mark_entry_vars = {} # {mark_type_id: StringVar}
        super().__init__(parent, title)

    def body(self, master):
        main_frame = ttk.Frame(master); main_frame.pack(fill=tk.BOTH,side=tk.LEFT, expand=True)

        # Homework Type / Template Selection
        type_frame = ttk.Frame(main_frame); type_frame.pack(pady=5, fill=tk.X)
        ttk.Label(type_frame, text="Homework Type/Name:").pack(side=tk.TOP, padx=5)
        self.homework_type_var = tk.StringVar()
        # Combine log behaviors and template names for the combobox
        combined_options = sorted(list(set(self.all_homework_types + [tpl['name'] for tpl_id, tpl in self.homework_templates.items()])))
        
        
        #self.homework_type_combobox.bind("<MouseWheel>", lambda event: "break")


        # Number of Items (relevant if marks are enabled and not using a template that defines it)
        self.num_items_frame = ttk.Frame(main_frame); # Packed later if needed
        ttk.Label(self.num_items_frame, text="Number of Items/Questions:").pack(side=tk.LEFT, padx=5)
        self.num_items_var = tk.StringVar(value="10") # Default
        self.num_items_spinbox = ttk.Spinbox(self.num_items_frame, from_=1, to=200, textvariable=self.num_items_var, width=5)
        self.num_items_spinbox.pack(side=tk.LEFT, padx=5)
        theheight=0
        for i in range(0,len(combined_options),4): theheight += 75
        btn_canvas = tk.Canvas(type_frame, borderwidth=0,width=660,height= theheight)
        btn_scrollbar = ttk.Scrollbar(type_frame, orient="vertical", command=btn_canvas.yview)
        scrollable_frame_for_buttons = ttk.Frame(btn_canvas)

        scrollable_frame_for_buttons.bind("<Configure>", lambda e: btn_canvas.configure(scrollregion=btn_canvas.bbox("all")))
        

        # Marks Frame (if enabled)
        self.marks_widgets_frame = ttk.LabelFrame(main_frame, text="Marks Details") # Packed later
        if self.log_marks_enabled and self.homework_mark_types:
            
            self.homework_type_combobox = ttk.Combobox(type_frame, textvariable=self.homework_type_var, values=combined_options, width=30, state="readonly")
            
            
            if combined_options: self.homework_type_combobox.set(combined_options[0])
            self.homework_type_combobox.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
            
            self.homework_type_combobox.bind("<<ComboboxSelected>>", self.on_template_select)
            self.num_items_frame.pack(pady=5, fill=tk.X)
            self.marks_widgets_frame.pack(pady=10, padx=5, fill=tk.BOTH, expand=True)
            self.target_entry2 = []
            self.target_entry3 = []
            self.cte = -1
            cols = 2 # Two columns for marks
            current_col = 0; current_row_marks = 0
            for i, mark_type in enumerate(self.homework_mark_types):
                mark_id = mark_type["id"]; mark_name = mark_type["name"]
                ttk.Label(self.marks_widgets_frame, text=f"{mark_name}:").grid(row=current_row_marks, column=current_col*2, sticky=tk.W, padx=5, pady=3)
                var = tk.StringVar()
                # Try to set default value from mark_type if it makes sense (e.g., for non-point entries or typical scores)
                # For point-based ones, usually user input is expected.
                if "default_value" in mark_type : var.set(str(mark_type["default_value"]))

                entry = ttk.Entry(self.marks_widgets_frame, textvariable=var, width=8)
                entry.grid(row=current_row_marks, column=current_col*2 + 1, sticky=tk.EW, padx=5, pady=3)
                self.mark_entry_vars[mark_id] = var
                entry.bind("<FocusIn>", lambda x=i,x2=i: self.set_numpad(x,x2))
                self.target_entry2.append(var)
                self.target_entry3.append(mark_type["name"])
                self.cte = i
                self.tte = i
                current_col += 1
                if current_col >= cols: current_col = 0; current_row_marks +=1
            #print(main_frame.slaves)
            self.keypad_frame = ttk.Frame(master, relief='sunken',borderwidth=10)
            self.keypad_frame.pack(side=tk.RIGHT)
            self.keypad_frame.grid_propagate(True)
            self.allow_decimal = True
            self.mark_vars2 = {} # {mark_type_id: tk.StringVar()}
            for mt in self.mark_vars2:
                self.mark_vars2[mt["id"]] = tk.StringVar()
            
            for i, mt_config in enumerate(self.mark_vars2):
                label_text = mt_config["name"]
                if mt_config.get("is_extra_credit"): label_text += " (Bonus)"
            #print(self.mark_vars2)
            self.target_entry = self.mark_entry_vars[mark_type["id"]]
            self.target_entry_name = mark_type["name"]
            
            self.target_name_label = ttk.Label(self.keypad_frame,text=self.target_entry_name)
            self.target_name_label.grid(column=0,row=0,columnspan=9)
            
            buttons = [('7',1,0),('8',1,1),('9',1,2),('4',2,0),('5',2,1),('6',2,2),('1',3,0),('2',3,1),('3',3,2),('0',4,0),('.',4,1) if self.allow_decimal else (' ',4,1),('⌫',4,2),('/',2,3)]
            button_font = ('TkDefaultFont', 15)
            self._next_entry()
            for i in range(4): self.rowconfigure(i, minsize=2000,weight=1)
            for i in range(4): self.columnconfigure(i, weight=1)
            for (text, r, c) in buttons:
                if text == ' ': continue
                action = lambda x=text: self._on_press(x)
                ttk.Button(self.keypad_frame, text=text, command=action, style="Keypad.TButton",padding=17,width=5).grid(row=r, column=c, padx=1, pady=1, sticky="nsew")
            ttk.Button(self.keypad_frame, text="Clear", command=self._clear_entry, style="Keypad.TButton").grid(row=1, column=3, padx=1, pady=1, sticky="nsew")
            ttk.Button(self.keypad_frame, text='Next', command=self._next_entry, style="Keypad.TButton").grid(row=4,column=3,padx=1,pady=1,sticky='nsew')
            ttk.Button(self.keypad_frame, text='Previous', command=self._previous_entry, style="Keypad.TButton").grid(row=3,column=3,padx=1,pady=1,sticky='nsew')
            s = ttk.Style(); s.configure("Keypad.TButton", font=button_font, padding=(5,10),height=100)
        
            
                
            

            #print(self.target_entry_name)
            
            for i in range(cols*2): # Configure columns to expand
                 self.marks_widgets_frame.grid_columnconfigure(i, weight=1 if i%2==1 else 0)
            




        else:
            
            btn_canvas.pack(side=tk.BOTTOM)
            btn_canvas.create_window((0, 0), window=scrollable_frame_for_buttons, anchor="nw")
            btn_canvas.configure(yscrollcommand=btn_scrollbar.set)
            row_idx, col_idx = 0, 0
            max_cols = 4 # Adjust number of columns as desired
            button_width_chars = 14 # Approximate characters
            button_height_lines = 5 # Approximate lines
            button_width = 17
            button_height = 5
            wraplength=button_width_chars*7
            button_font = ('TkDefaultFont', 15)
            b = ttk.Style(); b.configure("Keypad.TButton", font=button_font, wraplength=button_width_chars*10,justify='center',padding=(0))
            for i, behavior_text in enumerate(combined_options):
                btn = ttk.Button(scrollable_frame_for_buttons, text=behavior_text,
                                padding=(0,19),# Rough estimate for pixel width
                                width=button_width_chars, style="Keypad.TButton",
                                command=lambda n=behavior_text: self.on_button_select(n))
                btn.grid(row=row_idx, column=col_idx, sticky="nsew", padx=2, pady=3)
                type_frame.grid_columnconfigure(col_idx, weight=1) # Allow columns to expand
                col_idx += 1
                if col_idx >= max_cols:
                    col_idx = 0
                    row_idx += 1
            if col_idx != 0: # Ensure last row also configures row weight if partially filled
                scrollable_frame_for_buttons.grid_rowconfigure(row_idx, weight=1)

        ttk.Label(main_frame, text="Comment (Optional):").pack(pady=5, anchor=tk.W, padx=5)
        self.comment_entry = ttk.Entry(main_frame, width=45)
        self.comment_entry.pack(pady=5, fill=tk.X, padx=5)
        
        self.on_template_select(None) # Initialize based on default selection
        return self.homework_type_combobox if self.log_marks_enabled else self.comment_entry

    def on_button_select(self, name):
        self.homework_type_var.set(name)
        self.apply()
        self.ok()

    def _on_press(self, key):
        current_text = self.target_entry.get()
        if key == '⌫': n = len(current_text)-1; self.target_entry.set(current_text[0:n])
        elif key == '.':
            if self.allow_decimal and '.' not in current_text: self.target_entry.set(self.target_entry.get()+ key)
        else: self.target_entry.set(self.target_entry.get()+ key)
        
    def _clear_entry(self): self.target_entry.set("")
    def _next_entry(self):
        if self.cte < self.tte: self.cte +=1
        else: self.cte = 0
        self.target_entry = self.target_entry2[(self.cte)]
        self.target_entry_name = self.target_entry3[(self.cte)]
        self.target_name_label.configure(text=self.target_entry_name)
        
    def _previous_entry(self):
        if self.cte < self.tte and self.cte > -2: self.cte -=1
        else: self.cte = (self.tte-1)
        self.target_entry = self.target_entry2[(self.cte)]
        self.target_entry_name = self.target_entry3[(self.cte)]
        self.target_name_label.configure(text=self.target_entry_name)
        
    def set_numpad(self, event, x):
        #print(x)
        self.cte = x
        self.target_entry = self.target_entry2[(self.cte)]
        self.target_entry_name = self.target_entry3[(self.cte)]
        self.target_name_label.configure(text=self.target_entry_name)
    



    def on_template_select(self, event):
        selected_name = self.homework_type_var.get()
        template = next((tpl for tpl_id, tpl in self.homework_templates.items() if tpl['name'] == selected_name), None)

        if template:
            self.num_items_var.set(str(template.get("num_items", 10)))
            if self.num_items_frame.winfo_ismapped(): self.num_items_spinbox.config(state=tk.DISABLED) # Disable if template provides it

            if self.log_marks_enabled and "default_marks" in template:
                for mark_id, var in self.mark_entry_vars.items():
                    var.set(str(template["default_marks"].get(mark_id, ""))) # Use template's default marks
        else: # Not a template, or template doesn't have these fields
            if self.num_items_frame.winfo_ismapped(): self.num_items_spinbox.config(state=tk.NORMAL)
            # Clear mark entries if not a template or template has no defaults
            # for var in self.mark_entry_vars.values(): var.set("") # Or set to individual mark_type defaults


    def apply(self):
        homework_type = self.homework_type_var.get().strip()
        comment = self.comment_entry.get().strip()
        num_items_val = None
        marks_data = {}

        if not homework_type:
            messagebox.showwarning("Input Required", "Please select or enter a homework type/name.", parent=self)
            self.result = None; return

        if self.log_marks_enabled:
            try: num_items_val = int(self.num_items_var.get())
            except ValueError:
                messagebox.showwarning("Invalid Input", "Number of items must be a valid integer.", parent=self)
                self.result = None; return
            if num_items_val <= 0 :
                messagebox.showwarning("Invalid Input", "Number of items must be positive.", parent=self)
                self.result = None; return

            for mark_id, var in self.mark_entry_vars.items():
                val_str = var.get().strip()
                if val_str: # Only store if a value is entered
                    # Try to convert to float if possible, else store as string
                    try: marks_data[mark_id] = float(val_str)
                    except ValueError: marks_data[mark_id] = val_str
        
        self.result = (homework_type, comment, marks_data if self.log_marks_enabled else None, num_items_val if self.log_marks_enabled else None)

"""
# --- QuizScoreDialog needs significant changes for v51 ---
class QuizScoreDialog(simpledialog.Dialog):
    def __init__(self, parent, title, initial_quiz_name, mark_type_configs, quiz_templates, default_num_questions, initial_num_questions):
        self.initial_quiz_name = initial_quiz_name
        self.mark_type_configs = sorted(mark_type_configs, key=lambda mt: (mt.get("is_extra_credit", False), mt.get("name", ""))) # Sort for consistent display
        self.quiz_templates = quiz_templates
        self.default_num_questions = default_num_questions
        initial_num_questions_var = initial_num_questions
                

        self.quiz_name_var = tk.StringVar(value=initial_quiz_name)
        if initial_num_questions_var == default_num_questions:
            self.num_questions_var = tk.IntVar(value=default_num_questions)
        else:
            self.num_questions_var = tk.IntVar(value=initial_num_questions)
        self.template_var = tk.StringVar() # For selected template name
        self.initial_num_questions_var = tk.IntVar(value=initial_num_questions)
        self.mark_vars2 = {} # {mark_type_id: tk.StringVar()}
        for mt in self.mark_vars2:
            self.mark_vars2[mt["id"]] = tk.StringVar()
        
        for i, mt_config in enumerate(self.mark_vars2):
            label_text = mt_config["name"]
            if mt_config.get("is_extra_credit"): label_text += " (Bonus)"
            
            self.mark_vars2[mt_config["id"]]

        self.mark_vars = {} # {mark_type_id: tk.StringVar()}
        for mt in self.mark_type_configs:
            self.mark_vars[mt["id"]] = tk.StringVar()

        self.comment_text_widget = None
        super().__init__(parent, title)

    def body(self, master):
        main_frame = ttk.Frame(master); main_frame.grid(sticky="nsew",column=0,row=0, padx=10, pady=10)

        # Quiz Name and Number of Questions
        name_q_frame = ttk.Frame(main_frame); name_q_frame.pack(fill=tk.X, pady=(0,10))
        ttk.Label(name_q_frame, text="Quiz Name:").grid(row=0, column=0, sticky="w", padx=(0,5))
        self.quiz_name_entry = ttk.Entry(name_q_frame, textvariable=self.quiz_name_var, width=30)
        self.quiz_name_entry.grid(row=0, column=1, sticky="ew", padx=(0,10))
        name_q_frame.grid_columnconfigure(1, weight=1)

        ttk.Label(name_q_frame, text="# Questions:").grid(row=0, column=2, sticky="w", padx=(10,5))
        self.num_questions_spinbox = ttk.Spinbox(name_q_frame, from_=1, to=200, textvariable=self.num_questions_var, width=5)
        self.num_questions_spinbox.grid(row=0, column=3, sticky="w")

        # Quiz Templates (Optional)
        if self.quiz_templates:
            template_frame = ttk.Frame(main_frame); template_frame.pack(fill=tk.X, pady=(0,10))
            ttk.Label(template_frame, text="Load Template:").grid(row=0, column=0, sticky="w", padx=(0,5))
            template_names = [""] + [tpl["name"] for tpl_id, tpl in sorted(self.quiz_templates.items(), key=lambda item: item[1]['name'])]
            self.template_combo = ttk.Combobox(template_frame, textvariable=self.template_var, values=template_names, width=28, state="readonly")
            self.template_combo.grid(row=0, column=1, sticky="ew")
            template_frame.grid_columnconfigure(1, weight=1)
            self.template_combo.bind("<<ComboboxSelected>>", self.apply_quiz_template)
        
        # Mark Entry Fields (Dynamically created)
        marks_frame = ttk.LabelFrame(main_frame, text="Enter Scores by Mark Type", padding=10)
        marks_frame.pack(fill=tk.BOTH, expand=True, pady=(0,10))
        marks_frame.grid_columnconfigure(1, weight=1) # Allow entry fields to expand
        self.target_entry2 = []
        self.target_entry3 = []
        self.cte = -1
        
        for i, mt_config in enumerate(self.mark_type_configs):
            
            label_text = mt_config["name"]
            if mt_config.get("is_extra_credit"): label_text += " (Bonus)"
            ttk.Label(marks_frame, text=f"{label_text}:").grid(row=i, column=0, sticky="w", pady=3, padx=5)
            entry = ttk.Entry(marks_frame, textvariable=self.mark_vars[mt_config["id"]], width=10)
            entry.grid(row=i, column=1, sticky="ew", pady=3, padx=5)
            entry.bind("<FocusIn>", lambda x=i,x2=i: self.set_numpad(x,x2))
            self.target_entry2.append(self.mark_vars[mt_config["id"]])
            self.target_entry3.append(mt_config["name"])
            self.cte = i
            self.tte = i
            
            
            # Could add default points as placeholder or helper text here
        #print(self.tte)
        #self.cte -=1
        #print((self.mark_type_configs))
        #print(self.mark_vars[mt["mark2"]]) # type: ignore
        #print(self.mark_vars[mt_config["id"]]) # type: ignore
        
        #self.keypad_frame = NumericKeypad(master, self.mark_vars[mt_config["id"]],self.mark_type_configs, allow_decimal=True) # type: ignore
        #self.keypad_frame.pack(padx=(10,5), pady=5,side=tk.RIGHT, fill=tk.BOTH)
        self.keypad_frame = ttk.Frame(master, relief='sunken',borderwidth=10)
        self.keypad_frame.grid(sticky="nsew",column=1,row=0)
        self.keypad_frame.grid_propagate(True)
        self.allow_decimal = True
        self.mark_vars2 = {} # {mark_type_id: tk.StringVar()}
        for mt in self.mark_vars2:
            self.mark_vars2[mt["id"]] = tk.StringVar()
        
        for i, mt_config in enumerate(self.mark_vars2):
            label_text = mt_config["name"]
            if mt_config.get("is_extra_credit"): label_text += " (Bonus)"
        
        self.target_entry = self.mark_vars[mt_config["id"]]
        self.target_entry_name = mt_config["name"]
        #print(self.target_entry_name)
        self.target_name_label = ttk.Label(self.keypad_frame,text=self.target_entry_name)
        self.target_name_label.grid(column=0,row=0,columnspan=9)
        self._next_entry()
        
        buttons = [('7',1,0),('8',1,1),('9',1,2),('4',2,0),('5',2,1),('6',2,2),('1',3,0),('2',3,1),('3',3,2),('0',4,0),('.',4,1) if self.allow_decimal else (' ',4,1),('⌫',4,2),('/',2,3)]
        button_font = ('TkDefaultFont', 15)
        
        for i in range(4): self.rowconfigure(i, minsize=2000,weight=1)
        for i in range(4): self.columnconfigure(i, weight=1)
        for (text, r, c) in buttons:
            if text == ' ': continue
            action = lambda x=text: self._on_press(x)
            ttk.Button(self.keypad_frame, text=text, command=action, style="Keypad.TButton",padding=17,width=5).grid(row=r, column=c, padx=1, pady=1, sticky="nsew")
        ttk.Button(self.keypad_frame, text="Clear", command=self._clear_entry, style="Keypad.TButton").grid(row=1, column=3, padx=1, pady=1, sticky="nsew")
        ttk.Button(self.keypad_frame, text='Next', command=self._next_entry, style="Keypad.TButton").grid(row=4,column=3,padx=1,pady=1,sticky='nsew')
        ttk.Button(self.keypad_frame, text='Previous', command=self._previous_entry, style="Keypad.TButton").grid(row=3,column=3,padx=1,pady=1,sticky='nsew')
        s = ttk.Style(); s.configure("Keypad.TButton", font=button_font, padding=(5,10),height=100)
    
        
        
        
        #print(self.mark_type_configs)
        #main_frame.columnconfigure(2, weight=0)
        # Comment
        comment_frame = ttk.LabelFrame(main_frame, text="Comment (Optional)", padding=5)
        comment_frame.pack(fill=tk.X, pady=(0,5))
        self.comment_text_widget = tk.Text(comment_frame, width=45, height=3, wrap=tk.WORD)
        self.comment_text_widget.pack(fill=tk.X, expand=True)

        self.quiz_name_entry.focus_set()
        return self.quiz_name_entry
    def _on_press(self, key):
        current_text = self.target_entry.get()
        if key == '⌫': n = len(current_text)-1; self.target_entry.set(current_text[0:n])
        elif key == '.':
            if self.allow_decimal and '.' not in current_text: self.target_entry.set(self.target_entry.get()+ key)
        else: self.target_entry.set(self.target_entry.get()+ key)
        
    def _clear_entry(self): self.target_entry.set("")
    def _next_entry(self):
        if self.cte < self.tte: self.cte +=1
        else: self.cte = 0
        self.target_entry = self.target_entry2[(self.cte)]
        self.target_entry_name = self.target_entry3[(self.cte)]
        self.target_name_label.configure(text=self.target_entry_name)
        
    def _previous_entry(self):
        if self.cte < self.tte and self.cte > -2: self.cte -=1
        else: self.cte = (self.tte-1)
        self.target_entry = self.target_entry2[(self.cte)]
        self.target_entry_name = self.target_entry3[(self.cte)]
        self.target_name_label.configure(text=self.target_entry_name)
        
    def set_numpad(self, event, x):
        #print(x)
        self.cte = x
        self.target_entry = self.target_entry2[(self.cte)]
        self.target_entry_name = self.target_entry3[(self.cte)]
        self.target_name_label.configure(text=self.target_entry_name)
        
    def apply_quiz_template(self, event=None):
        selected_template_name = self.template_var.get()
        if not selected_template_name: return # No template selected

        template_id_found = None
        for tpl_id, tpl_data in self.quiz_templates.items():
            if tpl_data["name"] == selected_template_name:
                template_id_found = tpl_id
                break

        if template_id_found:
            template = self.quiz_templates[template_id_found]
            self.quiz_name_var.set(template.get("quiz_name_override", self.quiz_name_var.get())) # Update quiz name if in template
            self.num_questions_var.set(template.get("num_questions", self.num_questions_var.get()))
            for mt_id, mark_var in self.mark_vars.items():
                mark_var.set(template.get("default_marks", {}).get(mt_id, "")) # Set default scores for marks

    def validate(self):
        quiz_name = self.quiz_name_var.get().strip()
        if not quiz_name:
            messagebox.showwarning("Input Error", "Quiz name is required.", parent=self)
            return False
        try:
            num_q = self.num_questions_var.get()
            if num_q < 1:
                messagebox.showwarning("Input Error", "Number of questions must be at least 1.", parent=self)
                return False
        except tk.TclError:
            messagebox.showwarning("Input Error", "Number of questions must be a valid number.", parent=self)
            return False

        # Validate mark entries (ensure they are numbers if filled)
        for mt_id, mark_var in self.mark_vars.items():
            val_str = mark_var.get().strip()
            if val_str: # Only validate if not empty
                try:
                    float(val_str) # Check if it can be converted to a number
                except ValueError:
                    mt_name = next((mtc["name"] for mtc in self.mark_type_configs if mtc["id"] == mt_id), mt_id)
                    messagebox.showwarning("Input Error", f"Score for '{mt_name}' must be a number if entered.", parent=self)
                    return False
        return True

    def apply(self):
        if not self.validate():
            self.result = None
            return

        quiz_name_final = self.quiz_name_var.get().strip()
        comment_final = self.comment_text_widget.get("1.0", tk.END).strip()
        num_questions_final = self.num_questions_var.get()
        
        marks_data_final = {}
        for mt_id, mark_var in self.mark_vars.items():
            val_str = mark_var.get().strip()
            if val_str: # Only include marks that were actually entered
                try:
                    marks_data_final[mt_id] = float(val_str) # Store as float
                except ValueError:
                    marks_data_final[mt_id] = val_str # Should not happen if validate passed
            # If a mark type is not entered, it's omitted from marks_data_final

        self.result = (quiz_name_final, marks_data_final, comment_final, num_questions_final)
"""




class QuizScoreDialog(simpledialog.Dialog):
    # ... (same as v51)
    def __init__(self, parent, title, initial_quiz_name, mark_types, quiz_templates, default_num_questions, initial_num_questions_val):
        self.initial_quiz_name = initial_quiz_name
        self.mark_types = mark_types # List of dicts: {"id", "name", "default_points", ...}
        self.quiz_templates = quiz_templates # Dict: {template_id: {"name", "num_questions", "default_marks"}}
        self.default_num_questions = default_num_questions
        self.initial_num_questions_val = initial_num_questions_val
        self.result = None
        self.mark_entry_vars = {} # {mark_type_id: StringVar}
        super().__init__(parent, title)

    def body(self, master):
        main_frame = ttk.Frame(master); main_frame.pack(fill=tk.BOTH, expand=True)
        top_frame = ttk.Frame(main_frame); top_frame.grid(pady=5, sticky="ew", columnspan=3)
        ttk.Label(top_frame, text="Quiz Name/Template:").pack(side=tk.LEFT, padx=5)
        self.quiz_name_var = tk.StringVar(value=self.initial_quiz_name)
        
        # Combine quiz name suggestions and template names for the combobox
        template_names = [tpl['name'] for tpl_id, tpl in self.quiz_templates.items()]
        # Simple quiz name suggestions (can be expanded)
        name_suggestions = list(set([self.initial_quiz_name, "Pop Quiz", "Chapter Test"] + template_names))

        self.quiz_name_combobox = ttk.Combobox(top_frame, textvariable=self.quiz_name_var, values=sorted(name_suggestions), width=30)
        self.quiz_name_combobox.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.quiz_name_combobox.bind("<<ComboboxSelected>>", self.on_template_select)
        self.quiz_name_combobox.bind("<MouseWheel>", lambda event: "break")


        num_q_frame = ttk.Frame(main_frame); num_q_frame.grid(pady=5, sticky="ew")
        ttk.Label(num_q_frame, text="Number of Questions:").pack(side=tk.LEFT, padx=5)
        self.num_questions_var = tk.StringVar(value=str(self.initial_num_questions_val if self.initial_num_questions_val else self.default_num_questions))
        self.num_questions_spinbox = ttk.Spinbox(num_q_frame, from_=1, to=200, textvariable=self.num_questions_var, width=5)
        self.num_questions_spinbox.pack(side=tk.LEFT, padx=5)

        self.target_entry2 = []
        self.target_entry3 = []
        self.cte = -1

        marks_frame = ttk.LabelFrame(main_frame, text="Marks Breakdown"); marks_frame.grid(pady=10, column=0, row=2, sticky="nsew", padx=5)
        cols = 2 # Number of columns for mark entries
        current_col_mark = 0; current_row_mark = 0
        for i, mark_type in enumerate(self.mark_types):
            mark_id = mark_type["id"]; mark_name = mark_type["name"]
            ttk.Label(marks_frame, text=f"{mark_name}:").grid(row=current_row_mark, column=current_col_mark*2, sticky=tk.W, padx=5, pady=3)
            var = tk.StringVar()
            self.mark_entry_vars[mark_id] = var
            entry = ttk.Entry(marks_frame, textvariable=var, width=8)
            entry.grid(row=current_row_mark, column=current_col_mark*2 + 1, sticky=tk.EW, padx=5, pady=3)
            entry.bind("<FocusIn>", lambda x=i,x2=i: self.set_numpad(x,x2))
            self.target_entry2.append(var)
            self.target_entry3.append(mark_type["name"])
            self.cte = i
            self.tte = i
            current_col_mark += 1
            if current_col_mark >= cols: current_col_mark = 0; current_row_mark += 1
            
        self.keypad_frame = ttk.Frame(main_frame, relief='sunken',borderwidth=10)
        self.keypad_frame.grid(column=1,row=1, rowspan=4)
        self.keypad_frame.grid_propagate(True)
        self.allow_decimal = True
        self.mark_vars2 = {} # {mark_type_id: tk.StringVar()}
        for mt in self.mark_vars2:
            self.mark_vars2[mt["id"]] = tk.StringVar()
        
        for i, mt_config in enumerate(self.mark_vars2):
            label_text = mt_config["name"]
            if mt_config.get("is_extra_credit"): label_text += " (Bonus)"
        #print(self.mark_vars2)
        self.target_entry = self.mark_entry_vars[mark_type["id"]]
        self.target_entry_name = mark_type["name"]
        
        self.target_name_label = ttk.Label(self.keypad_frame,text=self.target_entry_name)
        self.target_name_label.grid(column=0,row=0,columnspan=9)
        
        buttons = [('7',1,0),('8',1,1),('9',1,2),('4',2,0),('5',2,1),('6',2,2),('1',3,0),('2',3,1),('3',3,2),('0',4,0),('.',4,1) if self.allow_decimal else (' ',4,1),('⌫',4,2),('/',2,3)]
        button_font = ('TkDefaultFont', 15)
        self._next_entry()
        for i in range(4): self.rowconfigure(i, minsize=2000,weight=1)
        for i in range(4): self.columnconfigure(i, weight=1)
        for (text, r, c) in buttons:
            if text == ' ': continue
            action = lambda x=text: self._on_press(x)
            ttk.Button(self.keypad_frame, text=text, command=action, style="Keypad.TButton",padding=17,width=5).grid(row=r, column=c, padx=1, pady=1, sticky="nsew")
        ttk.Button(self.keypad_frame, text="Clear", command=self._clear_entry, style="Keypad.TButton").grid(row=1, column=3, padx=1, pady=1, sticky="nsew")
        ttk.Button(self.keypad_frame, text='Next', command=self._next_entry, style="Keypad.TButton").grid(row=4,column=3,padx=1,pady=1,sticky='nsew')
        ttk.Button(self.keypad_frame, text='Previous', command=self._previous_entry, style="Keypad.TButton").grid(row=3,column=3,padx=1,pady=1,sticky='nsew')
        s = ttk.Style(); s.configure("Keypad.TButton", font=button_font, padding=(5,10),height=100)
        
            
                
            
            
        for i in range(cols*2): marks_frame.grid_columnconfigure(i, weight=1 if i%2==1 else 0)

        ttk.Label(main_frame, text="Comment (Optional):").grid(pady=(0,0), sticky="sw", column=0, row=3, padx=5)
        self.comment_entry = ttk.Entry(main_frame, width=45); self.comment_entry.grid(pady=(0), column=0, row=4, sticky="ews", padx=5)
        
        self.on_template_select(None) # Initialize based on current quiz name (might be a template)
        return self.quiz_name_combobox

    # Functions for NumPad
    def _on_press(self, key):
        current_text = self.target_entry.get()
        if key == '⌫': n = len(current_text)-1; self.target_entry.set(current_text[0:n])
        elif key == '.':
            if self.allow_decimal and '.' not in current_text: self.target_entry.set(self.target_entry.get()+ key)
        else: self.target_entry.set(self.target_entry.get()+ key)

    def _clear_entry(self): self.target_entry.set("")
    
    def _next_entry(self):
        if self.cte < self.tte: self.cte +=1
        else: self.cte = 0
        self.target_entry = self.target_entry2[(self.cte)]
        self.target_entry_name = self.target_entry3[(self.cte)]
        self.target_name_label.configure(text=self.target_entry_name)
        
    def _previous_entry(self):
        if self.cte < self.tte and self.cte > -2: self.cte -=1
        else: self.cte = (self.tte-1)
        self.target_entry = self.target_entry2[(self.cte)]
        self.target_entry_name = self.target_entry3[(self.cte)]
        self.target_name_label.configure(text=self.target_entry_name)
        
    def set_numpad(self, event, x):
        #print(x)
        self.cte = x
        self.target_entry = self.target_entry2[(self.cte)]
        self.target_entry_name = self.target_entry3[(self.cte)]
        self.target_name_label.configure(text=self.target_entry_name)
    
    # Continued code



    def on_template_select(self, event):
        selected_name = self.quiz_name_var.get()
        template = next((tpl for tpl_id, tpl in self.quiz_templates.items() if tpl['name'] == selected_name), None)
        if template:
            self.num_questions_var.set(str(template.get("num_questions", self.default_num_questions)))
            self.num_questions_spinbox.config(state=tk.DISABLED) # Disable if template provides it
            if "default_marks" in template:
                for mark_id, var in self.mark_entry_vars.items():
                    var.set(str(template["default_marks"].get(mark_id, ""))) # Use template's default marks
        else: # Not a template, or template doesn't have these fields
            self.num_questions_spinbox.config(state=tk.NORMAL)
            # Optionally clear marks or set to individual mark_type defaults when not a template
            # for mark_id, var in self.mark_entry_vars.items(): var.set("")

    def apply(self):
        quiz_name = self.quiz_name_var.get().strip(); comment = self.comment_entry.get().strip()
        try: num_questions_actual = int(self.num_questions_var.get())
        except ValueError: messagebox.showwarning("Invalid Input", "Number of questions must be an integer.", parent=self); self.result = None; return
        if num_questions_actual <= 0: messagebox.showwarning("Invalid Input", "Number of questions must be positive.", parent=self); self.result = None; return

        marks_data = {}
        total_marks_entered = 0 # Sum of counts for each mark type
        for mark_id, var in self.mark_entry_vars.items():
            val_str = var.get().strip()
            if val_str:
                try: 
                    val_int = int(val_str)
                    if val_int < 0: messagebox.showwarning("Invalid Input", f"Mark for '{mark_id}' cannot be negative.",parent=self); self.result=None; return
                    marks_data[mark_id] = val_int
                    # Only sum if it's a primary mark contributing to total questions (heuristic)
                    mark_type_obj = next((mt for mt in self.mark_types if mt["id"] == mark_id), None)
                    if mark_type_obj and mark_type_obj.get("contributes_to_total", True):
                        total_marks_entered += val_int
                except ValueError: messagebox.showwarning("Invalid Input", f"Marks for '{mark_id}' must be integers.", parent=self); self.result = None; return
        
        # Validate that sum of primary marks entered does not exceed num_questions_actual
        # This logic assumes 'contributes_to_total' correctly identifies marks that sum up to total questions.
        if total_marks_entered > num_questions_actual:
             messagebox.showwarning("Marks Exceed Questions", f"The sum of primary marks ({total_marks_entered}) exceeds the total number of questions ({num_questions_actual}).\nPlease check your entries.", parent=self)
             self.result = None; return

        if quiz_name: self.result = (quiz_name, marks_data, comment, num_questions_actual)
        else: messagebox.showwarning("Input Required", "Quiz name cannot be empty.", parent=self); self.result = None





class LiveQuizMarkDialog(simpledialog.Dialog):
    # ... (same as v51)
    def __init__(self, parent, student_id, app_instance, session_type="Quiz"): # session_type can be Quiz or Homework
        self.student_id = student_id
        self.app = app_instance
        self.student_name = self.app.students[student_id]['full_name']
        self.session_type_display = session_type
        self.result = None # "correct", "incorrect", "skip"
        super().__init__(parent, f"Mark {session_type} for {self.student_name}")

    def body(self, master):
        if self.session_type_display == "Quiz":
            current_score_info = self.app.live_quiz_scores.get(self.student_id, {"correct": 0, "total_asked": 0})
            score_text = f"Current Score: {current_score_info['correct']} / {current_score_info['total_asked']}"
            ttk.Label(master, text=score_text, font=("", 10)).pack(pady=(5,0))
            ttk.Label(master, text=f"Mark next question for {self.student_name}:").pack(pady=(5,10))
        else: # Homework or other types could be added
            ttk.Label(master, text=f"Update {self.session_type_display} status for {self.student_name}:").pack(pady=(5,10))
        return master # No specific focus needed as it's button driven

    def buttonbox(self): # Override to provide custom buttons
        button_frame = ttk.Frame(self)
        button_frame.pack(pady=10)
        if self.session_type_display == "Quiz":
            ttk.Button(button_frame, text="Correct ✔️", command=lambda: self.set_result_and_close("correct"), width=12).pack(side=tk.LEFT, padx=5)
            ttk.Button(button_frame, text="Incorrect ❌", command=lambda: self.set_result_and_close("incorrect"), width=12).pack(side=tk.LEFT, padx=5)
            ttk.Button(button_frame, text="Skip/Pass ⏭️", command=lambda: self.set_result_and_close("skip"), width=12).pack(side=tk.LEFT, padx=5)
        # Add buttons for other session types if needed
        ttk.Button(button_frame, text="Cancel", command=self.cancel, width=10).pack(side=tk.LEFT, padx=5)
        self.bind("<Escape>", lambda e: self.cancel())
        # Could bind 1,2,3 to correct, incorrect, skip for faster input

    def set_result_and_close(self, res_val):
        self.result = res_val
        self.destroy() # Close the dialog

    def apply(self): pass # Not strictly needed due to custom button actions


class LiveHomeworkMarkDialog(simpledialog.Dialog): # New
    def __init__(self, parent, student_id, app_instance, session_mode, current_hw_data):
        self.student_id = student_id
        self.app = app_instance
        self.student_name = self.app.students[student_id]['full_name']
        self.session_mode = session_mode # "Yes/No" or "Select"
        self.current_hw_data = current_hw_data # Existing data for this student in this session
        self.result_actions = None # Will store dict for Yes/No or list for Select
        
        # For Yes/No mode:
        self.homework_item_vars = {} # {homework_type_id: StringVar}
        # For Select mode:
        self.homework_option_vars = {} # {option_name: BooleanVar}

        super().__init__(parent, f"Mark Homework: {self.student_name} ({self.session_mode} Mode)")

    def body(self, master):
        main_frame = ttk.Frame(master); main_frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
        
        current_session_name_label = ttk.Label(main_frame, text=f"Session: {self.app.current_live_homework_name}", font=("", 11, "italic"))
        current_session_name_label.pack(pady=(0,10))

        if self.session_mode == "Yes/No":
            # Display a list of homework types defined in settings for Yes/No mode
            # self.app.all_homework_session_types (list of dicts: {"id", "name"})'
            select_options = self.app.settings.get("live_homework_select_mode_options", DEFAULT_HOMEWORK_SESSION_BUTTONS.copy())
            if not self.app.all_homework_session_types:
                 ttk.Label(main_frame, text="No homework types configured for 'Yes/No' mode in settings.").pack()
                 return master

            yes_no_frame = ttk.Frame(main_frame)
            yes_no_frame.pack(expand=True, fill=tk.BOTH)
            #print(self.app.all_homework_session_types)
            for hw_type_item in self.app.all_homework_session_types:
                hw_id = hw_type_item["id"]; hw_name = hw_type_item["name"]
                item_frame = ttk.Frame(yes_no_frame); item_frame.pack(fill=tk.X, pady=2)
                ttk.Label(item_frame, text=f"{hw_name}:", width=20, anchor=tk.W).pack(side=tk.LEFT, padx=5)
                
                var = tk.StringVar(value=self.current_hw_data.get(hw_id, "Pending").lower()) # Default to current or "Pending"
                self.homework_item_vars[hw_id] = var
                
                rb_yes = ttk.Radiobutton(item_frame, text="Yes", variable=var, value="yes")
                rb_yes.pack(side=tk.LEFT, padx=3)
                rb_no = ttk.Radiobutton(item_frame, text="No", variable=var, value="no")
                rb_no.pack(side=tk.LEFT, padx=3)
                rb_clear = ttk.Radiobutton(item_frame, text="Pending", variable=var, value="pending") # Clear/Pending option
                rb_clear.pack(side=tk.LEFT, padx=3)


        elif self.session_mode == "Select":
            # Display buttons/checkboxes based on DEFAULT_HOMEWORK_SESSION_BUTTONS or a customizable list from settings
            select_options = self.app.settings.get("live_homework_select_mode_options", DEFAULT_HOMEWORK_SESSION_BUTTONS.copy())
            if not select_options:
                 ttk.Label(main_frame, text="No options configured for 'Select' mode in settings.").pack()
                 return master
            
            select_frame = ttk.Frame(main_frame)
            select_frame.pack(expand=True, fill=tk.BOTH)
            ttk.Label(select_frame, text="Select applicable statuses:").grid(sticky="nw", pady=(0,5))

            # Get current selections for this student
            current_selected_options = self.current_hw_data.get("selected_options", [])

            # Use Checkbuttons for multi-select
            cols = 2; current_col = 0; current_row_sel = 1
            for option_info in select_options: # option_info is a dict, e.g., {"name": "Done"}
                option_name = option_info["name"]
                var = tk.BooleanVar(value=(option_name in current_selected_options))
                self.homework_option_vars[option_name] = var
                cb = ttk.Checkbutton(select_frame, text=option_name, variable=var)
                cb.grid(row=current_row_sel, column=current_col, sticky=tk.W, padx=5, pady=2)
                current_col += 1
                if current_col >= cols: current_col = 0; current_row_sel +=1
            for i in range(cols): select_frame.grid_columnconfigure(i, weight=1)

        return master

    def apply(self):
        if self.session_mode == "Yes/No":
            self.result_actions = {}
            for hw_id, var in self.homework_item_vars.items():
                status = var.get()
                if status != "pending": # Only store "yes" or "no"
                    self.result_actions[hw_id] = status
        elif self.session_mode == "Select":
            self.result_actions = []
            for option_name, var in self.homework_option_vars.items():
                if var.get():
                    self.result_actions.append(option_name)
        # If result_actions is empty (e.g. all pending in Yes/No, or nothing selected in Select)
        # it will be handled by the command to potentially clear the student's entry.

class ExitConfirmationDialog(simpledialog.Dialog): # Same as v50
    def __init__(self, parent, title): self.choice = None; super().__init__(parent, title)
    def body(self, master): ttk.Label(master, text="Do you want to save changes before quitting?").pack(pady=10, padx=10); return None
    def buttonbox(self):
        box = ttk.Frame(self)
        box.columnconfigure(0, weight=1); box.columnconfigure(1, weight=1); box.columnconfigure(2, weight=1)
        ttk.Button(box, text="Save and Quit", width=15, command=self.save_quit).grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        ttk.Button(box, text="Don't Save and Quit", width=20, command=self.no_save_quit).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(box, text="Cancel", width=10, command=self.cancel).grid(row=0, column=2, padx=5, pady=5, sticky="ew")
        self.bind("<Return>", lambda e: self.save_quit()); self.bind("<Escape>", self.cancel)
        box.pack(fill=tk.X, padx=5, pady=5)
    def save_quit(self): self.result = "save_quit"; self.destroy()
    def no_save_quit(self): self.result = "no_save_quit"; self.destroy()



class ImportExcelOptionsDialog(simpledialog.Dialog): # Same as v50, but ensure app_instance is passed
    def __init__(self, parent, app_instance):
        self.app_instance = app_instance
        self.file_path_var = tk.StringVar()
        self.import_incidents_var = tk.BooleanVar(value=False)
        self.student_sheet_var = tk.StringVar()
        self.workbook_sheet_names = []
        super().__init__(parent, "Import Data from Excel")
    def body(self, master):
        ttk.Label(master, text="Excel File:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.file_entry = ttk.Entry(master, textvariable=self.file_path_var, width=40, state="readonly")
        self.file_entry.grid(row=0, column=1, padx=5, pady=2)
        ttk.Button(master, text="Browse...", command=self._browse_file).grid(row=0, column=2, padx=5, pady=2)
        ttk.Label(master, text="Student Info Sheet:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.student_sheet_combo = ttk.Combobox(master, textvariable=self.student_sheet_var, width=37, state="disabled")
        self.student_sheet_combo.grid(row=1, column=1, padx=5, pady=2)
        ttk.Checkbutton(master, text="Import incidents from individual student sheets\n(matches sheet names to current students)", variable=self.import_incidents_var).grid(row=2, column=0, columnspan=3, sticky="w", padx=5, pady=5)
        return self.file_entry
    def _browse_file(self):
        path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        if path:
            self.file_path_var.set(path)
            try:
                wb = load_workbook(filename=path, read_only=True, data_only=True)
                self.workbook_sheet_names = wb.sheetnames
                self.student_sheet_combo['values'] = [""] + self.workbook_sheet_names # Add blank option
                common_sheet_names = ["Students", "Student List", "Roster", "Students Info", "Sheet1"] # Common names
                found_common = next((name for name in common_sheet_names if name in self.workbook_sheet_names), "")
                self.student_sheet_var.set(found_common if found_common else (self.workbook_sheet_names[0] if self.workbook_sheet_names else ""))
                self.student_sheet_combo.config(state="readonly")
            except Exception as e:
                messagebox.showerror("Excel Error", f"Could not read Excel file: {e}", parent=self.app_instance.root)
                self.student_sheet_combo['values'] = [""]; self.student_sheet_var.set(""); self.student_sheet_combo.config(state="disabled")
    def validate(self):
        if not self.file_path_var.get() or not os.path.exists(self.file_path_var.get()):
            messagebox.showerror("File Error", "Please select a valid Excel file.", parent=self.app_instance.root); return False
        # No need to check workbook_sheet_names here as _browse_file handles errors.
        # Student sheet can be blank if user doesn't want to import students.
        return True
    def apply(self):
        if not self.validate(): self.result = None; return
        self.result = (self.file_path_var.get(), self.import_incidents_var.get(), self.student_sheet_var.get() or None) # Return None if sheet is blank


class SizeInputDialog(simpledialog.Dialog): # Same as v50
    def __init__(self, parent, title, initial_w, initial_h, status):
        self.status = status
        self.initial_w, self.initial_h = initial_w, initial_h
        self.width_var = tk.IntVar(value=initial_w); self.height_var = tk.IntVar(value=initial_h)
        super().__init__(parent, title)
    def body(self, master):
        ttk.Label(master, text="Width:").grid(row=0, column=0, sticky="W", padx=5, pady=5)
        self.width_entry = ttk.Spinbox(master, from_=MIN_STUDENT_BOX_WIDTH, to=1000, textvariable=self.width_var, width=7)
        self.width_entry.grid(row=0, column=1, padx=5, pady=5)
        self.width_reset_entry = ttk.Button(master,command=self.reset_width ,text="Reset Width")
        self.width_reset_entry.grid(row=0,column=2,padx=5,pady=5)
        
        ttk.Label(master, text="Height:").grid(row=1, column=0, sticky="W", padx=5, pady=5)
        self.height_entry = ttk.Spinbox(master, from_=MIN_STUDENT_BOX_HEIGHT, to=1000, textvariable=self.height_var, width=7)
        self.height_entry.grid(row=1, column=1, padx=5, pady=5)
        self.height_reset_entry = ttk.Button(master,command=self.reset_height, text="Reset Height")
        self.height_reset_entry.grid(row=1,column=2,padx=5,pady=5)
        return self.width_entry
    
    def reset_width(self):
        if self.status:
            self.width_var.set(DEFAULT_STUDENT_BOX_WIDTH)
        else:
            self.width_var.set(REBBI_DESK_WIDTH)
        
    def reset_height(self):
        #self.height_var.set(DEFAULT_STUDENT_BOX_HEIGHT)
        if self.status:
            self.height_var.set(DEFAULT_STUDENT_BOX_HEIGHT)
        else:
            self.height_var.set(REBBI_DESK_HEIGHT)
    
    def validate(self):
        try:
            w, h = self.width_var.get(), self.height_var.get()
            if w < MIN_STUDENT_BOX_WIDTH or h < MIN_STUDENT_BOX_HEIGHT or w > 1000 or h > 1000:
                messagebox.showerror("Invalid Size", f"Width ({MIN_STUDENT_BOX_WIDTH}-1000), Height ({MIN_STUDENT_BOX_HEIGHT}-1000).", parent=self); return False
            return True
        except tk.TclError: messagebox.showerror("Invalid Input", "Valid numbers for W/H.", parent=self); return False
    def apply(self):
        if not self.validate(): self.result = None; return
        self.result = (self.width_var.get(), self.height_var.get())




class StudentStyleDialog(simpledialog.Dialog):
    # ... (same as v51)
    def __init__(self, parent, title, student_data, app):
        self.student_data = student_data
        self.app = app # For default settings and font list
        self.result = [] # List of (property, old_value, new_value) tuples for Command
        self.initial_overrides = student_data.get("style_overrides", {}).copy()
        super().__init__(parent, title)

    def body(self, master):
        prop_frame = ttk.Frame(master); prop_frame.pack(padx=10,pady=10)
        row_idx = 0
        # Fill Color
        ttk.Label(prop_frame, text="Box Fill Color:").grid(row=row_idx, column=0, sticky=tk.W, pady=3)
        self.fill_color_var = tk.StringVar(value=self.initial_overrides.get("fill_color", ""))
        self.fill_color_entry = ttk.Entry(prop_frame, textvariable=self.fill_color_var, width=15)
        self.fill_color_entry.grid(row=row_idx, column=1, pady=3, padx=2)
        ttk.Button(prop_frame, text="Choose...", command=lambda v=self.fill_color_var: self.choose_color(v)).grid(row=row_idx, column=2, pady=3, padx=2)
        ttk.Button(prop_frame, text="Default", command=lambda v=self.fill_color_var, k="fill_color": self.reset_to_default(v,k)).grid(row=row_idx, column=3, pady=3, padx=2)
        row_idx+=1
        # Outline Color
        ttk.Label(prop_frame, text="Box Outline Color:").grid(row=row_idx, column=0, sticky=tk.W, pady=3)
        self.outline_color_var = tk.StringVar(value=self.initial_overrides.get("outline_color", ""))
        self.outline_color_entry = ttk.Entry(prop_frame, textvariable=self.outline_color_var, width=15)
        self.outline_color_entry.grid(row=row_idx, column=1, pady=3, padx=2)
        ttk.Button(prop_frame, text="Choose...", command=lambda v=self.outline_color_var: self.choose_color(v)).grid(row=row_idx, column=2, pady=3, padx=2)
        ttk.Button(prop_frame, text="Default", command=lambda v=self.outline_color_var, k="outline_color": self.reset_to_default(v,k)).grid(row=row_idx, column=3, pady=3, padx=2)
        row_idx+=1
        # Font Family
        ttk.Label(prop_frame, text="Font Family:").grid(row=row_idx, column=0, sticky=tk.W, pady=3)
        self.font_family_var = tk.StringVar(value=self.initial_overrides.get("font_family", ""))
        available_fonts = self.app.settings.get("available_fonts", [DEFAULT_FONT_FAMILY])
        self.font_family_combo = ttk.Combobox(prop_frame, textvariable=self.font_family_var, values=available_fonts, width=20, state="readonly")
        self.font_family_combo.grid(row=row_idx, column=1, columnspan=2, pady=3, padx=2, sticky=tk.EW)
        self.font_family_combo.bind("<MouseWheel>", lambda event: "break")
        ttk.Button(prop_frame, text="Default", command=lambda v=self.font_family_var, k="font_family": self.reset_to_default(v,k)).grid(row=row_idx, column=3, pady=3, padx=2)
        row_idx+=1
        # Font Size
        ttk.Label(prop_frame, text="Font Size (pts):").grid(row=row_idx, column=0, sticky=tk.W, pady=3)
        self.font_size_var = tk.IntVar(value=self.initial_overrides.get("font_size", 0) or 0) # Ensure it's int, 0 for default placeholder
        self.font_size_spinbox = ttk.Spinbox(prop_frame, from_=6, to=30, textvariable=self.font_size_var, width=5)
        self.font_size_spinbox.grid(row=row_idx, column=1, pady=3, padx=2, sticky=tk.W)
        ttk.Button(prop_frame, text="Default", command=lambda v=self.font_size_var, k="font_size": self.reset_to_default(v,k)).grid(row=row_idx, column=3, pady=3, padx=2)
        row_idx+=1
        # Font Color
        ttk.Label(prop_frame, text="Font Color:").grid(row=row_idx, column=0, sticky=tk.W, pady=3)
        self.font_color_var = tk.StringVar(value=self.initial_overrides.get("font_color", ""))
        self.font_color_entry = ttk.Entry(prop_frame, textvariable=self.font_color_var, width=15)
        self.font_color_entry.grid(row=row_idx, column=1, pady=3, padx=2)
        ttk.Button(prop_frame, text="Choose...", command=lambda v=self.font_color_var: self.choose_color(v)).grid(row=row_idx, column=2, pady=3, padx=2)
        ttk.Button(prop_frame, text="Default", command=lambda v=self.font_color_var, k="font_color": self.reset_to_default(v,k)).grid(row=row_idx, column=3, pady=3, padx=2)
        return self.fill_color_entry # Initial focus
    def choose_color(self, var_to_set):
        initial_color = var_to_set.get() if var_to_set.get() else None
        color_code = colorchooser.askcolor(title="Choose color", initialcolor=initial_color, parent=self)
        if color_code and color_code[1]: var_to_set.set(color_code[1])
    def reset_to_default(self, var_to_set, key):
        # For string vars (colors, font family), set to empty string to signify "use app default"
        # For int var (font size), set to 0 to signify "use app default"
        if isinstance(var_to_set, tk.StringVar): var_to_set.set("")
        elif isinstance(var_to_set, tk.IntVar): var_to_set.set(0)
    def apply(self):
        style_props_vars = {
            "fill_color": self.fill_color_var, "outline_color": self.outline_color_var,
            "font_family": self.font_family_var, "font_size": self.font_size_var,
            "font_color": self.font_color_var
        }
        for prop_key, tk_var in style_props_vars.items():
            new_val = tk_var.get()
            # Standardize "default" representation: None for deletion from overrides
            # Empty string for colors/font family, 0 for font size, means use default
            if isinstance(new_val, str) and not new_val.strip(): final_new_val = None
            elif isinstance(new_val, int) and new_val == 0: final_new_val = None
            else: final_new_val = new_val
            
            old_val_from_overrides = self.initial_overrides.get(prop_key) # This will be None if key didn't exist
            if final_new_val != old_val_from_overrides:
                self.result.append((prop_key, old_val_from_overrides, final_new_val))



class ExportFilterDialog(simpledialog.Dialog):
    # ... (updated for homework filters)
    def __init__(self, parent, students_dict, all_behaviors_list, all_homework_types_list, default_settings):
        self.students_dict = students_dict
        self.all_behaviors_list = sorted(list(set(all_behaviors_list)))
        self.all_homework_types_list = ((all_homework_types_list)) # New
        self.default_settings = default_settings
        self.result = None
        super().__init__(parent, "Export Log Options")

    def body(self, master):
        frame = ttk.Frame(master); frame.pack(padx=10, pady=10)
        # Date Range
        date_frame = ttk.LabelFrame(frame, text="Date Range"); date_frame.grid(pady=5,column=0,row=0,columnspan=3, sticky="ew")
        ttk.Label(date_frame, text="Start Date:").grid(row=0, column=0, padx=5, pady=3, sticky=tk.W)
        self.start_date_var = tk.StringVar()
        if DateEntry: self.start_date_entry = DateEntry(date_frame, textvariable=self.start_date_var, date_pattern='yyyy-mm-dd', width=12);
        else: self.start_date_entry = ttk.Entry(date_frame, textvariable=self.start_date_var, width=12)
        self.start_date_entry.grid(row=0, column=1, padx=5, pady=3)
        ttk.Label(date_frame, text="End Date:").grid(row=1, column=0, padx=5, pady=3, sticky=tk.W)
        self.end_date_var = tk.StringVar()
        if DateEntry: self.end_date_entry = DateEntry(date_frame, textvariable=self.end_date_var, date_pattern='yyyy-mm-dd', width=12)
        else: self.end_date_entry = ttk.Entry(date_frame, textvariable=self.end_date_var, width=12)
        self.end_date_entry.grid(row=1, column=1, padx=5, pady=3)
        ttk.Button(date_frame, text="Clear Dates", command=self.clear_dates).grid(row=0, column=2, rowspan=2, padx=5, pady=3)

        # Students
        student_frame = ttk.LabelFrame(frame, text="Students"); student_frame.grid(pady=5,column=0,row=1, sticky="nsew")
        self.student_filter_var = tk.StringVar(value="all")
        ttk.Radiobutton(student_frame, text="All Students", variable=self.student_filter_var, value="all", command=self.toggle_student_list_state).pack(anchor=tk.W)
        ttk.Radiobutton(student_frame, text="Selected Students:", variable=self.student_filter_var, value="specific", command=self.toggle_student_list_state).pack(anchor=tk.W)
        self.student_list_frame = ttk.Frame(student_frame)
        self.student_list_frame.pack(fill=tk.X, padx=(20,0))
        self.student_listbox = tk.Listbox(self.student_list_frame, selectmode=tk.MULTIPLE, height=5, exportselection=False)
        sorted_students = sorted(self.students_dict.values(), key=lambda s: (s['last_name'], s['first_name']))
        self.student_listbox_map = {} # display_name -> student_id
        for i, s_data in enumerate(sorted_students):
            display_name = f"{s_data['last_name']}, {s_data['first_name']}" + (f" ({s_data.get('nickname')})" if s_data.get('nickname') else "")
            self.student_listbox.insert(tk.END, display_name)
            self.student_listbox_map[display_name] = s_data["id"]
        self.student_listbox.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=(0,3))
        student_scroll = ttk.Scrollbar(self.student_list_frame, orient=tk.VERTICAL, command=self.student_listbox.yview)
        student_scroll.pack(side=tk.RIGHT, fill=tk.Y, pady=(0,3)); self.student_listbox.config(yscrollcommand=student_scroll.set)

        # Behaviors (for Behavior and Quiz logs)
        behavior_frame = ttk.LabelFrame(frame, text="Behavior/Quiz Types"); behavior_frame.grid(pady=5, padx=5, column=1,row=1, sticky="nsew")
        self.behavior_filter_var = tk.StringVar(value="all")
        ttk.Radiobutton(behavior_frame, text="All Behavior/Quiz Types", variable=self.behavior_filter_var, value="all", command=self.toggle_behavior_list_state).pack(anchor=tk.W)
        ttk.Radiobutton(behavior_frame, text="Selected Behavior/Quiz Types:", variable=self.behavior_filter_var, value="specific", command=self.toggle_behavior_list_state).pack(anchor=tk.W)
        self.behavior_list_frame = ttk.Frame(behavior_frame)
        self.behavior_list_frame.pack(fill=tk.X, padx=(20,0))
        self.behavior_listbox = tk.Listbox(self.behavior_list_frame, selectmode=tk.MULTIPLE, height=5, exportselection=False)
        for b_name in self.all_behaviors_list: self.behavior_listbox.insert(tk.END, b_name)
        self.behavior_listbox.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=(0,3))
        bh_scroll = ttk.Scrollbar(self.behavior_list_frame, orient=tk.VERTICAL, command=self.behavior_listbox.yview)
        bh_scroll.pack(side=tk.RIGHT, fill=tk.Y, pady=(0,3)); self.behavior_listbox.config(yscrollcommand=bh_scroll.set)

        # Homework Types (New)
        homework_frame = ttk.LabelFrame(frame, text="Homework Types"); homework_frame.grid(pady=5,padx=5,column=0,row=2,rowspan=2, sticky="ew")
        self.homework_filter_var = tk.StringVar(value="all")
        ttk.Radiobutton(homework_frame, text="All Homework Types", variable=self.homework_filter_var, value="all", command=self.toggle_homework_list_state).pack(anchor=tk.W)
        ttk.Radiobutton(homework_frame, text="Selected Homework Types:", variable=self.homework_filter_var, value="specific", command=self.toggle_homework_list_state).pack(anchor=tk.W)
        self.homework_list_frame = ttk.Frame(homework_frame)
        self.homework_list_frame.pack(fill=tk.X, padx=(20,0))
        self.homework_listbox = tk.Listbox(self.homework_list_frame, selectmode=tk.MULTIPLE, height=5, exportselection=False)
        for hw_name in self.all_homework_types_list: self.homework_listbox.insert(tk.END, hw_name) # Use combined list of names
        self.homework_listbox.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=(0,3))
        hw_scroll = ttk.Scrollbar(self.homework_list_frame, orient=tk.VERTICAL, command=self.homework_listbox.yview)
        hw_scroll.pack(side=tk.RIGHT, fill=tk.Y, pady=(0,3)); self.homework_listbox.config(yscrollcommand=hw_scroll.set)


        # Log Type Inclusion
        include_frame = ttk.LabelFrame(frame, text="Include Log Types"); include_frame.grid(pady=5,column=1,row=2, sticky="nsew")
        self.include_behavior_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(include_frame, text="Behavior Logs", variable=self.include_behavior_var).pack(anchor=tk.W, padx=5)
        self.include_quiz_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(include_frame, text="Quiz Logs", variable=self.include_quiz_var).pack(anchor=tk.W, padx=5)
        self.include_homework_var = tk.BooleanVar(value=True) # New
        ttk.Checkbutton(include_frame, text="Homework Logs", variable=self.include_homework_var).pack(anchor=tk.W, padx=5)

        # Output Options
        output_options_frame = ttk.LabelFrame(frame, text="Excel Output Options"); output_options_frame.grid(pady=5, column=1,row=3, sticky="nsew")
        self.separate_sheets_var = tk.BooleanVar(value=self.default_settings.get("excel_export_separate_sheets_by_default", True))
        ttk.Checkbutton(output_options_frame, text="Separate sheets for Behavior, Quiz, Homework", variable=self.separate_sheets_var).pack(anchor=tk.W, padx=5)
        self.include_summaries_var = tk.BooleanVar(value=self.default_settings.get("excel_export_include_summaries_by_default", True))
        ttk.Checkbutton(output_options_frame, text="Include summary sheet", variable=self.include_summaries_var).pack(anchor=tk.W, padx=5)

        self.toggle_student_list_state(); self.toggle_behavior_list_state(); self.toggle_homework_list_state()
        return frame

    def clear_dates(self): self.start_date_var.set(""); self.end_date_var.set("")
    def toggle_student_list_state(self): self.student_listbox.config(state=tk.NORMAL if self.student_filter_var.get() == "specific" else tk.DISABLED)
    def toggle_behavior_list_state(self): self.behavior_listbox.config(state=tk.NORMAL if self.behavior_filter_var.get() == "specific" else tk.DISABLED)
    def toggle_homework_list_state(self): self.homework_listbox.config(state=tk.NORMAL if self.homework_filter_var.get() == "specific" else tk.DISABLED)

    def apply(self):
        start_dt, end_dt = None, None
        try:
            if self.start_date_var.get(): start_dt = datetime.strptime(self.start_date_var.get(), '%Y-%m-%d').date()
            if self.end_date_var.get(): end_dt = datetime.strptime(self.end_date_var.get(), '%Y-%m-%d').date()
            if start_dt and end_dt and start_dt > end_dt:
                messagebox.showerror("Invalid Dates", "Start date cannot be after end date.", parent=self); return
        except ValueError: messagebox.showerror("Invalid Date Format", "Please use YYYY-MM-DD for dates.", parent=self); return

        selected_s_ids = [self.student_listbox_map[self.student_listbox.get(i)] for i in self.student_listbox.curselection()] if self.student_filter_var.get() == "specific" else []
        selected_b_names = [self.behavior_listbox.get(i) for i in self.behavior_listbox.curselection()] if self.behavior_filter_var.get() == "specific" else []
        selected_hw_names = [self.homework_listbox.get(i) for i in self.homework_listbox.curselection()] if self.homework_filter_var.get() == "specific" else [] # New

        if not (self.include_behavior_var.get() or self.include_quiz_var.get() or self.include_homework_var.get()):
            messagebox.showwarning("No Log Types", "Please select at least one log type to include.", parent=self); return

        self.result = {
            "start_date": start_dt, "end_date": end_dt,
            "selected_students": self.student_filter_var.get(), "student_ids": selected_s_ids,
            "selected_behaviors": self.behavior_filter_var.get(), "behaviors_list": selected_b_names,
            "selected_homework_types": self.homework_filter_var.get(), "homework_types_list": selected_hw_names, # New
            "include_behavior_logs": self.include_behavior_var.get(),
            "include_quiz_logs": self.include_quiz_var.get(),
            "include_homework_logs": self.include_homework_var.get(), # New
            "separate_sheets_by_log_type": self.separate_sheets_var.get(),
            "include_summaries": self.include_summaries_var.get()
        }



class AttendanceReportDialog(simpledialog.Dialog):
    def __init__(self, parent, students_dict):
        self.students_dict = students_dict
        self.result = None
        super().__init__(parent, "Generate Attendance Report")

    def body(self, master):
        frame = ttk.Frame(master); frame.pack(padx=10, pady=10)

        date_frame = ttk.LabelFrame(frame, text="Report Date Range"); date_frame.pack(pady=5, fill=tk.X)
        ttk.Label(date_frame, text="Start Date:").grid(row=0, column=0, padx=5, pady=3, sticky=tk.W)
        self.start_date_var = tk.StringVar(value=(datetime_date.today() - timedelta(days=7)).strftime('%Y-%m-%d'))
        if DateEntry:
            self.start_date_entry = DateEntry(date_frame, textvariable=self.start_date_var, date_pattern='yyyy-mm-dd', width=12)
        else:
            self.start_date_entry = ttk.Entry(date_frame, textvariable=self.start_date_var, width=12)
        self.start_date_entry.grid(row=0, column=1, padx=5, pady=3)

        ttk.Label(date_frame, text="End Date:").grid(row=1, column=0, padx=5, pady=3, sticky=tk.W)
        self.end_date_var = tk.StringVar(value=datetime_date.today().strftime('%Y-%m-%d'))
        if DateEntry:
            self.end_date_entry = DateEntry(date_frame, textvariable=self.end_date_var, date_pattern='yyyy-mm-dd', width=12)
        else:
            self.end_date_entry = ttk.Entry(date_frame, textvariable=self.end_date_var, width=12)
        self.end_date_entry.grid(row=1, column=1, padx=5, pady=3)

        student_frame = ttk.LabelFrame(frame, text="Select Students"); student_frame.pack(pady=5, fill=tk.X)
        self.student_listbox = tk.Listbox(student_frame, selectmode=tk.MULTIPLE, height=8, exportselection=False)
        self.sorted_student_list_for_dialog = sorted(self.students_dict.values(), key=lambda s: (s['last_name'], s['first_name']))
        self.student_id_map_for_dialog = {} # display_name -> student_id
        for i, s_data in enumerate(self.sorted_student_list_for_dialog):
            display_name = f"{s_data['last_name']}, {s_data['first_name']}" + (f" ({s_data.get('nickname')})" if s_data.get('nickname') else "")
            self.student_listbox.insert(tk.END, display_name)
            self.student_id_map_for_dialog[display_name] = s_data["id"]
            self.student_listbox.selection_set(i) # Select all by default

        self.student_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=3, padx=3)
        student_scroll = ttk.Scrollbar(student_frame, orient=tk.VERTICAL, command=self.student_listbox.yview)
        student_scroll.pack(side=tk.RIGHT, fill=tk.Y, pady=3, padx=3)
        self.student_listbox.config(yscrollcommand=student_scroll.set)
        
        select_buttons_frame = ttk.Frame(student_frame)
        select_buttons_frame.pack(fill=tk.X, pady=(0,3))
        ttk.Button(select_buttons_frame, text="Select All", command=lambda: self.student_listbox.selection_set(0, tk.END)).pack(side=tk.LEFT, padx=5)
        ttk.Button(select_buttons_frame, text="Deselect All", command=lambda: self.student_listbox.selection_clear(0, tk.END)).pack(side=tk.LEFT, padx=5)


        return frame

    def apply(self):
        start_dt, end_dt = None, None
        try:
            if self.start_date_var.get(): start_dt = datetime.strptime(self.start_date_var.get(), '%Y-%m-%d').date()
            if self.end_date_var.get(): end_dt = datetime.strptime(self.end_date_var.get(), '%Y-%m-%d').date()
            if not start_dt or not end_dt:
                messagebox.showerror("Missing Dates", "Start and End dates are required.", parent=self); return
            if start_dt > end_dt:
                messagebox.showerror("Invalid Dates", "Start date cannot be after end date.", parent=self); return
        except ValueError:
            messagebox.showerror("Invalid Date Format", "Please use YYYY-MM-DD for dates.", parent=self); return

        selected_student_ids = [self.student_id_map_for_dialog[self.student_listbox.get(i)] for i in self.student_listbox.curselection()]
        if not selected_student_ids:
            messagebox.showwarning("No Students Selected", "Please select at least one student for the report.", parent=self); return

        self.result = (start_dt, end_dt, selected_student_ids)

class ConditionalFormattingRuleDialog(simpledialog.Dialog):
    def __init__(self, parent, app, rule_to_edit=None):
        self.app = app
        self.rule = rule_to_edit or {} # Existing rule or new empty dict
        self.result = None
        title = "Edit Conditional Formatting Rule" if rule_to_edit else "Add Conditional Formatting Rule"
        super().__init__(parent, title)

    def body(self, master):
        frame = ttk.Frame(master); frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        # Rule Type
        ttk.Label(frame, text="Rule applies to:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=3)
        self.rule_type_var = tk.StringVar(value=self.rule.get("type", "group"))
        type_options = ["group", "behavior_count", "quiz_score_threshold"] # Add more types later
        self.type_combo = ttk.Combobox(frame, textvariable=self.rule_type_var, values=type_options, state="readonly", width=25)
        self.type_combo.grid(row=0, column=1, columnspan=2, padx=5, pady=3, sticky=tk.EW)
        self.type_combo.bind("<<ComboboxSelected>>", self.on_rule_type_change)

        # Condition Frame (changes based on rule_type)
        self.condition_frame = ttk.Frame(frame)
        self.condition_frame.grid(row=1, column=0, columnspan=3, pady=5, sticky=tk.NSEW)

        # Formatting Options
        format_frame = ttk.LabelFrame(frame, text="Formatting Actions");
        format_frame.grid(row=2, column=0, columnspan=3, pady=10, sticky=tk.EW)

        ttk.Label(format_frame, text="Set Box Fill Color:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=3)
        self.fill_color_var = tk.StringVar(value=self.rule.get("color", ""))
        self.fill_color_entry = ttk.Entry(format_frame, textvariable=self.fill_color_var, width=12)
        self.fill_color_entry.grid(row=0, column=1, padx=2, pady=3)
        ttk.Button(format_frame, text="Choose...", command=lambda: self.choose_color_for_var(self.fill_color_var)).grid(row=0, column=2, padx=2, pady=3)
        ttk.Button(format_frame, text="Clear", command=lambda: self.fill_color_var.set("")).grid(row=0, column=3, padx=2, pady=3)


        ttk.Label(format_frame, text="Set Box Outline Color:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=3)
        self.outline_color_var = tk.StringVar(value=self.rule.get("outline", ""))
        self.outline_color_entry = ttk.Entry(format_frame, textvariable=self.outline_color_var, width=12)
        self.outline_color_entry.grid(row=1, column=1, padx=2, pady=3)
        ttk.Button(format_frame, text="Choose...", command=lambda: self.choose_color_for_var(self.outline_color_var)).grid(row=1, column=2, padx=2, pady=3)
        ttk.Button(format_frame, text="Clear", command=lambda: self.outline_color_var.set("")).grid(row=1, column=3, padx=2, pady=3)

        # Initialize condition frame based on current/default rule type
        self.on_rule_type_change(None)
        return self.type_combo

    def choose_color_for_var(self, color_var):
        initial = color_var.get() if color_var.get() else None
        color_code = colorchooser.askcolor(title="Choose color", initialcolor=initial, parent=self)
        if color_code and color_code[1]: color_var.set(color_code[1])

    def on_rule_type_change(self, event):
        for widget in self.condition_frame.winfo_children(): widget.destroy() # Clear previous condition widgets
        rule_type = self.rule_type_var.get()

        if rule_type == "group":
            ttk.Label(self.condition_frame, text="Select Group:").pack(side=tk.LEFT, padx=5)
            self.group_var = tk.StringVar(value=self.rule.get("group_id", ""))
            group_names = {gid: gdata["name"] for gid, gdata in self.app.student_groups.items()}
            self.group_id_map_cond = {name: gid for gid, name in group_names.items()} # name to id
            self.group_combo_cond = ttk.Combobox(self.condition_frame, textvariable=self.group_var,
                                            values=[""] + sorted(group_names.values()), state="readonly", width=20)
            if self.rule.get("group_id") and self.rule["group_id"] in self.app.student_groups:
                 self.group_var.set(self.app.student_groups[self.rule["group_id"]]["name"])
            self.group_combo_cond.pack(side=tk.LEFT, padx=5)

        elif rule_type == "behavior_count":
            ttk.Label(self.condition_frame, text="Behavior:").pack(side=tk.LEFT, padx=5)
            self.behavior_name_var = tk.StringVar(value=self.rule.get("behavior_name", ""))
            self.behavior_combo_cond = ttk.Combobox(self.condition_frame, textvariable=self.behavior_name_var,
                                               values=[""] + self.app.all_behaviors, width=18)
            self.behavior_combo_cond.pack(side=tk.LEFT, padx=2)

            ttk.Label(self.condition_frame, text="Count >=:").pack(side=tk.LEFT, padx=5)
            self.behavior_count_var = tk.IntVar(value=self.rule.get("count_threshold", 1))
            ttk.Spinbox(self.condition_frame, from_=1, to=100, textvariable=self.behavior_count_var, width=4).pack(side=tk.LEFT, padx=2)

            ttk.Label(self.condition_frame, text="In Last (Hours):").pack(side=tk.LEFT, padx=5)
            self.behavior_hours_var = tk.IntVar(value=self.rule.get("time_window_hours", 24))
            ttk.Spinbox(self.condition_frame, from_=1, to=720, textvariable=self.behavior_hours_var, width=4).pack(side=tk.LEFT, padx=2)

        elif rule_type == "quiz_score_threshold":
            ttk.Label(self.condition_frame, text="Quiz Name (contains):").pack(side=tk.LEFT, padx=5)
            self.quiz_name_contains_var = tk.StringVar(value=self.rule.get("quiz_name_contains", ""))
            ttk.Entry(self.condition_frame, textvariable=self.quiz_name_contains_var, width=15).pack(side=tk.LEFT, padx=2)

            ttk.Label(self.condition_frame, text="Score (%):").pack(side=tk.LEFT, padx=5)
            self.quiz_op_var = tk.StringVar(value=self.rule.get("operator", "<="))
            ttk.Combobox(self.condition_frame, textvariable=self.quiz_op_var, values=["<=", ">=", "==", "<", ">"], width=3, state="readonly").pack(side=tk.LEFT, padx=2)
            self.quiz_score_thresh_var = tk.DoubleVar(value=self.rule.get("score_threshold_percent", 50.0))
            ttk.Spinbox(self.condition_frame, from_=0, to=100, increment=1, textvariable=self.quiz_score_thresh_var, width=5).pack(side=tk.LEFT, padx=2)


    def apply(self):
        final_rule = {"type": self.rule_type_var.get()}
        rule_type = final_rule["type"]

        fill = self.fill_color_var.get().strip()
        outline = self.outline_color_var.get().strip()
        if not fill and not outline:
            messagebox.showerror("No Action", "Please specify at least one formatting action (fill or outline color).", parent=self)
            return
        if fill: final_rule["color"] = fill
        if outline: final_rule["outline"] = outline

        if rule_type == "group":
            selected_group_name = self.group_var.get()
            if not selected_group_name: messagebox.showerror("Missing Info", "Please select a group.", parent=self); return
            final_rule["group_id"] = self.group_id_map_cond.get(selected_group_name)
            if not final_rule["group_id"]: messagebox.showerror("Error", "Selected group not found.", parent=self); return

        elif rule_type == "behavior_count":
            b_name = self.behavior_name_var.get().strip()
            if not b_name: messagebox.showerror("Missing Info", "Please select a behavior.", parent=self); return
            final_rule["behavior_name"] = b_name
            final_rule["count_threshold"] = self.behavior_count_var.get()
            final_rule["time_window_hours"] = self.behavior_hours_var.get()

        elif rule_type == "quiz_score_threshold":
            final_rule["quiz_name_contains"] = self.quiz_name_contains_var.get().strip() # Can be empty for any quiz
            final_rule["operator"] = self.quiz_op_var.get()
            final_rule["score_threshold_percent"] = self.quiz_score_thresh_var.get()
        
        self.result = final_rule


class ManageStudentGroupsDialog(simpledialog.Dialog):
    def __init__(self, parent, student_groups_data, students_data, app, default_colors):
        self.student_groups = student_groups_data # This is a reference to app.student_groups, modified directly
        self.students = students_data # App.students, for assigning
        self.app = app # To get new group ID, and redraw
        self.default_colors = default_colors
        self.groups_changed_flag = False # Set to True if any persistent change is made
        super().__init__(parent, "Manage Student Groups")

    def body(self, master):
        self.master_frame = master
        top_frame = ttk.Frame(master); top_frame.pack(pady=5, padx=5, fill=tk.X)
        ttk.Button(top_frame, text="Add New Group", command=self.add_group).pack(side=tk.LEFT, padx=5)

        self.canvas_groups = tk.Canvas(master, borderwidth=0, background="#ffffff")
        self.groups_scrollable_frame = ttk.Frame(self.canvas_groups)
        self.scrollbar_groups = ttk.Scrollbar(master, orient="vertical", command=self.canvas_groups.yview)
        self.canvas_groups.configure(yscrollcommand=self.scrollbar_groups.set)

        self.scrollbar_groups.pack(side="right", fill="y")
        self.canvas_groups.pack(side="left", fill="both", expand=True)
        self.canvas_groups_window = self.canvas_groups.create_window((0,0), window=self.groups_scrollable_frame, anchor="nw")

        self.groups_scrollable_frame.bind("<Configure>", lambda e: self.canvas_groups.configure(scrollregion=self.canvas_groups.bbox("all")))
        self.canvas_groups.bind('<MouseWheel>', self._on_mousewheel_groups)

        self.populate_groups_list()
        return self.groups_scrollable_frame # For initial focus target (though dynamic)

    def _on_mousewheel_groups(self, event):
        if event.delta: self.canvas_groups.yview_scroll(int(-1*(event.delta/120)), "units")
        else: self.canvas_groups.yview_scroll(1 if event.num == 5 else -1, "units")


    def populate_groups_list(self):
        for widget in self.groups_scrollable_frame.winfo_children(): widget.destroy()
        row_idx = 0
        if not self.student_groups:
            ttk.Label(self.groups_scrollable_frame, text="No groups created yet.").pack(pady=10)
            return

        sorted_groups = sorted(self.student_groups.items(), key=lambda item: item[1]['name'])

        for group_id, group_data in sorted_groups:
            group_frame = ttk.Frame(self.groups_scrollable_frame, padding=5, relief=tk.RIDGE, borderwidth=1)
            group_frame.pack(fill=tk.X, pady=3, padx=3)

            name_var = tk.StringVar(value=group_data["name"])
            color_var = tk.StringVar(value=group_data.get("color", self.default_colors[0]))

            ttk.Label(group_frame, text="Name:").grid(row=0, column=0, sticky=tk.W)
            name_entry = ttk.Entry(group_frame, textvariable=name_var, width=20)
            name_entry.grid(row=0, column=1, padx=3)
            name_entry.bind("<FocusOut>", lambda e, gid=group_id, nv=name_var: self.update_group_name(gid, nv.get()))

            ttk.Label(group_frame, text="Color:").grid(row=0, column=2, sticky=tk.W, padx=(10,0))
            color_entry = ttk.Entry(group_frame, textvariable=color_var, width=10)
            color_entry.grid(row=0, column=3, padx=3)
            color_btn = ttk.Button(group_frame, text="...", width=3, command=lambda cv=color_var, gid=group_id: self.choose_group_color(cv, gid))
            color_btn.grid(row=0, column=4)
            color_preview = tk.Label(group_frame, text="  ", bg=color_var.get(), width=2, relief=tk.SUNKEN)
            color_preview.grid(row=0, column=5, padx=2)
            color_var.trace_add("write", lambda *args, cp=color_preview, cv=color_var: cp.config(bg=cv.get()))


            ttk.Button(group_frame, text="Assign Students...", command=lambda gid=group_id, gname=group_data["name"]: self.assign_students_to_group_dialog(gid, gname)).grid(row=0, column=6, padx=5)
            ttk.Button(group_frame, text="Delete Group", command=lambda gid=group_id: self.delete_group(gid)).grid(row=0, column=7, padx=5)

            # Show number of students in group
            count = sum(1 for sid, sdata in self.students.items() if sdata.get("group_id") == group_id)
            ttk.Label(group_frame, text=f"({count} student{'s' if count !=1 else ''})").grid(row=0, column=8, padx=5, sticky=tk.E)
            group_frame.grid_columnconfigure(8, weight=1) # Push delete button to right a bit


    def add_group(self):
        new_group_name = simpledialog.askstring("New Group", "Enter name for the new group:", parent=self)
        if new_group_name and new_group_name.strip():
            group_id_str, next_id_val = self.app.get_new_group_id() # Get ID from app
            # Check for name collision BEFORE updating app's next_group_id_num state
            if any(g['name'].lower() == new_group_name.strip().lower() for g in self.student_groups.values()):
                 messagebox.showwarning("Duplicate Name", f"A group named '{new_group_name.strip()}' already exists.", parent=self)
                 return # Do not consume the ID from app if name is duplicate

            self.app.next_group_id_num = next_id_val # Commit ID usage
            new_color_index = (self.app.next_group_id_num -1) % len(self.default_colors) # Cycle through default colors
            self.student_groups[group_id_str] = {"name": new_group_name.strip(), "color": self.default_colors[new_color_index]}
            self.groups_changed_flag = True
            self.populate_groups_list()
        elif new_group_name is not None: # User entered empty string
             messagebox.showwarning("Invalid Name", "Group name cannot be empty.", parent=self)


    def update_group_name(self, group_id, new_name):
        new_name = new_name.strip()
        if not new_name:
            messagebox.showwarning("Invalid Name", "Group name cannot be empty. Reverting.", parent=self)
            self.populate_groups_list(); return # Revert by repopulating
        
        # Check for name collision with OTHER groups
        if any(g_id != group_id and g_data['name'].lower() == new_name.lower() for g_id, g_data in self.student_groups.items()):
            messagebox.showwarning("Duplicate Name", f"Another group named '{new_name}' already exists. Reverting.", parent=self)
            self.populate_groups_list(); return

        if self.student_groups[group_id]["name"] != new_name:
            self.student_groups[group_id]["name"] = new_name
            self.groups_changed_flag = True
            # No need to repopulate for just name change if var is linked, but good for consistency if other things change
            self.populate_groups_list() # Repopulate to reflect changes and resort


    def choose_group_color(self, color_var, group_id):
        initial_color = color_var.get()
        new_color = colorchooser.askcolor(initial_color, title=f"Choose color for group", parent=self)
        if new_color and new_color[1]:
            if self.student_groups[group_id].get("color") != new_color[1]:
                color_var.set(new_color[1]) # This will trigger trace and update preview
                self.student_groups[group_id]["color"] = new_color[1]
                self.groups_changed_flag = True
                # No repopulate needed just for color if var is traced.

    def delete_group(self, group_id):
        group_name = self.student_groups[group_id]["name"]
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete group '{group_name}'?\nStudents in this group will be unassigned.", parent=self):
            # Unassign students from this group
            for student_id, student_data in self.students.items():
                if student_data.get("group_id") == group_id:
                    student_data["group_id"] = None # Or del student_data["group_id"]
            del self.student_groups[group_id]
            self.groups_changed_flag = True
            self.populate_groups_list()

    def assign_students_to_group_dialog(self, group_id, group_name):
        dialog = AssignStudentsToGroupSubDialog(self, group_id, group_name, self.students, self.student_groups)
        if dialog.assignments_changed:
            self.groups_changed_flag = True # Signal main app that changes occurred
            self.populate_groups_list() # Repopulate to update student counts

    def apply(self): # Called when OK is pressed
        # Changes are applied directly, so just set the flag if it was ever true
        self.result = self.groups_changed_flag
        if self.groups_changed_flag:
            self.app.draw_all_items(check_collisions_on_redraw=True) # Redraw if groups changed
            self.app.save_student_groups() # Save groups if changed


class AssignStudentsToGroupSubDialog(simpledialog.Dialog):
    def __init__(self, parent_dialog, group_id, group_name, all_students_data, all_groups_data):
        self.parent_dialog_ref = parent_dialog # Reference to ManageStudentGroupsDialog
        self.group_id_to_assign = group_id
        self.group_name = group_name
        self.all_students = all_students_data # app.students
        self.all_groups = all_groups_data # app.student_groups
        self.assignments_changed = False
        super().__init__(parent_dialog, f"Assign Students to '{group_name}'")

    def body(self, master):
        # Listbox for available students (not in any group OR in a different group)
        # Listbox for students currently in THIS group
        frame = ttk.Frame(master); frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        
        available_frame = ttk.LabelFrame(frame, text="Available Students"); available_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        self.available_lb = tk.Listbox(available_frame, selectmode=tk.EXTENDED, exportselection=False, height=15)
        self.available_lb.pack(fill=tk.BOTH, expand=True, pady=2)
        self.available_students_map = {} # display name -> student_id

        buttons_frame = ttk.Frame(frame); buttons_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5)
        ttk.Button(buttons_frame, text=">> Add >>", command=self.add_to_group).pack(pady=10)
        ttk.Button(buttons_frame, text="<< Remove <<", command=self.remove_from_group).pack(pady=10)

        assigned_frame = ttk.LabelFrame(frame, text=f"Students in '{self.group_name}'"); assigned_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        self.assigned_lb = tk.Listbox(assigned_frame, selectmode=tk.EXTENDED, exportselection=False, height=15)
        self.assigned_lb.pack(fill=tk.BOTH, expand=True, pady=2)
        self.assigned_students_map = {} # display name -> student_id

        self.populate_student_lists()
        return self.available_lb

    def populate_student_lists(self):
        self.available_lb.delete(0, tk.END); self.assigned_lb.delete(0, tk.END)
        self.available_students_map.clear(); self.assigned_students_map.clear()

        sorted_all_students = sorted(self.all_students.values(), key=lambda s: (s['last_name'], s['first_name']))

        for s_data in sorted_all_students:
            s_id = s_data["id"]
            display_name = f"{s_data['last_name']}, {s_data['first_name']}" + (f" ({s_data.get('nickname')})" if s_data.get('nickname') else "")
            
            if s_data.get("group_id") == self.group_id_to_assign:
                self.assigned_lb.insert(tk.END, display_name)
                self.assigned_students_map[display_name] = s_id
            elif not s_data.get("group_id") or s_data.get("group_id") == "NONE_GROUP_SENTINEL": # Truly unassigned
                self.available_lb.insert(tk.END, display_name)
                self.available_students_map[display_name] = s_id
            # Else: student is in a *different* group, don't show in "Available" unless we want a "move from group X" feature

    def add_to_group(self):
        selected_indices = self.available_lb.curselection()
        if not selected_indices: return
        for i in reversed(selected_indices): # Iterate reversed to handle index changes on delete
            display_name = self.available_lb.get(i)
            student_id = self.available_students_map.get(display_name)
            if student_id and student_id in self.all_students:
                self.all_students[student_id]["group_id"] = self.group_id_to_assign
                self.assignments_changed = True
        self.populate_student_lists() # Refresh both listboxes

    def remove_from_group(self):
        selected_indices = self.assigned_lb.curselection()
        if not selected_indices: return
        for i in reversed(selected_indices):
            display_name = self.assigned_lb.get(i)
            student_id = self.assigned_students_map.get(display_name)
            if student_id and student_id in self.all_students:
                self.all_students[student_id]["group_id"] = None # Unassign
                self.assignments_changed = True
        self.populate_student_lists()

    def apply(self):
        # Changes were made directly to self.all_students (app.students)
        # The self.assignments_changed flag will be checked by the parent dialog
        pass

# --- Quiz Template Management ---
class ManageQuizTemplatesDialog(simpledialog.Dialog):
    def __init__(self, parent, app_instance):
        self.app = app_instance # Main application instance
        self.templates_changed_flag = False
        super().__init__(parent, "Manage Quiz Templates")

    def body(self, master):
        self.master_frame = master
        top_frame = ttk.Frame(master); top_frame.pack(pady=5, padx=5, fill=tk.X)
        ttk.Button(top_frame, text="Add New Quiz Template", command=self.add_template).pack(side=tk.LEFT, padx=5)

        self.canvas_templates = tk.Canvas(master, borderwidth=0, background="#ffffff")
        self.templates_scrollable_frame = ttk.Frame(self.canvas_templates)
        self.scrollbar_templates = ttk.Scrollbar(master, orient="vertical", command=self.canvas_templates.yview)
        self.canvas_templates.configure(yscrollcommand=self.scrollbar_templates.set)

        self.scrollbar_templates.pack(side="right", fill="y")
        self.canvas_templates.pack(side="left", fill="both", expand=True)
        self.canvas_templates.create_window((0,0), window=self.templates_scrollable_frame, anchor="nw", tags="self.templates_scrollable_frame")

        self.templates_scrollable_frame.bind("<Configure>", lambda e: self.canvas_templates.configure(scrollregion=self.canvas_templates.bbox("all")))
        self.canvas_templates.bind('<MouseWheel>', self._on_mousewheel_templates)
        
        self.populate_templates_list()
        return self.templates_scrollable_frame

    def _on_mousewheel_templates(self, event):
        if event.delta: self.canvas_templates.yview_scroll(int(-1*(event.delta/120)), "units")
        else: self.canvas_templates.yview_scroll(1 if event.num == 5 else -1, "units")

    def populate_templates_list(self):
        for widget in self.templates_scrollable_frame.winfo_children(): widget.destroy()
        if not self.app.quiz_templates:
            ttk.Label(self.templates_scrollable_frame, text="No quiz templates created yet.").pack(pady=10)
            return

        sorted_templates = sorted(self.app.quiz_templates.items(), key=lambda item: item[1]['name'])
        for tpl_id, tpl_data in sorted_templates:
            tpl_frame = ttk.Frame(self.templates_scrollable_frame, padding=5, relief=tk.RIDGE, borderwidth=1)
            tpl_frame.pack(fill=tk.X, pady=3, padx=3)
            
            summary = f"{tpl_data['name']} ({tpl_data.get('num_questions', 'N/A')} Qs)"
            ttk.Label(tpl_frame, text=summary, width=40, anchor=tk.W).pack(side=tk.LEFT, padx=5)
            ttk.Button(tpl_frame, text="Edit", command=lambda tid=tpl_id: self.edit_template(tid)).pack(side=tk.LEFT, padx=3)
            ttk.Button(tpl_frame, text="Delete", command=lambda tid=tpl_id: self.delete_template(tid)).pack(side=tk.LEFT, padx=3)

    def add_template(self):
        dialog = QuizTemplateEditDialog(self, self.app, template_data=None) # Pass self as parent
        if dialog.result_template_data:
            template_id_str, next_id_val = self.app.get_new_quiz_template_id()
            
            # Check for name collision BEFORE committing ID
            if any(t['name'].lower() == dialog.result_template_data['name'].lower() for t in self.app.quiz_templates.values()):
                messagebox.showwarning("Duplicate Name", f"A quiz template named '{dialog.result_template_data['name']}' already exists.", parent=self)
                return
            
            self.app.next_quiz_template_id_num = next_id_val # Commit ID usage
            self.app.quiz_templates[template_id_str] = dialog.result_template_data
            self.templates_changed_flag = True
            self.populate_templates_list()

    def edit_template(self, template_id):
        if template_id not in self.app.quiz_templates: return
        current_template_data = self.app.quiz_templates[template_id]
        dialog = QuizTemplateEditDialog(self, self.app, template_data=current_template_data.copy(), existing_template_id=template_id)
        if dialog.result_template_data:
            # Check for name collision with OTHER templates
            if any(tid != template_id and t_data['name'].lower() == dialog.result_template_data['name'].lower() for tid, t_data in self.app.quiz_templates.items()):
                 messagebox.showwarning("Duplicate Name", f"Another quiz template named '{dialog.result_template_data['name']}' already exists. Edit cancelled.", parent=self)
                 return

            self.app.quiz_templates[template_id] = dialog.result_template_data
            self.templates_changed_flag = True
            self.populate_templates_list()

    def delete_template(self, template_id):
        if template_id not in self.app.quiz_templates: return
        tpl_name = self.app.quiz_templates[template_id]["name"]
        if messagebox.askyesno("Confirm Delete", f"Delete quiz template '{tpl_name}'?", parent=self):
            del self.app.quiz_templates[template_id]
            self.templates_changed_flag = True
            self.populate_templates_list()

    def apply(self): # OK button
        self.result = self.templates_changed_flag


class QuizTemplateEditDialog(simpledialog.Dialog):
    def __init__(self, parent_dialog, app_instance, template_data=None, existing_template_id=None):
        self.app = app_instance
        self.template_data_initial = template_data or {} # If editing, this is a copy
        self.existing_template_id = existing_template_id # Used to avoid name collision with self
        self.result_template_data = None # Populated on successful apply
        self.mark_entry_vars_tpl = {} # {mark_type_id: StringVar}
        title = "Edit Quiz Template" if template_data else "Add New Quiz Template"
        super().__init__(parent_dialog, title)

    def body(self, master):
        main_frame = ttk.Frame(master); main_frame.pack(padx=10, pady=10, expand=True, fill=tk.BOTH)
        
        name_frame = ttk.Frame(main_frame); name_frame.pack(fill=tk.X, pady=3)
        ttk.Label(name_frame, text="Template Name:").pack(side=tk.LEFT, padx=5)
        self.tpl_name_var = tk.StringVar(value=self.template_data_initial.get("name", ""))
        self.tpl_name_entry = ttk.Entry(name_frame, textvariable=self.tpl_name_var, width=35)
        self.tpl_name_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        num_q_frame = ttk.Frame(main_frame); num_q_frame.pack(fill=tk.X, pady=3)
        ttk.Label(num_q_frame, text="Number of Questions:").pack(side=tk.LEFT, padx=5)
        self.tpl_num_q_var = tk.StringVar(value=str(self.template_data_initial.get("num_questions", self.app.settings.get("default_quiz_questions",10))))
        self.tpl_num_q_spinbox = ttk.Spinbox(num_q_frame, from_=1, to=200, textvariable=self.tpl_num_q_var, width=5)
        self.tpl_num_q_spinbox.pack(side=tk.LEFT, padx=5)

        marks_frame = ttk.LabelFrame(main_frame, text="Default Marks (per question, if applicable)"); marks_frame.pack(pady=5, fill=tk.BOTH, expand=True)
        cols = 2
        current_col, current_row = 0,0
        default_marks_initial = self.template_data_initial.get("default_marks", {})
        for mt in self.app.settings.get("quiz_mark_types", []):
            mt_id = mt["id"]; mt_name = mt["name"]
            ttk.Label(marks_frame, text=f"{mt_name}:").grid(row=current_row, column=current_col*2, sticky=tk.W, padx=5, pady=2)
            var = tk.StringVar(value=str(default_marks_initial.get(mt_id, ""))) # "" if not set, or default_points
            self.mark_entry_vars_tpl[mt_id] = var
            entry = ttk.Entry(marks_frame, textvariable=var, width=8)
            entry.grid(row=current_row, column=current_col*2 + 1, sticky=tk.EW, padx=5, pady=2)
            current_col += 1
            if current_col >= cols: current_col = 0; current_row += 1
        for i in range(cols*2): marks_frame.grid_columnconfigure(i, weight=1 if i%2==1 else 0)
        
        return self.tpl_name_entry # Initial focus

    def apply(self):
        name = self.tpl_name_var.get().strip()
        if not name: messagebox.showerror("Input Required", "Template name cannot be empty.", parent=self); return
        try: num_q = int(self.tpl_num_q_var.get())
        except ValueError: messagebox.showerror("Invalid Input", "Number of questions must be an integer.", parent=self); return
        if num_q <=0: messagebox.showerror("Invalid Input", "Number of questions must be positive.", parent=self); return
        
        default_marks = {}
        for mt_id, var in self.mark_entry_vars_tpl.items():
            val_str = var.get().strip()
            if val_str: # Only store if set
                try: default_marks[mt_id] = int(val_str) # Store as int
                except ValueError: messagebox.showerror("Invalid Mark", f"Default mark for '{mt_id}' must be an integer or empty.", parent=self); return
        
        self.result_template_data = {"name": name, "num_questions": num_q, "default_marks": default_marks}

# --- Homework Template Management (New) ---
class ManageHomeworkTemplatesDialog(simpledialog.Dialog):
    def __init__(self, parent, app_instance):
        self.app = app_instance
        self.templates_changed_flag = False
        super().__init__(parent, "Manage Homework Templates")

    def body(self, master):
        self.master_frame = master
        top_frame = ttk.Frame(master); top_frame.pack(pady=5, padx=5, fill=tk.X)
        ttk.Button(top_frame, text="Add New Homework Template", command=self.add_template).pack(side=tk.LEFT, padx=5)

        self.canvas_templates = tk.Canvas(master, borderwidth=0, background="#ffffff")
        self.templates_scrollable_frame = ttk.Frame(self.canvas_templates)
        self.scrollbar_templates = ttk.Scrollbar(master, orient="vertical", command=self.canvas_templates.yview)
        self.canvas_templates.configure(yscrollcommand=self.scrollbar_templates.set)
        self.scrollbar_templates.pack(side="right", fill="y")
        self.canvas_templates.pack(side="left", fill="both", expand=True)
        self.canvas_templates.create_window((0,0), window=self.templates_scrollable_frame, anchor="nw", tags="self.templates_scrollable_frame")
        self.templates_scrollable_frame.bind("<Configure>", lambda e: self.canvas_templates.configure(scrollregion=self.canvas_templates.bbox("all")))
        self.canvas_templates.bind('<MouseWheel>', self._on_mousewheel_templates)
        self.populate_templates_list()
        return self.templates_scrollable_frame

    def _on_mousewheel_templates(self, event):
        if event.delta: self.canvas_templates.yview_scroll(int(-1*(event.delta/120)), "units")
        else: self.canvas_templates.yview_scroll(1 if event.num == 5 else -1, "units")

    def populate_templates_list(self):
        for widget in self.templates_scrollable_frame.winfo_children(): widget.destroy()
        if not self.app.homework_templates:
            ttk.Label(self.templates_scrollable_frame, text="No homework templates created yet.").pack(pady=10)
            return
        sorted_templates = sorted(self.app.homework_templates.items(), key=lambda item: item[1]['name'])
        for tpl_id, tpl_data in sorted_templates:
            tpl_frame = ttk.Frame(self.templates_scrollable_frame, padding=5, relief=tk.RIDGE, borderwidth=1)
            tpl_frame.pack(fill=tk.X, pady=3, padx=3)
            summary = f"{tpl_data['name']} ({tpl_data.get('num_items', 'N/A')} items)"
            ttk.Label(tpl_frame, text=summary, width=40, anchor=tk.W).pack(side=tk.LEFT, padx=5)
            ttk.Button(tpl_frame, text="Edit", command=lambda tid=tpl_id: self.edit_template(tid)).pack(side=tk.LEFT, padx=3)
            ttk.Button(tpl_frame, text="Delete", command=lambda tid=tpl_id: self.delete_template(tid)).pack(side=tk.LEFT, padx=3)

    def add_template(self):
        dialog = HomeworkTemplateEditDialog(self, self.app, template_data=None)
        if dialog.result_template_data:
            template_id_str, next_id_val = self.app.get_new_homework_template_id()
            if any(t['name'].lower() == dialog.result_template_data['name'].lower() for t in self.app.homework_templates.values()):
                messagebox.showwarning("Duplicate Name", f"A homework template named '{dialog.result_template_data['name']}' already exists.", parent=self)
                return
            self.app.next_homework_template_id_num = next_id_val
            self.app.homework_templates[template_id_str] = dialog.result_template_data
            self.templates_changed_flag = True
            self.populate_templates_list()

    def edit_template(self, template_id):
        if template_id not in self.app.homework_templates: return
        current_template_data = self.app.homework_templates[template_id]
        dialog = HomeworkTemplateEditDialog(self, self.app, template_data=current_template_data.copy(), existing_template_id=template_id)
        if dialog.result_template_data:
            if any(tid != template_id and t_data['name'].lower() == dialog.result_template_data['name'].lower() for tid, t_data in self.app.homework_templates.items()):
                 messagebox.showwarning("Duplicate Name", f"Another homework template named '{dialog.result_template_data['name']}' already exists. Edit cancelled.", parent=self)
                 return
            self.app.homework_templates[template_id] = dialog.result_template_data
            self.templates_changed_flag = True
            self.populate_templates_list()

    def delete_template(self, template_id):
        if template_id not in self.app.homework_templates: return
        tpl_name = self.app.homework_templates[template_id]["name"]
        if messagebox.askyesno("Confirm Delete", f"Delete homework template '{tpl_name}'?", parent=self):
            del self.app.homework_templates[template_id]
            self.templates_changed_flag = True
            self.populate_templates_list()

    def apply(self):
        self.result = self.templates_changed_flag


class HomeworkTemplateEditDialog(simpledialog.Dialog):
    def __init__(self, parent_dialog, app_instance, template_data=None, existing_template_id=None):
        self.app = app_instance
        self.template_data_initial = template_data or {}
        self.existing_template_id = existing_template_id
        self.result_template_data = None
        self.mark_entry_vars_hw_tpl = {} # {homework_mark_type_id: StringVar}
        title = "Edit Homework Template" if template_data else "Add New Homework Template"
        super().__init__(parent_dialog, title)

    def body(self, master):
        main_frame = ttk.Frame(master); main_frame.pack(padx=10, pady=10, expand=True, fill=tk.BOTH)
        name_frame = ttk.Frame(main_frame); name_frame.pack(fill=tk.X, pady=3)
        ttk.Label(name_frame, text="Template Name:").pack(side=tk.LEFT, padx=5)
        self.tpl_name_var = tk.StringVar(value=self.template_data_initial.get("name", ""))
        self.tpl_name_entry = ttk.Entry(name_frame, textvariable=self.tpl_name_var, width=35)
        self.tpl_name_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        num_items_frame = ttk.Frame(main_frame); num_items_frame.pack(fill=tk.X, pady=3)
        ttk.Label(num_items_frame, text="Number of Items/Tasks:").pack(side=tk.LEFT, padx=5)
        self.tpl_num_items_var = tk.StringVar(value=str(self.template_data_initial.get("num_items", 10))) # Default if not set
        self.tpl_num_items_spinbox = ttk.Spinbox(num_items_frame, from_=0, to=200, textvariable=self.tpl_num_items_var, width=5) # 0 if not applicable
        self.tpl_num_items_spinbox.pack(side=tk.LEFT, padx=5)

        marks_frame = ttk.LabelFrame(main_frame, text="Default Marks/Statuses"); marks_frame.pack(pady=5, fill=tk.BOTH, expand=True)
        cols = 2
        current_col, current_row = 0,0
        default_marks_initial = self.template_data_initial.get("default_marks", {})
        for hmt in self.app.settings.get("homework_mark_types", []):
            hmt_id = hmt["id"]; hmt_name = hmt["name"]
            ttk.Label(marks_frame, text=f"{hmt_name}:").grid(row=current_row, column=current_col*2, sticky=tk.W, padx=5, pady=2)
            var = tk.StringVar(value=str(default_marks_initial.get(hmt_id, ""))) # "" or default_points
            self.mark_entry_vars_hw_tpl[hmt_id] = var
            entry = ttk.Entry(marks_frame, textvariable=var, width=10) # Wider for potential text status
            entry.grid(row=current_row, column=current_col*2 + 1, sticky=tk.EW, padx=5, pady=2)
            current_col += 1
            if current_col >= cols: current_col = 0; current_row += 1
        for i in range(cols*2): marks_frame.grid_columnconfigure(i, weight=1 if i%2==1 else 0)
        return self.tpl_name_entry

    def apply(self):
        name = self.tpl_name_var.get().strip()
        if not name: messagebox.showerror("Input Required", "Template name cannot be empty.", parent=self); return
        try: num_items = int(self.tpl_num_items_var.get())
        except ValueError: messagebox.showerror("Invalid Input", "Number of items must be an integer (or 0 if not applicable).", parent=self); return
        if num_items < 0: messagebox.showerror("Invalid Input", "Number of items cannot be negative.", parent=self); return
        
        default_marks = {}
        for hmt_id, var in self.mark_entry_vars_hw_tpl.items():
            val_str = var.get().strip()
            if val_str: # Only store if set
                # Try to convert to float/int if numeric, else store as string (for statuses like "Done")
                try: default_marks[hmt_id] = float(val_str)
                except ValueError: default_marks[hmt_id] = val_str # Store as string if not purely numeric
        
        self.result_template_data = {"name": name, "num_items": num_items, "default_marks": default_marks}



class SettingsDialog(simpledialog.Dialog):
    def __init__(self, parent, current_settings, custom_behaviors, all_behaviors, app,
                 custom_homework_log_behaviors, all_homework_log_behaviors,
                 custom_homework_session_types, all_homework_session_types,
                 password_manager_instance, theme, custom_canvas_color):
        self.settings = current_settings # This is a reference to app.settings, modified directly by some controls
        self.custom_behaviors_ref = custom_behaviors # Reference to app.custom_behaviors
        self.all_behaviors_ref = all_behaviors # Reference, updated by app later
        self.custom_homework_log_behaviors_ref = custom_homework_log_behaviors # New
        self.all_homework_log_behaviors_ref = all_homework_log_behaviors # New
        self.custom_homework_session_types_ref = custom_homework_session_types # New for Live Yes/No mode
        self.all_homework_session_types_ref = all_homework_session_types # New
        self.app = app # For callbacks like saving, getting new IDs
        self.password_manager = password_manager_instance
        self.theme = tk.StringVar(value=theme)
        self.theme2 = self.theme.get()
        self.custom_canvas_color = tk.StringVar(value= custom_canvas_color if custom_canvas_color != None else "Default")
        self.settings_changed_flag = False # True if any setting that requires save/redraw is changed
        self.initial_settings_snapshot = {k: (v.copy() if isinstance(v, (dict, list)) else v) for k,v in current_settings.items()}
        super().__init__(parent, f"Settings - {APP_NAME}")

    def body(self, master):
        self.master_frame = master # Keep a reference for rebuilding tabs if needed
        self.notebook = ttk.Notebook(master)
        
        # --- General Tab ---
        general_tab = ttk.Frame(self.notebook, padding=10); self.notebook.add(general_tab, text="General")
        self.create_general_tab(general_tab)

        # --- Student Display Tab ---
        student_display_tab = ttk.Frame(self.notebook, padding=10); self.notebook.add(student_display_tab, text="Student Boxes")
        self.create_student_display_tab(student_display_tab)

        # --- Behavior/Quiz Log Tab ---
        behavior_log_tab = ttk.Frame(self.notebook, padding=10); self.notebook.add(behavior_log_tab, text="Behavior & Quiz")
        self.create_behavior_log_tab(behavior_log_tab)

        # --- Homework Log Tab (New) ---
        homework_log_tab = ttk.Frame(self.notebook, padding=10); self.notebook.add(homework_log_tab, text="Homework")
        self.create_homework_log_tab(homework_log_tab)

        # --- Data & Export Tab ---
        data_export_tab = ttk.Frame(self.notebook, padding=10); self.notebook.add(data_export_tab, text="Data & Export")
        self.create_data_export_tab(data_export_tab)

        # --- Security Tab ---
        security_tab = ttk.Frame(self.notebook, padding=10); self.notebook.add(security_tab, text="Security")
        self.create_security_tab(security_tab)

        # --- Advanced/Hidden Tab (Optional) ---
        # self.create_advanced_tab(advanced_tab)

        self.notebook.grid(column=0,row=0,columnspan=2)
        self.notebook.grid_propagate(True)
        # No specific focus needed, first field in first tab will get it.
        return self.notebook

    def create_general_tab(self, tab_frame):
        lf = ttk.LabelFrame(tab_frame, text="Application Behavior", padding=10); lf.pack(fill=tk.X, pady=5)
        # Autosave interval
        ttk.Label(lf, text="Autosave Interval (seconds):").grid(row=0, column=0, sticky=tk.W, padx=5, pady=3)
        self.autosave_interval_var = tk.IntVar(value=self.settings.get("autosave_interval_ms", 30000) // 1000)
        ttk.Spinbox(lf, from_=10, to=300, increment=10, textvariable=self.autosave_interval_var, width=5).grid(row=0, column=1, sticky=tk.W, padx=5, pady=3)

        # Grid snap
        self.grid_snap_var = tk.BooleanVar(value=self.settings.get("grid_snap_enabled", False))
        ttk.Checkbutton(lf, text="Enable Snap to Grid during Drag", variable=self.grid_snap_var).grid(row=1, column=0, columnspan=2, sticky=tk.W, padx=5, pady=3)
        ttk.Label(lf, text="Grid Size (pixels):").grid(row=2, column=0, sticky=tk.W, padx=5, pady=3)
        self.grid_size_var = tk.IntVar(value=self.settings.get("grid_size", DEFAULT_GRID_SIZE))
        ttk.Spinbox(lf, from_=5, to=100, increment=5, textvariable=self.grid_size_var, width=5).grid(row=2, column=1, sticky=tk.W, padx=5, pady=3)

        # Student Groups Enabled
        self.groups_enabled_var = tk.BooleanVar(value=self.settings.get("student_groups_enabled", True))
        ttk.Checkbutton(lf, text="Enable Student Groups Feature", variable=self.groups_enabled_var).grid(row=3, column=0, columnspan=2, sticky=tk.W, padx=5, pady=3)
        
        # Zoom Level Display
        self.show_zoom_var = tk.BooleanVar(value=self.settings.get("show_zoom_level_display", True))
        ttk.Checkbutton(lf, text="Show Zoom Level % Display on Main Screen", variable=self.show_zoom_var).grid(row=4, column=0, columnspan=2, sticky=tk.W, padx=5, pady=3)

        # Max Undo History Days
        ttk.Label(lf, text="Max Undo History (days):").grid(row=5, column=0, sticky=tk.W, padx=5, pady=3)
        self.max_undo_days_var = tk.IntVar(value=self.settings.get("max_undo_history_days", MAX_UNDO_HISTORY_DAYS))
        ttk.Spinbox(lf, from_=1, to=90, textvariable=self.max_undo_days_var, width=5).grid(row=5, column=1, sticky=tk.W, padx=5, pady=3)
        
        # Theme
        ttk.Label(lf, text = "Theme: ").grid(row=6,column=0,sticky='W', padx=5, pady=3)
        
        theme_combo = ttk.Combobox(lf, values = THEME_LIST, textvariable= self.theme, state='readonly')
        theme_combo.grid(row=6, column=1, sticky="W", padx=5, pady=3)
        theme_combo.bind("<<ComboboxSelected>>", self.theme_set)
        theme_combo.set(self.theme.get())
        
        # Canvas Color
        ttk.Label(lf, text = "Canvas color (background): ").grid(row=7,column=0,sticky='W', padx=5, pady=3)
        
        canvas_color_entry = ttk.Entry(lf, textvariable= self.custom_canvas_color)
        canvas_color_entry.grid(row=7, column=1, sticky="W", padx=5, pady=3)
        
        
        #ttk.Label(parent_frame, text="Default Fill Color:").grid(row=start_row,column=0,sticky=tk.W,padx=5,pady=3)
        #fill_var = tk.StringVar(value=self.settings.get(fill_key, DEFAULT_BOX_FILL_COLOR))
        #setattr(self, f"{fill_key}_var", fill_var) # Store var on self
        #ttk.Entry(, textvariable=fill_var, width=12).grid(row=start_row,column=1,sticky=tk.W,padx=5,pady=3)
        
        if self.custom_canvas_color.get() == "":
            #print("Hi")
            #self.custom_canvas_color
            pass
        else:
            #print("Hi2", self.custom_canvas_color.get())
            #print(self.custom_canvas_color)
            self.custom2 = tk.StringVar(value=self.custom_canvas_color.get())
        ttk.Button(lf, text="Choose...", command=lambda v=self.custom_canvas_color: self.choose_color_for_canvas(v)).grid(row=7,column=2,sticky=tk.W,padx=2,pady=3)
        ttk.Button(lf, text="Default", command=lambda v=self.custom_canvas_color: self.reset_canvas_color(v)).grid(row=7,column=3, sticky='W', padx=5, pady=3)
        #theme_combo.bind("<<ComboboxSelected>>", self.theme_set)
        #theme_combo.set(self.theme.get())
        #print(self.theme.get(), "initlial")

    def create_student_display_tab(self, tab_frame):
        lf_defaults = ttk.LabelFrame(tab_frame, text="Default Student Box Appearance", padding=10)
        lf_defaults.grid(sticky="nsew", column=0,row=0, pady=5)
        # Default size
        ttk.Label(lf_defaults, text="Default Width:").grid(row=0,column=0,sticky=tk.W,padx=5,pady=3)
        self.def_stud_w_var = tk.IntVar(value=self.settings.get("default_student_box_width", DEFAULT_STUDENT_BOX_WIDTH))
        ttk.Spinbox(lf_defaults, from_=MIN_STUDENT_BOX_WIDTH, to=500, textvariable=self.def_stud_w_var, width=5).grid(row=0,column=1,sticky=tk.W,padx=5,pady=3)
        ttk.Label(lf_defaults, text="Default Height:").grid(row=1,column=0,sticky=tk.W,padx=5,pady=3)
        self.def_stud_h_var = tk.IntVar(value=self.settings.get("default_student_box_height", DEFAULT_STUDENT_BOX_HEIGHT))
        ttk.Spinbox(lf_defaults, from_=MIN_STUDENT_BOX_HEIGHT, to=300, textvariable=self.def_stud_h_var, width=5).grid(row=1,column=1,sticky=tk.W,padx=5,pady=3)
        # Default colors and font
        self.create_color_font_settings_ui(lf_defaults, 2, "student_box_fill_color", "student_box_outline_color", "student_font_family", "student_font_size", "student_font_color")

        lf_cond_format = ttk.LabelFrame(tab_frame, text="Conditional Formatting Rules", padding=10, width=1000)
        lf_cond_format.grid(sticky="nse", pady=5, padx=5, column=1, columnspan=3, row=0)
        lf_cond_format.grid_anchor("e")
        ttk.Button(lf_cond_format, text="Add Rule...", command=self.add_conditional_rule).pack(pady=3, anchor=tk.W)
        self.rules_listbox = tk.Listbox(lf_cond_format, height=7, exportselection=False, width=75)
        self.rules_listbox.pack(fill=tk.X, expand=True, pady=2)
        self.populate_conditional_rules_listbox()
        rule_btns_frame = ttk.Frame(lf_cond_format); rule_btns_frame.pack(fill=tk.X)
        ttk.Button(rule_btns_frame, text="Edit Selected", command=self.edit_selected_conditional_rule).pack(side=tk.LEFT, padx=2)
        ttk.Button(rule_btns_frame, text="Remove Selected", command=self.remove_selected_conditional_rule).pack(side=tk.LEFT, padx=2)


    def create_behavior_log_tab(self, tab_frame):
        # Recent Incidents Display
        lf_recent = ttk.LabelFrame(tab_frame, text="Recent Incidents on Student Boxes (Behavior/Quiz)", padding=10); lf_recent.grid(sticky="nsew",column=0,row=0, pady=5)
        self.show_recent_var = tk.BooleanVar(value=self.settings.get("show_recent_incidents_on_boxes", True))
        ttk.Checkbutton(lf_recent, text="Show recent incidents on student boxes", variable=self.show_recent_var).grid(row=0,column=0,columnspan=2,sticky=tk.W, padx=5,pady=3)
        ttk.Label(lf_recent, text="Number to show:").grid(row=1,column=0,sticky=tk.W,padx=5,pady=3)
        self.num_recent_var = tk.IntVar(value=self.settings.get("num_recent_incidents_to_show", 2))
        ttk.Spinbox(lf_recent, from_=0, to=10, textvariable=self.num_recent_var, width=3).grid(row=1,column=1,sticky=tk.W,padx=5,pady=3)
        ttk.Label(lf_recent, text="Time window (hours):").grid(row=2,column=0,sticky=tk.W,padx=5,pady=3)
        self.time_window_var = tk.IntVar(value=self.settings.get("recent_incident_time_window_hours", 24))
        ttk.Spinbox(lf_recent, from_=1, to=168, textvariable=self.time_window_var, width=4).grid(row=2,column=1,sticky=tk.W,padx=5,pady=3)
        self.show_full_recent_var = tk.BooleanVar(value=self.settings.get("show_full_recent_incidents", False))
        ttk.Checkbutton(lf_recent, text="Show full behavior names (not initials)", variable=self.show_full_recent_var).grid(row=3,column=0,columnspan=2,sticky=tk.W,padx=5,pady=3)
        self.reverse_order_var = tk.BooleanVar(value=self.settings.get("reverse_incident_order", True))
        ttk.Checkbutton(lf_recent, text="Show most recent incident last (chronological)", variable=self.reverse_order_var).grid(row=4,column=0,columnspan=2,sticky=tk.W,padx=5,pady=3)

        # Custom Behaviors
        lf_custom_b = ttk.LabelFrame(tab_frame, text="Custom Behaviors & Initials", padding=10); lf_custom_b.grid(sticky="nsew",column=1,row=0, pady=5)
        custom_b_btns_frame = ttk.Frame(lf_custom_b); custom_b_btns_frame.pack(fill=tk.X)
        ttk.Button(custom_b_btns_frame, text="Add Behavior...", command=self.add_custom_behavior).pack(side=tk.LEFT, padx=2, pady=3)
        ttk.Button(custom_b_btns_frame, text="Manage Behavior/Quiz Initials...", command=self.manage_behavior_initials).pack(side=tk.LEFT, padx=2, pady=3)
        ttk.Button(custom_b_btns_frame, text="Manage Quiz Mark Types...", command=self.manage_quiz_mark_types).pack(side=tk.LEFT, padx=2, pady=3)


        self.custom_behaviors_listbox = tk.Listbox(lf_custom_b, height=5, exportselection=False)
        self.custom_behaviors_listbox.pack(fill=tk.X, expand=True, pady=(5,2))
        self.populate_custom_behaviors_listbox()
        custom_b_edit_btns_frame = ttk.Frame(lf_custom_b); custom_b_edit_btns_frame.pack(fill=tk.X)
        ttk.Button(custom_b_edit_btns_frame, text="Edit Selected", command=self.edit_selected_custom_behavior).pack(side=tk.LEFT, padx=2)
        ttk.Button(custom_b_edit_btns_frame, text="Remove Selected", command=self.remove_selected_custom_behavior).pack(side=tk.LEFT, padx=2)

        # Quiz Settings
        lf_quiz = ttk.LabelFrame(tab_frame, text="Quiz Logging & Session Settings", padding=10); lf_quiz.grid(sticky="nsew",column=0,row=1, pady=5)
        ttk.Label(lf_quiz, text="Default Quiz Name:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=3)
        self.def_quiz_name_var = tk.StringVar(value=self.settings.get("default_quiz_name", "Pop Quiz"))
        ttk.Entry(lf_quiz, textvariable=self.def_quiz_name_var, width=20).grid(row=0,column=1,sticky=tk.W,padx=5,pady=3)

        ttk.Label(lf_quiz, text="Default #Questions (Manual Log):").grid(row=1, column=0, sticky=tk.W, padx=5, pady=3)
        self.def_quiz_q_var = tk.IntVar(value=self.settings.get("default_quiz_questions",10))
        ttk.Spinbox(lf_quiz, from_=1, to=100, textvariable=self.def_quiz_q_var, width=5).grid(row=1,column=1,sticky=tk.W,padx=5,pady=3)

        ttk.Label(lf_quiz, text="Quiz Name Memory Timeout (mins):").grid(row=2, column=0, sticky=tk.W, padx=5, pady=3)
        self.quiz_timeout_var = tk.IntVar(value=self.settings.get("last_used_quiz_name_timeout_minutes", 60))
        ttk.Spinbox(lf_quiz, from_=0, to=1440, textvariable=self.quiz_timeout_var, width=5).grid(row=2,column=1,sticky=tk.W,padx=5,pady=3)

        self.show_inc_quiz_var = tk.BooleanVar(value=self.settings.get("show_recent_incidents_during_quiz", True))
        ttk.Checkbutton(lf_quiz, text="Show recent behaviors during live quiz", variable=self.show_inc_quiz_var).grid(row=3,column=0,columnspan=2,sticky=tk.W, padx=5,pady=3)
        
        self.combine_marks_display_var = tk.BooleanVar(value=self.settings.get("combine_marks_for_display", True))
        # ttk.Checkbutton(lf_quiz, text="Combine mark counts for log display (e.g., Correct: 8/10)", variable=self.combine_marks_display_var).grid(row=4,column=0,columnspan=2,sticky=tk.W, padx=5,pady=3) # Removed for now, logic complex

        ttk.Button(lf_quiz, text="Quiz Templates...", command=self.app.manage_quiz_templates_dialog).grid(row=0,column=2, padx=10, pady=3, sticky=tk.E)
        lf_quiz.grid_columnconfigure(2, weight=1)


    def create_homework_log_tab(self, tab_frame): # New Tab
        # Recent Homework Display
        lf_recent_hw = ttk.LabelFrame(tab_frame, text="Recent Homework on Student Boxes", padding=10)
        lf_recent_hw.grid(sticky="nsew",column=0,row=0, pady=0, padx=10)
        self.show_recent_hw_var = tk.BooleanVar(value=self.settings.get("show_recent_homeworks_on_boxes", True))
        ttk.Checkbutton(lf_recent_hw, text="Show recent homework logs on student boxes", variable=self.show_recent_hw_var).grid(row=0,column=0,columnspan=2,sticky=tk.W, padx=5,pady=3)
        ttk.Label(lf_recent_hw, text="Number to show:").grid(row=1,column=0,sticky=tk.W,padx=5,pady=3)
        self.num_recent_hw_var = tk.IntVar(value=self.settings.get("num_recent_homeworks_to_show", 2))
        ttk.Spinbox(lf_recent_hw, from_=0, to=10, textvariable=self.num_recent_hw_var, width=3).grid(row=1,column=1,sticky=tk.W,padx=5,pady=3)
        ttk.Label(lf_recent_hw, text="Time window (hours):").grid(row=2,column=0,sticky=tk.W,padx=5,pady=3)
        self.time_window_hw_var = tk.IntVar(value=self.settings.get("recent_homework_time_window_hours", 24))
        ttk.Spinbox(lf_recent_hw, from_=1, to=168, textvariable=self.time_window_hw_var, width=4).grid(row=2,column=1,sticky=tk.W,padx=5,pady=3)
        self.show_full_recent_hw_var = tk.BooleanVar(value=self.settings.get("show_full_recent_homeworks", False))
        ttk.Checkbutton(lf_recent_hw, text="Show full homework names (not initials)", variable=self.show_full_recent_hw_var).grid(row=3,column=0,columnspan=2,sticky=tk.W,padx=5,pady=3)
        self.reverse_hw_order_var = tk.BooleanVar(value=self.settings.get("reverse_homework_order", True))
        ttk.Checkbutton(lf_recent_hw, text="Show most recent homework last (chronological)", variable=self.reverse_hw_order_var).grid(row=4,column=0,columnspan=2,sticky=tk.W,padx=5,pady=3)

        # Custom Homework Log Behaviors (for manual logging options like "Done", "Not Done")
        lf_custom_hw_log = ttk.LabelFrame(tab_frame, text="Custom Homework Log Options & Initials", padding=10)
        lf_custom_hw_log.grid(sticky="nsew", column=1,row=0, pady=0)
        custom_hw_log_btns_frame = ttk.Frame(lf_custom_hw_log); custom_hw_log_btns_frame.pack(fill=tk.X)
        ttk.Button(custom_hw_log_btns_frame, text="Add Log Option...", command=self.add_custom_homework_log_behavior).pack(side=tk.LEFT, padx=2, pady=3)
        ttk.Button(custom_hw_log_btns_frame, text="Manage Homework Log Initials...", command=self.manage_homework_initials).pack(side=tk.LEFT, padx=2, pady=3)
        ttk.Button(custom_hw_log_btns_frame, text="Manage Homework Mark Types...", command=self.manage_homework_mark_types).pack(side=tk.LEFT, padx=2, pady=3)


        self.custom_hw_log_behaviors_listbox = tk.Listbox(lf_custom_hw_log, height=4, exportselection=False)
        self.custom_hw_log_behaviors_listbox.pack(fill=tk.X, expand=True, pady=(5,2))
        self.populate_custom_homework_log_behaviors_listbox()
        custom_hw_log_edit_btns_frame = ttk.Frame(lf_custom_hw_log); custom_hw_log_edit_btns_frame.pack(fill=tk.X)
        ttk.Button(custom_hw_log_edit_btns_frame, text="Edit Selected", command=self.edit_selected_custom_homework_log_behavior).pack(side=tk.LEFT, padx=2)
        ttk.Button(custom_hw_log_edit_btns_frame, text="Remove Selected", command=self.remove_selected_custom_homework_log_behavior).pack(side=tk.LEFT, padx=2)

        # Live Homework Session Settings
        lf_live_hw = ttk.LabelFrame(tab_frame, text="Live Homework Session Settings", padding=10)
        lf_live_hw.grid(sticky="nsew", column=0, row=1, pady=0, padx=5)
        ttk.Label(lf_live_hw, text="Default Session Name:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=3)
        self.def_hw_session_name_var = tk.StringVar(value=self.settings.get("default_homework_name", "Homework Check"))
        ttk.Entry(lf_live_hw, textvariable=self.def_hw_session_name_var, width=20).grid(row=0,column=1,sticky=tk.W,padx=5,pady=3)

        ttk.Label(lf_live_hw, text="Session Mode:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=3)
        self.live_hw_mode_var = tk.StringVar(value=self.settings.get("live_homework_session_mode", "Yes/No"))
        hw_mode_combo = ttk.Combobox(lf_live_hw, textvariable=self.live_hw_mode_var, values=["Yes/No", "Select"], state="readonly", width=10)
        hw_mode_combo.grid(row=1, column=1, sticky=tk.W, padx=5, pady=3)
        hw_mode_combo.bind("<<ComboboxSelected>>", self.on_live_hw_mode_change)


        # Settings specific to "Yes/No" mode
        self.yes_no_settings_frame = ttk.Frame(lf_live_hw)
        self.yes_no_settings_frame.grid(row=2, column=0, columnspan=3, sticky=tk.EW, padx=5, pady=3)
        # Custom Homework Session Types (for Yes/No mode list)
        lf_custom_hw_session_types = ttk.LabelFrame(self.yes_no_settings_frame, text="Custom Homework Types for 'Yes/No' Session", padding=5)
        lf_custom_hw_session_types.pack(fill=tk.X, pady=3)
        custom_hw_session_btns_frame = ttk.Frame(lf_custom_hw_session_types); custom_hw_session_btns_frame.pack(fill=tk.X)
        ttk.Button(custom_hw_session_btns_frame, text="Add Type...", command=self.add_custom_homework_session_type).pack(side=tk.LEFT, padx=2, pady=2)
        self.custom_hw_session_types_listbox = tk.Listbox(lf_custom_hw_session_types, height=3, exportselection=False)
        self.custom_hw_session_types_listbox.pack(fill=tk.X, expand=True, pady=(3,2))
        self.populate_custom_homework_session_types_listbox()
        custom_hw_session_edit_btns_frame = ttk.Frame(lf_custom_hw_session_types); custom_hw_session_edit_btns_frame.pack(fill=tk.X)
        ttk.Button(custom_hw_session_edit_btns_frame, text="Edit Selected", command=self.edit_selected_custom_homework_session_type).pack(side=tk.LEFT, padx=2)
        ttk.Button(custom_hw_session_edit_btns_frame, text="Remove Selected", command=self.remove_selected_custom_homework_session_type).pack(side=tk.LEFT, padx=2)


        # Settings specific to "Select" mode
        self.select_mode_settings_frame = ttk.Frame(lf_live_hw)
        # self.select_mode_settings_frame.grid(row=2, column=0, columnspan=3, sticky=tk.EW, padx=5, pady=3) # Positioned by on_live_hw_mode_change
        lf_select_options = ttk.LabelFrame(self.select_mode_settings_frame, text="Options for 'Select' Session Mode", padding=5)
        lf_select_options.pack(fill=tk.X, pady=3)
        # Add UI to manage self.settings["live_homework_select_mode_options"] (list of dicts {"name": "..."})
        # For now, it uses DEFAULT_HOMEWORK_SESSION_BUTTONS. A more complex UI would allow user to customize these.
        ttk.Button(lf_select_options, text="Manage 'Select' Options...", command=self.manage_live_homework_select_options).pack(pady=3, anchor=tk.W)


        # General Homework Settings
        self.log_hw_marks_var = tk.BooleanVar(value=self.settings.get("log_homework_marks_enabled", True))
        ttk.Checkbutton(lf_live_hw, text="Enable Detailed Marks for Manual Homework Logging", variable=self.log_hw_marks_var).grid(row=3,column=0,columnspan=3,sticky=tk.W, padx=5,pady=3)
        ttk.Button(lf_live_hw, text="Homework Templates...", command=self.app.manage_homework_templates_dialog).grid(row=0,column=2, padx=10, pady=3, sticky=tk.E)
        lf_live_hw.grid_columnconfigure(2, weight=1)

        self.on_live_hw_mode_change(None) # Show/hide mode-specific frames


    def on_live_hw_mode_change(self, event):
        mode = self.live_hw_mode_var.get()
        if mode == "Yes/No":
            self.select_mode_settings_frame.grid_remove()
            self.yes_no_settings_frame.grid(row=2, column=0, columnspan=3, sticky=tk.EW, padx=5, pady=3)
        elif mode == "Select":
            self.yes_no_settings_frame.grid_remove()
            self.select_mode_settings_frame.grid(row=2, column=0, columnspan=3, sticky=tk.EW, padx=5, pady=3)
        else:
            self.yes_no_settings_frame.grid_remove()
            self.select_mode_settings_frame.grid_remove()

    def create_data_export_tab(self, tab_frame):
        lf_excel = ttk.LabelFrame(tab_frame, text="Excel Export Defaults", padding=10); lf_excel.pack(fill=tk.X, pady=5)
        self.excel_sep_sheets_var = tk.BooleanVar(value=self.settings.get("excel_export_separate_sheets_by_default", True))
        ttk.Checkbutton(lf_excel, text="Separate log types into different sheets by default", variable=self.excel_sep_sheets_var).pack(anchor=tk.W, padx=5, pady=2)
        self.excel_inc_summary_var = tk.BooleanVar(value=self.settings.get("excel_export_include_summaries_by_default", True))
        ttk.Checkbutton(lf_excel, text="Include summary sheet by default", variable=self.excel_inc_summary_var).pack(anchor=tk.W, padx=5, pady=2)

        lf_autosave_excel = ttk.LabelFrame(tab_frame, text="Excel Log Autosave (Experimental)", padding=10); lf_autosave_excel.pack(fill=tk.X, pady=5)
        self.enable_excel_autosave_var = tk.BooleanVar(value=self.settings.get("enable_excel_autosave", False))
        ttk.Checkbutton(lf_autosave_excel, text=f"Enable autosaving log to Excel file ({os.path.basename(AUTOSAVE_EXCEL_FILE)})", variable=self.enable_excel_autosave_var).pack(anchor=tk.W, padx=5, pady=2)
        ttk.Label(lf_autosave_excel, text="Note: This uses current export filters if set, or exports all data. File is overwritten each time.").pack(anchor=tk.W, padx=5, pady=2)


    def create_security_tab(self, tab_frame):
        lf_password = ttk.LabelFrame(tab_frame, text="Application Password", padding=10)
        lf_password.pack(fill=tk.X, pady=5)

        current_pw_set = self.password_manager.is_password_set()
        self.current_pw_status_label = ttk.Label(lf_password, text="Status: Password IS SET" if current_pw_set else "Status: Password NOT SET")
        self.current_pw_status_label.grid(row=0, column=0, columnspan=2, sticky=tk.W, padx=5, pady=5)

        ttk.Label(lf_password, text="New Password:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=3)
        self.new_pw_var = tk.StringVar()
        new_pw_entry = ttk.Entry(lf_password, textvariable=self.new_pw_var, show="*", width=25)
        new_pw_entry.grid(row=1, column=1, sticky=tk.W, padx=5, pady=3)

        ttk.Label(lf_password, text="Confirm New Password:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=3)
        self.confirm_pw_var = tk.StringVar()
        confirm_pw_entry = ttk.Entry(lf_password, textvariable=self.confirm_pw_var, show="*", width=25)
        confirm_pw_entry.grid(row=2, column=1, sticky=tk.W, padx=5, pady=3)

        set_pw_btn = ttk.Button(lf_password, text="Set/Change Password", command=self.set_or_change_password)
        set_pw_btn.grid(row=3, column=0, padx=5, pady=10)
        remove_pw_btn = ttk.Button(lf_password, text="Remove Password", command=self.remove_password, state=tk.NORMAL if current_pw_set else tk.DISABLED)
        remove_pw_btn.grid(row=3, column=1, padx=5, pady=10, sticky=tk.W)
        self.remove_pw_button_ref = remove_pw_btn # To update state

        lf_pw_options = ttk.LabelFrame(tab_frame, text="Password Options", padding=10)
        lf_pw_options.pack(fill=tk.X, pady=5)
        self.pw_on_open_var = tk.BooleanVar(value=self.settings.get("password_on_open", False))
        ttk.Checkbutton(lf_pw_options, text="Require password on application open", variable=self.pw_on_open_var).pack(anchor=tk.W, padx=5, pady=2)
        self.pw_on_edit_var = tk.BooleanVar(value=self.settings.get("password_on_edit_action", False))
        ttk.Checkbutton(lf_pw_options, text="Require password for sensitive actions (add/edit/delete items, layout changes)", variable=self.pw_on_edit_var).pack(anchor=tk.W, padx=5, pady=2)
        
        auto_lock_frame = ttk.Frame(lf_pw_options); auto_lock_frame.pack(fill=tk.X, pady=2)
        self.pw_auto_lock_var = tk.BooleanVar(value=self.settings.get("password_auto_lock_enabled", False))
        ttk.Checkbutton(auto_lock_frame, text="Auto-lock application after inactivity for", variable=self.pw_auto_lock_var).pack(side=tk.LEFT, anchor=tk.W, padx=5)
        self.pw_auto_lock_timeout_var = tk.IntVar(value=self.settings.get("password_auto_lock_timeout_minutes", 15))
        ttk.Spinbox(auto_lock_frame, from_=1, to=120, textvariable=self.pw_auto_lock_timeout_var, width=4).pack(side=tk.LEFT, padx=2)
        ttk.Label(auto_lock_frame, text="minutes").pack(side=tk.LEFT)

        ttk.Label(lf_pw_options, text="For the Master Recovery Password, ask Yaakov Maimon (see Help)", foreground="blue", wraplength=420).pack(anchor=tk.W, padx=5, pady=5)


    def theme_set(self, event):
        #print(self.app.theme_style_using, "old")
        self.app.theme_style_using = self.theme.get()
        self.settings_changed_flag = True
        #print("Theme: ", self.theme.get())
        self.theme2 = self.theme.get()
        #print("theme2", self.theme2)


    def set_or_change_password(self):
        new_pw = self.new_pw_var.get()
        confirm_pw = self.confirm_pw_var.get()
        if not new_pw:
            messagebox.showerror("Password Error", "New password cannot be empty.", parent=self)
            return
        if new_pw != confirm_pw:
            messagebox.showerror("Password Error", "Passwords do not match.", parent=self)
            return
        if len(new_pw) < 4: # Basic length check
            messagebox.showwarning("Weak Password", "Password should be at least 4 characters.", parent=self)
            # Allow user to proceed if they insist
        
        self.password_manager.set_password(new_pw)
        self.settings_changed_flag = True # Settings (hash) changed
        self.new_pw_var.set(""); self.confirm_pw_var.set("")
        self.current_pw_status_label.config(text="Status: Password IS SET")
        self.remove_pw_button_ref.config(state=tk.NORMAL)
        messagebox.showinfo("Password Set", "Application password has been set/changed.", parent=self)

    def prompt_for_password(self, title, prompt_message, for_editing=False):
        if self.password_manager.is_locked:
             if not hasattr(self, '_lock_screen_active') or not self._lock_screen_active.winfo_exists(): self.show_lock_screen()
             return not self.password_manager.is_locked
        if for_editing and not self.settings.get("password_on_edit_action", False) and not self.password_manager.is_password_set(): return True
        if not self.password_manager.is_password_set(): return True
        dialog = PasswordPromptDialog(self.master, title, prompt_message, self.password_manager)
        return dialog.result

    def remove_password(self):
        if self.password_manager.is_password_set():
            if self.prompt_for_password("Confirm Password", "Enter current password to confirm identity", for_editing=True):
                if messagebox.askyesno("Confirm Removal", "Are you sure you want to remove the application password?", parent=self):
                    self.password_manager.set_password(None) # Set to None effectively removes it
                    self.settings_changed_flag = True
                    self.current_pw_status_label.config(text="Status: Password NOT SET")
                    self.remove_pw_button_ref.config(state=tk.DISABLED)
                    self.pw_on_open_var.set(False) # Disable options that require a password
                    self.pw_on_edit_var.set(False)
                    self.pw_auto_lock_var.set(False)
                    messagebox.showinfo("Password Removed", "Application password has been removed.", parent=self)
        else:
            messagebox.showinfo("No Password", "No application password is currently set.", parent=self)
    

    # --- UI Helper for color/font settings ---
    def create_color_font_settings_ui(self, parent_frame, start_row, fill_key, outline_key, font_fam_key, font_size_key, font_color_key):
        # Fill Color
        ttk.Label(parent_frame, text="Default Fill Color:").grid(row=start_row,column=0,sticky=tk.W,padx=5,pady=3)
        fill_var = tk.StringVar(value=self.settings.get(fill_key, DEFAULT_BOX_FILL_COLOR))
        setattr(self, f"{fill_key}_var", fill_var) # Store var on self
        ttk.Entry(parent_frame, textvariable=fill_var, width=12).grid(row=start_row,column=1,sticky=tk.W,padx=5,pady=3)
        ttk.Button(parent_frame, text="Choose...", command=lambda v=fill_var: self.choose_color_for_var(v)).grid(row=start_row,column=2,sticky=tk.W,padx=2,pady=3)
        # Outline Color
        ttk.Label(parent_frame, text="Default Outline Color:").grid(row=start_row+1,column=0,sticky=tk.W,padx=5,pady=3)
        outline_var = tk.StringVar(value=self.settings.get(outline_key, DEFAULT_BOX_OUTLINE_COLOR))
        setattr(self, f"{outline_key}_var", outline_var)
        ttk.Entry(parent_frame, textvariable=outline_var, width=12).grid(row=start_row+1,column=1,sticky=tk.W,padx=5,pady=3)
        ttk.Button(parent_frame, text="Choose...", command=lambda v=outline_var: self.choose_color_for_var(v)).grid(row=start_row+1,column=2,sticky=tk.W,padx=2,pady=3)
        # Font Family
        ttk.Label(parent_frame, text="Default Font Family:").grid(row=start_row+2,column=0,sticky=tk.W,padx=5,pady=3)
        font_fam_var = tk.StringVar(value=self.settings.get(font_fam_key, DEFAULT_FONT_FAMILY))
        setattr(self, f"{font_fam_key}_var", font_fam_var)
        available_fonts = self.settings.get("available_fonts", [DEFAULT_FONT_FAMILY])
        ff_combo = ttk.Combobox(parent_frame, textvariable=font_fam_var, values=available_fonts, width=20, state="readonly")
        ff_combo.grid(row=start_row+2,column=1,columnspan=2,sticky=tk.EW,padx=5,pady=3)
        ff_combo.bind("<MouseWheel>", lambda event: "break") # Prevent main canvas scroll
        # Font Size
        ttk.Label(parent_frame, text="Default Font Size (pts):").grid(row=start_row+3,column=0,sticky=tk.W,padx=5,pady=3)
        font_size_var = tk.IntVar(value=self.settings.get(font_size_key, DEFAULT_FONT_SIZE))
        setattr(self, f"{font_size_key}_var", font_size_var)
        ttk.Spinbox(parent_frame, from_=6, to=24, textvariable=font_size_var, width=5).grid(row=start_row+3,column=1,sticky=tk.W,padx=5,pady=3)
        # Font Color
        ttk.Label(parent_frame, text="Default Font Color:").grid(row=start_row+4,column=0,sticky=tk.W,padx=5,pady=3)
        font_color_var = tk.StringVar(value=self.settings.get(font_color_key, DEFAULT_FONT_COLOR))
        setattr(self, f"{font_color_key}_var", font_color_var)
        ttk.Entry(parent_frame, textvariable=font_color_var, width=12).grid(row=start_row+4,column=1,sticky=tk.W,padx=5,pady=3)
        ttk.Button(parent_frame, text="Choose...", command=lambda v=font_color_var: self.choose_color_for_var(v)).grid(row=start_row+4,column=2,sticky=tk.W,padx=2,pady=3)

    def choose_color_for_var(self, color_var): # Helper for color choosers in settings
        initial_color = color_var.get()
        if not initial_color: # If empty, pick a default to show in chooser
            if "fill" in color_var._name: initial_color = DEFAULT_BOX_FILL_COLOR
            elif "outline" in color_var._name: initial_color = DEFAULT_BOX_OUTLINE_COLOR
            else: initial_color = DEFAULT_FONT_COLOR
        
        color_code = colorchooser.askcolor(initial_color, title="Choose color", parent=self)
        if color_code and color_code[1]: color_var.set(color_code[1])

    def choose_color_for_canvas(self, color_var): # Helper for color choosers in settings
        initial_color = color_var.get()
        if initial_color == "Default": initial_color = None
        if not initial_color: # If empty, pick a default to show in chooser
            if "fill" in color_var._name: initial_color = DEFAULT_BOX_FILL_COLOR
            elif "outline" in color_var._name: initial_color = DEFAULT_BOX_OUTLINE_COLOR
            else: initial_color = DEFAULT_FONT_COLOR
        
        color_code = colorchooser.askcolor(initial_color, title="Choose color", parent=self)
        if color_code and color_code[1]: color_var.set(color_code[1])
        
    def reset_canvas_color(self, button):
        button.set("Default")

    # --- Methods for managing custom lists ---
    def populate_conditional_rules_listbox(self):
        self.rules_listbox.delete(0, tk.END)
        for i, rule in enumerate(self.settings.get("conditional_formatting_rules", [])):
            desc = f"Rule {i+1}: Type='{rule['type']}'"
            if rule['type'] == 'group': desc += f", Group='{self.app.student_groups.get(rule.get('group_id'),{}).get('name','Unknown')}'"
            elif rule['type'] == 'behavior_count': desc += f", Behavior='{rule.get('behavior_name')}', Count>={rule.get('count_threshold')}, Hours={rule.get('time_window_hours')}"
            elif rule['type'] == 'quiz_score_threshold': desc += f", Quiz~'{rule.get('quiz_name_contains','')}', Score {rule.get('operator','<=')} {rule.get('score_threshold_percent','')}%"
            desc += f" -> Fill='{rule.get('color','')}', Outline='{rule.get('outline','')}'"
            self.rules_listbox.insert(tk.END, desc)
    def add_conditional_rule(self):
        dialog = ConditionalFormattingRuleDialog(self, self.app) # Pass app
        if dialog.result:
            if "conditional_formatting_rules" not in self.settings: self.settings["conditional_formatting_rules"] = []
            self.settings["conditional_formatting_rules"].append(dialog.result)
            self.settings_changed_flag = True
            self.populate_conditional_rules_listbox()
    def edit_selected_conditional_rule(self):
        selected_idx = self.rules_listbox.curselection()
        if not selected_idx: messagebox.showinfo("No Selection", "Please select a rule to edit.", parent=self); return
        idx_to_edit = selected_idx[0]
        rule_copy = self.settings["conditional_formatting_rules"][idx_to_edit].copy()
        dialog = ConditionalFormattingRuleDialog(self, self.app, rule_to_edit=rule_copy)
        if dialog.result:
            self.settings["conditional_formatting_rules"][idx_to_edit] = dialog.result
            self.settings_changed_flag = True
            self.populate_conditional_rules_listbox()
    def remove_selected_conditional_rule(self):
        selected_idx = self.rules_listbox.curselection()
        if not selected_idx: messagebox.showinfo("No Selection", "Please select a rule to remove.", parent=self); return
        if messagebox.askyesno("Confirm Remove", "Remove selected conditional formatting rule?", parent=self):
            del self.settings["conditional_formatting_rules"][selected_idx[0]]
            self.settings_changed_flag = True
            self.populate_conditional_rules_listbox()


    def _manage_custom_list_generic(self, listbox, custom_list_ref, item_type_name, add_func_name, edit_func_name):
        # This is a placeholder for a more generic dialog if needed, for now specific ones are used
        pass

    # Custom Behaviors (for Log Behavior dialog)
    def populate_custom_behaviors_listbox(self):
        self.custom_behaviors_listbox.delete(0, tk.END)
        for behavior_item in self.custom_behaviors_ref: # List of dicts or old strings
            name = behavior_item["name"] if isinstance(behavior_item, dict) else behavior_item
            self.custom_behaviors_listbox.insert(tk.END, name)
    def add_custom_behavior(self):
        if len(self.custom_behaviors_ref) >= MAX_CUSTOM_TYPES:
            messagebox.showwarning("Limit Reached", f"Maximum of {MAX_CUSTOM_TYPES} custom {item_type_name.lower()}s allowed.", parent=self); return
        name = simpledialog.askstring("Add Custom Behavior", "Enter name for the new behavior:", parent=self)
        if name and name.strip():
            name = name.strip()
            if any(cb_item == name or (isinstance(cb_item, dict) and cb_item.get("name") == name) for cb_item in self.custom_behaviors_ref):
                 messagebox.showwarning("Duplicate", f"Behavior '{name}' already exists.", parent=self); return
            self.custom_behaviors_ref.append({"name": name}) # Store as dict
            self.settings_changed_flag = True; self.app.save_custom_behaviors(); self.populate_custom_behaviors_listbox()
    def edit_selected_custom_behavior(self):
        sel_idx = self.custom_behaviors_listbox.curselection()
        if not sel_idx: messagebox.showinfo("No Selection", "Please select a behavior to edit.", parent=self); return
        idx = sel_idx[0]
        old_item = self.custom_behaviors_ref[idx]
        old_name = old_item["name"] if isinstance(old_item, dict) else old_item
        new_name = simpledialog.askstring("Edit Custom Behavior", "Enter new name:", initialvalue=old_name, parent=self)
        if new_name and new_name.strip():
            new_name = new_name.strip()
            if new_name != old_name and any(cb_item == new_name or (isinstance(cb_item, dict) and cb_item.get("name") == new_name and (idx != i if isinstance(cb_item,dict) else True) ) for i, cb_item in enumerate(self.custom_behaviors_ref)):
                 messagebox.showwarning("Duplicate", f"Behavior '{new_name}' already exists.", parent=self); return
            self.custom_behaviors_ref[idx] = {"name": new_name}
            self.settings_changed_flag = True; self.app.save_custom_behaviors(); self.populate_custom_behaviors_listbox()
    def remove_selected_custom_behavior(self):
        sel_idx = self.custom_behaviors_listbox.curselection()
        if not sel_idx: messagebox.showinfo("No Selection", "Please select a behavior to remove.", parent=self); return
        if messagebox.askyesno("Confirm Remove", "Remove selected custom behavior?", parent=self):
            del self.custom_behaviors_ref[sel_idx[0]]
            self.settings_changed_flag = True; self.app.save_custom_behaviors(); self.populate_custom_behaviors_listbox()

    # Custom Homework Log Behaviors (for Manual Homework Log dialog options)
    def populate_custom_homework_log_behaviors_listbox(self):
        self.custom_hw_log_behaviors_listbox.delete(0, tk.END)
        for item in self.custom_homework_log_behaviors_ref:
            self.custom_hw_log_behaviors_listbox.insert(tk.END, item["name"])
    def add_custom_homework_log_behavior(self):
        if len(self.custom_homework_log_behaviors_ref) >= MAX_CUSTOM_TYPES:
             messagebox.showwarning("Limit Reached", f"Maximum of {MAX_CUSTOM_TYPES} custom homework log options allowed.", parent=self); return
        name = simpledialog.askstring("Add Homework Log Option", "Enter name for the new option (e.g., 'Excellent Effort'):", parent=self)
        if name and name.strip():
            name = name.strip()
            if any(item["name"] == name for item in self.custom_homework_log_behaviors_ref):
                 messagebox.showwarning("Duplicate", f"Option '{name}' already exists.", parent=self); return
            self.custom_homework_log_behaviors_ref.append({"name": name})
            self.settings_changed_flag = True; self.app.save_custom_homework_log_behaviors(); self.populate_custom_homework_log_behaviors_listbox()
    def edit_selected_custom_homework_log_behavior(self):
        sel_idx = self.custom_hw_log_behaviors_listbox.curselection()
        if not sel_idx: messagebox.showinfo("No Selection", "Please select an option to edit.", parent=self); return
        idx = sel_idx[0]; old_name = self.custom_homework_log_behaviors_ref[idx]["name"]
        new_name = simpledialog.askstring("Edit Homework Log Option", "Enter new name:", initialvalue=old_name, parent=self)
        if new_name and new_name.strip():
            new_name = new_name.strip()
            if new_name != old_name and any(item["name"] == new_name for i, item in enumerate(self.custom_homework_log_behaviors_ref) if i != idx):
                 messagebox.showwarning("Duplicate", f"Option '{new_name}' already exists.", parent=self); return
            self.custom_homework_log_behaviors_ref[idx]["name"] = new_name
            self.settings_changed_flag = True; self.app.save_custom_homework_log_behaviors(); self.populate_custom_homework_log_behaviors_listbox()
    def remove_selected_custom_homework_log_behavior(self):
        sel_idx = self.custom_hw_log_behaviors_listbox.curselection()
        if not sel_idx: messagebox.showinfo("No Selection", "Please select an option to remove.", parent=self); return
        if messagebox.askyesno("Confirm Remove", "Remove selected homework log option?", parent=self):
            del self.custom_homework_log_behaviors_ref[sel_idx[0]]
            self.settings_changed_flag = True; self.app.save_custom_homework_log_behaviors(); self.populate_custom_homework_log_behaviors_listbox()

    # Custom Homework Session Types (for Live Homework "Yes/No" mode)
    def populate_custom_homework_session_types_listbox(self):
        self.custom_hw_session_types_listbox.delete(0, tk.END)
        for item in self.custom_homework_session_types_ref: # list of {"id", "name"}
            self.custom_hw_session_types_listbox.insert(tk.END, item["name"])
    def add_custom_homework_session_type(self):
        if len(self.custom_homework_session_types_ref) >= MAX_CUSTOM_TYPES:
             messagebox.showwarning("Limit Reached", f"Maximum of {MAX_CUSTOM_TYPES} custom homework types for sessions allowed.", parent=self); return
        name = simpledialog.askstring("Add Homework Session Type", "Enter name for the new type (e.g., 'Project Milestone 1'):", parent=self)
        if name and name.strip():
            name = name.strip()
            if any(item["name"] == name for item in self.custom_homework_session_types_ref):
                 messagebox.showwarning("Duplicate", f"Type '{name}' already exists.", parent=self); return
            type_id_str, next_id_val = self.app.get_new_custom_homework_type_id()
            self.app.settings["next_custom_homework_type_id_num"] = next_id_val # Commit ID usage
            self.custom_homework_session_types_ref.append({"id": type_id_str, "name": name})
            self.settings_changed_flag = True; self.app.save_custom_homework_session_types(); self.populate_custom_homework_session_types_listbox()
    def edit_selected_custom_homework_session_type(self):
        sel_idx = self.custom_hw_session_types_listbox.curselection()
        if not sel_idx: messagebox.showinfo("No Selection", "Please select a type to edit.", parent=self); return
        idx = sel_idx[0]; old_name = self.custom_homework_session_types_ref[idx]["name"]
        new_name = simpledialog.askstring("Edit Homework Session Type", "Enter new name:", initialvalue=old_name, parent=self)
        if new_name and new_name.strip():
            new_name = new_name.strip()
            if new_name != old_name and any(item["name"] == new_name for i, item in enumerate(self.custom_homework_session_types_ref) if i != idx):
                 messagebox.showwarning("Duplicate", f"Type '{new_name}' already exists.", parent=self); return
            self.custom_homework_session_types_ref[idx]["name"] = new_name
            self.settings_changed_flag = True; self.app.save_custom_homework_session_types(); self.populate_custom_homework_session_types_listbox()
    def remove_selected_custom_homework_session_type(self):
        sel_idx = self.custom_hw_session_types_listbox.curselection()
        if not sel_idx: messagebox.showinfo("No Selection", "Please select a type to remove.", parent=self); return
        if messagebox.askyesno("Confirm Remove", "Remove selected homework session type?", parent=self):
            del self.custom_homework_session_types_ref[sel_idx[0]]
            self.settings_changed_flag = True; self.app.save_custom_homework_session_types(); self.populate_custom_homework_session_types_listbox()

    def manage_behavior_initials(self):
        dialog = ManageInitialsDialog(self, self.settings["behavior_initial_overrides"], self.app.all_behaviors, "Behavior/Quiz")
        if dialog.initials_changed: self.settings_changed_flag = True
    def manage_homework_initials(self): # New
        dialog = ManageInitialsDialog(self, self.settings["homework_initial_overrides"], self.app.all_homework_log_behaviors, "Homework Log")
        if dialog.initials_changed: self.settings_changed_flag = True
    def manage_quiz_mark_types(self):
        dialog = ManageMarkTypesDialog(self, self.settings["quiz_mark_types"], "Quiz Mark Types", DEFAULT_QUIZ_MARK_TYPES)
        if dialog.mark_types_changed: self.settings_changed_flag = True
    def manage_homework_mark_types(self): # New
        dialog = ManageMarkTypesDialog(self, self.settings["homework_mark_types"], "Homework Mark Types", DEFAULT_HOMEWORK_MARK_TYPES)
        if dialog.mark_types_changed: self.settings_changed_flag = True
    def manage_live_homework_select_options(self):
        dialog = ManageLiveSelectOptionsDialog(self, self.settings.get("live_homework_select_mode_options", DEFAULT_HOMEWORK_SESSION_BUTTONS.copy()))
        if dialog.options_changed_flag:
            self.settings["live_homework_select_mode_options"] = dialog.current_options
            self.settings_changed_flag = True

    """def buttonbox(self):
        ttk.Button(self, text= "Ok", command=self.ok).grid(column=0,row=1)
        ttk.Button(self, text="Cancel", command=self.cancel).grid(column=1,row=1, padx=10)
    """
    
    
    def apply(self): # OK button of SettingsDialog
        # General Tab
        self.settings["autosave_interval_ms"] = self.autosave_interval_var.get() * 1000
        self.settings["grid_snap_enabled"] = self.grid_snap_var.get()
        self.settings["grid_size"] = self.grid_size_var.get()
        self.settings["student_groups_enabled"] = self.groups_enabled_var.get()
        self.settings["show_zoom_level_display"] = self.show_zoom_var.get()
        self.settings["max_undo_history_days"] = self.max_undo_days_var.get()
        self.settings["theme"] = self.theme.get()
        self.settings["canvas_color"] = self.custom_canvas_color.get()
        self.app.theme_style_using = self.theme2
        self.app.custom_canvas_color = self.custom_canvas_color.get()
        #print("Theme2", self.theme2)
        #print(self.theme.get(), "Get")
        # Student Display Tab
        self.settings["default_student_box_width"]=self.def_stud_w_var.get()
        self.settings["default_student_box_height"]=self.def_stud_h_var.get()
        self.settings["student_box_fill_color"]=self.student_box_fill_color_var.get()
        self.settings["student_box_outline_color"]=self.student_box_outline_color_var.get()
        self.settings["student_font_family"]=self.student_font_family_var.get()
        self.settings["student_font_size"]=self.student_font_size_var.get()
        self.settings["student_font_color"]=self.student_font_color_var.get()
        # Behavior/Quiz Log Tab
        self.settings["show_recent_incidents_on_boxes"] = self.show_recent_var.get()
        self.settings["num_recent_incidents_to_show"] = self.num_recent_var.get()
        self.settings["recent_incident_time_window_hours"] = self.time_window_var.get()
        self.settings["show_full_recent_incidents"] = self.show_full_recent_var.get()
        self.settings["reverse_incident_order"] = self.reverse_order_var.get()
        self.settings["default_quiz_name"] = self.def_quiz_name_var.get()
        self.settings["default_quiz_questions"] = self.def_quiz_q_var.get()
        self.settings["last_used_quiz_name_timeout_minutes"] = self.quiz_timeout_var.get()
        self.settings["show_recent_incidents_during_quiz"] = self.show_inc_quiz_var.get()
        # self.settings["combine_marks_for_display"] = self.combine_marks_display_var.get()
        # Homework Log Tab
        self.settings["show_recent_homeworks_on_boxes"] = self.show_recent_hw_var.get()
        self.settings["num_recent_homeworks_to_show"] = self.num_recent_hw_var.get()
        self.settings["recent_homework_time_window_hours"] = self.time_window_hw_var.get()
        self.settings["show_full_recent_homeworks"] = self.show_full_recent_hw_var.get()
        self.settings["reverse_homework_order"] = self.reverse_hw_order_var.get()
        self.settings["default_homework_name"] = self.def_hw_session_name_var.get()
        self.settings["live_homework_session_mode"] = self.live_hw_mode_var.get()
        self.settings["log_homework_marks_enabled"] = self.log_hw_marks_var.get()
        # Data & Export Tab
        self.settings["excel_export_separate_sheets_by_default"] = self.excel_sep_sheets_var.get()
        self.settings["excel_export_include_summaries_by_default"] = self.excel_inc_summary_var.get()
        self.settings["enable_excel_autosave"] = self.enable_excel_autosave_var.get()
        # Security Tab
        self.settings["password_on_open"] = self.pw_on_open_var.get()
        self.settings["password_on_edit_action"] = self.pw_on_edit_var.get()
        self.settings["password_auto_lock_enabled"] = self.pw_auto_lock_var.get()
        self.settings["password_auto_lock_timeout_minutes"] = self.pw_auto_lock_timeout_var.get()

        # Check if any significant setting actually changed by comparing with snapshot
        for key, initial_val in self.initial_settings_snapshot.items():
            current_val = self.settings.get(key)
            if isinstance(current_val, (list, dict)): # For mutable types, content comparison is needed
                if initial_val != current_val: # This might not catch all deep changes if not careful
                    self.settings_changed_flag = True; break
            elif initial_val != current_val:
                self.settings_changed_flag = True; break
        # If any custom list (behaviors, marks, etc.) was modified, their specific dialogs
        # would have set self.settings_changed_flag = True already.
        self.result = self.settings_changed_flag # Simpledialog will check this



class ManageInitialsDialog(simpledialog.Dialog):
    def __init__(self, parent, initials_overrides_dict, all_names_list, item_type_display_name):
        self.initials_overrides = initials_overrides_dict # Direct reference to app.settings[key]
        self.all_names = sorted(list(set(all_names_list)))
        self.item_type_name = item_type_display_name # e.g., "Behavior/Quiz" or "Homework Log"
        self.initials_changed = False
        self.entry_vars = {} # name -> StringVar
        super().__init__(parent, f"Manage Initials for {self.item_type_name}")

    def body(self, master):
        info_label = ttk.Label(master, text=f"Set custom initials for {self.item_type_name} types.\nLeave blank to use auto-generated initials.", justify=tk.LEFT)
        info_label.pack(pady=5, padx=5, anchor=tk.W)
        
        canvas_frame = ttk.Frame(master); canvas_frame.pack(fill=tk.BOTH, expand=True)
        self.canvas_initials = tk.Canvas(canvas_frame, borderwidth=0)
        scrollable_content_frame = ttk.Frame(self.canvas_initials)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=self.canvas_initials.yview)
        self.canvas_initials.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y"); self.canvas_initials.pack(side="left", fill="both", expand=True)
        self.canvas_initials.create_window((0,0), window=scrollable_content_frame, anchor="nw")
        scrollable_content_frame.bind("<Configure>", lambda e: self.canvas_initials.configure(scrollregion=self.canvas_initials.bbox("all")))
        self.canvas_initials.bind('<MouseWheel>', lambda e: self.canvas_initials.yview_scroll(int(-1*(e.delta/120)), "units"))

        for name in self.all_names:
            item_frame = ttk.Frame(scrollable_content_frame); item_frame.pack(fill=tk.X, pady=1)
            ttk.Label(item_frame, text=name, width=30, anchor=tk.W).pack(side=tk.LEFT, padx=5)
            var = tk.StringVar(value=self.initials_overrides.get(name, ""))
            self.entry_vars[name] = var
            entry = ttk.Entry(item_frame, textvariable=var, width=5)
            entry.pack(side=tk.LEFT, padx=5)
        return canvas_frame # Or some specific entry if needed for focus

    def apply(self):
        for name, var in self.entry_vars.items():
            new_initial = var.get().strip()[:3] # Max 3 chars for initials
            if new_initial:
                if self.initials_overrides.get(name) != new_initial:
                    self.initials_overrides[name] = new_initial
                    self.initials_changed = True
            elif name in self.initials_overrides: # If blanked out and was set
                del self.initials_overrides[name]
                self.initials_changed = True
        # self.initials_overrides dictionary is modified directly.
        # self.initials_changed flag will be checked by SettingsDialog.


class ManageMarkTypesDialog(simpledialog.Dialog):
    def __init__(self, parent, current_mark_types_list, item_type_display_name, default_mark_types):
        self.mark_types_ref = current_mark_types_list # Direct reference, e.g., app.settings["quiz_mark_types"]
        self.item_type_name = item_type_display_name # "Quiz Mark Types" or "Homework Mark Types"
        self.default_mark_types_const = default_mark_types # The constant list for reset
        self.mark_types_changed = False
        super().__init__(parent, f"Manage {self.item_type_name}")

    def body(self, master):
        self.main_frame = master # To rebuild list
        button_frame = ttk.Frame(master); button_frame.pack(fill=tk.X, pady=5)
        ttk.Button(button_frame, text="Add New Mark Type", command=self.add_mark_type).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Reset to Defaults", command=self.reset_to_defaults).pack(side=tk.LEFT, padx=5)

        self.list_frame = ttk.Frame(master); self.list_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.populate_mark_types_ui()
        return self.main_frame # For focus

    def populate_mark_types_ui(self):
        for widget in self.list_frame.winfo_children(): widget.destroy()
        
        headers = ["ID", "Name", "Default Points", "To Total?", "Bonus?"]
        for c, h_text in enumerate(headers):
            ttk.Label(self.list_frame, text=h_text, font=("",9,"bold")).grid(row=0,column=c,padx=3,pady=3,sticky=tk.W)

        self.mark_type_widgets = [] # Store refs to entry vars/widgets if needed for direct update
        
        for r_idx, mt_dict in enumerate(self.mark_types_ref, start=1):
            widgets_row = {}
            # ID (display only for defaults, editable for custom?) - For now, mostly fixed post-creation
            id_val = mt_dict.get("id", f"custom_{r_idx}")
            widgets_row["id_label"] = ttk.Label(self.list_frame, text=id_val, width=15); widgets_row["id_label"].grid(row=r_idx, column=0, padx=3, sticky=tk.W)
            
            # Name
            name_var = tk.StringVar(value=mt_dict.get("name","")); widgets_row["name_var"] = name_var
            name_entry = ttk.Entry(self.list_frame, textvariable=name_var, width=20); name_entry.grid(row=r_idx, column=1, padx=3, sticky=tk.EW)
            
            # Default Points
            points_var = tk.DoubleVar(value=mt_dict.get("default_points",0.0)); widgets_row["points_var"] = points_var
            points_spin = ttk.Spinbox(self.list_frame, from_=-100, to=100, increment=0.1, textvariable=points_var, width=6); points_spin.grid(row=r_idx, column=2, padx=3)

            # Contributes to Total (Bool)
            to_total_var = tk.BooleanVar(value=mt_dict.get("contributes_to_total",True)); widgets_row["to_total_var"] = to_total_var
            ttk.Checkbutton(self.list_frame, variable=to_total_var).grid(row=r_idx, column=3, padx=3)
            
            # Is Extra Credit (Bool)
            is_bonus_var = tk.BooleanVar(value=mt_dict.get("is_extra_credit",False)); widgets_row["is_bonus_var"] = is_bonus_var
            ttk.Checkbutton(self.list_frame, variable=is_bonus_var).grid(row=r_idx, column=4, padx=3)

            del_btn = ttk.Button(self.list_frame, text="X", command=lambda idx=r_idx-1: self.delete_mark_type_at_index(idx), width=3)
            del_btn.grid(row=r_idx, column=5, padx=3)
            self.mark_type_widgets.append(widgets_row) # Store the dict of vars/widgets for this row
        self.list_frame.grid_columnconfigure(1, weight=1) # Allow name entry to expand


    def add_mark_type(self):
        if len(self.mark_types_ref) >= MAX_CUSTOM_TYPES: # Or a specific limit for mark types
            messagebox.showwarning("Limit Reached", "Maximum number of mark types reached.", parent=self); return
        
        new_id_base = "custom_mark"
        new_id_suffix = 1
        while f"{new_id_base}_{new_id_suffix}" in (mt.get("id") for mt in self.mark_types_ref):
            new_id_suffix +=1
        
        new_mark = {
            "id": f"{new_id_base}_{new_id_suffix}", "name": "New Mark Type", "default_points": 0.0,
            "contributes_to_total": False, "is_extra_credit": False
        }
        self.mark_types_ref.append(new_mark)
        self.mark_types_changed = True
        self.populate_mark_types_ui()

    def delete_mark_type_at_index(self, index):
        if 0 <= index < len(self.mark_types_ref):
            # Check if it's a default one, prevent deletion (or handle carefully)
            # For now, allow deletion, user can reset to defaults
            if messagebox.askyesno("Confirm Delete", f"Delete mark type '{self.mark_types_ref[index]['name']}'?", parent=self):
                del self.mark_types_ref[index]
                self.mark_types_changed = True
                self.populate_mark_types_ui()
    
    def reset_to_defaults(self):
        if messagebox.askyesno("Confirm Reset", f"Reset all {self.item_type_name.lower()} to application defaults?", parent=self):
            self.mark_types_ref.clear()
            for default_item in self.default_mark_types_const:
                self.mark_types_ref.append(default_item.copy()) # Add copies
            self.mark_types_changed = True
            self.populate_mark_types_ui()

    def apply(self): # OK button
        # Update the list of dicts from the UI widgets
        updated_list = []
        for row_widgets in self.mark_type_widgets:
            # Get ID from label (it's not editable here, but needed for the dict)
            current_id = row_widgets["id_label"].cget("text")
            name = row_widgets["name_var"].get().strip()
            if not name:
                messagebox.showerror("Invalid Name", f"Mark type name for ID '{current_id}' cannot be empty.", parent=self)
                # To prevent dialog closing on error, simpledialog needs validate() to return false.
                # This is tricky here as apply() is called after validate().
                # For now, we'll allow it but it might lead to an empty name. Better: prevent empty.
                return # Or handle error state
            
            # Check for duplicate names before saving
            if any(item['name'] == name and item['id'] != current_id for item in updated_list):
                messagebox.showerror("Duplicate Name", f"Mark type name '{name}' is already used. Names must be unique.", parent=self)
                return

            updated_list.append({
                "id": current_id,
                "name": name,
                "default_points": row_widgets["points_var"].get(),
                "contributes_to_total": row_widgets["to_total_var"].get(),
                "is_extra_credit": row_widgets["is_bonus_var"].get()
            })
        
        # Check if actual changes were made before setting the flag
        if self.mark_types_ref != updated_list: # Simple list comparison
            self.mark_types_ref[:] = updated_list # Replace content of original list
            self.mark_types_changed = True
        
        # self.mark_types_changed flag is now set if there were modifications.

class ManageLiveSelectOptionsDialog(simpledialog.Dialog):
    def __init__(self, parent, current_options_list):
        self.current_options = [opt.copy() for opt in current_options_list] # Work on a copy
        self.options_changed_flag = False
        super().__init__(parent, "Manage 'Select' Mode Options for Live Homework")

    def body(self, master):
        self.main_frame = master
        ttk.Label(master, text="Define the buttons/options available in 'Select' mode for live homework sessions.", wraplength=350).pack(pady=5)
        
        button_frame = ttk.Frame(master); button_frame.pack(fill=tk.X, pady=5)
        ttk.Button(button_frame, text="Add New Option", command=self.add_option).pack(side=tk.LEFT, padx=5)

        self.list_frame = ttk.Frame(master); self.list_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.populate_options_ui()
        return self.main_frame

    def populate_options_ui(self):
        for widget in self.list_frame.winfo_children(): widget.destroy()
        
        self.option_entry_vars = [] # List of StringVars for names

        for r_idx, option_dict in enumerate(self.current_options):
            option_frame = ttk.Frame(self.list_frame); option_frame.pack(fill=tk.X, pady=2)
            
            name_var = tk.StringVar(value=option_dict.get("name","")); self.option_entry_vars.append(name_var)
            name_entry = ttk.Entry(option_frame, textvariable=name_var, width=30)
            name_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
            
            del_btn = ttk.Button(option_frame, text="Remove", command=lambda idx=r_idx: self.delete_option_at_index(idx), width=8)
            del_btn.pack(side=tk.LEFT, padx=5)
        
        if not self.current_options:
            ttk.Label(self.list_frame, text="No options defined. Click 'Add New Option'.").pack(pady=10)

    def add_option(self):
        if len(self.current_options) >= MAX_CUSTOM_TYPES: # Use a reasonable limit
            messagebox.showwarning("Limit Reached", "Maximum number of 'Select' options reached.", parent=self); return
        
        new_name = simpledialog.askstring("New Option", "Enter name for the new option (e.g., 'Signed', 'Returned Late'):", parent=self)
        if new_name and new_name.strip():
            new_name_clean = new_name.strip()
            if any(opt.get("name","").lower() == new_name_clean.lower() for opt in self.current_options):
                messagebox.showwarning("Duplicate Name", f"An option named '{new_name_clean}' already exists.", parent=self); return
            self.current_options.append({"name": new_name_clean})
            self.options_changed_flag = True
            self.populate_options_ui()

    def delete_option_at_index(self, index):
        if 0 <= index < len(self.current_options):
            if messagebox.askyesno("Confirm Delete", f"Delete option '{self.current_options[index]['name']}'?", parent=self):
                del self.current_options[index]
                self.options_changed_flag = True
                self.populate_options_ui()

    def apply(self):
        updated_options = []
        for name_var in self.option_entry_vars:
            name = name_var.get().strip()
            if not name:
                messagebox.showerror("Invalid Name", "Option names cannot be empty.", parent=self); return
            if any(opt.get("name","") == name for opt in updated_options): # Check for duplicates within the new list
                messagebox.showerror("Duplicate Name", f"Option name '{name}' is duplicated in the list. Names must be unique.", parent=self); return
            updated_options.append({"name": name})
        
        # Compare with original to see if actual changes occurred
        # Simple comparison might fail if order changed but content is same.
        # For now, if any entry var changed from initial or list length changed, consider it changed.
        # The self.options_changed_flag handles additions/deletions. This apply covers edits.
        if self.current_options != updated_options: # If direct edits changed things
            self.options_changed_flag = True
        self.current_options = updated_options # Finalize list


class HelpDialog(simpledialog.Dialog):
    def __init__(self, parent, app_version):
        self.app_version = app_version
        super().__init__(parent, f"Help & About - {APP_NAME}")

    def body(self, master):
        main_frame = ttk.Frame(master); main_frame.pack(expand=True, fill="both")
        
        header_label = ttk.Label(main_frame, text=f"{APP_NAME} - Version {self.app_version}", font=("", 14, "bold"))
        header_label.pack(pady=(10,5))
        
        notebook = ttk.Notebook(main_frame)
        
        # --- General Info Tab ---
        info_tab = ttk.Frame(notebook, padding=10)
        info_text_content = f"""Welcome to {APP_NAME}!

This application helps you track student behaviors, quiz scores, and homework completion in a visual classroom layout.

Key Features:
- Visual seating chart: Drag and drop students and furniture.
- Behavior Logging: Quickly log predefined or custom behaviors for students.
- Quiz Logging: Record quiz scores, including detailed marks.
- Homework Logging: Track homework status via manual logs or live sessions.
- Live Sessions: Conduct real-time quiz or homework checks directly on the seating chart.
- Customization: Define your own behaviors, homework types, mark types, and layout templates.
- Data Export: Export logs to Excel or CSV for reporting and analysis.
- Undo/Redo: Most actions can be undone and redone.
- Password Protection: Secure your application data.
- Data Backup & Restore: Keep your classroom data safe.

Navigation & Interaction Tips:
- Canvas Panning: Middle-click and drag to pan the seating chart.
- Canvas Zoom: Ctrl + Mouse Wheel to zoom in/out. Shift + Mouse Wheel for horizontal scroll.
- Multi-Select: Ctrl + Left-click to select multiple items.
- Context Menus: Right-click on items or the canvas for quick actions.
- Edit Mode: Toggle "Edit Mode (Resize)" to resize items by dragging their bottom-right corner.
- Combobox Scrolling: You can often use the mouse wheel to scroll through options in dropdown (combobox) lists, even if a scrollbar isn't visible.
- Data Files: Your classroom data is stored locally. Use 'File > Open Data Folder' to access these files. Regularly backup using 'File > Backup All Application Data'.
"""
        info_text = tk.Text(info_tab, wrap="word", height=20, width=70, relief=tk.FLAT, font=('Arial',11))
        info_text.insert("1.0", info_text_content)
        info_text.config(state="disabled")
        info_text.pack(pady=5, fill="both", expand=True)
        notebook.add(info_tab, text="General Info")

        # --- Modes Tab ---
        modes_tab = ttk.Frame(notebook, padding=10)
        modes_text_content = f"""Application Modes:

The application operates in three main modes, selectable from the top toolbar:

1. Behavior Mode:
   - Default mode for general classroom management.
   - Clicking a student allows logging of behaviors (e.g., "Talking", "Great Participation").
   - Recent behavior logs can be displayed on student boxes.

2. Quiz Mode:
   - Focused on quiz-related activities.
   - Clicking a student allows logging of a quiz score with detailed marks.
   - Live Quiz Session:
     - Start a "Class Quiz" session to mark students' answers (Correct/Incorrect/Skip) in real-time.
     - Scores are displayed live on student boxes.
     - Session data is logged upon ending the session.
   - Quiz Templates: Define reusable quiz structures (name, number of questions, default marks) in Settings.

3. Homework Mode (New!):
   - Dedicated to tracking homework.
   - Clicking a student allows:
     - Manual Logging: Log detailed homework status (e.g., "Done", "Not Done", "Signed") and optionally record marks/scores if enabled in settings.
     - Live Homework Session:
       - Start a "Homework Session" (e.g., "Daily Homework Check").
       - Session Mode (configurable in Settings > Homework):
         - "Yes/No" Mode: Quickly mark predefined homework types (e.g., "Reading Assignment: Yes/No", "Math Worksheet: Yes/No"). Define these types in Settings > Homework.
         - "Select" Mode: Choose from a list of predefined statuses (e.g., "Done", "Signed", "Missing"). Define these options in Settings > Homework.
       - Statuses are displayed live on student boxes.
       - Session data is logged upon ending.
   - Homework Templates: Define reusable homework assignments (name, number of items, default marks/statuses) in Settings.
   - Recent homework logs can be displayed on student boxes.

Switching Modes:
- If a Live Quiz or Live Homework session is active, you will be prompted to end or discard the session before switching modes or closing the application.
"""
        modes_text = tk.Text(modes_tab, wrap="word", height=20, width=70, relief=tk.FLAT, font=('Arial',11))
        modes_text.insert("1.0", modes_text_content)
        modes_text.config(state="disabled")
        modes_text.pack(pady=5, fill="both", expand=True)
        notebook.add(modes_tab, text="Application Modes")

        # --- Settings & Customization Tab ---
        settings_tab = ttk.Frame(notebook, padding=10)
        settings_text_content = f"""Settings & Customization:

The "Settings" dialog allows extensive customization:

- General: Autosave interval, grid snapping, student group feature toggle, max undo history.
- Student Boxes: Default appearance (size, colors, font), and Conditional Formatting rules (e.g., change box color if a student is in a specific group or has many recent incidents).
- Behavior & Quiz:
    - Recent Incidents Display: Control how behavior/quiz logs appear on student boxes.
    - Custom Behaviors: Add your own behavior types.
    - Behavior/Quiz Initials: Customize the initials displayed for behaviors/quizzes.
    - Quiz Mark Types: Define the categories for quiz marks (e.g., Correct, Incorrect, Bonus) and their properties.
    - Quiz Session Defaults: Default quiz name, number of questions for manual log.
    - Quiz Templates: Manage reusable quiz structures.
- Homework (New!):
    - Recent Homework Display: Control how homework logs appear on student boxes.
    - Custom Homework Log Options: Define statuses for manual homework logging (e.g., "Done", "Not Done", "Late").
    - Homework Log Initials: Customize initials for homework log entries.
    - Homework Mark Types: Define categories for homework marks (e.g., Complete, Incomplete, Effort Score) and their properties.
    - Live Homework Session:
        - Default session name.
        - Session Mode: Choose between "Yes/No" or "Select".
        - Custom Types for "Yes/No" Mode: Define the list of homeworks to check (e.g., "Reading", "Math").
        - Options for "Select" Mode: Define the buttons available (e.g., "Done", "Signed").
    - Enable Detailed Marks: Toggle whether to log detailed scores/marks for manual homework entries.
    - Homework Templates: Manage reusable homework assignment structures.
- Data & Export: Default options for Excel export (separate sheets, include summary). Enable experimental Excel autosave.
- Security: Set an application password, configure password prompts (on open, for sensitive actions), and auto-lock settings.

Remember to save your settings after making changes!
"""
        settings_text = tk.Text(settings_tab, wrap="word", height=20, width=70, relief=tk.FLAT, font=('Arial',11))
        settings_text.insert("1.0", settings_text_content)
        settings_text.config(state="disabled")
        settings_text.pack(pady=5, fill="both", expand=True)
        notebook.add(settings_tab, text="Settings & Customization")
        
         # --- Settings & Customization Tab ---
        remarks_tab = ttk.Frame(notebook, padding=10)
        remarks_text_content = f"""
        Please note the following:
        
- The 'Export Layout as Image' function currently only takes a screenshot of the entire window. Hopefully soon I will be able to make it functional.
- I am still working on the homework logging and exporting - so expect to see more features, and don't be surprised if something doesn't work as expected.
- The Conditional Formatting feature currently doesn't work for quizzes. If you have two rules that apply to the same student box, the first one will take precedent.
    -Yaakov Maimon
"""        
        remarks_text = tk.Text(remarks_tab, wrap="word", height=20, width=70, relief=tk.FLAT, font=('Arial',11))
        remarks_text.insert("1.0", remarks_text_content)
        remarks_text.config(state="disabled")
        remarks_text.pack(pady=5, fill="both", expand=True)
        notebook.add(remarks_tab, text="Remarks & Notices")
        
        # --- Feedback & Contact Tab ---
        feedback_tab = ttk.Frame(notebook, padding=10)
        feedback_text_content = f"""                                                    Contact the developer:
        
Yaakov Maimon
Email: yaakovmaimon592@gmail.com
Phone: +1 206-750-5557

"""
        feedback_text = tk.Text(feedback_tab, wrap="word", height=20, width=70, relief=tk.FLAT, font=('Arial',11))
        feedback_text.insert("1.0", feedback_text_content)
        feedback_text.config(state="disabled")
        feedback_text.pack(pady=5, fill="both", expand=True)
        notebook.add(feedback_tab, text="Feedback")
        

        notebook.pack(expand=True, fill="both", pady=5)
        return main_frame # No specific focus needed

    def buttonbox(self): # Override to only show an OK button
        box = ttk.Frame(self)
        ok_button = ttk.Button(box, text="OK", width=10, command=self.ok, default=tk.ACTIVE)
        ok_button.pack(pady=5)
        self.bind("<Return>", self.ok)
        self.bind("<Escape>", self.ok)
        box.pack()

# --- Main Execution ---
if __name__ == "__main__":
    root = tk.Tk()
    # Apply a theme if available and desired
    try:
        # Examples: 'clam', 'alt', 'default', 'classic'
        # Some themes might require python -m tkinter to see available ones on your system
        # Or use ttkthemes for more options: from ttkthemes import ThemedTk
        # root = ThemedTk(theme="arc") # Example using ttkthemes
        #style = ttk.Style(root)
        #available_themes = style.theme_names() # ('winnative', 'clam', 'alt', 'default', 'classic', 'vista', 'xpnative') on Windows
        # print("Available themes:", available_themes)
        sv_ttk.set_theme("Light")
        #print("Startup")
        #if 'vista' in available_themes: style.theme_use('vista')
        #elif 'xpnative' in available_themes: style.theme_use('xpnative')

    except Exception as e_theme:
        print(f"Could not apply custom theme: {e_theme}")

    app = SeatingChartApp(root)
    
    try:
        t = threading.Thread(target=darkdetect.listener, args=(app.theme_auto, ))
        t.daemon = True
        t.start()
    except: pass
    root.mainloop()