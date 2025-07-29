import tkinter as tk
from tkinter import ttk, simpledialog, messagebox, filedialog, colorchooser

import os
import sys

from datetime import datetime, timedelta, date as datetime_date
from openpyxl import load_workbook




# def listener(callback: typing.Callable[[str], None]) -> None: ...

# TODO: make conditional formatting work by quizzes. add thing for homework also.
try:
    from tkcalendar import DateEntry
    from tkcalendar import DateEntryCustom
except ImportError:
    DateEntry = None
    #DateEntryCustom = None
    print("Warning: tkcalendar library not found. Date pickers in export filter will be simple text entries.")
    print("Consider installing it: pip install tkcalendar")



# --- Application Constants ---
APP_NAME = "BehaviorLogger"
APP_VERSION = "v57.0" # Version incremented
CURRENT_DATA_VERSION_TAG = "v10" # Incremented for new homework/marks features

# --- Default Configuration ---
DEFAULT_STUDENT_BOX_WIDTH = 130
DEFAULT_STUDENT_BOX_HEIGHT = 80
MIN_STUDENT_BOX_WIDTH = 60
MIN_STUDENT_BOX_HEIGHT = 40
REBBI_DESK_WIDTH = 200
REBBI_DESK_HEIGHT = 100

DEFAULT_FONT_FAMILY = "TkDefaultFont"
DEFAULT_FONT_SIZE = 10
DEFAULT_FONT_COLOR = "black"
DEFAULT_BOX_FILL_COLOR = "skyblue"
DEFAULT_BOX_OUTLINE_COLOR = "blue"
DEFAULT_QUIZ_SCORE_FONT_COLOR = "darkgreen"
DEFAULT_QUIZ_SCORE_FONT_STYLE_BOLD = True
DEFAULT_HOMEWORK_SCORE_FONT_COLOR = "purple" # New for homework scores
DEFAULT_HOMEWORK_SCORE_FONT_STYLE_BOLD = True # New
GROUP_COLOR_INDICATOR_SIZE = 12
DEFAULT_THEME = "System"
THEME_LIST = ["Light", "Dark", "System"]

DRAG_THRESHOLD = 5
DEFAULT_GRID_SIZE = 20
MAX_UNDO_HISTORY_DAYS = 90
LAYOUT_COLLISION_OFFSET = 5
RESIZE_HANDLE_SIZE = 10 # World units for resize handle

# --- Path Handling ---
def get_app_data_path(filename):
    try:
        # Determine base path based on whether the app is frozen (packaged) or running from script
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            # Running as a PyInstaller bundle
            if os.name == 'win32': # Windows
                base_path = os.path.join(os.getenv('APPDATA'), APP_NAME)
            elif sys.platform == 'darwin': # macOS
                base_path = os.path.join(os.path.expanduser('~'), 'Library', 'Application Support', APP_NAME)
            else: # Linux and other Unix-like
                xdg_config_home = os.getenv('XDG_CONFIG_HOME')
                if xdg_config_home:
                    base_path = os.path.join(xdg_config_home, APP_NAME)
                else:
                    base_path = os.path.join(os.path.expanduser('~'), '.config', APP_NAME)
        else:
            # Running as a script, use the script's directory
            base_path = os.path.dirname(os.path.abspath(__file__))

        # Create the base directory if it doesn't exist
        if not os.path.exists(base_path):
            os.makedirs(base_path, exist_ok=True)
        return os.path.join(base_path, filename)
    except Exception as e:
        # Fallback to current working directory if standard paths fail
        print(f"Warning: Could not determine standard app data path due to {e}. Using current working directory as fallback.")
        fallback_path = os.path.join(os.getcwd(), APP_NAME) # Create a subfolder in CWD
        if not os.path.exists(fallback_path):
             os.makedirs(fallback_path, exist_ok=True)
        return os.path.join(fallback_path, filename)

DATA_FILE_PATTERN = f"classroom_data_{CURRENT_DATA_VERSION_TAG}.json"
CUSTOM_BEHAVIORS_FILE_PATTERN = f"custom_behaviors_{CURRENT_DATA_VERSION_TAG}.json"
# NEW: For customizable homework types like "Reading Assignment", "Worksheet"
CUSTOM_HOMEWORK_TYPES_FILE_PATTERN = f"custom_homework_types_{CURRENT_DATA_VERSION_TAG}.json" 
# RENAMED: For customizable homework statuses like "Done", "Not Done"
CUSTOM_HOMEWORK_STATUSES_FILE_PATTERN = f"custom_homework_statuses_{CURRENT_DATA_VERSION_TAG}.json" 
AUTOSAVE_EXCEL_FILE_PATTERN = f"autosave_log_{CURRENT_DATA_VERSION_TAG}.xlsx" # Renamed for clarity
LAYOUT_TEMPLATES_DIR_NAME = "layout_templates"
STUDENT_GROUPS_FILE_PATTERN = f"student_groups_{CURRENT_DATA_VERSION_TAG}.json"
QUIZ_TEMPLATES_FILE_PATTERN = f"quiz_templates_{CURRENT_DATA_VERSION_TAG}.json"
HOMEWORK_TEMPLATES_FILE_PATTERN = f"homework_templates_{CURRENT_DATA_VERSION_TAG}.json" # New

DATA_FILE = get_app_data_path(DATA_FILE_PATTERN)
CUSTOM_BEHAVIORS_FILE = get_app_data_path(CUSTOM_BEHAVIORS_FILE_PATTERN)
CUSTOM_HOMEWORK_TYPES_FILE = get_app_data_path(CUSTOM_HOMEWORK_TYPES_FILE_PATTERN) # NEW
CUSTOM_HOMEWORK_STATUSES_FILE = get_app_data_path(CUSTOM_HOMEWORK_STATUSES_FILE_PATTERN) # RENAMED
AUTOSAVE_EXCEL_FILE = get_app_data_path(AUTOSAVE_EXCEL_FILE_PATTERN)
LAYOUT_TEMPLATES_DIR = get_app_data_path(LAYOUT_TEMPLATES_DIR_NAME)
STUDENT_GROUPS_FILE = get_app_data_path(STUDENT_GROUPS_FILE_PATTERN)
QUIZ_TEMPLATES_FILE = get_app_data_path(QUIZ_TEMPLATES_FILE_PATTERN)
HOMEWORK_TEMPLATES_FILE = get_app_data_path(HOMEWORK_TEMPLATES_FILE_PATTERN) # New
LOCK_FILE_PATH = get_app_data_path(f"{APP_NAME}.lock") # Lock file

if not os.path.exists(LAYOUT_TEMPLATES_DIR):
    os.makedirs(LAYOUT_TEMPLATES_DIR, exist_ok=True)

DEFAULT_BEHAVIORS_LIST = [
    "Talking", "Off Task", "Out of Seat", "Uneasy", "Placecheck",
    "Great Participation", "Called On", "Complimented", "Fighting", "Other"
]


# for manual detailed logging, i would want these | currently used for sessions (yes/no)
DEFAULT_HOMEWORK_TYPES_LIST = [ # For live session "Yes/No" mode AND now for manual logging
    "Reading Assignment", "Worksheet", "Math Problems", "Project Work", "Study for Test"
]

# these would only be for behavior-like logging (not detailed)
DEFAULT_HOMEWORK_LOG_BEHAVIORS = [ # For manual homework logging (like behavior logging)
    "Done", "Not Done", "Partially Done", "Signed", "Returned", "Late", "Excellent Work"
]


# Make these be used for the select AND Yes/No Live Homework session modes.
DEFAULT_HOMEWORK_SESSION_BUTTONS = [ # For live session "Select" mode
    {"name": "Done"}, {"name": "Not Done"}, {"name": "Signed"}, {"name": "Returned"}
]

DEFAULT_HOMEWORK_SESSION_BUTTONS2 = [ # For live session "Select" mode
    "Done", "Not Done", "Signed", "Returned"
]


DEFAULT_HOMEWORK_STATUSES = [
    "Done", "Not Done", "Partially Done", "Signed", "Returned", "Late", "Excellent Work"
]


DEFAULT_GROUP_COLORS = ["#FFADAD", "#FFD6A5", "#FDFFB6", "#CAFFBF", "#9BF6FF", "#A0C4FF", "#BDB2FF", "#FFC6FF", "#E0E0E0"]

DEFAULT_QUIZ_MARK_TYPES = [
    {"id": "mark_correct", "name": "Correct", "contributes_to_total": True, "is_extra_credit": False, "default_points": 1},
    {"id": "mark_incorrect", "name": "Incorrect", "contributes_to_total": True, "is_extra_credit": False, "default_points": 0},
    {"id": "mark_partial", "name": "Partial Credit", "contributes_to_total": True, "is_extra_credit": False, "default_points": 0.5},
    {"id": "extra_credit", "name": "Bonus", "contributes_to_total": False, "is_extra_credit": True, "default_points": 1}
]
DEFAULT_HOMEWORK_MARK_TYPES = [ # New for homework marks
    {"id": "hmark_complete", "name": "Complete", "default_points": 10},
    {"id": "hmark_incomplete", "name": "Incomplete", "default_points": 5},
    {"id": "hmark_notdone", "name": "Not Done", "default_points": 0},
    {"id": "hmark_effort", "name": "Effort Score (1-5)", "default_points": 3} # Example
]

MAX_CUSTOM_TYPES = 90 # Max for custom behaviors, homeworks, mark types


# --- Dialog Classes ---
class PasswordPromptDialog(simpledialog.Dialog):
    def __init__(self, parent, title, prompt, password_manager_instance):
        self.prompt = prompt
        self.password_manager = password_manager_instance
        self.result = False # True if password correct or not set, False otherwise
        super().__init__(parent, title)

    def body(self, master):
        ttk.Label(master, text=self.prompt, wraplength=280).pack(pady=5)
        self.password_entry = ttk.Entry(master, show="*", width=30)
        self.password_entry.pack(pady=5)
        self.status_label = ttk.Label(master, text="", foreground="red")
        self.status_label.pack(pady=2)
        return self.password_entry # initial focus

    def buttons(self):
        box = ttk.Frame(self)
        ttk.Button(box, text="OK", width=10, command=self.ok, default=tk.ACTIVE).pack(side=tk.LEFT, padx=5, pady=5)
        ttk.Button(box, text="Cancel", width=10, command=self.cancel).pack(side=tk.LEFT, padx=5, pady=5)
        self.bind("<Return>", lambda e: self.ok())
        self.bind("<Escape>", lambda e: self.cancel())
        box.pack()

    def apply(self):
        password_attempt = self.password_entry.get()
        if self.password_manager.unlock_application(password_attempt): # unlock_application handles recovery pw too
            self.result = True
        else:
            self.status_label.config(text="Incorrect password.")
            self.password_entry.focus_set()
            self.password_entry.select_range(0, tk.END)
            self.result = False # Explicitly set to false, stay open
            # To prevent dialog from closing on incorrect password, we don't call super().ok() or cancel
            # Instead, the dialog stays open. The caller of this dialog will check self.result.
            # This means the typical simpledialog auto-close on OK needs to be managed carefully.
            # For this use case, we want it to close only on success or cancel.
            # The `ok` method will be called. If result is False, the dialog won't close.
            # We need to ensure `ok` can prevent closing.
            # A simple way is to override `ok` more directly or ensure `validate` fails.
            # simpledialog.Dialog closes if validate() returns true.

    def ok(self, event=None): # Override ok to control closing
        if not self.validate(): # validate calls apply
            self.password_entry.focus_set()
            return # Don't close if validation (password check) fails
        self.withdraw()
        self.update_idletasks()
        # apply has already been called by validate
        self.parent.focus_set() # Give focus back to parent
        self.destroy()


    def validate(self): # This is called by simpledialog's ok method
        self.apply() # Calls our apply which sets self.result
        return self.result # If True, dialog closes. If False, stays open due to logic in apply.

class AddEditStudentDialog(simpledialog.Dialog):
    # ... (similar to v51, but needs group dropdown to be populated from app.student_groups)
    def __init__(self, parent, title, student_data=None, app=None):
        self.student_data = student_data
        self.app = app # To access student_groups
        self.result = None
        super().__init__(parent, title)

    def body(self, master):
        ttk.Label(master, text="First Name:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        self.fn_var = tk.StringVar(value=self.student_data["first_name"] if self.student_data else "")
        self.fn_entry = ttk.Entry(master, textvariable=self.fn_var, width=30); self.fn_entry.grid(row=0, column=1, padx=5, pady=2)

        ttk.Label(master, text="Last Name:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        self.ln_var = tk.StringVar(value=self.student_data["last_name"] if self.student_data else "")
        self.ln_entry = ttk.Entry(master, textvariable=self.ln_var, width=30); self.ln_entry.grid(row=1, column=1, padx=5, pady=2)

        ttk.Label(master, text="Nickname (Optional):").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        self.nick_var = tk.StringVar(value=self.student_data.get("nickname", "") if self.student_data else "")
        self.nick_entry = ttk.Entry(master, textvariable=self.nick_var, width=30); self.nick_entry.grid(row=2, column=1, padx=5, pady=2)

        ttk.Label(master, text="Gender:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=2)
        self.gender_var = tk.StringVar(value=self.student_data.get("gender", "Boy") if self.student_data else "Boy")
        gender_frame = ttk.Frame(master)
        ttk.Radiobutton(gender_frame, text="Boy", variable=self.gender_var, value="Boy").pack(side=tk.LEFT)
        ttk.Radiobutton(gender_frame, text="Girl", variable=self.gender_var, value="Girl").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(gender_frame, text="Other", variable=self.gender_var, value="Other").pack(side=tk.LEFT)
        gender_frame.grid(row=3, column=1, sticky=tk.W, padx=5, pady=2)

        if self.app and self.app.settings.get("student_groups_enabled", True):
            ttk.Label(master, text="Group:").grid(row=4, column=0, sticky=tk.W, padx=5, pady=2)
            self.group_var = tk.StringVar()
            group_options = {"NONE_GROUP_SENTINEL": "No Group"} # Sentinel for no group
            for gid, gdata in sorted(self.app.student_groups.items(), key=lambda item: item[1]['name']):
                group_options[gid] = gdata['name']
            
            self.group_combobox = ttk.Combobox(master, textvariable=self.group_var, values=list(group_options.values()), state="readonly", width=28)
            self.group_combobox_map = {name: gid for gid, name in group_options.items()} # Map display name back to ID
            
            current_group_id = self.student_data.get("group_id") if self.student_data else None
            if current_group_id and current_group_id in self.app.student_groups:
                self.group_var.set(self.app.student_groups[current_group_id]['name'])
            else:
                self.group_var.set("No Group")
            self.group_combobox.grid(row=4, column=1, padx=5, pady=2, sticky=tk.W)
            self.group_combobox.bind("<MouseWheel>", lambda event: "break") # Prevent main canvas scroll

        return self.fn_entry

    def apply(self):
        first_name = self.fn_var.get().strip()
        last_name = self.ln_var.get().strip()
        nickname = self.nick_var.get().strip()
        gender = self.gender_var.get()
        group_id_selection = None
        if self.app and self.app.settings.get("student_groups_enabled", True) and hasattr(self, 'group_var'):
            selected_group_name = self.group_var.get()
            group_id_selection = self.group_combobox_map.get(selected_group_name)
            if group_id_selection == "NONE_GROUP_SENTINEL": group_id_selection = None

        if first_name and last_name:
            self.result = (first_name, last_name, nickname, gender, group_id_selection)
        else:
            messagebox.showwarning("Missing Information", "First Name and Last Name are required.", parent=self)
            self.result = None # Stay open

class AddFurnitureDialog(simpledialog.Dialog):
    # ... (same as v51)
    def __init__(self, parent, title, furniture_data=None):
        self.furniture_data = furniture_data
        self.result = None
        super().__init__(parent, title)
    def body(self, master):
        ttk.Label(master, text="Name:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        self.name_var = tk.StringVar(value=self.furniture_data["name"] if self.furniture_data else "Desk")
        self.name_entry = ttk.Entry(master, textvariable=self.name_var); self.name_entry.grid(row=0, column=1, padx=5, pady=2)
        ttk.Label(master, text="Type:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        self.type_var = tk.StringVar(value=self.furniture_data["type"] if self.furniture_data else "Rebbi's Desk")
        self.type_entry = ttk.Entry(master, textvariable=self.type_var); self.type_entry.grid(row=1, column=1, padx=5, pady=2)
        ttk.Label(master, text="Width:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        self.width_var = tk.IntVar(value=self.furniture_data["width"] if self.furniture_data else REBBI_DESK_WIDTH)
        self.width_spinbox = ttk.Spinbox(master, from_=20, to=1000, textvariable=self.width_var, width=5); self.width_spinbox.grid(row=2, column=1, sticky=tk.W, padx=5, pady=2)
        ttk.Label(master, text="Height:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=2)
        self.height_var = tk.IntVar(value=self.furniture_data["height"] if self.furniture_data else REBBI_DESK_HEIGHT)
        self.height_spinbox = ttk.Spinbox(master, from_=20, to=1000, textvariable=self.height_var, width=5); self.height_spinbox.grid(row=3, column=1, sticky=tk.W, padx=5, pady=2)
        return self.name_entry
    def apply(self):
        name = self.name_var.get().strip(); item_type = self.type_var.get().strip()
        width = self.width_var.get(); height = self.height_var.get()
        if name and item_type and width > 0 and height > 0: self.result = (name, item_type, width, height)
        else: messagebox.showwarning("Invalid Input", "Name, Type, Width, and Height are required and must be positive.", parent=self); self.result = None

class BehaviorDialog(simpledialog.Dialog):
    # ... (same as v51)
    def __init__(self, parent, title, all_behaviors, custom_behaviors, initial_value=None): # Added initial_value
        self.all_behaviors = all_behaviors
        self.custom_behaviors = custom_behaviors
        self.result = None
        self.selected_behavior_var = tk.StringVar()
        if initial_value and initial_value in self.all_behaviors: # Pre-select if valid
            self.selected_behavior_var.set(initial_value)
        elif initial_value: # If provided but not in list, allow it to be typed (if combobox was used) or just store it for apply
             self.selected_behavior_var.set(initial_value) # For button-based, it will be the one "pressed"

        super().__init__(parent, title)
    def body(self, master):
        
        master.grid_columnconfigure(0, weight=1) # Allow master frame to expand
        top_label = ttk.Label(master, text="Select Behavior:")
        top_label.grid(row=0, column=0, columnspan=6, sticky="nw", pady=2, padx=5)

        # Frame for buttons with scrollbar
        button_frame_outer = ttk.Frame(master)
        button_frame_outer.grid(row=1, column=0, columnspan=6, sticky="nsew", pady=(5,10), )
        master.grid_rowconfigure(1, weight=1) # Allow button area to expand

        theheight=0
        for i in range(0,len(self.all_behaviors),6): theheight += 115
        
        btn_canvas = tk.Canvas(button_frame_outer, borderwidth=0,width=1000,height= theheight)
        btn_scrollbar = ttk.Scrollbar(button_frame_outer, orient="vertical", command=btn_canvas.yview)
        scrollable_frame_for_buttons = ttk.Frame(btn_canvas)

        scrollable_frame_for_buttons.bind("<Configure>", lambda e: btn_canvas.configure(scrollregion=btn_canvas.bbox("all")))
        btn_canvas.create_window((0, 0), window=scrollable_frame_for_buttons, anchor="nw")
        btn_canvas.configure(yscrollcommand=btn_scrollbar.set)

        btn_canvas.pack(side="left", fill="both", expand=True)
        btn_scrollbar.pack(side="right", fill="y")

        # Populate buttons
        row_idx, col_idx = 0, 0
        max_cols = 6 # Adjust number of columns as desired
        button_width_chars = 14 # Approximate characters
        button_height_lines = 5 # Approximate lines
        button_width = 17
        button_height = 5
        wraplength=button_width_chars*7
        button_font = ('TkDefaultFont', 15)
        b = ttk.Style(); b.configure("Keypad.TButton", font=button_font, wraplength=button_width_chars*10,justify='center',padding=(0))
        for i, behavior_text in enumerate(self.all_behaviors):
            btn = ttk.Button(scrollable_frame_for_buttons, text=behavior_text,
                            padding=(0,19),# Rough estimate for pixel width
                            width=button_width_chars, style="Keypad.TButton",
                            command=lambda n=behavior_text: self.behavior_selected(n))
            btn.grid(row=row_idx, column=col_idx, sticky="nsew", padx=2, pady=3)
            scrollable_frame_for_buttons.grid_columnconfigure(col_idx, weight=1) # Allow columns to expand
            col_idx += 1
            if col_idx >= max_cols:
                col_idx = 0
                row_idx += 1
        if col_idx != 0: # Ensure last row also configures row weight if partially filled
             scrollable_frame_for_buttons.grid_rowconfigure(row_idx, weight=1)
        

        comment_label = ttk.Label(master, text="Additional Comment (Optional):")
        comment_label.grid(row=2,column=0,columnspan=6, sticky="sw", pady=(8,2), padx=5)
        self.comment_text_widget = tk.Text(master, width=50, height=4, wrap=tk.WORD)
        self.comment_text_widget.grid(row=3, column=0,columnspan=6, pady=(0,5), padx=5, sticky="sew")
        master.grid_rowconfigure(3, weight=0) # Comment text doesn't expand as much

        # Set initial focus if desired, e.g., to the first button or comment box
        # For now, default Tk focus handling.
        return self.comment_text_widget # Or another widget for initial focus


        """        
        
        ttk.Label(master, text="Select Behavior:").pack(pady=5)
        self.behavior_var = tk.StringVar()
        self.behavior_combobox = ttk.Combobox(master, textvariable=self.behavior_var, values=self.all_behaviors, width=30)
        if self.all_behaviors: self.behavior_combobox.set(self.all_behaviors[0])
        self.behavior_combobox.pack(pady=5)
        self.behavior_combobox.bind("<MouseWheel>", lambda event: "break")
        ttk.Label(master, text="Comment (Optional):").pack(pady=5)
        self.comment_entry = ttk.Entry(master, width=33); self.comment_entry.pack(pady=5)
        return self.behavior_combobox
        """
    
    
    
    def behavior_selected(self, behavior_name):
        self.selected_behavior_var.set(behavior_name)
        self.ok() # Trigger apply and close

    def buttonbox(self): # Standard OK/Cancel are usually handled by simpledialog.Dialog
        # We only need a cancel button as OK is triggered by behavior selection
        cancel_button_frame = ttk.Frame(self)
        cancel_button_frame.pack(fill=tk.X, padx=5, pady=(0,5)) # Pad bottom
        ttk.Button(cancel_button_frame, text="Cancel", width=10, command=self.cancel).pack(side=tk.RIGHT, padx=5)
        self.bind("<Escape>", self.cancel)

    def apply(self):
        behavior = self.selected_behavior_var.get()
        comment = self.comment_text_widget.get("1.0", tk.END).strip()
        if not behavior: # Should not happen if ok() is only called on selection
            messagebox.showwarning("Input Error", "Please select a behavior.", parent=self)
            self.result = None
            return
        self.result = (behavior, comment)
    
    
    #def apply(self):
    #    behavior = self.behavior_var.get().strip(); comment = self.comment_entry.get().strip()
    #    if behavior: self.result = (behavior, comment)
    #    else: messagebox.showwarning("Input Required", "Please select or enter a behavior.", parent=self); self.result = None

class ManualHomeworkLogDialog(simpledialog.Dialog):
    def __init__(self, parent, title, all_homework_types, custom_homework_types, log_marks_enabled, homework_mark_types, homework_templates, app, initial_homework_name=None, initial_num_items=None):
        self.all_homework_types = all_homework_types # List of strings (for the combobox)
        self.custom_homework_types = custom_homework_types
        self.log_marks_enabled = log_marks_enabled
        self.homework_mark_types = homework_mark_types
        self.homework_templates = homework_templates
        self.app = app
        self.result = None
        self.mark_entry_vars = {}
        self.initial_homework_name_passed = initial_homework_name # Store for use in body
        self.initial_num_items_passed = initial_num_items # Store for use in body
        super().__init__(parent, title)

    def body(self, master):
        main_frame = ttk.Frame(master); main_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=True, padx=10, pady=10)

        # Homework Type / Template Selection
        type_frame = ttk.LabelFrame(main_frame, text="Homework Assignment"); type_frame.pack(pady=5, fill=tk.X)
        ttk.Label(type_frame, text="Name/Type:").pack(side=tk.TOP, padx=5, anchor='w')
        self.homework_type_var = tk.StringVar()
        
        # MODIFICATION: Populate combobox with homework types and template names
        combined_options = sorted(list(set(self.all_homework_types + [tpl['name'] for tpl_id, tpl in self.homework_templates.items()])))
        
        self.homework_type_combobox = ttk.Combobox(type_frame, textvariable=self.homework_type_var, values=combined_options, width=40)

        # Set initial value for combobox
        if self.initial_homework_name_passed and self.initial_homework_name_passed in combined_options:
            self.homework_type_combobox.set(self.initial_homework_name_passed)
        elif self.initial_homework_name_passed: # If it's a custom name not in templates/defaults yet
            self.homework_type_combobox.set(self.initial_homework_name_passed) # Allow typing it
        elif combined_options: # Fallback to first option if initial not found or not provided
            self.homework_type_combobox.set(combined_options[0])

        self.homework_type_combobox.pack(side=tk.LEFT, padx=5, pady=(0,5), fill=tk.X, expand=True)
        self.homework_type_combobox.bind("<<ComboboxSelected>>", self.on_template_select)
        self.homework_type_combobox.bind("<MouseWheel>", lambda event: "break")

        # Number of Items
        self.num_items_frame = ttk.Frame(main_frame) # Will be packed by on_template_select if needed
        ttk.Label(self.num_items_frame, text="Number of Items/Questions:").pack(side=tk.LEFT, padx=5)
        self.num_items_var = tk.StringVar(value=str(self.initial_num_items_passed if self.initial_num_items_passed is not None else self.app.settings.get("default_homework_items_for_yes_no_mode", 5)))
        self.num_items_spinbox = ttk.Spinbox(self.num_items_frame, from_=1, to=200, textvariable=self.num_items_var, width=5)
        self.num_items_spinbox.pack(side=tk.LEFT, padx=5)


        theheight=0
        for i in range(0,len(combined_options),4): theheight += 75
        btn_canvas = tk.Canvas(type_frame, borderwidth=0,width=660,height= theheight)
        btn_scrollbar = ttk.Scrollbar(type_frame, orient="vertical", command=btn_canvas.yview)
        scrollable_frame_for_buttons = ttk.Frame(btn_canvas)

        scrollable_frame_for_buttons.bind("<Configure>", lambda e: btn_canvas.configure(scrollregion=btn_canvas.bbox("all")))


        # Marks Frame (conditionally packed by on_template_select or if log_marks_enabled)
        self.marks_widgets_frame = ttk.LabelFrame(main_frame, text="Marks Details")
        self.marks_widgets_frame.pack(pady=10, padx=5, fill=tk.BOTH, expand=True)
        

        # Marks Frame (if enabled)
        self.marks_widgets_frame = ttk.LabelFrame(main_frame, text="Marks Details") # Packed later
        if self.log_marks_enabled and self.homework_mark_types:
            
            #self.homework_type_combobox = ttk.Combobox(type_frame, textvariable=self.homework_type_var, values=combined_options, width=30, state="readonly")
            
            
            #if combined_options: self.homework_type_combobox.set(combined_options[0])
            #self.homework_type_combobox.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
            
            self.homework_type_combobox.bind("<<ComboboxSelected>>", self.on_template_select)
            self.num_items_frame.pack(pady=5, fill=tk.X)
            self.marks_widgets_frame.pack(pady=10, padx=5, fill=tk.BOTH, expand=True)
            self.target_entry2 = []
            self.target_entry3 = []
            self.cte = -1
            cols = 2 # Two columns for marks
            current_col = 0; current_row_marks = 0
            for i, mark_type in enumerate(self.homework_mark_types):
                mark_id = mark_type["id"]; mark_name = mark_type["name"]
                ttk.Label(self.marks_widgets_frame, text=f"{mark_name}:").grid(row=current_row_marks, column=current_col*2, sticky=tk.W, padx=5, pady=3)
                var = tk.StringVar()
                # Try to set default value from mark_type if it makes sense (e.g., for non-point entries or typical scores)
                # For point-based ones, usually user input is expected.
                if "default_value" in mark_type : var.set(str(mark_type["default_value"]))

                entry = ttk.Entry(self.marks_widgets_frame, textvariable=var, width=8)
                entry.grid(row=current_row_marks, column=current_col*2 + 1, sticky=tk.EW, padx=5, pady=3)
                self.mark_entry_vars[mark_id] = var
                entry.bind("<FocusIn>", lambda x=i,x2=i: self.set_numpad(x,x2))
                self.target_entry2.append(var)
                self.target_entry3.append(mark_type["name"])
                self.cte = i
                self.tte = i
                current_col += 1
                if current_col >= cols: current_col = 0; current_row_marks +=1
            #print(main_frame.slaves)
            self.keypad_frame = ttk.Frame(master, relief='sunken',borderwidth=10)
            self.keypad_frame.pack(side=tk.RIGHT)
            self.keypad_frame.grid_propagate(True)
            self.allow_decimal = True
            self.mark_vars2 = {} # {mark_type_id: tk.StringVar()}
            for mt in self.mark_vars2:
                self.mark_vars2[mt["id"]] = tk.StringVar()
            
            for i, mt_config in enumerate(self.mark_vars2):
                label_text = mt_config["name"]
                if mt_config.get("is_extra_credit"): label_text += " (Bonus)"
            #print(self.mark_vars2)
            self.target_entry = self.mark_entry_vars[mark_type["id"]]
            self.target_entry_name = mark_type["name"]
            
            self.target_name_label = ttk.Label(self.keypad_frame,text=self.target_entry_name)
            self.target_name_label.grid(column=0,row=0,columnspan=9)
            
            buttons = [('7',1,0),('8',1,1),('9',1,2),('4',2,0),('5',2,1),('6',2,2),('1',3,0),('2',3,1),('3',3,2),('0',4,0),('.',4,1) if self.allow_decimal else (' ',4,1),('⌫',4,2),('/',2,3)]
            button_font = ('TkDefaultFont', 15)
            self._next_entry()
            for i in range(4): self.rowconfigure(i, minsize=2000,weight=1)
            for i in range(4): self.columnconfigure(i, weight=1)
            for (text, r, c) in buttons:
                if text == ' ': continue
                action = lambda x=text: self._on_press(x)
                ttk.Button(self.keypad_frame, text=text, command=action, style="Keypad.TButton",padding=17,width=5).grid(row=r, column=c, padx=1, pady=1, sticky="nsew")
            ttk.Button(self.keypad_frame, text="Clear", command=self._clear_entry, style="Keypad.TButton").grid(row=1, column=3, padx=1, pady=1, sticky="nsew")
            ttk.Button(self.keypad_frame, text='Next', command=self._next_entry, style="Keypad.TButton").grid(row=4,column=3,padx=1,pady=1,sticky='nsew')
            ttk.Button(self.keypad_frame, text='Previous', command=self._previous_entry, style="Keypad.TButton").grid(row=3,column=3,padx=1,pady=1,sticky='nsew')
            s = ttk.Style(); s.configure("Keypad.TButton", font=button_font, padding=(5,10),height=100)
        
            
                
            

            #print(self.target_entry_name)
            
            for i in range(cols*2): # Configure columns to expand
                 self.marks_widgets_frame.grid_columnconfigure(i, weight=1 if i%2==1 else 0)
            




        else:
            btn_canvas = tk.Canvas(type_frame, borderwidth=0,width=1000,height=200)
            btn_scrollbar = ttk.Scrollbar(type_frame, orient="vertical", command=btn_canvas.yview)
            scrollable_frame_for_buttons = ttk.Frame(btn_canvas)

            scrollable_frame_for_buttons.bind("<Configure>", lambda e: btn_canvas.configure(scrollregion=btn_canvas.bbox("all")))
            btn_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            btn_canvas.pack(side=tk.BOTTOM)
            btn_canvas.create_window((0, 0), window=scrollable_frame_for_buttons, anchor="nw")
            btn_canvas.configure(yscrollcommand=btn_scrollbar.set)
            row_idx, col_idx = 0, 0
            max_cols = 4 # Adjust number of columns as desired
            button_width_chars = 14 # Approximate characters
            button_height_lines = 5 # Approximate lines
            button_width = 17
            button_height = 5
            wraplength=button_width_chars*7
            button_font = ('TkDefaultFont', 15)
            b = ttk.Style(); b.configure("Keypad.TButton", font=button_font, wraplength=button_width_chars*10,justify='center',padding=(0))
            for i, behavior_text in enumerate(combined_options):
                btn = ttk.Button(scrollable_frame_for_buttons, text=behavior_text,
                                padding=(0,19),# Rough estimate for pixel width
                                width=button_width_chars, style="Keypad.TButton",
                                command=lambda n=behavior_text: self.on_button_select(n))
                btn.grid(row=row_idx, column=col_idx, sticky="nsew", padx=2, pady=3)
                type_frame.grid_columnconfigure(col_idx, weight=1) # Allow columns to expand
                col_idx += 1
                if col_idx >= max_cols:
                    col_idx = 0
                    row_idx += 1
            if col_idx != 0: # Ensure last row also configures row weight if partially filled
                scrollable_frame_for_buttons.grid_rowconfigure(row_idx, weight=1)

        ttk.Label(main_frame, text="Comment (Optional):").pack(pady=5, anchor=tk.W, padx=5)
        self.comment_entry = ttk.Entry(main_frame, width=45)
        self.comment_entry.pack(pady=5, fill=tk.X, padx=5)
        
        self.on_template_select(None) # Initialize based on default selection
        return self.homework_type_combobox if self.log_marks_enabled else self.comment_entry

    def on_button_select(self, name):
        self.homework_type_var.set(name)
        self.apply()
        self.ok()

    def _on_press(self, key):
        current_text = self.target_entry.get()
        if key == '⌫': n = len(current_text)-1; self.target_entry.set(current_text[0:n])
        elif key == '.':
            if self.allow_decimal and '.' not in current_text: self.target_entry.set(self.target_entry.get()+ key)
        else: self.target_entry.set(self.target_entry.get()+ key)
        
    def _clear_entry(self): self.target_entry.set("")
    def _next_entry(self):
        if self.cte < self.tte: self.cte +=1
        else: self.cte = 0
        self.target_entry = self.target_entry2[(self.cte)]
        self.target_entry_name = self.target_entry3[(self.cte)]
        self.target_name_label.configure(text=self.target_entry_name)
        
    def _previous_entry(self):
        if self.cte < self.tte and self.cte > -2: self.cte -=1
        else: self.cte = (self.tte-1)
        self.target_entry = self.target_entry2[(self.cte)]
        self.target_entry_name = self.target_entry3[(self.cte)]
        self.target_name_label.configure(text=self.target_entry_name)
        
    def set_numpad(self, event, x):
        #print(x)
        self.cte = x
        self.target_entry = self.target_entry2[(self.cte)]
        self.target_entry_name = self.target_entry3[(self.cte)]
        self.target_name_label.configure(text=self.target_entry_name)
    



    def on_template_select(self, event):
        selected_name = self.homework_type_var.get()
        template = next((tpl for tpl_id, tpl in self.homework_templates.items() if tpl['name'] == selected_name), None)

        if template:
            self.num_items_var.set(str(template.get("num_items", 10)))
            if self.num_items_frame.winfo_ismapped(): self.num_items_spinbox.config(state=tk.DISABLED) # Disable if template provides it

            if self.log_marks_enabled and "default_marks" in template:
                for mark_id, var in self.mark_entry_vars.items():
                    var.set(str(template["default_marks"].get(mark_id, ""))) # Use template's default marks
        else: # Not a template, or template doesn't have these fields
            if self.num_items_frame.winfo_ismapped(): self.num_items_spinbox.config(state=tk.NORMAL)
            # Clear mark entries if not a template or template has no defaults
            # for var in self.mark_entry_vars.values(): var.set("") # Or set to individual mark_type defaults


    def apply(self):
        homework_type = self.homework_type_var.get().strip()
        comment = self.comment_entry.get().strip()
        num_items_val = None
        marks_data = {}

        if not homework_type:
            messagebox.showwarning("Input Required", "Please select or enter a homework type/name.", parent=self)
            self.result = None; return

        if self.log_marks_enabled:
            try: num_items_val = int(self.num_items_var.get())
            except ValueError:
                messagebox.showwarning("Invalid Input", "Number of items must be a valid integer.", parent=self)
                self.result = None; return
            if num_items_val <= 0 :
                messagebox.showwarning("Invalid Input", "Number of items must be positive.", parent=self)
                self.result = None; return

            for mark_id, var in self.mark_entry_vars.items():
                val_str = var.get().strip()
                if val_str: # Only store if a value is entered
                    # Try to convert to float if possible, else store as string
                    try: marks_data[mark_id] = float(val_str)
                    except ValueError: marks_data[mark_id] = val_str
        
        self.result = (homework_type, comment, marks_data if self.log_marks_enabled else None, num_items_val if self.log_marks_enabled else None)

class QuizScoreDialog(simpledialog.Dialog):
    # ... (same as v51)
    def __init__(self, parent, title, initial_quiz_name, mark_types, quiz_templates, default_num_questions, initial_num_questions_val):
        self.initial_quiz_name = initial_quiz_name
        self.mark_types = mark_types # List of dicts: {"id", "name", "default_points", ...}
        self.quiz_templates = quiz_templates # Dict: {template_id: {"name", "num_questions", "default_marks"}}
        self.default_num_questions = default_num_questions
        self.initial_num_questions_val = initial_num_questions_val
        self.result = None
        self.mark_entry_vars = {} # {mark_type_id: StringVar}
        super().__init__(parent, title)

    def body(self, master):
        main_frame = ttk.Frame(master); main_frame.pack(fill=tk.BOTH, expand=True)
        top_frame = ttk.Frame(main_frame); top_frame.grid(pady=5, sticky="ew", columnspan=3)
        ttk.Label(top_frame, text="Quiz Name/Template:").pack(side=tk.LEFT, padx=5)
        self.quiz_name_var = tk.StringVar(value=self.initial_quiz_name)
        
        # Combine quiz name suggestions and template names for the combobox
        template_names = [tpl['name'] for tpl_id, tpl in self.quiz_templates.items()]
        # Simple quiz name suggestions (can be expanded)
        name_suggestions = list(set([self.initial_quiz_name, "Pop Quiz", "Chapter Test"] + template_names))

        self.quiz_name_combobox = ttk.Combobox(top_frame, textvariable=self.quiz_name_var, values=sorted(name_suggestions), width=30)
        self.quiz_name_combobox.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.quiz_name_combobox.bind("<<ComboboxSelected>>", self.on_template_select)
        self.quiz_name_combobox.bind("<MouseWheel>", lambda event: "break")


        num_q_frame = ttk.Frame(main_frame); num_q_frame.grid(pady=5, sticky="ew")
        ttk.Label(num_q_frame, text="Number of Questions:").pack(side=tk.LEFT, padx=5)
        self.num_questions_var = tk.StringVar(value=str(self.initial_num_questions_val if self.initial_num_questions_val else self.default_num_questions))
        self.num_questions_spinbox = ttk.Spinbox(num_q_frame, from_=1, to=200, textvariable=self.num_questions_var, width=5)
        self.num_questions_spinbox.pack(side=tk.LEFT, padx=5)

        self.target_entry2 = []
        self.target_entry3 = []
        self.cte = -1

        marks_frame = ttk.LabelFrame(main_frame, text="Marks Breakdown"); marks_frame.grid(pady=10, column=0, row=2, sticky="nsew", padx=5)
        cols = 2 # Number of columns for mark entries
        current_col_mark = 0; current_row_mark = 0
        for i, mark_type in enumerate(self.mark_types):
            mark_id = mark_type["id"]; mark_name = mark_type["name"]
            ttk.Label(marks_frame, text=f"{mark_name}:").grid(row=current_row_mark, column=current_col_mark*2, sticky=tk.W, padx=5, pady=3)
            var = tk.StringVar()
            self.mark_entry_vars[mark_id] = var
            entry = ttk.Entry(marks_frame, textvariable=var, width=8)
            entry.grid(row=current_row_mark, column=current_col_mark*2 + 1, sticky=tk.EW, padx=5, pady=3)
            entry.bind("<FocusIn>", lambda x=i,x2=i: self.set_numpad(x,x2))
            self.target_entry2.append(var)
            self.target_entry3.append(mark_type["name"])
            self.cte = i
            self.tte = i
            current_col_mark += 1
            if current_col_mark >= cols: current_col_mark = 0; current_row_mark += 1
            
        self.keypad_frame = ttk.Frame(main_frame, relief='sunken',borderwidth=10)
        self.keypad_frame.grid(column=1,row=1, rowspan=4)
        self.keypad_frame.grid_propagate(True)
        self.allow_decimal = True
        self.mark_vars2 = {} # {mark_type_id: tk.StringVar()}
        for mt in self.mark_vars2:
            self.mark_vars2[mt["id"]] = tk.StringVar()
        
        for i, mt_config in enumerate(self.mark_vars2):
            label_text = mt_config["name"]
            if mt_config.get("is_extra_credit"): label_text += " (Bonus)"
        #print(self.mark_vars2)
        self.target_entry = self.mark_entry_vars[mark_type["id"]]
        self.target_entry_name = mark_type["name"]
        
        self.target_name_label = ttk.Label(self.keypad_frame,text=self.target_entry_name)
        self.target_name_label.grid(column=0,row=0,columnspan=9)
        
        buttons = [('7',1,0),('8',1,1),('9',1,2),('4',2,0),('5',2,1),('6',2,2),('1',3,0),('2',3,1),('3',3,2),('0',4,0),('.',4,1) if self.allow_decimal else (' ',4,1),('⌫',4,2),('/',2,3)]
        button_font = ('TkDefaultFont', 15)
        self._next_entry()
        for i in range(4): self.rowconfigure(i, minsize=2000,weight=1)
        for i in range(4): self.columnconfigure(i, weight=1)
        for (text, r, c) in buttons:
            if text == ' ': continue
            action = lambda x=text: self._on_press(x)
            ttk.Button(self.keypad_frame, text=text, command=action, style="Keypad.TButton",padding=17,width=5).grid(row=r, column=c, padx=1, pady=1, sticky="nsew")
        ttk.Button(self.keypad_frame, text="Clear", command=self._clear_entry, style="Keypad.TButton").grid(row=1, column=3, padx=1, pady=1, sticky="nsew")
        ttk.Button(self.keypad_frame, text='Next', command=self._next_entry, style="Keypad.TButton").grid(row=4,column=3,padx=1,pady=1,sticky='nsew')
        ttk.Button(self.keypad_frame, text='Previous', command=self._previous_entry, style="Keypad.TButton").grid(row=3,column=3,padx=1,pady=1,sticky='nsew')
        s = ttk.Style(); s.configure("Keypad.TButton", font=button_font, padding=(5,10),height=100)
        
            
                
            
            
        for i in range(cols*2): marks_frame.grid_columnconfigure(i, weight=1 if i%2==1 else 0)

        ttk.Label(main_frame, text="Comment (Optional):").grid(pady=(0,0), sticky="sw", column=0, row=3, padx=5)
        self.comment_entry = ttk.Entry(main_frame, width=45); self.comment_entry.grid(pady=(0), column=0, row=4, sticky="ews", padx=5)
        
        self.on_template_select(None) # Initialize based on current quiz name (might be a template)
        return self.quiz_name_combobox

    # Functions for NumPad
    def _on_press(self, key):
        current_text = self.target_entry.get()
        if key == '⌫': n = len(current_text)-1; self.target_entry.set(current_text[0:n])
        elif key == '.':
            if self.allow_decimal and '.' not in current_text: self.target_entry.set(self.target_entry.get()+ key)
        else: self.target_entry.set(self.target_entry.get()+ key)

    def _clear_entry(self): self.target_entry.set("")
    
    def _next_entry(self):
        if self.cte < self.tte: self.cte +=1
        else: self.cte = 0
        self.target_entry = self.target_entry2[(self.cte)]
        self.target_entry_name = self.target_entry3[(self.cte)]
        self.target_name_label.configure(text=self.target_entry_name)
        
    def _previous_entry(self):
        if self.cte < self.tte and self.cte > -2: self.cte -=1
        else: self.cte = (self.tte-1)
        self.target_entry = self.target_entry2[(self.cte)]
        self.target_entry_name = self.target_entry3[(self.cte)]
        self.target_name_label.configure(text=self.target_entry_name)
        
    def set_numpad(self, event, x):
        #print(x)
        self.cte = x
        self.target_entry = self.target_entry2[(self.cte)]
        self.target_entry_name = self.target_entry3[(self.cte)]
        self.target_name_label.configure(text=self.target_entry_name)
    
    # Continued code



    def on_template_select(self, event):
        selected_name = self.quiz_name_var.get()
        template = next((tpl for tpl_id, tpl in self.quiz_templates.items() if tpl['name'] == selected_name), None)
        if template:
            self.num_questions_var.set(str(template.get("num_questions", self.default_num_questions)))
            self.num_questions_spinbox.config(state=tk.DISABLED) # Disable if template provides it
            if "default_marks" in template:
                for mark_id, var in self.mark_entry_vars.items():
                    var.set(str(template["default_marks"].get(mark_id, ""))) # Use template's default marks
        else: # Not a template, or template doesn't have these fields
            self.num_questions_spinbox.config(state=tk.NORMAL)
            # Optionally clear marks or set to individual mark_type defaults when not a template
            # for mark_id, var in self.mark_entry_vars.items(): var.set("")

    def apply(self):
        quiz_name = self.quiz_name_var.get().strip(); comment = self.comment_entry.get().strip()
        try: num_questions_actual = int(self.num_questions_var.get())
        except ValueError: messagebox.showwarning("Invalid Input", "Number of questions must be an integer.", parent=self); self.result = None; return
        if num_questions_actual <= 0: messagebox.showwarning("Invalid Input", "Number of questions must be positive.", parent=self); self.result = None; return

        marks_data = {}
        total_marks_entered = 0 # Sum of counts for each mark type
        for mark_id, var in self.mark_entry_vars.items():
            val_str = var.get().strip()
            if val_str:
                try: 
                    val_int = int(val_str)
                    if val_int < 0: messagebox.showwarning("Invalid Input", f"Mark for '{mark_id}' cannot be negative.",parent=self); self.result=None; return
                    marks_data[mark_id] = val_int
                    # Only sum if it's a primary mark contributing to total questions (heuristic)
                    mark_type_obj = next((mt for mt in self.mark_types if mt["id"] == mark_id), None)
                    if mark_type_obj and mark_type_obj.get("contributes_to_total", True):
                        total_marks_entered += val_int
                except ValueError: messagebox.showwarning("Invalid Input", f"Marks for '{mark_id}' must be integers.", parent=self); self.result = None; return
        
        # Validate that sum of primary marks entered does not exceed num_questions_actual
        # This logic assumes 'contributes_to_total' correctly identifies marks that sum up to total questions.
        if total_marks_entered > num_questions_actual:
             messagebox.showwarning("Marks Exceed Questions", f"The sum of primary marks ({total_marks_entered}) exceeds the total number of questions ({num_questions_actual}).\nPlease check your entries.", parent=self)
             self.result = None; return

        if quiz_name: self.result = (quiz_name, marks_data, comment, num_questions_actual)
        else: messagebox.showwarning("Input Required", "Quiz name cannot be empty.", parent=self); self.result = None

class LiveQuizMarkDialog(simpledialog.Dialog):
    # ... (same as v51)
    def __init__(self, parent, student_id, app_instance, session_type="Quiz"): # session_type can be Quiz or Homework
        self.student_id = student_id
        self.app = app_instance
        self.student_name = self.app.students[student_id]['full_name']
        self.session_type_display = session_type
        self.result = None # "correct", "incorrect", "skip"
        super().__init__(parent, f"Mark {session_type} for {self.student_name}")

    def body(self, master):
        if self.session_type_display == "Quiz":
            current_score_info = self.app.live_quiz_scores.get(self.student_id, {"correct": 0, "total_asked": 0})
            score_text = f"Current Score: {current_score_info['correct']} / {current_score_info['total_asked']}"
            ttk.Label(master, text=score_text, font=("", 10)).pack(pady=(5,0))
            ttk.Label(master, text=f"Mark next question for {self.student_name}:").pack(pady=(5,10))
        else: # Homework or other types could be added
            ttk.Label(master, text=f"Update {self.session_type_display} status for {self.student_name}:").pack(pady=(5,10))
        return master # No specific focus needed as it's button driven

    def buttonbox(self): # Override to provide custom buttons
        button_frame = ttk.Frame(self)
        button_frame.pack(pady=10)
        if self.session_type_display == "Quiz":
            ttk.Button(button_frame, text="Correct ✔️", command=lambda: self.set_result_and_close("correct"), width=12).pack(side=tk.LEFT, padx=5)
            ttk.Button(button_frame, text="Incorrect ❌", command=lambda: self.set_result_and_close("incorrect"), width=12).pack(side=tk.LEFT, padx=5)
            ttk.Button(button_frame, text="Skip/Pass ⏭️", command=lambda: self.set_result_and_close("skip"), width=12).pack(side=tk.LEFT, padx=5)
        # Add buttons for other session types if needed
        ttk.Button(button_frame, text="Cancel", command=self.cancel, width=10).pack(side=tk.LEFT, padx=5)
        self.bind("<Escape>", lambda e: self.cancel())
        # Could bind 1,2,3 to correct, incorrect, skip for faster input

    def set_result_and_close(self, res_val):
        self.result = res_val
        self.destroy() # Close the dialog

    def apply(self): pass # Not strictly needed due to custom button actions

class LiveHomeworkMarkDialog(simpledialog.Dialog): # New
    def __init__(self, parent, student_id, app_instance, session_mode, current_hw_data):
        self.student_id = student_id
        self.app = app_instance
        self.student_name = self.app.students[student_id]['full_name']
        self.session_mode = session_mode # "Yes/No" or "Select"
        self.current_hw_data = current_hw_data # Existing data for this student in this session
        self.result_actions = None # Will store dict for Yes/No or list for Select
        
        # For Yes/No mode:
        self.homework_item_vars = {} # {homework_type_id: StringVar}
        # For Select mode:
        self.homework_option_vars = {} # {option_name: BooleanVar}

        super().__init__(parent, f"Mark Homework: {self.student_name} ({self.session_mode} Mode)")

    def body(self, master):
        main_frame = ttk.Frame(master); main_frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
        
        current_session_name_label = ttk.Label(main_frame, text=f"Session: {self.app.current_live_homework_name}", font=("", 11, "italic"))
        current_session_name_label.pack(pady=(0,10))

        if self.session_mode == "Yes/No":
            # Display a list of homework types defined in settings for Yes/No mode
            # self.app.all_homework_session_types (list of dicts: {"id", "name"})'
            select_options = self.app.settings.get("live_homework_select_mode_options", DEFAULT_HOMEWORK_SESSION_BUTTONS.copy())
            if not self.app.all_homework_session_types:
                 ttk.Label(main_frame, text="No homework types configured for 'Yes/No' mode in settings.").pack()
                 return master

            yes_no_frame = ttk.Frame(main_frame)
            yes_no_frame.pack(expand=True, fill=tk.BOTH)
            #print(self.app.all_homework_session_types)
            for hw_type_item in self.app.all_homework_session_types:
                hw_id = hw_type_item["id"]; hw_name = hw_type_item["name"]
                item_frame = ttk.Frame(yes_no_frame); item_frame.pack(fill=tk.X, pady=2)
                ttk.Label(item_frame, text=f"{hw_name}:", width=20, anchor=tk.W).pack(side=tk.LEFT, padx=5)
                
                var = tk.StringVar(value=self.current_hw_data.get(hw_id, "Pending").lower()) # Default to current or "Pending"
                self.homework_item_vars[hw_id] = var
                
                rb_yes = ttk.Radiobutton(item_frame, text="Yes", variable=var, value="yes")
                rb_yes.pack(side=tk.LEFT, padx=3)
                rb_no = ttk.Radiobutton(item_frame, text="No", variable=var, value="no")
                rb_no.pack(side=tk.LEFT, padx=3)
                rb_clear = ttk.Radiobutton(item_frame, text="Pending", variable=var, value="pending") # Clear/Pending option
                rb_clear.pack(side=tk.LEFT, padx=3)


        elif self.session_mode == "Select":
            # Display buttons/checkboxes based on DEFAULT_HOMEWORK_SESSION_BUTTONS or a customizable list from settings
            select_options = self.app.settings.get("live_homework_select_mode_options", DEFAULT_HOMEWORK_SESSION_BUTTONS.copy())
            if not select_options:
                 ttk.Label(main_frame, text="No options configured for 'Select' mode in settings.").pack()
                 return master
            
            select_frame = ttk.Frame(main_frame)
            select_frame.pack(expand=True, fill=tk.BOTH)
            ttk.Label(select_frame, text="Select applicable statuses:").grid(sticky="nw", pady=(0,5))

            # Get current selections for this student
            current_selected_options = self.current_hw_data.get("selected_options", [])

            # Use Checkbuttons for multi-select
            cols = 2; current_col = 0; current_row_sel = 1
            for option_info in select_options: # option_info is a dict, e.g., {"name": "Done"}
                option_name = option_info["name"]
                var = tk.BooleanVar(value=(option_name in current_selected_options))
                self.homework_option_vars[option_name] = var
                cb = ttk.Checkbutton(select_frame, text=option_name, variable=var)
                cb.grid(row=current_row_sel, column=current_col, sticky=tk.W, padx=5, pady=2)
                current_col += 1
                if current_col >= cols: current_col = 0; current_row_sel +=1
            for i in range(cols): select_frame.grid_columnconfigure(i, weight=1)

        return master

    def apply(self):
        if self.session_mode == "Yes/No":
            self.result_actions = {}
            for hw_id, var in self.homework_item_vars.items():
                status = var.get()
                if status != "pending": # Only store "yes" or "no"
                    self.result_actions[hw_id] = status
        elif self.session_mode == "Select":
            self.result_actions = []
            for option_name, var in self.homework_option_vars.items():
                if var.get():
                    self.result_actions.append(option_name)
        # If result_actions is empty (e.g. all pending in Yes/No, or nothing selected in Select)
        # it will be handled by the command to potentially clear the student's entry.

class ExitConfirmationDialog(simpledialog.Dialog): # Same as v50
    def __init__(self, parent, title): self.choice = None; super().__init__(parent, title)
    def body(self, master): ttk.Label(master, text="Do you want to save changes before quitting?").pack(pady=10, padx=10); return None
    def buttonbox(self):
        box = ttk.Frame(self)
        box.columnconfigure(0, weight=1); box.columnconfigure(1, weight=1); box.columnconfigure(2, weight=1)
        ttk.Button(box, text="Save and Quit", width=15, command=self.save_quit).grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        ttk.Button(box, text="Don't Save and Quit", width=20, command=self.no_save_quit).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(box, text="Cancel", width=10, command=self.cancel).grid(row=0, column=2, padx=5, pady=5, sticky="ew")
        self.bind("<Return>", lambda e: self.save_quit()); self.bind("<Escape>", self.cancel)
        box.pack(fill=tk.X, padx=5, pady=5)
    def save_quit(self): self.result = "save_quit"; self.destroy()
    def no_save_quit(self): self.result = "no_save_quit"; self.destroy()

class ImportExcelOptionsDialog(simpledialog.Dialog): # Same as v50, but ensure app_instance is passed
    def __init__(self, parent, app_instance):
        self.app_instance = app_instance
        self.file_path_var = tk.StringVar()
        self.import_incidents_var = tk.BooleanVar(value=False)
        self.student_sheet_var = tk.StringVar()
        self.workbook_sheet_names = []
        super().__init__(parent, "Import Data from Excel")
    def body(self, master):
        ttk.Label(master, text="Excel File:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.file_entry = ttk.Entry(master, textvariable=self.file_path_var, width=40, state="readonly")
        self.file_entry.grid(row=0, column=1, padx=5, pady=2)
        ttk.Button(master, text="Browse...", command=self._browse_file).grid(row=0, column=2, padx=5, pady=2)
        ttk.Label(master, text="Student Info Sheet:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.student_sheet_combo = ttk.Combobox(master, textvariable=self.student_sheet_var, width=37, state="disabled")
        self.student_sheet_combo.grid(row=1, column=1, padx=5, pady=2)
        ttk.Checkbutton(master, text="Import incidents from individual student sheets\n(matches sheet names to current students)", variable=self.import_incidents_var).grid(row=2, column=0, columnspan=3, sticky="w", padx=5, pady=5)
        return self.file_entry
    def _browse_file(self):
        path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        if path:
            self.file_path_var.set(path)
            try:
                wb = load_workbook(filename=path, read_only=True, data_only=True)
                self.workbook_sheet_names = wb.sheetnames
                self.student_sheet_combo['values'] = [""] + self.workbook_sheet_names # Add blank option
                common_sheet_names = ["Students", "Student List", "Roster", "Students Info", "Sheet1"] # Common names
                found_common = next((name for name in common_sheet_names if name in self.workbook_sheet_names), "")
                self.student_sheet_var.set(found_common if found_common else (self.workbook_sheet_names[0] if self.workbook_sheet_names else ""))
                self.student_sheet_combo.config(state="readonly")
            except Exception as e:
                messagebox.showerror("Excel Error", f"Could not read Excel file: {e}", parent=self.app_instance.root)
                self.student_sheet_combo['values'] = [""]; self.student_sheet_var.set(""); self.student_sheet_combo.config(state="disabled")
    def validate(self):
        if not self.file_path_var.get() or not os.path.exists(self.file_path_var.get()):
            messagebox.showerror("File Error", "Please select a valid Excel file.", parent=self.app_instance.root); return False
        # No need to check workbook_sheet_names here as _browse_file handles errors.
        # Student sheet can be blank if user doesn't want to import students.
        return True
    def apply(self):
        if not self.validate(): self.result = None; return
        self.result = (self.file_path_var.get(), self.import_incidents_var.get(), self.student_sheet_var.get() or None) # Return None if sheet is blank

class SizeInputDialog(simpledialog.Dialog): # Same as v50
    def __init__(self, parent, title, initial_w, initial_h, status):
        self.status = status
        self.initial_w, self.initial_h = initial_w, initial_h
        self.width_var = tk.IntVar(value=initial_w); self.height_var = tk.IntVar(value=initial_h)
        super().__init__(parent, title)
    def body(self, master):
        ttk.Label(master, text="Width:").grid(row=0, column=0, sticky="W", padx=5, pady=5)
        self.width_entry = ttk.Spinbox(master, from_=MIN_STUDENT_BOX_WIDTH, to=1000, textvariable=self.width_var, width=7)
        self.width_entry.grid(row=0, column=1, padx=5, pady=5)
        self.width_reset_entry = ttk.Button(master,command=self.reset_width ,text="Reset Width")
        self.width_reset_entry.grid(row=0,column=2,padx=5,pady=5)
        
        ttk.Label(master, text="Height:").grid(row=1, column=0, sticky="W", padx=5, pady=5)
        self.height_entry = ttk.Spinbox(master, from_=MIN_STUDENT_BOX_HEIGHT, to=1000, textvariable=self.height_var, width=7)
        self.height_entry.grid(row=1, column=1, padx=5, pady=5)
        self.height_reset_entry = ttk.Button(master,command=self.reset_height, text="Reset Height")
        self.height_reset_entry.grid(row=1,column=2,padx=5,pady=5)
        return self.width_entry
    
    def reset_width(self):
        if self.status:
            self.width_var.set(DEFAULT_STUDENT_BOX_WIDTH)
        else:
            self.width_var.set(REBBI_DESK_WIDTH)
        
    def reset_height(self):
        #self.height_var.set(DEFAULT_STUDENT_BOX_HEIGHT)
        if self.status:
            self.height_var.set(DEFAULT_STUDENT_BOX_HEIGHT)
        else:
            self.height_var.set(REBBI_DESK_HEIGHT)
    
    def validate(self):
        try:
            w, h = self.width_var.get(), self.height_var.get()
            if w < MIN_STUDENT_BOX_WIDTH or h < MIN_STUDENT_BOX_HEIGHT or w > 1000 or h > 1000:
                messagebox.showerror("Invalid Size", f"Width ({MIN_STUDENT_BOX_WIDTH}-1000), Height ({MIN_STUDENT_BOX_HEIGHT}-1000).", parent=self); return False
            return True
        except tk.TclError: messagebox.showerror("Invalid Input", "Valid numbers for W/H.", parent=self); return False
    def apply(self):
        if not self.validate(): self.result = None; return
        self.result = (self.width_var.get(), self.height_var.get())

class StudentStyleDialog(simpledialog.Dialog):
    # ... (same as v51)
    def __init__(self, parent, title, student_data, app):
        self.student_data = student_data
        self.app = app # For default settings and font list
        self.result = [] # List of (property, old_value, new_value) tuples for Command
        self.initial_overrides = student_data.get("style_overrides", {}).copy()
        super().__init__(parent, title)

    def body(self, master):
        prop_frame = ttk.Frame(master); prop_frame.pack(padx=10,pady=10)
        row_idx = 0
        # Fill Color
        ttk.Label(prop_frame, text="Box Fill Color:").grid(row=row_idx, column=0, sticky=tk.W, pady=3)
        self.fill_color_var = tk.StringVar(value=self.initial_overrides.get("fill_color", ""))
        self.fill_color_entry = ttk.Entry(prop_frame, textvariable=self.fill_color_var, width=15)
        self.fill_color_entry.grid(row=row_idx, column=1, pady=3, padx=2)
        ttk.Button(prop_frame, text="Choose...", command=lambda v=self.fill_color_var: self.choose_color(v)).grid(row=row_idx, column=2, pady=3, padx=2)
        ttk.Button(prop_frame, text="Default", command=lambda v=self.fill_color_var, k="fill_color": self.reset_to_default(v,k)).grid(row=row_idx, column=3, pady=3, padx=2)
        row_idx+=1
        # Outline Color
        ttk.Label(prop_frame, text="Box Outline Color:").grid(row=row_idx, column=0, sticky=tk.W, pady=3)
        self.outline_color_var = tk.StringVar(value=self.initial_overrides.get("outline_color", ""))
        self.outline_color_entry = ttk.Entry(prop_frame, textvariable=self.outline_color_var, width=15)
        self.outline_color_entry.grid(row=row_idx, column=1, pady=3, padx=2)
        ttk.Button(prop_frame, text="Choose...", command=lambda v=self.outline_color_var: self.choose_color(v)).grid(row=row_idx, column=2, pady=3, padx=2)
        ttk.Button(prop_frame, text="Default", command=lambda v=self.outline_color_var, k="outline_color": self.reset_to_default(v,k)).grid(row=row_idx, column=3, pady=3, padx=2)
        row_idx+=1
        # Font Family
        ttk.Label(prop_frame, text="Font Family:").grid(row=row_idx, column=0, sticky=tk.W, pady=3)
        self.font_family_var = tk.StringVar(value=self.initial_overrides.get("font_family", ""))
        available_fonts = self.app.settings.get("available_fonts", [DEFAULT_FONT_FAMILY])
        self.font_family_combo = ttk.Combobox(prop_frame, textvariable=self.font_family_var, values=available_fonts, width=20, state="readonly")
        self.font_family_combo.grid(row=row_idx, column=1, columnspan=2, pady=3, padx=2, sticky=tk.EW)
        self.font_family_combo.bind("<MouseWheel>", lambda event: "break")
        ttk.Button(prop_frame, text="Default", command=lambda v=self.font_family_var, k="font_family": self.reset_to_default(v,k)).grid(row=row_idx, column=3, pady=3, padx=2)
        row_idx+=1
        # Font Size
        ttk.Label(prop_frame, text="Font Size (pts):").grid(row=row_idx, column=0, sticky=tk.W, pady=3)
        self.font_size_var = tk.IntVar(value=self.initial_overrides.get("font_size", 0) or 0) # Ensure it's int, 0 for default placeholder
        self.font_size_spinbox = ttk.Spinbox(prop_frame, from_=6, to=30, textvariable=self.font_size_var, width=5)
        self.font_size_spinbox.grid(row=row_idx, column=1, pady=3, padx=2, sticky=tk.W)
        ttk.Button(prop_frame, text="Default", command=lambda v=self.font_size_var, k="font_size": self.reset_to_default(v,k)).grid(row=row_idx, column=3, pady=3, padx=2)
        row_idx+=1
        # Font Color
        ttk.Label(prop_frame, text="Font Color:").grid(row=row_idx, column=0, sticky=tk.W, pady=3)
        self.font_color_var = tk.StringVar(value=self.initial_overrides.get("font_color", ""))
        self.font_color_entry = ttk.Entry(prop_frame, textvariable=self.font_color_var, width=15)
        self.font_color_entry.grid(row=row_idx, column=1, pady=3, padx=2)
        ttk.Button(prop_frame, text="Choose...", command=lambda v=self.font_color_var: self.choose_color(v)).grid(row=row_idx, column=2, pady=3, padx=2)
        ttk.Button(prop_frame, text="Default", command=lambda v=self.font_color_var, k="font_color": self.reset_to_default(v,k)).grid(row=row_idx, column=3, pady=3, padx=2)
        return self.fill_color_entry # Initial focus
    def choose_color(self, var_to_set):
        initial_color = var_to_set.get() if var_to_set.get() else None
        color_code = colorchooser.askcolor(title="Choose color", initialcolor=initial_color, parent=self)
        if color_code and color_code[1]: var_to_set.set(color_code[1])
    def reset_to_default(self, var_to_set, key):
        # For string vars (colors, font family), set to empty string to signify "use app default"
        # For int var (font size), set to 0 to signify "use app default"
        if isinstance(var_to_set, tk.StringVar): var_to_set.set("")
        elif isinstance(var_to_set, tk.IntVar): var_to_set.set(0)
    def apply(self):
        style_props_vars = {
            "fill_color": self.fill_color_var, "outline_color": self.outline_color_var,
            "font_family": self.font_family_var, "font_size": self.font_size_var,
            "font_color": self.font_color_var
        }
        for prop_key, tk_var in style_props_vars.items():
            new_val = tk_var.get()
            # Standardize "default" representation: None for deletion from overrides
            # Empty string for colors/font family, 0 for font size, means use default
            if isinstance(new_val, str) and not new_val.strip(): final_new_val = None
            elif isinstance(new_val, int) and new_val == 0: final_new_val = None
            else: final_new_val = new_val
            
            old_val_from_overrides = self.initial_overrides.get(prop_key) # This will be None if key didn't exist
            if final_new_val != old_val_from_overrides:
                self.result.append((prop_key, old_val_from_overrides, final_new_val))









class AttendanceReportDialog(simpledialog.Dialog):
    def __init__(self, parent, students_dict):
        self.students_dict = students_dict
        self.result = None
        super().__init__(parent, "Generate Attendance Report")

    def body(self, master):
        frame = ttk.Frame(master); frame.pack(padx=10, pady=10)

        date_frame = ttk.LabelFrame(frame, text="Report Date Range"); date_frame.pack(pady=5, fill=tk.X)
        ttk.Label(date_frame, text="Start Date:").grid(row=0, column=0, padx=5, pady=3, sticky=tk.W)
        self.start_date_var = tk.StringVar(value=(datetime_date.today() - timedelta(days=7)).strftime('%Y-%m-%d'))
        if DateEntry:
            self.start_date_entry = DateEntry(date_frame, textvariable=self.start_date_var, date_pattern='yyyy-mm-dd', width=12)
        else:
            self.start_date_entry = ttk.Entry(date_frame, textvariable=self.start_date_var, width=12)
        self.start_date_entry.grid(row=0, column=1, padx=5, pady=3)

        ttk.Label(date_frame, text="End Date:").grid(row=1, column=0, padx=5, pady=3, sticky=tk.W)
        self.end_date_var = tk.StringVar(value=datetime_date.today().strftime('%Y-%m-%d'))
        if DateEntry:
            self.end_date_entry = DateEntry(date_frame, textvariable=self.end_date_var, date_pattern='yyyy-mm-dd', width=12)
        else:
            self.end_date_entry = ttk.Entry(date_frame, textvariable=self.end_date_var, width=12)
        self.end_date_entry.grid(row=1, column=1, padx=5, pady=3)

        student_frame = ttk.LabelFrame(frame, text="Select Students"); student_frame.pack(pady=5, fill=tk.X)
        self.student_listbox = tk.Listbox(student_frame, selectmode=tk.MULTIPLE, height=8, exportselection=False)
        self.sorted_student_list_for_dialog = sorted(self.students_dict.values(), key=lambda s: (s['last_name'], s['first_name']))
        self.student_id_map_for_dialog = {} # display_name -> student_id
        for i, s_data in enumerate(self.sorted_student_list_for_dialog):
            display_name = f"{s_data['last_name']}, {s_data['first_name']}" + (f" ({s_data.get('nickname')})" if s_data.get('nickname') else "")
            self.student_listbox.insert(tk.END, display_name)
            self.student_id_map_for_dialog[display_name] = s_data["id"]
            self.student_listbox.selection_set(i) # Select all by default

        self.student_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=3, padx=3)
        student_scroll = ttk.Scrollbar(student_frame, orient=tk.VERTICAL, command=self.student_listbox.yview)
        student_scroll.pack(side=tk.RIGHT, fill=tk.Y, pady=3, padx=3)
        self.student_listbox.config(yscrollcommand=student_scroll.set)
        
        select_buttons_frame = ttk.Frame(student_frame)
        select_buttons_frame.pack(fill=tk.X, pady=(0,3))
        ttk.Button(select_buttons_frame, text="Select All", command=lambda: self.student_listbox.selection_set(0, tk.END)).pack(side=tk.LEFT, padx=5)
        ttk.Button(select_buttons_frame, text="Deselect All", command=lambda: self.student_listbox.selection_clear(0, tk.END)).pack(side=tk.LEFT, padx=5)


        return frame

    def apply(self):
        start_dt, end_dt = None, None
        try:
            if self.start_date_var.get(): start_dt = datetime.strptime(self.start_date_var.get(), '%Y-%m-%d').date()
            if self.end_date_var.get(): end_dt = datetime.strptime(self.end_date_var.get(), '%Y-%m-%d').date()
            if not start_dt or not end_dt:
                messagebox.showerror("Missing Dates", "Start and End dates are required.", parent=self); return
            if start_dt > end_dt:
                messagebox.showerror("Invalid Dates", "Start date cannot be after end date.", parent=self); return
        except ValueError:
            messagebox.showerror("Invalid Date Format", "Please use YYYY-MM-DD for dates.", parent=self); return

        selected_student_ids = [self.student_id_map_for_dialog[self.student_listbox.get(i)] for i in self.student_listbox.curselection()]
        if not selected_student_ids:
            messagebox.showwarning("No Students Selected", "Please select at least one student for the report.", parent=self); return

        self.result = (start_dt, end_dt, selected_student_ids)

class ConditionalFormattingRuleDialog(simpledialog.Dialog):
    def __init__(self, parent, app, rule_to_edit=None):
        self.app = app
        self.rule = rule_to_edit or {} # Existing rule or new empty dict
        self.result = None
        title = "Edit Conditional Formatting Rule" if rule_to_edit else "Add Conditional Formatting Rule"
        super().__init__(parent, title)

    def body(self, master):
        frame = ttk.Frame(master); frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        # Rule Type
        ttk.Label(frame, text="Rule applies to:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=3)
        self.rule_type_var = tk.StringVar(value=self.rule.get("type", "group"))
        type_options = ["group", "behavior_count", "quiz_score_threshold", "quiz_mark_count",
                        "live_quiz_response", "live_homework_yes_no", "live_homework_select"]
        self.type_combo = ttk.Combobox(frame, textvariable=self.rule_type_var, values=type_options, state="readonly", width=30) # Increased width
        self.type_combo.grid(row=0, column=1, columnspan=2, padx=5, pady=3, sticky=tk.EW)
        self.type_combo.bind("<<ComboboxSelected>>", self.on_rule_type_change)

        # Condition Frame (changes based on rule_type)
        self.condition_frame = ttk.Frame(frame)
        self.condition_frame.grid(row=1, column=0, columnspan=3, pady=5, sticky=tk.NSEW)

        # Formatting Options
        format_frame = ttk.LabelFrame(frame, text="Formatting Actions")
        format_frame.grid(row=2, column=0, columnspan=3, pady=10, sticky=tk.EW)

        ttk.Label(format_frame, text="Set Box Fill Color:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=3)
        self.fill_color_var = tk.StringVar(value=self.rule.get("color", ""))
        self.fill_color_entry = ttk.Entry(format_frame, textvariable=self.fill_color_var, width=12)
        self.fill_color_entry.grid(row=0, column=1, padx=2, pady=3)
        ttk.Button(format_frame, text="Choose...", command=lambda: self.choose_color_for_var(self.fill_color_var)).grid(row=0, column=2, padx=2, pady=3)
        ttk.Button(format_frame, text="Clear", command=lambda: self.fill_color_var.set("")).grid(row=0, column=3, padx=2, pady=3)


        ttk.Label(format_frame, text="Set Box Outline Color:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=3)
        self.outline_color_var = tk.StringVar(value=self.rule.get("outline", ""))
        self.outline_color_entry = ttk.Entry(format_frame, textvariable=self.outline_color_var, width=12)
        self.outline_color_entry.grid(row=1, column=1, padx=2, pady=3)
        ttk.Button(format_frame, text="Choose...", command=lambda: self.choose_color_for_var(self.outline_color_var)).grid(row=1, column=2, padx=2, pady=3)
        ttk.Button(format_frame, text="Clear", command=lambda: self.outline_color_var.set("")).grid(row=1, column=3, padx=2, pady=3)

        ttk.Label(format_frame, text="Application Style:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=3)
        self.application_style_var = tk.StringVar(value=self.rule.get("application_style", "stripe"))
        style_combo = ttk.Combobox(format_frame, textvariable=self.application_style_var, values=["stripe", "override"], state="readonly", width=10)
        style_combo.grid(row=2, column=1, padx=2, pady=3, sticky=tk.W)


        # Initialize condition frame based on current/default rule type
        self.on_rule_type_change(None)

        # --- Enable/Disable Rule Checkbox ---
        self.enabled_var = tk.BooleanVar(value=self.rule.get("enabled", True))
        ttk.Checkbutton(frame, text="Rule Enabled", variable=self.enabled_var).grid(row=3, column=0, columnspan=3, sticky=tk.W, padx=5, pady=(10,0))

        # --- Active Times Frame ---
        times_frame = ttk.LabelFrame(frame, text="Active Times (Optional - if none, active all times)");
        times_frame.grid(row=4, column=0, columnspan=3, pady=5, sticky=tk.NSEW)

        # Entry for new time slot
        new_time_slot_frame = ttk.Frame(times_frame)
        new_time_slot_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(new_time_slot_frame, text="Start (HH:MM):").pack(side=tk.LEFT)
        self.start_time_var = tk.StringVar()
        ttk.Entry(new_time_slot_frame, textvariable=self.start_time_var, width=6).pack(side=tk.LEFT, padx=2)
        ttk.Label(new_time_slot_frame, text="End (HH:MM):").pack(side=tk.LEFT, padx=2)
        self.end_time_var = tk.StringVar()
        ttk.Entry(new_time_slot_frame, textvariable=self.end_time_var, width=6).pack(side=tk.LEFT, padx=2)

        self.days_vars = [tk.BooleanVar(value=True) for _ in range(7)] # Mon-Sun
        days_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        days_frame = ttk.Frame(new_time_slot_frame)
        days_frame.pack(side=tk.LEFT, padx=5)
        for i, day_name in enumerate(days_names):
            ttk.Checkbutton(days_frame, text=day_name, variable=self.days_vars[i]).pack(side=tk.LEFT)

        ttk.Button(new_time_slot_frame, text="Add Time Slot", command=self.add_time_slot).pack(side=tk.LEFT, padx=5)

        # Listbox for existing time slots
        self.times_listbox = tk.Listbox(times_frame, height=3, exportselection=False)
        self.times_listbox.pack(fill=tk.X, expand=True, padx=5, pady=2)
        ttk.Button(times_frame, text="Remove Selected Slot", command=self.remove_time_slot).pack(pady=2, padx=5, anchor=tk.W)
        self._populate_times_listbox()


        # --- Active Modes Frame ---
        modes_frame = ttk.LabelFrame(frame, text="Active Modes (Optional - if none, active in all modes)");
        modes_frame.grid(row=5, column=0, columnspan=3, pady=5, sticky=tk.NSEW)
        self.available_modes = ["behavior", "quiz", "homework", "quiz_session", "homework_session"]
        self.mode_vars = {mode_name: tk.BooleanVar(value=(mode_name in self.rule.get("active_modes", []))) for mode_name in self.available_modes}

        modes_checkbox_frame = ttk.Frame(modes_frame)
        modes_checkbox_frame.pack(fill=tk.X, padx=5, pady=5)
        for i, mode_name in enumerate(self.available_modes):
            ttk.Checkbutton(modes_checkbox_frame, text=mode_name.replace("_", " ").capitalize(), variable=self.mode_vars[mode_name]).grid(row=i//3, column=i%3, sticky=tk.W, padx=5, pady=2)


        return self.type_combo

    def _populate_times_listbox(self):
        self.times_listbox.delete(0, tk.END)
        days_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        for slot in self.rule.get("active_times", []):
            days_str = ", ".join([days_names[i] for i in slot.get("days_of_week", [])])
            self.times_listbox.insert(tk.END, f"{slot['start_time']} - {slot['end_time']} ({days_str if days_str else 'All Days'})")

    def add_time_slot(self):
        start_time = self.start_time_var.get().strip()
        end_time = self.end_time_var.get().strip()
        if ":" not in start_time:
            start_time = start_time[:2] + ":" + start_time[2:]
        if ":" not in end_time:
            end_time = end_time[:2] + ":" + end_time[2:]
        try:
            datetime.strptime(start_time, "%H:%M")
            datetime.strptime(end_time, "%H:%M")
            if start_time >= end_time:
                messagebox.showerror("Invalid Time", "Start time must be before end time.", parent=self)
                return
        except ValueError:
            messagebox.showerror("Invalid Time Format", "Please use HH:MM format for times (e.g., 09:00, 14:30).", parent=self)
            return

        selected_days = [i for i, var in enumerate(self.days_vars) if var.get()]
        if not selected_days: # Default to all days if none explicitly selected by user for this new slot
            selected_days = list(range(7))

        new_slot = {"start_time": start_time, "end_time": end_time, "days_of_week": selected_days}

        if "active_times" not in self.rule: self.rule["active_times"] = []
        self.rule["active_times"].append(new_slot)
        self._populate_times_listbox()
        self.start_time_var.set("")
        self.end_time_var.set("")
        for var in self.days_vars: var.set(True) # Reset day checkboxes

    def remove_time_slot(self):
        selected_indices = self.times_listbox.curselection()
        if not selected_indices: return
        # Iterate reversed to avoid index issues when removing multiple
        for i in reversed(selected_indices):
            del self.rule["active_times"][i]
        self._populate_times_listbox()

    def choose_color_for_var(self, color_var):
        initial = color_var.get() if color_var.get() else None
        color_code = colorchooser.askcolor(title="Choose color", initialcolor=initial, parent=self)
        if color_code and color_code[1]: color_var.set(color_code[1])

    def on_rule_type_change(self, event):
        for widget in self.condition_frame.winfo_children(): widget.destroy() # Clear previous condition widgets
        rule_type = self.rule_type_var.get()

        if rule_type == "group":
            ttk.Label(self.condition_frame, text="Select Group:").pack(side=tk.LEFT, padx=5)
            self.group_id_selection_var = tk.StringVar() # Stores the group ID

            group_options_display = {"": "Select Group..."} # Display name -> ID
            for gid, gdata in sorted(self.app.student_groups.items(), key=lambda item: item[1]['name']):
                group_options_display[gdata['name']] = gid

            self.group_combo_cond = ttk.Combobox(self.condition_frame, textvariable=self.group_id_selection_var,
                                                 values=list(group_options_display.keys()), state="readonly", width=20)

            # Set initial value for combobox (display name)
            current_group_id_in_rule = self.rule.get("group_id")
            display_name_to_set = "Select Group..."
            for name, gid_map in group_options_display.items():
                if gid_map == current_group_id_in_rule:
                    display_name_to_set = name
                    break
            self.group_id_selection_var.set(display_name_to_set) # Set the display name

            # Store the map for apply method
            self.group_display_to_id_map = group_options_display
            self.group_combo_cond.pack(side=tk.LEFT, padx=5)

        elif rule_type == "behavior_count":
            ttk.Label(self.condition_frame, text="Behavior:").pack(side=tk.LEFT, padx=5)
            self.behavior_name_var = tk.StringVar(value=self.rule.get("behavior_name", ""))
            self.behavior_combo_cond = ttk.Combobox(self.condition_frame, textvariable=self.behavior_name_var,
                                               values=[""] + self.app.all_behaviors, width=18)
            self.behavior_combo_cond.pack(side=tk.LEFT, padx=2)

            ttk.Label(self.condition_frame, text="Count >=:").pack(side=tk.LEFT, padx=5)
            self.behavior_count_var = tk.IntVar(value=self.rule.get("count_threshold", 1))
            ttk.Spinbox(self.condition_frame, from_=1, to=100, textvariable=self.behavior_count_var, width=4).pack(side=tk.LEFT, padx=2)

            ttk.Label(self.condition_frame, text="In Last (Hours):").pack(side=tk.LEFT, padx=5)
            self.behavior_hours_var = tk.IntVar(value=self.rule.get("time_window_hours", 24))
            ttk.Spinbox(self.condition_frame, from_=1, to=720, textvariable=self.behavior_hours_var, width=4).pack(side=tk.LEFT, padx=2)

        elif rule_type == "quiz_score_threshold":
            ttk.Label(self.condition_frame, text="Quiz Name (contains):").pack(side=tk.LEFT, padx=5)
            self.quiz_name_contains_var = tk.StringVar(value=self.rule.get("quiz_name_contains", ""))
            ttk.Entry(self.condition_frame, textvariable=self.quiz_name_contains_var, width=15).pack(side=tk.LEFT, padx=2)

            ttk.Label(self.condition_frame, text="Score (%):").pack(side=tk.LEFT, padx=5)
            self.quiz_op_var = tk.StringVar(value=self.rule.get("operator", "<="))
            ttk.Combobox(self.condition_frame, textvariable=self.quiz_op_var, values=["<=", ">=", "==", "<", ">"], width=3, state="readonly").pack(side=tk.LEFT, padx=2)
            self.quiz_score_thresh_var = tk.DoubleVar(value=self.rule.get("score_threshold_percent", 50.0))
            ttk.Spinbox(self.condition_frame, from_=0, to=100, increment=1, textvariable=self.quiz_score_thresh_var, width=5).pack(side=tk.LEFT, padx=2)

        elif rule_type == "quiz_mark_count":
            qmc_frame = ttk.Frame(self.condition_frame); qmc_frame.pack(fill=tk.X, pady=2)
            ttk.Label(qmc_frame, text="Quiz Name (contains):").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
            self.qmc_quiz_name_var = tk.StringVar(value=self.rule.get("quiz_name_contains", ""))
            ttk.Entry(qmc_frame, textvariable=self.qmc_quiz_name_var, width=20).grid(row=0, column=1, padx=5, pady=2, sticky=tk.EW)

            ttk.Label(qmc_frame, text="Mark Type:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
            self.qmc_mark_type_var = tk.StringVar(value=self.rule.get("mark_type_id", "")) # Stores ID

            # Prepare mark type options for combobox
            self.qmc_mark_type_options_map = {} # Maps display name to ID
            mark_type_display_names = [""] # Start with a blank option
            for mt in self.app.settings.get("quiz_mark_types", []):
                self.qmc_mark_type_options_map[mt["name"]] = mt["id"]
                mark_type_display_names.append(mt["name"])

            self.qmc_mark_type_combo = ttk.Combobox(qmc_frame, textvariable=self.qmc_mark_type_var,
                                                    values=mark_type_display_names, state="readonly", width=18)
            # Set initial value for combobox (find name by ID)
            initial_mark_id = self.rule.get("mark_type_id", "")
            initial_mark_name = ""
            for name, mid in self.qmc_mark_type_options_map.items():
                if mid == initial_mark_id:
                    initial_mark_name = name; break
            self.qmc_mark_type_var.set(initial_mark_name) # Set display name
            self.qmc_mark_type_combo.grid(row=1, column=1, padx=5, pady=2, sticky=tk.EW)


            ttk.Label(qmc_frame, text="Operator:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
            self.qmc_operator_var = tk.StringVar(value=self.rule.get("mark_operator", ">="))
            ttk.Combobox(qmc_frame, textvariable=self.qmc_operator_var, values=[">=", "<=", "==", ">", "<", "!="], width=3, state="readonly").grid(row=2, column=1, sticky=tk.W, padx=5, pady=2)

            ttk.Label(qmc_frame, text="Count:").grid(row=2, column=2, sticky=tk.W, padx=(10,2), pady=2)
            self.qmc_count_var = tk.IntVar(value=self.rule.get("mark_count_threshold", 1))
            ttk.Spinbox(qmc_frame, from_=0, to=100, textvariable=self.qmc_count_var, width=4).grid(row=2, column=3, sticky=tk.W, padx=2, pady=2)
            qmc_frame.grid_columnconfigure(1, weight=1)

        elif rule_type == "live_quiz_response":
            lqr_frame = ttk.Frame(self.condition_frame); lqr_frame.pack(fill=tk.X, pady=2)
            ttk.Label(lqr_frame, text="Quiz Response:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
            self.lqr_response_var = tk.StringVar(value=self.rule.get("quiz_response", "Correct"))
            ttk.Combobox(lqr_frame, textvariable=self.lqr_response_var, values=["Correct", "Incorrect"], state="readonly", width=15).grid(row=0, column=1, padx=5, pady=2, sticky=tk.W)

        elif rule_type == "live_homework_yes_no":
            lhy_frame = ttk.Frame(self.condition_frame); lhy_frame.pack(fill=tk.X, pady=2)
            ttk.Label(lhy_frame, text="Homework Type:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)

            self.lhy_hw_type_id_var = tk.StringVar(value=self.rule.get("homework_type_id", "")) # Stores ID
            self.lhy_hw_type_options_map = {item['name']: item['id'] for item in self.app.all_homework_session_types}
            lhy_hw_type_display_names = [""] + sorted(self.lhy_hw_type_options_map.keys())

            self.lhy_hw_type_combo = ttk.Combobox(lhy_frame, textvariable=self.lhy_hw_type_id_var, values=lhy_hw_type_display_names, state="readonly", width=25)
            # Set initial display name
            initial_hw_type_id = self.rule.get("homework_type_id", "")
            initial_hw_type_name = ""
            for name, hid in self.lhy_hw_type_options_map.items():
                if hid == initial_hw_type_id:
                    initial_hw_type_name = name; break
            self.lhy_hw_type_id_var.set(initial_hw_type_name) # Set display name
            self.lhy_hw_type_combo.grid(row=0, column=1, padx=5, pady=2, sticky=tk.EW)

            ttk.Label(lhy_frame, text="Response:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
            self.lhy_response_var = tk.StringVar(value=self.rule.get("homework_response", "yes"))
            ttk.Combobox(lhy_frame, textvariable=self.lhy_response_var, values=["yes", "no"], state="readonly", width=10).grid(row=1, column=1, padx=5, pady=2, sticky=tk.W)
            lhy_frame.grid_columnconfigure(1, weight=1)

        elif rule_type == "live_homework_select":
            lhs_frame = ttk.Frame(self.condition_frame); lhs_frame.pack(fill=tk.X, pady=2)
            ttk.Label(lhs_frame, text="Homework Option:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
            self.lhs_option_name_var = tk.StringVar(value=self.rule.get("homework_option_name", ""))

            select_mode_options = [opt['name'] for opt in self.app.settings.get("live_homework_select_mode_options", [])]
            self.lhs_option_combo = ttk.Combobox(lhs_frame, textvariable=self.lhs_option_name_var, values=[""] + sorted(select_mode_options), state="readonly", width=25)
            self.lhs_option_combo.grid(row=0, column=1, padx=5, pady=2, sticky=tk.EW)
            lhs_frame.grid_columnconfigure(1, weight=1)


    def apply(self):
        final_rule = {"type": self.rule_type_var.get()}
        final_rule["application_style"] = self.application_style_var.get() # Common to all
        rule_type = final_rule["type"]
        
        active_modes = []
        for mode in self.mode_vars:
            if self.mode_vars[mode].get() != False:
                active_modes.append(mode)
        final_rule["active_modes"] = active_modes # type: ignore
        final_rule["active_times"] = self.rule.get("active_times") #["active_times"]
        
        fill = self.fill_color_var.get().strip()
        outline = self.outline_color_var.get().strip()
        if not fill and not outline:
            messagebox.showerror("No Action", "Please specify at least one formatting action (fill or outline color).", parent=self)
            return
        if fill: final_rule["color"] = fill
        if outline: final_rule["outline"] = outline

        if rule_type == "group":
            selected_group_name = self.group_var.get()
            if not selected_group_name: messagebox.showerror("Missing Info", "Please select a group.", parent=self); return
            final_rule["group_id"] = self.group_id_map_cond.get(selected_group_name)
            if not final_rule["group_id"]: messagebox.showerror("Error", "Selected group not found.", parent=self); return

        elif rule_type == "behavior_count":
            b_name = self.behavior_name_var.get().strip()
            if not b_name: messagebox.showerror("Missing Info", "Please select a behavior.", parent=self); return
            final_rule["behavior_name"] = b_name
            final_rule["count_threshold"] = self.behavior_count_var.get()
            final_rule["time_window_hours"] = self.behavior_hours_var.get()

        elif rule_type == "quiz_score_threshold":
            final_rule["quiz_name_contains"] = self.quiz_name_contains_var.get().strip() # Can be empty for any quiz
            final_rule["operator"] = self.quiz_op_var.get()
            final_rule["score_threshold_percent"] = self.quiz_score_thresh_var.get()

        elif rule_type == "quiz_mark_count":
            final_rule["quiz_name_contains"] = self.qmc_quiz_name_var.get().strip() # Can be empty
            selected_mark_name = self.qmc_mark_type_var.get()
            if not selected_mark_name: messagebox.showerror("Missing Info", "Please select a Mark Type.", parent=self); return
            final_rule["mark_type_id"] = self.qmc_mark_type_options_map.get(selected_mark_name)
            if not final_rule["mark_type_id"]: messagebox.showerror("Error", "Selected mark type is invalid.", parent=self); return
            final_rule["mark_operator"] = self.qmc_operator_var.get()
            try:
                count_val = self.qmc_count_var.get()
                if count_val < 0: messagebox.showerror("Invalid Input", "Count cannot be negative.", parent=self); return
                final_rule["mark_count_threshold"] = count_val
            except tk.TclError: messagebox.showerror("Invalid Input", "Count must be a valid integer.", parent=self); return

        elif rule_type == "live_quiz_response":
            response = self.lqr_response_var.get()
            if not response: messagebox.showerror("Missing Info", "Please select a quiz response.", parent=self); return
            final_rule["quiz_response"] = response

        elif rule_type == "live_homework_yes_no":
            selected_hw_type_name = self.lhy_hw_type_id_var.get() # This var now holds the name
            if not selected_hw_type_name: messagebox.showerror("Missing Info", "Please select a homework type.", parent=self); return
            final_rule["homework_type_id"] = self.lhy_hw_type_options_map.get(selected_hw_type_name) # Get ID from name
            if not final_rule["homework_type_id"]: messagebox.showerror("Error", "Selected homework type is invalid.", parent=self); return

            response = self.lhy_response_var.get()
            if not response: messagebox.showerror("Missing Info", "Please select a homework response (Yes/No).", parent=self); return
            final_rule["homework_response"] = response

        elif rule_type == "live_homework_select":
            option_name = self.lhs_option_name_var.get()
            if not option_name: messagebox.showerror("Missing Info", "Please select a homework option.", parent=self); return
            final_rule["homework_option_name"] = option_name
        
        self.result = final_rule

class ManageStudentGroupsDialog(simpledialog.Dialog):
    def __init__(self, parent, student_groups_data, students_data, app, default_colors):
        self.student_groups = student_groups_data # This is a reference to app.student_groups, modified directly
        self.students = students_data # App.students, for assigning
        self.app = app # To get new group ID, and redraw
        self.default_colors = default_colors
        self.groups_changed_flag = False # Set to True if any persistent change is made
        super().__init__(parent, "Manage Student Groups")

    def body(self, master):
        self.master_frame = master
        top_frame = ttk.Frame(master); top_frame.pack(pady=5, padx=5, fill=tk.X)
        ttk.Button(top_frame, text="Add New Group", command=self.add_group).pack(side=tk.LEFT, padx=5)

        self.canvas_groups = tk.Canvas(master, borderwidth=0, background="#ffffff")
        self.groups_scrollable_frame = ttk.Frame(self.canvas_groups)
        self.scrollbar_groups = ttk.Scrollbar(master, orient="vertical", command=self.canvas_groups.yview)
        self.canvas_groups.configure(yscrollcommand=self.scrollbar_groups.set)

        self.scrollbar_groups.pack(side="right", fill="y")
        self.canvas_groups.pack(side="left", fill="both", expand=True)
        self.canvas_groups_window = self.canvas_groups.create_window((0,0), window=self.groups_scrollable_frame, anchor="nw")

        self.groups_scrollable_frame.bind("<Configure>", lambda e: self.canvas_groups.configure(scrollregion=self.canvas_groups.bbox("all")))
        self.canvas_groups.bind('<MouseWheel>', self._on_mousewheel_groups)

        self.populate_groups_list()
        return self.groups_scrollable_frame # For initial focus target (though dynamic)

    def _on_mousewheel_groups(self, event):
        if event.delta: self.canvas_groups.yview_scroll(int(-1*(event.delta/120)), "units")
        else: self.canvas_groups.yview_scroll(1 if event.num == 5 else -1, "units")


    def populate_groups_list(self):
        for widget in self.groups_scrollable_frame.winfo_children(): widget.destroy()
        row_idx = 0
        if not self.student_groups:
            ttk.Label(self.groups_scrollable_frame, text="No groups created yet.").pack(pady=10)
            return

        sorted_groups = sorted(self.student_groups.items(), key=lambda item: item[1]['name'])

        for group_id, group_data in sorted_groups:
            group_frame = ttk.Frame(self.groups_scrollable_frame, padding=5, relief=tk.RIDGE, borderwidth=1)
            group_frame.pack(fill=tk.X, pady=3, padx=3)

            name_var = tk.StringVar(value=group_data["name"])
            color_var = tk.StringVar(value=group_data.get("color", self.default_colors[0]))

            ttk.Label(group_frame, text="Name:").grid(row=0, column=0, sticky=tk.W)
            name_entry = ttk.Entry(group_frame, textvariable=name_var, width=20)
            name_entry.grid(row=0, column=1, padx=3)
            name_entry.bind("<FocusOut>", lambda e, gid=group_id, nv=name_var: self.update_group_name(gid, nv.get()))

            ttk.Label(group_frame, text="Color:").grid(row=0, column=2, sticky=tk.W, padx=(10,0))
            color_entry = ttk.Entry(group_frame, textvariable=color_var, width=10)
            color_entry.grid(row=0, column=3, padx=3)
            color_btn = ttk.Button(group_frame, text="...", width=3, command=lambda cv=color_var, gid=group_id: self.choose_group_color(cv, gid))
            color_btn.grid(row=0, column=4)
            color_preview = tk.Label(group_frame, text="  ", bg=color_var.get(), width=2, relief=tk.SUNKEN)
            color_preview.grid(row=0, column=5, padx=2)
            color_var.trace_add("write", lambda *args, cp=color_preview, cv=color_var: cp.config(bg=cv.get()))


            ttk.Button(group_frame, text="Assign Students...", command=lambda gid=group_id, gname=group_data["name"]: self.assign_students_to_group_dialog(gid, gname)).grid(row=0, column=6, padx=5)
            ttk.Button(group_frame, text="Delete Group", command=lambda gid=group_id: self.delete_group(gid)).grid(row=0, column=7, padx=5)

            # Show number of students in group
            count = sum(1 for sid, sdata in self.students.items() if sdata.get("group_id") == group_id)
            ttk.Label(group_frame, text=f"({count} student{'s' if count !=1 else ''})").grid(row=0, column=8, padx=5, sticky=tk.E)
            group_frame.grid_columnconfigure(8, weight=1) # Push delete button to right a bit


    def add_group(self):
        new_group_name = simpledialog.askstring("New Group", "Enter name for the new group:", parent=self)
        if new_group_name and new_group_name.strip():
            group_id_str, next_id_val = self.app.get_new_group_id() # Get ID from app
            # Check for name collision BEFORE updating app's next_group_id_num state
            if any(g['name'].lower() == new_group_name.strip().lower() for g in self.student_groups.values()):
                 messagebox.showwarning("Duplicate Name", f"A group named '{new_group_name.strip()}' already exists.", parent=self)
                 return # Do not consume the ID from app if name is duplicate

            self.app.next_group_id_num = next_id_val # Commit ID usage
            new_color_index = (self.app.next_group_id_num -1) % len(self.default_colors) # Cycle through default colors
            self.student_groups[group_id_str] = {"name": new_group_name.strip(), "color": self.default_colors[new_color_index]}
            self.groups_changed_flag = True
            self.populate_groups_list()
        elif new_group_name is not None: # User entered empty string
             messagebox.showwarning("Invalid Name", "Group name cannot be empty.", parent=self)


    def update_group_name(self, group_id, new_name):
        new_name = new_name.strip()
        if not new_name:
            messagebox.showwarning("Invalid Name", "Group name cannot be empty. Reverting.", parent=self)
            self.populate_groups_list(); return # Revert by repopulating
        
        # Check for name collision with OTHER groups
        if any(g_id != group_id and g_data['name'].lower() == new_name.lower() for g_id, g_data in self.student_groups.items()):
            messagebox.showwarning("Duplicate Name", f"Another group named '{new_name}' already exists. Reverting.", parent=self)
            self.populate_groups_list(); return

        if self.student_groups[group_id]["name"] != new_name:
            self.student_groups[group_id]["name"] = new_name
            self.groups_changed_flag = True
            # No need to repopulate for just name change if var is linked, but good for consistency if other things change
            self.populate_groups_list() # Repopulate to reflect changes and resort


    def choose_group_color(self, color_var, group_id):
        initial_color = color_var.get()
        new_color = colorchooser.askcolor(initial_color, title=f"Choose color for group", parent=self)
        if new_color and new_color[1]:
            if self.student_groups[group_id].get("color") != new_color[1]:
                color_var.set(new_color[1]) # This will trigger trace and update preview
                self.student_groups[group_id]["color"] = new_color[1]
                self.groups_changed_flag = True
                # No repopulate needed just for color if var is traced.

    def delete_group(self, group_id):
        group_name = self.student_groups[group_id]["name"]
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete group '{group_name}'?\nStudents in this group will be unassigned.", parent=self):
            # Unassign students from this group
            for student_id, student_data in self.students.items():
                if student_data.get("group_id") == group_id:
                    student_data["group_id"] = None # Or del student_data["group_id"]
            del self.student_groups[group_id]
            self.groups_changed_flag = True
            self.populate_groups_list()

    def assign_students_to_group_dialog(self, group_id, group_name):
        dialog = AssignStudentsToGroupSubDialog(self, group_id, group_name, self.students, self.student_groups)
        if dialog.assignments_changed:
            self.groups_changed_flag = True # Signal main app that changes occurred
            self.populate_groups_list() # Repopulate to update student counts

    def apply(self): # Called when OK is pressed
        # Changes are applied directly, so just set the flag if it was ever true
        self.result = self.groups_changed_flag
        if self.groups_changed_flag:
            self.app.draw_all_items(check_collisions_on_redraw=True) # Redraw if groups changed
            self.app.save_student_groups() # Save groups if changed

class AssignStudentsToGroupSubDialog(simpledialog.Dialog):
    def __init__(self, parent_dialog, group_id, group_name, all_students_data, all_groups_data):
        self.parent_dialog_ref = parent_dialog # Reference to ManageStudentGroupsDialog
        self.group_id_to_assign = group_id
        self.group_name = group_name
        self.all_students = all_students_data # app.students
        self.all_groups = all_groups_data # app.student_groups
        self.assignments_changed = False
        super().__init__(parent_dialog, f"Assign Students to '{group_name}'")

    def body(self, master):
        # Listbox for available students (not in any group OR in a different group)
        # Listbox for students currently in THIS group
        frame = ttk.Frame(master); frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        
        available_frame = ttk.LabelFrame(frame, text="Available Students"); available_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        self.available_lb = tk.Listbox(available_frame, selectmode=tk.EXTENDED, exportselection=False, height=15)
        self.available_lb.pack(fill=tk.BOTH, expand=True, pady=2)
        self.available_students_map = {} # display name -> student_id

        buttons_frame = ttk.Frame(frame); buttons_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5)
        ttk.Button(buttons_frame, text=">> Add >>", command=self.add_to_group).pack(pady=10)
        ttk.Button(buttons_frame, text="<< Remove <<", command=self.remove_from_group).pack(pady=10)

        assigned_frame = ttk.LabelFrame(frame, text=f"Students in '{self.group_name}'"); assigned_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        self.assigned_lb = tk.Listbox(assigned_frame, selectmode=tk.EXTENDED, exportselection=False, height=15)
        self.assigned_lb.pack(fill=tk.BOTH, expand=True, pady=2)
        self.assigned_students_map = {} # display name -> student_id

        self.populate_student_lists()
        return self.available_lb

    def populate_student_lists(self):
        self.available_lb.delete(0, tk.END); self.assigned_lb.delete(0, tk.END)
        self.available_students_map.clear(); self.assigned_students_map.clear()

        sorted_all_students = sorted(self.all_students.values(), key=lambda s: (s['last_name'], s['first_name']))

        for s_data in sorted_all_students:
            s_id = s_data["id"]
            display_name = f"{s_data['last_name']}, {s_data['first_name']}" + (f" ({s_data.get('nickname')})" if s_data.get('nickname') else "")
            
            if s_data.get("group_id") == self.group_id_to_assign:
                self.assigned_lb.insert(tk.END, display_name)
                self.assigned_students_map[display_name] = s_id
            elif not s_data.get("group_id") or s_data.get("group_id") == "NONE_GROUP_SENTINEL": # Truly unassigned
                self.available_lb.insert(tk.END, display_name)
                self.available_students_map[display_name] = s_id
            # Else: student is in a *different* group, don't show in "Available" unless we want a "move from group X" feature

    def add_to_group(self):
        selected_indices = self.available_lb.curselection()
        if not selected_indices: return
        for i in reversed(selected_indices): # Iterate reversed to handle index changes on delete
            display_name = self.available_lb.get(i)
            student_id = self.available_students_map.get(display_name)
            if student_id and student_id in self.all_students:
                self.all_students[student_id]["group_id"] = self.group_id_to_assign
                self.assignments_changed = True
        self.populate_student_lists() # Refresh both listboxes

    def remove_from_group(self):
        selected_indices = self.assigned_lb.curselection()
        if not selected_indices: return
        for i in reversed(selected_indices):
            display_name = self.assigned_lb.get(i)
            student_id = self.assigned_students_map.get(display_name)
            if student_id and student_id in self.all_students:
                self.all_students[student_id]["group_id"] = None # Unassign
                self.assignments_changed = True
        self.populate_student_lists()

    def apply(self):
        # Changes were made directly to self.all_students (app.students)
        # The self.assignments_changed flag will be checked by the parent dialog
        pass

class ManageMarkTypesDialog(simpledialog.Dialog):
    def __init__(self, parent, current_mark_types_list, item_type_display_name, default_mark_types):
        self.mark_types_ref = current_mark_types_list # Direct reference, e.g., app.settings["quiz_mark_types"]
        self.item_type_name = item_type_display_name # "Quiz Mark Types" or "Homework Mark Types"
        self.default_mark_types_const = default_mark_types # The constant list for reset
        self.mark_types_changed = False
        super().__init__(parent, f"Manage {self.item_type_name}")

    def body(self, master):
        self.main_frame = master # To rebuild list
        button_frame = ttk.Frame(master); button_frame.pack(fill=tk.X, pady=5)
        ttk.Button(button_frame, text="Add New Mark Type", command=self.add_mark_type).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Reset to Defaults", command=self.reset_to_defaults).pack(side=tk.LEFT, padx=5)

        self.list_frame = ttk.Frame(master); self.list_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.populate_mark_types_ui()
        return self.main_frame # For focus

    def populate_mark_types_ui(self):
        for widget in self.list_frame.winfo_children(): widget.destroy()

        headers = ["ID", "Name", "Default Points", "To Total?", "Bonus?"]
        for c, h_text in enumerate(headers):
            ttk.Label(self.list_frame, text=h_text, font=("",9,"bold")).grid(row=0,column=c,padx=3,pady=3,sticky=tk.W)

        self.mark_type_widgets = [] # Store refs to entry vars/widgets if needed for direct update

        for r_idx, mt_dict in enumerate(self.mark_types_ref, start=1):
            widgets_row = {}
            # ID (display only for defaults, editable for custom?) - For now, mostly fixed post-creation
            id_val = mt_dict.get("id", f"custom_{r_idx}")
            widgets_row["id_label"] = ttk.Label(self.list_frame, text=id_val, width=15); widgets_row["id_label"].grid(row=r_idx, column=0, padx=3, sticky=tk.W)

            # Name
            name_var = tk.StringVar(value=mt_dict.get("name","")); widgets_row["name_var"] = name_var
            name_entry = ttk.Entry(self.list_frame, textvariable=name_var, width=20); name_entry.grid(row=r_idx, column=1, padx=3, sticky=tk.EW)

            # Default Points
            points_var = tk.DoubleVar(value=mt_dict.get("default_points",0.0)); widgets_row["points_var"] = points_var
            points_spin = ttk.Spinbox(self.list_frame, from_=-100, to=100, increment=0.1, textvariable=points_var, width=6); points_spin.grid(row=r_idx, column=2, padx=3)

            # Contributes to Total (Bool)
            to_total_var = tk.BooleanVar(value=mt_dict.get("contributes_to_total",True)); widgets_row["to_total_var"] = to_total_var
            ttk.Checkbutton(self.list_frame, variable=to_total_var).grid(row=r_idx, column=3, padx=3)

            # Is Extra Credit (Bool)
            is_bonus_var = tk.BooleanVar(value=mt_dict.get("is_extra_credit",False)); widgets_row["is_bonus_var"] = is_bonus_var
            ttk.Checkbutton(self.list_frame, variable=is_bonus_var).grid(row=r_idx, column=4, padx=3)

            del_btn = ttk.Button(self.list_frame, text="X", command=lambda idx=r_idx-1: self.delete_mark_type_at_index(idx), width=3)
            del_btn.grid(row=r_idx, column=5, padx=3)
            self.mark_type_widgets.append(widgets_row) # Store the dict of vars/widgets for this row
        self.list_frame.grid_columnconfigure(1, weight=1) # Allow name entry to expand


    def add_mark_type(self):
        if len(self.mark_types_ref) >= 90: # Or a specific limit for mark types
            messagebox.showwarning("Limit Reached", "Maximum number of mark types reached.", parent=self); return

        new_id_base = "custom_mark"
        new_id_suffix = 1
        while f"{new_id_base}_{new_id_suffix}" in (mt.get("id") for mt in self.mark_types_ref):
            new_id_suffix +=1

        new_mark = {
            "id": f"{new_id_base}_{new_id_suffix}", "name": "New Mark Type", "default_points": 0.0,
            "contributes_to_total": False, "is_extra_credit": False
        }
        self.mark_types_ref.append(new_mark)
        self.mark_types_changed = True
        self.populate_mark_types_ui()

    def delete_mark_type_at_index(self, index):
        if 0 <= index < len(self.mark_types_ref):
            # Check if it's a default one, prevent deletion (or handle carefully)
            # For now, allow deletion, user can reset to defaults
            if messagebox.askyesno("Confirm Delete", f"Delete mark type '{self.mark_types_ref[index]['name']}'?", parent=self):
                del self.mark_types_ref[index]
                self.mark_types_changed = True
                self.populate_mark_types_ui()

    def reset_to_defaults(self):
        if messagebox.askyesno("Confirm Reset", f"Reset all {self.item_type_name.lower()} to application defaults?", parent=self):
            self.mark_types_ref.clear()
            for default_item in self.default_mark_types_const:
                self.mark_types_ref.append(default_item.copy()) # Add copies
            self.mark_types_changed = True
            self.populate_mark_types_ui()

    def apply(self): # OK button
        # Update the list of dicts from the UI widgets
        updated_list = []
        for row_widgets in self.mark_type_widgets:
            # Get ID from label (it's not editable here, but needed for the dict)
            current_id = row_widgets["id_label"].cget("text")
            name = row_widgets["name_var"].get().strip()
            if not name:
                messagebox.showerror("Invalid Name", f"Mark type name for ID '{current_id}' cannot be empty.", parent=self)
                # To prevent dialog closing on error, simpledialog needs validate() to return false.
                # This is tricky here as apply() is called after validate().
                # For now, we'll allow it but it might lead to an empty name. Better: prevent empty.
                return # Or handle error state

            # Check for duplicate names before saving
            if any(item['name'] == name and item['id'] != current_id for item in updated_list):
                messagebox.showerror("Duplicate Name", f"Mark type name '{name}' is already used. Names must be unique.", parent=self)
                return

            updated_list.append({
                "id": current_id,
                "name": name,
                "default_points": row_widgets["points_var"].get(),
                "contributes_to_total": row_widgets["to_total_var"].get(),
                "is_extra_credit": row_widgets["is_bonus_var"].get()
            })

        # Check if actual changes were made before setting the flag
        if self.mark_types_ref != updated_list: # Simple list comparison
            self.mark_types_ref[:] = updated_list # Replace content of original list
            self.mark_types_changed = True

        # self.mark_types_changed flag is now set if there were modifications.

class BulkEditConditionalRulesDialog(simpledialog.Dialog):
    def __init__(self, parent, app, rules_being_edited_copies, original_indices):
        self.app = app
        self.rules_copies = rules_being_edited_copies # These are copies
        self.original_indices = original_indices     # Indices into app.settings["conditional_formatting_rules"]
        self.changes_applied_flag = False

        # Vars for UI choices
        self.enabled_action_var = tk.StringVar(value="no_change") # no_change, set_enabled, set_disabled

        self.times_action_var = tk.StringVar(value="no_change") # no_change, replace, add, clear
        self.new_active_times_for_bulk = [] # List of time slots defined in this dialog
        self.bulk_start_time_var = tk.StringVar()
        self.bulk_end_time_var = tk.StringVar()
        self.bulk_days_vars = [tk.BooleanVar(value=True) for _ in range(7)]

        self.modes_action_var = tk.StringVar(value="no_change") # no_change, replace, add_selected, remove_selected, clear
        self.available_modes_for_bulk = ["behavior", "quiz", "homework", "quiz_session", "homework_session"]
        self.bulk_mode_vars = {mode_name: tk.BooleanVar(value=False) for mode_name in self.available_modes_for_bulk}

        super().__init__(parent, f"Bulk Edit {len(self.rules_copies)} Conditional Rules")

    def body(self, master):
        main_frame = ttk.Frame(master); main_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        notebook = ttk.Notebook(main_frame); notebook.pack(fill=tk.BOTH, expand=True)

        # --- Enabled Tab ---
        enabled_tab = ttk.Frame(notebook); notebook.add(enabled_tab, text="Enabled Status")
        ttk.Radiobutton(enabled_tab, text="Make no change to 'Enabled' status", variable=self.enabled_action_var, value="no_change").pack(anchor=tk.W, pady=2)
        ttk.Radiobutton(enabled_tab, text="Set ALL selected rules to ENABLED", variable=self.enabled_action_var, value="set_enabled").pack(anchor=tk.W, pady=2)
        ttk.Radiobutton(enabled_tab, text="Set ALL selected rules to DISABLED", variable=self.enabled_action_var, value="set_disabled").pack(anchor=tk.W, pady=2)

        # --- Active Times Tab ---
        times_tab = ttk.Frame(notebook); notebook.add(times_tab, text="Active Times")
        ttk.Radiobutton(times_tab, text="Make no change to 'Active Times'", variable=self.times_action_var, value="no_change", command=self._toggle_times_ui).pack(anchor=tk.W, pady=2)
        ttk.Radiobutton(times_tab, text="REPLACE existing time slots with the new set below", variable=self.times_action_var, value="replace", command=self._toggle_times_ui).pack(anchor=tk.W, pady=2)
        ttk.Radiobutton(times_tab, text="ADD the new set below to each rule's existing time slots", variable=self.times_action_var, value="add", command=self._toggle_times_ui).pack(anchor=tk.W, pady=2)
        ttk.Radiobutton(times_tab, text="CLEAR all time slots from selected rules", variable=self.times_action_var, value="clear", command=self._toggle_times_ui).pack(anchor=tk.W, pady=2)

        self.bulk_times_def_frame = ttk.LabelFrame(times_tab, text="Define New Time Slot(s) for Bulk Operation")
        self.bulk_times_def_frame.pack(fill=tk.X, padx=5, pady=5)

        new_ts_frame_bulk = ttk.Frame(self.bulk_times_def_frame)
        new_ts_frame_bulk.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(new_ts_frame_bulk, text="Start (HH:MM):").pack(side=tk.LEFT)
        ttk.Entry(new_ts_frame_bulk, textvariable=self.bulk_start_time_var, width=6).pack(side=tk.LEFT, padx=2)
        ttk.Label(new_ts_frame_bulk, text="End (HH:MM):").pack(side=tk.LEFT, padx=2)
        ttk.Entry(new_ts_frame_bulk, textvariable=self.bulk_end_time_var, width=6).pack(side=tk.LEFT, padx=2)

        bulk_days_frame = ttk.Frame(new_ts_frame_bulk)
        bulk_days_frame.pack(side=tk.LEFT, padx=5)
        days_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        for i, day_name in enumerate(days_names):
            ttk.Checkbutton(bulk_days_frame, text=day_name, variable=self.bulk_days_vars[i]).pack(side=tk.LEFT)
        ttk.Button(new_ts_frame_bulk, text="Add to New Set", command=self._bulk_add_time_slot_to_list).pack(side=tk.LEFT, padx=5)

        self.bulk_times_listbox = tk.Listbox(self.bulk_times_def_frame, height=3, exportselection=False)
        self.bulk_times_listbox.pack(fill=tk.X, expand=True, padx=5, pady=2)
        ttk.Button(self.bulk_times_def_frame, text="Remove from New Set", command=self._bulk_remove_time_slot_from_list).pack(pady=2, padx=5, anchor=tk.W)

        # --- Active Modes Tab ---
        modes_tab = ttk.Frame(notebook); notebook.add(modes_tab, text="Active Modes")
        ttk.Radiobutton(modes_tab, text="Make no change to 'Active Modes'", variable=self.modes_action_var, value="no_change", command=self._toggle_modes_ui).pack(anchor=tk.W, pady=2)
        ttk.Radiobutton(modes_tab, text="REPLACE existing active modes with the selection below", variable=self.modes_action_var, value="replace", command=self._toggle_modes_ui).pack(anchor=tk.W, pady=2)
        ttk.Radiobutton(modes_tab, text="ADD selected modes to each rule's existing active modes", variable=self.modes_action_var, value="add_selected", command=self._toggle_modes_ui).pack(anchor=tk.W, pady=2)
        ttk.Radiobutton(modes_tab, text="REMOVE selected modes from each rule's existing active modes", variable=self.modes_action_var, value="remove_selected", command=self._toggle_modes_ui).pack(anchor=tk.W, pady=2)
        ttk.Radiobutton(modes_tab, text="CLEAR all active modes from selected rules", variable=self.modes_action_var, value="clear", command=self._toggle_modes_ui).pack(anchor=tk.W, pady=2)

        self.bulk_modes_selection_frame = ttk.LabelFrame(modes_tab, text="Select Modes for Bulk Operation")
        self.bulk_modes_selection_frame.pack(fill=tk.X, padx=5, pady=5)
        bulk_modes_cb_frame = ttk.Frame(self.bulk_modes_selection_frame)
        bulk_modes_cb_frame.pack(fill=tk.X, padx=5, pady=5)
        for i, mode_name in enumerate(self.available_modes_for_bulk):
            ttk.Checkbutton(bulk_modes_cb_frame, text=mode_name.replace("_", " ").capitalize(), variable=self.bulk_mode_vars[mode_name]).grid(row=i//3, column=i%3, sticky=tk.W, padx=5, pady=2)

        self._toggle_times_ui() # Initial UI state
        self._toggle_modes_ui() # Initial UI state
        return main_frame

    def _toggle_times_ui(self):
        action = self.times_action_var.get()
        state = tk.NORMAL if action in ["replace", "add"] else tk.DISABLED
        for child in self.bulk_times_def_frame.winfo_children():
            try: child.config(state=state)
            except tk.TclError: pass # Some widgets like LabelFrame itself might not have state
        if state == tk.DISABLED:
            self.bulk_times_listbox.config(state=tk.DISABLED)
        else:
            self.bulk_times_listbox.config(state=tk.NORMAL)


    def _toggle_modes_ui(self):
        action = self.modes_action_var.get()
        state = tk.NORMAL if action in ["replace", "add_selected", "remove_selected"] else tk.DISABLED
        for child in self.bulk_modes_selection_frame.winfo_children():
            # Iterate through checkboxes inside the checkbox frame
            if isinstance(child, ttk.Frame): # This is bulk_modes_cb_frame
                for cb in child.winfo_children():
                    try: cb.config(state=state)
                    except tk.TclError: pass
            else: # Other direct children like the LabelFrame title (which has no state)
                pass


    def _bulk_add_time_slot_to_list(self):
        start_time = self.bulk_start_time_var.get().strip()
        end_time = self.bulk_end_time_var.get().strip()
        try:
            datetime.strptime(start_time, "%H:%M"); datetime.strptime(end_time, "%H:%M")
            if start_time >= end_time: messagebox.showerror("Invalid Time", "Start time must be before end time.", parent=self); return
        except ValueError: messagebox.showerror("Invalid Time Format", "Use HH:MM.", parent=self); return

        selected_days = [i for i, var in enumerate(self.bulk_days_vars) if var.get()]
        if not selected_days: selected_days = list(range(7))

        new_slot = {"start_time": start_time, "end_time": end_time, "days_of_week": selected_days}
        self.new_active_times_for_bulk.append(new_slot)
        self._refresh_bulk_times_listbox()
        self.bulk_start_time_var.set(""); self.bulk_end_time_var.set("")
        for var in self.bulk_days_vars: var.set(True)

    def _bulk_remove_time_slot_from_list(self):
        sel = self.bulk_times_listbox.curselection()
        if not sel: return
        for i in reversed(sel): del self.new_active_times_for_bulk[i]
        self._refresh_bulk_times_listbox()

    def _refresh_bulk_times_listbox(self):
        self.bulk_times_listbox.delete(0, tk.END)
        days_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        for slot in self.new_active_times_for_bulk:
            days_str = ", ".join([days_names[i] for i in slot.get("days_of_week", [])])
            self.bulk_times_listbox.insert(tk.END, f"{slot['start_time']} - {slot['end_time']} ({days_str})")

    def apply(self):
        # Apply changes to the original rules in app.settings
        # The self.rules_copies were just for initial display/safety, not directly modified.

        enabled_action = self.enabled_action_var.get()
        times_action = self.times_action_var.get()
        modes_action = self.modes_action_var.get()

        selected_modes_for_op = [mode for mode, var in self.bulk_mode_vars.items() if var.get()]

        for rule_idx in self.original_indices:
            rule_in_settings = self.app.settings["conditional_formatting_rules"][rule_idx]
            self.changes_applied_flag = True # Assume a change is made if apply is called from OK

            # Enabled status
            if enabled_action == "set_enabled": rule_in_settings["enabled"] = True
            elif enabled_action == "set_disabled": rule_in_settings["enabled"] = False
            # else "no_change"

            # Active Times
            if times_action == "replace":
                rule_in_settings["active_times"] = [ts.copy() for ts in self.new_active_times_for_bulk]
            elif times_action == "add":
                if "active_times" not in rule_in_settings: rule_in_settings["active_times"] = []
                for new_slot in self.new_active_times_for_bulk:
                    if new_slot not in rule_in_settings["active_times"]: # Avoid duplicates if desired
                        rule_in_settings["active_times"].append(new_slot.copy())
            elif times_action == "clear":
                rule_in_settings["active_times"] = []
            # else "no_change"

            # Active Modes
            current_rule_modes = set(rule_in_settings.get("active_modes", []))
            if modes_action == "replace":
                rule_in_settings["active_modes"] = selected_modes_for_op.copy()
            elif modes_action == "add_selected":
                current_rule_modes.update(selected_modes_for_op)
                rule_in_settings["active_modes"] = list(current_rule_modes)
            elif modes_action == "remove_selected":
                current_rule_modes.difference_update(selected_modes_for_op)
                rule_in_settings["active_modes"] = list(current_rule_modes)
            elif modes_action == "clear":
                rule_in_settings["active_modes"] = []
            # else "no_change"

        # If no actual changes were made by any of the operations,
        # the self.changes_applied_flag might be set too eagerly.
        # A more robust check would compare original rules with final state.
        # For now, if user clicks OK, we assume intent to change if actions were selected.


# --- Main Execution ---
if __name__ == "__main__":
    root = tk.Tk()
    from seatingchartmain import SeatingChartApp
    app = SeatingChartApp(root)
    try:
        import darkdetect; import threading
        t = threading.Thread(target=darkdetect.listener, args=(app.theme_auto, ))
        t.daemon = True; t.start()
    except: pass
    root.mainloop()