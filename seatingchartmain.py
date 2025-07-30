import tkinter as tk
from tkinter import ttk, simpledialog, messagebox, filedialog, font as tkfont
import json
import os
import sys
import subprocess
from datetime import datetime, timedelta, date as datetime_date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font as OpenpyxlFont, Alignment as OpenpyxlAlignment
from openpyxl.utils import get_column_letter
import re
import shutil
import zipfile
import csv
import PIL
from PIL import Image
from settingsdialog import SettingsDialog
from commands import Command, DeleteGuideCommand, MoveItemsCommand, AddItemCommand, DeleteItemCommand, LogEntryCommand, \
    LogHomeworkEntryCommand, EditItemCommand, ChangeItemsSizeCommand, MarkLiveQuizQuestionCommand, \
        MarkLiveHomeworkCommand, ChangeStudentStyleCommand, ManageStudentGroupCommand, MoveGuideCommand, AddGuideCommand
from dialogs import PasswordPromptDialog, AddEditStudentDialog, AddFurnitureDialog, BehaviorDialog, \
    ManualHomeworkLogDialog, QuizScoreDialog, LiveQuizMarkDialog, LiveHomeworkMarkDialog, ExitConfirmationDialog, \
        ImportExcelOptionsDialog, SizeInputDialog, StudentStyleDialog,  AttendanceReportDialog, ManageStudentGroupsDialog
from quizhomework import ManageQuizTemplatesDialog, ManageHomeworkTemplatesDialog
from other import FileLockManager, PasswordManager, HelpDialog
from exportdialog import ExportFilterDialog
from data_locker import unlock_file, DATA_FILE
import json
from data_encryption import encrypt_data, decrypt_data
# Replace with your actual path to gswinXXc.exe
#EpsImagePlugin.gs_windows_binary = "C:\\Program Files\\gs\\gs10.05.1\bin\\gswin64c.exe" 
# Only use this ^ if something really doesn't work. Otherwise, it works even with just installing Ghostscript regularly, without any additional steps.
import sv_ttk # For themed widgets
import darkdetect # For dark mode detection
# Conditional import for platform-specific screenshot capability
import threading
import io
import tempfile
import cryptography.fernet # For making sure that the program can properly handle encrypted and non-encrypted data files
try:
    if sys.platform == "win32":
        import win32gui
        import win32ui
        import win32con
    else:
        # For non-Windows, these modules won't be available or needed for the current screenshot method.
        # The postscript-based method is platform-independent.
        win32gui = None
except ImportError:
    win32gui = None # Explicitly set to None if import fails
    print("Warning: win32gui/win32ui/win32con not found. Full window screenshot (deprecated) might not work if called.")

# def listener(callback: typing.Callable[[str], None]) -> None: ...
# TODO: make conditional formatting work by quizzes. add thing for homework also.

# --- Application Constants ---
APP_NAME = "BehaviorLogger"
APP_VERSION = "v57.0" # Version incremented
CURRENT_DATA_VERSION_TAG = "v10" # Incremented for guide saving

# --- Default Configuration ---
DEFAULT_STUDENT_BOX_WIDTH = 130
DEFAULT_STUDENT_BOX_HEIGHT = 80
MIN_STUDENT_BOX_WIDTH = 60
MIN_STUDENT_BOX_HEIGHT = 40
REBBI_DESK_WIDTH = 200
REBBI_DESK_HEIGHT = 100

DEFAULT_FONT_FAMILY = "TkDefaultFont"
DEFAULT_FONT_SIZE = 10
DEFAULT_FONT_COLOR = "black"
DEFAULT_BOX_FILL_COLOR = "skyblue"
DEFAULT_BOX_OUTLINE_COLOR = "blue"
DEFAULT_QUIZ_SCORE_FONT_COLOR = "darkgreen"
DEFAULT_QUIZ_SCORE_FONT_STYLE_BOLD = True
DEFAULT_HOMEWORK_SCORE_FONT_COLOR = "purple" # New for homework scores
DEFAULT_HOMEWORK_SCORE_FONT_STYLE_BOLD = True # New
GROUP_COLOR_INDICATOR_SIZE = 12
DEFAULT_THEME = "System"
THEME_LIST = ["Light", "Dark", "System"]

DRAG_THRESHOLD = 5
DEFAULT_GRID_SIZE = 20
MAX_UNDO_HISTORY_DAYS = 90
LAYOUT_COLLISION_OFFSET = 5
RESIZE_HANDLE_SIZE = 10 # World units for resize handle

# --- Path Handling ---
def get_app_data_path(filename):
    try:
        # Determine base path based on whether the app is frozen (packaged) or running from script
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            # Running as a PyInstaller bundle
            if os.name == 'win32': # Windows
                base_path = os.path.join(os.getenv('APPDATA'), APP_NAME)
            elif sys.platform == 'darwin': # macOS
                base_path = os.path.join(os.path.expanduser('~'), 'Library', 'Application Support', APP_NAME)
            else: # Linux and other Unix-like
                xdg_config_home = os.getenv('XDG_CONFIG_HOME')
                if xdg_config_home:
                    base_path = os.path.join(xdg_config_home, APP_NAME)
                else:
                    base_path = os.path.join(os.path.expanduser('~'), '.config', APP_NAME)
        else:
            # Running as a script, use the script's directory
            base_path = os.path.dirname(os.path.abspath(__file__))

        # Create the base directory if it doesn't exist
        if not os.path.exists(base_path):
            os.makedirs(base_path, exist_ok=True)
        return os.path.join(base_path, filename)
    except Exception as e:
        # Fallback to current working directory if standard paths fail
        print(f"Warning: Could not determine standard app data path due to {e}. Using current working directory as fallback.")
        fallback_path = os.path.join(os.getcwd(), APP_NAME) # Create a subfolder in CWD
        if not os.path.exists(fallback_path):
             os.makedirs(fallback_path, exist_ok=True)
        return os.path.join(fallback_path, filename)

DATA_FILE_PATTERN = f"classroom_data_{CURRENT_DATA_VERSION_TAG}.json"
CUSTOM_BEHAVIORS_FILE_PATTERN = f"custom_behaviors_{CURRENT_DATA_VERSION_TAG}.json"
# NEW: For customizable homework types like "Reading Assignment", "Worksheet"
CUSTOM_HOMEWORK_TYPES_FILE_PATTERN = f"custom_homework_types_{CURRENT_DATA_VERSION_TAG}.json" 
# RENAMED: For customizable homework statuses like "Done", "Not Done"
CUSTOM_HOMEWORK_STATUSES_FILE_PATTERN = f"custom_homework_statuses_{CURRENT_DATA_VERSION_TAG}.json" 
AUTOSAVE_EXCEL_FILE_PATTERN = f"autosave_log_{CURRENT_DATA_VERSION_TAG}.xlsx" # Renamed for clarity
LAYOUT_TEMPLATES_DIR_NAME = "layout_templates"
STUDENT_GROUPS_FILE_PATTERN = f"student_groups_{CURRENT_DATA_VERSION_TAG}.json"
QUIZ_TEMPLATES_FILE_PATTERN = f"quiz_templates_{CURRENT_DATA_VERSION_TAG}.json"
HOMEWORK_TEMPLATES_FILE_PATTERN = f"homework_templates_{CURRENT_DATA_VERSION_TAG}.json" # New

DATA_FILE = get_app_data_path(DATA_FILE_PATTERN)
CUSTOM_BEHAVIORS_FILE = get_app_data_path(CUSTOM_BEHAVIORS_FILE_PATTERN)
CUSTOM_HOMEWORK_TYPES_FILE = get_app_data_path(CUSTOM_HOMEWORK_TYPES_FILE_PATTERN) # NEW
CUSTOM_HOMEWORK_STATUSES_FILE = get_app_data_path(CUSTOM_HOMEWORK_STATUSES_FILE_PATTERN) # RENAMED
AUTOSAVE_EXCEL_FILE = get_app_data_path(AUTOSAVE_EXCEL_FILE_PATTERN)
LAYOUT_TEMPLATES_DIR = get_app_data_path(LAYOUT_TEMPLATES_DIR_NAME)
STUDENT_GROUPS_FILE = get_app_data_path(STUDENT_GROUPS_FILE_PATTERN)
QUIZ_TEMPLATES_FILE = get_app_data_path(QUIZ_TEMPLATES_FILE_PATTERN)
HOMEWORK_TEMPLATES_FILE = get_app_data_path(HOMEWORK_TEMPLATES_FILE_PATTERN) # New
LOCK_FILE_PATH = get_app_data_path(f"{APP_NAME}.lock") # Lock file
IMAGENAMEW = "export_layout_as_image_helper"

if not os.path.exists(LAYOUT_TEMPLATES_DIR):
    os.makedirs(LAYOUT_TEMPLATES_DIR, exist_ok=True)

DEFAULT_BEHAVIORS_LIST = [
    "Talking", "Off Task", "Out of Seat", "Uneasy", "Placecheck",
    "Great Participation", "Called On", "Complimented", "Fighting", "Other"
]

# for manual detailed logging, i would want these | currently used for sessions (yes/no)
DEFAULT_HOMEWORK_TYPES_LIST = [ # For live session "Yes/No" mode AND now for manual logging
    "Reading Assignment", "Worksheet", "Math Problems", "Project Work", "Study for Test"
]

# these would only be for behavior-like logging (not detailed)
DEFAULT_HOMEWORK_LOG_BEHAVIORS = [ # For manual homework logging (like behavior logging)
    "Done", "Not Done", "Partially Done", "Signed", "Returned", "Late", "Excellent Work"
]

# Make these be used for the select AND Yes/No Live Homework session modes.
DEFAULT_HOMEWORK_SESSION_BUTTONS = [ # For live session "Select" mode
    {"name": "Done"}, {"name": "Not Done"}, {"name": "Signed"}, {"name": "Returned"}
]

DEFAULT_HOMEWORK_SESSION_BUTTONS2 = [ # For live session "Select" mode
    "Done", "Not Done", "Signed", "Returned"
]

DEFAULT_HOMEWORK_STATUSES = [
    "Done", "Not Done", "Partially Done", "Signed", "Returned", "Late", "Excellent Work"
]

DEFAULT_GROUP_COLORS = ["#FFADAD", "#FFD6A5", "#FDFFB6", "#CAFFBF", "#9BF6FF", "#A0C4FF", "#BDB2FF", "#FFC6FF", "#E0E0E0"]

DEFAULT_QUIZ_MARK_TYPES = [
    {"id": "mark_correct", "name": "Correct", "contributes_to_total": True, "is_extra_credit": False, "default_points": 1},
    {"id": "mark_incorrect", "name": "Incorrect", "contributes_to_total": True, "is_extra_credit": False, "default_points": 0},
    {"id": "mark_partial", "name": "Partial Credit", "contributes_to_total": True, "is_extra_credit": False, "default_points": 0.5},
    {"id": "extra_credit", "name": "Bonus", "contributes_to_total": False, "is_extra_credit": True, "default_points": 1}
]
DEFAULT_HOMEWORK_MARK_TYPES = [ # New for homework marks
    {"id": "hmark_complete", "name": "Complete", "default_points": 10},
    {"id": "hmark_incomplete", "name": "Incomplete", "default_points": 5},
    {"id": "hmark_notdone", "name": "Not Done", "default_points": 0},
    {"id": "hmark_effort", "name": "Effort Score (1-5)", "default_points": 3} # Example
]

MAX_CUSTOM_TYPES = 90 # Max for custom behaviors, homeworks, mark types

MASTER_RECOVERY_PASSWORD_HASH = "5bf881cb69863167a3172fda5c552694a3328548a43c7ee258d6d7553fc0e1a1a8bad378fb131fbe10e37efbd9e285b22c29b75d27dcc2283d48d8edf8063292" # SHA256 of "RecoverMyData123!" # It's actually SHA3_512 or something else, which only Yaakov Maimon (me) has, although you can add your own if you like

def levenshtein_distance(s1, s2):
    if len(s1) < len(s2):
        return levenshtein_distance(s2, s1)
    if len(s2) == 0:
        return len(s1)
    previous_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        current_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = previous_row[j + 1] + 1
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row
    return previous_row[-1]

def name_similarity_ratio(s1, s2):
    """Calculates similarity ratio between 0 and 1 based on Levenshtein distance."""
    if not s1 and not s2: return 1.0 # Both empty
    if not s1 or not s2: return 0.0   # One empty
    distance = levenshtein_distance(s1.lower(), s2.lower())
    max_len = max(len(s1), len(s2))
    if max_len == 0: return 1.0 # Should be caught by above, but defensive
    return 1.0 - (distance / max_len)

# --- Main Application Class ---
class SeatingChartApp:
    def __init__(self, root_window):
        # ... (initial part of __init__ is the same) ...
        self.root = root_window 
        self.root.title(f"Classroom Behavior Tracker - {APP_NAME} - {APP_VERSION}")
        self.root.geometry("1400x980")
        if sys.platform == "win32":
            self.root.state('zoomed') # Maximizes the window on Windows
        self.file_lock_manager = FileLockManager(LOCK_FILE_PATH)
        if not self.file_lock_manager.acquire_lock():
            self.root.destroy()
            sys.exit(1)

        self.is_beginning = True
        
        self.students = {}
        self.furniture = {}
        self.behavior_log = []
        self.homework_log = []
        self.student_groups = {}
        self.quiz_templates = {}
        self.homework_templates = {}

        self.next_student_id_num = 1
        self.next_furniture_id_num = 1
        self.next_group_id_num = 1
        self.next_quiz_template_id_num = 1
        self.next_homework_template_id_num = 1

        self.all_behaviors = []
        self.custom_behaviors = []
        
        # RENAMED/REPURPOSED
        self.all_homework_types = [] # For "Reading Assignment", "Worksheet", etc.
        self.custom_homework_types = []

        self.all_homework_statuses = [] # For "Done", "Not Done", etc.
        self.custom_homework_statuses = []
        
        # This list is now derived from all_homework_types
        self.all_homework_session_types = [] 

        self.last_excel_export_path = None
        self.selected_items = set()
        self.undo_stack = []
        self.redo_stack = []
        self.type_theme = "sv_ttk"
        try:
            self.theme_style_using = sv_ttk.get_theme()
        except:
            self.theme_style_using = "System"
        
        self.settings = self._get_default_settings()
        self.password_manager = PasswordManager(self.settings)

        self.canvas_frame = None; self.canvas = None; self.h_scrollbar = None; self.v_scrollbar = None
        self.status_bar_label = None; self.zoom_display_label = None
        self.mode_var = tk.StringVar(value=self.settings["current_mode"])
        self.edit_mode_var = tk.BooleanVar(value=False)

        self.drag_data = {"x": 0, "y": 0, "item_id": None, "item_type": None,
                          "start_x_world": 0, "start_y_world": 0, # Start of drag in world coords
                          "original_positions": {}, "is_resizing": False, "resize_handle_corner": None,
                          "original_size_world": {}} # Original W/H in world coords
        self._potential_click_target = None
        self._drag_started_on_item = False
        self._recent_incidents_hidden_globally = False
        self._recent_homeworks_hidden_globally = False # New
        self._per_student_last_cleared = {}

        self.last_used_quiz_name = ""
        self.initial_num_questions = "" # For quiz
        self.last_used_quiz_name_timestamp = None

        # Attributes for remembering the last used homework name and items for manual logging
        self.last_used_homework_name = ""
        self.initial_num_homework_items = "" # For manual log, if needed in future
        self.last_used_homework_name_timestamp = None

        self.is_live_quiz_active = False
        self.current_live_quiz_name = ""
        self.live_quiz_scores = {}

        self.is_live_homework_active = False # New
        self.current_live_homework_name = "" # New
        self.live_homework_scores = {} # New {student_id: {"homework_type_name": "yes/no/selected_option", ...} or {"selected_options": [...]}}

        self.current_zoom_level = 1.0
        self.canvas_orig_width = 2000
        self.canvas_orig_height = 1500
        self.custom_canvas_color = None
        
        self.zoom_level = 1.0
        self.pan_x = 0.0
        self.pan_y = 0.0

        # Ruler and guide attributes
        self.guides = {} # To store {'id': str, 'type': 'v'/'h', 'world_coord': float, 'canvas_item_id': int}
        self.next_guide_id_num = 1
        self.add_guide_mode: Optional[str] = None # 'vertical', 'horizontal', or None
        self.active_guide_button: Optional[ttk.Button] = None # To manage button visual state

        self.ruler_thickness = 35  # pixels
        self.ruler_bg_color = "#f0f0f0"
        self.ruler_line_color = "#555555"
        self.ruler_text_color = "#333333"
        self.active_ruler_guide_coord_x: Optional[float] = None
        self.active_ruler_guide_coord_y: Optional[float] = None
        self.temporary_guides: List[Dict[str, Any]] = [] # List of {'type': 'h'/'v', 'world_coord': float, 'canvas_id': int}
        
        
        self.load_custom_behaviors()
        self.load_custom_homework_types() # NEW
        self.load_custom_homework_statuses() # RENAMED
        self.load_student_groups()
        self.load_quiz_templates()
        self.load_homework_templates()
        
        self.update_all_behaviors()
        self.update_all_homework_types() # NEW
        self.update_all_homework_statuses() # RENAMED
        self.update_all_homework_session_types() # This now depends on the others

        self.load_data() # Loads main data, including settings
        self.settings["available_fonts"] = sorted(list(tkfont.families()))
        self._ensure_next_ids()
        self.theme_auto(init=True)

        self.guide_line_color = self.settings.get("guides_color", "blue")
        self.setup_ui()
        # self.root.after_idle(self.draw_all_items) # Defer initial draw until window is mapped
        self.update_status(f"Application started. Data loaded from: {os.path.dirname(DATA_FILE)}") # type: ignore
        self.update_undo_redo_buttons_state()
        self.toggle_mode(initial=True) # Apply initial mode
        self.root.after(30000, self.periodic_checks)
        self.root.after(self.settings.get("autosave_interval_ms", 30000), self.autosave_data_wrapper)
        
        # Schedule the first time-based formatting update to align with the clock.
        now = datetime.now()
        seconds_until_first_minute = 60 - now.second
        milliseconds_until_first_minute = (seconds_until_first_minute * 1000) + 50 # Add 50ms buffer
        self.root.after(milliseconds_until_first_minute, self.update_time_based_formatting)

        self.root.protocol("WM_DELETE_WINDOW", self.on_exit_protocol)
        
        if self.password_manager.is_password_set() and self.settings.get("password_on_open", False):
            self.root.withdraw()
            if not self.prompt_for_password("Application Locked", "Enter password to open:"):
                self.on_exit_protocol(force_quit=True) # Ensure lock is released if exit fails here
            self.root.deiconify()

    def on_canvas_configure(self, event):
        """
        Called when the canvas is first configured or resized.
        We use this to trigger the very first draw_all_items() call,
        ensuring the canvas has its final size.
        """
        self.draw_all_items()
        
    def capture_tkinter_window(self, filename="tkinter_screenshot.png"):
        """
        Captures the content of a specific Tkinter window by its handle.
        This method is now deprecated in favor of export_layout_as_image for full canvas capture.
        """
        root_window = self.root
        hwnd = root_window.winfo_id() # Get the window handle (HWND) of the Tkinter root

        # Get the window dimensions
        left, top, right, bot = win32gui.GetWindowRect(hwnd)
        width = right - left
        height = bot - top

        # Create a device context (DC) for the window
        hdc = win32gui.GetWindowDC(hwnd)
        dc_obj = win32ui.CreateDCFromHandle(hdc)
        mem_dc = dc_obj.CreateCompatibleDC()

        # Create a bitmap object
        bitmap = win32ui.CreateBitmap()
        bitmap.CreateCompatibleBitmap(dc_obj, width, height)
        mem_dc.SelectObject(bitmap)

        # Copy the window content to the bitmap
        # Note: PrintWindow is generally better as it can capture minimized/obscured windows,
        # but GetWindowDC + BitBlt works well for visible windows.
        # For PrintWindow, you might need to find the exact flags.
        mem_dc.BitBlt((0, 0), (width, height), dc_obj, (0, 0), win32con.SRCCOPY)

        # Get the bitmap bits and create a PIL Image
        bmpinfo = bitmap.GetInfo()
        bmpstr = bitmap.GetBitmapBits(True)
        img = Image.frombuffer(
            'RGB',
            (bmpinfo['bmWidth'], bmpinfo['bmHeight']),
            bmpstr, 'raw', 'BGRX', 0, 1)

        # Clean up the DCs
        dc_obj.DeleteDC()
        mem_dc.DeleteDC()
        win32gui.ReleaseDC(hwnd, hdc)

        #print(os.listdir(os.path.dirname(get_app_data_path(filename))))
        default_filename = f"app_screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        file_path = filedialog.asksaveasfilename(defaultextension=".png", initialfile=default_filename + ".png",
                                                       filetypes=[("Image files", "*.png"), ("All Files", "*.*")], parent=self.root)
        
        #if filename in os.listdir(os.path.dirname(get_app_data_path(filename))): print(filename, 1); filename = "tkinter_screenshot3.png"; print(filename, 2) 
            
        output_dpi = int(self.settings.get("output_dpi", 600))
        #print(path)
        #img.save(output_image_file, dpi=(output_dpi, output_dpi))
        img.save(file_path, dpi=(output_dpi, output_dpi))
        print(f"Screenshot saved to {filename}")
        self.update_status(f"Screenshot saved to {file_path}")
        if messagebox.askyesno("Export Successful", f"Layout image saved to:\n{file_path}\n\nDo you want to open the file location?", parent=self.root): 
            self.open_specific_export_folder(file_path)

    def _read_and_decrypt_file(self, file_path):
        """Reads a file, attempts to decrypt it, and loads the JSON data."""
        if not os.path.exists(file_path):
            return None
        try:
            with open(file_path, 'rb') as f:
                file_content = f.read()
            
            if not file_content: # File is empty
                return None

            try:
                # Attempt to decrypt first
                decrypted_data_string = decrypt_data(file_content)
            except cryptography.fernet.InvalidToken:
                # If decryption fails, it's likely plaintext (or corrupt)
                # Assume it's a UTF-8 encoded string.
                decrypted_data_string = file_content.decode('utf-8')

            return json.loads(decrypted_data_string)

        except (json.JSONDecodeError, IOError, UnicodeDecodeError) as e:
            print(f"Error loading and decoding file {os.path.basename(file_path)}: {e}")
            return None

    def _encrypt_and_write_file(self, file_path, data_to_write):
        """Encodes data to JSON, encrypts if enabled, and writes to a file."""
        try:
            json_data_string = json.dumps(data_to_write, indent=4)
            
            # Use the app's setting to decide whether to encrypt
            if self.settings.get("encrypt_data_files", True):
                data_to_write_bytes = encrypt_data(json_data_string)
            else:
                data_to_write_bytes = json_data_string.encode('utf-8')

            with open(file_path, 'wb') as f:
                f.write(data_to_write_bytes)

        except IOError as e:
            print(f"Error saving file {os.path.basename(file_path)}: {e}")
        except Exception as e:
            print(f"An unexpected error occurred while saving {os.path.basename(file_path)}: {e}")

    def _get_default_settings(self):
        return {
            "show_recent_incidents_on_boxes": True,
            "num_recent_incidents_to_show": 2,
            "recent_incident_time_window_hours": 24,
            "show_full_recent_incidents": False,
            "reverse_incident_order": True,
            "selected_recent_behaviors_filter": None, # List of behavior names, or None for all

            "show_recent_homeworks_on_boxes": True, # New
            "num_recent_homeworks_to_show": 2, # New
            "recent_homework_time_window_hours": 24, # New
            "show_full_recent_homeworks": False, # New
            "reverse_homework_order": True, # New
            "selected_recent_homeworks_filter": None, # New

            "autosave_interval_ms": 30000,
            "default_student_box_width": DEFAULT_STUDENT_BOX_WIDTH,
            "default_student_box_height": DEFAULT_STUDENT_BOX_HEIGHT,
            "student_box_fill_color": DEFAULT_BOX_FILL_COLOR,
            "student_box_outline_color": DEFAULT_BOX_OUTLINE_COLOR,
            "student_font_family": DEFAULT_FONT_FAMILY,
            "student_font_size": DEFAULT_FONT_SIZE,
            "student_font_color": DEFAULT_FONT_COLOR,
            "grid_snap_enabled": False,
            "grid_size": DEFAULT_GRID_SIZE,
            "behavior_initial_overrides": {},
            "homework_initial_overrides": {}, # New for homework display initials
            "current_mode": "behavior", # "behavior", "quiz", or "homework"
            "max_undo_history_days": MAX_UNDO_HISTORY_DAYS,
            "conditional_formatting_rules": [], # Each rule will be a dict. See ConditionalFormattingRuleDialog
            # Example rule:
            # {
            #  "type": "group", "group_id": "group_1", "color": "#FF0000", "outline": "#AA0000",
            #  "enabled": True, "active_times": [], "active_modes": []
            # }
            # {
            #  "type": "behavior_count", "behavior_name": "Talking", "count_threshold": 3, "time_window_hours": 2,
            #  "color": "#FFFF00", "outline": null,
            #  "enabled": True, "active_times": [{"start_time": "09:00", "end_time": "10:30", "days_of_week": [0,1,2,3,4]}],
            #  "active_modes": ["behavior"]
            # }
            "student_groups_enabled": True,
            "show_zoom_level_display": True,
            "available_fonts": [], # Updated: populated later

            # Quiz specific
            "default_quiz_name": "Pop Quiz",
            "last_used_quiz_name_timeout_minutes": 60, # Timeout for remembering quiz name
            "show_recent_incidents_during_quiz": True,
            "live_quiz_score_font_color": DEFAULT_QUIZ_SCORE_FONT_COLOR,
            "live_quiz_score_font_style_bold": DEFAULT_QUIZ_SCORE_FONT_STYLE_BOLD,
            "quiz_mark_types": DEFAULT_QUIZ_MARK_TYPES.copy(),
            "default_quiz_questions": 10,
            "quiz_score_calculation": "percentage",
            "combine_marks_for_display": True,
            "live_quiz_questions": 5,
            "live_quiz_initial_color": "#FF0000",
            "live_quiz_final_color": "#00FF00",

            # Homework specific (New)
            "default_homework_name": "Homework Check", # Default name for manual log & live session
            "last_used_homework_name_timeout_minutes": 60, # Timeout for remembering homework name (manual log)
            "behavior_log_font_size": DEFAULT_FONT_SIZE -1, # Specific font size for behavior log text
            "quiz_log_font_size": DEFAULT_FONT_SIZE,       # Specific font size for quiz log text
            "homework_log_font_size": DEFAULT_FONT_SIZE -1, # Specific font size for homework log text
            "live_homework_session_mode": "Yes/No", # "Yes/No" or "Select"
            "log_homework_marks_enabled": True, # Enable/disable detailed marks for manual log
            "homework_mark_types": DEFAULT_HOMEWORK_MARK_TYPES.copy(),
            "default_homework_items_for_yes_no_mode": 5, # For live session "Yes/No"
            "live_homework_score_font_color": DEFAULT_HOMEWORK_SCORE_FONT_COLOR,
            "live_homework_score_font_style_bold": DEFAULT_HOMEWORK_SCORE_FONT_STYLE_BOLD,


            # Password settings
            "app_password_hash": None,
            "password_on_open": False,
            "password_on_edit_action": False,
            "password_auto_lock_enabled": False,
            "password_auto_lock_timeout_minutes": 15,

            # Next ID counters (managed by _ensure_next_ids but good to have defaults)
            "next_student_id_num": 1,
            "next_furniture_id_num": 1,
            "next_group_id_num": 1,
            "next_quiz_template_id_num": 1,
            "next_homework_template_id_num": 1, # New
            "next_custom_homework_type_id_num": 1, # For custom homework types in Yes/No mode

            # Internal state storage (prefixed with underscore)
            "_last_used_quiz_name_for_session": "", # Stores last used quiz name for manual log
            "_last_used_quiz_name_timestamp_for_session": None, # Timestamp for quiz name timeout
            "_last_used_q_num_for_session": 10, # Stores last used num questions for manual quiz log

            "_last_used_homework_name_for_session": "", # Stores last used homework name for manual log
            "_last_used_homework_name_timestamp_for_session": None, # Timestamp for homework name timeout
            "_last_used_hw_items_for_session": 5, # Stores last used num items for manual homework log
            "theme": "System", # Newer
            "enable_text_background_panel": True, # Default for the new setting
            "show_rulers": False, # Default for rulers
            "show_grid": False, # Default for grid visibility
            "grid_color": "#000000", # Default light gray for grid lines
            "save_guides_to_file": True, # New setting for guides
            "guides_stay_when_rulers_hidden": True, # New setting for guides
            "next_guide_id_num": 1, # Added in migration, also good here
            "guides_color": "blue", # Default color for guides
            "hidden_default_homework_types": [], # New for hiding default homework types
        }

    def _ensure_next_ids(self):
        # Student IDs
        max_s_id = 0
        for sid in self.students:
            if sid.startswith("student_"):
                try: max_s_id = max(max_s_id, int(sid.split("_")[1]))
                except (ValueError, IndexError): pass
        self.next_student_id_num = max(self.settings.get("next_student_id_num", 1), max_s_id + 1)
        self.settings["next_student_id_num"] = self.next_student_id_num

        # Furniture IDs
        max_f_id = 0
        for fid in self.furniture:
            if fid.startswith("furniture_"):
                try: max_f_id = max(max_f_id, int(fid.split("_")[1]))
                except (ValueError, IndexError): pass
        self.next_furniture_id_num = max(self.settings.get("next_furniture_id_num", 1), max_f_id + 1)
        self.settings["next_furniture_id_num"] = self.next_furniture_id_num

        # Group IDs
        max_g_id = 0
        for gid in self.student_groups:
            if gid.startswith("group_"):
                try: max_g_id = max(max_g_id, int(gid.split("_")[1]))
                except (ValueError, IndexError): pass
        self.next_group_id_num = max(self.settings.get("next_group_id_num", 1), max_g_id + 1)
        self.settings["next_group_id_num"] = self.next_group_id_num

        # Quiz Template IDs
        max_qt_id = 0
        for qtid in self.quiz_templates:
            if qtid.startswith("quiztemplate_"):
                try: max_qt_id = max(max_qt_id, int(qtid.split("_")[1]))
                except (ValueError, IndexError): pass
        self.next_quiz_template_id_num = max(self.settings.get("next_quiz_template_id_num", 1), max_qt_id + 1)
        self.settings["next_quiz_template_id_num"] = self.next_quiz_template_id_num

        # Homework Template IDs (New)
        max_ht_id = 0
        for htid in self.homework_templates:
            if htid.startswith("hwtemplate_"):
                try: max_ht_id = max(max_ht_id, int(htid.split("_")[1]))
                except (ValueError, IndexError): pass
        self.next_homework_template_id_num = max(self.settings.get("next_homework_template_id_num", 1), max_ht_id + 1)
        self.settings["next_homework_template_id_num"] = self.next_homework_template_id_num

        # Custom Homework Type IDs (MODIFIED)
        max_chwt_id = 0
        # Use the new custom_homework_types attribute here
        for chwt in self.custom_homework_types:
            if isinstance(chwt, dict) and chwt.get('id', '').startswith("hwtype_"):
                try: max_chwt_id = max(max_chwt_id, int(chwt['id'].split("_")[1]))
                except (ValueError, IndexError): pass
        self.settings["next_custom_homework_type_id_num"] = max(self.settings.get("next_custom_homework_type_id_num", 1), max_chwt_id + 1)

        # Guide IDs
        max_g_id_num = 0
        # Ensure self.guides exists and is iterable; it might not be populated if loading very old data before migration
        if hasattr(self, 'guides') and isinstance(self.guides, dict):
            for guide_info in self.guides:
                guide_id_str = self.guides[guide_info].get('id', '') #guide_info.get('id', '')
                if guide_id_str.startswith("guide_v_") or guide_id_str.startswith("guide_h_"):
                    try: max_g_id_num = max(max_g_id_num, int(guide_id_str.split("_")[-1]))
                    except (ValueError, IndexError, TypeError): pass

        # self.next_guide_id_num is now a direct attribute, not in settings dictionary initially
        # It will be loaded from the main data file if present, otherwise defaults to 1.
        # This ensures it's correctly set after loading data that might contain guides.
        self.next_guide_id_num = max(getattr(self, 'next_guide_id_num', 1), max_g_id_num + 1)

    def periodic_checks(self):
        self.password_manager.check_auto_lock()
        if self.password_manager.is_locked and not hasattr(self, '_lock_screen_active'):
            self.show_lock_screen()
        self.root.after(30000, self.periodic_checks)

    def update_time_based_formatting(self):
        """Periodically redraws all items to update time-based conditional formatting. Runs on the minute."""
        # This check is to avoid redrawing if no time-based rules exist.
        if any(rule.get("active_times") for rule in self.settings.get("conditional_formatting_rules", [])):
            self.draw_all_items()
        
        # Schedule the next run for the start of the next minute.
        now = datetime.now()
        seconds_until_next_minute = 60 - now.second
        milliseconds_until_next_minute = (seconds_until_next_minute * 1000) + 50 # Add 50ms buffer
        self.root.after(milliseconds_until_next_minute, self.update_time_based_formatting)

    def show_lock_screen(self):
        if hasattr(self, '_lock_screen_active') and self._lock_screen_active.winfo_exists(): return
        self._lock_screen_active = tk.Toplevel(self.root)
        self._lock_screen_active.title("Application Locked")
        self._lock_screen_active.transient(self.root)
        self._lock_screen_active.grab_set()
        self._lock_screen_active.protocol("WM_DELETE_WINDOW", lambda: None)
        self.root.update_idletasks()
        px, py, pw, ph = self.root.winfo_x(), self.root.winfo_y(), self.root.winfo_width(), self.root.winfo_height()
        dw, dh = 300, 150
        x, y = px + (pw // 2) - (dw // 2), py + (ph // 2) - (dh // 2)
        self._lock_screen_active.geometry(f"{dw}x{dh}+{x}+{y}")
        self._lock_screen_active.resizable(False, False)
        ttk.Label(self._lock_screen_active, text="Application is locked. Enter password:", font=("", 12)).pack(pady=10)
        password_var = tk.StringVar()
        password_entry = ttk.Entry(self._lock_screen_active, textvariable=password_var, show="*", width=30)
        password_entry.pack(pady=5); password_entry.focus_set()
        status_label_lock = ttk.Label(self._lock_screen_active, text="", foreground="red"); status_label_lock.pack(pady=2)
        def attempt_unlock():
            if self.password_manager.unlock_application(password_var.get()):
                self._lock_screen_active.destroy(); del self._lock_screen_active
                self.update_status("Application unlocked.")
            else:
                status_label_lock.config(text="Incorrect password."); password_entry.select_range(0, tk.END)
        ttk.Button(self._lock_screen_active, text="Unlock", command=attempt_unlock).pack(pady=10)
        self._lock_screen_active.bind('<Return>', lambda e: attempt_unlock())

    def prompt_for_password(self, title, prompt_message, for_editing=False):
        if self.password_manager.is_locked:
             if not hasattr(self, '_lock_screen_active') or not self._lock_screen_active.winfo_exists(): self.show_lock_screen()
             return not self.password_manager.is_locked
        if for_editing and not self.settings.get("password_on_edit_action", False) and not self.password_manager.is_password_set(): return True
        if not self.password_manager.is_password_set(): return True
        dialog = PasswordPromptDialog(self.root, title, prompt_message, self.password_manager)
        return dialog.result

    def execute_command(self, command: Command):
        is_sensitive_edit = isinstance(command, (AddItemCommand, DeleteItemCommand, EditItemCommand, ChangeItemsSizeCommand, ManageStudentGroupCommand))
        if is_sensitive_edit and self.settings.get("password_on_edit_action", False) and self.password_manager.is_password_set():
            if not self.prompt_for_password("Confirm Action", "Enter password to make this change:", for_editing=True):
                self.update_status("Action cancelled: Password not provided or incorrect."); return
        try:
            command.execute()
            self.undo_stack.append(command)
            self.redo_stack.clear()
            self.update_undo_redo_buttons_state()
            if not isinstance(command, (MarkLiveQuizQuestionCommand, MarkLiveHomeworkCommand)):
                self.save_data_wrapper(source="command_execution")
            self.password_manager.record_activity()
        except Exception as e:
            messagebox.showerror("Command Error", f"Error executing command: {e}\nCommand Type: {type(command).__name__}", parent=self.root)
            print(f"Command execution error: {e}\n{type(command)}")

    def undo_last_action(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock Required", "Enter password to undo action:"): return # type: ignore
        if self.undo_stack:
            command = self.undo_stack.pop()
            try:
                command.undo()
                self.redo_stack.append(command)
                self.update_undo_redo_buttons_state()
                if not isinstance(command, (MarkLiveQuizQuestionCommand, MarkLiveHomeworkCommand)):
                    self.save_data_wrapper(source="undo_command")
                self.draw_all_items()
                self.password_manager.record_activity()
            except Exception as e:
                messagebox.showerror("Undo Error", f"Error undoing action: {e}", parent=self.root)
                self.undo_stack.append(command); print(f"Undo error: {e}\n{type(command)}")

    def redo_last_action(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock Required", "Enter password to redo action:"): return
        if self.redo_stack:
            command = self.redo_stack.pop()
            try:
                command.execute()
                self.undo_stack.append(command)
                self.update_undo_redo_buttons_state()
                if not isinstance(command, (MarkLiveQuizQuestionCommand, MarkLiveHomeworkCommand)):
                    self.save_data_wrapper(source="redo_command")
                self.draw_all_items()
                self.password_manager.record_activity()
            except Exception as e:
                messagebox.showerror("Redo Error", f"Error redoing action: {e}", parent=self.root)
                self.redo_stack.append(command); print(f"Redo error: {e}\n{type(command)}")

    def update_undo_redo_buttons_state(self):
        if hasattr(self, 'undo_btn'): self.undo_btn.config(state=tk.NORMAL if self.undo_stack else tk.DISABLED)
        if hasattr(self, 'redo_btn'): self.redo_btn.config(state=tk.NORMAL if self.redo_stack else tk.DISABLED)

    def get_new_student_id(self):
        current_id_to_assign = self.next_student_id_num
        return f"student_{current_id_to_assign}", self.next_student_id_num + 1
    def get_new_furniture_id(self):
        current_id_to_assign = self.next_furniture_id_num
        return f"furniture_{current_id_to_assign}", self.next_furniture_id_num + 1
    def get_new_group_id(self):
        current_id_to_assign = self.next_group_id_num
        return f"group_{current_id_to_assign}", self.next_group_id_num + 1
    def get_new_quiz_template_id(self):
        current_id_to_assign = self.next_quiz_template_id_num
        return f"quiztemplate_{current_id_to_assign}", self.next_quiz_template_id_num + 1
    def get_new_homework_template_id(self): # New
        current_id_to_assign = self.next_homework_template_id_num
        return f"hwtemplate_{current_id_to_assign}", self.next_homework_template_id_num + 1
    def get_new_custom_homework_type_id(self): # New
        current_id_to_assign = self.settings.get("next_custom_homework_type_id_num", 1)
        return f"hwtype_{current_id_to_assign}", current_id_to_assign + 1

    def update_status(self, message):
        if self.status_bar_label: self.status_bar_label.configure(text=message)

    def setup_ui(self):
        self.main_frame = ttk.Frame(self.root, padding="5"); self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        self.top_frame = ScrollableToolbar(self.main_frame); self.top_frame.pack(side=tk.TOP, fill=tk.X, pady=(0,2))
        
        top_controls_frame_row1 = ttk.Frame(self.top_frame.interior); top_controls_frame_row1.pack(side=tk.TOP, fill=tk.X, pady=(0, 2))
        
        self.undo_btn = ttk.Button(top_controls_frame_row1, text="Undo", command=self.undo_last_action, state=tk.DISABLED)
        self.undo_btn.pack(side=tk.LEFT, padx=2)
        self.redo_btn = ttk.Button(top_controls_frame_row1, text="Redo", command=self.redo_last_action, state=tk.DISABLED)
        self.redo_btn.pack(side=tk.LEFT, padx=2)

        self.file_menu_btn = ttk.Menubutton(top_controls_frame_row1, text="File"); self.file_menu = tk.Menu(self.file_menu_btn, tearoff=0)
        self.file_menu.add_command(label="Save Now", command=self.save_data_wrapper, accelerator="Ctrl+S")
        self.file_menu.add_command(label="Import Students from Excel...", command=self.import_students_from_excel_dialog)
        self.file_menu.add_separator(); self.file_menu.add_command(label="Open Data Folder", command=self.open_data_folder)
        self.open_export_folder_menu_entry_index = self.file_menu.index(tk.END)
        self.file_menu.add_command(label="Open Last Export Folder (None)", command=self.open_last_export_folder, state=tk.DISABLED)
        self.open_export_folder_menu_entry_index = self.file_menu.index(tk.END)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Backup All Application Data (.zip)...", command=self.backup_all_data_dialog)
        self.file_menu.add_command(label="Restore All Application Data...", command=self.restore_all_data_dialog)
        self.file_menu.add_separator(); self.file_menu.add_command(label="Reset Application (Caution!)...", command=self.reset_application_dialog)
        self.file_menu.add_command(label="Import data from json (Caution!)...", command=self.import_data)
        self.file_menu.add_separator(); self.file_menu.add_command(label="Exit", command=self.on_exit_protocol, accelerator="Ctrl+Q")
        self.file_menu_btn["menu"] = self.file_menu
        self.file_menu_btn.pack(side=tk.LEFT, padx=2)
        self.update_open_last_export_folder_menu_item()
        self.root.bind_all("<Control-s>", lambda event: self.save_data_wrapper())
        self.root.bind_all("<Control-q>", lambda event: self.save_and_quit_app())
        self.root.bind_all("<Control-Q>", lambda event: self.save_and_quit_app())
        self.root.bind_all("<Control-z>", lambda event: self.undo_last_action())
        self.root.bind_all("<Control-y>", lambda event: self.redo_last_action())
        self.root.bind_all("<Control-Shift-Z>", lambda event: self.redo_last_action()) # Common alternative for redo
        self.root.bind_all("<Control-r>", lambda event: self.reload_canvas())
        self.root.bind_all("<Control-R>", lambda event: self.reload_canvas())
        self.root.bind_all("<S>", lambda event: self.open_settings_dialog())
        self.root.bind_all("<p>", lambda event: self.show_help_dialog())
        self.root.bind_all("<Control-plus>", lambda event: self.zoom_canvas(1.1))
        self.root.bind_all("<Control-equal>", lambda event: self.zoom_canvas(1.1))
        self.root.bind_all("<Control-minus>", lambda event: self.zoom_canvas(0.9))
        self.root.bind_all("<Control-0>", lambda event: self.zoom_canvas(0))
        self.root.bind_all("<E>", lambda event: self.toggle_edit_mode_shortcut())
        self.root.bind_all("<B>", lambda event: self.toggle_mode_("behavior"))
        self.root.bind_all("<Q>", lambda event: self.toggle_mode_("quiz"))
        self.root.bind_all("<H>", lambda event: self.toggle_mode_("homework"))
        
        # Edit Menu (for Undo History)
        # self.edit_menu_btn = ttk.Menubutton(top_controls_frame_row1, text="Edit")
        # self.edit_menu = tk.Menu(self.edit_menu_btn, tearoff=0)
        # self.edit_menu.add_command(label="Undo", command=self.undo_last_action, accelerator="Ctrl+Z")
        # self.edit_menu.add_command(label="Redo", command=self.redo_last_action, accelerator="Ctrl+Y")
        # self.edit_menu.add_separator()
        # self.edit_menu.add_command(label="Show Undo History...", command=self.show_undo_history_dialog)
        # self.edit_menu_btn["menu"] = self.edit_menu
        # self.edit_menu_btn.pack(side=tk.LEFT, padx=2)

        self.export_menu_btn = ttk.Menubutton(top_controls_frame_row1, text="Export Log"); self.export_menu = tk.Menu(self.export_menu_btn, tearoff=0)
        self.export_menu.add_command(label="To Excel (.xlsx)", command=lambda: self.export_log_dialog_with_filter(export_type="xlsx"))
        self.export_menu.add_command(label="To Excel Macro-Enabled (.xlsm)", command=lambda: self.export_log_dialog_with_filter(export_type="xlsm"))
        self.export_menu.add_command(label="To CSV Files (.zip)", command=lambda: self.export_log_dialog_with_filter(export_type="csv"))
        self.export_menu.add_separator()
        self.export_menu.add_command(label="Export Layout as Image (see Help)...", command=self.export_layout_as_image) #self.export_layout_as_image)
        self.export_menu.add_command(label="Generate Attendance Report...", command=self.generate_attendance_report_dialog)
        self.export_menu_btn["menu"] = self.export_menu
        self.export_menu_btn.pack(side=tk.LEFT, padx=2)
        settings_btn = ttk.Button(top_controls_frame_row1, text="Settings", underline=0, command=self.open_settings_dialog)
        settings_btn.pack(side=tk.LEFT, padx=2)
        
        self.mode_frame = ttk.LabelFrame(top_controls_frame_row1, text="Mode", padding=2)
        self.mode_frame.pack(side=tk.LEFT, padx=2)
        ttk.Radiobutton(self.mode_frame, text="Behavior", underline=0, variable=self.mode_var, value="behavior", command=self.toggle_mode).pack(side=tk.LEFT)
        ttk.Radiobutton(self.mode_frame, text="Quiz", underline=0, variable=self.mode_var, value="quiz", command=self.toggle_mode).pack(side=tk.LEFT)
        ttk.Radiobutton(self.mode_frame, text="Homework", underline=0, variable=self.mode_var, value="homework", command=self.toggle_mode).pack(side=tk.LEFT) # New Homework mode

        self.live_quiz_button_frame = ttk.LabelFrame(top_controls_frame_row1, text="Class Quiz")
        self.live_quiz_button_frame.pack(side=tk.LEFT, padx=2)
        self.start_live_quiz_btn = ttk.Button(self.live_quiz_button_frame, text="Start Session", command=self.start_live_quiz_session_dialog); self.start_live_quiz_btn.pack(side=tk.LEFT, padx=3, pady=3)
        self.end_live_quiz_btn = ttk.Button(self.live_quiz_button_frame, text="End Session", command=self.end_live_quiz_session, state=tk.DISABLED); self.end_live_quiz_btn.pack(side=tk.LEFT, padx=3, pady=3)

        self.live_homework_button_frame = ttk.LabelFrame(top_controls_frame_row1, text="Homework Session") # New
        self.live_homework_button_frame.pack(side=tk.LEFT, padx=2)
        self.start_live_homework_btn = ttk.Button(self.live_homework_button_frame, text="Start Session", command=self.start_live_homework_session_dialog); self.start_live_homework_btn.pack(side=tk.LEFT, padx=3, pady=3)
        self.end_live_homework_btn = ttk.Button(self.live_homework_button_frame, text="End Session", command=self.end_live_homework_session, state=tk.DISABLED); self.end_live_homework_btn.pack(side=tk.LEFT, padx=3, pady=3)

        

        
        self.zoom_var = tk.StringVar(value=str(float(self.current_zoom_level)*100))
        view_controls_frame = ttk.LabelFrame(top_controls_frame_row1, text="View & Edit", padding=2)
        view_controls_frame.pack(side=tk.LEFT, padx=2)
        zoom_in_btn = ttk.Button(view_controls_frame, text="In", command=lambda: self.zoom_canvas(1.1)); zoom_in_btn.pack(side=tk.LEFT, padx=2)
        self.zoom_display_label = ttk.Entry(view_controls_frame, textvariable=self.zoom_var, width=5)
        if self.settings.get("show_zoom_level_display", True): self.zoom_display_label.pack(side=tk.LEFT, padx=1)
        zoom_out_btn = ttk.Button(view_controls_frame, text="Out", command=lambda: self.zoom_canvas(0.9)); zoom_out_btn.pack(side=tk.LEFT, padx=2)
        zoom_reset_btn = ttk.Button(view_controls_frame, text="Reset", command=lambda: self.zoom_canvas(0)); zoom_reset_btn.pack(side=tk.LEFT, padx=2)
        self.edit_mode_checkbutton = ttk.Checkbutton(view_controls_frame, text="Edit Mode", underline=0, variable=self.edit_mode_var, command=self.toggle_edit_mode); self.edit_mode_checkbutton.pack(side=tk.LEFT, padx=5)
        self.toggle_incidents_btn = ttk.Button(view_controls_frame, text="Hide Recent Logs", command=self.toggle_global_recent_logs_visibility); self.toggle_incidents_btn.pack(side=tk.LEFT, padx=2) # Renamed
        self.update_toggle_incidents_button_text()

        self.toggle_rulers_btn = ttk.Button(view_controls_frame, text="Toggle Rulers", command=self.toggle_rulers_visibility)
        self.toggle_rulers_btn.pack(side=tk.LEFT, padx=2)

        self.toggle_grid_btn = ttk.Button(view_controls_frame, text="Toggle Grid", command=self.toggle_grid_visibility)
        self.toggle_grid_btn.pack(side=tk.LEFT, padx=2)

        # Add Guide Buttons
        self.add_v_guide_btn = ttk.Button(view_controls_frame, text="Add V Guide", command=lambda: self.toggle_add_guide_mode("vertical", self.add_v_guide_btn))
        self.add_v_guide_btn.pack(side=tk.LEFT, padx=2)
        self.add_h_guide_btn = ttk.Button(view_controls_frame, text="Add H Guide", command=lambda: self.toggle_add_guide_mode("horizontal", self.add_h_guide_btn))
        self.add_h_guide_btn.pack(side=tk.LEFT, padx=2)

        self.lock_app_btn = ttk.Button(top_controls_frame_row1, text="Lock", command=self.lock_application_ui_triggered)
        self.lock_app_btn.pack(side=tk.LEFT, padx=2)
        self.update_lock_button_state()
        self.root.bind_all("<Control-l>", lambda e: self.lock_application_ui_triggered())
        ttk.Button(top_controls_frame_row1, text="Help", underline=3, command=self.show_help_dialog).pack(side=tk.RIGHT, padx=2)
        
        self.top_controls_frame_row2 = ttk.Frame(self.top_frame.interior, height=1); self.top_controls_frame_row2.pack(side=tk.BOTTOM, expand=False, fill=tk.X, pady=(2, 5), anchor="sw")

        # --- Row 2 Widgets (parented to self.top_controls_frame_row2) ---
        self.manage_boxes_frame = ttk.Frame(self.top_controls_frame_row2)
        self.manage_boxes_frame.pack(side=tk.LEFT, padx=2)
        
        layout_tools_frame = ttk.LabelFrame(self.manage_boxes_frame, text="Layout Tools", padding=2)
        layout_tools_frame.pack(side=tk.LEFT, padx=2)
        ttk.Button(layout_tools_frame, text="Align Top", command=lambda: self.align_selected_items("top")).pack(side=tk.LEFT,pady=1, padx=1)
        ttk.Button(layout_tools_frame, text="Align Bottom", command=lambda: self.align_selected_items("bottom")).pack(side=tk.LEFT,pady=1, padx=1)
        ttk.Button(layout_tools_frame, text="Align Left", command=lambda: self.align_selected_items("left")).pack(side=tk.LEFT,pady=1, padx=1)
        ttk.Button(layout_tools_frame, text="Align Right", command=lambda: self.align_selected_items("right")).pack(side=tk.LEFT,pady=1, padx=1)
        ttk.Button(layout_tools_frame, text="Distribute H", command=lambda: self.distribute_selected_items_evenly("horizontal")).pack(side=tk.LEFT, pady=1, padx=1)
        ttk.Button(layout_tools_frame, text="Distribute V", command=lambda: self.distribute_selected_items_evenly("vertical")).pack(side=tk.LEFT, pady=1, padx=1)

        templates_groups_frame = ttk.LabelFrame(self.manage_boxes_frame, text="Layout & Groups", padding=2)
        templates_groups_frame.pack(side=tk.LEFT, padx=2)
        ttk.Button(templates_groups_frame, text="Save Layout...", command=self.save_layout_template_dialog).pack(side=tk.LEFT,pady=1, padx=1)
        ttk.Button(templates_groups_frame, text="Load Layout...", command=self.load_layout_template_dialog).pack(side=tk.LEFT,pady=1, padx=1)
        self.manage_groups_btn = ttk.Button(templates_groups_frame, text="Manage Groups...", command=self.manage_student_groups_dialog); self.manage_groups_btn.pack(side=tk.LEFT,pady=1, padx=1)        
        
        # Toggle Dragging Button
        self.toggle_dragging_btn = ttk.Button(self.top_controls_frame_row2, text="Disable Dragging", command=self.toggle_dragging_allowed)
        self.toggle_dragging_btn.pack(side=tk.LEFT, padx=2)
        self._update_toggle_dragging_button_text() # Initialize button text
        
        show_undo_history_btn = ttk.Button(self.top_controls_frame_row2, text="Show undo history", command=self.show_undo_history_dialog)
        show_undo_history_btn.pack(side=tk.LEFT, padx=2)
        
        add_student_btn = ttk.Button(self.top_controls_frame_row2, text="Add Student", command=self.add_student_dialog)
        add_student_btn.pack(side=tk.LEFT, padx=2)
        add_furniture_btn = ttk.Button(self.top_controls_frame_row2, text="Add Furniture", command=self.add_furniture_dialog)
        add_furniture_btn.pack(side=tk.LEFT, padx=2)
        
        self.zoom_display_label.bind("<FocusOut>", lambda e: self.update_zoom_display2())
        self.zoom_display_label.bind("<Return>", lambda e: self.update_zoom_display2())
        
        

        self.theme_auto(init=True)
            
        self.canvas_frame = ttk.Frame(self.main_frame); self.canvas_frame.pack(fill=tk.BOTH, after=self.top_frame, expand=True)
        self.h_scrollbar = ttk.Scrollbar(self.canvas_frame, orient=tk.HORIZONTAL, command=self.canvas_xview_custom)
        self.v_scrollbar = ttk.Scrollbar(self.canvas_frame, orient=tk.VERTICAL, command=self.canvas_yview_custom) #else "#1F1F1F"
        self.canvas = tk.Canvas(self.canvas_frame, bg=self.canvas_color, relief=tk.SUNKEN, borderwidth=1, xscrollcommand=self.h_scrollbar.set, yscrollcommand=self.v_scrollbar.set) # type: ignore
        self.canvas.bind("<Configure>", self.on_canvas_configure)
        self.h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X); self.v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y); self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.canvas.config(scrollregion=(0, 0, self.canvas_orig_width * self.current_zoom_level, self.canvas_orig_height * self.current_zoom_level))
        
        self.canvas.bind("<ButtonPress-1>", self.on_canvas_left_press); self.canvas.bind("<ButtonPress-3>", self.on_canvas_right_press)
        self.canvas.bind("<B1-Motion>", self.on_canvas_drag); self.canvas.bind("<ButtonRelease-1>", self.on_canvas_release)
        self.canvas.bind("<Control-Button-1>", self.on_canvas_ctrl_click); self.canvas.bind("<KeyPress-Delete>", self.on_delete_key_press); self.canvas.bind("<BackSpace>", self.on_delete_key_press)
        self.canvas.bind("<Control-MouseWheel>", self.on_mousewheel_zoom); self.canvas.bind("<Control-Button-4>", self.on_mousewheel_zoom); self.canvas.bind("<Control-Button-5>", self.on_mousewheel_zoom)
        self.canvas.bind("<ButtonPress-2>", self.on_pan_start); self.canvas.bind("<B2-Motion>", self.on_pan_move); self.canvas.bind("<ButtonRelease-2>", self.on_pan_end)
        self.canvas.bind("<MouseWheel>", self.on_mousewheel_scroll); self.canvas.bind("<Button-4>", self.on_mousewheel_scroll); self.canvas.bind("<Button-5>", self.on_mousewheel_scroll)
        if sys.platform == "darwin": self.canvas.bind("<Shift-MouseWheel>", self.on_mousewheel_scroll_horizontal_mac)
        else: self.canvas.bind("<Shift-MouseWheel>", self.on_mouse_wheel_horizontal) # For Windows/Linux with Shift

        self.status_bar_label = ttk.Label(self.root, text="Welcome!", relief=tk.SUNKEN, anchor=tk.W, padding=5); self.status_bar_label.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.focus_set()
        self.toggle_student_groups_ui_visibility()
        self.toggle_manage_boxes_visibility()

    def canvas_xview_custom(self, *args): self.canvas.xview(*args); self.password_manager.record_activity()
    def canvas_yview_custom(self, *args): self.canvas.yview(*args); self.password_manager.record_activity()
    def on_mousewheel_scroll(self, event):
        if self.password_manager.is_locked: return
        self.password_manager.record_activity()
        if event.num == 5 or event.delta < 0: self.canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta > 0: self.canvas.yview_scroll(-1, "units")
    def on_mouse_wheel_horizontal(self, event): # For Shift+Wheel on Windows/Linux
        if self.password_manager.is_locked: return
        self.password_manager.record_activity()
        if event.delta < 0: self.canvas.xview_scroll(1, "units") # Scroll right
        elif event.delta > 0: self.canvas.xview_scroll(-1, "units") # Scroll left
    def on_mousewheel_scroll_horizontal_mac(self, event):
        if self.password_manager.is_locked: return
        self.password_manager.record_activity()
        if event.delta < 0: self.canvas.xview_scroll(1, "units")
        elif event.delta > 0: self.canvas.xview_scroll(-1, "units")

    def lock_application_ui_triggered(self):
        if self.password_manager.is_password_set():
            if self.password_manager.lock_application():
                self.update_status("Application locked."); self.show_lock_screen(); self.update_lock_button_state()
            else: self.update_status("Failed to lock: No password set or already locked.")
        else:
            messagebox.showinfo("Password Not Set", "Please set an application password in Settings first.", parent=self.root)
            self.open_settings_dialog()
    def update_lock_button_state(self):
        if hasattr(self, 'lock_app_btn'): self.lock_app_btn.config(state=tk.NORMAL if self.password_manager.is_password_set() else tk.DISABLED)
    def save_and_quit_app(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Save & Quit", "Enter password to save and quit:"): return
        if self.is_live_quiz_active and not self.prompt_end_live_session_on_mode_switch("quiz"): return
        if self.is_live_homework_active and not self.prompt_end_live_session_on_mode_switch("homework"): return # New
        self.save_data_wrapper(source="save_and_quit")
        self.on_exit_protocol(force_quit=True) # Call main exit to release lock

    def on_delete_key_press(self, event=None):
        if self.password_manager.is_locked: return
        if self.selected_items: self.delete_selected_items_confirm()
        self.password_manager.record_activity()
    def update_open_last_export_folder_menu_item(self):
        if hasattr(self, 'file_menu') and self.file_menu:
            label_text, state = "Open Last Export Folder (None)", tk.DISABLED
            if self.last_excel_export_path and os.path.exists(os.path.dirname(self.last_excel_export_path)):
                label_text, state = f"Open Last Export Folder ({os.path.basename(os.path.dirname(self.last_excel_export_path))})", tk.NORMAL
            try:
                if self.open_export_folder_menu_entry_index is not None and self.open_export_folder_menu_entry_index <= self.file_menu.index(tk.END):
                    self.file_menu.entryconfigure(self.open_export_folder_menu_entry_index, label=label_text, state=state)
            except tk.TclError as e: print(f"Error updating 'Open Last Export Folder' menu item: {e}.")

    def prompt_end_live_session_on_mode_switch(self, session_type_to_check): # "quiz" or "homework"
        is_active_flag = self.is_live_quiz_active if session_type_to_check == "quiz" else self.is_live_homework_active
        current_name = self.current_live_quiz_name if session_type_to_check == "quiz" else self.current_live_homework_name
        end_function = self.end_live_quiz_session if session_type_to_check == "quiz" else self.end_live_homework_session
        start_btn = self.start_live_quiz_btn if session_type_to_check == "quiz" else self.start_live_homework_btn
        end_btn = self.end_live_quiz_btn if session_type_to_check == "quiz" else self.end_live_homework_btn
        scores_dict = self.live_quiz_scores if session_type_to_check == "quiz" else self.live_homework_scores

        if is_active_flag:
            msg = f"A Class {session_type_to_check.capitalize()} session ('{current_name}') is active.\n\nDo you want to end and log this session?"
            response = messagebox.askyesnocancel(f"Class {session_type_to_check.capitalize()} Active", msg, parent=self.root)
            if response is True: end_function(confirm=False); return True
            elif response is False: # Discard
                if session_type_to_check == "quiz": self.is_live_quiz_active = False; self.current_live_quiz_name = ""
                else: self.is_live_homework_active = False; self.current_live_homework_name = ""
                scores_dict.clear()
                start_btn.config(state=tk.NORMAL); end_btn.config(state=tk.DISABLED)
                self.update_status(f"Class {session_type_to_check.capitalize()} session discarded.")
                self.draw_all_items(check_collisions_on_redraw=True); return True
            else: return False # Cancel
        return True # No active session of this type

    def toggle_mode_(self, mode):
        self.mode_var.set(mode)
        self.toggle_mode()

    def toggle_mode(self, initial=False):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Change Mode", "Enter password to change mode:"):
                self.mode_var.set(self.settings["current_mode"]); return

        new_mode = self.mode_var.get()
        old_mode = self.settings["current_mode"]

        if old_mode == "quiz" and new_mode != "quiz" and self.is_live_quiz_active:
            if not self.prompt_end_live_session_on_mode_switch("quiz"): self.mode_var.set("quiz"); return
        if old_mode == "homework" and new_mode != "homework" and self.is_live_homework_active: # New
            if not self.prompt_end_live_session_on_mode_switch("homework"): self.mode_var.set("homework"); return


        self.settings["current_mode"] = new_mode
        if not initial:
            self.update_status(f"Mode switched to {new_mode.capitalize()}.")

        if hasattr(self, 'live_quiz_button_frame'):
            if new_mode == "quiz":
                self.live_quiz_button_frame.pack(side=tk.LEFT, padx=(0,3), after=self.mode_frame)
                self.start_live_quiz_btn.config(state=tk.DISABLED if self.is_live_quiz_active else tk.NORMAL)
                self.end_live_quiz_btn.config(state=tk.NORMAL if self.is_live_quiz_active else tk.DISABLED)
            else: self.live_quiz_button_frame.pack_forget()

        if hasattr(self, 'live_homework_button_frame'): # New
            if new_mode == "homework":
                self.live_homework_button_frame.pack(side=tk.LEFT, padx=(0,3), after=self.mode_frame)
                self.start_live_homework_btn.config(state=tk.DISABLED if self.is_live_homework_active else tk.NORMAL)
                self.end_live_homework_btn.config(state=tk.NORMAL if self.is_live_homework_active else tk.DISABLED)
            else: self.live_homework_button_frame.pack_forget()

        self.draw_all_items(check_collisions_on_redraw=True)
        self.save_data_wrapper(source="toggle_mode")
        self.password_manager.record_activity()

    def toggle_edit_mode_shortcut(self):
        self.edit_mode_var.set(value=True if self.edit_mode_var.get() != True else False)
        self.toggle_edit_mode()
        
    def toggle_edit_mode(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Toggle Edit Mode", "Enter password to change edit mode:"):
                self.edit_mode_var.set(not self.edit_mode_var.get()); return
        is_edit_mode = self.edit_mode_var.get()
        self.update_status(f"Edit Mode {'Enabled. Click item corners to resize' if is_edit_mode else 'Disabled'}.")
        self.toggle_manage_boxes_visibility()
        self.draw_all_items(check_collisions_on_redraw=True)
        self.password_manager.record_activity()

    def start_live_quiz_session_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Start Quiz", "Enter password to start quiz session:"): return
        if self.is_live_quiz_active:
            messagebox.showinfo("Class Quiz Active", "A Class Quiz session is already active.", parent=self.root); return
        quiz_name = simpledialog.askstring("Start Class Quiz", "Enter a name for this quiz session:", initialvalue=self.settings.get("default_quiz_name", "Class Quiz"), parent=self.root)
        if quiz_name and quiz_name.strip():
            self.current_live_quiz_name = quiz_name.strip()
            self.is_live_quiz_active = True; self.live_quiz_scores.clear()
            self.start_live_quiz_btn.config(state=tk.DISABLED); self.end_live_quiz_btn.config(state=tk.NORMAL)
            self.update_status(f"Class Quiz '{self.current_live_quiz_name}' started. Click a student to mark.")
            self.draw_all_items(check_collisions_on_redraw=True); self.password_manager.record_activity()
        else: self.update_status("Class Quiz start cancelled.")

    def end_live_quiz_session(self, confirm=True):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to End Quiz", "Enter password to end quiz session:"): return
        if not self.is_live_quiz_active: self.update_status("No active Class Quiz session to end."); return
        if confirm and not messagebox.askyesno("End Class Quiz", f"End Class Quiz session: '{self.current_live_quiz_name}'?\nScores will be logged.", parent=self.root): return

        log_commands = []
        for student_id, score_data in self.live_quiz_scores.items():
            if student_id in self.students and score_data["total_asked"] > 0:
                student = self.students[student_id]
                log_entry = {"timestamp": datetime.now().isoformat(), "student_id": student_id,
                             "student_first_name": student["first_name"], "student_last_name": student["last_name"],
                             "behavior": self.current_live_quiz_name, "score_details": score_data.copy(),
                             "comment": "From Class Quiz session.", "type": "quiz", "day": datetime.now().strftime('%A')}
                log_commands.append(LogEntryCommand(self, log_entry, student_id))
        for cmd in log_commands: self.execute_command(cmd)
        if log_commands: self.save_data_wrapper(source="end_live_quiz")
        self.update_status(f"Class Quiz '{self.current_live_quiz_name}' ended. {len(log_commands)} student scores logged.")
        self.is_live_quiz_active = False; self.current_live_quiz_name = ""; self.live_quiz_scores.clear()
        self.start_live_quiz_btn.config(state=tk.NORMAL); self.end_live_quiz_btn.config(state=tk.DISABLED)
        self.draw_all_items(check_collisions_on_redraw=True); self.password_manager.record_activity()

    def handle_live_quiz_tap(self, student_id):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Mark Quiz", "Enter password to mark quiz:"): return
        if not self.is_live_quiz_active or student_id not in self.students: return
        dialog = LiveQuizMarkDialog(self.root, student_id, self, session_type="Quiz")
        if dialog.result:
            self.execute_command(MarkLiveQuizQuestionCommand(self, student_id, dialog.result))
            self.password_manager.record_activity()

    # --- Live Homework Session Methods (New) ---
    def start_live_homework_session_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Start Homework Session", "Enter password to start homework session:"): return
        if self.is_live_homework_active:
            messagebox.showinfo("Homework Session Active", "A Homework session is already active.", parent=self.root); return

        homework_session_name = simpledialog.askstring("Start Homework Session", "Enter a name for this homework session:",
                                                       initialvalue=self.settings.get("default_homework_name", "Homework Check"), parent=self.root)
        if homework_session_name and homework_session_name.strip():
            self.current_live_homework_name = homework_session_name.strip()
            self.is_live_homework_active = True
            self.live_homework_scores.clear() # {student_id: {hw_type_id: "yes"/"no"} or {student_id: {"selected_options": [...]}}}
            self.start_live_homework_btn.config(state=tk.DISABLED)
            self.end_live_homework_btn.config(state=tk.NORMAL)
            self.update_status(f"Homework Session '{self.current_live_homework_name}' started. Click a student to mark.")
            self.draw_all_items(check_collisions_on_redraw=True)
            self.password_manager.record_activity()
        else:
            self.update_status("Homework Session start cancelled.")

    def end_live_homework_session(self, confirm=True):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to End Homework Session", "Enter password to end homework session:"): return
        if not self.is_live_homework_active:
            self.update_status("No active Homework Session to end."); return
        if confirm and not messagebox.askyesno("End Homework Session", f"End Homework Session: '{self.current_live_homework_name}'?\nData will be logged.", parent=self.root):
            return

        log_commands = []
        for student_id, hw_data in self.live_homework_scores.items():
            if student_id in self.students and hw_data: # Ensure there's some data to log
                student = self.students[student_id]
                # Log entry structure depends on session mode
                if self.settings.get('live_homework_session_mode') == "Yes/No":
                    type2 = "y"
                else: type2 = "s"
                log_entry = {
                    "timestamp": datetime.now().isoformat(), "student_id": student_id,
                    "student_first_name": student["first_name"], "student_last_name": student["last_name"],
                    "behavior": self.current_live_homework_name, # Session name
                    "homework_details": hw_data.copy(), # Store the collected data
                    "comment": f"From Live Homework Session ({self.settings.get('live_homework_session_mode', 'Yes/No')} mode).",
                    "type": f"homework_session_{type2}", # Distinguish from manual homework log
                    "day": datetime.now().strftime('%A')
                }
                log_commands.append(LogHomeworkEntryCommand(self, log_entry, student_id)) # Use homework log command

        for cmd in log_commands: self.execute_command(cmd)
        if log_commands: self.save_data_wrapper(source="end_live_homework_session")

        self.update_status(f"Homework Session '{self.current_live_homework_name}' ended. {len(log_commands)} student entries logged.")
        self.is_live_homework_active = False; self.current_live_homework_name = ""; self.live_homework_scores.clear()
        self.start_live_homework_btn.config(state=tk.NORMAL); self.end_live_homework_btn.config(state=tk.DISABLED)
        self.draw_all_items(check_collisions_on_redraw=True); self.password_manager.record_activity()

    def handle_live_homework_tap(self, student_id):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Mark Homework", "Enter password to mark homework:"): return
        if not self.is_live_homework_active or student_id not in self.students: return

        session_mode = self.settings.get("live_homework_session_mode", "Yes/No")
        current_student_hw_data = self.live_homework_scores.get(student_id, {}).copy()

        dialog = LiveHomeworkMarkDialog(self.root, student_id, self,
                                        session_mode=session_mode,
                                        current_hw_data=current_student_hw_data)
        if dialog.result_actions is not None: # Dialog returns actions or None if cancelled
            cmd = MarkLiveHomeworkCommand(self, student_id, dialog.result_actions, session_mode)
            self.execute_command(cmd)
            self.password_manager.record_activity()

    def add_student_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Add Student", "Enter password to add student:"): return
        dialog = AddEditStudentDialog(self.root, "Add Student", app=self)
        if dialog.result:
            first_name, last_name, nickname, gender, group_id_selection = dialog.result
            if first_name and last_name:
                old_next_student_id_num_for_command = self.next_student_id_num
                student_id_str, next_id_val_for_app_state_after_this = self.get_new_student_id()
                full_name = f"{first_name} \"{nickname}\" {last_name}" if nickname else f"{first_name} {last_name}"
                x, y = self.canvas_to_world_coords(50 + (len(self.students) % 10) * 20, 50 + ((len(self.students) // 10) * 20))
                student_data = {"first_name": first_name, "last_name": last_name, "nickname": nickname, "full_name": full_name, "gender": gender,
                                "x": x, "y": y, "id": student_id_str, "width": self.settings.get("default_student_box_width"),
                                "height": self.settings.get("default_student_box_height"), "original_next_id_num_after_add": next_id_val_for_app_state_after_this,
                                "group_id": group_id_selection if group_id_selection else None, "style_overrides": {}}
                self.execute_command(AddItemCommand(self, student_id_str, 'student', student_data, old_next_student_id_num_for_command))
                self.password_manager.record_activity()
            else: messagebox.showwarning("Invalid Name", "First and Last names cannot be empty.", parent=self.root)

    def add_furniture_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Add Furniture", "Enter password to add furniture:"): return
        dialog = AddFurnitureDialog(self.root, "Add Furniture Item")
        if dialog.result:
            name, item_type, width, height = dialog.result
            if name and item_type:
                old_next_furniture_id_num_for_command = self.next_furniture_id_num
                furniture_id_str, next_id_val_for_app_state_after_this = self.get_new_furniture_id()
                x, y = self.canvas_to_world_coords(70 + (len(self.furniture) % 10) * 20, 70 + ((len(self.furniture) // 10) * 20))
                furniture_data = {"name": name, "type": item_type, "x": x, "y": y, "id": furniture_id_str, "width": width, "height": height,
                                  "fill_color": "lightgray", "outline_color": "dimgray", "original_next_id_num_after_add": next_id_val_for_app_state_after_this}
                self.execute_command(AddItemCommand(self, furniture_id_str, 'furniture', furniture_data, old_next_furniture_id_num_for_command))
                self.password_manager.record_activity()

    def toggle_global_recent_logs_visibility(self): # Renamed for clarity
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Change Setting", "Enter password to change log visibility:"): return
        # This now toggles both behavior and homework visibility together for simplicity,
        # or could be split into two separate global toggles if needed.
        self._recent_incidents_hidden_globally = not self._recent_incidents_hidden_globally
        self._recent_homeworks_hidden_globally = self._recent_incidents_hidden_globally # Link them for now
        self.draw_all_items(check_collisions_on_redraw=True)
        self.update_toggle_incidents_button_text() # Button text reflects combined state
        status_msg = "Recent behavior/homework logs hidden globally." if self._recent_incidents_hidden_globally else "Recent behavior/homework logs shown globally."
        self.update_status(status_msg)
        self.password_manager.record_activity()

    def update_toggle_incidents_button_text(self): # Renamed
        if hasattr(self, "toggle_incidents_btn"):
            # Text reflects combined state of behavior and homework logs
            text = "Show Recent Logs" if self._recent_incidents_hidden_globally or self._recent_homeworks_hidden_globally else "Hide Recent Logs"
            self.toggle_incidents_btn.config(text=text)

    def clear_recent_logs_for_student(self, student_id): # Renamed
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Clear Logs", "Enter password to clear logs for student:"): return
        self._per_student_last_cleared[student_id] = datetime.now().isoformat()
        self.draw_single_student(student_id, check_collisions=True)
        student = self.students.get(student_id)
        if student: self.update_status(f"Recent behavior/homework logs cleared for {student['full_name']}.")
        self.save_data_wrapper(); self.password_manager.record_activity()

    def show_recent_logs_for_student(self, student_id): # Renamed
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Show Logs", "Enter password to show logs for student:"): return
        if student_id in self._per_student_last_cleared:
            del self._per_student_last_cleared[student_id]
            self.draw_single_student(student_id, check_collisions=True)
            student = self.students.get(student_id)
            if student: self.update_status(f"Recent behavior/homework logs will now show for {student['full_name']}.")
            self.save_data_wrapper(); self.password_manager.record_activity()

    def _get_recent_logs_for_student(self, student_id, log_type_key): # "behavior" or "homework"
        student = self.students.get(student_id)
        if not student: return []

        summary_lines_list = []
        log_source = self.behavior_log if log_type_key == "behavior" else self.homework_log
        setting_prefix = "recent_incidents" if log_type_key == "behavior" else "recent_homeworks" # For settings keys
        global_hidden_flag = self._recent_incidents_hidden_globally if log_type_key == "behavior" else self._recent_homeworks_hidden_globally
        behavior_key_in_log = "behavior" if log_type_key == "behavior" else "homework_type" # Or "behavior" for manual homework log

        if not global_hidden_flag and self.settings.get(f"show_{setting_prefix}_on_boxes", True):
            num_to_show = self.settings.get(f"num_{setting_prefix}_to_show", 0)
            if num_to_show > 0:
                time_window_hours = self.settings.get(f"{setting_prefix}_time_window_hours", 24)
                cutoff_time = datetime.now() - timedelta(hours=time_window_hours)
                last_cleared_iso = self._per_student_last_cleared.get(student_id)
                cleared_dt = datetime.fromisoformat(last_cleared_iso) if last_cleared_iso else None

                # Adjust filter for log type. Behavior logs have "type":"behavior", quiz logs have "type":"quiz".
                # Homework logs have "type":"homework" or "type":"homework_session".
                type_filter_values = []
                if log_type_key == "behavior": type_filter_values = ["behavior", "quiz"] # Behavior tab shows both
                elif log_type_key == "homework": type_filter_values = ["homework", "homework_session_y", "homework_session_s"]


                all_recent_logs = sorted(
                    [log for log in log_source
                     if log["student_id"] == student_id and log.get("type") in type_filter_values and
                        datetime.fromisoformat(log["timestamp"]) >= cutoff_time and
                        (not cleared_dt or datetime.fromisoformat(log["timestamp"]) > cleared_dt)],
                    key=lambda x: x["timestamp"], reverse=True
                )

                specific_filter_list = self.settings.get(f"selected_{setting_prefix}_filter", None)
                filtered_logs = []
                if specific_filter_list is None: filtered_logs = all_recent_logs
                elif isinstance(specific_filter_list, list) and not specific_filter_list: filtered_logs = [] # Empty list means filter all
                elif isinstance(specific_filter_list, list):
                    filtered_logs = [log for log in all_recent_logs if log.get(behavior_key_in_log, log.get("behavior")) in specific_filter_list]


                recent_to_display = filtered_logs[:num_to_show]
                if self.settings.get(f"reverse_{setting_prefix.replace('s', '')}_order", True): recent_to_display.reverse()

                initial_overrides_key = "behavior_initial_overrides" if log_type_key == "behavior" else "homework_initial_overrides"
                initial_overrides = self.settings.get(initial_overrides_key, {})
                
                #string_for_get = 
                
                if self.settings.get(f"show_full_{setting_prefix}", False):
                    summary_lines_list = [log.get(behavior_key_in_log, log.get("behavior")) for log in recent_to_display if log.get(behavior_key_in_log, log.get("behavior"))]
                else:
                    temp_initials = []
                    for entry in recent_to_display:
                        name = entry.get(behavior_key_in_log, entry.get("behavior"))
                        if name in initial_overrides and initial_overrides[name]: temp_initials.append(initial_overrides[name])
                        elif name: temp_initials.append(''.join(part[0].upper() for part in name.split() if part))
                        else: temp_initials.append("?")
                    if temp_initials: summary_lines_list.append('  '.join(temp_initials))
        return summary_lines_list

    def _get__logs_for_student(self, student_id, log_type_key, num_max, window, name_of_spec): # "behavior" or "homework"
        student = self.students.get(student_id)
        if not student: return []
        #print(num_max)
        summary_lines_list = []
        log_source = self.behavior_log if log_type_key == "behavior" else self.homework_log
        setting_prefix = "recent_incidents" if log_type_key == "behavior" else "recent_homeworks" # For settings keys
        global_hidden_flag = self._recent_incidents_hidden_globally if log_type_key == "behavior" else self._recent_homeworks_hidden_globally
        behavior_key_in_log = "behavior" if log_type_key == "behavior" else "homework_type" # Or "behavior" for manual homework log

        if not global_hidden_flag and self.settings.get(f"show_{setting_prefix}_on_boxes", True):
            num_to_show = num_max
            if num_to_show > 0:
                time_window_hours = window
                cutoff_time = datetime.now() - timedelta(hours=time_window_hours)

                # Adjust filter for log type. Behavior logs have "type":"behavior", quiz logs have "type":"quiz".
                # Homework logs have "type":"homework" or "type":"homework_session".
                type_filter_values = []
                if log_type_key == "behavior": type_filter_values = ["behavior", "quiz"] # Behavior tab shows both
                elif log_type_key == "homework": type_filter_values = ["homework", "homework_session_y", "homework_session_s"]


                all_recent_logs = sorted(
                    [log for log in log_source
                     if log["student_id"] == student_id and log.get("type") in type_filter_values and
                        datetime.fromisoformat(log["timestamp"]) >= cutoff_time],
                    key=lambda x: x["timestamp"], reverse=True
                )

                specific_filter_list = [name_of_spec]
                filtered_logs = []
                if specific_filter_list is None: filtered_logs = all_recent_logs
                elif isinstance(specific_filter_list, list) and not specific_filter_list: filtered_logs = [] # Empty list means filter all
                elif isinstance(specific_filter_list, list):
                    filtered_logs = [log for log in all_recent_logs if log.get(behavior_key_in_log, log.get("behavior")) in specific_filter_list]


                recent_to_display = filtered_logs[:num_to_show]
                
                
                summary_lines_list = [log.get(behavior_key_in_log, log.get("behavior")) for log in recent_to_display if log.get(behavior_key_in_log, log.get("behavior"))]
                
        return len(summary_lines_list)

    def update_student_display_text(self, student_id):
        student = self.students.get(student_id)
        if not student: return

        name_to_display = student.get('nickname') or student['first_name']
        main_content_lines = [name_to_display, student['last_name']]
        incident_display_lines = [] # List of dicts: {"text": "...", "type": "incident/quiz_score/homework_score"}

        current_mode = self.mode_var.get()
        if current_mode == "quiz" and self.is_live_quiz_active:
            score_text = f"{self.current_live_quiz_name}: (Pending)"
            if student_id in self.live_quiz_scores:
                score_info = self.live_quiz_scores[student_id]
                score_text = f"{self.current_live_quiz_name}: {score_info['correct']} of {score_info['total_asked']}"
            incident_display_lines.append({"text": score_text, "type": "quiz_score"})
            if self.settings.get("show_recent_incidents_during_quiz", True):
                for line_text in self._get_recent_logs_for_student(student_id, "behavior"):
                    if line_text: incident_display_lines.append({"text": line_text, "type": "incident"})

        elif current_mode == "homework" and self.is_live_homework_active: # New for live homework
            hw_session_mode = self.settings.get("live_homework_session_mode", "Yes/No")
            hw_score_data = self.live_homework_scores.get(student_id, {})
            hw_display_texts = []
            if hw_session_mode == "Yes/No":
                for hw_type_id, status in hw_score_data.items():
                    # Find homework type name from custom_homework_session_types or defaults
                    hw_type_obj = next((ht for ht in self.all_homework_session_types if isinstance(ht, dict) and ht.get('id') == hw_type_id), None)
                    hw_name_display = hw_type_obj['name'] if hw_type_obj else hw_type_id
                    hw_display_texts.append(f"{hw_name_display}: {status.capitalize()}")
            elif hw_session_mode == "Select" and "selected_options" in hw_score_data:
                hw_display_texts = list(hw_score_data["selected_options"])

            if hw_display_texts:
                incident_display_lines.append({"text": f"{self.current_live_homework_name}:", "type": "homework_score_header"})
                for text in hw_display_texts:
                    incident_display_lines.append({"text": f"  {text}", "type": "homework_score_item"})
            else:
                incident_display_lines.append({"text": f"{self.current_live_homework_name}: (Pending)", "type": "homework_score_header"})


        else: # Behavior mode or non-live quiz/homework mode
            # Show recent behavior/quiz logs
            for line_text in self._get_recent_logs_for_student(student_id, "behavior"):
                if line_text: incident_display_lines.append({"text": line_text, "type": "incident"})
            # Show recent homework logs
            homework_lines = self._get_recent_logs_for_student(student_id, "homework")
            if homework_lines and incident_display_lines: # Add separator if both types of logs exist
                incident_display_lines.append({"text": "--- Homework ---", "type": "separator"})
            for line_text in homework_lines:
                if line_text: incident_display_lines.append({"text": line_text, "type": "homework_log"})


        student["display_lines"] = main_content_lines
        student["incident_display_lines"] = incident_display_lines

    def applies_to_conditional(self, student_id, rule):
        student_data = self.students.get(student_id)
        if not student_data: return False # Student data is essential

        # Preliminary checks based on new rule fields
        if not rule.get("enabled", True): # Default to enabled if key is missing (should be set by load_data)
            return False

        # Check active_modes
        active_modes = rule.get("active_modes", [])
        if active_modes: # If list is not empty, mode must match
            current_app_mode = self.mode_var.get() # "behavior", "quiz", "homework"
            effective_mode = current_app_mode # Base mode

            # Determine more specific "session" modes
            if current_app_mode == "quiz" and self.is_live_quiz_active:
                effective_mode = "quiz_session"
            elif current_app_mode == "homework" and self.is_live_homework_active:
                effective_mode = "homework_session"

            if effective_mode not in active_modes:
                return False

        # Check active_times
        active_times = rule.get("active_times", [])
        if active_times: # If list is not empty, time must match
            now = datetime.now()
            current_time_str = now.strftime("%H:%M")
            current_day_of_week = now.weekday() # Monday is 0 and Sunday is 6

            time_match_found = False
            for time_slot in active_times:
                slot_days = time_slot.get("days_of_week", list(range(7))) # Default to all days if not specified
                if current_day_of_week in slot_days:
                    start_time_str = time_slot.get("start_time")
                    end_time_str = time_slot.get("end_time")
                    if start_time_str and end_time_str:
                        if start_time_str <= current_time_str < end_time_str:
                            time_match_found = True
                            break
            if not time_match_found:
                return False

        # If all preliminary checks passed, proceed to rule-specific logic
        rule_type = rule.get("type")

        # --- Live Session Conditional Formatting Rules ---
        if rule_type == "live_quiz_response":
            if not self.is_live_quiz_active or student_id not in self.live_quiz_scores:
                return False
            student_live_score = self.live_quiz_scores[student_id]
            last_response_mark_id = student_live_score.get("last_response_details") # Assuming this is set by MarkLiveQuizQuestionCommand
            if not last_response_mark_id: return False

            effective_response_type = ""
            for mt in self.settings.get("quiz_mark_types", []):
                if mt["id"] == last_response_mark_id:
                    if "correct" in mt["name"].lower() or "bonus" in mt["name"].lower() or \
                       (mt.get("default_points", 0) > 0 and not mt.get("is_extra_credit", False)) or \
                       (mt.get("default_points", 0) > 0 and mt.get("is_extra_credit", True)):
                        effective_response_type = "Correct"
                    elif "incorrect" in mt["name"].lower() or mt.get("default_points", 0) == 0:
                         effective_response_type = "Incorrect"
                    break

            rule_quiz_response = rule.get("quiz_response")
            return bool(effective_response_type and rule_quiz_response and effective_response_type == rule_quiz_response)

        elif rule_type == "live_homework_yes_no":
            if not self.is_live_homework_active or self.settings.get("live_homework_session_mode") != "Yes/No" or \
               student_id not in self.live_homework_scores:
                return False
            student_hw_data = self.live_homework_scores.get(student_id, {})
            rule_hw_type_id = rule.get("homework_type_id")
            rule_hw_response = rule.get("homework_response")
            return bool(rule_hw_type_id in student_hw_data and student_hw_data[rule_hw_type_id] == rule_hw_response)

        elif rule_type == "live_homework_select":
            if not self.is_live_homework_active or self.settings.get("live_homework_session_mode") != "Select" or \
               student_id not in self.live_homework_scores:
                return False
            student_hw_data = self.live_homework_scores.get(student_id, {})
            rule_option_name = rule.get("homework_option_name")
            return bool("selected_options" in student_hw_data and rule_option_name in student_hw_data["selected_options"])

        # --- Standard Conditional Formatting Rules (Non-Live Session) ---
        if rule_type == "behavior_count":
            time_window_hours = rule.get("time_window_hours", 24) # Default if not set
            count_threshold = rule.get("count_threshold", 1)
            behavior_name = rule.get("behavior_name", "")
            if not behavior_name: # Behavior name is essential for this rule type
                return False

            # _get__logs_for_student returns the count of matching logs to be displayed
            # The third argument to _get__logs_for_student is num_max (for display),
            # but here it's used as the threshold for comparison.
            # This might be okay if _get__logs_for_student correctly counts beyond num_max internally
            # or if the intent is to check against the *displayed* count.
            # For now, assuming existing logic of _get__logs_for_student is what's intended.
            actual_count = self._get__logs_for_student(student_id, "behavior", count_threshold, time_window_hours, behavior_name)
            
            # The comparison was actual_count >= count_threshold.
            # If _get__logs_for_student returns a count that is capped by its num_max (which is count_threshold here),
            # then (result >= count_threshold) would only be true if result == count_threshold.
            # This needs careful check of _get__logs_for_student's behavior or adjustment here.
            # For now, sticking to the original comparison structure.
            if actual_count >= count_threshold:
                return True
            return False # Explicitly return False if condition not met

        elif rule_type == "quiz_score_threshold":
            operator = rule.get("operator", "<=")
            quiz_name_contains = rule.get("quiz_name_contains", "")
            score_threshold_percent = rule.get("score_threshold_percent", 50.0)
            
            # Placeholder for actual quiz score logic - to be implemented next
            # This will iterate through relevant quiz logs for the student
            # calculate score % and compare.
            for log_entry in self.behavior_log:
                if log_entry.get("student_id") == student_id and log_entry.get("type") == "quiz":
                    # Check quiz name
                    if quiz_name_contains and quiz_name_contains.lower() not in log_entry.get("behavior", "").lower():
                        continue

                    current_score_percentage = self._calculate_quiz_score_percentage(log_entry)
                    if current_score_percentage is None:
                        continue

                    # Compare with threshold
                    if operator == "<=" and current_score_percentage <= score_threshold_percent: return True
                    elif operator == ">=" and current_score_percentage >= score_threshold_percent: return True
                    elif operator == "==" and abs(current_score_percentage - score_threshold_percent) < 0.01 : return True # Using tolerance for float comparison
                    elif operator == "<" and current_score_percentage < score_threshold_percent: return True
                    elif operator == ">" and current_score_percentage > score_threshold_percent: return True
            return False # No matching quiz log found or condition not met

        elif rule_type == "quiz_mark_count":
            quiz_name_contains = rule.get("quiz_name_contains", "").lower()
            mark_type_id_to_check = rule.get("mark_type_id")
            operator = rule.get("mark_operator", ">=") # Default operator
            count_threshold = rule.get("mark_count_threshold", 1)

            if not mark_type_id_to_check: # Mark type ID is essential
                return False

            for log_entry in self.behavior_log:
                if log_entry.get("student_id") == student_id and log_entry.get("type") == "quiz":
                    # Check quiz name
                    if quiz_name_contains and quiz_name_contains not in log_entry.get("behavior", "").lower():
                        continue

                    marks_data = log_entry.get("marks_data", {})
                    actual_count = marks_data.get(mark_type_id_to_check, 0) # Default to 0 if mark_type not in log

                    if not isinstance(actual_count, (int, float)): # Ensure we are comparing numbers
                        actual_count = 0

                    # Compare with threshold
                    if operator == ">=" and actual_count >= count_threshold: return True
                    elif operator == "<=" and actual_count <= count_threshold: return True
                    elif operator == "==" and actual_count == count_threshold: return True
                    elif operator == ">" and actual_count > count_threshold: return True
                    elif operator == "<" and actual_count < count_threshold: return True
                    elif operator == "!=" and actual_count != count_threshold: return True
            return False # No matching quiz log found or condition not met

        elif rule_type == "group":
            # Group rules are handled directly in draw_single_student before this method is called.
            # If it reaches here, it means it's not a group rule being evaluated by this function.
            return False
        
        return False # Default for unknown or unhandled rule types
        #if student_data.get()
        
    def draw_single_student(self, student_id, check_collisions=False):
        # ... (largely same as v51, but needs to handle new "homework_score_header/item" and "separator" types for drawing)
        # This method is long, so I'll highlight the key change area for incident_display_lines
        try:
            student_data = self.students.get(student_id)
            if not student_data: return
            self.update_student_display_text(student_id)

            style_overrides = student_data.get("style_overrides", {})
            world_x, world_y = student_data["x"], student_data["y"]
            world_width = style_overrides.get("width", student_data.get("width", self.settings.get("default_student_box_width")))
            world_base_height = style_overrides.get("height", student_data.get("height", self.settings.get("default_student_box_height")))
            canvas_x, canvas_y = self.world_to_canvas_coords(world_x, world_y)
            canvas_width = world_width * self.current_zoom_level
            canvas_base_height = world_base_height * self.current_zoom_level

            fill_color = style_overrides.get("fill_color", self.settings.get("student_box_fill_color"))
            outline_color_orig = style_overrides.get("outline_color", self.settings.get("student_box_outline_color"))

            if self.is_live_quiz_active and student_id in self.live_quiz_scores:
                score_info = self.live_quiz_scores[student_id]
                total_questions = self.settings.get("live_quiz_questions", 5)
                questions_answered = score_info['total_asked']

                initial_color_hex = self.settings.get("live_quiz_initial_color", "#FF0000")
                final_color_hex = self.settings.get("live_quiz_final_color", "#00FF00")

                if questions_answered >= total_questions:
                    outline_color_orig = final_color_hex
                else:
                    # Interpolate color
                    try:
                        initial_r, initial_g, initial_b = int(initial_color_hex[1:3], 16), int(initial_color_hex[3:5], 16), int(initial_color_hex[5:7], 16)
                        final_r, final_g, final_b = int(final_color_hex[1:3], 16), int(final_color_hex[3:5], 16), int(final_color_hex[5:7], 16)

                        progress = questions_answered / total_questions

                        r = int(initial_r + (final_r - initial_r) * progress)
                        g = int(initial_g + (final_g - initial_g) * progress)
                        b = int(initial_b + (final_b - initial_b) * progress)

                        outline_color_orig = f"#{r:02x}{g:02x}{b:02x}"
                    except ValueError:
                        outline_color_orig = self.settings.get("student_box_outline_color") # Fallback

            font_family = style_overrides.get("font_family", self.settings.get("student_font_family"))
            font_size_world = style_overrides.get("font_size", self.settings.get("student_font_size"))
            font_size_canvas = int(max(6, font_size_world * self.current_zoom_level))
            font_color = style_overrides.get("font_color", self.settings.get("student_font_color"))

            group_id = student_data.get("group_id"); group_indicator_color = None
            if self.settings.get("student_groups_enabled", True) and group_id and group_id in self.student_groups:
                group_data = self.student_groups[group_id]
                group_indicator_color = group_data.get("color")
                # Apply the first matching group rule to the base fill_color and outline_color_orig
                for rule in self.settings.get("conditional_formatting_rules", []):
                    if rule.get("type") == "group" and rule.get("group_id") == group_id:
                        if rule.get("color"): # Check if color is not empty or None
                            fill_color = rule["color"]
                        if rule.get("outline"): # Check if outline is not empty or None
                            outline_color_orig = rule["outline"]
                        break # First matching group rule takes precedence for base colors

            # Now, collect all other (non-group) applicable conditional formatting rules
            active_rules_colors = []
            # Base colors that will be used if no other rules apply or for the first stripe.
            # These might have been set by a group rule already.
            # If active_rules_colors remains empty, these base colors will be used for the whole box.
            
            # Store the base colors derived from defaults, overrides, or group rules.
            # These will be used if no other active_rules_colors are found, or as the first "layer" if we decide to.
            # For now, if active_rules_colors gets populated, those will define the stripes.
            # If active_rules_colors is empty, the single fill_color/outline_color_orig will be used.

            if self.settings.get("conditional_formatting_rules", []):
                for rule in self.settings.get("conditional_formatting_rules", []):
                    if rule.get("type") == "group":
                        continue  # Group rules already processed for base color

                    if self.applies_to_conditional(student_id, rule):
                        rule_fill = rule.get("color")
                        rule_outline = rule.get("outline")
                        if rule_fill or rule_outline: # Only add if the rule specifies a color
                            active_rules_colors.append({
                                "fill": rule_fill if rule_fill else None,
                                "outline": rule_outline if rule_outline else None
                            })
                            # No break here, collect all matching non-group rules.
            
            # If active_rules_colors is empty, the drawing logic
            # will use the existing fill_color and outline_color_orig for the single rectangle.
            # If populated, it will draw stripes based on these collected colors.
            # The original fill_color and outline_color_orig (potentially set by a group rule)
            # can be considered the "base" if no other rules apply.

            live_override_applied = False
            # Process override rules first for live sessions
            if self.is_live_quiz_active or self.is_live_homework_active:
                for rule in self.settings.get("conditional_formatting_rules", []):
                    if rule.get("type") in ["live_quiz_response", "live_homework_yes_no", "live_homework_select"] and \
                       rule.get("application_style") == "override":
                        if self.applies_to_conditional(student_id, rule):
                            if rule.get("color"): fill_color = rule["color"]
                            if rule.get("outline"): outline_color_orig = rule["outline"]
                            active_rules_colors = []  # Clear any standard stripes if overridden
                            live_override_applied = True
                            break

            # Collect stripe rules (standard and live, if no live override took place)
            if not live_override_applied:
                active_rules_colors = [] # Ensure it's clean before collecting stripes
                for rule in self.settings.get("conditional_formatting_rules", []):
                    rule_type = rule.get("type")
                    is_live_rule = rule_type in ["live_quiz_response", "live_homework_yes_no", "live_homework_select"]

                    if rule_type == "group":
                        continue

                    applies_now = False
                    if is_live_rule:
                        if rule.get("application_style") == "stripe":
                            if (self.is_live_quiz_active and rule_type == "live_quiz_response") or \
                               (self.is_live_homework_active and rule_type in ["live_homework_yes_no", "live_homework_select"]):
                                applies_now = self.applies_to_conditional(student_id, rule)
                    else:
                        applies_now = self.applies_to_conditional(student_id, rule)

                    if applies_now:
                        rule_fill = rule.get("color")
                        rule_outline = rule.get("outline")
                        if rule_fill or rule_outline:
                            active_rules_colors.append({
                                "fill": rule_fill if rule_fill else None,
                                "outline": rule_outline if rule_outline else None
                            })

            # Font setup using new specific settings
            name_font_obj = tkfont.Font(family=font_family, size=font_size_canvas, weight="bold")

            behavior_log_font_size_canvas = int(max(5, self.settings.get("behavior_log_font_size", DEFAULT_FONT_SIZE -1) * self.current_zoom_level))
            incident_font_obj = tkfont.Font(family=font_family, size=behavior_log_font_size_canvas)

            quiz_log_font_size_canvas = int(max(5, self.settings.get("quiz_log_font_size", DEFAULT_FONT_SIZE) * self.current_zoom_level))
            quiz_score_font_color_setting = self.settings.get("live_quiz_score_font_color")
            quiz_score_font_bold_setting = self.settings.get("live_quiz_score_font_style_bold")
            quiz_score_font_weight = "bold" if quiz_score_font_bold_setting else "normal"
            quiz_score_font_obj = tkfont.Font(family=font_family, size=quiz_log_font_size_canvas, weight=quiz_score_font_weight)

            homework_log_font_size_canvas = int(max(5, self.settings.get("homework_log_font_size", DEFAULT_FONT_SIZE -1) * self.current_zoom_level))
            hw_score_font_color_setting = self.settings.get("live_homework_score_font_color", DEFAULT_HOMEWORK_SCORE_FONT_COLOR)
            hw_score_font_bold_setting = self.settings.get("live_homework_score_font_style_bold", DEFAULT_HOMEWORK_SCORE_FONT_STYLE_BOLD)
            hw_score_font_weight = "bold" if hw_score_font_bold_setting else "normal"
            # For homework_score_header, use the dedicated homework_log_font_size
            hw_score_font_obj = tkfont.Font(family=font_family, size=homework_log_font_size_canvas, weight=hw_score_font_weight)
            # For homework_score_item, also use homework_log_font_size (or could be a new setting if finer control is needed)
            hw_score_item_font_obj = tkfont.Font(family=font_family, size=homework_log_font_size_canvas, weight=hw_score_font_weight)
            
            self.canvas.delete(student_id)
            rect_tag = ("student_item", student_id, "rect")

            world_padding = 5; canvas_padding = world_padding * self.current_zoom_level
            current_y_offset_for_calc_world = world_padding
            for name_line_text in student_data.get("display_lines", []):
                font_for_calc = tkfont.Font(family=font_family, size=font_size_world, weight="bold")
                current_y_offset_for_calc_world += font_for_calc.metrics('linespace')

            if student_data.get("incident_display_lines"):
                current_y_offset_for_calc_world += world_padding / 2
                for line_info in student_data.get("incident_display_lines", []):
                    line_text, line_type = line_info["text"], line_info["type"]
                    current_font_world_calc = tkfont.Font(family=font_family, size=font_size_world -1)
                    if line_type == "quiz_score": current_font_world_calc = tkfont.Font(family=font_family, size=font_size_world, weight=quiz_score_font_weight)
                    elif line_type == "homework_score_header": current_font_world_calc = tkfont.Font(family=font_family, size=font_size_world, weight=hw_score_font_weight)
                    elif line_type == "homework_score_item": current_font_world_calc = tkfont.Font(family=font_family, size=max(5, font_size_world -1), weight=hw_score_font_weight)
                    elif line_type == "separator": current_font_world_calc = tkfont.Font(family=font_family, size=max(4, font_size_world -2)) # Smaller for separator

                    text_width_pixels_world = current_font_world_calc.measure(line_text)
                    visual_lines_world = 1
                    available_width_for_text_world = world_width - 2 * world_padding
                    if available_width_for_text_world > 0 and text_width_pixels_world > available_width_for_text_world:
                        visual_lines_world = -(-text_width_pixels_world // available_width_for_text_world)
                    current_y_offset_for_calc_world += visual_lines_world * current_font_world_calc.metrics('linespace')

            world_text_content_height_with_padding = current_y_offset_for_calc_world + world_padding
            world_dynamic_height = max(world_base_height, world_text_content_height_with_padding)
            canvas_dynamic_height = world_dynamic_height * self.current_zoom_level
            student_data['_current_world_height'] = world_dynamic_height
            student_data['_current_world_width'] = world_width

            # Box drawing logic:
            if not active_rules_colors:
                # No specific non-group rules apply, draw a single box with base/group colors
                self.canvas.create_rectangle(canvas_x, canvas_y, canvas_x + canvas_width, canvas_y + canvas_dynamic_height,
                                             fill=fill_color, outline=outline_color_orig, width=max(1, int(2 * self.current_zoom_level)), tags=rect_tag)
            else:
                num_effective_rules = min(len(active_rules_colors), 3) # Max 3 stripes
                stripe_height_canvas = canvas_dynamic_height / num_effective_rules

                base_default_fill = self.settings.get("student_box_fill_color")
                base_default_outline = self.settings.get("student_box_outline_color")

                for i in range(num_effective_rules):
                    rule_colors = active_rules_colors[i]
                    stripe_fill = rule_colors.get("fill") if rule_colors.get("fill") else fill_color # Fallback to base fill_color (from group/default)
                    stripe_outline = rule_colors.get("outline") if rule_colors.get("outline") else outline_color_orig # Fallback to base outline_color_orig

                    # If even the fallback is None (e.g. student default is empty string), provide a very basic default.
                    if not stripe_fill: stripe_fill = base_default_fill
                    if not stripe_outline: stripe_outline = base_default_outline

                    stripe_y_start = canvas_y + (i * stripe_height_canvas)
                    stripe_y_end = canvas_y + ((i + 1) * stripe_height_canvas)

                    # For the last stripe, ensure it goes exactly to the bottom edge to avoid floating point issues
                    if i == num_effective_rules - 1:
                        stripe_y_end = canvas_y + canvas_dynamic_height

                    self.canvas.create_rectangle(canvas_x, stripe_y_start, canvas_x + canvas_width, stripe_y_end,
                                                 fill=stripe_fill,
                                                 outline=stripe_outline,
                                                 width=max(1, int(1 * self.current_zoom_level)), # Thinner outline for stripes
                                                 tags=rect_tag + (f"stripe_{i}",))

            # --- Text Drawing ---
            current_y_text_draw_canvas = canvas_y + canvas_padding
            available_text_width_canvas = canvas_width - 2 * canvas_padding
            colored = True if self.settings.get("always_show_text_background_panel", False) or active_rules_colors else False
            if self.settings.get("enable_text_background_panel", True) and colored:
                text_panel_fill = "#F0F0F0" # Light gray for text background
                text_panel_internal_padding = 2 * self.current_zoom_level # Small padding around text within its panel

                # Panel for Name Lines
                name_lines_content_for_panel = student_data.get("display_lines", [])
                if name_lines_content_for_panel:
                    name_block_height_canvas = 0
                    max_name_width_pixels = 0
                    for name_line_text_calc in name_lines_content_for_panel:
                        name_block_height_canvas += name_font_obj.metrics('linespace')
                        max_name_width_pixels = max(max_name_width_pixels, name_font_obj.measure(name_line_text_calc))

                    name_panel_width = min(max_name_width_pixels + 2 * text_panel_internal_padding, available_text_width_canvas - 2 * text_panel_internal_padding)
                    name_panel_height = name_block_height_canvas

                    name_panel_x0 = canvas_x + (canvas_width - name_panel_width) / 2
                    name_panel_y0 = (canvas_y + canvas_padding) - text_panel_internal_padding
                    name_panel_x1 = name_panel_x0 + name_panel_width
                    name_panel_y1 = name_panel_y0 + name_panel_height + 2 * text_panel_internal_padding

                    if name_panel_y1 < (canvas_y + canvas_dynamic_height - canvas_padding * 0.5):
                        self.canvas.create_rectangle(name_panel_x0, name_panel_y0, name_panel_x1, name_panel_y1,
                                                     fill=text_panel_fill, outline="",
                                                     tags=("student_item", student_id, "text_background_name"))

                # Panel for Incident/Score Lines
                incident_lines_content_for_panel = student_data.get("incident_display_lines", [])
                if incident_lines_content_for_panel:
                    incident_block_start_y_for_panel = (canvas_y + canvas_padding) + \
                                                       sum(name_font_obj.metrics('linespace') for _ in name_lines_content_for_panel) + \
                                                       (canvas_padding / 2 if name_lines_content_for_panel else 0)
                    incident_block_height_canvas = 0
                    max_incident_width_pixels = 0

                    for line_info_calc in incident_lines_content_for_panel:
                        line_text_calc, line_type_calc = line_info_calc["text"], line_info_calc["type"]
                        current_font_for_calc = incident_font_obj
                        if line_type_calc == "quiz_score": current_font_for_calc = quiz_score_font_obj
                        elif line_type_calc == "homework_score_header": current_font_for_calc = hw_score_font_obj
                        elif line_type_calc == "homework_score_item": current_font_for_calc = hw_score_item_font_obj
                        elif line_type_calc == "separator": current_font_for_calc = tkfont.Font(family=font_family, size=max(4, int((font_size_world-2)*self.current_zoom_level)))

                        text_width_pixels_canvas_calc = current_font_for_calc.measure(line_text_calc)
                        available_incident_text_width_calc = available_text_width_canvas - (text_panel_internal_padding if line_type_calc == "homework_score_item" else 0)
                        visual_lines_calc = 1
                        if available_incident_text_width_calc > 0 and text_width_pixels_canvas_calc > available_incident_text_width_calc:
                            visual_lines_calc = -(-text_width_pixels_canvas_calc // available_incident_text_width_calc)
                        incident_block_height_canvas += visual_lines_calc * current_font_for_calc.metrics('linespace')
                        max_incident_width_pixels = max(max_incident_width_pixels, min(text_width_pixels_canvas_calc, available_incident_text_width_calc))

                    if incident_block_height_canvas > 0:
                        incident_panel_width = min(max_incident_width_pixels + 2 * text_panel_internal_padding, available_text_width_canvas - 2 * text_panel_internal_padding)
                        incident_panel_height = incident_block_height_canvas

                        inc_panel_x0 = canvas_x + (canvas_width - incident_panel_width) / 2
                        inc_panel_y0 = incident_block_start_y_for_panel - text_panel_internal_padding
                        inc_panel_x1 = inc_panel_x0 + incident_panel_width
                        inc_panel_y1 = inc_panel_y0 + incident_panel_height + 2 * text_panel_internal_padding

                        if inc_panel_y1 < (canvas_y + canvas_dynamic_height - canvas_padding * 0.5):
                             self.canvas.create_rectangle(inc_panel_x0, inc_panel_y0, inc_panel_x1, inc_panel_y1,
                                                         fill=text_panel_fill, outline="",
                                                         tags=("student_item", student_id, "text_background_incidents"))

            # Draw Name Lines (always drawn, panel is conditional)
            name_lines_content = student_data.get("display_lines", [])
            for name_line_text in name_lines_content:
                self.canvas.create_text(canvas_x + canvas_width / 2, current_y_text_draw_canvas, text=name_line_text,
                                        fill=font_color, font=name_font_obj, tags=("student_item", student_id, "text", "student_name"),
                                        anchor=tk.N, width=max(1, available_text_width_canvas), justify=tk.CENTER)
                current_y_text_draw_canvas += name_font_obj.metrics('linespace')

            # Draw Incident/Score Lines (always drawn, panel is conditional)
            incident_lines_content = student_data.get("incident_display_lines", [])
            if incident_lines_content:
                current_y_text_draw_canvas += canvas_padding / 2 # Space before incidents

                for line_info in incident_lines_content:
                    line_text, line_type = line_info["text"], line_info["type"]
                    current_font_canvas_draw, current_color_canvas_draw = incident_font_obj, font_color
                    text_anchor_canvas, text_justify_canvas = tk.N, tk.CENTER
                    text_x_pos_canvas = canvas_x + canvas_width / 2

                    if line_type == "quiz_score": current_font_canvas_draw, current_color_canvas_draw = quiz_score_font_obj, quiz_score_font_color_setting
                    elif line_type == "homework_score_header": current_font_canvas_draw, current_color_canvas_draw = hw_score_font_obj, hw_score_font_color_setting
                    elif line_type == "homework_score_item":
                        current_font_canvas_draw, current_color_canvas_draw = hw_score_item_font_obj, hw_score_font_color_setting
                        text_anchor_canvas, text_justify_canvas = tk.NW, tk.LEFT
                        text_x_pos_canvas = canvas_x + canvas_padding
                    elif line_type == "separator":
                        current_font_canvas_draw = tkfont.Font(family=font_family, size=max(4, int((font_size_world-2)*self.current_zoom_level)))
                        current_color_canvas_draw = "gray"

                    self.canvas.create_text(text_x_pos_canvas, current_y_text_draw_canvas, text=line_text,
                                            fill=current_color_canvas_draw, font=current_font_canvas_draw,
                                            tags=("student_item", student_id, "text", f"student_{line_type}"),
                                            anchor=text_anchor_canvas, width=max(1, available_text_width_canvas if text_anchor_canvas == tk.N else available_text_width_canvas - canvas_padding),
                                            justify=text_justify_canvas)
                    text_width_pixels_canvas = current_font_canvas_draw.measure(line_text)
                    visual_lines_canvas = 1
                    if available_text_width_canvas > 0 and text_width_pixels_canvas > available_text_width_canvas:
                        visual_lines_canvas = -(-text_width_pixels_canvas // available_text_width_canvas)
                    current_y_text_draw_canvas += visual_lines_canvas * current_font_canvas_draw.metrics('linespace')

            if self.settings.get("student_groups_enabled", True) and group_indicator_color:
                indicator_size_canvas = GROUP_COLOR_INDICATOR_SIZE * self.current_zoom_level
                indicator_padding_canvas = 2 * self.current_zoom_level
                indicator_x = canvas_x + canvas_width - indicator_size_canvas - indicator_padding_canvas
                indicator_y = canvas_y + indicator_padding_canvas
                self.canvas.create_rectangle(indicator_x, indicator_y, indicator_x + indicator_size_canvas, indicator_y + indicator_size_canvas,
                                            fill=group_indicator_color, outline=outline_color_orig, tags=("student_item", student_id, "group_indicator"))
            if student_id in self.selected_items:
                sel_outline_width = max(1, int(2 * self.current_zoom_level))
                self.canvas.create_rectangle(canvas_x - sel_outline_width, canvas_y - sel_outline_width,
                                             canvas_x + canvas_width + sel_outline_width, canvas_y + canvas_dynamic_height + sel_outline_width,
                                             outline="red", width=sel_outline_width, tags=("student_item", student_id, "selection_highlight"))
            if self.edit_mode_var.get() and student_id in self.selected_items:
                handle_size_canvas = RESIZE_HANDLE_SIZE * self.current_zoom_level
                br_x = canvas_x + canvas_width - handle_size_canvas / 2 # Center handle on corner
                br_y = canvas_y + canvas_dynamic_height - handle_size_canvas / 2
                self.canvas.create_rectangle(br_x - handle_size_canvas/2, br_y - handle_size_canvas/2,
                                             br_x + handle_size_canvas/2, br_y + handle_size_canvas/2,
                                             fill="gray", outline="black", tags=("student_item", student_id, "resize_handle", "br_handle"))
            if check_collisions and self.settings.get("check_for_collisions", True): self.handle_layout_collision(student_id)
        except AttributeError: pass # Canvas might not be fully initialized during early calls

    def draw_single_furniture(self, furniture_id):
        # ... (same as v51)
        item_data = self.furniture.get(furniture_id)
        if not item_data: return
        world_x, world_y = item_data["x"], item_data["y"]
        world_width = item_data.get("width", DEFAULT_STUDENT_BOX_WIDTH)
        world_height = item_data.get("height", DEFAULT_STUDENT_BOX_HEIGHT)
        canvas_x, canvas_y = self.world_to_canvas_coords(world_x, world_y)
        canvas_width = world_width * self.current_zoom_level
        canvas_height = world_height * self.current_zoom_level
        item_data['_current_world_height'] = world_height
        item_data['_current_world_width'] = world_width
        fill_color = item_data.get("fill_color", "lightgrey")
        outline_color = item_data.get("outline_color", "dimgray")
        name = item_data.get("name", "Furniture")
        try:
            self.canvas.delete(furniture_id)
            rect_tag = ("furniture_item", furniture_id, "rect")
            self.canvas.create_rectangle(canvas_x, canvas_y, canvas_x + canvas_width, canvas_y + canvas_height,
                                         fill=fill_color, outline=outline_color, width=max(1, int(2*self.current_zoom_level)), tags=rect_tag)
            font_size_canvas = int(max(6, (self.settings.get("student_font_size", DEFAULT_FONT_SIZE) -1) * self.current_zoom_level))
            font_spec = (self.settings.get("student_font_family", DEFAULT_FONT_FAMILY), font_size_canvas)
            self.canvas.create_text(canvas_x + canvas_width / 2, canvas_y + canvas_height / 2, text=name,
                                    fill=self.settings.get("student_font_color", DEFAULT_FONT_COLOR), font=font_spec,
                                    tags=("furniture_item", furniture_id, "text"), anchor=tk.CENTER,
                                    width=max(1, canvas_width - int(10*self.current_zoom_level)), justify=tk.CENTER)
            if furniture_id in self.selected_items:
                sel_outline_width = max(1, int(2 * self.current_zoom_level))
                self.canvas.create_rectangle(canvas_x - sel_outline_width, canvas_y - sel_outline_width,
                                             canvas_x + canvas_width + sel_outline_width, canvas_y + canvas_height + sel_outline_width,
                                             outline="red", width=sel_outline_width, tags=("furniture_item", furniture_id, "selection_highlight"))
            if self.edit_mode_var.get() and furniture_id in self.selected_items:
                handle_size_canvas = RESIZE_HANDLE_SIZE * self.current_zoom_level
                br_x = canvas_x + canvas_width - handle_size_canvas / 2
                br_y = canvas_y + canvas_height - handle_size_canvas / 2
                self.canvas.create_rectangle(br_x - handle_size_canvas/2, br_y - handle_size_canvas/2,
                                             br_x + handle_size_canvas/2, br_y + handle_size_canvas/2,
                                             fill="gray", outline="black", tags=("furniture_item", furniture_id, "resize_handle", "br_handle"))
        except AttributeError: pass

    def draw_all_items(self, check_collisions_on_redraw=False):
        if not self.canvas: return
        self.canvas.delete("all") # Clear canvas before redrawing everything

        if self.settings.get("show_grid", False):
            self.draw_grid()

        if self.settings.get("show_rulers", False):
            self.draw_rulers()

        # Draw temporary guides first, so they are under items if needed (though typically on top)
        #self.draw_temporary_guides() # Guides will be drawn after items for better visibility
        # The new self.draw_guides() is called after items.

        all_items_data = list(self.students.values()) + list(self.furniture.values())
        
        if ((self.edit_mode_var.get() == True or self.settings.get("always_show_box_management", False) == True) and self.settings.get("show_canvas_border_lines", False) == True) or self.settings.get("force_canvas_border_lines", False) == True:
            self.canvas.create_line(0,0,1,2000, tags=("border_line", "border_vertical")) # These seem to be fixed debug lines, not dynamic with canvas/zoom
            self.canvas.create_line(0,0,2000,1, tags=("border_line", "border_horizontal")) # Consider removing or making them dynamic if kept.
        if not all_items_data:
            try:
                default_sr_w = self.canvas_orig_width * self.current_zoom_level; default_sr_h = self.canvas_orig_height * self.current_zoom_level
                self.canvas.configure(scrollregion=(0, 0, default_sr_w, default_sr_h))
            except AttributeError: pass
        else:
            min_x_world, min_y_world = float('inf'), float('inf'); max_x_world_br, max_y_world_br = float('-inf'), float('-inf')
            for item_data in all_items_data:
                item_world_x, item_world_y = item_data['x'], item_data['y']
                item_world_w = item_data.get('_current_world_width', item_data.get('width', DEFAULT_STUDENT_BOX_WIDTH))
                item_world_h = item_data.get('_current_world_height', item_data.get('height', DEFAULT_STUDENT_BOX_HEIGHT))
                min_x_world = min(min_x_world, item_world_x); min_y_world = min(min_y_world, item_world_y)
                max_x_world_br = max(max_x_world_br, item_world_x + item_world_w); max_y_world_br = max(max_y_world_br, item_world_y + item_world_h)
            padding_world = 100
            scroll_min_x_canvas, scroll_min_y_canvas = self.world_to_canvas_coords(min_x_world - padding_world, min_y_world - padding_world)
            scroll_max_x_canvas, scroll_max_y_canvas = self.world_to_canvas_coords(max_x_world_br + padding_world, max_y_world_br + padding_world)
            try:
                visible_canvas_width = self.canvas.winfo_width(); visible_canvas_height = self.canvas.winfo_height()
                final_scroll_max_x = max(scroll_max_x_canvas, scroll_min_x_canvas + visible_canvas_width)
                final_scroll_max_y = max(scroll_max_y_canvas, scroll_min_y_canvas + visible_canvas_height)
            except AttributeError: final_scroll_max_x, final_scroll_max_y = scroll_max_x_canvas, scroll_max_y_canvas # Fallback if winfo fails
            final_scroll_min_x = min(scroll_min_x_canvas, 0); final_scroll_min_y = min(scroll_min_y_canvas, 0)
            try: self.canvas.config(scrollregion=(final_scroll_min_x, final_scroll_min_y, final_scroll_max_x, final_scroll_max_y))
            except AttributeError: pass
        for student_id in self.students: self.draw_single_student(student_id, check_collisions=check_collisions_on_redraw)
        for furniture_id in self.furniture: self.draw_single_furniture(furniture_id)

        self.draw_guides() # Draw guides on top of items
        self.update_toggle_incidents_button_text(); self.update_zoom_display()
        self.update_toggle_rulers_button_text()
        self.update_toggle_grid_button_text()

    def draw_guides(self):
        """Draws all stored guides on the canvas."""
        if not self.canvas: return
        self.canvas.delete("guide") # Delete only items tagged 'guide'

        for guide_info in self.guides:
            guide_type = self.guides[guide_info].get('type')
            world_coord = self.guides[guide_info].get('world_coord') #guide_info['world_coord']
            guide_id_tag = self.guides[guide_info].get('id') #guide_info['id'] # e.g., "guide_v_1"

            canvas_item_id = None
            if guide_type == 'h': # Horizontal guide
                _, screen_y = self.world_to_canvas_coords(0, world_coord)
                canvas_item_id = self.canvas.create_line(
                    0, screen_y, self.canvas.winfo_width(), screen_y,
                    fill=self.guide_line_color, tags=("guide", guide_id_tag, "guide_h"), width=1, dash=(4, 2)
                )
            elif guide_type == 'v': # Vertical guide
                screen_x, _ = self.world_to_canvas_coords(world_coord, 0)
                canvas_item_id = self.canvas.create_line(
                    screen_x, 0, screen_x, self.canvas.winfo_height(),
                    fill=self.guide_line_color, tags=("guide", guide_id_tag, "guide_v"), width=1, dash=(4, 2)
                )
            self.guides[guide_info]['canvas_item_id'] = canvas_item_id #guide_info['canvas_item_id'] = canvas_item_id # Store/update Tkinter canvas item ID

    def toggle_grid_visibility(self):
        self.settings["show_grid"] = not self.settings.get("show_grid", False)
        self.draw_all_items()
        self.update_toggle_grid_button_text()
        self.update_status(f"Grid {'shown' if self.settings['show_grid'] else 'hidden'}.")

    def reload_canvas(self, event=None):
        self.draw_all_items()
        self.update_status("Reloaded")

    def update_toggle_grid_button_text(self):
        if hasattr(self, 'toggle_grid_btn'):
            text = "Hide Grid" if self.settings.get("show_grid", False) else "Show Grid"
            self.toggle_grid_btn.config(text=text)

    def toggle_add_guide_mode(self, mode: str, button_pressed: ttk.Button):
        """Toggles the mode for adding guides and updates button visuals."""
        if self.add_guide_mode == mode: # Toggle off
            self.add_guide_mode = None
            if self.active_guide_button:
                self.active_guide_button.state(['!pressed', '!focus']) # Remove pressed and focus state
            self.active_guide_button = None
            self.update_status("Add guide mode deactivated.")
        else: # Toggle on for the given mode
            if self.active_guide_button: # Deactivate any other active guide button
                self.active_guide_button.state(['!pressed', '!focus'])

            self.add_guide_mode = mode
            self.active_guide_button = button_pressed
            self.active_guide_button.state(['pressed', 'focus']) # Set pressed and focus state
            self.update_status(f"Click on canvas to add a {mode} guide. Click button again to cancel.")

        # Ensure other button is not in pressed state
        if mode == "vertical" and hasattr(self, 'add_h_guide_btn') and self.add_h_guide_btn != button_pressed:
            self.add_h_guide_btn.state(['!pressed', '!focus'])
        elif mode == "horizontal" and hasattr(self, 'add_v_guide_btn') and self.add_v_guide_btn != button_pressed:
            self.add_v_guide_btn.state(['!pressed', '!focus'])

    def draw_grid(self):
        if not self.canvas: return
        grid_size = self.settings.get("grid_size", DEFAULT_GRID_SIZE)
        grid_color = self.settings.get("grid_color", "#d3d3d3")
        if grid_size <= 0: return

        canvas_width_screen = self.canvas.winfo_width()
        canvas_height_screen = self.canvas.winfo_height()

        # Get the visible world coordinates
        world_x_start, world_y_start = self.canvas_to_world_coords(0, 0)
        world_x_end, world_y_end = self.canvas_to_world_coords(canvas_width_screen, canvas_height_screen)

        # Adjust start coordinates to the nearest lower grid line
        start_grid_x_world = int(world_x_start / grid_size) * grid_size
        start_grid_y_world = int(world_y_start / grid_size) * grid_size

        # Vertical lines
        for world_x in range(start_grid_x_world, int(world_x_end) + grid_size, grid_size):
            canvas_x, _ = self.world_to_canvas_coords(world_x, world_y_start)
            # Draw line across the current visible canvas height, adjusted for ruler if present
            line_y_start_on_canvas = self.ruler_thickness if self.settings.get("show_rulers", False) else 0
            self.canvas.create_line(canvas_x, line_y_start_on_canvas, canvas_x, canvas_height_screen,
                                    fill=grid_color, tags="grid_line", width=1, dash=(2,4))

        # Horizontal lines
        for world_y in range(start_grid_y_world, int(world_y_end) + grid_size, grid_size):
            _, canvas_y = self.world_to_canvas_coords(world_x_start, world_y)
            # Draw line across the current visible canvas width, adjusted for ruler if present
            line_x_start_on_canvas = self.ruler_thickness if self.settings.get("show_rulers", False) else 0
            self.canvas.create_line(line_x_start_on_canvas, canvas_y, canvas_width_screen, canvas_y,
                                    fill=grid_color, tags="grid_line", width=1, dash=(2,4))

    def toggle_rulers_visibility(self):
        self.settings["show_rulers"] = not self.settings.get("show_rulers", False)

        if not self.settings["show_rulers"]: # Rulers are being hidden
            self.active_ruler_guide_coord_x = None # Always cancel pending guide placement
            self.active_ruler_guide_coord_y = None
            if not self.settings.get("guides_stay_when_rulers_hidden", True):
                self.clear_temporary_guides() # Clear data and canvas items
            else:
                # Keep data, just delete canvas items and nullify canvas_id
                if self.canvas:
                    for guide_info in self.temporary_guides:
                        if guide_info.get('canvas_id') is not None:
                            self.canvas.delete(guide_info['canvas_id'])
                            guide_info['canvas_id'] = None

        self.draw_all_items() # This will redraw rulers if shown, and guides if data exists and rulers shown
        self.update_toggle_rulers_button_text()
        self.update_status(f"Rulers {'shown' if self.settings['show_rulers'] else 'hidden'}.")

    def update_toggle_rulers_button_text(self):
        if hasattr(self, 'toggle_rulers_btn'):
            text = "Hide Rulers" if self.settings.get("show_rulers", False) else "Show Rulers"
            self.toggle_rulers_btn.config(text=text)

    def draw_rulers(self):
        if not self.canvas: return
        # Horizontal Ruler (Top)
        self.canvas.create_rectangle(0, 0, self.canvas.winfo_width(), self.ruler_thickness,
                                    #  fill=self.ruler_bg_color,
                                     outline=self.ruler_line_color, tags="ruler_bg")
        # Vertical Ruler (Left)
        self.canvas.create_rectangle(0, self.ruler_thickness, self.ruler_thickness, self.canvas.winfo_height(),
                                    #  fill=self.ruler_bg_color, 
                                     outline=self.ruler_line_color, tags="ruler_bg")

        # Markings
        canvas_width = self.canvas.winfo_width()
        canvas_height = self.canvas.winfo_height()

        # Determine visible world coordinates
        world_x_start, world_y_start = self.canvas_to_world_coords(0,0)
        world_x_end, world_y_end = self.canvas_to_world_coords(canvas_width, canvas_height)

        # Adjust for ruler thickness when calculating visible world range for markings
        world_x_for_hruler_start, _ = self.canvas_to_world_coords(self.ruler_thickness, self.ruler_thickness)
        world_x_for_hruler_end, _ = self.canvas_to_world_coords(canvas_width, self.ruler_thickness)

        _, world_y_for_vruler_start = self.canvas_to_world_coords(self.ruler_thickness, self.ruler_thickness)
        _, world_y_for_vruler_end = self.canvas_to_world_coords(self.ruler_thickness, canvas_height)


        # Dynamic interval based on zoom - simplified
        interval = 50
        if self.current_zoom_level < 0.5: interval = 100
        elif self.current_zoom_level > 2: interval = 20

        # Horizontal Markings
        start_mark_x = int(world_x_for_hruler_start / interval) * interval
        for world_x in range(start_mark_x, int(world_x_for_hruler_end) + interval, interval):
            canvas_x, _ = self.world_to_canvas_coords(world_x, world_y_start) # Use world_y_start for consistency
            if canvas_x >= self.ruler_thickness and canvas_x <= canvas_width:
                tick_len = 5 if world_x % (interval * 2) != 0 else 10
                self.canvas.create_line(canvas_x, self.ruler_thickness - tick_len, canvas_x, self.ruler_thickness,
                                        fill=self.ruler_line_color, tags="ruler_marking")
                if tick_len == 10:
                    self.canvas.create_text(canvas_x, self.ruler_thickness - tick_len - 5, text=str(world_x),
                                            fill=self.ruler_text_color, anchor=tk.S, tags="ruler_marking_text", font=(DEFAULT_FONT_FAMILY, 8))
        # Vertical Markings
        start_mark_y = int(world_y_for_vruler_start / interval) * interval
        for world_y in range(start_mark_y, int(world_y_for_vruler_end) + interval, interval):
            _, canvas_y = self.world_to_canvas_coords(world_x_start, world_y) # Use world_x_start for consistency
            if canvas_y >= self.ruler_thickness and canvas_y <= canvas_height:
                tick_len = 5 if world_y % (interval * 2) != 0 else 10
                self.canvas.create_line(self.ruler_thickness - tick_len, canvas_y, self.ruler_thickness, canvas_y,
                                        fill=self.ruler_line_color, tags="ruler_marking")
                if tick_len == 10:
                    self.canvas.create_text(self.ruler_thickness - tick_len - 5, canvas_y, text=str(world_y),
                                            fill=self.ruler_text_color, anchor=tk.E, tags="ruler_marking_text", font=(DEFAULT_FONT_FAMILY, 8))

    def draw_temporary_guides(self):
        if not self.canvas: return
        self.canvas.delete("temporary_guide") # Clear old guides before redrawing

        # Redraw persistent guides based on self.temporary_guides list
        # Note: This loop is for redrawing guides that were previously added and stored.
        # The actual 'canvas_id' stored in the dictionary is not used here for redrawing,
        # as we are clearing all "temporary_guide" tagged items and redrawing.
        # If we wanted to selectively update/remove, canvas_id would be useful.
        for guide_info in self.temporary_guides:
            guide_type = guide_info['type']
            world_coord = guide_info['world_coord']

            guide_canvas_id = None
            if guide_type == 'h': # Horizontal guide
                # Convert world_coord (y) to current screen coordinates
                _, screen_y = self.world_to_canvas_coords(0, world_coord) # x doesn't matter for horizontal line screen y
                guide_canvas_id = self.canvas.create_line(0, screen_y, self.canvas.winfo_width(), screen_y,
                                                          fill=self.guide_line_color, tags=("temporary_guide", guide_info.get('id', 'unknown_guide')), width=1, dash=(4, 2))
            elif guide_type == 'v': # Vertical guide
                # Convert world_coord (x) to current screen coordinates
                screen_x, _ = self.world_to_canvas_coords(world_coord, 0) # y doesn't matter for vertical line screen x
                guide_canvas_id = self.canvas.create_line(screen_x, 0, screen_x, self.canvas.winfo_height(),
                                                          fill=self.guide_line_color, tags=("temporary_guide", guide_info.get('id', 'unknown_guide')), width=1, dash=(4, 2))
            if guide_canvas_id:
                guide_info['canvas_id'] = guide_canvas_id # Store the canvas ID

    def clear_temporary_guides(self):
        self.temporary_guides.clear()
        if self.canvas:
            self.canvas.delete("temporary_guide")

    def _calculate_quiz_score_percentage(self, log_entry):
        """Calculates the score percentage for a given quiz log entry."""
        if not log_entry or log_entry.get("type") != "quiz":
            return None

        marks_data = log_entry.get("marks_data", {})
        num_questions = log_entry.get("num_questions", 0)
        score_details = log_entry.get("score_details") # For live quiz sessions

        if num_questions <= 0 and not score_details: # Cannot calculate percentage if no questions or score details
            return None

        # Handle live quiz session scores first
        if score_details and isinstance(score_details, dict):
            correct = score_details.get("correct", 0)
            total_asked = score_details.get("total_asked", 0)
            if total_asked > 0:
                return (correct / total_asked) * 100
            return 0 # Or None if no questions asked yet

        # Handle manually logged quiz scores with marks_data
        if not marks_data or num_questions <= 0: # Need marks_data and num_questions for this path
             return None

        total_earned_points = 0
        total_possible_points_main = 0 # Points from questions that contribute to the main total (not bonus)

        quiz_mark_types_settings = self.settings.get("quiz_mark_types", DEFAULT_QUIZ_MARK_TYPES)

        # Determine the point value of a single "fully correct" non-bonus question.
        # This is used as the basis for the percentage calculation if num_questions is the primary driver.
        # A common approach is to find the "correct" mark type.
        default_points_per_main_question = 1 # Fallback
        correct_mark_type = next((mt for mt in quiz_mark_types_settings if mt.get("id") == "mark_correct" and mt.get("contributes_to_total")), None)
        if correct_mark_type:
            default_points_per_main_question = correct_mark_type.get("default_points", 1)

        total_possible_points_main = num_questions * default_points_per_main_question

        for mark_id, count in marks_data.items():
            if not isinstance(count, (int, float)) or count == 0:
                continue # Skip if count is not a number or zero

            mark_config = next((mt for mt in quiz_mark_types_settings if mt.get("id") == mark_id), None)
            if not mark_config:
                continue

            points_for_this_mark_type = count * mark_config.get("default_points", 0)

            if not mark_config.get("is_extra_credit", False):
                # Only add to total_earned_points if it's not extra credit,
                # as extra credit is handled separately to potentially exceed 100%.
                # This logic assumes that points for main questions are summed up here.
                total_earned_points += points_for_this_mark_type
            else: # Is extra credit
                total_earned_points += points_for_this_mark_type # Add extra credit to earned points

        if total_possible_points_main > 0:
            return (total_earned_points / total_possible_points_main) * 100
        elif total_earned_points > 0 : # e.g. only extra credit was scored, and no main possible points
            return 100.0 # Or handle as a special case, like "Bonus Achieved"

        return 0.0 # Default to 0% if no points possible or earned meaningfully

    def handle_layout_collision(self, moved_item_id):
        # ... (same as v51)
        if moved_item_id not in self.students: return
        moved_item_data = self.students[moved_item_id]
        moved_x1, moved_y1 = moved_item_data['x'], moved_item_data['y']
        moved_w = moved_item_data.get('_current_world_width', moved_item_data.get('width', DEFAULT_STUDENT_BOX_WIDTH))
        moved_h = moved_item_data.get('_current_world_height', moved_item_data.get('height', DEFAULT_STUDENT_BOX_HEIGHT))
        moved_x2, moved_y2 = moved_x1 + moved_w, moved_y1 + moved_h
        items_to_shift_data = []
        all_other_items = []
        for sid, sdata in self.students.items():
            if sid != moved_item_id: all_other_items.append({'id': sid, 'type': 'student', 'data': sdata})
        for fid, fdata in self.furniture.items(): all_other_items.append({'id': fid, 'type': 'furniture', 'data': fdata})
        for other_item_info in all_other_items:
            other_id, other_data = other_item_info['id'], other_item_info['data']
            other_x1, other_y1 = other_data['x'], other_data['y']
            other_w = other_data.get('_current_world_width', other_data.get('width', DEFAULT_STUDENT_BOX_WIDTH))
            other_h = other_data.get('_current_world_height', other_data.get('height', DEFAULT_STUDENT_BOX_HEIGHT))
            other_x2, other_y2 = other_x1 + other_w, other_y1 + other_h
            is_colliding = not (moved_x2 <= other_x1 or moved_x1 >= other_x2 or moved_y2 <= other_y1 or moved_y1 >= other_y2)
            if is_colliding:
                vertical_overlap = moved_y2 - other_y1
                if vertical_overlap > 0 :
                    shift_by = vertical_overlap + LAYOUT_COLLISION_OFFSET
                    items_to_shift_data.append({'id': other_id, 'type': other_item_info['type'], 'old_x': other_x1, 'old_y': other_y1, 'new_x': other_x1, 'new_y': other_y1 + shift_by})
        if items_to_shift_data:
            self.execute_command(MoveItemsCommand(self, items_to_shift_data))
            self.update_status(f"Adjusted layout for {len(items_to_shift_data)} items due to overlap with {moved_item_data['full_name']}.")

    def world_to_canvas_coords(self, world_x, world_y):
        """
        Converts world coordinates to the "infinite" virtual canvas coordinates
        for drawing. This is the forward transformation.
        """
        canvas_x = (world_x * self.current_zoom_level) + self.pan_x
        canvas_y = (world_y * self.current_zoom_level) + self.pan_y
        return canvas_x, canvas_y
    
    def world_to_canvas_coords_guides(self, world_x, world_y):
        """
        Converts world coordinates to the "infinite" virtual canvas coordinates
        for drawing. This is the forward transformation.
        """
        canvas_x = (world_x * self.current_zoom_level) + self.pan_x
        canvas_y = (world_y * self.current_zoom_level) + self.pan_y
        return canvas_x, canvas_y
    
    def canvas_to_world_coords(self, screen_x, screen_y):
        """
        Converts screen coordinates to world coordinates, accounting for
        pan, zoom, and canvas scrolling.
        """
        # Step 1: Account for canvas scrolling (if any)
        # This gives the coordinate on the "infinite" virtual canvas
        true_canvas_x = self.canvas.canvasx(screen_x)
        true_canvas_y = self.canvas.canvasy(screen_y)

        # Step 2: Account for panning and zooming (the inverse of drawing)
        # This is the crucial step you were missing.
        world_x = (true_canvas_x - self.pan_x) / self.zoom_level
        world_y = (true_canvas_y - self.pan_y) / self.zoom_level

        return world_x, world_y
    
    def canvas_to_world_coords_guides(self, screen_x, screen_y):
        """
        Converts screen coordinates to world coordinates, accounting for
        pan, zoom, and canvas scrolling.
        """
        # Step 1: Account for canvas scrolling (if any)
        # This gives the coordinate on the "infinite" virtual canvas
        true_canvas_x = self.canvas.canvasx(screen_x)
        true_canvas_y = self.canvas.canvasy(screen_y)

        # Step 2: Account for panning and zooming (the inverse of drawing)
        # This is the crucial step you were missing.
        world_x = (true_canvas_x - self.pan_x) / self.current_zoom_level
        world_y = (true_canvas_y - self.pan_y) / self.current_zoom_level

        return world_x, world_y
    

    def update_zoom_display(self):
        if hasattr(self, 'zoom_display_label') and self.zoom_display_label:
            if self.settings.get("show_zoom_level_display", True):
                #self.zoom_display_label.config(text=f"{self.current_zoom_level*100:.0f}%")
                self.zoom_var.set(value=str(self.current_zoom_level*100.0))
                if not self.zoom_display_label.winfo_ismapped():
                    zoom_in_btn = next((child for child in self.zoom_display_label.master.winfo_children() if isinstance(child, ttk.Button) and "Zoom In" in child.cget("text")), None)
                    if zoom_in_btn: self.zoom_display_label.pack(side=tk.LEFT, padx=1, after=zoom_in_btn)
                    else: self.zoom_display_label.pack(side=tk.LEFT, padx=1)
            elif self.zoom_display_label.winfo_ismapped(): self.zoom_display_label.pack_forget()

    def update_zoom_display2(self):
        #print("Return")
        self.current_zoom_level = float(self.zoom_var.get())/100.0
        self.zoom_var.set(value=str(float(self.current_zoom_level)*100.0))
        self.zoom_canvas(1)

    def zoom_canvas(self, factor):
        # ... (same as v51)
        if self.password_manager.is_locked: return
        world_center_x_before, world_center_y_before = self.canvas_to_world_coords(self.canvas.winfo_width() // 2, self.canvas.winfo_height() // 2)
        if factor == 0: self.current_zoom_level = 1.0
        else: self.current_zoom_level = max(0.1, min(self.current_zoom_level * factor, 10.0))
        self.draw_all_items(check_collisions_on_redraw=False)
        # Centering logic after zoom could be added here if desired, similar to v50/v51
        self.update_status(f"Zoom level: {self.current_zoom_level:.2f}x"); self.update_zoom_display(); self.password_manager.record_activity()
        self.zoom_var.set(value=str(self.current_zoom_level*100.0))
        self.zoom_display_label.configure(textvariable=self.zoom_var)
        #print(self.zoom_var.get())

    def on_mousewheel_zoom(self, event):
        if self.password_manager.is_locked: return
        factor = 0.9 if (event.num == 5 or event.delta < 0) else 1.1
        self.zoom_canvas(factor)
    def on_pan_start(self, event):
        # ... (same as v51)
        if self.password_manager.is_locked: return
        world_event_x, world_event_y = self.canvas_to_world_coords(event.x, event.y)
        item_ids_under_cursor = self.canvas.find_overlapping(event.x-1, event.y-1, event.x+1, event.y+1) # Use screen coords for find_overlapping
        clicked_on_item = False
        for item_canvas_id in item_ids_under_cursor:
            tags = self.canvas.gettags(item_canvas_id)
            if any(t.startswith("student_") or t.startswith("furniture_") for t in tags if "rect" in tags or "resize_handle" in tags):
                clicked_on_item = True; break
        if not clicked_on_item:
            self.canvas.scan_mark(event.x, event.y); self.update_status("Panning canvas..."); self._drag_started_on_item = False
        else: self._drag_started_on_item = True
        self.password_manager.record_activity()

    def on_pan_move(self, event):
        if self.password_manager.is_locked: return
        if not self._drag_started_on_item: self.canvas.scan_dragto(event.x, event.y, gain=1)
        self.password_manager.record_activity()
    def on_pan_end(self, event):
        if self.password_manager.is_locked: return
        if not self._drag_started_on_item: self.update_status("Canvas panned.")
        self._drag_started_on_item = False; self.password_manager.record_activity()
    def on_canvas_ctrl_click(self, event):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Select", "Enter password to select items:"): return
        world_click_x, world_click_y = self.canvas_to_world_coords(event.x, event.y)
        item_canvas_ids = self.canvas.find_overlapping(world_click_x -1, world_click_y -1, world_click_x +1, world_click_y+1) # Use screen coords
        topmost_item_id, topmost_item_type = None, None
        for item_c_id in reversed(item_canvas_ids):
            tags = self.canvas.gettags(item_c_id); current_item_id, current_item_type = None, None
            for tag in tags:
                if tag.startswith("student_") and tag in self.students: current_item_id, current_item_type = tag, "student"; break
                elif tag.startswith("furniture_") and tag in self.furniture: current_item_id, current_item_type = tag, "furniture"; break
            if current_item_id and any("rect" in t for t in tags): topmost_item_id, topmost_item_type = current_item_id, current_item_type; break
        if topmost_item_id:
            if topmost_item_id in self.selected_items: self.selected_items.remove(topmost_item_id)
            else: self.selected_items.add(topmost_item_id)
            if topmost_item_type == "student": self.draw_single_student(topmost_item_id)
            elif topmost_item_type == "furniture": self.draw_single_furniture(topmost_item_id)
            self.update_status(f"{len(self.selected_items)} items selected.")
        self.password_manager.record_activity()

    def _get_guide_at_canvas_coords(self, event_x, event_y, tolerance=10):
        """
        Checks if a click at canvas coordinates (event_x, event_y) is on/near a guide.
        Returns the ID of the guide if hit, otherwise None.
        Uses a tolerance for easier clicking.
        """
        for guide_info in reversed(self.guides): # Check topmost guides first
            if self.guides[guide_info].get('canvas_item_id') is None: #guide_info.get('canvas_item_id') is None:
                continue # Guide not drawn or ID missing

            # Get current screen coordinates of the guide line
            # For Tkinter, self.canvas.coords(item_id) returns [x1, y1, x2, y2]
            try:
                coords = self.canvas.coords(self.guides[guide_info].get('canvas_item_id')) #guide_info['canvas_item_id'])
                if not coords: continue
            except tk.TclError: # Item might have been deleted unexpectedly
                continue

            guide_type = self.guides[guide_info].get('type')#guide_info['type']

            if guide_type == 'h': # Horizontal guide
                # coords are [x1, y1, x2, y1] - check y-coordinate proximity
                guide_screen_y = coords[1]
                if abs(event_y - guide_screen_y) < tolerance and \
                   min(coords[0], coords[2]) <= event_x <= max(coords[0], coords[2]):
                    return self.guides[guide_info].get('id') #guide_info['id']
            elif guide_type == 'v': # Vertical guide
                # coords are [x1, y1, x1, y2] - check x-coordinate proximity
                guide_screen_x = coords[0]
                if abs(event_x - guide_screen_x) < tolerance and \
                   min(coords[1], coords[3]) <= event_y <= max(coords[1], coords[3]):
                    return self.guides[guide_info].get('id') #guide_info['id']
        return None

    def on_canvas_left_press(self, event):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Interact", "Enter password to interact with canvas:"): return
        self.canvas.focus_set()

        # Check if dragging is allowed # This is obsolete (I removed it's function)
        if False:#not self.settings.get("allow_box_dragging", True):
            # If dragging is disabled, still allow selection clicks but not drag initiation for items.
            # Guide interaction and ruler interaction might still be allowed or handled separately.
            world_event_x_no_drag, world_event_y_no_drag = self.canvas_to_world_coords(event.x, event.y)
            item_canvas_ids_no_drag = self.canvas.find_overlapping(world_event_x_no_drag -1, world_event_y_no_drag -1, world_event_x_no_drag +1, world_event_y_no_drag +1)
            clicked_item_id_no_drag, clicked_item_type_no_drag = None, None
            for item_c_id_nd in reversed(item_canvas_ids_no_drag):
                tags_nd = self.canvas.gettags(item_c_id_nd); temp_id_nd, temp_type_nd, is_main_rect_nd = None, None, False
                for tag_nd in tags_nd:
                    if tag_nd.startswith("student_") and tag_nd in self.students: temp_id_nd, temp_type_nd = tag_nd, "student"
                    elif tag_nd.startswith("furniture_") and tag_nd in self.furniture: temp_id_nd, temp_type_nd = tag_nd, "furniture"
                    if "rect" in tag_nd: is_main_rect_nd = True
                if temp_id_nd and is_main_rect_nd: clicked_item_id_no_drag, clicked_item_type_no_drag = temp_id_nd, temp_type_nd; break

            if clicked_item_id_no_drag: # An item was clicked
                if not (event.state & 0x0004): # Ctrl NOT pressed
                    if clicked_item_id_no_drag not in self.selected_items:
                        self.deselect_all_items()
                        self.selected_items.add(clicked_item_id_no_drag)
                    # If it's already selected and Ctrl is not pressed, it remains selected (standard behavior)
                else: # Ctrl IS pressed
                    if clicked_item_id_no_drag in self.selected_items:
                        self.selected_items.remove(clicked_item_id_no_drag)
                    else:
                        self.selected_items.add(clicked_item_id_no_drag)

                # Redraw the clicked item to show selection state
                if clicked_item_type_no_drag == "student": self.draw_single_student(clicked_item_id_no_drag)
                elif clicked_item_type_no_drag == "furniture": self.draw_single_furniture(clicked_item_id_no_drag)
                self.update_status(f"{len(self.selected_items)} items selected. Dragging disabled.")
                # Set _potential_click_target for on_canvas_release to handle logging
                self._potential_click_target = clicked_item_id_no_drag
                # Initialize drag_data minimally for on_canvas_release to correctly identify a click (not a drag)
                self.drag_data = {"x": world_event_x_no_drag, "y": world_event_y_no_drag, "item_id": None,
                                  "start_x_world": world_event_x_no_drag, "start_y_world": world_event_y_no_drag,
                                  "original_positions": {}, "is_resizing": False,
                                  "_actual_drag_initiated": False} # Ensure drag is not initiated
                self._drag_started_on_item = False # Important for on_canvas_drag
                return # Consume event, selection handled, drag explicitly not initiated

            # If click was not on an item, and dragging is disabled, do nothing more for left press.
            # General context menu (right click) is handled by on_canvas_right_press.
            # Guide/Ruler interactions might still proceed if their logic is before this check or separate.
            # For now, let's assume if dragging is off, item interaction is limited to selection.
            # self.update_status("Dragging is disabled.") # Optional status update
            # return # Consume the event to prevent other bindings if any

        x_coords = event.x; y_coords = event.y
        world_event_x, world_event_y = self.canvas_to_world_coords(event.x, event.y)
        # Guide Creation via Add Guide Mode
        if self.add_guide_mode:
            # Check if click is outside ruler areas to prevent placing guide when trying to use ruler for legacy placement
            on_h_ruler = self.settings.get("show_rulers", False) and world_event_y < self.ruler_thickness and world_event_x > self.ruler_thickness
            on_v_ruler = self.settings.get("show_rulers", False) and world_event_x < self.ruler_thickness and world_event_y > self.ruler_thickness

            if not on_h_ruler and not on_v_ruler:
                world_x, world_y = self.canvas_to_world_coords_guides(event.x, event.y)
                guide_type_to_add = self.add_guide_mode
                
                current_guide_id_num = self.next_guide_id_num
                self.next_guide_id_num += 1
                # self.settings["next_guide_id_num"] = self.next_guide_id_num # Will be saved with main data if guides are persistent

                guide_id_str = f"guide_{guide_type_to_add[0]}_{current_guide_id_num}"
                world_coord_to_use = world_x if guide_type_to_add == 'vertical' else world_y

                new_guide_data = {'id': guide_id_str, 'type': guide_type_to_add[0], 'world_coord': world_coord_to_use, 'canvas_item_id': None}
                #self.guides[guide_id_str] = (new_guide_data)

                self.execute_command(AddGuideCommand(self, guide_id_str, guide_type_to_add, new_guide_data, self.next_guide_id_num, ))
                
                #self.draw_all_items() # Redraws everything including the new guide via self.draw_guides()
                self.update_status(f"Added {guide_type_to_add} guide ({guide_id_str}) at {world_coord_to_use:.0f}.")

                # Deactivate add guide mode and reset button state
                if self.active_guide_button:
                    self.toggle_add_guide_mode(self.add_guide_mode, self.active_guide_button) # This will set mode to None and fix button
                else: # Should not happen if button toggling is correct
                    self.add_guide_mode = None
                return # Consume event
            # else: Click was on a ruler, let ruler interaction (legacy placement) handle it or do nothing.

        # Check for guide dragging (if not in add_guide_mode)
        if not self.add_guide_mode:
            clicked_guide_id = self._get_guide_at_canvas_coords(world_event_x, world_event_y)
            if clicked_guide_id:
                #guide_info = next((g for g in self.guides if g['id'] == clicked_guide_id), None)
                guide_info = self.guides.get(clicked_guide_id)
                if guide_info:
                    self.drag_data = {
                        "item_id": None, # Ensures student/furniture drag logic doesn't run
                        "is_dragging_guide": True,
                        "dragged_guide_id": clicked_guide_id,
                        "dragged_guide_type": guide_info['type'],
                        "original_world_coord": guide_info['world_coord'],
                        # Store canvas click coords, world coords will be calculated in on_canvas_drag relative to this
                        "start_click_canvas_x": event.x,
                        "start_click_canvas_y": event.y,
                        "start_drag_world_x": guide_info['world_coord'] if guide_info['type'] != 'v' else self.canvas_to_world_coords(event.x, event.y)[0],
                        "start_drag_world_y": guide_info['world_coord'] if guide_info['type'] != 'h' else self.canvas_to_world_coords(event.x, event.y)[1],
                    }
                    self._drag_started_on_item = True # To prevent canvas scan_mark
                    self.update_status(f"Dragging guide {clicked_guide_id}")
                    if guide_info['type'] == 'h': self.canvas.config(cursor="sb_v_double_arrow")
                    else: self.canvas.config(cursor="sb_h_double_arrow")
                    self.password_manager.record_activity()
                    return # Consume event, guide drag initiated

        # Ruler interaction (Legacy guide placement, can be kept or removed)
        if self.settings.get("show_rulers", False) and not self.add_guide_mode: # Only if not in button-activated add_guide_mode
            if world_event_y < self.ruler_thickness and world_event_x > self.ruler_thickness: # Horizontal ruler
                self.active_ruler_guide_coord_x, _ = self.canvas_to_world_coords(event.x, event.y)
                self.active_ruler_guide_coord_y = None
                self.toggle_add_guide_mode("vertical", self.add_v_guide_btn)
                self.update_status(f"Click on canvas to place vertical guide") #at x={self.active_ruler_guide_coord_x:.0f}")
                return # Consume click
            # Vertical ruler area (left)
            elif world_event_x < self.ruler_thickness and world_event_y > self.ruler_thickness:
                _, self.active_ruler_guide_coord_y = self.canvas_to_world_coords(event.x, event.y)
                self.active_ruler_guide_coord_x = None
                self.toggle_add_guide_mode("horizontal", self.add_h_guide_btn)
                self.update_status(f"Click on canvas to place horizontal guide") # at y={self.active_ruler_guide_coord_y:.0f}")
                return # Consume click

        # Place active guide if one is pending and click is on canvas proper
        if self.active_ruler_guide_coord_x is not None or self.active_ruler_guide_coord_y is not None:
            # Check if click is outside ruler areas
            if not (world_event_y < self.ruler_thickness and world_event_x > self.ruler_thickness) and \
               not (world_event_x < self.ruler_thickness and world_event_y > self.ruler_thickness):

                current_guide_id_num = self.next_guide_id_num
                self.next_guide_id_num += 1
                self.settings["next_guide_id_num"] = self.next_guide_id_num # Persist the incremented counter

                if self.active_ruler_guide_coord_x is not None:
                    guide_id = f"guide_v_{current_guide_id_num}"
                    self.temporary_guides.append({'id': guide_id, 'type': 'v', 'world_coord': self.active_ruler_guide_coord_x, 'canvas_id': None})
                    self.update_status(f"Placed vertical guide ({guide_id}) at x={self.active_ruler_guide_coord_x:.0f}. Guides are temporary.")
                elif self.active_ruler_guide_coord_y is not None:
                    guide_id = f"guide_h_{current_guide_id_num}"
                    self.temporary_guides.append({'id': guide_id, 'type': 'h', 'world_coord': self.active_ruler_guide_coord_y, 'canvas_id': None})
                    self.update_status(f"Placed horizontal guide ({guide_id}) at y={self.active_ruler_guide_coord_y:.0f}. Guides are temporary.")
                self.draw_all_items() # Redraw to show the new guide
            else: # Clicked on a ruler again, cancel placement
                 self.update_status("Guide placement cancelled.")
            self.active_ruler_guide_coord_x = None
            self.active_ruler_guide_coord_y = None
            return # Consume click

        # Check for guide dragging # This is unnecessary - I replaced this thing, so it can't be called
        """HIT_TOLERANCE = 5 # Pixels
        for guide_info in reversed(self.temporary_guides): # Check topmost first
            if guide_info.get('canvas_id') is None:
                continue

            coords = self.canvas.coords(guide_info['canvas_id'])
            if not coords: continue

            guide_type = guide_info['type']
            world_coord = guide_info['world_coord']

            is_hit = False
            if guide_type == 'h':
                # Horizontal guide: coords are [x1, y1, x2, y1]
                guide_screen_y = coords[1]
                if abs(world_event_y - guide_screen_y) < HIT_TOLERANCE and \
                   coords[0] <= world_event_x <= coords[2]:
                    is_hit = True
                    self.canvas.config(cursor="sb_v_double_arrow")
            elif guide_type == 'v':
                # Vertical guide: coords are [x1, y1, x1, y2]
                guide_screen_x = coords[0]
                if abs(world_event_x - guide_screen_x) < HIT_TOLERANCE and \
                   coords[1] <= world_event_x <= coords[3]:
                    is_hit = True
                    self.canvas.config(cursor="sb_h_double_arrow")

            if is_hit:
                self.drag_data = {
                    'is_dragging_guide': True,
                    'dragged_guide_id': guide_info['id'],
                    'dragged_guide_type': guide_type,
                    'original_world_coord': world_coord,
                    'start_click_canvas_x': world_event_x,
                    'start_click_canvas_y': world_event_y,
                    'item_id': None # Ensure other drag logic doesn't interfere
                }
                self._drag_started_on_item = True # Use this flag to indicate an active drag
                self.update_status(f"Dragging guide {guide_info['id']}")
                self.password_manager.record_activity()
                return # Consume event

        """        # End Unnecessary


        world_event_x, world_event_y = self.canvas_to_world_coords(event.x, event.y)
        self.drag_data = {"x": world_event_x, "y": world_event_y, "item_id": None, "item_type": None,
                          "start_x_world": world_event_x, "start_y_world": world_event_y, # Store start in world
                          "original_positions": {}, "is_resizing": False, "resize_handle_corner": None,
                          "original_size_world": {}}
        self._potential_click_target = None; self._drag_started_on_item = False

        if self.edit_mode_var.get():
            # Use screen coordinates for find_overlapping with canvas items
            item_ids_under_cursor_screen = self.canvas.find_overlapping(world_event_x - RESIZE_HANDLE_SIZE/2, world_event_y - RESIZE_HANDLE_SIZE/2,
                                                                        world_event_x + RESIZE_HANDLE_SIZE/2, world_event_y + RESIZE_HANDLE_SIZE/2)
            for item_c_id in reversed(item_ids_under_cursor_screen):
                tags = self.canvas.gettags(item_c_id)
                if any("resize_handle" in tag for tag in tags):
                    item_id, item_type, handle_corner = None, None, None
                    for tag in tags:
                        if tag.startswith("student_") and tag in self.students: item_id, item_type = tag, "student"
                        elif tag.startswith("furniture_") and tag in self.furniture: item_id, item_type = tag, "furniture"
                        if "br_handle" in tag: handle_corner = "br"; break
                    if item_id and item_id in self.selected_items:
                        self.drag_data.update({"item_id": item_id, "item_type": item_type, "is_resizing": True, "resize_handle_corner": handle_corner})
                        data_src = self.students if item_type == "student" else self.furniture
                        # Store original world width/height
                        orig_w = data_src[item_id].get("width", DEFAULT_STUDENT_BOX_WIDTH)
                        orig_h = data_src[item_id].get("height", DEFAULT_STUDENT_BOX_HEIGHT)
                        if item_type == "student":
                           orig_w = data_src[item_id].get("style_overrides",{}).get("width", orig_w)
                           orig_h = data_src[item_id].get("style_overrides",{}).get("height", orig_h)
                        self.drag_data["original_size_world"] = {"width": orig_w, "height": orig_h}
                        self._drag_started_on_item = True; self.update_status(f"Resizing {item_type} '{item_id}'..."); self.password_manager.record_activity(); return

        item_canvas_ids_click = self.canvas.find_overlapping(world_event_x -1, world_event_y -1, world_event_x +1, world_event_y +1) # Screen coords
        clicked_item_id, clicked_item_type = None, None
        for item_c_id in reversed(item_canvas_ids_click):
            tags = self.canvas.gettags(item_c_id); temp_id, temp_type, is_main_rect = None, None, False
            for tag in tags:
                if tag.startswith("student_") and tag in self.students: temp_id, temp_type = tag, "student"
                elif tag.startswith("furniture_") and tag in self.furniture: temp_id, temp_type = tag, "furniture"
                if "rect" in tag: is_main_rect = True
            if temp_id and is_main_rect: clicked_item_id, clicked_item_type = temp_id, temp_type; break
        clicked_on_selected_item = clicked_item_id and clicked_item_id in self.selected_items
        if not (event.state & 0x0004): # Ctrl NOT pressed
            if clicked_item_id and not clicked_on_selected_item:
                self.deselect_all_items(); self.selected_items.add(clicked_item_id)
                if clicked_item_type == "student": self.draw_single_student(clicked_item_id)
                else: self.draw_single_furniture(clicked_item_id)
            elif not clicked_item_id: self.deselect_all_items()
        if clicked_item_id:
            self._potential_click_target = clicked_item_id; self.drag_data["item_id"] = clicked_item_id; self.drag_data["item_type"] = clicked_item_type; self._drag_started_on_item = True
            self.drag_data["original_positions"] = {}
            for sel_id in self.selected_items:
                sel_item_type_drag = "student" if sel_id in self.students else "furniture"
                data_src_drag = self.students if sel_id in self.students else self.furniture
                if sel_id in data_src_drag: self.drag_data["original_positions"][sel_id] = {"x": data_src_drag[sel_id]["x"], "y": data_src_drag[sel_id]["y"], "type": sel_item_type_drag}
        self.password_manager.record_activity()

    def on_canvas_drag(self, event):
        if self.password_manager.is_locked: return

        # Prevent item drag if setting is off # This is obsolete, as i changed it so that it didn't log when you drag.
        #if not self.settings.get("allow_box_dragging", True) and self.drag_data.get("item_id") and not self.drag_data.get('is_dragging_guide'):
        #    # If dragging items is disabled, but a guide drag might have been initiated, allow guide drag.
        #    return # Do not process item drag

        if self.drag_data.get('is_dragging_guide'):
            dragged_guide_id = self.drag_data.get('dragged_guide_id')
            # guide_info = next((g for g in self.guides if g['id'] == dragged_guide_id), None)
            guide_info = self.guides.get(dragged_guide_id)
            if not guide_info or guide_info.get('canvas_item_id') is None:
                return # Guide not found or not drawn

            #current_event_world_x, current_event_world_y = self.canvas_to_world_coords_guides(event.x, event.y)

            original_guide_world_coord = self.drag_data['original_world_coord']
            #start_drag_world_x = self.drag_data['start_drag_world_x']
            #start_drag_world_y = self.drag_data['start_drag_world_y']

            new_world_coord = original_guide_world_coord

            # Get the change in mouse position in screen coordinates
            dx_screen = event.x - self.drag_data["start_click_canvas_x"]
            dy_screen = event.y - self.drag_data["start_click_canvas_y"]

            # Convert the screen delta to a world delta by dividing by the zoom level
            dx_world = dx_screen / self.current_zoom_level
            dy_world = dy_screen / self.current_zoom_level

            if guide_info['type'] == 'h':  # Horizontal guide
                new_world_coord = self.drag_data['original_world_coord'] + dy_world
                guide_info['world_coord'] = new_world_coord
                _, screen_y = self.world_to_canvas_coords(0, new_world_coord)
                self.canvas.coords(guide_info['canvas_item_id'], 0, screen_y, self.canvas.winfo_width(), screen_y)
            elif guide_info['type'] == 'v':  # Vertical guide
                new_world_coord = self.drag_data['original_world_coord'] + dx_world
                guide_info['world_coord'] = new_world_coord
                screen_x, _ = self.world_to_canvas_coords(new_world_coord, 0)
                self.canvas.coords(guide_info['canvas_item_id'], screen_x, 0, screen_x, self.canvas.winfo_height())

            self.password_manager.record_activity()
            return # Event handled, do not pass to student/furniture drag

        if not self.drag_data.get("item_id") or not self._drag_started_on_item: return

        world_event_x, world_event_y = self.canvas_to_world_coords(event.x, event.y)

        if self.drag_data.get("is_resizing"):
            # Check if dragging (which includes resizing) is allowed
            if not self.settings.get("allow_box_dragging", True):
                # If resizing is part of dragging and dragging is off, clear drag_data and return
                self.drag_data.clear() # Clear drag data to prevent further processing in on_canvas_release
                self._drag_started_on_item = False
                self.update_status("Resizing disabled.")
                self.draw_all_items(check_collisions_on_redraw=True) # Redraw to remove any visual cues of resize start
                return

            item_id, item_type = self.drag_data["item_id"], self.drag_data["item_type"]
            data_src = self.students if item_type == "student" else self.furniture
            item_data = data_src[item_id]

            # Calculate total delta from the drag start in world coordinates
            dx_total_world = world_event_x - self.drag_data["start_x_world"]
            dy_total_world = world_event_y - self.drag_data["start_y_world"]

            orig_world_w = self.drag_data["original_size_world"]["width"]
            orig_world_h = self.drag_data["original_size_world"]["height"]

            new_world_w = orig_world_w + dx_total_world
            new_world_h = orig_world_h + dy_total_world

            min_w = MIN_STUDENT_BOX_WIDTH if item_type == "student" else 20
            min_h = MIN_STUDENT_BOX_HEIGHT if item_type == "student" else 20
            new_world_w = max(min_w, new_world_w); new_world_h = max(min_h, new_world_h)

            if item_type == "student":
                if "style_overrides" not in item_data: item_data["style_overrides"] = {}
                item_data["style_overrides"]["width"] = new_world_w; item_data["style_overrides"]["height"] = new_world_h
                item_data["width"] = new_world_w; item_data["height"] = new_world_h # Sync base
            else: item_data["width"] = new_world_w; item_data["height"] = new_world_h
            if item_type == "student": self.draw_single_student(item_id)
            else: self.draw_single_furniture(item_id)
        else: # Moving
            if not self.drag_data.get("_actual_drag_initiated"):
                # Use world coordinates for drag threshold calculation
                total_dx_world_from_start = abs(world_event_x - self.drag_data["start_x_world"])
                total_dy_world_from_start = abs(world_event_y - self.drag_data["start_y_world"])
                # Convert threshold to world units (approximately, as zoom might affect perception)
                threshold_world = DRAG_THRESHOLD / self.current_zoom_level
                if total_dx_world_from_start > threshold_world or total_dy_world_from_start > threshold_world:
                    self.drag_data["_actual_drag_initiated"] = True; self._potential_click_target = None
                else: return

            # For moving, dx_world and dy_world are deltas from the *last* event's world position
            dx_world_move = world_event_x - self.drag_data["x"]
            dy_world_move = world_event_y - self.drag_data["y"]
            dx_canvas_move = dx_world_move * self.current_zoom_level
            dy_canvas_move = dy_world_move * self.current_zoom_level
            if self.settings.get("allow_box_dragging", True):
                for selected_id in self.selected_items: self.canvas.move(selected_id, dx_canvas_move, dy_canvas_move)

        self.drag_data["x"] = world_event_x # Update last world position for next delta
        self.drag_data["y"] = world_event_y
        self.password_manager.record_activity()

    def on_canvas_release(self, event):
        # ... (largely same as v51, but uses start_x_world/start_y_world for move calculations)
        if self.password_manager.is_locked: return

        if self.drag_data.get('is_dragging_guide'):
            dragged_guide_id = self.drag_data.get('dragged_guide_id')
            guide_info = self.guides.get(dragged_guide_id)
            
            if guide_info:
                # Final update of world_coord happened during on_canvas_drag
                # Optionally, snap to grid here if desired for guides
                # e.g., if self.settings.get("grid_snap_enabled", False):
                #   grid_size = self.settings.get("grid_size", DEFAULT_GRID_SIZE)
                #   guide_info['world_coord'] = round(guide_info['world_coord'] / grid_size) * grid_size
                #   # Redraw this specific guide to snap position if coords changed
                #   self.canvas.delete(guide_info['id']) # Delete by specific guide_id tag
                #   self._draw_single_guide(guide_info) # A hypothetical method to redraw one guide
                world_even_x, world_even_y = self.canvas_to_world_coords(event.x, event.y)
                item_moves_for_command = []
                original_pos_info = self.drag_data.get("original_world_coord", "")#.get(dragged_guide_id)
                # Use total delta from drag_start_world for accurate final position
                if guide_info["type"] == "v":
                    total_dx_world = world_even_x - self.drag_data["start_drag_world_x"]
                    new_world_x = original_pos_info + total_dx_world
                    # current_item_type = "vguide"#original_pos_info["type"]
                    
                    item_id_moved = dragged_guide_id
                    final_x_world= new_world_x
                    item_moves_for_command.append({'id': item_id_moved, 'old_coord': original_pos_info, 'new_coord': final_x_world})
                    
                else:
                    total_dy_world = world_even_y - self.drag_data["start_drag_world_y"]
                    new_world_y = original_pos_info + total_dy_world
                    # current_item_type = "hguide"#original_pos_info["type"]
                    
                    item_id_moved = dragged_guide_id
                    final_y_world = new_world_y
                    item_moves_for_command.append({'id': item_id_moved, 'old_coord': original_pos_info, 'new_coord': final_y_world})
                    
                self.execute_command(MoveGuideCommand(self, item_moves_for_command))
                
                self.update_status(f"Guide {guide_info['id']} position finalized at {guide_info['world_coord']:.0f}.")
                # If guides are persistent, consider saving here or flagging for save
                # self.save_data_wrapper(source="guide_drag_end")
            else:
                self.update_status("Guide drag finished.")

            self.canvas.config(cursor="") # Reset cursor to default
            self.drag_data.clear()
            self._drag_started_on_item = False # Reset general drag flag
            self.password_manager.record_activity()
            self.draw_all_items() # Redraw to ensure canvas is clean and guide is in final state
            return # Event handled

        clicked_item_id_at_press = self._potential_click_target
        dragged_item_id = self.drag_data.get("item_id")
        was_resizing = self.drag_data.get("is_resizing", False)
        actual_drag_initiated = self.drag_data.get("_actual_drag_initiated", False)
        world_event_x, world_event_y = self.canvas_to_world_coords(event.x, event.y)

        if was_resizing and dragged_item_id:
            item_type = self.drag_data["item_type"]
            data_src = self.students if item_type == "student" else self.furniture
            item_data = data_src[dragged_item_id]
            final_w = item_data.get("width"); final_h = item_data.get("height") # These were updated during drag
            if item_type == "student":
                final_w = item_data.get("style_overrides", {}).get("width", final_w)
                final_h = item_data.get("style_overrides", {}).get("height", final_h)
            old_w = self.drag_data["original_size_world"]["width"]; old_h = self.drag_data["original_size_world"]["height"]
            if final_w != old_w or final_h != old_h:
                size_change_info = [{'id': dragged_item_id, 'type': item_type, 'old_w': old_w, 'old_h': old_h, 'new_w': final_w, 'new_h': final_h}]
                self.execute_command(ChangeItemsSizeCommand(self, size_change_info))
            else: self.draw_all_items(check_collisions_on_redraw=True)
            self.update_status(f"Resized {item_type} '{dragged_item_id}'.")
        elif clicked_item_id_at_press and not actual_drag_initiated:
            item_type_of_clicked = "student" if clicked_item_id_at_press in self.students else "furniture"
            if self.edit_mode_var.get(): pass # No action on click in edit mode unless on handle
            elif item_type_of_clicked == "student":
                current_mode = self.mode_var.get()
                if current_mode == "quiz" and self.is_live_quiz_active: self.handle_live_quiz_tap(clicked_item_id_at_press)
                elif current_mode == "homework" and self.is_live_homework_active: self.handle_live_homework_tap(clicked_item_id_at_press) # New
                elif current_mode == "quiz": self.log_quiz_score_dialog(clicked_item_id_at_press)
                elif current_mode == "homework": self.log_homework_dialog(clicked_item_id_at_press) # New
                elif current_mode == "behavior": self.log_behavior_dialog(clicked_item_id_at_press)
        elif actual_drag_initiated and dragged_item_id and not was_resizing:
            grid_size_world = self.settings.get("grid_size", DEFAULT_GRID_SIZE); snap_to_grid = self.settings.get("grid_snap_enabled", False)
            items_moves_for_command = []
            if self.settings.get("allow_box_dragging", True):
                for item_id_moved in self.selected_items:
                    original_pos_info = self.drag_data.get("original_positions", {}).get(item_id_moved)
                    if not original_pos_info: continue
                    current_item_type = original_pos_info["type"]
                    # Use total delta from drag_start_world for accurate final position
                    total_dx_world = world_event_x - self.drag_data["start_x_world"]
                    total_dy_world = world_event_y - self.drag_data["start_y_world"]
                    new_world_x = original_pos_info["x"] + total_dx_world
                    new_world_y = original_pos_info["y"] + total_dy_world
                    final_x_world, final_y_world = new_world_x, new_world_y
                    if snap_to_grid and grid_size_world > 0:
                        final_x_world = round(new_world_x / grid_size_world) * grid_size_world
                        final_y_world = round(new_world_y / grid_size_world) * grid_size_world
                    if abs(final_x_world - original_pos_info["x"]) > 0.01 or abs(final_y_world - original_pos_info["y"]) > 0.01:
                        items_moves_for_command.append({'id': item_id_moved, 'type': current_item_type, 'old_x': original_pos_info["x"], 'old_y': original_pos_info["y"], 'new_x': final_x_world, 'new_y': final_y_world})
                    else: # No change after snap, ensure it's redrawn to its original spot
                        data_s = self.students if current_item_type == "student" else self.furniture
                        data_s[item_id_moved]['x'] = original_pos_info["x"]; data_s[item_id_moved]['y'] = original_pos_info["y"]
                if items_moves_for_command: self.execute_command(MoveItemsCommand(self, items_moves_for_command))
                else: self.draw_all_items(check_collisions_on_redraw=True)
            
        self.drag_data.clear(); self._potential_click_target = None; self._drag_started_on_item = False; self.password_manager.record_activity()

    def on_canvas_right_press(self, event):
        # ... (same as v51, but ensure context menu for homework log is added)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Open Menu", "Enter password to open context menu:"): return
        self.canvas.focus_set()
        world_event_x, world_event_y = self.canvas_to_world_coords(event.x, event.y)
        item_canvas_ids_context = self.canvas.find_overlapping(world_event_x -1, world_event_y -1, world_event_x +1, world_event_y +1) # Screen coords
        context_item_id, context_item_type = None, None
        for item_c_id in reversed(item_canvas_ids_context):
            tags = self.canvas.gettags(item_c_id); temp_id, temp_type, is_main_rect = None, None, False
            i=0
            for tag in tags:
                if tag.startswith("student_") and tag in self.students: temp_id, temp_type = tag, "student"
                elif tag.startswith("furniture_") and tag in self.furniture: temp_id, temp_type = tag, "furniture"
                elif tag.startswith("guide") and self.is_in_guides(tag): temp_id, temp_type = tag, "guide"; context_item_id = "guide"
                elif tag == ("current"): 
                    temp_id, temp_type = tags[1], "guide"
                    context_item_id = "guide"
                    break
                if "rect" in tag: is_main_rect = True
                i+=1
            if temp_id and is_main_rect: context_item_id, context_item_type = temp_id, temp_type; break
        if context_item_id:
            if context_item_id not in self.selected_items:
                self.deselect_all_items(); self.selected_items.add(context_item_id)
                if context_item_type == "student": self.draw_single_student(context_item_id)
                else: self.draw_single_furniture(context_item_id)
            if context_item_type == "student": self.show_student_context_menu(event, context_item_id)
            elif context_item_type == "furniture": self.show_furniture_context_menu(event, context_item_id)
            elif context_item_id == "guide":
                self.show_guide_context_menu(event, temp_id)
        else: self.show_general_context_menu(event)
        self.password_manager.record_activity()

    def show_general_context_menu(self, event):
        # ... (same as v51)
        context_menu = tk.Menu(self.canvas, tearoff=0)
        context_menu.add_command(label="Add Student...", command=self.add_student_dialog)
        context_menu.add_command(label="Add Furniture...", command=self.add_furniture_dialog)
        context_menu.add_separator()
        if self.selected_items:
            context_menu.add_command(label=f"Change Size of Selected ({len(self.selected_items)})...", command=self.change_size_selected_dialog)
            context_menu.add_command(label=f"Delete Selected ({len(self.selected_items)})", command=self.delete_selected_items_confirm)
            #context_menu.add_command(label=f"Assign selected ({len(self.selected_items)}) to group")
            if self.settings.get("student_groups_enabled", True) and self.student_groups and not ('furniture' in str(self.selected_items)):
                previous_group = ""
                #print(list(self.selected_items))
                for student in self.selected_items:
                    try:
                        student_data = self.students[student]; student_name = student_data["full_name"]
                        if student_data.get("group_id") == previous_group and previous_group != "":
                            pass
                        elif previous_group == "" and student_data.get("group_id"):
                            previous_group = student_data.get("group_id")
                        else:
                            previous_group = None
                            break
                    except KeyError: previous_group = None
                current_group_id = previous_group
                group_menu = tk.Menu(context_menu, tearoff=0) #; current_group_id = student_data.get("group_id")
                group_var_menu = tk.StringVar(value=current_group_id if current_group_id else "NONE_GROUP_SENTINEL")
                group_menu.add_radiobutton(label="No Group", variable=group_var_menu, value="NONE_GROUP_SENTINEL", command=lambda sid=self.selected_items: self.assign_students_to_group_via_menu(sid, None))
                for gid, gdata in sorted(self.student_groups.items(), key=lambda item: item[1]['name']):
                    group_menu.add_radiobutton(label=gdata['name'], variable=group_var_menu, value=gid, command=lambda sid=self.selected_items, new_gid=gid: self.assign_students_to_group_via_menu(sid, new_gid))
                context_menu.add_cascade(label="Assign selected to Group", menu=group_menu)
            context_menu.add_separator()
            
            if len(self.selected_items) > 1:
                align_menu = tk.Menu(context_menu, tearoff=0)
                align_menu.add_command(label="Align Top", command=lambda: self.align_selected_items("top"))
                align_menu.add_command(label="Align Bottom", command=lambda: self.align_selected_items("bottom"))
                align_menu.add_command(label="Align Left", command=lambda: self.align_selected_items("left"))
                align_menu.add_command(label="Align Right", command=lambda: self.align_selected_items("right"))
                align_menu.add_command(label="Align Horizontal Center", command=lambda: self.align_selected_items("center_h"))
                align_menu.add_command(label="Align Vertical Center", command=lambda: self.align_selected_items("center_v"))
                context_menu.add_cascade(label="Align Selected", menu=align_menu); context_menu.add_separator()
            num_students = 0
            for item in self.selected_items:
                if "student" in item:
                    num_students +=1
            if num_students > 0:
                context_menu.add_command(label=f"Log Behavior for {num_students} students", command= lambda:self.mass_log_behavior(num_students))
                
                
        context_menu.add_command(label="Select All Students", command=self.select_all_students)
        context_menu.add_command(label="Select All Furniture", command=self.select_all_furniture)
        context_menu.add_command(label="Select All Items", command=self.select_all_items)
        context_menu.add_command(label="Deselect All", command=self.deselect_all_items)
        if self.guides != {}: context_menu.add_command(label=f"Delete all guides", command=lambda: self.delete_all_guides())
        try: context_menu.tk_popup(event.x_root, event.y_root)
        finally: context_menu.grab_release()

    def show_student_context_menu(self, event, student_id):
        # ... (updated for homework mode)
        if student_id not in self.students: return
        student_data = self.students[student_id]; student_name = student_data["full_name"]
        context_menu = tk.Menu(self.canvas, tearoff=0); current_mode = self.mode_var.get()
        if current_mode == "quiz":
            if self.is_live_quiz_active: context_menu.add_command(label=f"Mark Quiz for {student_name}", command=lambda: self.handle_live_quiz_tap(student_id))
            else: context_menu.add_command(label=f"Log Quiz Score for {student_name}", command=lambda: self.log_quiz_score_dialog(student_id))
            context_menu.add_command(label=f"Log Behavior for {student_name}", command=lambda: self.log_behavior_dialog(student_id))
            context_menu.add_command(label=f"Log Homework for {student_name}", command=lambda: self.log_homework_dialog(student_id)) # New
        elif current_mode == "homework": # New
            if self.is_live_homework_active: context_menu.add_command(label=f"Mark Homework for {student_name}", command=lambda: self.handle_live_homework_tap(student_id))
            else: context_menu.add_command(label=f"Log Homework for {student_name}", command=lambda: self.log_homework_dialog(student_id))
            context_menu.add_command(label=f"Log Behavior for {student_name}", command=lambda: self.log_behavior_dialog(student_id))
            context_menu.add_command(label=f"Log Quiz Score for {student_name}", command=lambda: self.log_quiz_score_dialog(student_id))
        else: # Behavior mode
            context_menu.add_command(label=f"Log Behavior for {student_name}", command=lambda: self.log_behavior_dialog(student_id))
            context_menu.add_command(label=f"Log Homework for {student_name}", command=lambda: self.log_homework_dialog(student_id)) # New
            context_menu.add_command(label=f"Log Quiz Score for {student_name}", command=lambda: self.log_quiz_score_dialog(student_id))

        context_menu.add_command(label=f"Edit {student_name}...", command=lambda: self.edit_student_dialog(student_id))
        context_menu.add_command(label=f"Customize Style for {student_name}...", command=lambda: self.customize_student_style_dialog(student_id))
        context_menu.add_command(label=f"Change Size for {student_name}...", command=lambda: self.change_item_size_dialog(student_id, "student"))
        if self.settings.get("student_groups_enabled", True) and self.student_groups:
            group_menu = tk.Menu(context_menu, tearoff=0); current_group_id = student_data.get("group_id")
            group_var_menu = tk.StringVar(value=current_group_id if current_group_id else "NONE_GROUP_SENTINEL")
            group_menu.add_radiobutton(label="No Group", variable=group_var_menu, value="NONE_GROUP_SENTINEL", command=lambda sid=student_id: self.assign_student_to_group_via_menu(sid, None))
            for gid, gdata in sorted(self.student_groups.items(), key=lambda item: item[1]['name']):
                group_menu.add_radiobutton(label=gdata['name'], variable=group_var_menu, value=gid, command=lambda sid=student_id, new_gid=gid: self.assign_student_to_group_via_menu(sid, new_gid))
            context_menu.add_cascade(label="Assign to Group", menu=group_menu)
        context_menu.add_command(label=f"Delete {student_name}", command=lambda: self.delete_student_confirm(student_id))
        context_menu.add_separator()
        if student_id in self._per_student_last_cleared: context_menu.add_command(label="Show Recent Logs (This Student)", command=lambda: self.show_recent_logs_for_student(student_id))
        else: context_menu.add_command(label="Hide Recent Logs (This Student)", command=lambda: self.clear_recent_logs_for_student(student_id))
        context_menu.add_separator()
        global_toggle_text = "Show Recent Logs (All)" if self._recent_incidents_hidden_globally or self._recent_homeworks_hidden_globally else "Hide Recent Logs (All)"
        context_menu.add_command(label=global_toggle_text, command=self.toggle_global_recent_logs_visibility)
        try: context_menu.tk_popup(event.x_root, event.y_root)
        finally: context_menu.grab_release()

    def show_furniture_context_menu(self, event, furniture_id):
        # ... (same as v51)
        if furniture_id not in self.furniture: return
        item_data = self.furniture[furniture_id]; item_name = item_data["name"]
        context_menu = tk.Menu(self.canvas, tearoff=0)
        context_menu.add_command(label=f"Edit {item_name}...", command=lambda: self.edit_furniture_dialog(furniture_id))
        context_menu.add_command(label=f"Change Size for {item_name}...", command=lambda: self.change_item_size_dialog(furniture_id, "furniture"))
        context_menu.add_command(label=f"Delete {item_name}", command=lambda: self.delete_furniture_confirm(furniture_id))
        try: context_menu.tk_popup(event.x_root, event.y_root)
        finally: context_menu.grab_release()

    def is_in_guides(self, tag):
        for guide in self.guides:
            if self.guides[guide] == tag: return True
        return False
    
    def show_guide_context_menu(self, event, guide_id):
        # ... (updated for homework mode)
        #if guide_id not in self.guides: return
        #guide_data = self.guides[guide_id]
        print(guide_id)
        for guide in self.guides:
            if self.guides[guide] == guide_id: guide_data = self.guides[guide].get("id")
        #print(guide_id.get("id"))
        context_menu = tk.Menu(self.canvas, tearoff=0); current_mode = self.mode_var.get()
        
        context_menu.add_command(label=f"Delete guide", command=lambda: self.delete_guide(guide_id))
        context_menu.add_command(label=f"Delete all guides", command=lambda: self.delete_all_guides())
        try: context_menu.tk_popup(event.x_root, event.y_root)
        finally: context_menu.grab_release()

    def delete_guide(self, guide_id):
        #self.canvas.delete(guide_id)
        
        self.execute_command(DeleteGuideCommand(self, guide_id, self.guides[guide_id]))
        #self.guides.__delitem__(guide_id)
        
        self.password_manager.record_activity()

    def delete_all_guides(self):
        if messagebox.askyesno("Delete all guides", "Are you sure you want to delete ALL guides?"):
            commands_to_execute = []
            for guide in self.guides:
                commands_to_execute.append(DeleteGuideCommand(self, self.guides[guide].get("id"), self.guides[guide]))
            for cmd in commands_to_execute: self.execute_command(cmd)
            #self.canvas.delete("guide")
            #self.guides.clear()

    def _select_items_by_type(self, item_type_key):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Select", "Enter password to select items:"): return
        self.deselect_all_items(); items_to_select = []
        if item_type_key == "students" or item_type_key == "all": items_to_select.extend(self.students.keys())
        if item_type_key == "furniture" or item_type_key == "all": items_to_select.extend(self.furniture.keys())
        for item_id in items_to_select:
            self.selected_items.add(item_id)
            if item_id in self.students: self.draw_single_student(item_id)
            elif item_id in self.furniture: self.draw_single_furniture(item_id)
        self.update_status(f"{len(self.selected_items)} items selected."); self.password_manager.record_activity()

    def select_all_students(self): self._select_items_by_type("students")
    def select_all_furniture(self): self._select_items_by_type("furniture")
    def select_all_items(self): self._select_items_by_type("all")
    
    def deselect_all_items(self):
        # ... (same as v51)
        if self.password_manager.is_locked: pass
        items_to_redraw = list(self.selected_items); self.selected_items.clear()
        for item_id in items_to_redraw:
            if item_id in self.students: self.draw_single_student(item_id)
            elif item_id in self.furniture: self.draw_single_furniture(item_id)
        if items_to_redraw: self.update_status("Selection cleared.")

    def mass_log_behavior(self, num_students_selected):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Log Behavior", "Enter password to log behavior:"): return
        
        dialog = BehaviorDialog(self.root, f"Log Behavior for {num_students_selected} students", self.all_behaviors, self.custom_behaviors)
        
        if dialog.result:
            behavior, comment = dialog.result
            for student_id in self.selected_items:
                if "student" in student_id:
                    student = self.students.get(student_id)
                    if not student: continue
                    log_entry = {"timestamp": datetime.now().isoformat(), "student_id": student_id, "student_first_name": student["first_name"],
                                "student_last_name": student["last_name"], "behavior": behavior, "comment": comment, "type": "behavior", "day": datetime.now().strftime('%A')}
                    self.execute_command(LogEntryCommand(self, log_entry, student_id))
            self.update_status(f"Behavior {behavior} logged for {num_students_selected} students")
            self.draw_all_items(check_collisions_on_redraw=True); self.password_manager.record_activity()

    def change_item_size_dialog(self, item_id, item_type):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Change Size", "Enter password to change item size:"): return
        item_data_source = self.students if item_type == "student" else self.furniture
        if item_id not in item_data_source: return
        item_data = item_data_source[item_id]
        current_w = item_data.get("width", DEFAULT_STUDENT_BOX_WIDTH); current_h = item_data.get("height", DEFAULT_STUDENT_BOX_HEIGHT)
        status = False
        if item_type == "student":
            status = True
            style_overrides = item_data.get("style_overrides", {})
            current_w = style_overrides.get("width", current_w); current_h = style_overrides.get("height", current_h)
        dialog = SizeInputDialog(self.root, f"Set Size for {item_data.get('full_name', item_data.get('name'))}", current_w, current_h, status)
        if dialog.result:
            new_w, new_h = dialog.result
            min_w = MIN_STUDENT_BOX_WIDTH if item_type == "student" else 20; min_h = MIN_STUDENT_BOX_HEIGHT if item_type == "student" else 20
            final_w, final_h = max(min_w, new_w), max(min_h, new_h)
            if final_w != current_w or final_h != current_h:
                item_size_info = [{'id': item_id, 'type': item_type, 'old_w': current_w, 'old_h': current_h, 'new_w': final_w, 'new_h': final_h}]
                self.execute_command(ChangeItemsSizeCommand(self, item_size_info))
            self.password_manager.record_activity()

    def change_size_selected_dialog(self):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Change Size", "Enter password to change size of selected items:"): return
        if not self.selected_items: messagebox.showinfo("No Selection", "No items are selected to resize.", parent=self.root); return
        first_selected_id = next(iter(self.selected_items))
        item_type_first = "student" if first_selected_id in self.students else "furniture"
        item_data_source_first = self.students if item_type_first == "student" else self.furniture
        first_item_data = item_data_source_first[first_selected_id]
        current_w_default = first_item_data.get("width", DEFAULT_STUDENT_BOX_WIDTH); current_h_default = first_item_data.get("height", DEFAULT_STUDENT_BOX_HEIGHT)
        status = False
        if item_type_first == "student":
            status = True
            first_item_overrides = first_item_data.get("style_overrides", {})
            current_w_default = first_item_overrides.get("width", current_w_default); current_h_default = first_item_overrides.get("height", current_h_default)
        dialog = SizeInputDialog(self.root, f"Set Size for Selected ({len(self.selected_items)}) Items", current_w_default, current_h_default, status)
        if dialog.result:
            new_w_dialog, new_h_dialog = dialog.result; items_size_changes_for_command = []
            for item_id in self.selected_items:
                item_type_current = "student" if item_id in self.students else "furniture"
                data_src_current = self.students if item_type_current == "student" else self.furniture
                item_data_current = data_src_current[item_id]
                old_w_eff = item_data_current.get("width", DEFAULT_STUDENT_BOX_WIDTH); old_h_eff = item_data_current.get("height", DEFAULT_STUDENT_BOX_HEIGHT)
                if item_type_current == "student":
                    style_ov = item_data_current.get("style_overrides", {})
                    old_w_eff = style_ov.get("width", old_w_eff); old_h_eff = style_ov.get("height", old_h_eff)
                min_w, min_h = (MIN_STUDENT_BOX_WIDTH, MIN_STUDENT_BOX_HEIGHT) if item_type_current == "student" else (20, 20)
                final_w, final_h = max(min_w, new_w_dialog), max(min_h, new_h_dialog)
                if final_w != old_w_eff or final_h != old_h_eff:
                    items_size_changes_for_command.append({'id': item_id, 'type': item_type_current, 'old_w': old_w_eff, 'old_h': old_h_eff, 'new_w': final_w, 'new_h': final_h})
            if items_size_changes_for_command: self.execute_command(ChangeItemsSizeCommand(self, items_size_changes_for_command))
            self.password_manager.record_activity()

    def edit_student_dialog(self, student_id):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Edit", "Enter password to edit student:"): return
        student = self.students.get(student_id);
        if not student: return
        old_student_data_snapshot = student.copy()
        if "style_overrides" in old_student_data_snapshot: old_student_data_snapshot["style_overrides"] = old_student_data_snapshot["style_overrides"].copy()
        dialog = AddEditStudentDialog(self.root, f"Edit Student: {student['full_name']}", student_data=student, app=self)
        if dialog.result:
            fn, ln, nick, gend, grp_id = dialog.result; new_full_name = f"{fn} \"{nick}\" {ln}" if nick else f"{fn} {ln}"
            changes_for_command = {}
            if fn != student.get("first_name"): changes_for_command["first_name"] = fn
            if ln != student.get("last_name"): changes_for_command["last_name"] = ln
            if nick != student.get("nickname", ""): changes_for_command["nickname"] = nick
            if gend != student.get("gender", "Boy"): changes_for_command["gender"] = gend
            if grp_id != student.get("group_id"): changes_for_command["group_id"] = grp_id
            if new_full_name != student.get("full_name"): changes_for_command["full_name"] = new_full_name
            if changes_for_command: self.execute_command(EditItemCommand(self, student_id, "student", old_student_data_snapshot, changes_for_command))
            self.password_manager.record_activity()

    def edit_furniture_dialog(self, furniture_id):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Edit", "Enter password to edit furniture:"): return
        item = self.furniture.get(furniture_id);
        if not item: return
        old_item_data_snapshot = item.copy()
        dialog = AddFurnitureDialog(self.root, f"Edit Furniture: {item['name']}", furniture_data=item)
        if dialog.result:
            name, item_type, width, height = dialog.result; changes_for_command = {}
            if name != item.get("name"): changes_for_command["name"] = name
            if item_type != item.get("type"): changes_for_command["type"] = item_type
            if width != item.get("width"): changes_for_command["width"] = width
            if height != item.get("height"): changes_for_command["height"] = height
            if changes_for_command: self.execute_command(EditItemCommand(self, furniture_id, "furniture", old_item_data_snapshot, changes_for_command))
            self.password_manager.record_activity()

    def customize_student_style_dialog(self, student_id):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Customize Style", "Enter password to customize style:"): return
        student = self.students.get(student_id);
        if not student: return
        dialog = StudentStyleDialog(self.root, f"Customize Style: {student['full_name']}", student, self)
        if dialog.result:
            for prop, old_val, new_val in dialog.result:
                self.execute_command(ChangeStudentStyleCommand(self, student_id, prop, old_val, new_val))
            self.password_manager.record_activity()

    def delete_student_confirm(self, student_id):
        # ... (updated to include homework_log)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Delete", "Enter password to delete student:"): return
        if student_id not in self.students: return
        student_name = self.students[student_id]["full_name"]
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete {student_name}?\nThis will also remove their behavior, quiz, and homework log entries.", parent=self.root):
            student_data_copy = self.students[student_id].copy()
            if "style_overrides" in student_data_copy: student_data_copy["style_overrides"] = student_data_copy["style_overrides"].copy()
            associated_logs = [log.copy() for log in self.behavior_log if log["student_id"] == student_id]
            # DeleteItemCommand now handles associated_homework_logs internally
            cmd = DeleteItemCommand(self, student_id, "student", student_data_copy, associated_logs)
            self.execute_command(cmd); self.password_manager.record_activity()

    def delete_furniture_confirm(self, furniture_id):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Delete", "Enter password to delete furniture:"): return
        if furniture_id not in self.furniture: return
        item_name = self.furniture[furniture_id]["name"]
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete furniture item '{item_name}'?", parent=self.root):
            item_data_copy = self.furniture[furniture_id].copy()
            self.execute_command(DeleteItemCommand(self, furniture_id, "furniture", item_data_copy))
            self.password_manager.record_activity()

    def delete_selected_items_confirm(self):
        # ... (updated to include homework_log in message)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Delete", "Enter password to delete selected items:"): return
        if not self.selected_items: messagebox.showinfo("No Selection", "No items are selected to delete.", parent=self.root); return
        num_students = sum(1 for sid in self.selected_items if sid in self.students)
        num_furniture = sum(1 for fid in self.selected_items if fid in self.furniture)
        message = f"Are you sure you want to delete {len(self.selected_items)} selected items"
        details = []
        if num_students > 0: details.append(f"{num_students} student(s) (all logs will also be removed)") # Updated message
        if num_furniture > 0: details.append(f"{num_furniture} furniture item(s)")
        if details: message += f" ({', '.join(details)})"; message += "?"
        if messagebox.askyesno("Confirm Delete Selected", message, parent=self.root):
            commands_to_execute = []
            for item_id in list(self.selected_items):
                if item_id in self.students:
                    student_data_copy = self.students[item_id].copy()
                    if "style_overrides" in student_data_copy: student_data_copy["style_overrides"] = student_data_copy["style_overrides"].copy()
                    associated_logs = [log.copy() for log in self.behavior_log if log["student_id"] == item_id]
                    # DeleteItemCommand now handles associated_homework_logs internally
                    commands_to_execute.append(DeleteItemCommand(self, item_id, "student", student_data_copy, associated_logs))
                elif item_id in self.furniture:
                    item_data_copy = self.furniture[item_id].copy()
                    commands_to_execute.append(DeleteItemCommand(self, item_id, "furniture", item_data_copy))
            for cmd in commands_to_execute: self.execute_command(cmd)
            self.password_manager.record_activity()

    def log_behavior_dialog(self, student_id):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Log Behavior", "Enter password to log behavior:"): return
        student = self.students.get(student_id);
        if not student: return
        dialog = BehaviorDialog(self.root, f"Log Behavior for {student['full_name']}", self.all_behaviors, self.custom_behaviors)
        if dialog.result:
            behavior, comment = dialog.result
            log_entry = {"timestamp": datetime.now().isoformat(), "student_id": student_id, "student_first_name": student["first_name"],
                         "student_last_name": student["last_name"], "behavior": behavior, "comment": comment, "type": "behavior", "day": datetime.now().strftime('%A')}
            self.execute_command(LogEntryCommand(self, log_entry, student_id))
            self.draw_all_items(check_collisions_on_redraw=True); self.password_manager.record_activity()

    def log_homework_dialog(self, student_id):
        """
        Handles manual homework logging. Implements the new Simplified/Detailed view logic.
        - Simplified View (if marks are disabled): Two popups to select type and status.
        - Detailed View (if marks are enabled): Opens the full dialog to log marks.
        """
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Log Homework", "Enter password to log homework:"): return
        
        student = self.students.get(student_id)
        if not student: return

        initial_homework_name_to_pass = self.settings.get("default_homework_name", "Homework")
        initial_num_items_to_pass = self.settings.get("_last_used_hw_items_for_session", self.settings.get("default_homework_items_for_yes_no_mode", 5)) # Default to 5 if not set
        timeout_hw_minutes = self.settings.get("last_used_homework_name_timeout_minutes", 60)

        if self.last_used_homework_name and self.last_used_homework_name_timestamp:
            try:
                time_since_last_hw_use = (datetime.now() - datetime.fromisoformat(self.last_used_homework_name_timestamp)).total_seconds() / 60
                if time_since_last_hw_use < timeout_hw_minutes:
                    initial_homework_name_to_pass = self.last_used_homework_name
                    initial_num_items_to_pass = self.initial_num_homework_items # Use the stored number of items
            except ValueError:
                print("Warning: Could not parse last_used_homework_name_timestamp.")


        # --- BRANCHING LOGIC FOR SIMPLIFIED/DETAILED VIEW ---
        if not self.settings.get("log_homework_marks_enabled", True):
            # --- Simplified View ---
            # 1. First Popup: Select Homework Type (uses BehaviorDialog)
            # Pass the potentially remembered name to the dialog
            type_dialog = BehaviorDialog(self.root, f"Select Homework Type for {student['full_name']}",
                                         self.all_homework_types, [],
                                         initial_value=initial_homework_name_to_pass) # Pass initial value
            if not type_dialog.result:
                self.update_status("Homework log cancelled.")
                return
            homework_type, comment_type = type_dialog.result

            # 2. Second Popup: Select Homework Status
            status_dialog = BehaviorDialog(self.root, f"Select Status for '{homework_type}'", self.all_homework_statuses, [])
            if not status_dialog.result:
                self.update_status("Homework log cancelled.")
                return
            homework_status, comment_status = status_dialog.result
            
            final_comment = comment_type
            if comment_status:
                final_comment += (f" - {comment_status}" if final_comment else comment_status)

            log_entry = {
                "timestamp": datetime.now().isoformat(), "student_id": student_id,
                "student_first_name": student["first_name"], "student_last_name": student["last_name"],
                "behavior": f"{homework_type}: {homework_status}",
                "homework_type": homework_type,
                "homework_status": homework_status,
                "comment": final_comment, "type": "homework", "day": datetime.now().strftime('%A')
            }
            self.execute_command(LogHomeworkEntryCommand(self, log_entry, student_id))

            # Remember this homework type for next time
            self.last_used_homework_name = homework_type
            self.last_used_homework_name_timestamp = datetime.now().isoformat()
            # For simplified view, num_items isn't directly relevant for "remembering" but store consistently
            self.initial_num_homework_items = 1 # Or some other default for simplified
            self.settings["_last_used_homework_name_for_session"] = self.last_used_homework_name
            self.settings["_last_used_homework_name_timestamp_for_session"] = self.last_used_homework_name_timestamp
            self.settings["_last_used_hw_items_for_session"] = self.initial_num_homework_items

            self.draw_all_items(check_collisions_on_redraw=True)
            self.password_manager.record_activity()

        else:
            # --- Detailed Marks View ---
            dialog = ManualHomeworkLogDialog(
                self.root, f"Log Homework for {student['full_name']}",
                self.all_homework_types,
                self.custom_homework_types,
                log_marks_enabled=True,
                homework_mark_types=self.settings.get("homework_mark_types", DEFAULT_HOMEWORK_MARK_TYPES.copy()),
                homework_templates=self.homework_templates,
                app=self,
                initial_homework_name=initial_homework_name_to_pass, # Pass remembered name
                initial_num_items=initial_num_items_to_pass # Pass remembered items
            )
            if dialog.result:
                homework_type, comment, marks_data, num_items = dialog.result
                log_entry = {
                    "timestamp": datetime.now().isoformat(), "student_id": student_id,
                    "student_first_name": student["first_name"], "student_last_name": student["last_name"],
                    "behavior": homework_type, "homework_type": homework_type,
                    "comment": comment, "type": "homework", "day": datetime.now().strftime('%A')
                }
                if marks_data:
                    log_entry["marks_data"] = marks_data
                    log_entry["num_items"] = num_items

                self.execute_command(LogHomeworkEntryCommand(self, log_entry, student_id))

                # Remember this homework type and num_items for next time
                self.last_used_homework_name = homework_type
                self.last_used_homework_name_timestamp = datetime.now().isoformat()
                self.initial_num_homework_items = num_items if num_items is not None else self.settings.get("default_homework_items_for_yes_no_mode", 5)

                self.settings["_last_used_homework_name_for_session"] = self.last_used_homework_name
                self.settings["_last_used_homework_name_timestamp_for_session"] = self.last_used_homework_name_timestamp
                self.settings["_last_used_hw_items_for_session"] = self.initial_num_homework_items

                self.draw_all_items(check_collisions_on_redraw=True)
                self.password_manager.record_activity()

    def log_quiz_score_dialog(self, student_id):
        # ... (same as v51)
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Log Quiz Score", "Enter password to log quiz score:"): return
        student = self.students.get(student_id)
        if not student: return
        initial_quiz_name = self.settings.get("default_quiz_name", "Pop Quiz")
        initial_num_questions = self.settings.get("_last_used_q_num_for_session", self.settings.get("default_quiz_questions", 10))
        timeout_minutes = self.settings.get("last_used_quiz_name_timeout_minutes", 60)
        if self.last_used_quiz_name and self.last_used_quiz_name_timestamp:
            try:
                if (datetime.now() - datetime.fromisoformat(self.last_used_quiz_name_timestamp)).total_seconds() / 60 < timeout_minutes:
                    initial_quiz_name = self.last_used_quiz_name; initial_num_questions = self.initial_num_questions
            except ValueError: print("Warning: Could not parse last_used_quiz_name_timestamp.")
        dialog = QuizScoreDialog(self.root, f"Log Quiz Score for {student['full_name']}", initial_quiz_name,
                                 self.settings.get("quiz_mark_types", DEFAULT_QUIZ_MARK_TYPES.copy()), self.quiz_templates,
                                 self.settings.get("default_quiz_questions", 10), initial_num_questions)
        if dialog.result:
            quiz_name, marks_data, comment, num_questions_actual = dialog.result
            log_entry = {"timestamp": datetime.now().isoformat(), "student_id": student_id, "student_first_name": student["first_name"],
                         "student_last_name": student["last_name"], "behavior": quiz_name, "comment": comment, "marks_data": marks_data,
                         "num_questions": num_questions_actual, "type": "quiz", "day": datetime.now().strftime('%A')}
            self.execute_command(LogEntryCommand(self, log_entry, student_id))
            self.last_used_quiz_name = quiz_name; self.last_used_quiz_name_timestamp = datetime.now().isoformat(); self.initial_num_questions = num_questions_actual
            self.settings["_last_used_quiz_name_for_session"] = self.last_used_quiz_name
            self.settings["_last_used_quiz_name_timestamp_for_session"] = self.last_used_quiz_name_timestamp
            self.settings["_last_used_q_num_for_session"] = self.initial_num_questions
            self.password_manager.record_activity()
            self.draw_all_items()
    
    def save_data_wrapper(self, event=None, source="manual"):
        self._ensure_next_ids()
        serializable_undo_stack = [cmd.to_dict() for cmd in self.undo_stack]
        serializable_redo_stack = [cmd.to_dict() for cmd in self.redo_stack]

        data_to_save = {
            "students": self.students,
            "furniture": self.furniture,
            "behavior_log": self.behavior_log,
            "homework_log": self.homework_log,
            "settings": self.settings,
            "last_excel_export_path": self.last_excel_export_path,
            "_per_student_last_cleared": self._per_student_last_cleared,
            "undo_stack": serializable_undo_stack,
            "redo_stack": serializable_redo_stack,
            "guides": {}, 
            "next_guide_id_num": self.next_guide_id_num
        }

        guides_to_save = {}
        for guide_info in self.guides:
            guides_to_save[guide_info] = { # guides_to_save.append({
                'id': self.guides[guide_info].get('id'), #guide_info.get('id'),
                'type': self.guides[guide_info].get('type'), #guide_info.get('type'),
                'world_coord': self.guides[guide_info].get('world_coord'), #guide_info.get('world_coord')
            }
        data_to_save["guides"] = guides_to_save

        try:
            # Encrypt the data
            json_data_string = json.dumps(data_to_save, indent=4)
            if self.settings.get("encrypt_data_files", True):
                data = encrypt_data(json_data_string)
            else:
                data = json_data_string.encode('utf-8')
            with open(DATA_FILE, 'wb') as f: # Open in binary write mode
                f.write(data)

            verbose_save = source not in ["autosave", "command_execution", "undo_command", "redo_command", "toggle_mode", "end_live_quiz", "end_live_homework_session", "reset", "assign_group_menu", "load_template", "save_and_quit"]
            if verbose_save:
                self.update_status(f"Data saved to {os.path.basename(DATA_FILE)}")
            elif source == "autosave":
                self.update_status(f"Autosaved data at {datetime.now().strftime('%H:%M:%S')}")

        except IOError as e:
            self.update_status(f"Error saving data: {e}")
            messagebox.showerror("Save Error", f"Could not save data to {DATA_FILE}: {e}", parent=self.root)
        except Exception as e:
            print(e)
            
        # Call all individual config savers
        self.save_student_groups()
        self.save_custom_behaviors()
        self.save_custom_homework_types()
        self.save_custom_homework_statuses()
        self.save_quiz_templates()
        self.save_homework_templates()

    def _update_toggle_dragging_button_text(self):
        if hasattr(self, 'toggle_dragging_btn'):
            if self.settings.get("allow_box_dragging", True):
                self.toggle_dragging_btn.config(text="Disable Dragging")
            else:
                self.toggle_dragging_btn.config(text="Enable Dragging")

    def toggle_dragging_allowed(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Toggle Dragging", "Enter password to toggle dragging:"): return

        current_setting = self.settings.get("allow_box_dragging", True)
        self.settings["allow_box_dragging"] = not current_setting
        self._update_toggle_dragging_button_text()
        self.update_status(f"Box dragging {'enabled' if self.settings['allow_box_dragging'] else 'disabled'}.")
        self.save_data_wrapper(source="toggle_dragging_button") # Save settings immediately
        # No redraw needed unless there's a visual cue for draggable state on items themselves
        self.password_manager.record_activity()
    
    
    def import_data(self):
        if not self.prompt_for_password("Unlock to import data", "Enter password to import data:", True):
            return
        
        if messagebox.askokcancel("Import data", "Import data from JSON? This will reset application data!", icon='warning'):
            try:
                self.backup_all_data_dialog(force=True)
                messagebox.showinfo("Backup Created", "A backup of your current data has been created in the application's data folder.", parent=self.root)
            except Exception as e:
                messagebox.showerror("Backup Error", f"Could not create a backup. Import aborted. Error: {e}", parent=self.root)
                return

            file_path = filedialog.askopenfilename(title="Import JSON", filetypes=[("JSON", "*.json"), ("All Files", "*.*")])
            if file_path:
                try:
                    self.load_data(file_path=file_path)
                except Exception as e:
                    messagebox.showerror("Error loading data", f"Error loading data from json {e}:", icon="error")
                finally:
                    self.draw_all_items()
                    self.reload_canvas()
            
    
    def load_data(self, file_path=None, is_restore=False):
        # ... (updated migration chain)
        target_file = file_path or DATA_FILE
        default_settings_copy = self._get_default_settings()
        data_loaded_successfully = False

        if os.path.exists(target_file):
            try:
                with open(target_file, 'rb') as f: # Open in binary read mode
                    encrypted_data = f.read()
                try:
                    decrypted_data_string = decrypt_data(encrypted_data)
                except cryptography.fernet.InvalidToken:
                    decrypted_data_string = encrypted_data
                data = json.loads(decrypted_data_string)
                """try:
                    with open(target_file, 'r', encoding='utf-8') as f: data = json.load(f)"""
                data_version_from_filename = None
                file_basename = os.path.basename(target_file)
                if "_v3" in file_basename or "_v4" in file_basename or file_basename == f"classroom_data.json": data_version_from_filename = 3
                elif "_v5" in file_basename: data_version_from_filename = 5
                elif "_v6" in file_basename: data_version_from_filename = 6
                elif "_v7" in file_basename: data_version_from_filename = 7
                elif "_v8" in file_basename: data_version_from_filename = 8 # Previous version

                if data_version_from_filename is None or data_version_from_filename <= 3:
                    print(f"Migrating data from v3/v4 format (or older) from {target_file}")
                    data = self._migrate_v3_edited_data(data); data = self._migrate_v4_data(data); data = self._migrate_v5_data(data)
                    data = self._migrate_v6_data(data); data = self._migrate_v7_data(data); data = self._migrate_v8_data(data); data = self._migrate_v9_data(data)
                elif data_version_from_filename == 5:
                    print(f"Migrating data from v5 format from {target_file}")
                    data = self._migrate_v5_data(data); data = self._migrate_v6_data(data); data = self._migrate_v7_data(data); data = self._migrate_v8_data(data); data = self._migrate_v9_data(data)
                elif data_version_from_filename == 6:
                    print(f"Migrating data from v6 format from {target_file}")
                    data = self._migrate_v6_data(data); data = self._migrate_v7_data(data); data = self._migrate_v8_data(data); data = self._migrate_v9_data(data)
                elif data_version_from_filename == 7:
                    print(f"Migrating data from v7 format from {target_file}")
                    data = self._migrate_v7_data(data); data = self._migrate_v8_data(data); data = self._migrate_v9_data(data)
                elif data_version_from_filename == 8:
                    print(f"Migrating data from v8 format from {target_file}")
                    data = self._migrate_v8_data(data); data = self._migrate_v9_data(data)
                elif data_version_from_filename == 9: # New: If loading v9 data
                    print(f"Migrating data from v9 format from {target_file}")
                    data = self._migrate_v9_data(data)


                final_settings = default_settings_copy.copy(); final_settings.update(data.get("settings", {}))
                data["settings"] = final_settings
                self.students = data.get("students", {}); self.furniture = data.get("furniture", {})
                self.behavior_log = data.get("behavior_log", []); self.homework_log = data.get("homework_log", []) # Load homework log
                self.settings = data.get("settings", default_settings_copy)
                self.last_excel_export_path = data.get("last_excel_export_path", None)
                self._per_student_last_cleared = data.get("_per_student_last_cleared", {})
                self.last_used_quiz_name = self.settings.get("_last_used_quiz_name_for_session", "")
                self.last_used_quiz_name_timestamp = self.settings.get("_last_used_quiz_name_timestamp_for_session", None)
                self.initial_num_questions = self.settings.get("_last_used_q_num_for_session", 10)

                self.last_used_homework_name = self.settings.get("_last_used_homework_name_for_session", "")
                self.last_used_homework_name_timestamp = self.settings.get("_last_used_homework_name_timestamp_for_session", None)
                self.initial_num_homework_items = self.settings.get("_last_used_hw_items_for_session", self.settings.get("default_homework_items_for_yes_no_mode", 5))

                self.theme_style_using = self.settings.get("theme", "System") # Newer
                self.custom_canvas_color = self.settings.get("canvas_color", "Default")
                self.type_theme = self.settings.get("type_theme", "sv_ttk")

                # Load guides
                loaded_guides_raw = data.get("guides", []) # Load from "guides" key
                self.guides = {} # Initialize self.guides before populating
                for guide_data_raw in loaded_guides_raw:
                    # Ensure only expected keys are loaded and canvas_item_id is reset (will be set on draw)
                    self.guides[guide_data_raw] = {
                        'id': loaded_guides_raw.get(guide_data_raw)["id"], #.get('id'),
                        'type': loaded_guides_raw.get(guide_data_raw)["type"], #guide_data_raw.get('type'),
                        'world_coord': loaded_guides_raw.get(guide_data_raw)["world_coord"], #guide_data_raw.get('world_coord'),
                        'canvas_item_id': None
                    }

                # Load next_guide_id_num, defaulting to 1 if not present
                self.next_guide_id_num = data.get("next_guide_id_num", 1)


                self.undo_stack.clear(); self.redo_stack.clear()
                loaded_undo_stack = data.get("undo_stack", [])
                loaded_redo_stack = data.get("redo_stack", [])
                cutoff_date_iso = (datetime.now() - timedelta(days=self.settings.get("max_undo_history_days", MAX_UNDO_HISTORY_DAYS))).isoformat()
                for cmd_data in loaded_undo_stack:
                    if cmd_data.get('timestamp', '0') >= cutoff_date_iso:
                        cmd_obj = Command.from_dict(self, cmd_data)
                        if cmd_obj: self.undo_stack.append(cmd_obj)
                for cmd_data in loaded_redo_stack:
                    if cmd_data.get('timestamp', '0') >= cutoff_date_iso:
                        cmd_obj = Command.from_dict(self, cmd_data)
                        if cmd_obj: self.redo_stack.append(cmd_obj)
                self.update_undo_redo_buttons_state()
                self.password_manager = PasswordManager(self.settings) # Re-initialize with loaded settings
                self.update_lock_button_state()
                data_loaded_successfully = True
            except (json.JSONDecodeError, KeyError, IOError, TypeError) as e:
                print(f"Error loading data from {target_file}: {e}. Using defaults or attempting recovery.")
                if is_restore:
                    messagebox.showerror("Restore Error", f"Failed to load restored data from {target_file}: {e}\n\nApplication will use default data or attempt to load the standard data file.", parent=self.root)
                    # Fallback to loading standard DATA_FILE if restore failed and target_file was not DATA_FILE
                    if target_file != DATA_FILE: self.load_data(DATA_FILE, is_restore=False); return
                else:
                    messagebox.showwarning("Load Error", f"Error loading data file: {e}.\nDefault settings and empty classroom will be used.", parent=self.root)
                self.students, self.furniture, self.behavior_log, self.homework_log = {}, {}, [], []
                self.settings = default_settings_copy.copy()
                self.last_excel_export_path, self._per_student_last_cleared = None, {}
                self.undo_stack.clear(); self.redo_stack.clear()
        else:
            if not is_restore: print(f"Data file {target_file} not found. Using default settings and empty classroom.")
            self.students, self.furniture, self.behavior_log, self.homework_log = {}, {}, [], []
            self.settings = default_settings_copy.copy()
            self.last_excel_export_path, self._per_student_last_cleared = None, {}
            self.undo_stack.clear(); self.redo_stack.clear()

        # Ensure essential settings are present if a very old or corrupted file was loaded
        for key, value in default_settings_copy.items():
            if key not in self.settings: self.settings[key] = value

        # Specifically ensure new conditional formatting rule fields have defaults
        if "conditional_formatting_rules" in self.settings:
            for rule in self.settings["conditional_formatting_rules"]:
                rule.setdefault("enabled", True)
                rule.setdefault("active_times", [])
                rule.setdefault("active_modes", [])
        
        # Ensure next ID counters are robustly initialized/updated after data load
        self._ensure_next_ids()
        if data_loaded_successfully and not is_restore and file_path is None and \
           (os.path.basename(DATA_FILE) != f"classroom_data_{CURRENT_DATA_VERSION_TAG}.json" or \
            (data_version_from_filename is not None and data_version_from_filename < int(CURRENT_DATA_VERSION_TAG[1:]))):
            # If the main data file was from an older version, save it immediately in the new version format
            print(f"Data file loaded from an older version ({file_basename}). Saving in new format: classroom_data_{CURRENT_DATA_VERSION_TAG}.json")
            self.save_data_wrapper(source="migration_save")
            # Optionally, attempt to delete the old version file if migration was successful
            # Be cautious with this, maybe offer as a user option later
            if os.path.exists(target_file) and target_file != DATA_FILE:
                try:
                    # os.remove(target_file)
                    # print(f"Old data file {target_file} removed after successful migration.")
                    print(f"Old data file {target_file} can be manually removed if no longer needed.")
                except OSError as e_del:
                    print(f"Could not remove old data file {target_file}: {e_del}")
 
    def _migrate_v8_data(self, data): # New migration for v8 -> v9
        """Migration for data version 8 (APP_VERSION v51) to v9 (APP_VERSION v52)."""
        # Add new homework-related settings if missing
        if "settings" in data:
            data["settings"].setdefault("homework_log", []) # Initialize if not present
            data["settings"].setdefault("show_recent_homeworks_on_boxes", True)
            data["settings"].setdefault("num_recent_homeworks_to_show", 2)
            data["settings"].setdefault("recent_homework_time_window_hours", 24)
            data["settings"].setdefault("show_full_recent_homeworks", False)
            data["settings"].setdefault("reverse_homework_order", True)
            data["settings"].setdefault("selected_recent_homeworks_filter", None)
            data["settings"].setdefault("homework_initial_overrides", {})
            data["settings"].setdefault("default_homework_name", "Homework Check")
            data["settings"].setdefault("live_homework_session_mode", "Yes/No")
            data["settings"].setdefault("log_homework_marks_enabled", True)
            data["settings"].setdefault("homework_mark_types", DEFAULT_HOMEWORK_MARK_TYPES.copy())
            data["settings"].setdefault("default_homework_items_for_yes_no_mode", 5)
            data["settings"].setdefault("live_homework_score_font_color", DEFAULT_HOMEWORK_SCORE_FONT_COLOR)
            data["settings"].setdefault("live_homework_score_font_style_bold", DEFAULT_HOMEWORK_SCORE_FONT_STYLE_BOLD)
            data["settings"].setdefault("next_homework_template_id_num", 1)
            data["settings"].setdefault("next_custom_homework_type_id_num", 1)
            data["settings"].setdefault("_last_used_homework_name_for_session", "")
            data["settings"].setdefault("_last_used_homework_name_timestamp_for_session", None)
            data["settings"].setdefault("_last_used_hw_items_for_session", 5)
            data["settings"].setdefault("next_guide_id_num", 1) # For v10 guides

        # Ensure homework_log list exists at the top level of data
        if "homework_log" not in data:
            data["homework_log"] = []

        # Ensure guides list exists for v10, initialize if migrating from older
        if "guides" not in data:
            data["guides"] = []

        # Migrate existing behavior_log entries that might have been intended as homework
        # This is heuristic; might need adjustment based on how users previously logged homework.
        # For now, assume if behavior name contains "Homework" or "HW", it might be a homework log.
        # This is a very basic migration attempt.
        temp_behavior_log = []
        temp_homework_log = list(data.get("homework_log", [])) # Start with existing homework log

        for log_entry in data.get("behavior_log", []):
            behavior_name = log_entry.get("behavior", "").lower()
            log_type = log_entry.get("type", "behavior").lower()
            if ("homework" in behavior_name or "hw" in behavior_name) and log_type == "behavior":
                # Convert this to a homework log entry
                new_hw_entry = log_entry.copy()
                new_hw_entry["type"] = "homework" # Change type
                new_hw_entry["homework_type"] = log_entry.get("behavior") # Store original behavior as homework_type
                if "marks_data" not in new_hw_entry and self.settings.get("log_homework_marks_enabled", True):
                    # If marks are enabled but not present, add a placeholder or try to infer
                    new_hw_entry["marks_data"] = {} # Add empty marks_data
                temp_homework_log.append(new_hw_entry)
            else:
                temp_behavior_log.append(log_entry)

        data["behavior_log"] = temp_behavior_log
        data["homework_log"] = temp_homework_log
        print("Applied v8 (to v9) data migration.")
        return data

    def _migrate_v9_data(self, data):
        """Migration for data version 9 (APP_VERSION v54) to v10 (APP_VERSION v56)."""
        # Key change: Addition of persistent guides.
        # These settings are no longer used as guides are part of main data.
        # data["settings"].setdefault("save_guides_to_file", True) # Removed
        # data["settings"].setdefault("guides_stay_when_rulers_hidden", True) # Removed
        # data["settings"].setdefault("next_guide_id_num", 1) # Removed from settings, direct attribute now

        # Ensure 'guides' list and 'next_guide_id_num' exist at the top level of data
        if "guides" not in data:
            data["guides"] = []
        if "next_guide_id_num" not in data:
            data["next_guide_id_num"] = 1

        # If migrating from a version that used "temporary_guides" in settings (like Kivy v55)
        # and by chance that data structure is present, convert it.
        # This is unlikely for seatingchartmain.py's own data evolution but defensive.
        if "settings" in data and "temporary_guides" in data["settings"]:
            if not data["guides"]: # Only if 'guides' isn't already populated
                 data["guides"] = data["settings"].pop("temporary_guides", [])
            else: # guides already exists, just remove the old settings one
                data["settings"].pop("temporary_guides", None)

        if "settings" in data and "next_guide_id_num" in data["settings"]:
            if data.get("next_guide_id_num",1) == 1: # Only if main data one is still default
                data["next_guide_id_num"] = data["settings"].pop("next_guide_id_num",1)
            else: # next_guide_id_num already has a value, remove from settings
                 data["settings"].pop("next_guide_id_num", None)


        print("Applied v9 (to v10) data migration (guides integrated into main data structure).")
        return data

    def _migrate_v7_data(self, data):
        """Migration for data version 7 (APP_VERSION v50) to v8 (APP_VERSION v51)."""
        # For v50 -> v51 (data v7 -> v8)
        if "settings" in data:
            data["settings"].setdefault("student_groups_enabled", True)
            data["settings"].setdefault("show_zoom_level_display", True)
            data["settings"].setdefault("next_group_id_num", 1) # Initialize if missing
            data["settings"].setdefault("next_quiz_template_id_num", 1) # Initialize if missing
            data["settings"].setdefault("available_fonts", sorted(list(tkfont.families())))
            data["settings"].setdefault("default_quiz_questions", 10)
            data["settings"].setdefault("quiz_score_calculation", "percentage")
            data["settings"].setdefault("combine_marks_for_display", True)

            # Password settings
            data["settings"].setdefault("app_password_hash", None)
            data["settings"].setdefault("password_on_open", False)
            data["settings"].setdefault("password_on_edit_action", False)
            data["settings"].setdefault("password_auto_lock_enabled", False)
            data["settings"].setdefault("password_auto_lock_timeout_minutes", 15)
             # Quiz session internal state storage
            data["settings"].setdefault("_last_used_quiz_name_for_session", "")
            data["settings"].setdefault("_last_used_quiz_name_timestamp_for_session", None)
            data["settings"].setdefault("_last_used_q_num_for_session", 10)

            # Quiz Mark Types (ensure all fields are present)
            migrated_mark_types = []
            default_mark_type_map = {d["id"]: d for d in DEFAULT_QUIZ_MARK_TYPES}
            for mt in data["settings"].get("quiz_mark_types", []):
                # If it's an old string-based mark type, skip or attempt conversion if possible (not done here)
                if isinstance(mt, dict) and "id" in mt and "name" in mt:
                    default_entry = default_mark_type_map.get(mt["id"])
                    if default_entry: # Use defaults for missing fields if id matches a default
                        migrated_mt = default_entry.copy()
                        migrated_mt.update(mt) # Override with user's existing values
                    else: # Custom entry, ensure all fields are present
                        migrated_mt = mt.copy()
                        migrated_mt.setdefault("contributes_to_total", True)
                        migrated_mt.setdefault("is_extra_credit", False)
                        migrated_mt.setdefault("default_points", 1 if mt.get("name","").lower() == "correct" else (0 if mt.get("name","").lower() == "incorrect" else 0.5))
                    migrated_mark_types.append(migrated_mt)
            if not migrated_mark_types and not data["settings"].get("quiz_mark_types"): # If empty or missing
                 data["settings"]["quiz_mark_types"] = DEFAULT_QUIZ_MARK_TYPES.copy()
            elif migrated_mark_types:
                 data["settings"]["quiz_mark_types"] = migrated_mark_types


        if "students" in data:
            for student_id, student_data in data["students"].items():
                student_data.setdefault("group_id", None)
                student_data.setdefault("style_overrides", {})

        # Remove obsolete _undo_history_file from settings if present
        if "settings" in data and "_undo_history_file" in data["settings"]:
            del data["settings"]["_undo_history_file"]

        # Migrate undo/redo stack if it exists (it might be from a very old direct load)
        if "undo_stack" not in data: data["undo_stack"] = []
        if "redo_stack" not in data: data["redo_stack"] = []

        print("Applied v7 (to v8) data migration.")
        return data

    def _migrate_v6_data(self, data):
        """Migration for data version 6 (APP_VERSION v0.6.x) to v7 (APP_VERSION v50)."""
        # For v0.6.x -> v50 (data v6 -> v7)
        if "settings" in data:
            data["settings"].setdefault("conditional_formatting_rules", [])
            data["settings"].setdefault("live_quiz_score_font_color", DEFAULT_QUIZ_SCORE_FONT_COLOR)
            data["settings"].setdefault("live_quiz_score_font_style_bold", DEFAULT_QUIZ_SCORE_FONT_STYLE_BOLD)
            data["settings"].setdefault("quiz_mark_types", DEFAULT_QUIZ_MARK_TYPES) # Previously just strings
            data["settings"].setdefault("last_used_quiz_name_timeout_minutes", 60)
            data["settings"].setdefault("show_recent_incidents_during_quiz", True)

        if "behavior_log" in data:
            for log_entry in data["behavior_log"]:
                log_entry.setdefault("type", "behavior") # Old logs are behavior
                if log_entry["type"] == "quiz" and "marks_data" not in log_entry:
                    # Simple migration for old quiz logs to new marks_data structure
                    score_str = log_entry.get("score", "") # e.g., "7/10"
                    marks = {}
                    if score_str and "/" in score_str:
                        try:
                            correct, total = map(int, score_str.split('/'))
                            marks[DEFAULT_QUIZ_MARK_TYPES[0]["id"]] = correct # Assume first type is "Correct"
                            marks[DEFAULT_QUIZ_MARK_TYPES[1]["id"]] = total - correct # Assume second is "Incorrect"
                            log_entry["num_questions"] = total
                        except ValueError:
                            log_entry["num_questions"] = log_entry.get("num_questions", 0) # Keep if already exists
                    log_entry["marks_data"] = marks

        print("Applied v6 (to v7) data migration.")
        return data

    def _migrate_v5_data(self, data):
        """Migration for data version 5 (APP_VERSION v0.5.x) to v6 (APP_VERSION v0.6.x)."""
        if "settings" in data:
            data["settings"].setdefault("max_undo_history_days", MAX_UNDO_HISTORY_DAYS)
            data["settings"].setdefault("selected_recent_behaviors_filter", None) # None means all
            data["settings"].setdefault("current_mode", "behavior")
            data["settings"].setdefault("default_quiz_name", "Pop Quiz")
        if "students" in data:
            for student_id, student_data in data["students"].items():
                student_data.setdefault("nickname", "")
                student_data.setdefault("gender", "Boy") # Default to Boy if missing
        if "_per_student_last_cleared" not in data:
            data["_per_student_last_cleared"] = {}
        if "undo_stack" not in data: data["undo_stack"] = []
        if "redo_stack" not in data: data["redo_stack"] = []
        print("Applied v5 (to v6) data migration.")
        return data

    def _migrate_v4_data(self, data):
        """Migration for data version 4 (APP_VERSION v0.4.x) to v5."""
        if "settings" in data:
            data["settings"].setdefault("reverse_incident_order", True) # Default new setting
            data["settings"].setdefault("show_full_recent_incidents", False)
        print("Applied v4 (to v5) data migration.")
        return data

    def _migrate_v3_edited_data(self, data):
        """Migration for data version 3 and older (APP_VERSION < v0.3.x) to v4."""
        if "settings" in data:
            data["settings"].setdefault("show_recent_incidents_on_boxes", True)
            data["settings"].setdefault("num_recent_incidents_to_show", 2)
            data["settings"].setdefault("recent_incident_time_window_hours", 24)
            data["settings"].setdefault("autosave_interval_ms", 30000)
            data["settings"].setdefault("default_student_box_width", DEFAULT_STUDENT_BOX_WIDTH)
            data["settings"].setdefault("default_student_box_height", DEFAULT_STUDENT_BOX_HEIGHT)
            data["settings"].setdefault("student_box_fill_color", DEFAULT_BOX_FILL_COLOR)
            data["settings"].setdefault("student_box_outline_color", DEFAULT_BOX_OUTLINE_COLOR)
            data["settings"].setdefault("student_font_family", DEFAULT_FONT_FAMILY)
            data["settings"].setdefault("student_font_size", DEFAULT_FONT_SIZE)
            data["settings"].setdefault("student_font_color", DEFAULT_FONT_COLOR)
            data["settings"].setdefault("grid_snap_enabled", False)
            data["settings"].setdefault("grid_size", DEFAULT_GRID_SIZE)
            data["settings"].setdefault("behavior_initial_overrides", {})

            # Student ID migration: from simple numbers to student_NUM
            # Furniture ID migration: from simple numbers to furniture_NUM
            if "next_student_id" in data["settings"]:
                data["settings"]["next_student_id_num"] = data["settings"].pop("next_student_id")
            if "next_furniture_id" in data["settings"]:
                data["settings"]["next_furniture_id_num"] = data["settings"].pop("next_furniture_id")

        migrated_students = {}
        if "students" in data:
            for k, v in data["students"].items():
                new_id = f"student_{k}" if not str(k).startswith("student_") else str(k)
                v["id"] = new_id # Ensure 'id' field matches key
                if 'name' in v and 'full_name' not in v: v['full_name'] = v['name'] # Old format had 'name'
                if 'first_name' not in v or 'last_name' not in v: # Try to split 'full_name'
                    parts = v.get('full_name', '').split(' ', 1)
                    v['first_name'] = parts[0]
                    v['last_name'] = parts[1] if len(parts) > 1 else "Lastname"
                v.setdefault("nickname", "")
                v.setdefault("gender", "Boy")
                v.setdefault("width", data.get("settings", {}).get("default_student_box_width", DEFAULT_STUDENT_BOX_WIDTH))
                v.setdefault("height", data.get("settings", {}).get("default_student_box_height", DEFAULT_STUDENT_BOX_HEIGHT))
                migrated_students[new_id] = v
        data["students"] = migrated_students

        migrated_furniture = {}
        if "furniture" in data:
            for k, v in data["furniture"].items():
                new_id = f"furniture_{k}" if not str(k).startswith("furniture_") else str(k)
                v["id"] = new_id
                v.setdefault("width", REBBI_DESK_WIDTH) # Example default, adjust as needed
                v.setdefault("height", REBBI_DESK_HEIGHT)
                migrated_furniture[new_id] = v
        data["furniture"] = migrated_furniture

        if "behavior_log" in data:
            for entry in data["behavior_log"]:
                if "student_id" in entry and not str(entry["student_id"]).startswith("student_"):
                    entry["student_id"] = f"student_{entry['student_id']}"
                if "student_name" in entry and ("student_first_name" not in entry or "student_last_name" not in entry):
                    parts = entry["student_name"].split(" ", 1)
                    entry["student_first_name"] = parts[0]
                    entry["student_last_name"] = parts[1] if len(parts) > 1 else ""
                entry.setdefault("day", datetime.fromisoformat(entry["timestamp"]).strftime('%A') if "timestamp" in entry else "Unknown")
        print("Applied v3 (to v4) data migration (includes older formats).")
        return data

    def autosave_data_wrapper(self):
        self.save_data_wrapper(source="autosave")
        if hasattr(self, 'autosave_excel_log') and callable(self.autosave_excel_log):
             self.autosave_excel_log() # Call autosave for Excel if it exists
        self.root.after(self.settings.get("autosave_interval_ms", 30000), self.autosave_data_wrapper)

    def autosave_excel_log(self):
        if self.settings.get("enable_excel_autosave", False):
            if not self.behavior_log and not self.homework_log: # Don't save if empty
                print("Autosave Excel: Log is empty, skipping.")
                return

            filename = AUTOSAVE_EXCEL_FILE
            #try:
            # Use current settings for export filters if they exist
            filter_settings = {
                "start_date": None, "end_date": None,
                "selected_students": "all", "student_ids": [],
                "selected_behaviors": "all", "behaviors_list": [],
                "selected_homework_types": "all", "homework_types_list": [], # New
                "include_behavior_logs": True,
                "include_quiz_logs": True,
                "include_homework_logs": True, # New
                "include_summaries": self.settings.get("excel_export_include_summaries_by_default", True),
                "separate_sheets_by_log_type": self.settings.get("excel_export_separate_sheets_by_default", True),
                "excel_export_master_log_by_default": self.settings.get("excel_export_master_log_by_default", True)
            }
            self.export_data_to_excel(filename, "xlsx", filter_settings, is_autosave=True)
                # self.update_status(f"Log autosaved to {os.path.basename(filename)} at {datetime.now().strftime('%H:%M:%S')}")
            #except Exception as e:
            #    print(f"Error during Excel autosave: {e}")
            #   # self.update_status(f"Error during Excel autosave: {e}")
    
    def load_custom_behaviors(self):
        loaded_data = self._read_and_decrypt_file(CUSTOM_BEHAVIORS_FILE)
        self.custom_behaviors = loaded_data if isinstance(loaded_data, list) else []

    def save_custom_behaviors(self):
        self._encrypt_and_write_file(CUSTOM_BEHAVIORS_FILE, self.custom_behaviors)
    
    def load_custom_homework_types(self): # NEW
        """Loads customizable homework types (e.g., "Reading Assignment", "Worksheet")."""
        loaded_data = self._read_and_decrypt_file(CUSTOM_HOMEWORK_TYPES_FILE)
        self.custom_homework_types = loaded_data if isinstance(loaded_data, list) else []

    def save_custom_homework_types(self): # NEW
        self._encrypt_and_write_file(CUSTOM_HOMEWORK_TYPES_FILE, self.custom_homework_types)
    
    def load_custom_homework_statuses(self): # RENAMED
        """Loads customizable homework statuses (e.g., "Done", "Not Done", "Late")."""
        loaded_data = self._read_and_decrypt_file(CUSTOM_HOMEWORK_STATUSES_FILE)
        self.custom_homework_statuses = loaded_data if isinstance(loaded_data, list) else []

    def save_custom_homework_statuses(self): # RENAMED
        self._encrypt_and_write_file(CUSTOM_HOMEWORK_STATUSES_FILE, self.custom_homework_statuses)
    
    # Remove the old load_custom_homework_session_types and save_custom_homework_session_types
    # as CUSTOM_HOMEWORK_TYPES_FILE now serves this purpose.

    def update_all_behaviors(self):
        self.all_behaviors = DEFAULT_BEHAVIORS_LIST + [b["name"] if isinstance(b, dict) else str(b) for b in self.custom_behaviors]

    def update_all_homework_types(self): # NEW
        hidden_defaults = self.settings.get("hidden_default_homework_types", [])
        # Start with default types that are not hidden
        self.all_homework_types = [ht for ht in DEFAULT_HOMEWORK_TYPES_LIST if ht not in hidden_defaults]
        
        # Add custom types, ensuring no duplicates
        custom_names = [b["name"] for b in self.custom_homework_types if isinstance(b, dict) and "name" in b]
        for name in custom_names:
            if name not in self.all_homework_types:
                self.all_homework_types.append(name)

    def update_all_homework_statuses(self): # RENAMED
        """Combines default and custom homework statuses."""
        self.all_homework_statuses = DEFAULT_HOMEWORK_STATUSES + [item["name"] for item in self.custom_homework_statuses]

    def update_all_homework_session_types(self): # MODIFIED
        """This list is for the 'Yes/No' live session dialog. It's now based on the main homework types list."""
        # The list needs to be in dict format {"id", "name"} for the dialog to work.
        # We'll create temporary IDs for the default ones.
        default_as_dicts = [{"id": f"default_{name.lower().replace(' ','_')}", "name": name} for name in DEFAULT_HOMEWORK_TYPES_LIST]
        # Custom ones already have IDs.
        self.all_homework_session_types = default_as_dicts + [ct for ct in self.custom_homework_types if isinstance(ct, dict)]
    
    def load_student_groups(self):
        loaded_data = self._read_and_decrypt_file(STUDENT_GROUPS_FILE)
        self.student_groups = loaded_data if isinstance(loaded_data, dict) else {}
        
        if self.student_groups:
            max_g_id = 0
            for gid in self.student_groups:
                if gid.startswith("group_"):
                    try: max_g_id = max(max_g_id, int(gid.split("_")[1]))
                    except (ValueError, IndexError): pass
            self.next_group_id_num = max(self.settings.get("next_group_id_num",1), max_g_id + 1)
            self.settings["next_group_id_num"] = self.next_group_id_num

    def save_student_groups(self):
        self._encrypt_and_write_file(STUDENT_GROUPS_FILE, self.student_groups)

    def load_quiz_templates(self):
        loaded_data = self._read_and_decrypt_file(QUIZ_TEMPLATES_FILE)
        self.quiz_templates = loaded_data if isinstance(loaded_data, dict) else {}

        if self.quiz_templates:
            max_qt_id = 0
            for qtid in self.quiz_templates:
                if qtid.startswith("quiztemplate_"):
                    try: max_qt_id = max(max_qt_id, int(qtid.split("_")[1]))
                    except (ValueError, IndexError): pass
            self.next_quiz_template_id_num = max(self.settings.get("next_quiz_template_id_num",1), max_qt_id + 1)
            self.settings["next_quiz_template_id_num"] = self.next_quiz_template_id_num

    def save_quiz_templates(self):
        self._encrypt_and_write_file(QUIZ_TEMPLATES_FILE, self.quiz_templates)

    def load_homework_templates(self): # New
        loaded_data = self._read_and_decrypt_file(HOMEWORK_TEMPLATES_FILE)
        self.homework_templates = loaded_data if isinstance(loaded_data, dict) else {}

        if self.homework_templates:
            max_ht_id = 0
            for htid in self.homework_templates:
                if htid.startswith("hwtemplate_"):
                    try: max_ht_id = max(max_ht_id, int(htid.split("_")[1]))
                    except (ValueError, IndexError): pass
            self.next_homework_template_id_num = max(self.settings.get("next_homework_template_id_num",1), max_ht_id + 1)
            self.settings["next_homework_template_id_num"] = self.next_homework_template_id_num

    def save_homework_templates(self): # New
        self._encrypt_and_write_file(HOMEWORK_TEMPLATES_FILE, self.homework_templates)
    
    def update_all_homework_log_behaviors(self): # New
        self.all_homework_log_behaviors = DEFAULT_HOMEWORK_LOG_BEHAVIORS + [b["name"] for b in self.custom_homework_statuses if "name" in b]
   
    def get_earliest_log_date(self, type):
        log_source = self.behavior_log + self.homework_log
        
        #all_recent_logs = sorted(
        #            [log for log in log_source
        #             if log["student_id"] == student_id and log.get("type") in type_filter_values and
        #                datetime.fromisoformat(log["timestamp"]) >= cutoff_time],
        #            key=lambda x: x["timestamp"], reverse=True
        #        )
        
        
        all_recent_logs = sorted([log for log in log_source],
                                  key=lambda x: x["timestamp"], reverse=False)

        if all_recent_logs:
            v = all_recent_logs[0].get("timestamp")
            b = v[:v.find("T")]
            c = int(b.replace("-", ""))
            
        else:
            c = str(datetime.now())
            c = str(c.replace("-", ""))
        if type == "y":
            c = str(c)[:4]
        elif type == "m":
            c = str(c)[4:6]
        elif type == "d":
            c = str(c)[6:9]
        print(c)
        return c
    
    def export_log_dialog_with_filter(self, export_type="xlsx"):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Export", "Enter password to export log data:"): return
        date = datetime_date(int(self.get_earliest_log_date("y")), int(self.get_earliest_log_date("m")), int(self.get_earliest_log_date("d")))
        print(str(date))
        dialog = ExportFilterDialog(self.root, self.students, self.all_behaviors,
                                    self.all_homework_session_types + self.all_homework_statuses, # Combine all possible homework type names for filter
                                    default_settings=self.settings, earliest_date=date)
        if dialog.result:
            filter_settings = dialog.result
            default_filename = f"behavior_log_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            if export_type == "xlsx":
                file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_filename + ".xlsx",
                                                       filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")], parent=self.root)
            elif export_type == "xlsm":
                file_path = filedialog.asksaveasfilename(defaultextension=".xlsm", initialfile=default_filename + ".xlsm",
                                                       filetypes=[("Excel Macro-Enabled files", "*.xlsm"), ("All files", "*.*")], parent=self.root)
            elif export_type == "csv":
                 file_path = filedialog.asksaveasfilename(defaultextension=".zip", initialfile=default_filename + "_csv.zip",
                                                       filetypes=[("ZIP archives", "*.zip"), ("All files", "*.*")], parent=self.root)
            else: return

            if file_path:
                try:
                    if export_type in ["xlsx", "xlsm"]:
                        self.export_data_to_excel(file_path, export_type, filter_settings)
                    elif export_type == "csv":
                        self.export_data_to_csv_zip(file_path, filter_settings)

                    self.last_excel_export_path = file_path # Store path even for CSV for "Open Last Export Folder"
                    self.update_open_last_export_folder_menu_item()
                    self.save_data_wrapper(source="export_log")
                    self.update_status(f"Log exported to {os.path.basename(file_path)}")
                    if messagebox.askyesno("Export Successful", f"Log exported successfully to:\n{file_path}\n\nDo you want to open the file location?", parent=self.root):
                        self.open_last_export_folder()
                except Exception as e:
                    messagebox.showerror("Export Error", f"Failed to export log: {e}", parent=self.root)
                    self.update_status(f"Error exporting log: {e}")
                    print(f"Error: {e}")
            else: self.update_status("Export cancelled.")
            self.password_manager.record_activity()

    def _make_safe_sheet_name(self, name_str, id_fallback="Sheet"):
        invalid_chars = r'[\\/?*\[\]:]' # Excel invalid sheet name characters
        safe_name = re.sub(invalid_chars, '_', str(name_str))
        if not safe_name: safe_name = str(id_fallback)
        return safe_name[:31] # Max 31 chars for sheet names

    def export_data_to_excel(self, file_path, export_format="xlsx", filter_settings=None, is_autosave=False, export_all_students_info = True):
        # ... (substantially updated for new log types, summaries, and filtering)
        wb = Workbook()
        wb.remove(wb.active) # Remove default sheet
        mark_type_configs = self.settings.get("quiz_mark_types", [])
        mark_type_configs_h = self.settings.get("homework_mark_types", [])
        quiz_mark_type_headers = [mt["name"] for mt in mark_type_configs]
        homework_mark_type_headers = [mt["name"] for mt in mark_type_configs_h]
        homework_session_types_headers = [mt["name"] for mt in self.all_homework_session_types]

        student_data_for_export = {sid: {"first_name": s["first_name"], "last_name": s["last_name"], "full_name": s["full_name"]} for sid, s in self.students.items()}
        
        logs_to_process = []
        if filter_settings.get("include_behavior_logs", True):
            logs_to_process.extend([log for log in self.behavior_log if log.get("type") == "behavior"])
        if filter_settings.get("include_quiz_logs", True):
            logs_to_process.extend([log for log in self.behavior_log if log.get("type") == "quiz"])
        if filter_settings.get("include_homework_logs", True): # New
            logs_to_process.extend([log for log in self.homework_log if log.get("type") == "homework" or log.get("type") == "homework_session_y" or log.get("type") == "homework_session_s"])

        # Apply filters
        filtered_stud_ids = set()
        filtered_log = []
        start_date = filter_settings.get("start_date")
        end_date = filter_settings.get("end_date")
        selected_students_option = filter_settings.get("selected_students", "all")
        student_ids_filter = filter_settings.get("student_ids", [])
        selected_behaviors_option = filter_settings.get("selected_behaviors", "all")
        behaviors_list_filter = filter_settings.get("behaviors_list", [])
        selected_homework_types_option = filter_settings.get("selected_homework_types", "all") # New
        homework_types_list_filter = filter_settings.get("homework_types_list", []) # New
        for entry in logs_to_process:
            try:
                entry_date = datetime.fromisoformat(entry["timestamp"]).date()
                if start_date and entry_date < start_date: continue
                if end_date and entry_date > end_date: continue
            except ValueError: continue # Skip if timestamp is invalid

            if selected_students_option == "specific" and entry["student_id"] not in student_ids_filter: continue
            filtered_stud_ids.add(entry["student_id"])
            log_type = entry.get("type", "behavior")
            
            entry_name_field = entry.get("behavior") # Default for behavior and quiz
            if log_type == "homework" or log_type == "homework_session_y" or log_type == "homework_session_s":
                entry_name_field = entry.get("homework_type", entry.get("behavior")) # For homework logs
                #entry_name_field2 = entry.get("home")
            #print("list", homework_types_list_filter)
            #print("entry", entry_name_field)
            if log_type == "behavior" or log_type == "quiz":
                if selected_behaviors_option == "specific" and entry_name_field not in behaviors_list_filter: continue
            elif log_type == "homework" or log_type == "homework_session_s":
                if selected_homework_types_option == "specific" and entry_name_field not in homework_types_list_filter: continue
                elif selected_homework_types_option == "specific" and entry_name_field in homework_types_list_filter: continue
            elif log_type == "homework_session_y":
                if selected_homework_types_option == "specific" and entry_name_field not in homework_types_list_filter: continue#continue
                elif selected_homework_types_option == "specific" and entry_name_field in homework_types_list_filter: pass
            filtered_log.append(entry)
        
        filtered_log.sort(key=lambda x: x["timestamp"])

        # Determine sheet strategy
        separate_sheets = filter_settings.get("separate_sheets_by_log_type", True) # type: ignore
        master_log = filter_settings.get("include_master_log", True) if separate_sheets else False
        sheets_data = {} # {sheet_name: [entries]}

        if separate_sheets and not master_log:
            if filter_settings.get("include_behavior_logs", True): sheets_data["Behavior Log"] = []
            if filter_settings.get("include_quiz_logs", True): sheets_data["Quiz Log"] = []
            if filter_settings.get("include_homework_logs", True): sheets_data["Homework Log"] = [] # New
            for entry in filtered_log:
                log_type = entry.get("type")
                if log_type == "behavior" and "Behavior Log" in sheets_data: sheets_data["Behavior Log"].append(entry)
                elif log_type == "quiz" and "Quiz Log" in sheets_data: sheets_data["Quiz Log"].append(entry)
                elif (log_type == "homework" or log_type == "homework_session_y" or log_type == "homework_session_s") and "Homework Log" in sheets_data: sheets_data["Homework Log"].append(entry)
        elif separate_sheets and master_log:
            if filter_settings.get("include_behavior_logs", True): sheets_data["Behavior Log"] = []
            if filter_settings.get("include_quiz_logs", True): sheets_data["Quiz Log"] = []
            if filter_settings.get("include_homework_logs", True): sheets_data["Homework Log"] = [] # New
            if filter_settings.get("include_master_log", True): sheets_data["Master Log"] = [] # Newer
            for entry in filtered_log:
                log_type = entry.get("type")
                if log_type == "behavior" and "Behavior Log" in sheets_data: sheets_data["Behavior Log"].append(entry)
                elif log_type == "quiz" and "Quiz Log" in sheets_data: sheets_data["Quiz Log"].append(entry)
                elif (log_type == "homework" or log_type == "homework_session_y" or log_type == "homework_session_s") and "Homework Log" in sheets_data: sheets_data["Homework Log"].append(entry)
            sheets_data["Master Log"] = filtered_log
        else:
            sheets_data["Combined Log"] = filtered_log


        bold_font = OpenpyxlFont(bold=True)
        center_alignment = OpenpyxlAlignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = OpenpyxlAlignment(horizontal='left', vertical='center', wrap_text=True)
        right_alignment = OpenpyxlAlignment(horizontal='right', vertical='center', wrap_text=False)

        for sheet_name, entries_for_sheet in sheets_data.items():
            if not entries_for_sheet and ((sheet_name != "Combined Log" or sheet_name != "Master Log") or not filtered_log) : continue # Skip empty specific sheets

            ws = wb.create_sheet(title=sheet_name)
            headers = ["Timestamp", "Date", "Time", "Day", "Student ID", "First Name", "Last Name"]
            if sheet_name == "Behavior Log" or not separate_sheets or sheet_name == "Master Log": headers.append("Behavior")
            if sheet_name == "Quiz Log" or not separate_sheets or sheet_name == "Master Log":
                headers.extend(["Quiz Name", "Num Questions"])
                # Add headers for each mark type (e.g., Correct, Incorrect, Bonus)
                for mt in self.settings.get("quiz_mark_types", []): headers.append(mt["name"])
                headers.append("Quiz Score (%)")
            if sheet_name == "Homework Log" or not separate_sheets or sheet_name == "Master Log": # New headers for Homework
                headers.extend(["Homework Type/Session Name", "Num Items"])
                # Add headers for each homework mark type
                for hmt in self.settings.get("homework_mark_types", []): headers.append(hmt["name"])
                headers.extend(["Homework Score (Total Pts)", "Homework Effort"]) # Example summary fields
                headers.extend(homework_session_types_headers)
            headers.append("Comment")
            i=0
            for header in headers: 
                if header == "Complete": headers[i] = "Complete/Did"
                i += 1
            if not separate_sheets or sheet_name == "Master Log": headers.append("Log Type")


            for col_num, header_title in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header_title)
                cell.font = bold_font; cell.alignment = center_alignment
            ws.freeze_panes = 'A2'

            row_num = 2
            for entry in entries_for_sheet:
                student_info = student_data_for_export.get(entry["student_id"], {"first_name": "N/A", "last_name": "N/A"})
                try: dt_obj = datetime.fromisoformat(entry["timestamp"])
                except ValueError: dt_obj = datetime.now() # Fallback
                col_num = 1
                ws.cell(row=row_num, column=col_num, value=entry["timestamp"]); col_num+=1
                ws.cell(row=row_num, column=col_num, value=dt_obj.strftime('%Y-%m-%d')).alignment = right_alignment; col_num+=1
                ws.cell(row=row_num, column=col_num, value=dt_obj.strftime('%H:%M:%S')).alignment = right_alignment; col_num+=1
                ws.cell(row=row_num, column=col_num, value=entry.get("day", dt_obj.strftime('%A'))); col_num+=1
                ws.cell(row=row_num, column=col_num, value=entry["student_id"]); col_num+=1
                ws.cell(row=row_num, column=col_num, value=student_info["first_name"]); col_num+=1
                ws.cell(row=row_num, column=col_num, value=student_info["last_name"]); col_num+=1

                entry_type = entry.get("type", "behavior")

                if sheet_name == "Behavior Log" or ((not separate_sheets or sheet_name == "Master Log") and entry_type == "behavior"):
                    ws.cell(row=row_num, column=col_num, value=entry.get("behavior")); col_num+=1
                elif sheet_name == "Quiz Log" or ((not separate_sheets or sheet_name == "Master Log") and entry_type == "quiz"):
                    ws.cell(row=row_num, column=col_num, value=entry.get("behavior")); col_num+=1 # Quiz Name
                    num_q = entry.get("num_questions", 0); ws.cell(row=row_num, column=col_num, value=num_q).alignment = right_alignment; col_num+=1
                    marks_data = entry.get("marks_data", {})
                    total_possible_points_for_calc = 0; total_earned_points_for_calc = 0; extra_credit_earned = 0
                    for mt in self.settings.get("quiz_mark_types", []):
                        points = marks_data.get(mt["id"], 0)
                        ws.cell(row=row_num, column=col_num, value=points).alignment = right_alignment; col_num+=1
                        if mt.get("contributes_to_total", True): total_possible_points_for_calc += mt.get("default_points",1) * num_q # Simplified: assumes each question can get this mark type
                        if points > 0 : # Only add earned if student got this mark
                            if mt.get("is_extra_credit", False): extra_credit_earned += points * mt.get("default_points",1)
                            else: total_earned_points_for_calc += points * mt.get("default_points",1)
                    # More robust score calculation needed based on how num_questions and marks_data relate
                    score_percent = 0
                    if num_q > 0: # Use num_questions as the basis for total possible points from main Qs
                        # Calculate total possible for main questions based on default points of contributing mark types
                        # This is a simplification; assumes each question has a potential max based on one 'correct' type
                        main_q_total_possible = 0
                        correct_type = next((m for m in self.settings.get("quiz_mark_types",[]) if m.get("id") == "mark_correct"), None)
                        if correct_type: main_q_total_possible = correct_type.get("default_points", 1) * num_q

                        if main_q_total_possible > 0:
                            score_percent = ((total_earned_points_for_calc + extra_credit_earned) / main_q_total_possible) * 100
                        elif total_earned_points_for_calc + extra_credit_earned > 0 : # Scored only on EC or non-standard
                            score_percent = 100 # Or some other representation
                    ws.cell(row=row_num, column=col_num, value=round(score_percent,2) if score_percent else "").alignment = right_alignment; col_num+=1
                elif sheet_name == "Homework Log" or ((not separate_sheets or sheet_name == "Master Log") and (entry_type == "homework" or entry_type == "homework_session_y" or entry_type == "homework_session_s")): # New Homework
                    ws.cell(row=row_num, column=col_num, value=entry.get("homework_type", entry.get("behavior"))); col_num+=1 # Homework Type/Session Name
                    num_items = entry.get("num_items") # For manually logged with marks
                    if entry.get("type") == "homework_session_s": # For live sessions
                        # Try to count items from details if Yes/No mode
                        homework_details = entry.get("homework_details", {})
                        if not is_autosave:
                            num_items = len(homework_details.get("selected_options",[])) if isinstance(homework_details, dict) else 0
                    elif entry.get("type") == "homework_session_y":
                        num_items = None
                    
                    if separate_sheets and (sheet_name == "Combined Log") or sheet_name == "Master Log":
                        col_num += len(headers)-(len(homework_session_types_headers))-(col_num)-10
                    ws.cell(row=row_num, column=col_num, value=num_items if num_items is not None else "").alignment = right_alignment; col_num+=1
                    total_hw_points = 0; effort_score_val = "" # For summary columns
                    if entry_type == "homework" and "marks_data" in entry: # Graded manual log
                        
                        #if separate_sheets and (sheet_name == "Combined Log") or sheet_name == "Master Log":
                        #    col_num += len(headers)-(len(homework_session_types_headers))-(col_num)-9
                        
                        hw_marks_data = entry.get("marks_data", {})
                        for hmt in self.settings.get("homework_mark_types", []):
                            val = hw_marks_data.get(hmt["id"], "")
                            ws.cell(row=row_num, column=col_num, value=val).alignment = right_alignment; col_num+=1
                            if isinstance(val, (int,float)): total_hw_points += val # Sum points if numeric
                            if hmt["id"] == "hmark_effort": effort_score_val = val # Capture effort score
                            
                    elif entry_type == "homework_session_s" or entry_type == "homework_session_y": # Live session log
                        # For live sessions, fill placeholders for mark type columns or try to map
                        session_details = entry.get("homework_details", {})
                        live_session_mode = entry.get("type")
                        if live_session_mode == "homework_session_y":
                            """
                            for hmt in self.settings.get("homework_mark_types", []): # Fill placeholders
                                # Could try to map "Yes" to complete, "No" to not done, etc.
                                # For now, just leave blank or show raw status if one of the types matches the key
                                found_status_for_mark_type = ""
                                for hw_type_id_key, status_val in session_details.items():
                                    # This mapping is very approximate.
                                    #print(status_val)
                                    if hmt["name"].lower() in hw_type_id_key.lower() or hmt["name"].lower() == status_val.lower():
                                        found_status_for_mark_type = status_val
                                        break
                                    elif "complete" in hmt["name"].lower() and status_val.lower() == "yes":
                                        found_status_for_mark_type = "Yes" # Or map to points
                                        if "default_points" in hmt: total_hw_points += hmt["default_points"]
                                        break
                                    elif "not done" in hmt["name"].lower() and status_val.lower() == "no":
                                        found_status_for_mark_type = "No"
                                        if "default_points" in hmt: total_hw_points += hmt["default_points"]
                                        break
                                ws.cell(row=row_num, column=col_num, value=found_status_for_mark_type).alignment = right_alignment; col_num+=1
                            """
                            i = 0
                            #print(self.all_homework_session_types)
                            found_status_for_mark_type2 = ""
                            col_num += ((((len(headers)-col_num)-len(homework_session_types_headers))) if not is_autosave else (((len(headers)-col_num)-len(homework_session_types_headers)))) if "Master Log" not in sheet_name or "Combined Log" not in sheet_name else ((((len(headers)-col_num)-len(homework_session_types_headers))-1) if not is_autosave else (((len(headers)-col_num)-len(homework_session_types_headers))))
                            for typeh in entry.get("homework_details"):
                                #print(typeh)
                                for hwtype in self.all_homework_session_types:
                                    h_id = hwtype.get("id")
                                    name = hwtype.get("name")
                                    if typeh == h_id:
                                        found_status_for_mark_type2 = entry.get("homework_details").get(typeh).capitalize()
                                i += 1
                                ws.cell(row=row_num, column=col_num, value=found_status_for_mark_type2).alignment = right_alignment; col_num+=1
                        elif live_session_mode == "homework_session_s":
                            selected_options = session_details.get("selected_options", [])
                            
                            
                            s_correct = str(selected_options).removeprefix("[").removesuffix("]")
                            #s_total = len(selected_options)
                            #ws.cell(row=row_num, column=col_num, value=s_total).alignment = right_alignment; col_num+=1
                            ws.cell(row=row_num, column=col_num, value=s_correct).alignment = right_alignment; col_num+=1
                            """for hmt in self.settings.get("homework_mark_types", []): # Fill placeholders based on selected options
                                val_to_put = ""
                                if hmt["name"] in selected_options: # If a mark type name matches a selected option
                                    val_to_put = "Selected" # or hmt["default_points"]
                                    if "default_points" in hmt: total_hw_points += hmt["default_points"]
                                ws.cell(row=row_num, column=col_num, value=val_to_put).alignment = right_alignment; col_num+=1"""
                                
                                
                        else: # Unknown live mode or no details
                            for _ in self.settings.get("homework_mark_types", []): ws.cell(row=row_num, column=col_num, value="").alignment = right_alignment; col_num+=1

                    ws.cell(row=row_num, column=col_num, value=total_hw_points if total_hw_points else "").alignment = right_alignment; col_num+=1 # Total Points
                    ws.cell(row=row_num, column=col_num, value=effort_score_val).alignment = right_alignment; col_num+=1 # Effort

                comment_col = headers.index("Comment") + 1
                ws.cell(row=row_num, column=comment_col, value=entry.get("comment", "")).alignment = left_alignment
                if not separate_sheets or sheet_name == "Master Log":
                    log_type_col = headers.index("Log Type") + 1
                    ws.cell(row=row_num, column=log_type_col, value=entry.get("type", "behavior").capitalize())
                row_num += 1

            # Auto-size columns
            for col_letter in [get_column_letter(i) for i in range(1, ws.max_column + 1)]:
                max_length = 0
                column_values = [cell.value for cell in ws[col_letter]]
                for cell_val in column_values:
                    if cell_val is not None:
                        try: max_length = max(max_length, len(str(cell_val)))
                        except: pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[col_letter].width = min(max(adjusted_width, 10), 50) # Min/Max width

        log_data_to_export = filtered_log

        # --- Individual Student Log Sheets ---
        if export_all_students_info: # Only create these if full export
            student_worksheets = {} # {student_id: worksheet_object}
            student_headers = ["Timestamp", "Type", "Behavior/Homework/Quiz Name",
                               "Correct/Did", "Total Qs/Total Selected", "Percentage", "Comment", "Day"]
            student_headers.extend(quiz_mark_type_headers) # Add mark types here too
            student_headers.extend(homework_mark_type_headers)
            student_headers.extend(homework_session_types_headers)
            # Put this line below back if it puts something in those columns
            #student_headers.extend(["Homework Type/Session Name", "Num Items"])

            for entry in log_data_to_export:
                student_id = entry["student_id"]
                student_data = self.students.get(student_id)
                student_name_for_sheet = self._make_safe_sheet_name(
                    f"{student_data['first_name']}_{student_data['last_name']}" if student_data else f"Unknown_{student_id}",
                    student_id
                )
                s_homework_marks_data = [""] * len(homework_mark_type_headers)
                
                if entry.get("type") == "homework" and "marks_data" in entry: # Graded manual log
                    total_hw_points = 0; effort_score_val = "" # For summary columns
                    hw_marks_data = entry.get("marks_data", {})
                    i=0
                    for hmt in self.settings.get("homework_mark_types", []):
                        val = hw_marks_data.get(hmt["id"], "")
                        #ws.cell(row=row_num, column=col_num, value=val).alignment = right_alignment; col_num+=1
                        if isinstance(val, (int,float)): total_hw_points += val # Sum points if numeric
                        if hmt["id"] == "hmark_effort": effort_score_val = val # Capture effort score
                        #for i, mt_config_h in enumerate(mark_type_configs_h):
                        s_homework_marks_data[i] = hw_marks_data.get(hmt["id"], "")
                        i+=1
                    #print(s_homework_marks_data)
                    
                    # Add headers for each homework mark type
                    #for hmt in self.settings.get("homework_mark_types", []): student_headers.append(hmt["name"])
                    
                    
                if student_id not in student_worksheets:
                    ws_student = wb.create_sheet(title=student_name_for_sheet)
                    student_worksheets[student_id] = ws_student
                    ws_student.append(student_headers)
                    for col_num, header_text in enumerate(student_headers, 1):
                        cell = ws_student.cell(row=1, column=col_num)
                        cell.font = OpenpyxlFont(bold=True)
                        cell.alignment = OpenpyxlAlignment(horizontal="center")
                        width = len(header_text) + 5 # Basic width
                        if header_text == "Timestamp": width = 20
                        elif header_text == "Behavior/Homework/Quiz Name": width = 30
                        elif header_text == "Type": width = 20
                        elif header_text == "Comment": width = 40
                        elif header_text == "Day": width = 12
                        ws_student.column_dimensions[get_column_letter(col_num)].width = width

                ws_student = student_worksheets[student_id]
                ts_obj_s = datetime.fromisoformat(entry["timestamp"])
                s_correct, s_total, s_perc = "", "", ""
                s_quiz_marks_data = [""] * len(quiz_mark_type_headers)
                all_h_types = self.all_homework_session_types
                s_homework_marks_data_2 = [""] * len(all_h_types)
                #print(all_h_types)
                
                
                if entry.get("type") == "quiz":
                    s_marks_data = entry.get("marks_data")
                    s_num_q = entry.get("num_questions", self.settings.get("default_quiz_questions",10))
                    if "score_details" in entry: # Live quiz
                        s_correct = entry["score_details"].get("correct", "")
                        s_total = entry["score_details"].get("total_asked", "")
                        if isinstance(s_correct, (int, float)) and isinstance(s_total, (int, float)) and s_total > 0:
                            s_perc = f"{round((s_correct / s_total) * 100)}%"
                    elif isinstance(s_marks_data, dict):
                        primary_correct_id_s = next((mt["id"] for mt in mark_type_configs if mt["name"].lower() == "correct"), "mark1")
                        s_correct = s_marks_data.get(primary_correct_id_s, "")
                        s_total = s_num_q
                        if isinstance(s_correct, (int,float)) and isinstance(s_total, (int,float)) and s_total > 0:
                            s_perc = f"{round((s_correct / s_total) * 100)}%"
                        for i, mt_config_s in enumerate(mark_type_configs):
                            s_quiz_marks_data[i] = s_marks_data.get(mt_config_s["id"], "")
                elif entry.get("type") == "homework_session_s":
                    s_correct = str(entry["homework_details"].get("selected_options")).removeprefix("[").removesuffix("]")
                    s_total = len(entry["homework_details"]["selected_options"])
                    #pass #print("Homework_session")
                elif entry.get("type") == "homework_session_y":
                    i = 0
                    #print(self.all_homework_session_types)
                    for typeh in entry.get("homework_details"):
                        #print(typeh)
                        for hwtype in all_h_types:
                            h_id = hwtype.get("id")
                            name = hwtype.get("name")
                            if typeh == h_id:
                                s_homework_marks_data_2[i] = entry.get("homework_details").get(typeh).capitalize()
                        i += 1
                            
                s_row_base = [
                    ts_obj_s.strftime('%Y-%m-%d %H:%M:%S'),
                    entry.get("type", "behavior").capitalize(),
                    entry.get("behavior", ""),
                    #entry.get("homework", ""),
                    s_correct, s_total, s_perc,
                    entry.get("comment", "").replace("\n", " "),
                    entry.get("day", "")
                ]
                s_row_base.extend(s_quiz_marks_data)
                s_row_base.extend(s_homework_marks_data)
                s_row_base.extend(s_homework_marks_data_2)
                
                ws_student.append(s_row_base)

        # --- Student Information Sheet ---
        #print((filtered_stud_ids))
        if export_all_students_info and len(filtered_stud_ids) > 1:
            students_info_ws = wb.create_sheet(title="Students Info")
            student_info_headers = ["Student ID", "First Name", "Last Name", "Nickname", "Full Name", "Gender", "Group Name"]
            students_info_ws.append(student_info_headers)
            for col_num, header in enumerate(student_info_headers, 1):
                cell = students_info_ws.cell(row=1, column=col_num)
                cell.font = OpenpyxlFont(bold=True); cell.alignment = OpenpyxlAlignment(horizontal="center")
                info_widths = {"Student ID": 15, "First Name": 15, "Last Name": 15, "Nickname": 15,
                               "Full Name": 25, "Gender": 10, "Group Name": 20}
                students_info_ws.column_dimensions[get_column_letter(col_num)].width = info_widths.get(header, 12)
            
            sorted_students_info = sorted(self.students.values(), key=lambda s: (s.get("last_name", "").lower(), s.get("first_name", "").lower()))

            for student_data in sorted_students_info:
                if student_data.get("id", "") in filtered_stud_ids:
                    group_id = student_data.get("group_id")
                    group_name = ""
                    if self.settings.get("student_groups_enabled", True) and group_id and group_id in self.student_groups:
                        group_name = self.student_groups[group_id].get("name", "")

                    info_row = [
                        student_data.get("id", ""), student_data.get("first_name", ""),
                        student_data.get("last_name", ""), student_data.get("nickname", ""),
                        student_data.get("full_name", ""), student_data.get("gender", ""), group_name
                    ]
                    students_info_ws.append(info_row)


        # Add Summary Sheet if requested
        if filter_settings.get("include_summaries", True) and filtered_log:
            ws_summary = wb.create_sheet(title="Summary")
            current_row = 1
            ws_summary.cell(row=current_row, column=1, value="Log Summary").font = OpenpyxlFont(bold=True, size=14); current_row += 2

            # Behavior Summary
            if filter_settings.get("include_behavior_logs", True):
                ws_summary.cell(row=current_row, column=1, value="Behavior Summary by Student").font = bold_font; current_row += 1
                b_headers = ["Student", "Behavior", "Count"]
                for c_num, h_title in enumerate(b_headers, 1): ws_summary.cell(row=current_row, column=c_num, value=h_title).font = OpenpyxlFont(italic=True)
                current_row += 1
                behavior_counts = {} # {student_id: {behavior_name: count}}
                for entry in filtered_log:
                    if entry.get("type") == "behavior":
                        sid = entry["student_id"]; b_name = entry.get("behavior")
                        behavior_counts.setdefault(sid, {}).setdefault(b_name, 0)
                        behavior_counts[sid][b_name] += 1
                for sid in sorted(behavior_counts.keys(), key=lambda x: student_data_for_export.get(x, {}).get("last_name","")):
                    s_info = student_data_for_export.get(sid, {"full_name": "Unknown"})
                    for b_name, count in sorted(behavior_counts[sid].items()):
                        ws_summary.cell(row=current_row, column=1, value=s_info["full_name"])
                        ws_summary.cell(row=current_row, column=2, value=b_name)
                        ws_summary.cell(row=current_row, column=3, value=count).alignment = right_alignment
                        current_row +=1
                current_row +=1 # Spacer

            # Quiz Summary
            if filter_settings.get("include_quiz_logs", True):
                ws_summary.cell(row=current_row, column=1, value="Quiz Averages by Student").font = bold_font; current_row += 1
                q_headers = ["Student", "Quiz Name", "Avg Score (%)", "Times Taken"]
                for c_num, h_title in enumerate(q_headers, 1): ws_summary.cell(row=current_row, column=c_num, value=h_title).font = OpenpyxlFont(italic=True)
                current_row += 1
                quiz_scores_summary = {} # {student_id: {quiz_name: [scores]}}
                for entry in filtered_log:
                    if entry.get("type") == "quiz":
                        sid = entry["student_id"]; q_name = entry.get("behavior"); num_q_s = entry.get("num_questions",0)
                        marks_d = entry.get("marks_data", {})
                        total_earned_s = 0; extra_credit_s = 0
                        for mt_s in self.settings.get("quiz_mark_types", []):
                            pts_s = marks_d.get(mt_s["id"], 0)
                            if pts_s > 0:
                                if mt_s.get("is_extra_credit", False): extra_credit_s += pts_s * mt_s.get("default_points",1)
                                else: total_earned_s += pts_s * mt_s.get("default_points",1)
                        main_q_total_possible_s = 0
                        correct_type_s = next((m for m in self.settings.get("quiz_mark_types",[]) if m.get("id") == "mark_correct"), None)
                        if correct_type_s and num_q_s > 0: main_q_total_possible_s = correct_type_s.get("default_points", 1) * num_q_s
                        score_val = ((total_earned_s + extra_credit_s) / main_q_total_possible_s) * 100 if main_q_total_possible_s > 0 else (100 if total_earned_s + extra_credit_s > 0 else 0)
                        quiz_scores_summary.setdefault(sid, {}).setdefault(q_name, []).append(score_val)
                for sid in sorted(quiz_scores_summary.keys(), key=lambda x: student_data_for_export.get(x, {}).get("last_name","")):
                    s_info = student_data_for_export.get(sid, {"full_name": "Unknown"})
                    for q_name, scores_list in sorted(quiz_scores_summary[sid].items()):
                        avg_score = sum(scores_list) / len(scores_list) if scores_list else 0
                        ws_summary.cell(row=current_row, column=1, value=s_info["full_name"])
                        ws_summary.cell(row=current_row, column=2, value=q_name)
                        ws_summary.cell(row=current_row, column=3, value=f"{avg_score:.2f}%").alignment = right_alignment
                        ws_summary.cell(row=current_row, column=4, value=len(scores_list)).alignment = right_alignment
                        current_row+=1
                current_row +=1

            # Homework Summary (New)
            if filter_settings.get("include_homework_logs", True):
                ws_summary.cell(row=current_row, column=1, value="Homework Completion by Student").font = bold_font; current_row += 1
                hw_headers = ["Student", "Homework Type/Session", "Count", "Total Points (if applicable)"]
                for c_num, h_title in enumerate(hw_headers, 1): ws_summary.cell(row=current_row, column=c_num, value=h_title).font = OpenpyxlFont(italic=True)
                current_row += 1
                homework_summary = {} # {student_id: {hw_type: {"count": 0, "total_points": 0}}}
                for entry in filtered_log:
                    if entry.get("type") == "homework" or entry.get("type") == "homework_session_s" or entry.get("type") == "homework_session_y":
                        sid = entry["student_id"]
                        hw_name = entry.get("homework_type", entry.get("behavior"))
                        summary_entry = homework_summary.setdefault(sid, {}).setdefault(hw_name, {"count": 0, "total_points": 0.0})
                        summary_entry["count"] += 1
                        # Sum points from marks_data for "homework" type
                        if entry.get("type") == "homework" and "marks_data" in entry:
                            for mark_id, mark_val in entry["marks_data"].items():
                                if isinstance(mark_val, (int, float)): summary_entry["total_points"] += mark_val
                        # Sum points from live session details (approximate)
                        elif entry.get("type") == "homework_session_y" or entry.get("type") == "homework_session_s":
                            hw_details = entry.get("homework_details", {})
                            live_mode = entry.get("type")
                            if live_mode == "homework_session_y":
                                for ht_id_key, status_val in hw_details.items():
                                     if status_val.lower() == "yes": # Simplified: 'yes' adds default points of 'complete' mark type
                                        complete_mark_type = next((m for m in self.settings.get("homework_mark_types",[]) if m["id"] == "hmark_complete"), None)
                                        if complete_mark_type: summary_entry["total_points"] += complete_mark_type.get("default_points",0)
                            elif live_mode == "homework_session_s":
                                selected_opts = hw_details.get("selected_options", [])
                                for opt_name in selected_opts:
                                    opt_mark_type = next((m for m in self.settings.get("homework_mark_types",[]) if m["name"] == opt_name), None)
                                    if opt_mark_type: summary_entry["total_points"] += opt_mark_type.get("default_points",0)

                for sid in sorted(homework_summary.keys(), key=lambda x: student_data_for_export.get(x, {}).get("last_name","")):
                    s_info = student_data_for_export.get(sid, {"full_name": "Unknown"})
                    for hw_name, data in sorted(homework_summary[sid].items()):
                        ws_summary.cell(row=current_row, column=1, value=s_info["full_name"])
                        ws_summary.cell(row=current_row, column=2, value=hw_name)
                        ws_summary.cell(row=current_row, column=3, value=data["count"]).alignment = right_alignment
                        ws_summary.cell(row=current_row, column=4, value=f"{data['total_points']:.2f}" if data['total_points'] else "").alignment = right_alignment
                        current_row += 1
                current_row += 1

            for col_letter_s in [get_column_letter(i) for i in range(1, ws_summary.max_column + 1)]:
                 ws_summary.column_dimensions[col_letter_s].width = 25


        # Save workbook
        try:
            wb.save(filename=file_path)
        except PermissionError as e:
            if is_autosave:
                print(f"Autosave PermissionError: {e}. File might be open.")
                # Don't show messagebox for autosave, just print
            else:
                messagebox.showerror("Save Error", f"Permission denied. Could not save to '{file_path}'.\nPlease ensure the file is not open in another program and you have write permissions.", parent=self.root)
            raise # Re-raise to be caught by the calling function for status update
        except Exception as e_save:
            if is_autosave: print(f"Autosave error: {e_save}")
            else: messagebox.showerror("Save Error", f"An unexpected error occurred while saving Excel file: {e_save}", parent=self.root)
            raise

    def export_data_to_csv_zip(self, zip_file_path, filter_settings=None):
        # ... (updated for new log types and filtering)
        temp_dir = tempfile.mkdtemp()
        try:
            student_data_for_export = {sid: {"first_name": s["first_name"], "last_name": s["last_name"], "full_name": s["full_name"]} for sid, s in self.students.items()}
            logs_to_process_csv = []
            if filter_settings.get("include_behavior_logs", True): logs_to_process_csv.extend([log for log in self.behavior_log if log.get("type") == "behavior"])
            if filter_settings.get("include_quiz_logs", True): logs_to_process_csv.extend([log for log in self.behavior_log if log.get("type") == "quiz"])
            if filter_settings.get("include_homework_logs", True): logs_to_process_csv.extend([log for log in self.homework_log if log.get("type") == "homework" or log.get("type") == "homework_session_y" or log.get("type") == "homework_session_s"])

            filtered_log_csv = []
            start_date_csv, end_date_csv = filter_settings.get("start_date"), filter_settings.get("end_date")
            sel_students_opt_csv, student_ids_flt_csv = filter_settings.get("selected_students", "all"), filter_settings.get("student_ids", [])
            sel_behaviors_opt_csv, behaviors_flt_csv = filter_settings.get("selected_behaviors", "all"), filter_settings.get("behaviors_list", [])
            sel_hw_opt_csv, hw_flt_csv = filter_settings.get("selected_homework_types", "all"), filter_settings.get("homework_types_list", [])

            for entry in logs_to_process_csv:
                try:
                    entry_date = datetime.fromisoformat(entry["timestamp"]).date()
                    if start_date_csv and entry_date < start_date_csv: continue
                    if end_date_csv and entry_date > end_date_csv: continue
                except ValueError: continue
                if sel_students_opt_csv == "specific" and entry["student_id"] not in student_ids_flt_csv: continue
                log_type_csv = entry.get("type", "behavior")
                entry_name_csv = entry.get("behavior")
                if log_type_csv == "homework" or log_type_csv == "homework_session": entry_name_csv = entry.get("homework_type", entry.get("behavior"))

                if log_type_csv == "behavior" or log_type_csv == "quiz":
                    if sel_behaviors_opt_csv == "specific" and entry_name_csv not in behaviors_flt_csv: continue
                elif log_type_csv == "homework" or log_type_csv == "homework_session":
                    if sel_hw_opt_csv == "specific" and entry_name_csv not in hw_flt_csv: continue
                filtered_log_csv.append(entry)
            filtered_log_csv.sort(key=lambda x: x["timestamp"])

            # CSV file for all logs (or separate if preferred, but Excel handles separation better)
            all_logs_csv_path = os.path.join(temp_dir, "all_logs.csv")
            with open(all_logs_csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ["Timestamp", "Date", "Time", "Day", "Student_ID", "First_Name", "Last_Name",
                              "Log_Type", "Item_Name", "Comment", "Num_Questions_Items",
                              "Marks_Data_JSON", "Score_Details_JSON", "Homework_Details_JSON"]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, extrasaction='ignore')
                writer.writeheader()
                for entry in filtered_log_csv:
                    student_info = student_data_for_export.get(entry["student_id"], {"first_name": "N/A", "last_name": "N/A"})
                    try: dt_obj = datetime.fromisoformat(entry["timestamp"])
                    except ValueError: dt_obj = datetime.now()
                    row_data = {
                        "Timestamp": entry["timestamp"], "Date": dt_obj.strftime('%Y-%m-%d'), "Time": dt_obj.strftime('%H:%M:%S'),
                        "Day": entry.get("day", dt_obj.strftime('%A')), "Student_ID": entry["student_id"],
                        "First_Name": student_info["first_name"], "Last_Name": student_info["last_name"],
                        "Log_Type": entry.get("type", "").capitalize(),
                        "Item_Name": entry.get("behavior", entry.get("homework_type", "")),
                        "Comment": entry.get("comment", ""),
                        "Num_Questions_Items": entry.get("num_questions", entry.get("num_items")),
                        "Marks_Data_JSON": json.dumps(entry.get("marks_data")) if "marks_data" in entry else "",
                        "Score_Details_JSON": json.dumps(entry.get("score_details")) if "score_details" in entry else "",
                        "Homework_Details_JSON": json.dumps(entry.get("homework_details")) if "homework_details" in entry else ""
                    }
                    writer.writerow(row_data)

            # CSV file for student list
            students_csv_path = os.path.join(temp_dir, "students.csv")
            with open(students_csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames_s = ["Student_ID", "First_Name", "Last_Name", "Nickname", "Gender", "Group_ID"]
                writer_s = csv.DictWriter(csvfile, fieldnames=fieldnames_s, extrasaction='ignore')
                writer_s.writeheader()
                for sid, sdata in self.students.items():
                     writer_s.writerow({"Student_ID": sid, "First_Name": sdata["first_name"], "Last_Name": sdata["last_name"],
                                        "Nickname": sdata.get("nickname",""), "Gender": sdata.get("gender",""), "Group_ID": sdata.get("group_id","")})

            # Create ZIP file
            with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.write(all_logs_csv_path, arcname="all_logs.csv")
                zf.write(students_csv_path, arcname="students.csv")
                if filter_settings.get("include_summaries", False): # type: ignore # Basic summary text file
                    summary_txt_path = os.path.join(temp_dir, "summary.txt")
                    with open(summary_txt_path, 'w', encoding='utf-8')as f_sum:
                        f_sum.write(f"Log Export Summary - {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
                        f_sum.write(f"Date Range: {start_date_csv or 'Any'} to {end_date_csv or 'Any'}\n")
                        f_sum.write(f"Total Log Entries Exported: {len(filtered_log_csv)}\n")
                        # Further summary details could be added here
                    zf.write(summary_txt_path, arcname="summary.txt")

        finally: shutil.rmtree(temp_dir) # Clean up temp directory

    def export_layout_as_image(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Export Image", "Enter password to export layout as image:"): return
        file_path = filedialog.asksaveasfilename(defaultextension=".png", initialfile=f"layout_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
                                               filetypes=[("PNG Image", "*.png"), ("JPG Image", "*.jpg"), ("WebP Image", "*.webp"), ("TIFF Image", "*.tiff"), ("BMP Image", "*.bmp"), ("PPM Image", "*.ppm"), ("PGM Image", "*.pgm"),("GIF Image", "*.gif"), ("All files", "*.*")], parent=self.root)
        if not file_path: self.update_status("Image export cancelled."); return
        try:
            # Determine current bounds of drawn items on canvas (in canvas coordinates)
            # This uses the scrollregion which should be set by draw_all_items
            s_region = self.canvas.cget("scrollregion")
            if not s_region: # Fallback if scrollregion is not set (e.g. empty canvas)
                 x1, y1, x2, y2 = 0,0, self.canvas.winfo_width(), self.canvas.winfo_height()
            else:
                try: 
                    v= s_region.split()
                    x1 = float(v[0])
                    y1 = float(v[1])
                    x2 = float(v[2])
                    y2 = float(v[3])
                except Exception as e: x1, y1, x2, y2 = 0,0, self.canvas.winfo_width(), self.canvas.winfo_height()
            
            # Ensure x1, y1 are not negative for postscript (though typically they are 0 or positive)
            # If they are negative, it means content is scrolled left/up off screen.
            # We want to capture from the top-leftmost content.
            
            postscript_x_offset = -x1 if x1 < 0 else 0
            postscript_y_offset = -y1 if y1 < 0 else 0
            
            # Create PostScript of the entire scrollable region
            ps_io = io.BytesIO()
            timestamp = str(IMAGENAMEW)
            self.canvas.postscript( # type: ignore
                x=x1 + postscript_x_offset,
                y=y1 + postscript_y_offset,
                width=x2 - x1, # Width of the scrollable area
                height=y2 - y1, # Height of the scrollable area
                colormode='color',
                file=(timestamp) # Write to BytesIO object
            )
            ps_io.seek(0)
            
            output_dpi = int(self.settings.get("output_dpi", 600))
            
            try:
                img = Image.open(os.path.abspath(timestamp))
                ps_file = ps_io
                output_image_file = file_path
                #output_dpi = 600 
                scale_factor = output_dpi / 72.0
                try: img.load(scale=scale_factor)  # type: ignore
                except AttributeError:
                    print("Warning: img.load(scale=...) might not be directly supported for .ps files in your Pillow version in this way.")
                    print("Pillow will use Ghostscript's default rasterization or a pre-set one.")
                    # If direct scaling isn't working, you might need to use subprocess for full control (see advanced section).

                # Now save the image. The 'dpi' parameter here is metadata for formats like PNG/TIFF.
                # The actual pixel dimensions are determined by the rasterization step.
                img.save(output_image_file, dpi=(output_dpi, output_dpi))
                print(f"PostScript file '{timestamp}' converted to '{output_image_file}' at {output_dpi} DPI.")     
                
                
                img.save(file_path, "png")
                self.update_status(f"Layout exported as image: {os.path.basename(file_path)}")
                if messagebox.askyesno("Export Successful", f"Layout image saved to:\n{file_path}\n\nDo you want to open the file location?", parent=self.root):
                    self.open_specific_export_folder(file_path)
            except (OSError, PIL.UnidentifiedImageError) as e_pil:
                print(f"PIL error processing PostScript: {e_pil}")
                if "gs" in str(e_pil).lower() or "ghostscript" in str(e_pil).lower():
                     messagebox.showerror("Image Export Error", "Failed to convert PostScript to image. Ghostscript might not be installed or found in your system's PATH. Please install Ghostscript to enable image export.", parent=self.root)
                else:
                     messagebox.showerror("Image Export Error", f"Failed to save image: {e_pil}.\nEnsure you have image processing libraries like Pillow and its dependencies (e.g., Ghostscript for EPS/PS) installed.", parent=self.root)
            except Exception as e: print("e2", e)
            finally:
                ps_io.close()
                """try: img.close(); os.remove(os.path.abspath(IMAGENAMEW))
                except FileNotFoundError: pass
                except: pass"""

        except tk.TclError as e_tk:
            messagebox.showerror("Image Export Error", f"Tkinter error during PostScript generation: {e_tk}", parent=self.root)
        except Exception as e:
            messagebox.showerror("Image Export Error", f"An unexpected error occurred: {e}", parent=self.root); print("e", e)
        finally:
            self.password_manager.record_activity()
        
    def _import_data_from_excel_logic(self, file_path, import_incidents_flag, student_sheet_name_to_import):
        # This function needs significant updates if we want to import detailed quiz scores.
        # For now, it will import students and basic incident info as before.
        # Importing complex quiz scores from Excel would require a well-defined column mapping.
        workbook = load_workbook(filename=file_path, data_only=True)
        imported_student_count = 0

        # --- Import Students ---
        if student_sheet_name_to_import:
            if student_sheet_name_to_import not in workbook.sheetnames:
                messagebox.showerror("Import Error", f"Selected student sheet '{student_sheet_name_to_import}' not found.", parent=self.root)
                return 0, 0

            sheet = workbook[student_sheet_name_to_import]
            header_row_values = [str(cell.value).lower().strip() if cell.value else "" for cell in sheet[1]]

            # Try to find columns for student data
            col_indices = {}
            common_headers = {
                "first_name": ["first", "first name", "firstname"],
                "last_name": ["last", "last name", "lastname", "surname"],
                "full_name": ["full name", "name", "student name"],
                "nickname": ["nickname", "preferred name", "nick"],
                "gender": ["gender", "sex"],
                "group_name": ["group", "group name", "student group"] # For importing group assignment by name
            }
            for key, common_list in common_headers.items():
                for idx, header_val in enumerate(header_row_values):
                    if header_val in common_list:
                        col_indices[key] = idx; break

            # Fallback for full name if "name" exists
            if "name" in header_row_values and "full_name" not in col_indices:
                try: col_indices["full_name"] = header_row_values.index("name")
                except ValueError: pass

            # Basic name column check
            if "first_name" not in col_indices or "last_name" not in col_indices:
                if "full_name" not in col_indices:
                    if messagebox.askyesno("Column Ambiguity", f"Could not auto-detect specific name columns in '{student_sheet_name_to_import}'.\n"
                                           "Assume Col A = First, Col B = Last? \n(No assumes Col A = 'Last, First' or 'First Last')", parent=self.root):
                        col_indices["first_name"], col_indices["last_name"] = 0, 1
                    else: col_indices["full_name"] = 0


            existing_full_names_in_app = {s['full_name'].lower().strip(): s['id'] for s in self.students.values()}
            commands_to_add_students = []
            current_id_num_for_batch = self.next_student_id_num # Use app's current next ID

            for row_idx, row_values_tuple in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                row_values = list(row_values_tuple)
                first_name, last_name, nickname, gender, group_id_to_assign = None, None, "", "Boy", None

                def get_val(key_idx, default=""):
                    idx = col_indices.get(key_idx)
                    if idx is not None and idx < len(row_values) and row_values[idx] is not None:
                        return str(row_values[idx]).strip()
                    return default

                first_name = get_val("first_name", None)
                last_name = get_val("last_name", None)
                nickname = get_val("nickname", "")
                gender_str = get_val("gender", "Boy").lower()
                if gender_str in ["girl", "female", "f"]: gender = "Girl"

                group_name_from_excel = get_val("group_name", "")
                if group_name_from_excel and self.settings.get("student_groups_enabled", True):
                    # Find group_id by name
                    for gid, gdata in self.student_groups.items():
                        if gdata.get("name", "").lower() == group_name_from_excel.lower():
                            group_id_to_assign = gid; break


                if not first_name or not last_name: # Try parsing from full_name if specific cols missing
                    if "full_name" in col_indices:
                        full_name_str = get_val("full_name", None)
                        if full_name_str:
                            if "," in full_name_str:
                                parts = full_name_str.split(",", 1)
                                last_name = parts[0].strip()
                                first_name = parts[1].strip() if len(parts) > 1 else ""
                            elif " " in full_name_str: # Assume "First Last"
                                parts = full_name_str.split(" ", 1)
                                first_name = parts[0].strip()
                                last_name = parts[1].strip() if len(parts) > 1 else ""
                            else: first_name = full_name_str # Single word as first name

                if first_name: # Must have at least a first name
                    if not last_name: last_name = "" # Ensure last_name is a string

                    full_name_display = f"{first_name} \"{nickname}\" {last_name}" if nickname else f"{first_name} {last_name}"
                    full_name_key = full_name_display.lower().strip()

                    if full_name_key not in existing_full_names_in_app:
                        student_id_str = f"student_{current_id_num_for_batch}"
                        old_next_id_for_command = current_id_num_for_batch # For AddItemCommand's undo logic
                        next_id_for_app_after_this = current_id_num_for_batch + 1

                        x_pos = 50 + (imported_student_count % 10) * (self.settings.get("default_student_box_width", DEFAULT_STUDENT_BOX_WIDTH) + 10)
                        y_pos = 50 + (imported_student_count // 10) * (self.settings.get("default_student_box_height", DEFAULT_STUDENT_BOX_HEIGHT) + 10)

                        s_data = {
                            "first_name": first_name, "last_name": last_name, "nickname": nickname,
                            "gender": gender, "full_name": full_name_display,
                            "x": x_pos, "y": y_pos, "id": student_id_str,
                            "width": self.settings.get("default_student_box_width"),
                            "height": self.settings.get("default_student_box_height"),
                            "original_next_id_num_after_add": next_id_for_app_after_this,
                            "group_id": group_id_to_assign,
                            "style_overrides": {}
                        }
                        cmd = AddItemCommand(self, student_id_str, 'student', s_data, old_next_id_for_command)
                        commands_to_add_students.append(cmd)
                        existing_full_names_in_app[full_name_key] = student_id_str # Add to check for this batch
                        imported_student_count += 1
                        current_id_num_for_batch += 1
                    else: # Student exists, maybe update their group?
                        existing_student_id = existing_full_names_in_app[full_name_key]
                        if group_id_to_assign and self.students[existing_student_id].get("group_id") != group_id_to_assign:
                            # Create an EditItemCommand to update group_id
                            old_data_snapshot = self.students[existing_student_id].copy()
                            changes = {"group_id": group_id_to_assign}
                            # No need for AddItemCommand here, it's an edit.
                            # This part could be more complex if we want to batch these edits.
                            # For now, let's assume import primarily adds new students.
                            print(f"Student {full_name_display} already exists. Group update from Excel not yet fully implemented here.")


            for cmd in commands_to_add_students:
                self.execute_command(cmd) # This will update self.next_student_id_num via AddItemCommand

            if commands_to_add_students: # If any students were added
                 self.next_student_id_num = current_id_num_for_batch # Ensure app's counter is past the last used ID

        # --- Import Incidents (Simplified - does not import detailed quiz marks yet) ---
        imported_incident_count = 0
        if import_incidents_flag:
            incident_commands_to_add = []
            for sheet_name_excel in workbook.sheetnames:
                # Try to match Excel sheet name (e.g., "FirstName_LastName") to an existing student
                matched_student_id, matched_student_first_name, matched_student_last_name = None, "", ""
                normalized_excel_sheet_name_for_match = sheet_name_excel.replace("_", " ").lower()

                for s_id_app, s_data_app in self.students.items():
                    # Check against "FirstName LastName" and "FirstName_LastName" formats
                    app_student_full_name_match = s_data_app['full_name'].lower()
                    app_student_export_format_match = f"{s_data_app['first_name']}_{s_data_app['last_name']}".lower()
                    if normalized_excel_sheet_name_for_match == app_student_full_name_match or \
                       sheet_name_excel.lower() == app_student_export_format_match:
                        matched_student_id = s_id_app
                        matched_student_first_name = s_data_app['first_name']
                        matched_student_last_name = s_data_app['last_name']
                        break

                if matched_student_id:
                    student_sheet_incidents = workbook[sheet_name_excel]
                    s_header_values = [str(cell.value).lower().strip() if cell.value else "" for cell in student_sheet_incidents[1]]
                    col_map_incidents = {}
                    try: # Basic incident columns
                        col_map_incidents["ts"] = s_header_values.index("timestamp")
                        col_map_incidents["type"] = s_header_values.index("type")
                        col_map_incidents["beh_quiz_name"] = s_header_values.index("behavior/quiz name")
                        # For quiz scores, we'd look for "Correct", "Total Qs" or specific mark type columns
                        # This part is simplified for now.
                        col_map_incidents["score_correct"] = s_header_values.index("correct") if "correct" in s_header_values else -1
                        col_map_incidents["score_total_qs"] = s_header_values.index("total qs") if "total qs" in s_header_values else -1
                        col_map_incidents["comment"] = s_header_values.index("comment")
                        col_map_incidents["day"] = s_header_values.index("day")
                    except ValueError:
                        print(f"Skipping sheet '{sheet_name_excel}' for incident import: missing expected basic headers.")
                        continue

                    for row_idx_inc, s_row_values_tuple_inc in enumerate(student_sheet_incidents.iter_rows(min_row=2, values_only=True)):
                        s_row_values_inc = list(s_row_values_tuple_inc)
                        try:
                            timestamp_str = str(s_row_values_inc[col_map_incidents["ts"]]) if col_map_incidents["ts"] < len(s_row_values_inc) and s_row_values_inc[col_map_incidents["ts"]] else None
                            log_type_str = str(s_row_values_inc[col_map_incidents["type"]]).lower() if col_map_incidents["type"] < len(s_row_values_inc) and s_row_values_inc[col_map_incidents["type"]] else "behavior"
                            behavior_quiz_name_str = str(s_row_values_inc[col_map_incidents["beh_quiz_name"]]) if col_map_incidents["beh_quiz_name"] < len(s_row_values_inc) and s_row_values_inc[col_map_incidents["beh_quiz_name"]] else ""
                            comment_str = str(s_row_values_inc[col_map_incidents["comment"]]) if col_map_incidents["comment"] < len(s_row_values_inc) and s_row_values_inc[col_map_incidents["comment"]] else ""
                            day_str = str(s_row_values_inc[col_map_incidents["day"]]) if col_map_incidents["day"] < len(s_row_values_inc) and s_row_values_inc[col_map_incidents["day"]] else ""

                            if not timestamp_str or not behavior_quiz_name_str: continue

                            parsed_dt = None
                            if isinstance(s_row_values_inc[col_map_incidents["ts"]], datetime): parsed_dt = s_row_values_inc[col_map_incidents["ts"]]
                            else:
                                try: parsed_dt = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                                except ValueError:
                                    try: parsed_dt = datetime.fromisoformat(timestamp_str)
                                    except ValueError:
                                        print(f"Skipping incident in '{sheet_name_excel}', row {row_idx_inc+2}: bad timestamp format '{timestamp_str}'")
                                        continue
                            iso_timestamp = parsed_dt.isoformat()

                            # Check for duplicates before adding
                            is_duplicate = False
                            for existing_log in self.behavior_log:
                                if existing_log["student_id"] == matched_student_id and \
                                   existing_log["timestamp"] == iso_timestamp and \
                                   existing_log["behavior"] == behavior_quiz_name_str and \
                                   existing_log.get("type","behavior") == log_type_str:
                                    is_duplicate = True; break

                            if not is_duplicate:
                                log_entry_data = {
                                    "timestamp": iso_timestamp, "student_id": matched_student_id,
                                    "student_first_name": matched_student_first_name,
                                    "student_last_name": matched_student_last_name,
                                    "behavior": behavior_quiz_name_str, "comment": comment_str,
                                    "type": log_type_str, "day": day_str
                                }
                                # Simplified quiz score import (just basic correct/total if available)
                                if log_type_str == "quiz":
                                    correct_val_imp = None
                                    total_qs_val_imp = None
                                    if col_map_incidents["score_correct"] != -1 and col_map_incidents["score_correct"] < len(s_row_values_inc) and s_row_values_inc[col_map_incidents["score_correct"]]:
                                        try: correct_val_imp = int(s_row_values_inc[col_map_incidents["score_correct"]])
                                        except ValueError: pass
                                    if col_map_incidents["score_total_qs"] != -1 and col_map_incidents["score_total_qs"] < len(s_row_values_inc) and s_row_values_inc[col_map_incidents["score_total_qs"]]:
                                        try: total_qs_val_imp = int(s_row_values_inc[col_map_incidents["score_total_qs"]])
                                        except ValueError: pass

                                    if correct_val_imp is not None and total_qs_val_imp is not None:
                                        # Store as simple score_details for now, similar to live quiz
                                        log_entry_data["score_details"] = {"correct": correct_val_imp, "total_asked": total_qs_val_imp}
                                        log_entry_data["num_questions"] = total_qs_val_imp # Assume total_asked is num_questions
                                    elif correct_val_imp is not None: # If only correct is found, log it as a raw score string
                                        log_entry_data["score"] = str(correct_val_imp)


                                cmd = LogEntryCommand(self, log_entry_data, matched_student_id, timestamp=iso_timestamp)
                                incident_commands_to_add.append(cmd)
                                imported_incident_count += 1
                        except IndexError:
                            print(f"Skipping row {row_idx_inc + 2} in '{sheet_name_excel}' for incident import: missing data columns.")
                            continue
            for cmd in incident_commands_to_add:
                self.execute_command(cmd)

        return imported_student_count, imported_incident_count

    def import_students_from_excel_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Import", "Enter password to import from Excel:"): return

        dialog = ImportExcelOptionsDialog(self.root, app_instance=self)
        if dialog.result:
            file_path, import_incidents_flag, student_sheet_name = dialog.result
            if not file_path: return
            try:
                imported_student_count, imported_incident_count = self._import_data_from_excel_logic(file_path, import_incidents_flag, student_sheet_name)
                status_msg = f"Imported {imported_student_count} new students"
                if import_incidents_flag: status_msg += f" and {imported_incident_count} new incidents"
                status_msg += ". Duplicates were skipped."
                self.update_status(status_msg)
                self.draw_all_items(check_collisions_on_redraw=True)
                self.save_data_wrapper(source="import_excel") # Save after successful import
                self.password_manager.record_activity()
            except Exception as e:
                messagebox.showerror("Import Error", f"Failed to import from Excel: {e}", parent=self.root)
                self.update_status(f"Error during Excel import: {e}")
                import traceback
                traceback.print_exc()

    def save_layout_template_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Save Layout", "Enter password to save layout template:"): return
        template_name = simpledialog.askstring("Save Layout Template", "Enter a name for this layout template:", parent=self.root)
        if template_name and template_name.strip():
            filename = "".join(c if c.isalnum() or c in (' ', '_', '-') else '_' for c in template_name.strip()) + ".json"
            file_path = os.path.join(LAYOUT_TEMPLATES_DIR, filename)
            layout_data = {
                "students": {
                    sid: {
                        "x": s["x"], "y": s["y"],
                        "width": s.get("width"), "height": s.get("height"),
                        "style_overrides": s.get("style_overrides",{}).copy(),
                        # Add name details for robust loading
                        "first_name": s.get("first_name", ""),
                        "last_name": s.get("last_name", ""),
                        "nickname": s.get("nickname", "")
                    } for sid, s in self.students.items()
                },
                "furniture": {
                    fid: {
                        "x": f["x"], "y": f["y"],
                        "width": f.get("width"), "height": f.get("height")
                    } for fid, f in self.furniture.items()
                }
            }
            try:
                self._encrypt_and_write_file(file_path, layout_data)
                self.update_status(f"Layout template '{template_name}' saved.")
            except Exception as e: messagebox.showerror("Save Error", f"Could not save layout template: {e}", parent=self.root)
        else: self.update_status("Layout template save cancelled.")
        self.password_manager.record_activity()

    def load_layout_template_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Load Layout", "Enter password to load layout template:"): return
        if not os.path.exists(LAYOUT_TEMPLATES_DIR) or not os.listdir(LAYOUT_TEMPLATES_DIR):
            messagebox.showinfo("No Templates", "No layout templates found in default folder.", parent=self.root); 
        file_path = filedialog.askopenfilename(initialdir=LAYOUT_TEMPLATES_DIR, title="Select Layout Template",
                                               filetypes=[("JSON files", "*.json"), ("All files", "*.*")], parent=self.root)
        if file_path:
            try:
                template_data = self._read_and_decrypt_file(file_path)
                if not isinstance(template_data, dict):
                    raise json.JSONDecodeError("Invalid template format.", "", 0)

                if messagebox.askyesno("Confirm Load", "Loading this template will overwrite current item positions and sizes. Student data (names, logs) will be preserved. Continue?", parent=self.root):
                    # ... (rest of the logic remains the same)
                    move_commands_data = []
                    size_commands_data = []
                    template_students = template_data.get("students", {})
                    template_furniture = template_data.get("furniture", {})

                    applied_count = 0
                    skipped_count = 0
                    name_match_log = []
                    match_by_name = messagebox.askyesno("Layout Loading Options", "Load layout template by names of students (doesn't need to be exact) or by ID (not preferred-doesn't preserve student positions correctly)?\nYes is by names, no is by ID.")
                    for template_student_id, t_stud_data in template_students.items():
                        target_student_id = None
                        s_current = None
                        if match_by_name:
                            # 2. Secondary Match: Name (first, last, then nickname for disambiguation)
                            t_first = t_stud_data.get("first_name", "").lower()
                            t_last = t_stud_data.get("last_name", "").lower()
                            t_nick = t_stud_data.get("nickname", "").lower()

                            if not t_first or not t_last: # Cannot match by name if essential parts are missing
                                name_match_log.append(f"Skipped template student (ID: {template_student_id}, Name: {t_stud_data.get('full_name', 'N/A')}) due to missing name components in template.")
                                skipped_count +=1
                                continue

                            potential_matches = []
                            for c_sid, c_sdata in self.students.items():
                                if c_sdata.get("first_name", "").lower() == t_first and \
                                   c_sdata.get("last_name", "").lower() == t_last:
                                    potential_matches.append(c_sid)

                            if len(potential_matches) == 1:
                                target_student_id = potential_matches[0]
                                s_current = self.students[target_student_id]
                                name_match_log.append(f"Matched template's {t_stud_data.get('first_name')} {t_stud_data.get('last_name')} to classroom's {s_current['full_name']} by name.")
                            elif len(potential_matches) > 1:
                                # Attempt disambiguation with nickname
                                if t_nick:
                                    final_matches = [pid for pid in potential_matches if self.students[pid].get("nickname","").lower() == t_nick]
                                    if len(final_matches) == 1:
                                        target_student_id = final_matches[0]
                                        s_current = self.students[target_student_id]
                                        name_match_log.append(f"Matched template's {t_stud_data.get('first_name')} {t_stud_data.get('last_name')} ({t_nick}) to classroom's {s_current['full_name']} by exact name & nickname.")
                                    else: # No exact nickname match, or multiple after filtering by nickname
                                        name_match_log.append(f"Ambiguous exact name match for template's {t_stud_data.get('first_name')} {t_stud_data.get('last_name')} (Nickname: {t_nick}). Found {len(potential_matches)} with same first/last, {len(final_matches)} after nickname filter. Trying fuzzy match.")
                                        # Proceed to fuzzy matching for these potential_matches if final_matches was not unique
                                        potential_matches_for_fuzzy = final_matches if t_nick and final_matches else potential_matches
                                        # Fall through to fuzzy matching logic below if no unique exact match yet
                                else: # No nickname in template to disambiguate exact first/last name matches
                                    name_match_log.append(f"Ambiguous exact name match for template's {t_stud_data.get('first_name')} {t_stud_data.get('last_name')}. Found {len(potential_matches)} classroom students. Trying fuzzy match.")
                                    # Fall through to fuzzy matching logic below

                            # Fuzzy Matching Stage (if no unique exact match by ID or full name + nickname)
                            if not target_student_id: # Only if we haven't found a target yet
                                fuzzy_matches = []
                                # If potential_matches had some exact first/last name hits, fuzzy match within that subset first
                                students_to_search_fuzzy = [self.students[pid] for pid in potential_matches] if potential_matches else list(self.students.values())

                                for c_sdata_fuzzy in students_to_search_fuzzy:
                                    # Construct full names for comparison
                                    template_full_name_for_fuzzy = f"{t_first} {t_last}"
                                    classroom_full_name_for_fuzzy = f"{c_sdata_fuzzy.get('first_name','').lower()} {c_sdata_fuzzy.get('last_name','').lower()}"

                                    similarity = name_similarity_ratio(template_full_name_for_fuzzy, classroom_full_name_for_fuzzy)

                                    if similarity >= 0.85: # Similarity threshold
                                        fuzzy_matches.append({"id": c_sdata_fuzzy["id"], "similarity": similarity, "data": c_sdata_fuzzy})

                                if fuzzy_matches:
                                    fuzzy_matches.sort(key=lambda x: x["similarity"], reverse=True) # Sort by best match

                                    if len(fuzzy_matches) == 1 or fuzzy_matches[0]["similarity"] > fuzzy_matches[1]["similarity"] + 0.05: # Unique best fuzzy match or significantly better
                                        best_fuzzy_match = fuzzy_matches[0]
                                        target_student_id = best_fuzzy_match["id"]
                                        s_current = self.students[target_student_id]
                                        name_match_log.append(f"Fuzzy matched template's {t_stud_data.get('first_name')} {t_stud_data.get('last_name')} to classroom's {s_current['full_name']} (Similarity: {best_fuzzy_match['similarity']:.2f}).")
                                    else: # Multiple good fuzzy matches, try nickname disambiguation again
                                        if t_nick:
                                            final_fuzzy_nick_matches = [fm for fm in fuzzy_matches if fm["data"].get("nickname","").lower() == t_nick and fm["similarity"] >=0.85]
                                            if len(final_fuzzy_nick_matches) == 1:
                                                target_student_id = final_fuzzy_nick_matches[0]["id"]
                                                s_current = self.students[target_student_id]
                                                name_match_log.append(f"Fuzzy matched (with nickname) template's {t_stud_data.get('first_name')} {t_stud_data.get('last_name')} ({t_nick}) to classroom's {s_current['full_name']} (Similarity: {final_fuzzy_nick_matches[0]['similarity']:.2f}).")
                                            else:
                                                name_match_log.append(f"Ambiguous fuzzy match for template's {t_stud_data.get('first_name')} {t_stud_data.get('last_name')} ({t_nick}) after nickname. Skipped.")
                                                skipped_count += 1
                                        else:
                                            name_match_log.append(f"Ambiguous fuzzy match for template's {t_stud_data.get('first_name')} {t_stud_data.get('last_name')}. Skipped.")
                                            skipped_count += 1
                                elif not potential_matches : # Only log "no match" if there were no exact first/last name potential_matches initially
                                    name_match_log.append(f"No ID, exact name, or close fuzzy match for template student {t_stud_data.get('first_name')} {t_stud_data.get('last_name')} (ID: {template_student_id}). Skipped.")
                                    skipped_count += 1

                            # If after all matching attempts, still no target_student_id
                            if not target_student_id and not potential_matches : # Redundant check for skipped_count already done by fuzzy logic.
                                # This log might be duplicated if fuzzy also logged a skip.
                                # name_match_log.append(f"Final skip for template student {t_stud_data.get('first_name')} {t_stud_data.get('last_name')} (ID: {template_student_id}).")
                                # skipped_count +=1 # This might double count skips if fuzzy already counted it.
                                pass
                        else:   # 2. ID Match
                            if template_student_id in self.students:
                                target_student_id = template_student_id
                                s_current = self.students[target_student_id]

                        # If a student was found (either by ID, exact name, or fuzzy name)
                        if target_student_id and s_current:
                            applied_count +=1
                            # Position
                            old_x, old_y = s_current["x"], s_current["y"]
                            new_x, new_y = t_stud_data.get("x", old_x), t_stud_data.get("y", old_y)
                            if old_x != new_x or old_y != new_y:
                                move_commands_data.append({'id':target_student_id, 'type':'student', 'old_x':old_x, 'old_y':old_y, 'new_x':new_x, 'new_y':new_y})
                            
                            # Size
                            old_w = s_current.get("style_overrides",{}).get("width", s_current.get("width", DEFAULT_STUDENT_BOX_WIDTH))
                            old_h = s_current.get("style_overrides",{}).get("height", s_current.get("height", DEFAULT_STUDENT_BOX_HEIGHT))
                            new_w = t_stud_data.get("width", old_w)
                            new_h = t_stud_data.get("height", old_h)
                            if old_w != new_w or old_h != new_h:
                                size_commands_data.append({'id':target_student_id, 'type':'student', 'old_w':old_w, 'old_h':old_h, 'new_w':new_w, 'new_h':new_h})
                            
                            # Style Overrides
                            t_style_overrides = t_stud_data.get("style_overrides", {})
                            if t_style_overrides or (not t_style_overrides and s_current.get("style_overrides")): # Apply if template has styles OR if current has styles that need clearing
                                current_style_snapshot = s_current.get("style_overrides", {}).copy()

                                # Create a snapshot of the full student data before style change for EditItemCommand
                                full_old_student_data_for_style_cmd = s_current.copy()
                                full_old_student_data_for_style_cmd["style_overrides"] = current_style_snapshot

                                # The new_item_data_changes for EditItemCommand needs to be just the changes.
                                # Here, we are replacing the entire style_overrides dict from the template.
                                if current_style_snapshot != t_style_overrides:
                                     self.execute_command(EditItemCommand(self, target_student_id, "student", full_old_student_data_for_style_cmd, {"style_overrides": t_style_overrides.copy()}))


                    # Furniture (still by ID)
                    for item_id, t_data in template_furniture.items():
                         if item_id in self.furniture:
                            f_current = self.furniture[item_id]
                            old_x, old_y = f_current["x"], f_current["y"]
                            new_x, new_y = t_data.get("x", old_x), t_data.get("y", old_y)
                            if old_x != new_x or old_y != new_y : move_commands_data.append({'id':item_id, 'type':'furniture', 'old_x':old_x, 'old_y':old_y, 'new_x':new_x, 'new_y':new_y})

                            old_w = f_current.get("width", REBBI_DESK_WIDTH) ; old_h = f_current.get("height", REBBI_DESK_HEIGHT)
                            new_w = t_data.get("width", old_w); new_h = t_data.get("height", old_h)
                            if old_w != new_w or old_h != new_h: size_commands_data.append({'id':item_id, 'type':'furniture', 'old_w':old_w, 'old_h':old_h, 'new_w':new_w, 'new_h':new_h})

                    if move_commands_data: self.execute_command(MoveItemsCommand(self, move_commands_data))
                    if size_commands_data: self.execute_command(ChangeItemsSizeCommand(self, size_commands_data))

                    status_message = f"Layout '{os.path.basename(file_path)}' loaded. Applied to {applied_count} students."
                    if skipped_count > 0:
                        status_message += f" Skipped {skipped_count} template students (see console log for details)."
                    if name_match_log:
                        print("--- Layout Load Name Matching Log ---")
                        for log_line in name_match_log: print(log_line)
                        print("------------------------------------")

                    self.update_status(status_message)
                    self.draw_all_items(check_collisions_on_redraw=True)
                    self.save_data_wrapper(source="load_template")
            except (json.JSONDecodeError, IOError) as e: messagebox.showerror("Load Error", f"Could not load layout template: {e}", parent=self.root)
        else: self.update_status("Layout template load cancelled.")
        self.password_manager.record_activity()
    
    def generate_attendance_report_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Generate Report", "Enter password to generate attendance report:"): return

        dialog = AttendanceReportDialog(self.root, self.students)
        if dialog.result:
            start_date, end_date, selected_student_ids = dialog.result
            if not selected_student_ids:
                messagebox.showinfo("No Students", "No students selected for the report.", parent=self.root)
                return

            report_data = self.generate_attendance_data(start_date, end_date, selected_student_ids)
            if not report_data:
                messagebox.showinfo("No Data", "No attendance-relevant log data found for the selected criteria.", parent=self.root)
                return

            default_filename = f"attendance_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_filename,
                                                   filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")], parent=self.root)
            if file_path:
                try:
                    self.export_attendance_to_excel(file_path, report_data, start_date, end_date)
                    self.update_status(f"Attendance report saved to {os.path.basename(file_path)}.")
                    if messagebox.askyesno("Export Successful", f"Attendance report saved to:\n{file_path}\n\nDo you want to open the file location?", parent=self.root):
                        self.open_specific_export_folder(file_path)
                except Exception as e:
                    messagebox.showerror("Export Error", f"Failed to save attendance report: {e}", parent=self.root)
            else:
                self.update_status("Attendance report export cancelled.")
        self.password_manager.record_activity()

    def generate_attendance_data(self, start_date, end_date, student_ids):
        attendance = {} # {date_obj: {student_id: "Present"}}
        all_logs = self.behavior_log + self.homework_log # Combine logs for presence check

        current_date = start_date
        while current_date <= end_date:
            attendance[current_date] = {}
            for student_id in student_ids:
                # Check if any log entry exists for this student on this date
                present = any(
                    log["student_id"] == student_id and
                    datetime.fromisoformat(log["timestamp"]).date() == current_date
                    for log in all_logs
                )
                attendance[current_date][student_id] = "P" if present else "A" # Present / Absent
            current_date += timedelta(days=1)
        return attendance

    def export_attendance_to_excel(self, file_path, attendance_data, report_start_date, report_end_date):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        headers = ["Student Name"]
        date_columns_map = {} # date_obj -> column_index
        current_col = 2
        d_iter = report_start_date
        while d_iter <= report_end_date:
            headers.append(d_iter.strftime("%Y-%m-%d (%a)"))
            date_columns_map[d_iter] = current_col
            current_col += 1
        headers.append("Total Present")
        headers.append("Total Absent")

        for col_num, header_title in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header_title).font = OpenpyxlFont(bold=True)
            ws.column_dimensions[get_column_letter(col_num)].width = 15 if col_num > 1 else 25
        ws.freeze_panes = 'B2'

        current_row = 2
        sorted_student_ids = sorted(
            list(set(sid for day_data in attendance_data.values() for sid in day_data.keys())),
            key=lambda sid: (self.students.get(sid, {}).get("last_name", ""), self.students.get(sid, {}).get("first_name", ""))
        )

        for student_id in sorted_student_ids:
            student_name = self.students.get(student_id, {}).get("full_name", student_id)
            ws.cell(row=current_row, column=1, value=student_name)
            total_present = 0
            total_absent = 0
            for date_obj, col_idx in date_columns_map.items():
                status = attendance_data.get(date_obj, {}).get(student_id, "A")
                ws.cell(row=current_row, column=col_idx, value=status).alignment = OpenpyxlAlignment(horizontal='center')
                if status == "P": total_present += 1
                else: total_absent += 1
            ws.cell(row=current_row, column=len(headers)-1, value=total_present).alignment = OpenpyxlAlignment(horizontal='center')
            ws.cell(row=current_row, column=len(headers), value=total_absent).alignment = OpenpyxlAlignment(horizontal='center')
            current_row += 1
        wb.save(file_path)

    def align_selected_items(self, edge):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Align", "Enter password to align items:"): return
        if len(self.selected_items) < 2:
            self.update_status("Select at least two items to align."); return

        items_data_for_align = []
        for item_id in self.selected_items:
            item_type = "student" if item_id in self.students else "furniture"
            data_source = self.students if item_type == "student" else self.furniture
            if item_id in data_source:
                item = data_source[item_id]
                items_data_for_align.append({
                    "id": item_id, "type": item_type, "x": item["x"], "y": item["y"],
                    "width": item.get("_current_world_width", item.get("width", DEFAULT_STUDENT_BOX_WIDTH)), # Use dynamic if available
                    "height": item.get("_current_world_height", item.get("height", DEFAULT_STUDENT_BOX_HEIGHT))
                })
        if not items_data_for_align: return

        target_coord = 0
        if edge == "left": target_coord = min(item["x"] for item in items_data_for_align)
        elif edge == "right": target_coord = max(item["x"] + item["width"] for item in items_data_for_align)
        elif edge == "top": target_coord = min(item["y"] for item in items_data_for_align)
        elif edge == "bottom": target_coord = max(item["y"] + item["height"] for item in items_data_for_align)
        elif edge == "center_h": # Align to average horizontal center of the selection box
            min_x = min(it["x"] for it in items_data_for_align)
            max_x_br = max(it["x"] + it["width"] for it in items_data_for_align)
            target_coord = (min_x + max_x_br) / 2 # This is the center of the bounding box of selected items
        elif edge == "center_v": # Align to average vertical center
            min_y = min(it["y"] for it in items_data_for_align)
            max_y_br = max(it["y"] + it["height"] for it in items_data_for_align)
            target_coord = (min_y + max_y_br) / 2

        move_commands_for_align = []
        for item_to_align in items_data_for_align:
            old_x_align, old_y_align = item_to_align["x"], item_to_align["y"]
            new_x_align, new_y_align = old_x_align, old_y_align
            if edge == "left": new_x_align = target_coord
            elif edge == "right": new_x_align = target_coord - item_to_align["width"]
            elif edge == "top": new_y_align = target_coord
            elif edge == "bottom": new_y_align = target_coord - item_to_align["height"]
            elif edge == "center_h": new_x_align = target_coord - item_to_align["width"] / 2
            elif edge == "center_v": new_y_align = target_coord - item_to_align["height"] / 2

            if abs(new_x_align - old_x_align) > 0.01 or abs(new_y_align - old_y_align) > 0.01:
                move_commands_for_align.append({'id': item_to_align["id"], 'type': item_to_align["type"], 'old_x': old_x_align, 'old_y': old_y_align, 'new_x': new_x_align, 'new_y': new_y_align})

        if move_commands_for_align:
            self.execute_command(MoveItemsCommand(self, move_commands_for_align))
            self.update_status(f"Aligned {len(move_commands_for_align)} items to {edge}.")
        else: self.update_status("Items already aligned."); self.draw_all_items(check_collisions_on_redraw=True)
        self.password_manager.record_activity()

    def distribute_selected_items_evenly(self, direction='horizontal'):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Distribute", "Enter password to distribute items:"): return

        if len(self.selected_items) < 2:
            self.update_status("Select at least two items to distribute.")
            return

        items_to_distribute = []
        for item_id in self.selected_items:
            item_data = None
            item_type = None
            default_width = DEFAULT_STUDENT_BOX_WIDTH
            default_height = DEFAULT_STUDENT_BOX_HEIGHT

            if item_id in self.students:
                item_data = self.students[item_id]
                item_type = "student"
            elif item_id in self.furniture:
                item_data = self.furniture[item_id]
                item_type = "furniture"
                # Furniture might have different defaults, but let's assume student defaults for now if not specified.
                # Or better, use specific furniture defaults if available.
                default_width = REBBI_DESK_WIDTH
                default_height = REBBI_DESK_HEIGHT


            if item_data:
                # Use _current_world_width/height if available (from draw_single_student/furniture)
                # otherwise fallback to item's own width/height or defaults.
                width = item_data.get('_current_world_width', item_data.get('width', default_width))
                height = item_data.get('_current_world_height', item_data.get('height', default_height))

                # For students, width/height might be in style_overrides
                if item_type == "student":
                    style_overrides = item_data.get("style_overrides", {})
                    width = style_overrides.get("width", width)
                    height = style_overrides.get("height", height)

                items_to_distribute.append({
                    "id": item_id,
                    "type": item_type,
                    "x": float(item_data["x"]),
                    "y": float(item_data["y"]),
                    "width": float(width),
                    "height": float(height),
                })

        if not items_to_distribute:
            return

        moves_for_command = []

        if direction == 'horizontal':
            items_to_distribute.sort(key=lambda item: item['x'])

            min_x_overall = items_to_distribute[0]['x']
            # Max x-coordinate is the x of the rightmost item's right edge
            max_x_item_overall = items_to_distribute[-1]
            max_x_coord_overall = max_x_item_overall['x'] + max_x_item_overall['width']

            total_items_width = sum(item['width'] for item in items_to_distribute)
            total_span = max_x_coord_overall - min_x_overall

            if len(items_to_distribute) > 1:
                available_space_for_gaps = total_span - total_items_width
                # Prevent negative gap if items overlap significantly; ensure a minimal positive gap or zero.
                gap_size = max(0, available_space_for_gaps / (len(items_to_distribute) - 1))
            else:
                return # Should be caught by len < 2 check

            current_x = min_x_overall # Start placing the first item at its original position (or the leftmost edge)
            for i, item in enumerate(items_to_distribute):
                # Only create a move command if the item's position actually changes
                if abs(item['x'] - current_x) > 0.01: # Using a small tolerance for float comparison
                    moves_for_command.append({
                        'id': item['id'], 'type': item['type'],
                        'old_x': item['x'], 'old_y': item['y'],
                        'new_x': current_x, 'new_y': item['y'] # Keep original y
                    })
                current_x += item['width'] + gap_size

        elif direction == 'vertical':
            items_to_distribute.sort(key=lambda item: item['y'])

            min_y_overall = items_to_distribute[0]['y']
            max_y_item_overall = items_to_distribute[-1]
            max_y_coord_overall = max_y_item_overall['y'] + max_y_item_overall['height']

            total_items_height = sum(item['height'] for item in items_to_distribute)
            total_span = max_y_coord_overall - min_y_overall

            if len(items_to_distribute) > 1:
                available_space_for_gaps = total_span - total_items_height
                gap_size = max(0, available_space_for_gaps / (len(items_to_distribute) - 1))
            else:
                return

            current_y = min_y_overall
            for i, item in enumerate(items_to_distribute):
                if abs(item['y'] - current_y) > 0.01:
                    moves_for_command.append({
                        'id': item['id'], 'type': item['type'],
                        'old_x': item['x'], 'old_y': item['y'],
                        'new_x': item['x'], 'new_y': current_y # Keep original x
                    })
                current_y += item['height'] + gap_size

        if moves_for_command:
            self.execute_command(MoveItemsCommand(self, moves_for_command))
            self.update_status(f"Distributed {len(moves_for_command)} items {direction}ly.")
        else:
            self.update_status(f"Items already distributed {direction}ly or no change needed.")
        self.password_manager.record_activity()

    def assign_student_to_group_via_menu(self, student_id, group_id):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Assign Group", "Enter password to assign group:"): return
        student = self.students.get(student_id)
        if not student: return

        old_group_id = student.get("group_id")
        new_group_id_val = None if group_id == "NONE_GROUP_SENTINEL" else group_id

        if old_group_id != new_group_id_val:
            # For undo/redo, we need snapshots of all group assignments if ManageStudentGroupCommand is used broadly.
            # For a single assignment, EditItemCommand is simpler.
            old_student_data_snapshot = student.copy()
            if "style_overrides" in old_student_data_snapshot: old_student_data_snapshot["style_overrides"] = old_student_data_snapshot["style_overrides"].copy()
            
            changes_for_command = {"group_id": new_group_id_val}
            self.execute_command(EditItemCommand(self, student_id, "student", old_student_data_snapshot, changes_for_command))
            # Command will call draw_all_items, which will redraw the student
            group_name = self.student_groups[new_group_id_val]['name'] if new_group_id_val and new_group_id_val in self.student_groups else "No Group"
            self.update_status(f"Assigned {student['full_name']} to group: {group_name}.")
            self.save_data_wrapper(source="assign_group_menu") # Save student data which now includes group_id
        self.password_manager.record_activity()

    def assign_students_to_group_via_menu(self, student_ids, group_id):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Assign Group", "Enter password to assign group:"): return
        
        for student in student_ids:
            student_id = student
            student = self.students.get(student_id)
            if not student: return

            old_group_id = student.get("group_id")
            new_group_id_val = None if group_id == "NONE_GROUP_SENTINEL" else group_id

            if old_group_id != new_group_id_val:
                # For undo/redo, we need snapshots of all group assignments if ManageStudentGroupCommand is used broadly.
                # For a single assignment, EditItemCommand is simpler.
                old_student_data_snapshot = student.copy()
                if "style_overrides" in old_student_data_snapshot: old_student_data_snapshot["style_overrides"] = old_student_data_snapshot["style_overrides"].copy()
                
                changes_for_command = {"group_id": new_group_id_val}
                self.execute_command(EditItemCommand(self, student_id, "student", old_student_data_snapshot, changes_for_command))
                # Command will call draw_all_items, which will redraw the student
                group_name = self.student_groups[new_group_id_val]['name'] if new_group_id_val and new_group_id_val in self.student_groups else "No Group"
                self.update_status(f"Assigned {student['full_name']} to group: {group_name}.")
                self.save_data_wrapper(source="assign_group_menu") # Save student data which now includes group_id
        self.password_manager.record_activity()

    def manage_student_groups_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Manage Groups", "Enter password to manage student groups:"): return
        # Take snapshots for ManageStudentGroupCommand
        old_groups_snap = {gid: gdata.copy() for gid, gdata in self.student_groups.items()}
        old_student_assignments_snap = {sid: sdata.get("group_id") for sid, sdata in self.students.items() if "group_id" in sdata and sdata["group_id"] is not None}
        old_next_group_id_num_snap = self.next_group_id_num

        dialog = ManageStudentGroupsDialog(self.root, self.student_groups, self.students, self, default_colors=DEFAULT_GROUP_COLORS)
        if dialog.groups_changed_flag: # Check if dialog indicated changes were made
            # The dialog directly modifies self.student_groups, self.students (group_id), and self.next_group_id_num
            new_groups_snap = {gid: gdata.copy() for gid, gdata in self.student_groups.items()} # Current state after dialog
            new_student_assignments_snap = {sid: sdata.get("group_id") for sid, sdata in self.students.items() if "group_id" in sdata and sdata["group_id"] is not None}
            new_next_group_id_num_snap = self.next_group_id_num

            cmd = ManageStudentGroupCommand(self, old_groups_snap, new_groups_snap,
                                            old_student_assignments_snap, new_student_assignments_snap,
                                            old_next_group_id_num_snap, new_next_group_id_num_snap)
            # Execute command will apply the new state (which is already set by dialog) and handle saving/drawing
            # The command's execute() will effectively re-apply what the dialog did, which is fine.
            # The crucial part is that undo() will restore the old snapshots.
            self.execute_command(cmd) # This will also call save_student_groups and draw_all_items
            self.update_status("Student groups updated via dialog.") # Status from command is more generic
        else:
            self.update_status("Manage student groups cancelled or no changes made.")
        self.password_manager.record_activity()

    def toggle_student_groups_ui_visibility(self):
        enabled = self.settings.get("student_groups_enabled", True)
        if hasattr(self, 'manage_groups_btn'):
            self.manage_groups_btn.config(state=tk.NORMAL if enabled else tk.DISABLED)
        self.draw_all_items(check_collisions_on_redraw=False) # Redraw to show/hide indicators

    def toggle_manage_boxes_visibility(self):
        if self.edit_mode_var.get() or self.settings.get("always_show_box_management", False): self.top_controls_frame_row2.pack(side=tk.TOP, fill=tk.X, pady=(2, 5)); self.top_frame.height_adjusted = 110
        else: self.top_controls_frame_row2.pack_forget(); self.top_frame.height_adjusted = 50
            
    def manage_quiz_templates_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Manage Templates", "Enter password to manage quiz templates:"): return
        dialog = ManageQuizTemplatesDialog(self.root, self) # Pass app instance
        if dialog.templates_changed_flag:
            self.save_quiz_templates() # Dialog modifies self.quiz_templates directly for now
            self.update_status("Quiz templates updated.")
        else:
            self.update_status("Quiz template management cancelled or no changes made.")
        self.password_manager.record_activity()

    def manage_homework_templates_dialog(self): # New
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Manage Templates", "Enter password to manage homework templates:"): return
        dialog = ManageHomeworkTemplatesDialog(self.root, self) # Pass app instance
        if dialog.templates_changed_flag:
            self.save_homework_templates() # Dialog modifies self.homework_templates directly
            self.update_status("Homework templates updated.")
        else:
            self.update_status("Homework template management cancelled or no changes made.")
        self.password_manager.record_activity()
    
    def set_theme(self, theme, canvas_color):
        self.theme_style_using = theme
        self.theme_set()
        
        if canvas_color == "Default" or canvas_color == "" or canvas_color == None:
            canvas_color = None; self.custom_canvas_color = None
        else:
            self.custom_canvas_color = canvas_color
            self.canvas_color = canvas_color
        
        if self.custom_canvas_color: self.canvas_color = self.custom_canvas_color
        elif self.theme_style_using == "Dark": self.canvas_color = "#1F1F1F"
        elif self.theme_style_using == "System": self.canvas_color = "lightgrey" if darkdetect.theme() == "Light" else "#1F1F1F"
        else: self.canvas_color = "lightgrey"
        self.canvas.configure(bg=self.canvas_color)
    
    def _apply_canvas_color(self):
        """Applies the current canvas color based on theme and custom settings."""
        if self.custom_canvas_color and self.custom_canvas_color != "Default":
            self.canvas_color = self.custom_canvas_color
        elif self.theme_style_using == "Dark":
            self.canvas_color = "#1F1F1F"
        elif self.theme_style_using == "System":
            self.canvas_color = "lightgrey" if darkdetect.theme() == "Light" else "#1F1F1F"
        else: # Light theme
            self.canvas_color = "lightgrey"
        
        if hasattr(self, 'canvas') and self.canvas:
            self.canvas.configure(bg=self.canvas_color)
    
    def theme_set(self, theme=None): 
        if self.type_theme == "sv_ttk":
            if self.theme_style_using == "System":
                sv_ttk.set_theme(darkdetect.theme())
            else:
                if self.theme_style_using.lower() == "light" or self.theme_style_using.lower() == "dark":
                    sv_ttk.set_theme(self.theme_style_using)
                else:
                    self.theme_style_using = "Light"
                    sv_ttk.set_theme(self.theme_style_using)
        else:
            style = ttk.Style(self.root)
            
            style.theme_use(self.type_theme if "sun-valley" not in self.type_theme else f"{self.type_theme[:10]}-{self.theme_style_using.lower()}")
    
    def theme_auto(self, init=False):
        self.theme_set()
        if self.custom_canvas_color != "Default" and self.custom_canvas_color != None: self.canvas_color = self.custom_canvas_color
        elif self.theme_style_using == "Dark": self.canvas_color = "#1F1F1F"
        elif self.theme_style_using == "System": self.canvas_color = "lightgrey" if darkdetect.theme() == "Light" else "#1F1F1F"
        else: self.canvas_color = "lightgrey"
        
        if not init == True:
            print(init, self.canvas_color, self.custom_canvas_color)
            self.canvas.configure(bg=self.canvas_color) # type: ignore

    def open_settings_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Open Settings", "Enter password to open settings:"): return
        style = ttk.Style(self.root)
        self.styles = style.theme_names()
        # Store a copy of settings for potential revert or detailed change tracking for undo (complex)
        # For now, settings dialog applies changes directly and saves.
        # old_settings_snapshot = self.settings.copy() # For a potential future SettingsChangeCommand
        
        # Unbind root so that shorcuts can work in settings
        self.root.unbind_all("<Control-z>")
        self.root.unbind_all("<Control-y>")
        self.root.unbind_all("<Control-Shift-Z>")
        
        dialog = SettingsDialog(self.root, self.settings, self.custom_behaviors, self.all_behaviors, self,
                                self.custom_homework_statuses, self.all_homework_statuses, # Homework log behaviors
                                self.custom_homework_types, self.all_homework_session_types, # Homework session types (Yes/No mode)
                                self.password_manager, self.theme_style_using, self.custom_canvas_color, self.styles, self.type_theme)
        if dialog.settings_changed_flag: # Check if dialog indicated changes
            # Settings are applied directly by the dialog for most parts
            self.save_data_wrapper(source="settings_dialog") # Save all data as settings are part of it
            self.update_all_behaviors(); self.update_all_homework_log_behaviors(); self.update_all_homework_session_types()
            self.guide_line_color = self.settings.get("guides_color", "blue")
            self.draw_all_items(check_collisions_on_redraw=True)
            self.update_status("Settings updated.")
            self._update_toggle_dragging_button_text()
            self.update_zoom_display()
            self.update_lock_button_state()
            self.toggle_student_groups_ui_visibility()
            self.set_theme(self.theme_style_using, self.custom_canvas_color)
            self.toggle_manage_boxes_visibility()
            
            # Re-schedule autosave if interval changed
            self.root.after_cancel(self.autosave_data_wrapper) # Cancel existing if any (might need to store the after_id)
            self.root.after(self.settings.get("autosave_interval_ms", 30000), self.autosave_data_wrapper)
        else:
            try:
                self.update_status("Settings dialog closed, no changes applied through dialog confirm.")
            except: pass
        
        # Rebind root after settings closes
        self.root.bind_all("<Control-z>", lambda event: self.undo_last_action())
        self.root.bind_all("<Control-y>", lambda event: self.redo_last_action())
        self.root.bind_all("<Control-Shift-Z>", lambda event: self.redo_last_action()) # Common alternative for redo        

        self.password_manager.record_activity()

    def reset_application_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Reset Application", "Enter password to reset application:"): return
        msg = "This will reset ALL application data including students, furniture, logs, settings, custom behaviors, templates, and groups. This action CANNOT be undone.\n\nAre you absolutely sure you want to reset the application to its default state?"
        if messagebox.askyesno("Confirm Application Reset", msg, icon='error', parent=self.root, default=messagebox.NO):
            # Second, more direct confirmation
            if messagebox.askyesno("Final Confirmation", "Really reset everything? This is your last chance to cancel.", icon='error', parent=self.root, default=messagebox.NO):
                self._perform_reset()
            else: self.update_status("Application reset cancelled (final prompt).")
        else: self.update_status("Application reset cancelled.")
        self.password_manager.record_activity()

    def _perform_reset(self):
        try:
            self.backup_all_data_dialog(force=True)
        except Exception as e:
            print(e)
        try:
            # Clear current data in memory
            self.students.clear(); self.furniture.clear(); self.behavior_log.clear(); self.homework_log.clear()
            self.student_groups.clear(); self.quiz_templates.clear(); self.homework_templates.clear()
            self.custom_behaviors.clear(); self.custom_homework_statuses.clear(); #self.custom_homework_session_types.clear()
            self.undo_stack.clear(); self.redo_stack.clear()
            self._per_student_last_cleared.clear()
            self.last_excel_export_path = None
            self.settings = self._get_default_settings() # Reset to defaults
            self._ensure_next_ids() # Reset ID counters based on default settings
            self.password_manager = PasswordManager(self.settings) # Reset password manager with fresh settings
            self.guides.clear()
            # Delete data files
            files_to_delete = [
                DATA_FILE, CUSTOM_BEHAVIORS_FILE, 
                CUSTOM_HOMEWORK_TYPES_FILE, # NEW
                CUSTOM_HOMEWORK_STATUSES_FILE, # RENAMED
                STUDENT_GROUPS_FILE, QUIZ_TEMPLATES_FILE, HOMEWORK_TEMPLATES_FILE,
                AUTOSAVE_EXCEL_FILE
            ]
            # Attempt to delete old version files if they exist from previous versions
            for i in range(1, int(CURRENT_DATA_VERSION_TAG[1:])):
                files_to_delete.append(get_app_data_path(f"classroom_data_v{i}.json"))
                files_to_delete.append(get_app_data_path(f"custom_behaviors_v{i}.json"))
                files_to_delete.append(get_app_data_path(f"custom_homeworks_v{i}.json"))
                files_to_delete.append(get_app_data_path(f"student_groups_v{i}.json"))
                files_to_delete.append(get_app_data_path(f"quiz_templates_v{i}.json"))
                files_to_delete.append(get_app_data_path(f"homework_templates_v{i}.json"))
                files_to_delete.append(get_app_data_path(f"custom_homework_session_types_v{i}.json"))
                files_to_delete.append(get_app_data_path(f"autosave_log_v{i}.xlsx"))


            for f_path in files_to_delete:
                if os.path.exists(f_path):
                    try: os.remove(f_path)
                    except OSError as e: print(f"Warning: Could not delete file {f_path} during reset: {e}")
            
            # Delete layout templates directory contents
            if os.path.exists(LAYOUT_TEMPLATES_DIR):
                for item_name in os.listdir(LAYOUT_TEMPLATES_DIR):
                    item_path = os.path.join(LAYOUT_TEMPLATES_DIR, item_name)
                    try:
                        if os.path.isfile(item_path) or os.path.islink(item_path): os.unlink(item_path)
                        elif os.path.isdir(item_path): shutil.rmtree(item_path)
                    except Exception as e_del_layout: print(f"Failed to delete {item_path}: {e_del_layout}")


            # Save fresh default data (which will create new empty files)
            self.save_data_wrapper(source="reset")
            self.update_all_behaviors(); self.update_all_homework_statuses(); self.update_all_homework_session_types()
            self.draw_all_items(check_collisions_on_redraw=True)
            self.update_undo_redo_buttons_state()
            self.update_lock_button_state()
            self.update_status("Application has been reset to default state.")
            messagebox.showinfo("Reset Complete", "Application reset successfully. All data and settings are now at their defaults.", parent=self.root)
        except Exception as e:
            messagebox.showerror("Reset Error", f"An error occurred during application reset: {e}", parent=self.root)
            self.update_status(f"Error during reset: {e}")

    def backup_all_data_dialog(self, force=False):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Backup", "Enter password to create a backup:"): return
        default_filename = f"{APP_NAME}_Backup_{CURRENT_DATA_VERSION_TAG}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        if force == False:
            backup_zip_path = filedialog.asksaveasfilename(
                title="Save Backup As",
                defaultextension=".zip",
                initialfile=default_filename,
                filetypes=[("ZIP archive", "*.zip")],
                parent=self.root
            )
            if not backup_zip_path:
                self.update_status("Backup cancelled."); return
        elif force == True:
            backup_zip_path = os.path.abspath(os.path.join(os.path.dirname(DATA_FILE), default_filename))
        # Ensure latest data is saved before backup
        self.save_data_wrapper(source="backup_preparation")

        files_to_backup = [
            DATA_FILE, CUSTOM_BEHAVIORS_FILE, 
            CUSTOM_HOMEWORK_TYPES_FILE, # NEW
            CUSTOM_HOMEWORK_STATUSES_FILE, # RENAMED
            STUDENT_GROUPS_FILE,
            QUIZ_TEMPLATES_FILE, HOMEWORK_TEMPLATES_FILE,
        ]
        # Also include all files in LAYOUT_TEMPLATES_DIR
        layout_template_files = []
        if os.path.exists(LAYOUT_TEMPLATES_DIR):
            for fname in os.listdir(LAYOUT_TEMPLATES_DIR):
                fpath = os.path.join(LAYOUT_TEMPLATES_DIR, fname)
                if os.path.isfile(fpath): layout_template_files.append(fpath)
        
        try:
            with zipfile.ZipFile(backup_zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for file_path in files_to_backup:
                    if os.path.exists(file_path) and os.path.isfile(file_path):
                        zf.write(file_path, arcname=os.path.basename(file_path))
                for file_path in layout_template_files:
                     zf.write(file_path, arcname=os.path.join(LAYOUT_TEMPLATES_DIR_NAME, os.path.basename(file_path)))
            self.update_status(f"Backup created: {os.path.basename(backup_zip_path)}")
            messagebox.showinfo("Backup Successful", f"All application data backed up to:\n{backup_zip_path}", parent=self.root)
        except Exception as e:
            messagebox.showerror("Backup Error", f"Failed to create backup: {e}", parent=self.root)
            self.update_status(f"Error creating backup: {e}")
        finally:
            self.password_manager.record_activity()

    def restore_all_data_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Restore", "Enter password to restore data:"): return

        if not messagebox.askyesno("Confirm Restore", "Restoring data will OVERWRITE all current application data (students, logs, settings, etc.) with the contents of the backup.\nThis action CANNOT BE UNDONE directly from the backup itself.\n\nAre you sure you want to proceed?", parent=self.root, icon='warning', default=messagebox.NO):
            self.update_status("Restore cancelled."); return

        backup_zip_path = filedialog.askopenfilename(
            title="Select Backup File to Restore",
            filetypes=[("ZIP archives", "*.zip")],
            parent=self.root
        )
        if not backup_zip_path:
            self.update_status("Restore cancelled."); return

        app_data_dir = os.path.dirname(DATA_FILE) # Get the directory where app data is stored
        
        try:
            with zipfile.ZipFile(backup_zip_path, 'r') as zf:
                # Preliminary check for main data file to guess version compatibility (optional)
                main_data_filename_in_zip = None
                expected_main_data_filename = os.path.basename(DATA_FILE) # e.g. classroom_data_v9.json
                
                # List of all possible data file names across versions to check against
                possible_main_data_files = [f"classroom_data_v{i}.json" for i in range(1,11)] + ["classroom_data.json"] # Make sure to make the range one above the current data version tag, so that it catches itself

                found_compatible_main_file = False
                for name_in_zip in zf.namelist():
                    if os.path.basename(name_in_zip) in possible_main_data_files:
                        main_data_filename_in_zip = os.path.basename(name_in_zip)
                        found_compatible_main_file = True
                        break
                
                if not found_compatible_main_file:
                    messagebox.showerror("Restore Error", "The selected ZIP file does not appear to be a valid application backup (missing main data file).", parent=self.root)
                    return

                # Clear existing layout templates before extraction
                if os.path.exists(LAYOUT_TEMPLATES_DIR):
                    for item_name in os.listdir(LAYOUT_TEMPLATES_DIR):
                        item_path = os.path.join(LAYOUT_TEMPLATES_DIR, item_name)
                        try:
                            if os.path.isfile(item_path) or os.path.islink(item_path): os.unlink(item_path)
                            elif os.path.isdir(item_path): shutil.rmtree(item_path)
                        except Exception as e_del_layout: print(f"Failed to delete old layout item {item_path}: {e_del_layout}")
                else:
                    os.makedirs(LAYOUT_TEMPLATES_DIR, exist_ok=True)


                # Extract files directly into the application data directory
                # This will overwrite existing files with the same names.
                for member in zf.infolist():
                    # Handle paths correctly: extract to app_data_dir, but if member.filename includes a dir (like layout_templates), recreate that structure
                    target_path = os.path.join(app_data_dir, member.filename)
                    # Ensure parent directory exists for the target_path
                    target_dir_for_file = os.path.dirname(target_path)
                    if not os.path.exists(target_dir_for_file):
                        os.makedirs(target_dir_for_file, exist_ok=True)

                    if not member.is_dir(): # Check if it's a file
                        with open(target_path, "wb") as outfile:
                            outfile.write(zf.read(member.filename))
            
            # After extraction, reload data from the potentially new main data file.
            # The main data file name in the backup might be an older version.
            # The load_data method handles migration.
            path_to_load_after_restore = os.path.join(app_data_dir, main_data_filename_in_zip) if main_data_filename_in_zip else DATA_FILE

            self.load_data(file_path=path_to_load_after_restore, is_restore=True) # Reload all data from extracted files
            self.load_custom_behaviors(); self.load_custom_homework_statuses(); #self.load_custom_homework_session_types()
            self.load_student_groups(); self.load_quiz_templates(); self.load_homework_templates()
            self._ensure_next_ids() # Crucial after loading potentially old data
            self.update_all_behaviors(); self.update_all_homework_log_behaviors(); self.update_all_homework_session_types()
            self.draw_all_items(check_collisions_on_redraw=True)
            self.update_undo_redo_buttons_state()
            self.update_lock_button_state()
            self.toggle_student_groups_ui_visibility()
            self.mode_var.set(self.settings.get("current_mode", "behavior")); self.toggle_mode()


            self.update_status("Data restored successfully. Application reloaded.")
            messagebox.showinfo("Restore Successful", "Data restored from backup. The application has reloaded the restored data.", parent=self.root)

        except FileNotFoundError:
            messagebox.showerror("Restore Error", "Backup file not found.", parent=self.root)
        except zipfile.BadZipFile:
            messagebox.showerror("Restore Error", "Invalid or corrupted backup ZIP file.", parent=self.root)
        except Exception as e:
            messagebox.showerror("Restore Error", f"Failed to restore data: {e}", parent=self.root)
            self.update_status(f"Error restoring data: {e}")
            # Attempt to reload current (pre-restore attempt) data to stabilize
            self.load_data(DATA_FILE, is_restore=False)
            self.draw_all_items()
        finally:
            self.password_manager.record_activity()

    def open_data_folder(self):
        folder_path = os.path.dirname(DATA_FILE)
        try:
            if sys.platform == "win32": os.startfile(folder_path)
            elif sys.platform == "darwin": subprocess.Popen(["open", folder_path])
            else: subprocess.Popen(["xdg-open", folder_path])
            self.update_status(f"Opened data folder: {folder_path}")
        except Exception as e:
            self.update_status(f"Error opening data folder: {e}")
            messagebox.showerror("Error", f"Could not open data folder: {e}\nPath: {folder_path}", parent=self.root)

    def open_last_export_folder(self):
        if self.last_excel_export_path and os.path.exists(os.path.dirname(self.last_excel_export_path)):
            self.open_specific_export_folder(self.last_excel_export_path)
        else: self.update_status("Last export path not set or not found.")

    def open_specific_export_folder(self, file_path_in_folder):
        folder_path = os.path.dirname(file_path_in_folder)
        try:
            if sys.platform == "win32": os.startfile(folder_path)
            elif sys.platform == "darwin": subprocess.Popen(["open", folder_path])
            else: subprocess.Popen(["xdg-open", folder_path])
            self.update_status(f"Opened folder: {folder_path}")
        except Exception as e:
            self.update_status(f"Error opening folder {folder_path}: {e}")
            messagebox.showerror("Error", f"Could not open folder: {e}\nPath: {folder_path}", parent=self.root)

    def show_help_dialog(self):
        HelpDialog(self.root, APP_VERSION)

    def show_undo_history_dialog(self):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to View History", "Enter password to view undo history:"): return
        # Ensure dialogs module is available where UndoHistoryDialog is defined
        from undohistorydialog import UndoHistoryDialog
        # Check if a dialog is already open, if so, bring to front or recreate
        if hasattr(self, '_undo_history_dialog_instance') and self._undo_history_dialog_instance.winfo_exists():
            self._undo_history_dialog_instance.lift()
            self._undo_history_dialog_instance.populate_history() # Refresh content
        else:
            self._undo_history_dialog_instance = UndoHistoryDialog(self.root, self)
        self.password_manager.record_activity()

    def selective_redo_action(self, target_command_index_in_undo_stack):
        if self.password_manager.is_locked:
            if not self.prompt_for_password("Unlock to Redo Action", "Enter password to perform this redo action:"):
                return

        if not (0 <= target_command_index_in_undo_stack < len(self.undo_stack)):
            messagebox.showerror("Error", "Invalid action selected for redo.", parent=self.root)
            return

        # Commands to be undone to reach the target command (these come after the target in execution order)
        commands_to_undo_count = len(self.undo_stack) - 1 - target_command_index_in_undo_stack

        temp_undone_for_redo_stack = []

        # 1. Undo actions that occurred *after* the target command
        for _ in range(commands_to_undo_count):
            if not self.undo_stack: break # Should not happen if logic is correct
            command_to_temporarily_undo = self.undo_stack.pop()
            try:
                command_to_temporarily_undo.undo()
                temp_undone_for_redo_stack.append(command_to_temporarily_undo) # Keep them in order of undoin
            except Exception as e:
                messagebox.showerror("Selective Redo Error", f"Error undoing a subsequent action: {e}", parent=self.root)
                # Attempt to restore state might be complex; for now, stop and alert user.
                # Re-push commands that were successfully undone before error?
                # Or, more simply, acknowledge that the state might be partially changed.
                self.undo_stack.append(command_to_temporarily_undo) # Put it back if undo failed
                for cmd_to_re_push in reversed(temp_undone_for_redo_stack): # Re-push successfully undone ones
                    self.undo_stack.append(cmd_to_re_push)
                self.draw_all_items(check_collisions_on_redraw=True)
                return

        # 2. The target command is now at the top of the undo_stack. Pop it.
        if not self.undo_stack or len(self.undo_stack) -1 != target_command_index_in_undo_stack :
             messagebox.showerror("Error", "Undo stack state error during selective redo.", parent=self.root)
             # Restore temp_undone_for_redo_stack to undo_stack before returning
             for cmd_to_re_push in reversed(temp_undone_for_redo_stack): self.undo_stack.append(cmd_to_re_push)
             return

        target_command = self.undo_stack.pop()

        # 3. Undo the target command itself (to get its original pre-state for redo, and add to redo_stack)
        try:
            target_command.undo()
            # self.redo_stack.append(target_command) # Standard undo would do this.
                                                  # For selective redo, we are immediately re-executing it.
                                                  # The key is that subsequent history is invalidated.
        except Exception as e:
            messagebox.showerror("Selective Redo Error", f"Error undoing the target action: {e}", parent=self.root)
            self.undo_stack.append(target_command) # Put target back
            for cmd_to_re_push in reversed(temp_undone_for_redo_stack): self.undo_stack.append(cmd_to_re_push) # Put subsequent back
            self.draw_all_items(check_collisions_on_redraw=True)
            return

        # 4. Re-execute the target command
        try:
            target_command.execute()
            self.undo_stack.append(target_command) # Add it back to the undo_stack as the new latest action
        except Exception as e:
            messagebox.showerror("Selective Redo Error", f"Error re-executing the target action: {e}", parent=self.root)
            # State might be inconsistent. Try to restore the target command to its "undone" state.
            # This is tricky. Simplest is to inform user.
            # For now, we'll leave it as executed on the undo_stack and let user manually undo if needed.
            self.draw_all_items(check_collisions_on_redraw=True)
            return

        # 5. Invalidate subsequent history: Clear the redo_stack and the temp_undone_for_redo_stack is discarded.
        self.redo_stack.clear()
        # temp_undone_for_redo_stack is naturally discarded as it's a local variable.
        # These actions are now "lost" as a new history branch has been created.

        self.update_status(f"Redid action: {target_command.get_description()}. Subsequent history cleared.")
        self.draw_all_items(check_collisions_on_redraw=True)
        self.save_data_wrapper(source="selective_redo")
        self.password_manager.record_activity()
        # The UndoHistoryDialog should refresh itself.

    def on_exit_protocol(self, force_quit=False):

        #dialog = ExitConfirmationDialog(self.root, "Exit Confirmation")
        #if dialog.result == "save_quit":
        #    self.save_data_wrapper(source="exit_save")
        #    self.root.destroy()
        #elif dialog.result == "no_save_quit":
        #    self.root.destroy()
        # If dialog.result is None (Cancel), do nothing.
        
        try:
            if not force_quit:
                if self.password_manager.is_locked:
                    if not self.prompt_for_password("Unlock to Save & Quit", "Enter password to save and quit:"): return

                if self.is_live_quiz_active and not self.prompt_end_live_session_on_mode_switch("quiz"): return
                if self.is_live_homework_active and not self.prompt_end_live_session_on_mode_switch("homework"): return

                #if messagebox.askyesno("Exit", "Save changes and exit application?", parent=self.root, ):
                #    self.save_data_wrapper(source="exit_protocol")
                #else: # User chose not to save, but still wants to exit
                #    self.update_status("Exited without saving.")
                dialog = ExitConfirmationDialog(self.root, "Exit Confirmation")
                if dialog.result == "save_quit":
                    self.save_data_wrapper(source="exit_protocol")
                    self.root.destroy()
                    sys.exit(0) # Ensure clean exit
                elif dialog.result == "no_save_quit":
                    #if self.file_lock_manager: self.file_lock_manager.release_lock()
                    self.update_status("Exited without saving.")
                    self.root.destroy()
                    
                    sys.exit(0) # Ensure clean exit
            else: # Force quit (e.g. after save_and_quit or if lock fails)
                self.root.destroy()
                sys.exit(0) # Ensure clean exit # Data should have been saved by save_and_quit if called from there
        except Exception as e:
            print(f"Error during exit procedure: {e}") # Log error but proceed with exit
            self.root.destroy()
            sys.exit(0) # Ensure clean exit
        #finally:
        #    
        #    self.root.destroy()
        #    sys.exit(0) # Ensure clean exit

class ScrollableToolbar(ttk.Frame):
    """A horizontally scrollable frame, used for toolbars."""
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        
        self.height_adjusted = 50
        # The canvas that will contain the scrollable content
        self.canvas = tk.Canvas(self, highlightthickness=0, borderwidth=0, height=self.height_adjusted)
        
        # The scrollbar
        self.scrollbar = ttk.Scrollbar(self, orient="horizontal", command=self.canvas.xview)
        
        # The interior frame to hold the widgets
        self.interior = ttk.Frame(self.canvas)

        # Place the scrollbar and canvas
        self.scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Link canvas to scrollbar
        self.canvas.configure(xscrollcommand=self.scrollbar.set)

        # Create a window in the canvas that contains the interior frame
        self.canvas.create_window((0, 0), window=self.interior, anchor="nw")

        # Update scrollregion when the interior frame's size changes
        self.interior.bind('<Configure>', self._on_frame_configure)

        # Bind mousewheel scrolling for convenience
        self.canvas.bind('<Enter>', self._bind_mousewheel)
        self.canvas.bind('<Leave>', self._unbind_mousewheel)

    def _on_frame_configure(self, event=None):
        """Update the scrollregion of the canvas."""
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self.canvas.configure(height=self.height_adjusted)

    def _bind_mousewheel(self, event):
        """Bind mousewheel for horizontal scrolling."""
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _unbind_mousewheel(self, event):
        """Unbind mousewheel."""
        self.canvas.unbind_all("<MouseWheel>")

    def _on_mousewheel(self, event):
        """Handle horizontal scrolling with the mousewheel."""
        self.canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")




# --- Main Execution ---
if __name__ == "__main__":
    try:
        import pyi_splash
        # You can optionally update the splash screen text as things load
        pyi_splash.update_text("Loading UI...")
    except ImportError:
        pyi_splash = None # Will be None when not running from a PyInstaller bundle
    except RuntimeError: pass

    root = tk.Tk()
    # Apply a theme if available and desired
    try:
        # Examples: 'clam', 'alt', 'default', 'classic'
        # Some themes might require python -m tkinter to see available ones on your system
        # Or use ttkthemes for more options: from ttkthemes import ThemedTk
        # root = ThemedTk(theme="arc") # Example using ttkthemes
        style = ttk.Style(root)
        #available_themes = style.theme_names() # ('winnative', 'clam', 'alt', 'default', 'classic', 'vista', 'xpnative') on Windows
        # print("Available themes:", available_themes)
        sv_ttk.set_theme("Light")
        #if 'vista' in available_themes: style.theme_use('vista')
        #elif 'xpnative' in available_themes: style.theme_use('xpnative')

    except Exception as e_theme:
        print(f"Could not apply custom theme: {e_theme}")
        
    app = SeatingChartApp(root)
    
    try:
        t = threading.Thread(target=darkdetect.listener, args=(app.theme_auto, ))
        t.daemon = True
        t.start()
    except: pass

    # Close the splash screen once the main app is initialized and ready
    try: pyi_splash.close()
    except: pass
    
    root.mainloop()